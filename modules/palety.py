"""
Moduł palet — routes dla /palety/*, /produkt/* (edycja), /produkty/* (meta)
"""
from flask import Blueprint, request, redirect, session, flash, jsonify, Response, current_app, render_template, render_template_string
from flask_wtf.csrf import generate_csrf
from datetime import datetime
import os
import json

palety_bp = Blueprint('palety', __name__)


def _get_css():
    from modules.shared import CSS
    return CSS


def _get_gemini_client():
    """Pobiera GEMINI_CLIENT z głównego modułu app"""
    try:
        from app import GEMINI_CLIENT
        return GEMINI_CLIENT
    except ImportError:
        return None


def _get_extract_allegro_params():
    """Pobiera extract_allegro_params z głównego modułu app"""
    from app import extract_allegro_params
    return extract_allegro_params


def _get_auto_kategoryzuj():
    """Pobiera auto_kategoryzuj z shared (unika circular import)"""
    from modules.shared import auto_kategoryzuj
    return auto_kategoryzuj


def _clean_price(val):
    """Wyczyść wartość ceny — usuń symbole walut, spacje, zamień przecinek na kropkę."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s or s.lower() in ('n/a', 'nan', 'none', '', '-'):
        return 0.0
    # Usuń symbole walut i spacje
    import re
    s = re.sub(r'[złZŁ€$£¥₹\s]', '', s)
    s = s.replace(',', '.')
    # Usuń wszystko oprócz cyfr, kropki i minusa
    s = re.sub(r'[^\d.\-]', '', s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _dodaj_dostawca_options(selected=''):
    """Generuje opcje <option> dla selecta dostawcy (dynamiczne)"""
    from modules.database import get_dostawcy_list
    dostawcy = get_dostawcy_list()
    opts = ''
    for d in dostawcy:
        sel = ' selected' if d == selected else ''
        opts += f'<option value="{d}"{sel}>{d}</option>'
    opts += '<option value="__custom__">+ Dodaj nowego...</option>'
    return opts


# ============================================================
# MODULE CSS & JS (like magazynier pattern)
# ============================================================

_PALETY_CSS = '''
/* Palety module — Stitch Design System */
.paleta-card{background:rgba(22,26,33,0.7);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:14px;margin-bottom:10px;box-shadow:0 4px 16px rgba(0,0,0,0.2);transition:all 0.3s}
.paleta-card:hover{border-color:rgba(143,245,255,0.25);box-shadow:0 4px 20px rgba(143,245,255,0.1)}
.progress-bar-wrap{background:rgba(255,255,255,0.06);border-radius:6px;height:8px;overflow:hidden}
.progress-bar-fill{height:100%;border-radius:6px;transition:width 0.3s;background:linear-gradient(90deg,#8ff5ff,#beee00)}
.paleta-stats-box{background:rgba(22,26,33,0.5);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:14px;text-align:center}
.paleta-stats-num{font-size:1.6rem;font-weight:800;font-family:'Space Grotesk','Inter',sans-serif}
.paleta-stats-label{font-size:0.72rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.8px;font-weight:600}
.sale-banner{background:linear-gradient(135deg,rgba(190,238,0,0.1),rgba(143,245,255,0.05));border:1px solid rgba(190,238,0,0.3);border-radius:var(--radius);padding:18px;margin-bottom:15px}
.sale-banner-label{font-size:0.82rem;color:#beee00;text-transform:uppercase;letter-spacing:1.5px;font-weight:700;font-family:'Space Grotesk',sans-serif}
.sale-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.sale-cell{text-align:center}
.sale-cell-val{font-size:1.4rem;font-weight:800;font-family:'Space Grotesk','Inter',sans-serif}
.sale-cell-lbl{font-size:0.68rem;color:rgba(190,238,0,0.7);font-weight:600}
.import-link{display:block;border-radius:var(--radius);padding:16px;margin-bottom:10px;text-decoration:none;color:#fff;transition:all 0.3s}
.import-link:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,0.3)}
.import-link-inner{display:flex;align-items:center;gap:12px}
.import-link-icon{font-size:2rem}
.import-link-title{font-weight:700;font-size:1.1rem;font-family:'Space Grotesk',sans-serif}
.import-link-sub{font-size:0.82rem;opacity:0.9}
.import-link-arrow{margin-left:auto;font-size:1.5rem}
.modal-overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.6);backdrop-filter:blur(8px);z-index:1000;padding:20px}
.modal-box{background:rgba(22,26,33,0.9);backdrop-filter:blur(16px);border:1px solid rgba(143,245,255,0.15);border-radius:var(--radius);max-width:400px;margin:50px auto;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,0.5)}
.modal-box h3{color:#8ff5ff;margin:0 0 15px;font-family:'Space Grotesk',sans-serif}
.product-row{border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:14px;margin-bottom:10px;background:rgba(22,26,33,0.5);transition:all 0.2s}
.product-row:hover{border-color:rgba(143,245,255,0.15)}
.me-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}
.me-stat{background:rgba(22,26,33,0.5);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:14px;text-align:center;transition:all 0.2s}
.me-stat:hover{border-color:rgba(143,245,255,0.15)}
.me-stat-num{font-size:1.4rem;font-weight:800;font-family:'Space Grotesk','Inter',sans-serif}
.me-stat-label{font-size:0.72rem;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:0.5px}
.me-bottom{position:fixed;bottom:0;left:0;right:0;background:rgba(22,26,33,0.95);backdrop-filter:blur(16px);border-top:1px solid rgba(143,245,255,0.15);padding:12px 8px;z-index:100;box-shadow:0 -4px 20px rgba(0,0,0,0.3)}
.me-bottom-inner{display:flex;flex-direction:column;gap:8px;max-width:1800px;margin:0 auto}
.me-bottom-row{display:flex;gap:8px}
.me-btn{flex:1;margin:0;padding:14px 10px;font-size:0.95rem;font-weight:700;border:none;border-radius:10px;color:#fff;cursor:pointer;min-height:48px;text-align:center;text-decoration:none;display:flex;align-items:center;justify-content:center;transition:all 0.2s;font-family:'Space Grotesk','Inter',sans-serif}
.me-btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,0.3)}
.me-btn-back{background:rgba(255,255,255,0.04);color:var(--text);border:1px solid rgba(255,255,255,0.08);flex:0 0 auto;padding:14px 16px}
.me-btn-meta{background:linear-gradient(135deg,#ff6b9b,#8b5cf6)}
.me-btn-wystaw{background:linear-gradient(135deg,#beee00,#22c55e)}
.me-info{font-size:0.82rem;margin-bottom:12px;padding:12px;background:rgba(190,238,0,0.06);border:1px solid rgba(190,238,0,0.15);border-radius:var(--radius-sm);color:var(--text-muted)}
.menu-item{padding:10px 14px;cursor:pointer;border-radius:8px;font-size:0.9rem;transition:all 0.2s}
.menu-item:hover{background:rgba(143,245,255,0.06)}
/* Responsive */
@media(max-width:768px){
    .me-stats{grid-template-columns:repeat(2,1fr)}
    .me-stat-num{font-size:1.1rem}
    .me-bottom{padding:10px 6px}
    .me-bottom-row{gap:6px}
    .me-btn{padding:12px 8px;font-size:0.85rem;min-height:44px}
}
@media(max-width:480px){
    .me-stats{grid-template-columns:repeat(2,1fr);gap:6px}
}
/* === Legacy class overrides → Cyberpunk === */
.header{text-align:center;padding:20px 0;border-bottom:1px solid rgba(143,245,255,0.12);margin-bottom:18px}
.header h1,.header h2{font-family:'Space Grotesk',sans-serif;color:#8ff5ff;text-shadow:0 0 20px rgba(143,245,255,0.25);font-weight:800;display:flex;align-items:center;justify-content:center;gap:10px}
.header small{color:var(--text-muted);font-size:0.8rem;font-weight:400}
.card{backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:18px;margin-bottom:14px;transition:all 0.2s}
.card:hover{border-color:rgba(143,245,255,0.15)}
.card-header{margin-bottom:12px;padding-bottom:10px;border-bottom:1px solid rgba(255,255,255,0.06)}
.card-title{font-family:'Space Grotesk',sans-serif;font-weight:700;font-size:0.95rem;color:var(--text);display:flex;align-items:center;gap:8px}
.stat-row{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:14px}
.stat-box{background:rgba(13,15,26,0.8);border:1px solid rgba(255,255,255,0.06);border-left:3px solid rgba(143,245,255,0.2);padding:14px;text-align:center;transition:all 0.2s}
.stat-box:hover{border-left-color:#8ff5ff;background:rgba(13,15,26,0.95)}
.stat-val{font-size:1.4rem;font-weight:800;font-family:'Space Grotesk',sans-serif;color:var(--text)}
.stat-lbl,.stat-label{font-size:0.6rem;text-transform:uppercase;letter-spacing:1.2px;color:var(--text-muted);font-weight:600;margin-top:4px}
.form-group{margin-bottom:14px}
.form-group label{display:block;font-size:0.72rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-bottom:6px}
.form-control{width:100%;padding:10px 12px;background:rgba(14,14,20,0.8);border:1px solid rgba(143,245,255,0.10);border-radius:8px;color:var(--text);font-family:'Space Grotesk',sans-serif;font-size:0.9rem;transition:all 0.3s;box-sizing:border-box}
.form-control:focus{outline:none;border-color:#8ff5ff;box-shadow:0 0 16px rgba(143,245,255,0.12)}
select.form-control{appearance:auto}
textarea.form-control{min-height:80px;resize:vertical}
.btn{padding:10px 18px;border:none;border-radius:8px;font-weight:700;font-family:'Space Grotesk',sans-serif;cursor:pointer;transition:all 0.2s;font-size:0.85rem;display:inline-flex;align-items:center;gap:6px}
.btn:hover{transform:translateY(-1px)}
.btn-success,.btn-primary{background:rgba(190,238,0,0.12);border:1px solid rgba(190,238,0,0.3);color:#beee00}
.btn-success:hover,.btn-primary:hover{background:rgba(190,238,0,0.2);box-shadow:0 0 12px rgba(190,238,0,0.15)}
.btn-danger{background:rgba(239,68,68,0.12);border:1px solid rgba(239,68,68,0.3);color:#ff4d6a}
.btn-danger:hover{background:rgba(239,68,68,0.2);box-shadow:0 0 12px rgba(239,68,68,0.15)}
.btn-warning{background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.3);color:#f59e0b}
.btn-warning:hover{background:rgba(245,158,11,0.2)}
.btn-purple{background:rgba(139,92,246,0.12);border:1px solid rgba(139,92,246,0.3);color:#a78bfa}
.btn-purple:hover{background:rgba(139,92,246,0.2)}
.back{color:#8ff5ff;text-decoration:none;font-size:0.82rem;font-weight:600}
.back:hover{text-shadow:0 0 8px rgba(143,245,255,0.4)}
.alert{padding:12px 16px;border-radius:8px;margin-bottom:14px;font-size:0.85rem}
.alert-success{background:rgba(34,197,94,0.1);border:1px solid rgba(34,197,94,0.25);color:#22c55e}
.alert-danger,.alert-error{background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.25);color:#ef4444}
.alert-warning{background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.25);color:#f59e0b}
.toggle-row{display:flex;align-items:center;justify-content:space-between;padding:10px 0;border-bottom:1px solid rgba(255,255,255,0.04)}
table{width:100%;border-collapse:collapse}
table th{background:rgba(13,15,26,0.8);color:var(--text-muted);font-size:0.65rem;text-transform:uppercase;letter-spacing:1px;font-weight:700;padding:10px 8px;text-align:left;border-bottom:1px solid rgba(143,245,255,0.1)}
table td{padding:10px 8px;border-bottom:1px solid rgba(255,255,255,0.04);font-size:0.85rem;color:var(--text)}
table tr:hover td{background:rgba(143,245,255,0.02)}
'''

_PALETY_JS = '''
'''

_PALETY_TEMPLATE = '''{% extends "base.html" %}
{% block page_title %}{{ page_title }}{% endblock %}
{% block content %}
<style>{{ palety_css|safe }}</style>
{{ content_html|safe }}
<script>{{ palety_js|safe }}</script>
{% endblock %}'''


def render(content, page_title='Palety', extra_js=''):
    return render_template_string(
        _PALETY_TEMPLATE,
        content_html=content,
        page_title=page_title,
        palety_css=_PALETY_CSS,
        palety_js=_PALETY_JS + extra_js,
        version=current_app.config.get('VERSION', ''),
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        current_user=session.get('user')
    )


# ============================================================
# EXTRAKTOR ALLEGRO - REGENERUJ META TITLE
# ============================================================
@palety_bp.route('/produkty/<int:produkt_id>/regenerate-meta-title', methods=['POST', 'OPTIONS'])
def produkt_regenerate_meta_title(produkt_id):
    """Regeneruje meta_title dla pojedynczego produktu"""
    # CORS preflight
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response

    from modules.database import get_db, get_config

    try:
        # Sprawdź klucz Gemini z DB config
        gemini_key = get_config('gemini_api_key', '')
        if not gemini_key:
            response = jsonify({'success': False, 'error': 'Brak klucza Gemini API - ustaw w Ustawieniach'})
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        # Pobierz produkt
        conn = get_db()
        produkt = conn.execute('SELECT nazwa, ean, asin FROM produkty WHERE id = ?', (produkt_id,)).fetchone()

        if not produkt:
            response = jsonify({'success': False, 'error': 'Produkt nie znaleziony'})
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        # Użyj pełnego tytułu z Amazon (scraped.nazwa) jeśli dostępny
        asin = (produkt['asin'] or '').strip().upper()
        scraped = None
        if asin:
            scraped = conn.execute(
                'SELECT nazwa, bullet_points FROM scraped WHERE asin = ?', (asin,)
            ).fetchone()

        scraped_nazwa = (scraped['nazwa'] if scraped and scraped['nazwa'] else '').strip()
        bullet_pts = (scraped['bullet_points'] if scraped and scraped['bullet_points'] else '')
        prod_nazwa = (produkt['nazwa'] or '').strip()

        # Jeśli brak pełnej nazwy w scraped → spróbuj scrapować Amazon na żywo
        if asin and len(scraped_nazwa) < 25:
            try:
                from modules.utils import scrape_amazon_product
                amazon_data = scrape_amazon_product(asin)
                if amazon_data and amazon_data.get('title') and len(amazon_data['title']) > 20:
                    scraped_nazwa = amazon_data['title']
                    _bp = amazon_data.get('bullet_points', [])
                    if _bp:
                        bullet_pts = json.dumps(_bp) if isinstance(_bp, list) else str(_bp)
                    # Zapisz do scraped i produkty
                    if scraped:
                        conn.execute('UPDATE scraped SET nazwa=?, bullet_points=? WHERE asin=?',
                                    (scraped_nazwa, bullet_pts, asin))
                    else:
                        conn.execute('INSERT OR IGNORE INTO scraped (asin, nazwa, bullet_points, status) VALUES (?,?,?,?)',
                                    (asin, scraped_nazwa, bullet_pts, 'nowy'))
                    conn.execute('UPDATE produkty SET nazwa=? WHERE id=? AND LENGTH(COALESCE(nazwa,"")) < LENGTH(?)',
                                (scraped_nazwa, produkt_id, scraped_nazwa))
                    conn.commit()
                    print(f"[REGEN] Live scrape OK: {scraped_nazwa[:50]}")
            except Exception as _e:
                print(f"[REGEN] Live scrape failed: {_e}")

        # Użyj dłuższej nazwy
        amazon_nazwa = scraped_nazwa if len(scraped_nazwa) >= len(prod_nazwa) else (prod_nazwa or scraped_nazwa)

        # Generuj meta_title przez AI
        from modules.smart_importer import generate_meta_title
        meta_title = generate_meta_title(
            produkt_nazwa=amazon_nazwa,
            produkt_ean=produkt['ean'] or '',
            produkt_asin=asin,
            bullet_points=bullet_pts
        )

        if not meta_title:
            response = jsonify({'success': False, 'error': 'Nie udało się wygenerować tytułu'})
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        # Zapisz do bazy
        conn.execute('UPDATE produkty SET meta_title = ? WHERE id = ?', (meta_title, produkt_id))
        conn.commit()

        response = jsonify({'success': True, 'meta_title': meta_title})
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        return response

    except Exception as e:
        response = jsonify({'success': False, 'error': str(e)})
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        return response


# ============================================================
# EXTRAKTOR ALLEGRO - BATCH GENERATION
# ============================================================
@palety_bp.route('/api/generate_meta_title_batch', methods=['POST', 'OPTIONS'])
def generate_meta_title_batch():
    """Generuje meta_title dla wielu produktów naraz"""
    # CORS preflight
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response

    import time
    from modules.database import get_db

    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])

        # BATCH SIZE LIMIT (zwiększony dla paid tier)
        MAX_BATCH_SIZE = 100  # Zwiększone z 10 na 100 dla paid tier
        if len(product_ids) > MAX_BATCH_SIZE:
            response = jsonify({
                'success': False,
                'error': f'Zbyt dużo produktów! Max {MAX_BATCH_SIZE} na raz. Zaznacz mniej produktów lub podziel na mniejsze batche.'
            })
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        if not product_ids:
            response = jsonify({'success': False, 'error': 'Brak produktów do przetworzenia'})
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        # Sprawdź API key
        from modules.database import get_config
        gemini_key = get_config('gemini_api_key', '')
        if not gemini_key:
            response = jsonify({'success': False, 'error': 'Brak klucza Gemini API - ustaw w Ustawieniach'})
            response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
            return response

        conn = get_db()
        results = {
            'success': True,
            'total': len(product_ids),
            'generated': 0,
            'failed': 0,
            'details': []
        }

        # Generuj dla każdego produktu
        print(f"\n<span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> [BATCH START] Przetwarzam {len(product_ids)} produktów...")

        for idx, product_id in enumerate(product_ids, 1):
            try:
                print(f"\n[INVENTORY_2] [{idx}/{len(product_ids)}] Processing product ID: {product_id}")

                # Pobierz produkt
                produkt = conn.execute('SELECT nazwa, ean, asin FROM produkty WHERE id = ?', (product_id,)).fetchone()

                if not produkt:
                    print(f"   ✗ Produkt nie znaleziony w bazie!")
                    results['failed'] += 1
                    results['details'].append({
                        'id': product_id,
                        'status': 'error',
                        'error': 'Produkt nie znaleziony'
                    })
                    continue

                print(f"   → Nazwa z bazy: {produkt['nazwa'][:50]}...")

                # Użyj pełnego tytułu z Amazon (scraped.nazwa) jeśli dostępny
                _asin = (produkt['asin'] or '').strip().upper()
                _scraped = None
                if _asin:
                    _scraped = conn.execute(
                        'SELECT nazwa, tytul_seo, bullet_points FROM scraped WHERE asin = ?', (_asin,)
                    ).fetchone()

                # Jeśli jest już tytul_seo — użyj bezpośrednio
                if _scraped and _scraped['tytul_seo'] and len(_scraped['tytul_seo']) > 10:
                    _t = _scraped['tytul_seo']
                    if len(_t) > 75:
                        _t = _t[:75].rsplit(' ', 1)[0]
                    conn.execute('UPDATE produkty SET meta_title = ? WHERE id = ?', (_t, product_id))
                    conn.commit()
                    results['generated'] += 1
                    results['details'].append({'id': product_id, 'status': 'success', 'meta_title': _t})
                    print(f"   ✓ Użyto tytul_seo ze scraped: {_t}")
                    continue

                _amazon_nazwa = (_scraped['nazwa'] if _scraped and _scraped['nazwa'] else '') or produkt['nazwa'] or ''
                _bullet_pts = (_scraped['bullet_points'] if _scraped and _scraped['bullet_points'] else '')
                print(f"   → Amazon title: {_amazon_nazwa[:60]}")

                # Generuj meta_title
                from modules.smart_importer import generate_meta_title
                meta_title = generate_meta_title(
                    produkt_nazwa=_amazon_nazwa,
                    produkt_ean=produkt['ean'] or '',
                    produkt_asin=_asin,
                    bullet_points=_bullet_pts
                )

                print(f"   ← Otrzymano meta_title: {meta_title[:75] if meta_title else 'BRAK'}")

                if meta_title:
                    # Zapisz do bazy
                    conn.execute('UPDATE produkty SET meta_title = ? WHERE id = ?', (meta_title, product_id))
                    conn.commit()
                    print(f"   ✓ Zapisano do bazy")

                    results['generated'] += 1
                    results['details'].append({
                        'id': product_id,
                        'status': 'success',
                        'meta_title': meta_title
                    })
                else:
                    print(f"   ✗ Brak meta_title (puste)")
                    results['failed'] += 1
                    results['details'].append({
                        'id': product_id,
                        'status': 'error',
                        'error': 'Nie udało się wygenerować tytułu'
                    })

                # Delay dla rate limiting (WOLNIEJ = STABILNIEJ)
                if idx < len(product_ids):
                    # Start z większym delay dla stabilności
                    if not hasattr(generate_meta_title_batch, '_api_delay'):
                        generate_meta_title_batch._api_delay = 2.0  # 2s = ~30 req/min (BEZPIECZNY!)

                    print(f"   [HOURGLASS_TOP] Czekam {generate_meta_title_batch._api_delay}s przed następnym...")
                    time.sleep(generate_meta_title_batch._api_delay)

            except Exception as e:
                error_msg = str(e)

                # Sprawdź czy to błąd quota (429)
                if '429' in error_msg or 'quota' in error_msg.lower() or 'exceeded' in error_msg.lower():
                    # AUTO-SLOWDOWN: zwiększ delay
                    if not hasattr(generate_meta_title_batch, '_api_delay'):
                        generate_meta_title_batch._api_delay = 2.0  # Start z 2s

                    old_delay = generate_meta_title_batch._api_delay
                    generate_meta_title_batch._api_delay = min(old_delay * 2, 10.0)  # Max 10s

                    print(f"   [WARNING]  QUOTA EXCEEDED! Zwiększam delay: {old_delay}s → {generate_meta_title_batch._api_delay}s")
                    print(f"   [LIGH] WOLNIEJ = STABILNIEJ!")

                    results['failed'] += 1
                    results['details'].append({
                        'id': product_id,
                        'status': 'error',
                        'error': f'⏰ Quota exceeded! Zwiększono delay do {generate_meta_title_batch._api_delay}s. Upgrade do PAID = 2000 RPM (tylko dodaj kartę!)'
                    })
                    # NIE przerywaj - spróbuj dalej z większym delay
                    continue
                else:
                    results['failed'] += 1
                    results['details'].append({
                        'id': product_id,
                        'status': 'error',
                        'error': error_msg
                    })

        response = jsonify(results)
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        return response

    except Exception as e:
        response = jsonify({'success': False, 'error': str(e)})
        response.headers.add('Access-Control-Allow-Origin', request.host_url.rstrip('/'))
        return response


# ============================================================
# EXTRAKTOR ALLEGRO - UI
# ============================================================
@palety_bp.route('/produkty/<int:produkt_id>/extract-params')
def produkt_extract_params(produkt_id):
    """Strona z parametrami Allegro wygenerowanymi przez AI"""
    from modules.database import get_db
    GEMINI_CLIENT = _get_gemini_client()
    extract_allegro_params = _get_extract_allegro_params()

    conn = get_db()
    produkt = conn.execute('SELECT * FROM produkty WHERE id = ?', (produkt_id,)).fetchone()

    if not produkt:
        return redirect('/palety')

    # Sprawdź czy Gemini jest dostępne
    if not GEMINI_CLIENT:
        content = f'''
        <div class="header">
            <h1><span class=material-symbols-outlined>warning</span> Extraktor Allegro</h1>
            <small>Gemini AI niedostępne</small>
        </div>
        <div class="card" style="margin-bottom:20px">
            <p style="color:var(--red)">Aby użyć Extraktora Allegro, ustaw GEMINI_API_KEY w gemini_config.py:</p>
            <code style="background:var(--bg);padding:10px;display:block;margin-top:10px;color:var(--green)">
            GEMINI_API_KEY = 'twoj_klucz_api'
            </code>
            <p style="margin-top:15px;color:var(--text-muted);font-size:0.9rem">
            Klucz API możesz uzyskać na: <a href="https://aistudio.google.com/apikey" target="_blank" style="color:var(--blue)">Google AI Studio</a>
            </p>
        </div>
        <a href="/palety" class="back">&larr; Powrót do palet</a>
        '''
        return render(content, 'Extraktor Allegro')

    # Generuj parametry
    # Dociągnij bullet_points ze scraped żeby tytuł był lepszy
    bullet_points = []
    if produkt.get('asin'):
        import json as _json
        scraped_row = conn.execute('SELECT bullet_points FROM scraped WHERE asin = ?', (produkt['asin'],)).fetchone()
        if scraped_row and scraped_row['bullet_points']:
            try:
                bullet_points = _json.loads(scraped_row['bullet_points'])
            except:
                pass

    result = extract_allegro_params(
        produkt_nazwa=produkt['nazwa'] or '',
        produkt_ean=produkt['ean'] or '',
        produkt_asin=produkt['asin'] or '',
        bullet_points=bullet_points
    )

    # Sprawdź błędy
    if 'error' in result and result['error']:
        content = f'''
        <div class="header">
            <h1><span class=material-symbols-outlined>cancel</span> Błąd Extraktora</h1>
            <small>Produkt #{produkt_id}</small>
        </div>
        <div class="card" style="margin-bottom:20px">
            <p style="color:var(--red)">{result['error']}</p>
        </div>
        <a href="javascript:history.back()" class="back">&larr; Powrót</a>
        '''
        return render(content, 'Błąd Extraktora')

    meta_title = result.get('meta_title', '')
    params = result.get('params', {})

    # Buduj tabelkę parametrów
    params_html = ''
    for key, value in params.items():
        params_html += f'''
        <tr style="border-bottom:1px solid var(--border)">
            <td style="padding:12px;color:var(--text-muted);font-weight:600">{key}</td>
            <td style="padding:12px;color:var(--text)">{value}</td>
        </tr>
        '''

    if not params_html:
        params_html = '<tr><td colspan="2" style="padding:20px;text-align:center;color:var(--text-muted)">Brak parametrów</td></tr>'

    # Strona wyników
    content = f'''
    <div class="header">
        <h1><span class=material-symbols-outlined>smart_toy</span> Extraktor Allegro</h1>
        <small>Produkt #{produkt_id}</small>
    </div>

    <!-- META TITLE -->
    <div style="background:var(--green-soft);border:2px solid rgba(34,197,94,0.5);border-radius:12px;padding:20px;margin-bottom:20px">
        <div class="section-title"><span class=material-symbols-outlined>edit_note</span> META TYTUŁ ALLEGRO (Skopiuj poniżej)</div>
        <div style="background:var(--bg-card);padding:15px;border-radius:8px;font-size:1.1rem;font-weight:600;color:var(--green);cursor:pointer"
             onclick="navigator.clipboard.writeText(this.innerText); alert('Skopiowano do schowka!')">
            {meta_title or 'Brak tytułu'}
        </div>
        <div style="font-size:0.7rem;color:var(--text-muted);margin-top:8px"><span class=material-symbols-outlined>lightbulb</span> Kliknij aby skopiować</div>
    </div>

    <!-- ORYGINALNA NAZWA -->
    <div class="card" style="margin-bottom:20px">
        <div class="section-title"><span class=material-symbols-outlined>inventory_2</span> ORYGINALNA NAZWA</div>
        <div style="color:var(--text);font-size:0.9rem">{produkt['nazwa'] or 'Brak nazwy'}</div>
    </div>

    <!-- PARAMETRY TECHNICZNE -->
    <div class="section-title"><span class=material-symbols-outlined>settings</span> PARAMETRY TECHNICZNE</div>
    <div class="card" style="padding:0;overflow:hidden;margin-bottom:20px">
        <table style="width:100%;border-collapse:collapse">
            {params_html}
        </table>
    </div>

    <!-- PRZYCISKI -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:20px">
        <form action="/produkty/{produkt_id}/quick-draft" method="POST" style="margin:0">
            <input type="hidden" name="meta_title" value="{meta_title}">
            <button type="submit" class="btn btn-success" style="margin:0">
                <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Wystaw szkic
            </button>
        </form>
        <button onclick="window.print()" class="btn" style="background:var(--blue);margin:0">
            <span class=material-symbols-outlined>print</span> Drukuj
        </button>
        <button onclick="window.location.reload()" class="btn btn-purple" style="margin:0">
            <span class=material-symbols-outlined>sync</span> Regeneruj
        </button>
    </div>

    <a href="javascript:history.back()" class="back">&larr; Powrót</a>
    '''

    return render(content, 'Extraktor Allegro')


@palety_bp.route('/produkty/<int:produkt_id>/quick-draft', methods=['POST'])
def produkt_quick_draft(produkt_id):
    """Szybkie wystawienie szkicu na Allegro z wygenerowanym META_TITLE"""
    from modules.database import get_db
    from modules.allegro_api import create_offer, is_authenticated, upload_image_to_allegro
    import re

    meta_title = request.form.get('meta_title', '').strip()[:75]

    if not meta_title:
        return redirect(f'/produkty/{produkt_id}/extract-params?error=no_title')

    # Sprawdź autoryzację Allegro
    if not is_authenticated():
        content = f'''
        <div class="header">
            <h1><span class=material-symbols-outlined>cancel</span> Błąd Allegro</h1>
            <small>Produkt #{produkt_id}</small>
        </div>
        <div class="card" style="margin-bottom:20px">
            <p style="color:var(--red);font-weight:600">Nie jesteś zalogowany do Allegro!</p>
            <p style="margin-top:10px;color:var(--text-muted)">Musisz najpierw połączyć konto Allegro w ustawieniach.</p>
            <a href="/allegro/auth" class="btn btn-success" style="display:inline-block;width:auto;margin-top:15px;padding:12px 24px">
                Połącz Allegro
            </a>
        </div>
        <a href="javascript:history.back()" class="back">&larr; Powrót</a>
        '''
        return render(content, 'Błąd Allegro')

    # Pobierz dane produktu
    conn = get_db()
    produkt = conn.execute('SELECT * FROM produkty WHERE id = ?', (produkt_id,)).fetchone()

    if not produkt:
        return redirect('/palety')

    # Przygotuj dane oferty
    # Preferuj tytul_seo z scraped (wygenerowany przez AI z bullet points) nad meta_title z formularza
    tytul = meta_title
    if produkt.get('asin'):
        import json as _json
        scraped_seo = conn.execute('SELECT tytul_seo FROM scraped WHERE asin = ?', (produkt['asin'],)).fetchone()
        if scraped_seo and scraped_seo['tytul_seo'] and len(scraped_seo['tytul_seo']) > 10:
            tytul = scraped_seo['tytul_seo'][:75]
            print(f"   [TARGET] Używam tytul_seo ze scraped: {tytul}")
    cena = produkt['cena_allegro'] or 100.0
    ilosc = produkt['ilosc'] or 1
    ean = produkt['ean'] or None
    kategoria = produkt['kategoria'] or ''

    # --- ASIN DEDUP: jeśli jest już aktywna oferta z tym ASIN → dodaj ilość zamiast tworzyć nową ---
    prod_asin = produkt.get('asin')
    if prod_asin:
        existing_offer = conn.execute('''
            SELECT o.allegro_id, o.ilosc, o.tytul, p.paleta
            FROM oferty o
            JOIN produkty p ON o.produkt_id = p.id
            WHERE p.asin = ? AND o.status IN ('aktywna', 'wystawiona')
              AND o.produkt_id != ?
            ORDER BY o.data_wystawienia DESC
            LIMIT 1
        ''', (prod_asin, produkt_id)).fetchone()

        if existing_offer:
            from modules.allegro_api import update_offer_stock as _update_stock
            new_qty = (existing_offer['ilosc'] or 0) + ilosc
            result_upd, err_upd = _update_stock(existing_offer['allegro_id'], new_qty)
            if result_upd is not None or err_upd is None:
                conn.execute("UPDATE produkty SET status='dodano_do_oferty' WHERE id=?", (produkt_id,))
                conn.commit()
                content = f'''
                <div class="header">
                    <h1><span class=material-symbols-outlined>add_shopping_cart</span> Dodano do istniejącej oferty!</h1>
                    <small>ASIN: {prod_asin}</small>
                </div>
                <div style="background:rgba(34,197,94,0.08);border:2px solid rgba(34,197,94,0.4);border-radius:12px;padding:20px;margin-bottom:20px">
                    <div style="font-weight:600;color:var(--green);margin-bottom:10px">
                        <span class=material-symbols-outlined style="vertical-align:middle">check_circle</span>
                        Ilość zaktualizowana na istniejącej ofercie
                    </div>
                    <div style="color:var(--text)">
                        <strong>Oferta:</strong> {existing_offer['tytul'][:60]}...<br>
                        <strong>Skąd:</strong> paleta &ldquo;{existing_offer['paleta'] or '?'}&rdquo;<br>
                        <strong>Nowa ilość:</strong> {new_qty} szt. (+{ilosc} szt. z tej palety)<br>
                        <strong>ID Allegro:</strong> {existing_offer['allegro_id']}
                    </div>
                </div>
                <a href="/palety" class="back">&larr; Powrót do palet</a>
                '''
                return render(content, 'Dodano do oferty')
            else:
                # Błąd aktualizacji — idź normalną ścieżką (utwórz nową ofertę)
                print(f"[DEDUP] update_offer_stock error: {err_upd} — tworzę nową ofertę")

    # Generuj prosty opis (albo użyj istniejącego)
    opis = produkt['opis_ai'] if produkt['opis_ai'] else f'''
    <p><strong>{produkt['nazwa']}</strong></p>
    <p>Stan: {produkt['stan'] or 'Używany'}</p>
    <p>Ilość: {ilosc} szt.</p>
    '''

    # Pobierz zdjęcia z kolumny images (lokalne ścieżki) lub fallback na scraped/zdjecie_url
    zdjecia = []

    # Sposób 1: Pobierz z produkty.images (lokalne ścieżki)
    if produkt.get('images'):
        try:
            import json
            images_data = produkt['images']
            if isinstance(images_data, str):
                zdjecia = json.loads(images_data) if images_data and images_data != '[]' else []
            elif isinstance(images_data, list):
                zdjecia = images_data
            if zdjecia:
                print(f"   [PHOTO_CAMERA] [SOURCE] produkty.images: {len(zdjecia)} plików")
        except Exception as e:
            print(f"   [WARNING]  [ERROR] Parse images: {e}")

    # Sposób 2: FALLBACK na scraped.wszystkie_zdjecia (lokalne ścieżki przez ASIN)
    if not zdjecia and produkt.get('asin'):
        try:
            import json
            scraped = conn.execute('SELECT wszystkie_zdjecia FROM scraped WHERE asin = ?', (produkt['asin'],)).fetchone()
            if scraped and scraped['wszystkie_zdjecia']:
                try:
                    scraped_images = json.loads(scraped['wszystkie_zdjecia'])
                    if scraped_images and len(scraped_images) > 0:
                        zdjecia = scraped_images
                        print(f"   [PHOTO_CAMERA] [SOURCE] scraped.wszystkie_zdjecia: {len(zdjecia)} plików")
                except:
                    pass
        except Exception as e:
            print(f"   [WARNING]  [ERROR] Read scraped: {e}")

    # Sposób 3: Fallback na zdjecie_url
    if not zdjecia and produkt['zdjecie_url']:
        img_url = produkt['zdjecie_url']
        if 'media-amazon.com' in img_url:
            img_url = re.sub(r'\._[A-Z0-9_,]+_\.', '._AC_SL1500_.', img_url)
        zdjecia = [img_url]
        print(f"   [PHOTO_CAMERA] [SOURCE] produkty.zdjecie_url: 1 URL")

    # Zamknij połączenie dopiero teraz

    print(f"   [PHOTO_CAMERA] [TOTAL] {len(zdjecia)} zdjęć do uploadu")

    # Upload zdjęć do Allegro (LOKALNE PLIKI lub URL)
    uploaded_urls = []
    print(f"   [UPLOAD] Uploaduję {len(zdjecia[:8])} zdjęć do Allegro...")
    for idx, path_or_url in enumerate(zdjecia[:8], 1):
        try:
            # Sprawdź czy to lokalny plik czy URL
            if isinstance(path_or_url, str) and not path_or_url.startswith('http'):
                print(f"      [{idx}/{min(len(zdjecia), 8)}] Local file: {path_or_url}")
            else:
                print(f"      [{idx}/{min(len(zdjecia), 8)}] URL: {path_or_url[:60]}...")

            allegro_url = upload_image_to_allegro(path_or_url)
            if allegro_url:
                uploaded_urls.append(allegro_url)
                print(f"      ✓ [{idx}/{min(len(zdjecia), 8)}] Success!")
            else:
                print(f"      ✗ [{idx}/{min(len(zdjecia), 8)}] Failed")
        except Exception as e:
            print(f"      ✗ [{idx}/{min(len(zdjecia), 8)}] Error: {str(e)[:80]}")

    print(f"   [CHECK_CIRCLE] Uploaded {len(uploaded_urls)}/{min(len(zdjecia), 8)} zdjęć")

    # Utwórz ofertę jako szkic
    try:
        offer_data = {
            'name': tytul,
            'category': {'id': kategoria} if kategoria else None,
            'sellingMode': {
                'price': {
                    'amount': str(cena),
                    'currency': 'PLN'
                }
            },
            'stock': {
                'available': ilosc
            },
            'description': {
                'sections': [
                    {
                        'items': [
                            {
                                'type': 'TEXT',
                                'content': opis
                            }
                        ]
                    }
                ]
            },
            'images': [{'url': url} for url in uploaded_urls] if uploaded_urls else [],
            'publication': {
                'status': 'INACTIVE'  # Szkic
            }
        }

        # Dodaj EAN jeśli jest
        if ean:
            offer_data['ean'] = [ean]

        # Wywołaj Allegro API
        result = create_offer(offer_data)

        if result and 'id' in result:
            offer_id = result['id']

            # Zaktualizuj status w bazie
            conn = get_db()
            conn.execute('''
                UPDATE produkty
                SET status = 'szkic',
                    krotki_tytul = ?
                WHERE id = ?
            ''', (tytul, produkt_id))
            conn.commit()

            # Sukces!
            content = f'''
            <div class="header">
                <h1><span class=material-symbols-outlined>check_circle</span> Szkic utworzony!</h1>
                <small>Produkt #{produkt_id}</small>
            </div>

            <div style="background:var(--green-soft);border:2px solid rgba(34,197,94,0.5);border-radius:12px;padding:20px;margin-bottom:20px">
                <div style="font-size:1.2rem;font-weight:600;color:var(--green);margin-bottom:10px">[CELEBRATION] Oferta na Allegro!</div>
                <div style="color:var(--text);margin-bottom:15px">
                    <strong>Tytuł:</strong> {tytul}<br>
                    <strong>Cena:</strong> {cena:.2f} PLN<br>
                    <strong>ID Allegro:</strong> {offer_id}
                </div>
                <a href="https://allegro.pl/moje-allegro/sprzedaz/drafted/{offer_id}" target="_blank"
                   class="btn btn-success" style="display:inline-block;width:auto;padding:12px 24px">
                    <span class=material-symbols-outlined>edit_note</span> Zobacz szkic na Allegro
                </a>
            </div>

            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px">
                <a href="/produkty/{produkt_id}/extract-params" class="btn" style="background:var(--blue)">
                    <span class=material-symbols-outlined>sync</span> Wygeneruj ponownie
                </a>
                <a href="javascript:history.back()" class="btn btn-secondary">
                    ← Powrót
                </a>
            </div>
            '''
            return render(content, 'Szkic utworzony')
        else:
            raise Exception("Nie otrzymano ID oferty z Allegro")

    except Exception as e:
        # Błąd
        content = f'''
        <div class="header">
            <h1><span class=material-symbols-outlined>cancel</span> Błąd wystawiania</h1>
            <small>Produkt #{produkt_id}</small>
        </div>
        <div class="card" style="margin-bottom:20px">
            <p style="color:var(--red);font-weight:600">Nie udało się wystawić szkicu:</p>
            <p style="margin-top:10px;color:var(--text-muted)">{str(e)}</p>
        </div>
        <a href="javascript:history.back()" class="back">&larr; Powrót</a>
        '''
        return render(content, 'Błąd wystawiania')


@palety_bp.route('/palety/<int:paleta_id>/edit', methods=['GET', 'POST'])
def paleta_edit(paleta_id):
    """Edycja palety - formularz"""
    from modules.database import get_db

    conn = get_db()

    if request.method == 'POST':
        # Pobierz dane z formularza
        nazwa = request.form.get('nazwa', '').strip()
        dostawca = request.form.get('dostawca', '').strip()
        if dostawca == '__custom__':
            dostawca = request.form.get('dostawca_custom', '').strip()
            if dostawca:
                from modules.database import save_custom_dostawca
                save_custom_dostawca(dostawca)
        regal = request.form.get('regal', '').strip()
        cena_zakupu = float(request.form.get('cena_zakupu', 0))
        # cena_zakupu = brutto
        cena_zakupu_netto = round(cena_zakupu / 1.23, 2) if cena_zakupu > 0 else 0
        data_zakupu = request.form.get('data_zakupu', '')
        notatki = request.form.get('notatki', '').strip()
        koszt_jedn = float(request.form.get('koszt_jednostkowy', 0) or 0)
        ilosc_sztuk = int(request.form.get('ilosc_sztuk', 0) or 0)

        # Zaktualizuj paletę
        try:
            conn.execute('''
                UPDATE palety
                SET nazwa = ?, dostawca = ?, cena_zakupu = ?, cena_zakupu_netto = ?, data_zakupu = ?, notatki = ?, regal = ?, koszt_jednostkowy = ?, ilosc_sztuk = ?
                WHERE id = ?
            ''', (nazwa, dostawca, cena_zakupu, cena_zakupu_netto, data_zakupu, notatki, regal, koszt_jedn, ilosc_sztuk, paleta_id))
        except Exception:
            conn.execute('''
                UPDATE palety
                SET nazwa = ?, dostawca = ?, cena_zakupu = ?, cena_zakupu_netto = ?, data_zakupu = ?, notatki = ?, regal = ?, koszt_jednostkowy = ?
                WHERE id = ?
            ''', (nazwa, dostawca, cena_zakupu, cena_zakupu_netto, data_zakupu, notatki, regal, koszt_jedn, paleta_id))
        conn.commit()

        return redirect(f'/palety/{paleta_id}?success=updated')

    # GET - wyświetl formularz
    paleta = conn.execute('SELECT * FROM palety WHERE id = ?', (paleta_id,)).fetchone()

    if not paleta:
        return redirect('/palety')

    # Bezpieczne pobieranie regalu
    try:
        regal_value = paleta['regal'] or ''
    except (KeyError, TypeError):
        regal_value = ''

    # Koszt jednostkowy
    _kj_val = 0
    try:
        _kj_val = float(paleta['koszt_jednostkowy'] or 0)
    except:
        pass

    # Dostawcy - dynamiczna lista
    from modules.database import get_dostawcy_list
    _dostawcy_list = get_dostawcy_list()
    dostawca_opcje = '<option value="">— Wybierz —</option>'
    for d in _dostawcy_list:
        selected = ' selected' if d == paleta['dostawca'] else ''
        dostawca_opcje += f'<option value="{d}"{selected}>{d}</option>'
    # Opcja dodania nowego dostawcy
    dostawca_opcje += '<option value="__custom__">+ Dodaj nowego...</option>'

    # Buduj formularz
    content = f'''
    <div class="header"><h1>&#x270F; Edytuj Palete</h1><small>ID: {paleta_id}</small></div>

    <form method="POST" class="card">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <div class="form-group">
            <label>Nazwa palety</label>
            <input type="text" name="nazwa" value="{paleta['nazwa'] or ''}" class="form-control">
        </div>

        <div class="form-group">
            <label>Dostawca</label>
            <select name="dostawca" class="form-control" onchange="if(this.value==='__custom__'){{this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}}else{{this.nextElementSibling.style.display='none'}}">
                {dostawca_opcje}
            </select>
            <input type="text" name="dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px" class="form-control">
        </div>

        <div class="form-group">
            <label><span class=material-symbols-outlined>pin_drop</span> Regal / Lokalizacja</label>
            <input type="text" name="regal" value="{regal_value}" placeholder="np. Migło, Regał A1" class="form-control">
        </div>

        <div class="form-group">
            <label>Cena zakupu (PLN brutto)</label>
            <input type="number" name="cena_zakupu" value="{paleta['cena_zakupu'] or 0}" step="0.01" class="form-control">
        </div>

        <div class="form-group">
            <label>Koszt jednostkowy (netto/szt) - staly</label>
            <input type="number" name="koszt_jednostkowy" value="{_kj_val}" step="0.01" min="0"
                class="form-control" style="border-color:var(--orange)"
                placeholder="Zostaw 0 dla auto-obliczenia z ceny palety">
            <div style="font-size:0.7rem;color:var(--text-muted);margin-top:4px">Cena netto za 1 sztuke. Zostaw 0 = auto z ceny palety / ilosc sztuk</div>
        </div>

        <div class="form-group">
            <label>📦 Ilość sztuk przy zakupie (oryginalna)</label>
            <input type="number" name="ilosc_sztuk" value="{paleta['ilosc_sztuk'] or 0}" min="0" step="1"
                class="form-control" style="border-color:var(--blue)"
                placeholder="np. 60">
            <div style="font-size:0.7rem;color:var(--text-muted);margin-top:4px">Łączna liczba sztuk jaka była przy zakupie. Służy do liczenia sprzedanych (kupiłeś - zostało = sprzedane)</div>
        </div>

        <div class="form-group">
            <label>Data zakupu</label>
            <input type="date" name="data_zakupu" value="{paleta['data_zakupu'] or ''}" class="form-control">
        </div>

        <div class="form-group">
            <label>Notatki</label>
            <textarea name="notatki" rows="4" class="form-control">{paleta['notatki'] or ''}</textarea>
        </div>

        <div style="display:flex;gap:10px;margin-top:20px">
            <button type="submit" class="btn btn-success" style="flex:1">Zapisz zmiany</button>
            <a href="/palety/{paleta_id}" class="btn btn-danger" style="flex:1;text-decoration:none">Anuluj</a>
        </div>
    </form>

    <a href="/palety/{paleta_id}" class="back">&larr; Powrot do palety</a>
    '''

    return render(content, 'Edytuj Palete')


# ============================================================
# ZARZĄDZANIE PALETAMI
# ============================================================
@palety_bp.route('/palety/napraw-ceny')
def napraw_ceny_palet():
    """Uzupełnia brakujące ceny zakupu w paletach (cena_zakupu = 0) - zapisuje netto i brutto"""
    from modules.database import get_db

    conn = get_db()

    # Sprawdź czy kolumna cena_zakupu_netto istnieje
    kolumny = [desc[0] for desc in conn.execute('PRAGMA table_info(palety)').fetchall()]
    ma_kolumne_netto = 'cena_zakupu_netto' in kolumny

    # Jeśli nie ma - dodaj ją
    if not ma_kolumne_netto:
        try:
            conn.execute('ALTER TABLE palety ADD COLUMN cena_zakupu_netto REAL DEFAULT 0')
            conn.commit()
            ma_kolumne_netto = True
            print("[CHECK_CIRCLE] Dodano kolumnę cena_zakupu_netto")
        except:
            pass

    # Pobierz palety z cena_zakupu = 0
    palety = conn.execute('''
        SELECT id, nazwa FROM palety WHERE cena_zakupu IS NULL OR cena_zakupu = 0
    ''').fetchall()

    updated = 0

    for p in palety:
        # Oblicz sumę cen brutto produktów (cena_brutto to ŁĄCZNA cena za produkt, nie za sztukę)
        suma_brutto = conn.execute('''
            SELECT COALESCE(SUM(cena_brutto), 0) FROM produkty WHERE paleta_id = ?
        ''', (p['id'],)).fetchone()[0]
        suma_netto = round(suma_brutto / 1.23, 2) if suma_brutto > 0 else 0

        if suma_brutto > 0:
            # cena_zakupu = BRUTTO
            if ma_kolumne_netto:
                conn.execute('''
                    UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?
                ''', (suma_brutto, suma_netto, p['id']))
            else:
                conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (suma_brutto, p['id']))
            updated += 1
            print(f"[CHECK_CIRCLE] Naprawiono paletę {p['id']}: {p['nazwa']} -> {suma_netto:.0f} netto | {suma_brutto:.0f} brutto")

    conn.commit()

    content = f'''
    <div style="text-align:center;padding:60px 20px">
        <div style="font-size:3rem;margin-bottom:20px"><span class=material-symbols-outlined>check_circle</span></div>
        <div style="font-size:1.2rem">Naprawiono {updated} palet!</div>
        <div style="color:var(--text-muted);margin-top:10px">Ceny zakupu (netto + brutto) zostały uzupełnione</div>
        <a href="/palety" class="btn btn-primary" style="display:inline-block;width:auto;margin-top:20px;padding:12px 24px">Powrót do palet</a>
    </div>
    <script>setTimeout(function(){{ window.location='/palety'; }}, 2000);</script>
    '''
    return render(content, 'Napraw ceny')


@palety_bp.route('/palety/przelicz-brutto')
def przelicz_brutto_palet():
    """Przelicza WSZYSTKIE palety - cena_zakupu = suma netto produktów * 1.23"""
    from modules.database import get_db

    conn = get_db()

    # Pobierz WSZYSTKIE palety
    palety = conn.execute('SELECT id, nazwa, cena_zakupu FROM palety').fetchall()

    updated = 0

    for p in palety:
        # Oblicz sumę cen brutto produktów (cena_brutto = ŁĄCZNA za produkt)
        suma_brutto = conn.execute('''
            SELECT COALESCE(SUM(cena_brutto), 0) FROM produkty WHERE paleta_id = ?
        ''', (p['id'],)).fetchone()[0]
        suma_netto = round(suma_brutto / 1.23, 2) if suma_brutto > 0 else 0

        if suma_brutto > 0:
            stara_cena = p['cena_zakupu'] or 0
            conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?', (suma_brutto, suma_netto, p['id']))
            updated += 1
            print(f"[CHECK_CIRCLE] Paleta {p['id']}: {p['nazwa']} -> {stara_cena:.0f} → {suma_brutto:.0f} zł brutto")

    conn.commit()

    content = f'''
    <div style="text-align:center;padding:60px 20px">
        <div style="font-size:3rem;margin-bottom:20px"><span class=material-symbols-outlined>check_circle</span></div>
        <div style="font-size:1.2rem">Przeliczono {updated} palet!</div>
        <div style="color:var(--text-muted);margin-top:10px">Wszystkie ceny zakupu = suma netto × 1.23 (brutto)</div>
        <a href="/palety" class="btn btn-primary" style="display:inline-block;width:auto;margin-top:20px;padding:12px 24px">Powrót do palet</a>
    </div>
    <script>setTimeout(function(){{ window.location='/palety'; }}, 2000);</script>
    '''
    return render(content, 'Przelicz brutto')


@palety_bp.route('/palety')
def palety_lista():
    # Redirect na Magazynier — jedna strona palet
    return redirect('/magazyn/palety')
    from modules.database import get_palety_list, get_full_stats

    palety = get_palety_list(100)
    stats = get_full_stats()

    palety_html = ''
    for p in palety:
        data = p['data_zakupu'] if p['data_zakupu'] else 'Brak daty'
        # Bezpieczne pobieranie wartości - sqlite3.Row nie ma .get()
        try:
            wartosc_zakupu_prod = p['wartosc_zakupu_produktow'] or 0
        except (KeyError, TypeError):
            wartosc_zakupu_prod = 0

        # Bezpieczne pobieranie regalu
        try:
            regal = p['regal'] if p['regal'] else ''
        except (KeyError, TypeError):
            regal = ''

        # Statystyki sprzedaży
        try:
            sztuk_w_magazynie = p['sztuk_w_magazynie'] or 0
            sprzedano_status = p['sprzedano_status'] or 0
            sprzedano_tabela = p['sprzedano_tabela'] or 0
            try:
                sprzedano_offline = p['sprzedano_offline'] or 0  # sprzedane poza Allegro
            except:
                sprzedano_offline = 0
            try:
                przychod_offline = p['przychod_offline'] or 0  # przychód ze sprzedaży offline
            except:
                przychod_offline = 0
            sprzedano_wartosc_status = p['sprzedano_wartosc_status'] or 0
            sprzedano_wartosc_tabela = p['sprzedano_wartosc_tabela'] or 0
            # ZMIANA: Użyj ceny liczonej z produktów zamiast z tabeli palety
            koszt_palety = p['cena_zakupu'] or wartosc_zakupu_prod  # Cena z palety, fallback na produkty

            # FIX: sprzedano_tabela JUŻ ZAWIERA offline (kupujacy='offline')
            # więc NIE dodajemy sprzedano_offline/przychod_offline osobno!
            if sprzedano_tabela > 0:
                sprzedano_szt = sprzedano_tabela
                sprzedano_wartosc = sprzedano_wartosc_tabela
            else:
                sprzedano_szt = sprzedano_status + sprzedano_offline
                sprzedano_wartosc = sprzedano_wartosc_status + przychod_offline
            # FIX: użyj MAX z dwóch źródeł:
            # 1) sztuk_w_magazynie + sprzedano_szt (dla produktów z ilosc=0 po sprzedaży)
            # 2) SUM(ilosc) z produktów (dla produktów z zachowanym oryginalnym ilosc)
            # Np. Plecaki: ilosc=42 ale sprzedano 17 → max(0+17, 42) = 42
            # Np. Bieżnie: ilosc=0 (sprzedany) → max(0+13, 0) = 13
            stary_lacznie = sztuk_w_magazynie + sprzedano_szt
            try:
                ilosc_total = p['sztuk_lacznie_total'] or 0
            except (KeyError, TypeError):
                ilosc_total = 0
            sztuk_lacznie = max(stary_lacznie, ilosc_total)
            # Przelicz sztuk_w_magazynie na resztę
            sztuk_w_magazynie = max(0, sztuk_lacznie - sprzedano_szt)
        except (KeyError, TypeError):
            sztuk_lacznie = 0
            sprzedano_szt = 0
            sprzedano_wartosc = 0
            koszt_palety = 0

        # Oblicz koszt sprzedanych na podstawie średniej ceny za sztukę
        if sztuk_lacznie > 0 and koszt_palety > 0:
            srednia_cena_szt = koszt_palety / sztuk_lacznie
            sprzedano_koszt = sprzedano_szt * srednia_cena_szt
        else:
            sprzedano_koszt = 0

        # Oblicz zysk netto (przychód - koszt)
        zysk_netto = sprzedano_wartosc - sprzedano_koszt

        # Pasek postępu sprzedaży
        procent_sprzedane = (sprzedano_szt / sztuk_lacznie * 100) if sztuk_lacznie > 0 else 0
        progress_color = 'var(--green)' if procent_sprzedane >= 50 else 'var(--yellow)' if procent_sprzedane >= 20 else 'var(--text-muted)'

        palety_html += f'''
        <div class="paleta-card" id="paleta-card-{p['id']}">
            <div style="display:flex;justify-content:space-between;align-items:start">
                <div style="display:flex;align-items:start;gap:8px">
                    <input type="checkbox" class="paleta-cb" value="{p['id']}" onchange="updateBulkDelete()" style="margin-top:4px;width:18px;height:18px;cursor:pointer;accent-color:var(--red)">
                    <div>
                        <div style="font-weight:600">{p['nazwa'] or f"Paleta #{p['id']}"}</div>
                        <div style="font-size:0.8rem;color:var(--text-muted)"><span class="dostawca-name">{p['dostawca']}</span> • {data}</div>
                        {f'<div style="font-size:0.75rem;color:var(--purple);margin-top:2px"><span class=material-symbols-outlined>pin_drop</span> Regal: {regal}</div>' if regal else ''}
                    </div>
                </div>
                <div style="text-align:right">
                    <div style="font-weight:600;color:var(--red)">{koszt_palety:.0f} zł</div>
                    <div style="font-size:0.75rem;color:var(--text-muted)">{p['produktow']} prod.</div>
                </div>
            </div>

            <!-- PASEK SPRZEDAŻY -->
            <div style="margin-top:10px;background:var(--bg);border-radius:6px;padding:8px">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">
                    <span style="font-size:0.75rem;color:var(--text-secondary)"><span class=material-symbols-outlined>bar_chart</span> Sprzedano:</span>
                    <span style="font-size:0.85rem;font-weight:700;color:{progress_color}">{sprzedano_szt} / {sztuk_lacznie} szt</span>
                </div>
                <div class="progress-bar-wrap">
                    <div class="progress-bar-fill" style="background:{progress_color};width:{procent_sprzedane:.0f}%"></div>
                </div>
                {f'<div style="display:flex;justify-content:space-between;margin-top:6px;font-size:0.7rem"><span style="color:var(--green)"><span class=material-symbols-outlined>paid</span> Zysk: {zysk_netto:+.0f} zł</span><span style="color:var(--text-muted)">({procent_sprzedane:.0f}%)</span></div>' if sprzedano_szt > 0 else ''}
            </div>

            <div style="margin-top:8px;display:flex;justify-content:space-between;font-size:0.75rem">
                <span style="color:var(--red)"><span class=material-symbols-outlined>paid</span> Zakup: {koszt_palety:.0f} zł</span>
                <span style="color:var(--green)">Detal: {p['wartosc_detalu']:.0f} zł</span>
            </div>
            <a href="/palety/{p['id']}" style="display:block;text-align:center;color:var(--blue);margin-top:8px;font-size:0.8rem;text-decoration:none">Szczegóły →</a>
        </div>
        '''

    if not palety:
        palety_html = '<div style="text-align:center;color:var(--text-muted);padding:30px">Brak palet. Dodaj pierwszą!</div>'

    content = f'''
    <div class="header">
        <h1><span class=material-symbols-outlined>inventory_2</span> PALETY</h1>
        <small>Zarządzaj zakupami</small>
    </div>

    <div class="stat-row" style="margin-bottom:15px">
        <div class="stat-box">
            <div class="stat-val blue">{stats['palety_lacznie']}</div>
            <div class="stat-lbl">ŁĄCZNIE</div>
        </div>
        <div class="stat-box">
            <div class="stat-val blue">{stats['palety_miesiac']}</div>
            <div class="stat-lbl">TEN MSC</div>
        </div>
        <div class="stat-box">
            <div class="stat-val red">{stats['palety_lacznie_koszt']:.0f}</div>
            <div class="stat-lbl">WYDANE ZŁ</div>
        </div>
    </div>

    <div style="display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap">
        <a href="/palety/dodaj" class="btn btn-success"><span class=material-symbols-outlined>add</span> DODAJ PALETĘ</a>
        <button type="button" id="bulk-select-btn" onclick="toggleSelectAll()" class="btn" style="background:var(--bg-card);border:1px solid var(--border);font-size:0.8rem"><span class="material-symbols-outlined" style="font-size:1rem;vertical-align:middle">check_box</span> Zaznacz wszystkie</button>
    </div>

    <!-- Pasek masowego usuwania -->
    <div id="bulk-delete-bar" style="display:none;position:sticky;top:60px;z-index:50;background:linear-gradient(135deg,var(--red),#dc2626);border-radius:10px;padding:12px 16px;margin-bottom:12px;display:none;align-items:center;justify-content:space-between;box-shadow:0 4px 15px rgba(239,68,68,0.4)">
        <span style="color:#fff;font-weight:600;font-size:0.9rem" id="bulk-delete-count">0 zaznaczonych</span>
        <button type="button" onclick="bulkDeletePalety()" style="background:#fff;color:var(--red);border:none;padding:8px 20px;border-radius:8px;font-weight:700;cursor:pointer;font-size:0.85rem"><span class=material-symbols-outlined>delete</span> USUŃ ZAZNACZONE</button>
    </div>

    <a href="/palety/przelicz-brutto" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-bottom:15px;font-size:0.8rem" onclick="return confirm('Przeliczyć ceny zakupu wszystkich palet na brutto (netto × 1.23)?')"><span class=material-symbols-outlined>build</span> Przelicz ceny na brutto (+23% VAT)</a>

    <div class="section-title">OSTATNIE PALETY</div>

    {palety_html}

    <a href="/statystyki" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:15px">← Statystyki</a>

    <script>
    function updateBulkDelete() {{
        const checked = document.querySelectorAll('.paleta-cb:checked');
        const bar = document.getElementById('bulk-delete-bar');
        const count = document.getElementById('bulk-delete-count');
        if (checked.length > 0) {{
            bar.style.display = 'flex';
            count.textContent = checked.length + ' zaznaczonych';
        }} else {{
            bar.style.display = 'none';
        }}
    }}

    function toggleSelectAll() {{
        const cbs = document.querySelectorAll('.paleta-cb');
        const allChecked = [...cbs].every(cb => cb.checked);
        cbs.forEach(cb => cb.checked = !allChecked);
        updateBulkDelete();
    }}

    function bulkDeletePalety() {{
        const checked = document.querySelectorAll('.paleta-cb:checked');
        const ids = [...checked].map(cb => cb.value);
        if (ids.length === 0) return;
        if (!confirm('Na pewno usunąć ' + ids.length + ' palet z produktami? Tej operacji nie można cofnąć!')) return;

        fetch('/palety/bulk-delete', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{ids: ids}})
        }})
        .then(r => r.json())
        .then(data => {{
            if (data.ok) {{
                ids.forEach(id => {{
                    const card = document.getElementById('paleta-card-' + id);
                    if (card) card.remove();
                }});
                updateBulkDelete();
                alert('Usunięto ' + data.deleted + ' palet');
                location.reload();
            }} else {{
                alert('Błąd: ' + (data.error || 'Nieznany'));
            }}
        }})
        .catch(e => alert('Błąd: ' + e));
    }}
    </script>
    '''
    return render(content, 'Palety')


@palety_bp.route('/palety/bulk-delete', methods=['POST'])
def palety_bulk_delete():
    """Masowe usuwanie palet z produktami"""
    from modules.database import get_db
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'ok': False, 'error': 'Brak palet do usunięcia'})

        conn = get_db()
        deleted = 0
        for pid in ids:
            pid = int(pid)
            # Usuń produkty palety
            conn.execute('DELETE FROM produkty WHERE paleta_id = ?', (pid,))
            # Usuń paletę
            conn.execute('DELETE FROM palety WHERE id = ?', (pid,))
            deleted += 1
        conn.commit()
        return jsonify({'ok': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:100]})


@palety_bp.route('/palety/dodaj', methods=['GET', 'POST'])
def paleta_dodaj():
    from modules.database import add_paleta

    if request.method == 'POST':
        # Sprawdź limit triala
        from modules.plan_features import check_trial_limit
        from modules.database import get_db as _get_db
        _conn = _get_db()
        _count = _conn.execute('SELECT COUNT(*) FROM palety').fetchone()[0]
        allowed, limit, msg = check_trial_limit('palety', _count)
        if not allowed:
            flash(msg, 'error')
            return redirect('/palety')

        nazwa = request.form.get('nazwa', '')
        dostawca = request.form.get('dostawca', '')
        if dostawca == '__custom__':
            dostawca = request.form.get('dostawca_custom', '').strip()
            if dostawca:
                from modules.database import save_custom_dostawca
                save_custom_dostawca(dostawca)
        if not dostawca:
            dostawca = 'Nieznany'
        regal = request.form.get('regal', '')

        # Bezpieczna konwersja ceny
        cena_str = request.form.get('cena', '0')
        try:
            cena = float(cena_str) if cena_str else 0
        except:
            cena = 0

        data = request.form.get('data', '')
        notatki = request.form.get('notatki', '')

        # Debug log
        print(f"[INVENTORY_2] Dodaję paletę: nazwa={nazwa}, dostawca={dostawca}, cena={cena}")

        paleta_id = add_paleta(nazwa, dostawca, cena, data, notatki, regal)

        print(f"[CHECK_CIRCLE] Utworzono paletę ID: {paleta_id}")

        return redirect(f'/palety/{paleta_id}')

    content = '''
    <div class="header">
        <h1><span class=material-symbols-outlined>add</span> NOWA PALETA</h1>
        <small>Dodaj zakupioną paletę</small>
    </div>

    <!-- IMPORT Z EXCEL -->
    <a href="/palety/import-xlsx" class="import-link" style="background:linear-gradient(135deg,var(--green),#16a34a)">
        <div class="import-link-inner">
            <div class="import-link-icon"><span class=material-symbols-outlined>bar_chart</span></div>
            <div>
                <div class="import-link-title">IMPORT Z EXCEL</div>
                <div class="import-link-sub">Wrzuć plik XLSX z listą produktów</div>
            </div>
            <div class="import-link-arrow">→</div>
        </div>
    </a>

    <a href="/palety/bulk-import" class="import-link" style="background:linear-gradient(135deg,var(--blue),#2563eb)">
        <div class="import-link-inner">
            <div class="import-link-icon"><span class=material-symbols-outlined>inventory_2</span></div>
            <div>
                <div style="font-weight:600;font-size:1rem">BULK IMPORT (wiele palet)</div>
                <div style="font-size:0.75rem;opacity:0.9">Importuj kilka palet naraz z osobnymi plikami</div>
            </div>
            <div class="import-link-arrow">→</div>
        </div>
    </a>

    <div style="text-align:center;color:var(--text-muted);font-size:0.8rem;margin-bottom:15px">— lub dodaj ręcznie —</div>

    <form method="POST" class="card">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <div class="form-group">
            <label>Nazwa / Opis</label>
            <input type="text" name="nazwa" placeholder="np. Mix elektronika #15" class="form-control">
        </div>

        <div class="form-group">
            <label>Dostawca</label>
            <select name="dostawca" class="form-control" onchange="if(this.value==='__custom__'){this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}else{this.nextElementSibling.style.display='none'}">
                ''' + _dodaj_dostawca_options() + '''
            </select>
            <input type="text" name="dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px" class="form-control">
        </div>

        <div class="form-group">
            <label><span class=material-symbols-outlined>pin_drop</span> Regal / Lokalizacja</label>
            <input type="text" name="regal" placeholder="np. Migło, Regał A1, itp." class="form-control">
        </div>

        <div class="form-row" style="margin-bottom:12px">
            <div class="form-group">
                <label>Cena zakupu brutto (zł)</label>
                <input type="number" name="cena" placeholder="2500" step="0.01" class="form-control">
            </div>
            <div class="form-group">
                <label>Data zakupu</label>
                <input type="date" name="data" value="''' + datetime.now().strftime('%Y-%m-%d') + '''" class="form-control">
            </div>
        </div>

        <div class="form-group">
            <label>Notatki</label>
            <textarea name="notatki" rows="2" placeholder="Opcjonalne uwagi..." class="form-control" style="resize:vertical"></textarea>
        </div>

        <button type="submit" class="btn btn-success"><span class=material-symbols-outlined>save</span> ZAPISZ PALETĘ</button>
    </form>

    <a href="/palety" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:15px">← Anuluj</a>
    '''
    return render(content, 'Nowa paleta')


@palety_bp.route('/palety/import-xlsx', methods=['GET', 'POST'])
def paleta_import_xlsx():
    """Import palety z pliku Excel"""
    import pandas as pd
    from modules.database import get_db, add_paleta
    auto_kategoryzuj = _get_auto_kategoryzuj()

    if request.method == 'POST':
        # Obsługa uploadu pliku
        if 'file' not in request.files:
            return redirect('/palety/import-xlsx?error=no_file')

        file = request.files['file']
        if file.filename == '':
            return redirect('/palety/import-xlsx?error=no_file')

        if not file.filename.endswith(('.xlsx', '.xls')):
            return redirect('/palety/import-xlsx?error=wrong_format')

        try:
            # Wczytaj Excel
            df = pd.read_excel(file)

            # Pobierz dane palety z formularza
            nazwa = request.form.get('nazwa', file.filename)
            dostawca = request.form.get('dostawca', '')
            if dostawca == '__custom__':
                dostawca = request.form.get('dostawca_custom', '').strip()
                if dostawca:
                    from modules.database import save_custom_dostawca
                    save_custom_dostawca(dostawca)
            if not dostawca:
                dostawca = 'Nieznany'
            regal = request.form.get('regal', '')
            cena_zakupu = float(request.form.get('cena', 0) or 0)
            data_zakupu = request.form.get('data', datetime.now().strftime('%Y-%m-%d'))

            # Mapowanie kolumn (elastyczne)
            col_nazwa = request.form.get('col_nazwa', '')
            col_ean = request.form.get('col_ean', '')
            col_ilosc = request.form.get('col_ilosc', '')
            col_cena = request.form.get('col_cena', '')
            col_cena_detal = request.form.get('col_cena_detal', '')

            # Utwórz paletę
            paleta_id = add_paleta(nazwa, dostawca, cena_zakupu, data_zakupu, f'Import z: {file.filename}', regal)

            # Dodaj produkty
            conn = get_db()
            produkty_dodane = 0

            for idx, row in df.iterrows():
                try:
                    # Pobierz wartości z wybranych kolumn
                    prod_nazwa = str(row[col_nazwa]) if col_nazwa and col_nazwa in df.columns else f'Produkt {idx+1}'
                    prod_ean = str(row[col_ean]) if col_ean and col_ean in df.columns else ''
                    prod_ilosc = int(row[col_ilosc]) if col_ilosc and col_ilosc in df.columns and pd.notna(row[col_ilosc]) else 1
                    prod_cena = float(row[col_cena]) if col_cena and col_cena in df.columns and pd.notna(row[col_cena]) else 0
                    prod_cena_detal = float(row[col_cena_detal]) if col_cena_detal and col_cena_detal in df.columns and pd.notna(row[col_cena_detal]) else prod_cena * 2
                    # cena_brutto = cena_netto * 1.23 (VAT 23%)
                    prod_cena_brutto = round(prod_cena * 1.23, 2)

                    # Pomiń puste wiersze
                    if not prod_nazwa or prod_nazwa == 'nan' or prod_nazwa.strip() == '':
                        continue

                    # Auto-kategoryzacja na podstawie nazwy
                    prod_kategoria = auto_kategoryzuj(prod_nazwa)

                    # Zawsze INSERT — każda paleta ma własne stany ilościowe
                    conn.execute('''
                        INSERT INTO produkty (nazwa, ean, ilosc, cena_netto, cena_brutto, cena_allegro, paleta_id, dostawca, status, kategoria)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'magazyn', ?)
                    ''', (prod_nazwa[:200], prod_ean, prod_ilosc, prod_cena, prod_cena_brutto, prod_cena_detal, paleta_id, dostawca, prod_kategoria))
                    produkt_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

                    produkty_dodane += 1

                except Exception as e:
                    print(f"Błąd wiersza {idx}: {e}")
                    continue

            # Aktualizuj liczbę produktów w palecie
            conn.execute('UPDATE palety SET ilosc_produktow = ? WHERE id = ?', (produkty_dodane, paleta_id))

            # NIE przeliczaj z sumy - cena_zakupu = STAŁA od momentu importu
            stara_cena = conn.execute('SELECT COALESCE(cena_zakupu, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
            if stara_cena == 0:
                # Nowa paleta - ustaw z sumy cen brutto produktów (cena_brutto = ŁĄCZNA za produkt)
                suma_brutto = conn.execute('SELECT COALESCE(SUM(cena_brutto), 0) FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()[0]
                nowa_netto = round(suma_brutto / 1.23, 2) if suma_brutto > 0 else 0
                nowa_brutto = suma_brutto
            else:
                # Istniejąca paleta - nie ruszaj ceny, tylko dodaj nowe produkty
                nowe_netto = 0  # zostaje stara cena, nowe produkty były importowane przez paletomat który sam akumuluje
                nowa_netto = round(stara_cena / 1.23, 2)
                nowa_brutto = stara_cena

            kolumny = [desc[0] for desc in conn.execute('PRAGMA table_info(palety)').fetchall()]
            if 'cena_zakupu_netto' not in kolumny:
                try:
                    conn.execute('ALTER TABLE palety ADD COLUMN cena_zakupu_netto REAL DEFAULT 0')
                except:
                    pass

            try:
                conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?',
                    (nowa_brutto, nowa_netto, paleta_id))
            except:
                conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (nowa_brutto, paleta_id))
            print(f"[PAID] Cena zakupu palety (stała): {nowa_netto:.2f} netto | {nowa_brutto:.2f} brutto")

            conn.commit()

            return redirect(f'/palety/{paleta_id}?imported={produkty_dodane}')

        except Exception as e:
            return redirect(f'/palety/import-xlsx?error={str(e)[:50]}')

    # GET - pokaż formularz lub podgląd kolumn
    preview_html = ''
    columns = []

    # Jeśli jest plik w sesji - pokaż podgląd
    if 'xlsx_preview' in request.args:
        # TODO: obsługa podglądu
        pass

    error = request.args.get('error', '')
    error_html = ''
    if error:
        error_html = f'<div class="alert alert-error" style="margin-bottom:15px"><span class=material-symbols-outlined>warning</span> Błąd: {error}</div>'

    content = f'''
    <div class="header">
        <h1><span class=material-symbols-outlined>bar_chart</span> IMPORT Z EXCEL</h1>
        <small>Wrzuć plik XLSX z produktami</small>
    </div>

    {error_html}

    <form method="POST" enctype="multipart/form-data" class="card">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">

        <!-- PLIK -->
        <div class="form-group">
            <label><span class=material-symbols-outlined>folder</span> Plik Excel (.xlsx)</label>
            <input type="file" name="file" accept=".xlsx,.xls" required class="form-control">
        </div>

        <!-- DANE PALETY -->
        <div class="section-title" style="margin-top:20px"><span class=material-symbols-outlined>inventory_2</span> DANE PALETY</div>

        <div class="form-row" style="margin-bottom:12px">
            <div class="form-group">
                <label>Nazwa palety</label>
                <input type="text" name="nazwa" placeholder="np. Jobalots #15" class="form-control">
            </div>
            <div class="form-group">
                <label>Dostawca</label>
                <select name="dostawca" class="form-control" onchange="if(this.value==='__custom__'){{this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}}else{{this.nextElementSibling.style.display='none'}}">
                    {_dodaj_dostawca_options()}
                </select>
                <input type="text" name="dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px" class="form-control">
            </div>
        </div>

        <div class="form-row" style="margin-bottom:15px">
            <div class="form-group">
                <label>Cena zakupu brutto (zł)</label>
                <input type="number" name="cena" placeholder="2500" step="0.01" class="form-control">
            </div>
            <div class="form-group">
                <label>Data zakupu</label>
                <input type="date" name="data" value="{datetime.now().strftime('%Y-%m-%d')}" class="form-control">
            </div>
        </div>

        <div class="form-group">
            <label><span class=material-symbols-outlined>pin_drop</span> Regal / Lokalizacja</label>
            <input type="text" name="regal" placeholder="np. Migło, Regał A1, itp." class="form-control">
        </div>

        <!-- MAPOWANIE KOLUMN -->
        <div class="section-title" style="margin-top:20px"><span class=material-symbols-outlined>link</span> MAPOWANIE KOLUMN</div>
        <div style="font-size:0.8rem;color:var(--text-secondary);margin-bottom:12px">Wpisz nazwy kolumn z Twojego Excela (dokładnie jak w nagłówku)</div>

        <div class="form-row" style="margin-bottom:12px">
            <div class="form-group">
                <label>Kolumna z NAZWĄ *</label>
                <input type="text" name="col_nazwa" placeholder="np. Description" required class="form-control">
            </div>
            <div class="form-group">
                <label>Kolumna z EAN</label>
                <input type="text" name="col_ean" placeholder="np. EAN / Barcode" class="form-control">
            </div>
        </div>

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:15px">
            <div class="form-group">
                <label>Ilość</label>
                <input type="text" name="col_ilosc" placeholder="np. Qty" class="form-control">
            </div>
            <div class="form-group">
                <label>Cena zakupu</label>
                <input type="text" name="col_cena" placeholder="np. Unit Price" class="form-control">
            </div>
            <div class="form-group">
                <label>RRP / Detal</label>
                <input type="text" name="col_cena_detal" placeholder="np. RRP" class="form-control">
            </div>
        </div>

        <button type="submit" class="btn btn-success">
            <span class=material-symbols-outlined>download</span> IMPORTUJ PALETĘ
        </button>
    </form>

    <!-- PRZYKŁADOWE NAZWY KOLUMN -->
    <div class="card" style="margin-top:15px">
        <div style="font-weight:600;margin-bottom:10px;color:var(--text-secondary)"><span class=material-symbols-outlined>lightbulb</span> Przykładowe nazwy kolumn</div>
        <div style="font-size:0.85rem;color:var(--text-muted)">
            <b>Jobalots:</b> Description, EAN, Qty, Unit Price, RRP<br>
            <b>Warrington:</b> Item Description, Barcode, Quantity, Cost, Retail<br>
            <b>Miglo:</b> Nazwa, EAN, Ilość, Cena, Cena detal
        </div>
    </div>

    <a href="/palety/dodaj" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:15px">← Powrót</a>
    '''
    return render(content, 'Import z Excel')


# ═══════════════════════════════════════════════════════════════════════════
# BULK IMPORT - WIELE PALET NARAZ
# ═══════════════════════════════════════════════════════════════════════════

@palety_bp.route('/palety/bulk-import', methods=['GET', 'POST'])
def paleta_bulk_import():
    """Import wielu palet naraz - każda z osobnym plikiem XLSX i nazwą"""
    import openpyxl
    import io
    from modules.database import get_db, add_paleta

    auto_kategoryzuj = _get_auto_kategoryzuj()

    if request.method == 'POST':
        try:
            conn = get_db()

            # Pobierz wspólne ustawienia
            dostawca = request.form.get('dostawca', '')
            if dostawca == '__custom__':
                dostawca = request.form.get('dostawca_custom', '').strip()
                if dostawca:
                    from modules.database import save_custom_dostawca
                    save_custom_dostawca(dostawca)
            if not dostawca:
                dostawca = 'Nieznany'
            waluta = request.form.get('waluta', 'EUR').upper()

            # Kurs EUR→PLN
            eur_rate = 1.0
            if waluta == 'EUR':
                from modules.smart_importer import get_eur_pln_rate
                eur_rate = get_eur_pln_rate()

            wyniki = []

            # Iteruj po plikach (max 20)
            for i in range(20):
                file_key = f'file_{i}'
                name_key = f'nazwa_{i}'
                cena_key = f'cena_{i}'
                regal_key = f'regal_{i}'
                typ_key = f'typ_{i}'

                if file_key not in request.files:
                    continue

                file = request.files[file_key]
                if not file or file.filename == '':
                    continue

                if not file.filename.endswith(('.xlsx', '.xls', '.zip')):
                    wyniki.append({'nazwa': file.filename, 'status': 'error', 'msg': 'Nieprawidłowy format (xlsx/zip)'})
                    continue

                # Obsługa ZIP — rozpakuj i wrzuć WSZYSTKIE produkty do JEDNEJ palety
                if file.filename.endswith('.zip'):
                    import zipfile
                    zip_cena_raw = float(request.form.get(cena_key, 0) or 0)
                    zip_regal = request.form.get(regal_key, '').strip() if regal_key else ''
                    zip_data_zakupu = request.form.get('data', datetime.now().strftime('%Y-%m-%d'))
                    zip_typ = request.form.get(typ_key, 'paleta').strip()

                    # Nazwa palety — z formularza lub z nazwy ZIP
                    zip_nazwa = request.form.get(name_key, '').strip()
                    if not zip_nazwa:
                        zip_nazwa = file.filename.rsplit('.', 1)[0]

                    try:
                        zip_data = io.BytesIO(file.read())
                        with zipfile.ZipFile(zip_data) as zf:
                            excel_files = [n for n in zf.namelist() if n.endswith(('.xlsx', '.xls')) and not n.startswith('__MACOSX')]
                            if not excel_files:
                                wyniki.append({'nazwa': file.filename, 'status': 'error', 'msg': 'ZIP nie zawiera plików Excel'})
                                continue

                            # Utwórz JEDNĄ paletę dla całego ZIP
                            zip_cena = round(zip_cena_raw, 2)  # Cena zakupu palety zawsze w PLN
                            paleta_id = add_paleta(zip_nazwa, dostawca, zip_cena, zip_data_zakupu, f'ZIP: {file.filename} ({len(excel_files)} plików)', zip_regal, typ=zip_typ)
                            total_prod_count = 0
                            total_szt = 0
                            files_ok = 0

                            for excel_name in excel_files:
                                try:
                                    excel_data = zf.read(excel_name)
                                    wb_zip = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
                                    ws_zip = wb_zip.active
                                    rows_zip = list(ws_zip.iter_rows(values_only=True))
                                    if len(rows_zip) < 2:
                                        continue
                                    # Znajdź nagłówki
                                    z_headers = []
                                    z_header_row = 0
                                    for ri, rc in enumerate(rows_zip[:10]):
                                        if not rc:
                                            continue
                                        ne = [c for c in rc if c is not None and str(c).strip()]
                                        if len(ne) < 2:
                                            continue
                                        z_headers = [str(c).strip().lower() if c else '' for c in rc]
                                        z_header_row = ri
                                        break
                                    if not z_headers:
                                        continue
                                    # Auto-detect kolumn
                                    col_nazwa = col_ean = col_ilosc = col_cena = col_rrp = col_asin = -1
                                    col_image = -1
                                    for ci, h in enumerate(z_headers):
                                        hl = h.lower().strip()
                                        hc = hl.replace(' ', '')
                                        if any(k in hl for k in ['product title', 'description', 'nazwa', 'item']):
                                            if col_nazwa < 0: col_nazwa = ci
                                        elif hl in ('name', 'title', 'titel', 'bezeichnung', 'product name'):
                                            if col_nazwa < 0: col_nazwa = ci
                                        elif any(k in hl for k in ['barcode', 'ean', 'gtin', 'upc']):
                                            if col_ean < 0: col_ean = ci
                                        elif any(k in hc for k in ['qty', 'quantity', 'ilosc', 'menge']):
                                            if 'amount' not in hc:  # "amount" to zwykle cena
                                                if col_ilosc < 0: col_ilosc = ci
                                        elif hc == 'unitrrp' or hl == 'unit rrp':
                                            # Unit RRP = cena za 1 szt (Jobalot format) — to jest RRP
                                            col_cena = ci
                                            col_rrp = ci
                                        elif hc == 'totalrrp' or hl == 'total rrp':
                                            pass  # Ignoruj total — chcemy unit
                                        elif any(k in hl for k in ['unit price', 'cost', 'price', 'cena', 'preis', 'kosten']):
                                            if 'total' not in hl:
                                                if col_cena < 0: col_cena = ci
                                        elif any(k in hl for k in ['rrp', 'retail', 'msrp', 'uvp']):
                                            if 'total' not in hl:
                                                if col_rrp < 0: col_rrp = ci
                                        elif hl == 'asin':
                                            if col_asin < 0: col_asin = ci
                                        elif hl == 'image 1' or hl == 'image':
                                            col_image = ci
                                    if col_nazwa < 0:
                                        col_nazwa = 0
                                    # Sprawdź czy kolumna ceny to "total"
                                    z_price_is_total = False
                                    if col_cena >= 0 and col_cena < len(z_headers):
                                        hc = z_headers[col_cena].replace(' ', '').lower()
                                        if any(x in hc for x in ['total', 'amount', 'wartosc', 'wartość', 'lineamount', 'gesamt']):
                                            z_price_is_total = True
                                    # Parsuj produkty z tego Excela
                                    file_prod_count = 0
                                    for row_data in rows_zip[z_header_row + 1:]:
                                        if not row_data or all(c is None for c in row_data):
                                            continue
                                        prod_nazwa = str(row_data[col_nazwa] or '').strip() if col_nazwa >= 0 and col_nazwa < len(row_data) else ''
                                        if not prod_nazwa or len(prod_nazwa) < 2:
                                            continue
                                        # Pomiń wiersze podsumowujące
                                        nazwa_lower = prod_nazwa.lower().strip()
                                        if nazwa_lower in ('total', 'razem', 'sum', 'suma', 'gesamt', 'subtotal', 'podsumowanie'):
                                            continue
                                        prod_ean = str(row_data[col_ean] or '').strip() if col_ean >= 0 and col_ean < len(row_data) else ''
                                        if prod_ean in ('nan', 'None', 'none'):
                                            prod_ean = ''
                                        prod_ilosc = 1
                                        if col_ilosc >= 0 and col_ilosc < len(row_data):
                                            try: prod_ilosc = max(1, int(float(row_data[col_ilosc] or 1)))
                                            except: prod_ilosc = 1
                                        prod_cena = 0
                                        if col_cena >= 0 and col_cena < len(row_data):
                                            prod_cena = round(_clean_price(row_data[col_cena]) * eur_rate, 2)
                                        prod_rrp = 0
                                        if col_rrp >= 0 and col_rrp < len(row_data):
                                            prod_rrp = round(_clean_price(row_data[col_rrp]) * eur_rate, 2)
                                        prod_asin = ''
                                        if col_asin >= 0 and col_asin < len(row_data):
                                            prod_asin = str(row_data[col_asin] or '').strip()
                                            if prod_asin in ('nan', 'None', 'none'):
                                                prod_asin = ''
                                        # Jeśli cena to total — podziel na 1 szt
                                        if z_price_is_total and prod_ilosc > 1:
                                            prod_cena = round(prod_cena / prod_ilosc, 2)
                                            prod_rrp = round(prod_rrp / prod_ilosc, 2)
                                        # Zdjęcie z Excela (Jobalot: Image 1)
                                        prod_image = ''
                                        if col_image >= 0 and col_image < len(row_data):
                                            prod_image = str(row_data[col_image] or '').strip()
                                            if prod_image.lower() in ('n/a', 'nan', 'none', ''):
                                                prod_image = ''
                                        # cena_allegro = sugerowana cena sprzedaży (RRP lub cena × 2)
                                        prod_cena_allegro = prod_rrp if prod_rrp > 0 else (round(prod_cena * 2, 2) if prod_cena > 0 else 0)
                                        kategoria = auto_kategoryzuj(prod_nazwa) if auto_kategoryzuj else 'inne'
                                        conn.execute('''INSERT INTO produkty (nazwa, ean, asin, ilosc, cena_netto, cena_brutto, cena_allegro, kategoria, status, paleta_id, dostawca, zdjecie_url)
                                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
                                            (prod_nazwa, prod_ean, prod_asin, prod_ilosc, prod_cena, prod_rrp, prod_cena_allegro, kategoria, 'magazyn', paleta_id, dostawca, prod_image))
                                        file_prod_count += 1
                                        total_szt += prod_ilosc
                                    total_prod_count += file_prod_count
                                    if file_prod_count > 0:
                                        files_ok += 1
                                except Exception as ze:
                                    print(f"[WARNING] ZIP Excel error ({excel_name}): {ze}")
                                    continue

                            # Aktualizuj paletę
                            conn.execute('UPDATE palety SET ilosc_produktow = ? WHERE id = ?', (total_prod_count, paleta_id))
                            if zip_cena_raw == 0 and total_prod_count > 0:
                                auto_cena = conn.execute('SELECT COALESCE(SUM(cena_brutto * ilosc), 0) FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()[0]
                                conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (auto_cena, paleta_id))
                            conn.commit()

                            wyniki.append({
                                'nazwa': zip_nazwa, 'status': 'ok', 'paleta_id': paleta_id,
                                'produkty': total_prod_count, 'szt': total_szt,
                                'plik': f'{file.filename} ({files_ok}/{len(excel_files)} plików)',
                                'typ': zip_typ
                            })
                    except Exception as ze2:
                        wyniki.append({'nazwa': file.filename, 'status': 'error', 'msg': f'Błąd ZIP: {str(ze2)[:100]}'})
                    continue

                # Nazwa palety
                nazwa = request.form.get(name_key, '').strip()
                if not nazwa:
                    # Auto-nazwa z pliku
                    nazwa = file.filename.rsplit('.', 1)[0]

                cena_zakupu_raw = float(request.form.get(cena_key, 0) or 0)
                cena_zakupu = round(cena_zakupu_raw, 2)  # Cena zakupu palety zawsze w PLN
                regal = request.form.get(regal_key, '').strip()
                typ = request.form.get(typ_key, 'paleta').strip()
                data_zakupu = request.form.get('data', datetime.now().strftime('%Y-%m-%d'))

                try:
                    wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        wyniki.append({'nazwa': nazwa or file.filename, 'status': 'error', 'msg': 'Pusty plik (brak danych)'})
                        continue

                    # Auto-detekcja kolumn (skopiowane z paletomat.py)
                    # Szukaj wiersza z nagłówkami (max 10 pierwszych)
                    headers = []
                    header_row_idx = 0
                    for ri, row_check in enumerate(rows[:10]):
                        if not row_check:
                            continue
                        non_empty = [c for c in row_check if c is not None and str(c).strip()]
                        if len(non_empty) < 2:
                            continue
                        text_cells = [c for c in non_empty if not str(c).replace('.', '').replace(',', '').isdigit()]
                        if len(text_cells) < 2:
                            continue
                        headers = [str(c).lower().strip() if c else '' for c in row_check]
                        header_row_idx = ri
                        break

                    print(f"[LIST_ALT] Nagłówki Excel (wiersz {header_row_idx+1}): {[h for h in headers if h]}")

                    col_nazwa_i = -1
                    col_ean_i = -1
                    col_asin_i = -1
                    col_ilosc_i = -1
                    col_unit_price = -1
                    col_netto = -1
                    col_cost = -1
                    col_cena_i = -1
                    col_rrp_i = -1

                    for i, h in enumerate(headers):
                        h_clean = h.replace(' ', '').replace('_', '').replace('-', '').replace('ó', 'o').replace('ś', 's').replace('ć', 'c')
                        h_orig = h.lower()

                        # Nazwa / Description
                        if col_nazwa_i == -1 and any(x in h_clean for x in ['description', 'nazwa', 'name', 'titel', 'title', 'bezeichnung']):
                            if 'price' not in h_clean and 'cena' not in h_clean:
                                col_nazwa_i = i

                        # EAN
                        if col_ean_i == -1 and any(x in h_clean for x in ['ean', 'barcode', 'kodkreskowy', 'gtin']):
                            col_ean_i = i

                        # ASIN / SKU
                        if h_clean == 'asin':
                            col_asin_i = i
                        elif col_asin_i == -1 and 'product' not in h_clean and any(x in h_clean for x in ['sku', 'code', 'artikelnummer', 'article']):
                            col_asin_i = i

                        # Jobalot: "Unit RRP" = cena za 1 szt → użyj jako główną cenę
                        if h_clean == 'unitrrp' or h_orig == 'unit rrp':
                            col_unit_price = i
                            continue
                        # "Total RRP" = cena × ilość → ignoruj (nie unit price)
                        if h_clean == 'totalrrp' or h_orig == 'total rrp':
                            continue

                        # UNIKAJ kolumn z cenami rynkowymi!
                        if any(x in h_orig for x in ['regularn', 'rynkow', 'rrp', 'retail', 'msrp']):
                            if 'total' not in h_orig and 'jednostkow' not in h_orig:
                                col_rrp_i = i
                                continue

                        # NAJWYŻSZY PRIORYTET: Cena jednostkowa sprzedaży
                        if col_unit_price == -1 and 'jednostkow' in h_orig and any(x in h_orig for x in ['sprzeda', 'cena']):
                            col_unit_price = i

                        # WYSOKI: Cena sprzedaży netto
                        if col_netto == -1 and 'sprzeda' in h_orig and 'netto' in h_orig:
                            col_netto = i

                        # ŚREDNI: Unit Cost, Cost, Cena zakupu
                        if col_cost == -1 and any(x in h_clean for x in ['unitcost', 'cenazakupu', 'koszt', 'einkaufspreis', 'unitprice']):
                            col_cost = i

                        # NISKI: Cena sprzedaży
                        if col_cena_i == -1 and 'sprzeda' in h_orig and 'jednostkow' not in h_orig and 'netto' not in h_orig:
                            col_cena_i = i

                        # Ilość
                        if col_ilosc_i == -1 and any(x in h_clean for x in ['ilosc', 'ilość', 'qty', 'quantity', 'sztuk', 'szt', 'pcs', 'pieces', 'count', 'menge', 'anzahl', 'stueck', 'stuck']):
                            col_ilosc_i = i

                    # Wykryj kolumny "total" (cena × ilość, nie unit price)
                    col_total_price = -1
                    for i, h in enumerate(headers):
                        hc = h.replace(' ', '').replace('_', '').lower()
                        if any(x in hc for x in ['totalprice', 'totalcost', 'lineamount', 'linetotal', 'gesamtpreis', 'wartosc', 'wartość', 'amount']):
                            if 'qty' not in hc and 'quantity' not in hc and 'ilosc' not in hc:
                                col_total_price = i

                    # Wybierz najlepszą kolumnę ceny (priorytet jak paletomat)
                    price_is_netto = False
                    price_is_total = False
                    if col_unit_price >= 0:
                        col_cena_i = col_unit_price
                        price_is_netto = True
                    elif col_netto >= 0:
                        col_cena_i = col_netto
                        price_is_netto = True
                    elif col_cost >= 0:
                        col_cena_i = col_cost

                    # Fallback: szukaj ogólnie "price/cena/preis"
                    if col_cena_i == -1:
                        # Jeśli mamy total_price a nie unit — użyj go, ale oznacz jako total
                        if col_total_price >= 0:
                            col_cena_i = col_total_price
                            price_is_total = True
                        else:
                            for i, h in enumerate(headers):
                                hc = h.replace(' ', '').replace('_', '')
                                if any(x in hc for x in ['price', 'cena', 'preis']) and i != col_rrp_i:
                                    col_cena_i = i
                                    break

                    # Sprawdź czy wybrana kolumna to total (nagłówek zawiera total/amount/wartość)
                    if col_cena_i >= 0 and not price_is_total:
                        h_check = headers[col_cena_i].lower().replace(' ', '')
                        if any(x in h_check for x in ['total', 'amount', 'wartosc', 'wartość', 'lineamount', 'gesamt']):
                            price_is_total = True

                    print(f"[BAR_CHART] Bulk import kolumny: nazwa={col_nazwa_i} ean={col_ean_i} asin={col_asin_i} ilosc={col_ilosc_i} cena={col_cena_i} rrp={col_rrp_i} total={price_is_total}")

                    # Utwórz paletę/box
                    paleta_id = add_paleta(nazwa, dostawca, cena_zakupu, data_zakupu, f'Bulk import: {file.filename}', regal, typ=typ)

                    produkty_dodane = 0
                    data_rows = rows[header_row_idx + 1:]
                    for idx, row in enumerate(data_rows):
                        try:
                            # Pomiń puste wiersze
                            if not row or all(c is None or str(c).strip() == '' for c in row):
                                continue

                            # Nazwa — wymagana, pomiń jeśli brak
                            if col_nazwa_i >= 0 and col_nazwa_i < len(row) and row[col_nazwa_i] is not None:
                                prod_nazwa = str(row[col_nazwa_i]).strip()
                            else:
                                continue  # Brak nazwy = pomiń wiersz

                            prod_ean = str(row[col_ean_i]).strip() if col_ean_i >= 0 and col_ean_i < len(row) and row[col_ean_i] is not None else ''
                            prod_asin = str(row[col_asin_i]).strip() if col_asin_i >= 0 and col_asin_i < len(row) and row[col_asin_i] is not None else ''
                            try:
                                prod_ilosc = int(float(str(row[col_ilosc_i]).replace(',', '.'))) if col_ilosc_i >= 0 and col_ilosc_i < len(row) and row[col_ilosc_i] is not None else 1
                            except:
                                prod_ilosc = 1
                            prod_cena_raw = _clean_price(row[col_cena_i]) if col_cena_i >= 0 and col_cena_i < len(row) else 0
                            prod_cena_detal_raw = _clean_price(row[col_rrp_i]) if col_rrp_i >= 0 and col_rrp_i < len(row) else 0
                            if prod_cena_detal_raw == 0:
                                prod_cena_detal_raw = prod_cena_raw * 2

                            # Przelicz EUR→PLN (ceny za 1 szt)
                            prod_cena = round(prod_cena_raw * eur_rate, 2)
                            prod_cena_detal = round(prod_cena_detal_raw * eur_rate, 2)
                            # cena_brutto = cena_netto * 1.23 (VAT 23%) jeśli cena jest netto
                            if price_is_netto:
                                prod_cena_brutto = round(prod_cena * 1.23, 2)
                            else:
                                prod_cena_brutto = round(prod_cena, 2)  # Cena już brutto

                            # Jeśli kolumna ceny to "total" (nie unit) — podziel na 1 szt
                            if price_is_total and prod_ilosc > 1:
                                prod_cena = round(prod_cena / prod_ilosc, 2)
                                prod_cena_brutto = round(prod_cena_brutto / prod_ilosc, 2)
                                prod_cena_detal = round(prod_cena_detal / prod_ilosc, 2)

                            if not prod_nazwa or prod_nazwa in ('nan', 'None', '') or prod_nazwa.strip() == '':
                                continue
                            # Pomiń wiersze podsumowujące
                            nazwa_lower = prod_nazwa.lower().strip()
                            if nazwa_lower in ('total', 'razem', 'sum', 'suma', 'gesamt', 'subtotal', 'podsumowanie'):
                                continue

                            # Wyczyść EAN/ASIN z 'nan'
                            if prod_ean in ('nan', 'None', 'none'):
                                prod_ean = ''
                            if prod_asin in ('nan', 'None', 'none'):
                                prod_asin = ''

                            prod_kategoria = auto_kategoryzuj(prod_nazwa)

                            conn.execute('''
                                INSERT INTO produkty (nazwa, ean, asin, ilosc, cena_netto, cena_brutto, cena_allegro, paleta_id, dostawca, status, kategoria)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'magazyn', ?)
                            ''', (prod_nazwa[:200], prod_ean, prod_asin, prod_ilosc, prod_cena, prod_cena_brutto, prod_cena_detal, paleta_id, dostawca, prod_kategoria))

                            produkty_dodane += 1
                        except:
                            continue

                    # Aktualizuj liczbę, sztuki i cenę
                    _ilosc_sztuk_sum = conn.execute(
                        'SELECT COALESCE(SUM(ilosc), 0) FROM produkty WHERE paleta_id = ?', (paleta_id,)
                    ).fetchone()[0]
                    conn.execute('UPDATE palety SET ilosc_produktow = ?, ilosc_sztuk = ? WHERE id = ?',
                                 (produkty_dodane, _ilosc_sztuk_sum, paleta_id))

                    if cena_zakupu > 0:
                        # User podał cenę zakupu — ustaw ją (nadpisz ewentualny auto-oblicz)
                        cena_netto = round(cena_zakupu / 1.23, 2)
                        try:
                            conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?', (cena_zakupu, cena_netto, paleta_id))
                        except:
                            conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (cena_zakupu, paleta_id))
                    else:
                        # Auto-oblicz z produktów (cena_brutto = ŁĄCZNA za produkt)
                        suma_brutto = conn.execute('SELECT COALESCE(SUM(cena_brutto), 0) FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()[0]
                        suma_netto = round(suma_brutto / 1.23, 2) if suma_brutto > 0 else 0
                        try:
                            conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?', (suma_brutto, suma_netto, paleta_id))
                        except:
                            conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (suma_brutto, paleta_id))

                    wyniki.append({
                        'nazwa': nazwa, 'status': 'ok', 'paleta_id': paleta_id,
                        'produkty': produkty_dodane, 'plik': file.filename, 'typ': typ
                    })

                except Exception as e:
                    wyniki.append({'nazwa': nazwa or file.filename, 'status': 'error', 'msg': str(e)[:80]})

            conn.commit()

            # [INBOX] MERGE BOXES: jeśli user zaznaczył "Połącz wszystkie boxy w jeden"
            merge_boxes = request.form.get('merge_boxes', '0') == '1'
            if merge_boxes:
                box_wyniki = [w for w in wyniki if w['status'] == 'ok' and w.get('typ') == 'box']
                if len(box_wyniki) >= 2:
                    merged_nazwa = request.form.get('merged_box_nazwa', '').strip() or 'Box mix'
                    # Pierwszy box zostaje jako główny
                    main_box_id = box_wyniki[0]['paleta_id']
                    conn.execute('UPDATE palety SET nazwa = ? WHERE id = ?', (merged_nazwa, main_box_id))
                    # Przenieś produkty z pozostałych boxów do głównego
                    total_merged = 0
                    for bw in box_wyniki[1:]:
                        moved = conn.execute('UPDATE produkty SET paleta_id = ? WHERE paleta_id = ?',
                                           (main_box_id, bw['paleta_id'])).rowcount
                        total_merged += moved
                        # Dodaj cenę zakupu do głównego boxa
                        sub_cena = conn.execute('SELECT cena_zakupu FROM palety WHERE id = ?', (bw['paleta_id'],)).fetchone()
                        if sub_cena and sub_cena[0]:
                            conn.execute('UPDATE palety SET cena_zakupu = cena_zakupu + ? WHERE id = ?',
                                       (sub_cena[0], main_box_id))
                        # Usuń pusty box
                        conn.execute('DELETE FROM palety WHERE id = ?', (bw['paleta_id'],))
                        bw['status'] = 'merged'
                    # Przelicz ilości w głównym boxie
                    new_count = conn.execute('SELECT COUNT(*) FROM produkty WHERE paleta_id = ?', (main_box_id,)).fetchone()[0]
                    new_sztuk = conn.execute('SELECT COALESCE(SUM(ilosc), 0) FROM produkty WHERE paleta_id = ?', (main_box_id,)).fetchone()[0]
                    conn.execute('UPDATE palety SET ilosc_produktow = ?, ilosc_sztuk = ? WHERE id = ?',
                               (new_count, new_sztuk, main_box_id))
                    conn.commit()
                    # Zaktualizuj wynik głównego boxa
                    box_wyniki[0]['nazwa'] = merged_nazwa
                    box_wyniki[0]['produkty'] = new_count
                    print(f"[INBOX] Merged {len(box_wyniki)} boxes into '{merged_nazwa}' (id={main_box_id}, {new_count} products)")

            # <span class=material-symbols-outlined>list_alt</span> Dodaj produkty z ASIN do tabeli scraped (żeby pojawiły się w generatorze ofert)
            scrape_count = 0
            try:
                all_paleta_ids = [w['paleta_id'] for w in wyniki if w['status'] == 'ok']
                if all_paleta_ids:
                    placeholders = ','.join('?' * len(all_paleta_ids))
                    new_products = conn.execute(f'''
                        SELECT DISTINCT p.asin, p.nazwa, p.ean, p.cena_netto, p.cena_brutto, p.kategoria, p.zdjecie_url
                        FROM produkty p
                        WHERE p.paleta_id IN ({placeholders})
                        AND p.asin IS NOT NULL AND p.asin != '' AND p.asin != 'nan'
                    ''', all_paleta_ids).fetchall()

                    for np_row in new_products:
                        # Dodaj do scraped jeśli jeszcze nie istnieje
                        exists = conn.execute('SELECT 1 FROM scraped WHERE asin=?', (np_row['asin'],)).fetchone()
                        if not exists:
                            conn.execute('''
                                INSERT INTO scraped (asin, nazwa, kategoria, zdjecie_url, cena_amazon, status, data_scrape)
                                VALUES (?, ?, ?, ?, ?, 'nowy', datetime('now'))
                            ''', (np_row['asin'], np_row['nazwa'], np_row['kategoria'] or '', np_row['zdjecie_url'] or '', np_row['cena_brutto'] or 0))
                    conn.commit()

                    # [AGRI] AUTO-SCRAPING: odpal kombajn dla WSZYSTKICH produktów z ASIN
                    asins_rows = conn.execute(f'''
                        SELECT DISTINCT asin FROM produkty
                        WHERE paleta_id IN ({placeholders})
                        AND asin IS NOT NULL AND asin != '' AND asin != 'nan'
                    ''', all_paleta_ids).fetchall()
                    asins = [r['asin'] for r in asins_rows if r['asin'] and len(r['asin']) >= 5]
                    if asins:
                        from modules.paletomat import auto_process_products
                        auto_process_products(asins)
                        scrape_count = len(asins)
                        print(f"[AGRI] Bulk import → auto-scraping {scrape_count} produktów")
            except Exception as e:
                print(f"[WARNING] Auto-scraping/scraped insert error: {e}")

            # [AUTO-META] Generuj meta_title w tle dla produktów bez tytułu
            ok_paleta_ids = [w['paleta_id'] for w in wyniki if w['status'] == 'ok']
            if ok_paleta_ids:
                import threading as _threading
                def _auto_meta(pids):
                    try:
                        import time as _t
                        _t.sleep(10)  # poczekaj aż auto-scraping pobierze nazwy
                        from modules.database import get_db as _gdb3
                        from modules.smart_importer import _optimize_amazon_title
                        _conn3 = _gdb3()
                        _phs = ','.join('?' * len(pids))
                        _prods = _conn3.execute(
                            f"SELECT id, nazwa, meta_title FROM produkty WHERE paleta_id IN ({_phs}) AND LENGTH(COALESCE(nazwa,'')) >= 20",
                            pids
                        ).fetchall()
                        _upd = 0
                        for _p3 in _prods:
                            try:
                                _new = _optimize_amazon_title(_p3['nazwa'], 75)
                                if _new and len(_new) >= 20 and _new != _p3['meta_title']:
                                    _conn3.execute('UPDATE produkty SET meta_title = ? WHERE id = ?', (_new, _p3['id']))
                                    _upd += 1
                            except:
                                pass
                        if _upd:
                            _conn3.commit()
                        print(f"[AUTO-META] Wygenerowano meta_title dla {_upd}/{len(_prods)} produktów")
                    except Exception as _em:
                        print(f"[WARN] Auto-meta error: {_em}")
                _threading.Thread(target=_auto_meta, args=(ok_paleta_ids,), daemon=True).start()

            # Pokaż wyniki
            ok_count = sum(1 for w in wyniki if w['status'] == 'ok')
            err_count = sum(1 for w in wyniki if w['status'] == 'error')

            results_html = ''
            if waluta == 'EUR':
                results_html += f'<div style="padding:10px;background:var(--blue-soft);border-radius:10px;margin-bottom:12px;font-size:0.85rem"><span class=material-symbols-outlined>currency_exchange</span> Przeliczono ceny EUR → PLN po kursie NBP: <b>{eur_rate:.4f}</b></div>'
            if scrape_count > 0:
                results_html += f'<div style="padding:10px;background:var(--green-soft);border-radius:10px;margin-bottom:12px;font-size:0.85rem"><span class=material-symbols-outlined>agriculture</span> Auto-scraping uruchomiony dla <b>{scrape_count}</b> produktów z ASIN. Zdjęcia pojawią się w tle.</div>'
            for w in wyniki:
                if w['status'] == 'merged':
                    results_html += f'''
                    <div style="display:flex;align-items:center;gap:10px;padding:12px;background:#f59e0b11;border:1px solid #f59e0b44;border-radius:10px;margin-bottom:8px">
                        <div style="font-size:1.5rem"><span class=material-symbols-outlined>inbox</span></div>
                        <div style="flex:1">
                            <div style="font-weight:600;color:#f59e0b">{w['nazwa']}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted)">Połączono z głównym boxem • {w['plik']}</div>
                        </div>
                    </div>'''
                elif w['status'] == 'ok':
                    typ_icon = '<span class=material-symbols-outlined>inbox</span>' if w.get('typ') == 'box' else '<span class=material-symbols-outlined>check_circle</span>'
                    results_html += f'''
                    <div style="display:flex;align-items:center;gap:10px;padding:12px;background:var(--green-soft);border:1px solid rgba(34,197,94,0.3);border-radius:10px;margin-bottom:8px">
                        <div style="font-size:1.5rem">{typ_icon}</div>
                        <div style="flex:1">
                            <div style="font-weight:600">{w['nazwa']}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted)">{w['produkty']} produktów • {w['plik']}</div>
                        </div>
                        <a href="/palety/{w['paleta_id']}" style="color:var(--blue);text-decoration:none;font-size:0.85rem">Otwórz →</a>
                        <a href="/magazyn/produkty?paleta_id={w['paleta_id']}" style="color:#f59e0b;text-decoration:none;font-size:0.8rem;margin-left:8px" title="Zgrupuj produkty w box"><span class=material-symbols-outlined>inbox</span> Box</a>
                    </div>'''
                else:
                    results_html += f'''
                    <div style="display:flex;align-items:center;gap:10px;padding:12px;background:var(--red-soft);border:1px solid rgba(239,68,68,0.3);border-radius:10px;margin-bottom:8px">
                        <div style="font-size:1.5rem"><span class=material-symbols-outlined>cancel</span></div>
                        <div style="flex:1">
                            <div style="font-weight:600">{w['nazwa']}</div>
                            <div style="font-size:0.8rem;color:var(--red)">{w.get('msg', 'Błąd')}</div>
                        </div>
                    </div>'''

            content = f'''
            <div class="header">
                <h1><span class=material-symbols-outlined>bar_chart</span> WYNIKI IMPORTU</h1>
                <small>Zaimportowano {ok_count} palet{', błędy: ' + str(err_count) if err_count else ''}</small>
            </div>
            {results_html}
            <a href="/palety" class="btn" style="background:var(--blue);margin-top:15px"><span class=material-symbols-outlined>inventory_2</span> Przejdź do palet</a>
            <a href="/palety/bulk-import" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:10px"><span class=material-symbols-outlined>bar_chart</span> Importuj kolejne</a>
            '''
            return render(content, 'Wyniki importu')

        except Exception as e:
            return redirect(f'/palety/bulk-import?error={str(e)[:50]}')

    # === GET - formularz ===
    error = request.args.get('error', '')
    error_html = f'<div class="alert alert-error" style="margin-bottom:15px"><span class=material-symbols-outlined>warning</span> {error}</div>' if error else ''

    content = f'''
    <div class="header">
        <h1><span class=material-symbols-outlined>bar_chart</span> BULK IMPORT PALET</h1>
        <small>Importuj wiele palet naraz — każda z osobnym plikiem XLSX</small>
    </div>

    {error_html}

    <form method="POST" enctype="multipart/form-data" id="bulk-form">
    <input type="hidden" name="csrf_token" value="{generate_csrf()}">

    <!-- WSPÓLNE USTAWIENIA -->
    <div class="card" style="margin-bottom:15px">
        <div class="section-title"><span class=material-symbols-outlined>settings</span> WSPÓLNE USTAWIENIA</div>

        <div class="form-row" style="margin-bottom:12px">
            <div class="form-group">
                <label>Dostawca</label>
                <select name="dostawca" class="form-control" onchange="if(this.value==='__custom__'){{this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}}else{{this.nextElementSibling.style.display='none'}}">
                    {_dodaj_dostawca_options()}
                </select>
                <input type="text" name="dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px" class="form-control">
            </div>
            <div class="form-group">
                <label>Waluta cen w pliku</label>
                <select name="waluta" class="form-control" id="waluta-select">
                    <option value="EUR">EUR (przelicz na PLN)</option>
                    <option value="PLN">PLN (bez przeliczania)</option>
                </select>
            </div>
            <div class="form-group">
                <label>Data zakupu</label>
                <input type="date" name="data" value="{datetime.now().strftime('%Y-%m-%d')}" class="form-control">
            </div>
        </div>

        <!-- AUTO-DETEKCJA KOLUMN -->
        <div style="margin-top:15px;padding:12px;background:var(--bg);border-radius:10px;border:1px solid var(--border)">
            <div style="font-size:0.8rem;font-weight:600;color:var(--green);margin-bottom:6px"><span class=material-symbols-outlined>smart_toy</span> AUTO-DETEKCJA KOLUMN</div>
            <div style="font-size:0.75rem;color:var(--text-muted)">
                System automatycznie rozpozna kolumny z Excela:<br>
                <b>Nazwa</b> (Description, Name, Product...),
                <b>EAN</b> (Barcode, GTIN...),
                <b>Ilość</b> (Qty, Quantity...),
                <b>Cena</b> (Unit Price, Cost...),
                <b>RRP</b> (Retail, MSRP...)
            </div>
        </div>
    </div>

    <!-- ZIP UPLOAD -->
    <div class="card" style="margin-bottom:15px;border:2px dashed rgba(255,107,155,0.3);background:rgba(255,107,155,0.04)">
        <div style="text-align:center;padding:10px">
            <div style="font-size:1.1rem;font-weight:700;color:#ff6b9b;margin-bottom:6px"><span class=material-symbols-outlined>inventory_2</span> Szybki import z ZIP</div>
            <div style="font-size:0.8rem;color:var(--text-muted);margin-bottom:12px">Wrzuć plik .zip z wieloma Excelami — system automatycznie utworzy slot dla każdego</div>
            <input type="file" id="zip-upload" accept=".zip" style="display:none" onchange="handleZipUpload(this)">
            <button type="button" onclick="document.getElementById('zip-upload').click()" style="padding:10px 24px;background:linear-gradient(135deg,#ff6b9b,#8ff5ff);border:none;border-radius:10px;color:#fff;font-weight:600;cursor:pointer;font-size:0.9rem">
                <span class=material-symbols-outlined>folder</span> Wybierz plik ZIP
            </button>
            <span id="zip-status" style="margin-left:10px;font-size:0.8rem;color:var(--text-muted)"></span>
        </div>
    </div>

    <!-- PALETY -->
    <div class="section-title"><span class=material-symbols-outlined>inventory_2</span> PALETY DO IMPORTU</div>

    <div id="palety-container"></div>

    <button type="button" onclick="addPaleta()" style="width:100%;padding:14px;background:var(--bg-card);border:2px dashed var(--blue);border-radius:12px;color:var(--blue);font-weight:600;cursor:pointer;margin-bottom:15px;font-size:0.95rem">
        <span class=material-symbols-outlined>add</span> DODAJ PALETĘ
    </button>

    <div id="merge-box-option" style="display:none;margin-bottom:12px;padding:14px;background:#f59e0b11;border:1px solid #f59e0b44;border-radius:12px">
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:#f59e0b;font-weight:600">
            <input type="checkbox" name="merge_boxes" value="1" style="width:18px;height:18px;accent-color:#f59e0b">
            <span class=material-symbols-outlined>inbox</span> Połącz wszystkie boxy w jeden
        </label>
        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:5px;margin-left:28px">Produkty z wszystkich plików typu "Box" trafią do jednego boxa</div>
        <div style="margin-top:8px;margin-left:28px">
            <input type="text" name="merged_box_nazwa" placeholder="Nazwa boxa np. Box mix peruki" style="padding:8px 12px;background:rgba(22,26,33,0.7);border:1px solid rgba(255,255,255,0.06);border-radius:8px;color:var(--text);width:300px;font-size:0.85rem">
        </div>
    </div>

    <button type="submit" id="submit-btn" disabled class="btn btn-success" style="font-size:1.1rem;padding:16px;opacity:0.5">
        <span class=material-symbols-outlined>download</span> IMPORTUJ WSZYSTKIE
    </button>

    </form>

    <a href="/palety/dodaj" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:15px">← Powrót</a>
    '''

    bulk_js = '''
    let paletaCount = 0;

    function handleZipUpload(input) {
        if (!input.files.length) return;
        var file = input.files[0];
        var statusEl = document.getElementById('zip-status');
        statusEl.textContent = ' ' + file.name + ' — dodaję jako slot...';
        statusEl.style.color = '#fbbf24';
        // Dodaj jeden slot z plikiem ZIP
        addPaleta();
        var lastIdx = paletaCount - 1;
        var fileInput = document.querySelector('input[name="file_' + lastIdx + '"]');
        if (fileInput) {
            // Zmień accept na zip
            fileInput.accept = '.xlsx,.xls,.zip';
            // Użyj DataTransfer żeby ustawić plik
            var dt = new DataTransfer();
            dt.items.add(file);
            fileInput.files = dt.files;
            // Ustaw nazwę
            var nameField = document.getElementById('nazwa-' + lastIdx);
            if (nameField) nameField.placeholder = file.name.replace(/\\.[^.]+$/, '') + ' (ZIP)';
        }
        statusEl.textContent = ' ' + file.name + ' — gotowy do importu!';
        statusEl.style.color = '#22c55e';
        input.value = '';
    }

    function addPaleta() {
        const i = paletaCount++;
        const container = document.getElementById('palety-container');

        const div = document.createElement('div');
        div.className = 'paleta-row';
        div.id = 'paleta-' + i;
        div.style.cssText = 'background:var(--bg);border:1px solid var(--border);border-radius:14px;padding:15px;margin-bottom:10px;position:relative';

        div.innerHTML = `
            <button type="button" onclick="removePaleta(${i})" style="position:absolute;top:10px;right:10px;background:var(--red-soft);border:none;border-radius:8px;color:var(--red);padding:4px 10px;cursor:pointer;font-size:0.8rem">✕</button>

            <div id="slot-title-${i}" style="font-weight:600;color:var(--blue);margin-bottom:10px;font-size:0.9rem"><span class=material-symbols-outlined>inventory_2</span> Paleta #${i+1}</div>

            <div style="margin-bottom:10px">
                <label style="display:block;font-size:0.75rem;color:var(--text-secondary);margin-bottom:3px"><span class=material-symbols-outlined>folder</span> Plik Excel</label>
                <input type="file" name="file_${i}" accept=".xlsx,.xls,.zip" required onchange="updateFileName(this, ${i})"
                    class="form-control">
            </div>

            <div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr;gap:8px">
                <div>
                    <label style="display:block;font-size:0.75rem;color:var(--text-secondary);margin-bottom:3px">Nazwa palety</label>
                    <input type="text" name="nazwa_${i}" id="nazwa-${i}" placeholder="Auto z nazwy pliku" class="form-control">
                </div>
                <div>
                    <label style="display:block;font-size:0.75rem;color:var(--text-secondary);margin-bottom:3px"><span class=material-symbols-outlined>paid</span> Cena zakupu</label>
                    <input type="number" name="cena_${i}" placeholder="np. 144.80" step="0.01" class="form-control" title="Ile zapłaciłeś — NIE cena z Excela">
                </div>
                <div>
                    <label style="display:block;font-size:0.75rem;color:var(--text-secondary);margin-bottom:3px">Typ</label>
                    <select name="typ_${i}" class="form-control" onchange="updateSlotType(${i}, this.value)">
                        <option value="paleta">Paleta</option>
                        <option value="box">Box</option>
                    </select>
                </div>
                <div>
                    <label style="display:block;font-size:0.75rem;color:var(--text-secondary);margin-bottom:3px">Regał</label>
                    <input type="text" name="regal_${i}" placeholder="A1" class="form-control">
                </div>
            </div>
        `;

        container.appendChild(div);
        updateSubmitBtn();
    }

    function removePaleta(i) {
        const el = document.getElementById('paleta-' + i);
        if (el) el.remove();
        updateSubmitBtn();
    }

    function updateFileName(input, i) {
        const nameField = document.getElementById('nazwa-' + i);
        if (nameField && !nameField.value && input.files.length) {
            // Auto-fill nazwa z pliku (bez rozszerzenia)
            nameField.placeholder = input.files[0].name.replace(/\\.[^.]+$/, '');
        }
    }

    function updateSlotType(i, typ) {
        const title = document.getElementById('slot-title-' + i);
        if (title) {
            if (typ === 'box') {
                title.innerHTML = '<span class=material-symbols-outlined>inbox</span> Box #' + (i+1);
                title.style.color = '#f59e0b';
            } else {
                title.innerHTML = '<span class=material-symbols-outlined>inventory_2</span> Paleta #' + (i+1);
                title.style.color = 'var(--blue)';
            }
        }
        updateSubmitBtn();
    }

    function updateSubmitBtn() {
        const rows = document.querySelectorAll('.paleta-row');
        const btn = document.getElementById('submit-btn');
        btn.disabled = rows.length === 0;
        btn.style.opacity = rows.length === 0 ? '0.5' : '1';
        btn.textContent = rows.length === 0 ? ' DODAJ PALETY POWYŻEJ' : ' IMPORTUJ ' + rows.length + ' PALET';
        // Pokaż opcję "Połącz w jeden box" jeśli >1 slot z typem box
        let boxCount = 0;
        rows.forEach(function(row) {
            const sel = row.querySelector('select[name^="typ_"]');
            if (sel && sel.value === 'box') boxCount++;
        });
        const mergeDiv = document.getElementById('merge-box-option');
        if (mergeDiv) {
            mergeDiv.style.display = boxCount >= 2 ? 'block' : 'none';
        }
    }

    // Dodaj pierwszą od razu
    addPaleta();
    '''

    return render(content, 'Bulk import palet', extra_js=bulk_js)

# ═══════════════════════════════════════════════════════════════════════════
# MASOWA EDYCJA PALET - Adrian's custom feature v3.1.0
# ═══════════════════════════════════════════════════════════════════════════

@palety_bp.route('/palety/<int:paleta_id>/mass-edit')
def paleta_mass_edit(paleta_id):
    """Strona masowej edycji produktów z palety"""
    from modules.database import get_db

    conn = get_db()
    paleta = conn.execute('SELECT * FROM palety WHERE id = ?', (paleta_id,)).fetchone()

    if not paleta:
        return redirect('/palety')

    # Pobierz produkty z magazynu
    produkty = conn.execute('''
        SELECT * FROM produkty
        WHERE paleta_id = ?
        ORDER BY
            CASE
                WHEN status = 'wystawiony' THEN 1
                WHEN status = 'magazyn' THEN 2
                ELSE 3
            END,
            data_dodania DESC
    ''', (paleta_id,)).fetchall()

    # Stats
    stats = conn.execute('''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status = 'wystawiony' THEN 1 ELSE 0 END) as wystawione,
            SUM(CASE WHEN status = 'magazyn' THEN 1 ELSE 0 END) as magazyn,
            COALESCE(SUM(cena_allegro * ilosc), 0) as wartosc_total
        FROM produkty
        WHERE paleta_id = ?
    ''', (paleta_id,)).fetchone()


    if not produkty or len(produkty) == 0:
        content = f'''
        <div class="header">
            <h1><span class=material-symbols-outlined>warning</span> Brak produktów</h1>
            <small>Paleta #{paleta_id}</small>
        </div>
        <div class="alert alert-warning">
            Ta paleta nie ma jeszcze żadnych produktów. Najpierw zaimportuj produkty z Excel.
        </div>
        <a href="/magazyn/import?paleta_id={paleta_id}" class="btn btn-primary"><span class=material-symbols-outlined>download</span> Importuj produkty</a>
        <a href="/palety/{paleta_id}" class="btn btn-secondary">← Powrót</a>
        '''
        return render(content, 'Brak produktów')

    # Generuj HTML produktów
    produkty_html = ''
    wybrane_count = 0

    for p in produkty:
        # Kolory statusów
        if p['status'] == 'wystawiony':
            status_badge = '<span class="badge badge-success"><span class=material-symbols-outlined>check_circle</span> WYSTAWIONE</span>'
            row_bg = 'var(--green-soft)'
            checkbox_disabled = 'disabled'
            checkbox_checked = ''
        elif p['status'] == 'magazyn':
            status_badge = '<span class="badge" style="background:var(--blue-soft);color:var(--blue)">● MAGAZYN</span>'
            row_bg = 'var(--blue-soft)'
            checkbox_disabled = ''
            checkbox_checked = 'checked'
            wybrane_count += 1
        elif p['status'] == 'szkic':
            status_badge = '<span class="badge" style="background:var(--accent-soft);color:var(--purple)"><span class=material-symbols-outlined>edit_note</span> SZKIC</span>'
            row_bg = 'var(--accent-soft)'
            checkbox_disabled = ''
            checkbox_checked = 'checked'
            wybrane_count += 1
        else:
            status_badge = '<span class="badge" style="background:rgba(100,116,139,0.1);color:var(--text-muted)">● NOWY</span>'
            row_bg = 'rgba(100,116,139,0.05)'
            checkbox_disabled = ''
            checkbox_checked = ''

        # Cena jednostkowa zakupu - z palety (cena_zakupu / ilosc_sztuk), fallback na cena_brutto/ilosc
        paleta_ilosc_szt_w = 0
        try:
            paleta_ilosc_szt_w = paleta['ilosc_sztuk'] or 0
        except:
            pass
        paleta_cena_zak_w = paleta['cena_zakupu'] or 0
        if paleta_cena_zak_w > 0 and paleta_ilosc_szt_w > 0:
            brutto_szt_w = paleta_cena_zak_w / paleta_ilosc_szt_w
            netto_szt_w = round(brutto_szt_w / 1.23, 2)
            ceny_tekst = f"Za szt: {netto_szt_w:.2f} zł netto / {brutto_szt_w:.2f} zł brutto (z palety)"
        else:
            ilosc_produktu = p['ilosc'] if p['ilosc'] > 0 else 1
            brutto_total = p['cena_brutto'] if p['cena_brutto'] > 0 else 0
            netto_total = p['cena_netto'] if p['cena_netto'] > 0 else 0
            brutto_szt_w = brutto_total if brutto_total > 0 else 0  # już jednostkowa, nie dzielić!
            netto_szt_w = netto_total if netto_total > 0 else 0
            if netto_total > 0 and brutto_total > 0:
                ceny_tekst = f"Za szt: {netto_szt_w:.2f} zł netto / {brutto_szt_w:.2f} zł brutto"
            elif netto_total > 0:
                ceny_tekst = f"Za szt: {netto_szt_w:.2f} zł netto"
            else:
                ceny_tekst = f"Za szt: {brutto_szt_w:.2f} zł brutto"

        img_html = ''
        _img_url = p['zdjecie_url'] or ''
        if _img_url and _img_url.startswith('/static/downloads/'):
            import os as _os
            if not _os.path.exists(_img_url.lstrip('/')):
                _asin_c = (p['asin'] or '').strip().upper()
                _img_url = f'https://m.media-amazon.com/images/I/{_asin_c}._AC_SL1500_.jpg' if _asin_c and len(_asin_c) >= 10 else ''
        if _img_url:
            img_html = f'<img src="{_img_url}" style="width:50px;height:50px;object-fit:contain;border-radius:8px;background:#fff;margin-right:10px" onerror="this.style.display=\'none\'" loading="lazy">'

        cena_input = f'''
        <input type="number"
               class="price-input"
               data-product-id="{p['id']}"
               value="{p['cena_allegro']:.0f}"
               min="1"
               step="1"
               style="width:90px;padding:10px 8px;background:var(--bg);border:2px solid var(--border);border-radius:10px;color:var(--text);text-align:center;font-weight:700;font-size:1rem;min-height:42px"
               {checkbox_disabled}>
        '''

        # Meta title bar
        has_meta = 'meta_title' in p.keys() and p['meta_title']
        meta_bar = ''
        if has_meta:
            meta_bar = f'''<div class="bl-meta-bar bl-meta-ok">
                <span class="material-symbols-outlined" style="font-size:1rem;color:#beee00">edit_note</span>
                <span class="bl-meta-text">{str(p["meta_title"])[:80]}</span>
                <button onclick="regenerateMetaTitle({p["id"]}, this)" class="bl-meta-btn bl-meta-btn-regen">
                    <span class="material-symbols-outlined" style="font-size:.9rem">sync</span> Regeneruj
                </button>
            </div>'''
        else:
            meta_bar = f'''<div class="bl-meta-bar bl-meta-missing">
                <span class="material-symbols-outlined" style="font-size:1rem;color:#ff6b9b">warning</span>
                <span class="bl-meta-text" style="color:#ff6b9b">Brak META TITLE</span>
                <button onclick="regenerateMetaTitle({p["id"]}, this)" class="bl-meta-btn bl-meta-btn-gen">
                    <span class="material-symbols-outlined" style="font-size:.9rem">auto_awesome</span> Generuj
                </button>
            </div>'''

        # SKU code
        sku_code = p['asin'] if p['asin'] and p['asin'].upper() not in ('N/A','NAN','NONE') else (p['ean'] or '—')

        # Stock color
        stock_val = p['ilosc'] or 0
        if stock_val >= 10:
            stock_class = 'bl-stock-high'
        elif stock_val >= 5:
            stock_class = 'bl-stock-med'
        else:
            stock_class = 'bl-stock-low'

        produkty_html += f'''
        <tr class="bl-row" data-status="{p['status']}">
            <td class="bl-td-check">
                <label class="bl-checkbox-wrap">
                    <input type="checkbox" class="product-checkbox bl-checkbox" data-product-id="{p['id']}" value="{p['id']}" {checkbox_checked} {checkbox_disabled}>
                    <span class="bl-checkmark"></span>
                </label>
            </td>
            <td class="bl-td-img">
                {f'<img src="{_img_url}" class="bl-thumb" onerror="this.style.display=&apos;none&apos;" loading="lazy">' if _img_url else '<div class="bl-thumb-empty"><span class="material-symbols-outlined">image</span></div>'}
            </td>
            <td class="bl-td-name">
                <div class="bl-name">{p['nazwa'][:60]}</div>
                <div class="bl-category">{p['lokalizacja'] or '—'}</div>
                {meta_bar}
            </td>
            <td class="bl-td-sku"><code class="bl-sku">{sku_code}</code></td>
            <td class="bl-td-stock"><span class="bl-stock {stock_class}" data-qty-id="{p['id']}">{stock_val}</span></td>
            <td class="bl-td-price">
                {cena_input}
                <div class="bl-cost-info">{ceny_tekst}</div>
            </td>
            <td class="bl-td-status">{status_badge}</td>
        </tr>
        '''

    content = f'''
    <style>
    /* ═══ BULK LISTER — Cyberpunk Neon HUD ═══ */
    .bl-header{{text-align:center;margin-bottom:28px;position:relative}}
    .bl-header::after{{content:'';position:absolute;bottom:-8px;left:50%;transform:translateX(-50%);width:120px;height:2px;background:linear-gradient(90deg,transparent,#8ff5ff,transparent)}}
    .bl-title{{font-family:'Space Grotesk',sans-serif;font-size:1.8rem;font-weight:800;background:linear-gradient(135deg,#8ff5ff,#beee00);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin:0;letter-spacing:-0.5px}}
    .bl-subtitle{{font-family:'Manrope',sans-serif;font-size:0.85rem;color:rgba(143,245,255,0.6);margin-top:6px;letter-spacing:1.5px;text-transform:uppercase}}
    .bl-pallet-name{{font-family:'Space Grotesk',sans-serif;font-size:0.95rem;color:#ff6b9b;margin-top:4px}}

    .bl-stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:20px}}
    .bl-stat{{background:rgba(22,26,33,0.6);backdrop-filter:blur(12px);border:1px solid rgba(143,245,255,0.1);border-radius:14px;padding:16px 12px;text-align:center;transition:all 0.3s}}
    .bl-stat:hover{{border-color:rgba(143,245,255,0.3);box-shadow:0 0 20px rgba(143,245,255,0.08)}}
    .bl-stat-num{{font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800}}
    .bl-stat-label{{font-family:'Manrope',sans-serif;font-size:0.68rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:1px;margin-top:4px;font-weight:600}}

    .bl-panel{{background:rgba(22,26,33,0.5);backdrop-filter:blur(16px);border:1px solid rgba(143,245,255,0.1);border-radius:16px;overflow:hidden;margin-bottom:20px}}
    .bl-panel-header{{display:flex;align-items:center;justify-content:space-between;padding:16px 20px;border-bottom:1px solid rgba(143,245,255,0.08);background:rgba(143,245,255,0.03)}}
    .bl-panel-title{{font-family:'Space Grotesk',sans-serif;font-size:1rem;color:#8ff5ff;font-weight:700;display:flex;align-items:center;gap:8px}}
    .bl-select-all-btn{{display:flex;align-items:center;gap:6px;padding:8px 16px;background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);border-radius:8px;color:#8ff5ff;font-family:'Space Grotesk',sans-serif;font-size:0.8rem;font-weight:600;cursor:pointer;transition:all 0.2s;text-transform:uppercase;letter-spacing:0.5px}}
    .bl-select-all-btn:hover{{background:rgba(143,245,255,0.15);border-color:rgba(143,245,255,0.4)}}

    .bl-table{{width:100%;border-collapse:separate;border-spacing:0}}
    .bl-table thead th{{font-family:'Space Grotesk',sans-serif;font-size:0.7rem;color:rgba(143,245,255,0.5);text-transform:uppercase;letter-spacing:1.2px;padding:12px 14px;text-align:left;font-weight:700;border-bottom:1px solid rgba(143,245,255,0.08);white-space:nowrap}}
    .bl-row{{transition:all 0.2s}}
    .bl-row:hover{{background:rgba(143,245,255,0.04)}}
    .bl-row[data-status="wystawiony"]{{background:rgba(34,197,94,0.06)}}
    .bl-row td{{padding:12px 14px;border-bottom:1px solid rgba(255,255,255,0.04);vertical-align:middle}}

    .bl-checkbox-wrap{{display:flex;align-items:center;justify-content:center;cursor:pointer;position:relative;width:22px;height:22px}}
    .bl-checkbox{{opacity:0;position:absolute;cursor:pointer;width:22px;height:22px}}
    .bl-checkmark{{width:22px;height:22px;border:2px solid rgba(143,245,255,0.3);border-radius:6px;transition:all 0.2s;display:flex;align-items:center;justify-content:center;background:rgba(22,26,33,0.8)}}
    .bl-checkbox:checked+.bl-checkmark{{background:linear-gradient(135deg,#8ff5ff,#beee00);border-color:#8ff5ff}}
    .bl-checkbox:checked+.bl-checkmark::after{{content:'\\2713';color:#0a0e14;font-size:14px;font-weight:700}}
    .bl-checkbox:disabled+.bl-checkmark{{opacity:0.3;cursor:not-allowed}}

    .bl-thumb{{width:48px;height:48px;object-fit:contain;border-radius:10px;background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.08)}}
    .bl-thumb-empty{{width:48px;height:48px;border-radius:10px;background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;justify-content:center;color:var(--text-muted);font-size:1.2rem}}

    .bl-name{{font-family:'Manrope',sans-serif;font-weight:700;font-size:0.9rem;color:var(--text);line-height:1.3;margin-bottom:2px}}
    .bl-category{{font-family:'Manrope',sans-serif;font-size:0.75rem;color:var(--text-muted)}}
    .bl-sku{{font-family:'JetBrains Mono','Fira Code',monospace;font-size:0.75rem;color:rgba(143,245,255,0.7);background:rgba(143,245,255,0.06);padding:3px 8px;border-radius:6px;border:1px solid rgba(143,245,255,0.1)}}

    .bl-stock{{font-family:'Space Grotesk',sans-serif;font-weight:800;font-size:1rem;padding:4px 12px;border-radius:8px;display:inline-block;min-width:32px;text-align:center}}
    .bl-stock-high{{color:#22c55e;background:rgba(34,197,94,0.1);border:1px solid rgba(34,197,94,0.2)}}
    .bl-stock-med{{color:#eab308;background:rgba(234,179,8,0.1);border:1px solid rgba(234,179,8,0.2)}}
    .bl-stock-low{{color:#ff6b9b;background:rgba(255,107,155,0.1);border:1px solid rgba(255,107,155,0.2)}}

    .bl-td-price .price-input{{width:90px;padding:10px 8px;background:rgba(22,26,33,0.8);border:2px solid rgba(143,245,255,0.15);border-radius:10px;color:#beee00;text-align:center;font-weight:700;font-size:1rem;font-family:'Space Grotesk',sans-serif;min-height:42px;transition:all 0.3s}}
    .bl-td-price .price-input:focus{{border-color:#8ff5ff;box-shadow:0 0 12px rgba(143,245,255,0.2);outline:none}}
    .bl-cost-info{{font-family:'Manrope',sans-serif;font-size:0.65rem;color:var(--text-muted);margin-top:4px;max-width:140px}}

    .bl-meta-bar{{display:flex;align-items:center;gap:6px;margin-top:6px;padding:5px 8px;border-radius:6px;font-size:0.78rem}}
    .bl-meta-ok{{background:rgba(190,238,0,0.06);border:1px solid rgba(190,238,0,0.12)}}
    .bl-meta-missing{{background:rgba(255,107,155,0.06);border:1px solid rgba(255,107,155,0.12)}}
    .bl-meta-text{{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:rgba(190,238,0,0.8);font-family:'Manrope',sans-serif}}
    .bl-meta-btn{{padding:4px 10px;border:none;border-radius:6px;color:#fff;font-size:0.72rem;cursor:pointer;white-space:nowrap;font-family:'Space Grotesk',sans-serif;font-weight:600;display:flex;align-items:center;gap:4px;transition:all 0.2s}}
    .bl-meta-btn:hover{{transform:translateY(-1px)}}
    .bl-meta-btn-regen{{background:rgba(139,92,246,0.8)}}
    .bl-meta-btn-gen{{background:rgba(34,197,94,0.8)}}

    .bl-bottom{{position:fixed;bottom:0;left:0;right:0;background:rgba(10,14,20,0.92);backdrop-filter:blur(20px);border-top:1px solid rgba(143,245,255,0.12);padding:14px 16px;z-index:100;box-shadow:0 -8px 32px rgba(0,0,0,0.4)}}
    .bl-bottom-inner{{display:flex;align-items:center;gap:10px;max-width:1800px;margin:0 auto;flex-wrap:wrap}}
    .bl-btn{{padding:12px 20px;font-family:'Space Grotesk',sans-serif;font-size:0.9rem;font-weight:700;border:none;border-radius:10px;color:#fff;cursor:pointer;min-height:48px;display:flex;align-items:center;justify-content:center;gap:8px;transition:all 0.25s;text-decoration:none;white-space:nowrap}}
    .bl-btn:hover{{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,0.3)}}
    .bl-btn-back{{background:rgba(255,255,255,0.06);color:var(--text);border:1px solid rgba(255,255,255,0.1)}}
    .bl-btn-meta{{background:linear-gradient(135deg,#8b5cf6,#a855f7);box-shadow:0 4px 15px rgba(139,92,246,0.3)}}
    .bl-btn-meta:hover{{box-shadow:0 6px 24px rgba(139,92,246,0.5)}}
    .bl-btn-wystaw{{background:linear-gradient(135deg,#beee00,#22c55e);color:#0a0e14;box-shadow:0 4px 15px rgba(190,238,0,0.3)}}
    .bl-btn-wystaw:hover{{box-shadow:0 6px 24px rgba(190,238,0,0.5)}}
    .bl-bottom-spacer{{flex:1}}
    .bl-bottom-count{{font-family:'Space Grotesk',sans-serif;font-size:0.85rem;color:rgba(143,245,255,0.7);font-weight:600}}

    .bl-info{{font-family:'Manrope',sans-serif;font-size:0.8rem;margin-bottom:16px;padding:12px 16px;background:rgba(190,238,0,0.04);border:1px solid rgba(190,238,0,0.1);border-radius:10px;color:var(--text-muted);display:flex;align-items:center;gap:10px}}
    .bl-info .material-symbols-outlined{{color:#beee00;font-size:1.1rem}}

    /* Responsive */
    @media(max-width:768px){{
        .bl-stats{{grid-template-columns:repeat(2,1fr)}}
        .bl-table thead{{display:none}}
        .bl-row{{display:flex;flex-wrap:wrap;gap:8px;padding:14px!important;border-bottom:1px solid rgba(255,255,255,0.04)}}
        .bl-row td{{border:none;padding:4px 0}}
        .bl-td-check{{order:1}}
        .bl-td-img{{order:2}}
        .bl-td-name{{order:3;flex:1;min-width:calc(100% - 100px)}}
        .bl-td-sku{{order:5}}
        .bl-td-stock{{order:6}}
        .bl-td-price{{order:7}}
        .bl-td-status{{order:4}}
        .bl-bottom-inner{{flex-direction:column}}
        .bl-btn{{width:100%}}
        .bl-bottom-spacer{{display:none}}
    }}
    @media(max-width:480px){{
        .bl-title{{font-size:1.4rem}}
        .bl-stats{{gap:6px}}
        .bl-stat{{padding:12px 8px}}
        .bl-stat-num{{font-size:1.2rem}}
    }}
    </style>

    <div class="bl-header">
        <h1 class="bl-title"><span class="material-symbols-outlined" style="font-size:1.6rem;vertical-align:middle;-webkit-text-fill-color:#8ff5ff">deployed_code</span> Bulk Lister / Universal</h1>
        <div class="bl-subtitle">Select items for Allegro deployment</div>
        <div class="bl-pallet-name">{paleta['nazwa'] or f"Paleta #{paleta_id}"}</div>
    </div>

    <div class="bl-stats">
        <div class="bl-stat">
            <div class="bl-stat-num" style="color:#8ff5ff">{stats['total']}</div>
            <div class="bl-stat-label">Wszystkich</div>
        </div>
        <div class="bl-stat">
            <div class="bl-stat-num" style="color:#22c55e">{stats['wystawione']}</div>
            <div class="bl-stat-label">Wystawione</div>
        </div>
        <div class="bl-stat">
            <div class="bl-stat-num" style="color:#beee00" id="count-selected">{wybrane_count}</div>
            <div class="bl-stat-label">Zaznaczone</div>
        </div>
        <div class="bl-stat">
            <div class="bl-stat-num" style="color:#22c55e" id="value-total">{stats['wartosc_total']:.0f} zł</div>
            <div class="bl-stat-label">Wartość</div>
        </div>
    </div>

    <div class="bl-info">
        <span class="material-symbols-outlined">lightbulb</span>
        Zaznacz produkty → edytuj ceny → kliknij <b>Wystaw</b>. Wystawione (zielone) nie można zaznaczyć.
    </div>

    <div class="bl-panel" style="margin-bottom:160px">
        <div class="bl-panel-header">
            <div class="bl-panel-title">
                <span class="material-symbols-outlined" style="font-size:1.1rem">inventory_2</span>
                Produkty z palety
            </div>
            <button id="btn-select-all" class="bl-select-all-btn" onclick="toggleSelectAll()">
                <span class="material-symbols-outlined" style="font-size:1rem">select_check_box</span> Zaznacz wszystkie
            </button>
        </div>
        <div style="overflow-x:auto">
            <table class="bl-table">
                <thead>
                    <tr>
                        <th style="width:50px">
                            <label class="bl-checkbox-wrap">
                                <input type="checkbox" id="header-select-all" class="bl-checkbox" onchange="toggleSelectAll()">
                                <span class="bl-checkmark"></span>
                            </label>
                        </th>
                        <th style="width:60px">Zdjęcie</th>
                        <th>Produkt</th>
                        <th>SKU / ASIN</th>
                        <th style="width:80px">Stan</th>
                        <th style="width:160px">Cena</th>
                        <th style="width:120px">Status</th>
                    </tr>
                </thead>
                <tbody id="products-list">
                    {produkty_html}
                </tbody>
            </table>
        </div>
    </div>

    <div class="bl-bottom">
        <div class="bl-bottom-inner">
            <a href="/palety/{paleta_id}" class="bl-btn bl-btn-back">
                <span class="material-symbols-outlined" style="font-size:1.1rem">arrow_back</span> Powrót
            </a>
            <div class="bl-bottom-spacer"></div>
            <div class="bl-bottom-count"><span id="count-bottom">{wybrane_count}</span> zaznaczonych</div>
            <button id="btn-batch-meta" class="bl-btn bl-btn-meta" onclick="batchGenerateMetaTitles()">
                <span class="material-symbols-outlined" style="font-size:1.1rem">auto_awesome</span> Generuj META (<span id="count-meta-btn">{wybrane_count}</span>)
            </button>
            <button id="btn-wystaw" class="bl-btn bl-btn-wystaw" onclick="wystawZaznaczone()">
                <span class="material-symbols-outlined" style="font-size:1.1rem">rocket_launch</span> Wystaw (<span id="count-btn">{wybrane_count}</span>)
            </button>
        </div>
    </div>
    '''

    mass_edit_js = f'''
    function updateCounter() {{
        const checkboxes = document.querySelectorAll('.product-checkbox:checked:not(:disabled)');
        const count = checkboxes.length;
        document.getElementById('count-selected').textContent = count;
        document.getElementById('count-btn').textContent = count;
        document.getElementById('count-meta-btn').textContent = count;
        const countBottom = document.getElementById('count-bottom');
        if (countBottom) countBottom.textContent = count;
        document.getElementById('btn-wystaw').disabled = count === 0;
        document.getElementById('btn-batch-meta').disabled = count === 0;

        // Update header select-all checkbox state
        const allCbs = document.querySelectorAll('.product-checkbox:not(:disabled)');
        const headerCb = document.getElementById('header-select-all');
        if (headerCb && allCbs.length > 0) {{
            headerCb.checked = Array.from(allCbs).every(cb => cb.checked);
        }}

        let total = 0;
        checkboxes.forEach(cb => {{
            const productId = cb.dataset.productId;
            const priceInput = document.querySelector('.price-input[data-product-id="' + productId + '"]');
            const qtyEl = document.querySelector('[data-qty-id="' + productId + '"]');
            const qty = qtyEl ? (parseInt(qtyEl.textContent) || 1) : 1;
            if (priceInput) {{
                total += (parseFloat(priceInput.value) || 0) * qty;
            }}
        }});
        document.getElementById('value-total').textContent = total.toFixed(0) + ' zł';
    }}

    const priceInputs = document.querySelectorAll('.price-input');
    priceInputs.forEach(input => {{
        let timeout;
        input.addEventListener('input', function() {{
            clearTimeout(timeout);
            const productId = this.dataset.productId;
            const newPrice = parseFloat(this.value) || 0;
            this.style.borderColor = '#beee00';

            timeout = setTimeout(() => {{
                fetch('/palety/api/update-price', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{
                        product_id: productId,
                        price: newPrice
                    }})
                }})
                .then(r => r.json())
                .then(data => {{
                    if (data.success) {{
                        input.style.borderColor = '#22c55e';
                        setTimeout(() => input.style.borderColor = '', 1000);
                        updateCounter();
                    }} else {{
                        input.style.borderColor = '#ff6b9b';
                        alert('Błąd zapisu: ' + data.error);
                    }}
                }})
                .catch(err => {{
                    input.style.borderColor = '#ff6b9b';
                    console.error('Error:', err);
                }});
            }}, 800);
        }});
    }});

    const checkboxes = document.querySelectorAll('.product-checkbox');
    checkboxes.forEach(cb => {{
        cb.addEventListener('change', updateCounter);
    }});

    function wystawZaznaczone() {{
        const checked = document.querySelectorAll('.product-checkbox:checked:not(:disabled)');
        if (checked.length === 0) {{
            alert('Zaznacz przynajmniej 1 produkt!');
            return;
        }}
        const productIds = Array.from(checked).map(cb => cb.value);
        window.location.href = '/paletomat/generator/mass-create-from-paleta?paleta_id={paleta_id}&ids=' + productIds.join(',');
    }}

    function batchGenerateMetaTitles() {{
        const checked = document.querySelectorAll('.product-checkbox:checked:not(:disabled)');
        if (checked.length === 0) {{
            alert('Zaznacz przynajmniej 1 produkt!');
            return;
        }}

        const MAX_BATCH = 100;
        if (checked.length > MAX_BATCH) {{
            alert('Zbyt dużo produktów!\\n\\nZaznaczono: ' + checked.length + '\\nMax: ' + MAX_BATCH + '\\n\\nZaznacz mniej produktów lub podziel na mniejsze batche.');
            return;
        }}

        const estimatedTime = checked.length * 5;
        const minutes = Math.floor(estimatedTime / 60);
        const seconds = estimatedTime % 60;
        const timeStr = minutes > 0 ? minutes + 'min ' + seconds + 's' : seconds + 's';

        if (!confirm('Wygenerować META TITLE dla ' + checked.length + ' produktów?\\n\\nSzacowany czas: ~' + timeStr + '\\n5s opóźnienie między produktami (safe rate limiting)\\n\\nKontynuować?')) {{
            return;
        }}

        const productIds = Array.from(checked).map(cb => cb.value);
        const button = document.getElementById('btn-batch-meta');
        const originalText = button.innerHTML;

        button.disabled = true;
        button.innerHTML = '<span class="material-symbols-outlined" style="font-size:1.1rem">hourglass_top</span> Generuję 0/' + productIds.length + '...';

        fetch('/api/generate_meta_title_batch', {{
            method: 'POST',
            mode: 'cors',
            credentials: 'same-origin',
            headers: {{
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }},
            body: JSON.stringify({{ product_ids: productIds }})
        }})
        .then(res => res.json())
        .then(data => {{
            if (data.success) {{
                const quotaErrors = data.details ? data.details.filter(d => d.error && d.error.includes('Quota')).length : 0;
                let msg = 'Gotowe!\\n\\nWygenerowano: ' + data.generated + '\\nBłędy: ' + data.failed;
                if (quotaErrors > 0) {{
                    msg += '\\n\\nQuota exceeded! Poczekaj do jutra (reset o 9:00 AM) lub upgrade do paid tier.';
                }}
                alert(msg);
                location.reload();
            }} else {{
                let errorMsg = data.error || 'Nieznany błąd';
                if (errorMsg.includes('Zbyt dużo')) {{
                    errorMsg += '\\n\\nTIP: Zaznacz max 10 produktów lub poczekaj do jutra na reset quota.';
                }}
                alert('Błąd:\\n\\n' + errorMsg);
                button.disabled = false;
                button.innerHTML = originalText;
            }}
        }})
        .catch(err => {{
            alert('Błąd połączenia:\\n\\n' + err + '\\n\\nSprawdź console (F12) dla szczegółów.');
            button.disabled = false;
            button.innerHTML = originalText;
        }});
    }}

    function toggleSelectAll() {{
        const checkboxes = document.querySelectorAll('.product-checkbox:not(:disabled)');
        const allChecked = Array.from(checkboxes).every(cb => cb.checked);
        checkboxes.forEach(cb => {{ cb.checked = !allChecked; }});
        const btn = document.getElementById('btn-select-all');
        btn.innerHTML = allChecked
            ? '<span class="material-symbols-outlined" style="font-size:1rem">select_check_box</span> Zaznacz wszystkie'
            : '<span class="material-symbols-outlined" style="font-size:1rem">check_box_outline_blank</span> Odznacz wszystkie';
        updateCounter();
    }}

    function regenerateMetaTitle(productId, button) {{
        const originalText = button.innerHTML;
        button.disabled = true;
        button.innerHTML = '<span class="material-symbols-outlined" style="font-size:.9rem">hourglass_top</span> Generuję...';

        fetch('/produkty/' + productId + '/regenerate-meta-title', {{
            method: 'POST',
            mode: 'cors',
            credentials: 'same-origin',
            headers: {{
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }}
        }})
        .then(res => res.json())
        .then(data => {{
            if (data.success) {{
                location.reload();
            }} else {{
                alert('Błąd: ' + (data.error || 'Nieznany błąd'));
                button.disabled = false;
                button.innerHTML = originalText;
            }}
        }})
        .catch(err => {{
            alert('Błąd połączenia: ' + err);
            button.disabled = false;
            button.innerHTML = originalText;
        }});
    }}

    updateCounter();
    '''

    return render(content, 'Masowa edycja', extra_js=mass_edit_js)

@palety_bp.route('/palety/api/update-price', methods=['POST'])
def api_update_price():
    """API do aktualizacji ceny produktu"""
    from modules.database import get_db

    try:
        data = request.get_json()
        product_id = data.get('product_id')
        new_price = float(data.get('price', 0))

        if not product_id or new_price < 0:
            return jsonify({'success': False, 'error': 'Nieprawidłowe dane'})

        conn = get_db()

        # Pobierz starą cenę do historii
        old_product = conn.execute('SELECT cena_allegro, paleta_id FROM produkty WHERE id = ?', (product_id,)).fetchone()
        old_price = old_product['cena_allegro'] if old_product else 0

        conn.execute('UPDATE produkty SET cena_allegro = ? WHERE id = ?', (new_price, product_id))

        # Dodaj do historii jeśli cena się zmieniła
        if old_price != new_price:
            from modules.database import add_historia
            add_historia(product_id, 'zmiana_ceny', f'Zmiana ceny Allegro: {old_price:.0f} → {new_price:.0f} zł',
                {'stara_cena': old_price, 'nowa_cena': new_price})

        product = conn.execute('SELECT paleta_id FROM produkty WHERE id = ?', (product_id,)).fetchone()

        if product and product['paleta_id']:
            stats = conn.execute('''
                SELECT COALESCE(SUM(cena_allegro * ilosc), 0) as total
                FROM produkty WHERE paleta_id = ?
            ''', (product['paleta_id'],)).fetchone()
            conn.commit()
            return jsonify({'success': True, 'new_total': float(stats['total'])})

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        print(f"Error updating price: {e}")
        return jsonify({'success': False, 'error': str(e)})

@palety_bp.route('/palety/<int:paleta_id>')
def paleta_szczegoly(paleta_id):
    """Widok szczegółów palety z kolorami statusów"""
    from modules.database import get_db

    conn = get_db()
    paleta = conn.execute('SELECT * FROM palety WHERE id = ?', (paleta_id,)).fetchone()
    produkty = conn.execute('SELECT * FROM produkty WHERE paleta_id = ? ORDER BY data_dodania DESC', (paleta_id,)).fetchall()

    # MIGRACJA: przenieś przychod_offline z produkty -> sprzedaze (PRZED obliczaniem stats)
    try:
        from datetime import datetime as _dtm
        stare_offline = conn.execute("""
            SELECT p.id, p.nazwa, p.przychod_offline, p.sprzedano_offline, pal.data_zakupu
            FROM produkty p LEFT JOIN palety pal ON pal.id = p.paleta_id
            WHERE p.paleta_id = ? AND p.sprzedano_offline > 0 AND p.przychod_offline > 0
              AND NOT EXISTS (SELECT 1 FROM sprzedaze s WHERE s.produkt_id=p.id AND s.kupujacy='offline' AND s.cena>0)
        """, (paleta_id,)).fetchall()
        for row in stare_offline:
            data = row['data_zakupu'] or _dtm.now().strftime('%Y-%m-%dT%H:%M:%S')
            cena_szt = round(row['przychod_offline'] / max(row['sprzedano_offline'], 1), 2)
            conn.execute("DELETE FROM sprzedaze WHERE produkt_id=? AND kupujacy='offline' AND cena=0", (row['id'],))
            conn.execute("INSERT INTO sprzedaze (produkt_id,nazwa,cena,ilosc,status,data_sprzedazy,kupujacy,notified) VALUES (?,?,?,?,'sprzedana',?,'offline',1)",
                (row['id'], row['nazwa'] or f'Produkt #{row["id"]}', cena_szt, row['sprzedano_offline'], data))
            conn.execute("UPDATE produkty SET przychod_offline=0 WHERE id=?", (row['id'],))
        if stare_offline:
            conn.commit()
            print(f"[CHECK_CIRCLE] Migracja palety {paleta_id}: {len(stare_offline)} offline -> sprzedaze")
    except Exception as _em:
        print(f"[WARNING] Migracja palety: {_em}")

    # Zeruj przychod_offline I sprzedano_offline dla produktów które mają już rekord w sprzedaze (cleanup)
    # FIX: zeruj OBA pola — wcześniej tylko przychod_offline, co powodowało mismatch (+1 sprzedanych)
    try:
        conn.execute('''
            UPDATE produkty SET przychod_offline = 0, sprzedano_offline = 0
            WHERE paleta_id = ? AND (przychod_offline > 0 OR sprzedano_offline > 0)
              AND EXISTS (
                  SELECT 1 FROM sprzedaze s
                  WHERE s.produkt_id = produkty.id AND s.kupujacy = 'offline' AND s.cena > 0
              )
        ''', (paleta_id,))
        conn.commit()
    except:
        pass

    # Sprawdź czy kolumny offline istnieją
    has_offline_columns = False
    try:
        conn.execute("SELECT sprzedano_offline, przychod_offline FROM produkty LIMIT 1")
        has_offline_columns = True
    except:
        pass

    if has_offline_columns:
        stats = conn.execute('''
            SELECT COUNT(*) as cnt,
                   COALESCE(SUM(ilosc), 0) as sztuki,
                   COALESCE(SUM(CASE WHEN status IN ('wystawiony', 'szkic') THEN cena_allegro * ilosc ELSE 0 END), 0) as wartosc,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' AND (sprzedano_offline IS NULL OR sprzedano_offline = 0) THEN cena_allegro ELSE 0 END), 0) as sprzedano_wartosc,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' AND (sprzedano_offline IS NULL OR sprzedano_offline = 0) THEN 1 ELSE 0 END), 0) as sprzedane_produkty,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' THEN cena_brutto ELSE 0 END), 0) as sprzedano_koszt,
                   SUM(CASE WHEN status = 'wystawiony' THEN 1 ELSE 0 END) as wystawione,
                   SUM(CASE WHEN status = 'magazyn' THEN 1 ELSE 0 END) as magazyn,
                   SUM(CASE WHEN status = 'sprzedany' THEN 1 ELSE 0 END) as sprzedane_cnt,
                   COALESCE(SUM(cena_brutto), 0) as zakup_brutto_suma,
                   COALESCE(SUM(cena_netto), 0) as zakup_netto_suma,
                   COALESCE(SUM(sprzedano_offline), 0) as sprzedano_offline_suma,
                   COALESCE(SUM(przychod_offline), 0) as przychod_offline_suma
            FROM produkty WHERE paleta_id = ?
        ''', (paleta_id,)).fetchone()
    else:
        stats = conn.execute('''
            SELECT COUNT(*) as cnt,
                   COALESCE(SUM(ilosc), 0) as sztuki,
                   COALESCE(SUM(CASE WHEN status IN ('wystawiony', 'szkic') THEN cena_allegro * ilosc ELSE 0 END), 0) as wartosc,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' THEN cena_allegro ELSE 0 END), 0) as sprzedano_wartosc,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' THEN 1 ELSE 0 END), 0) as sprzedane_produkty,
                   COALESCE(SUM(CASE WHEN status = 'sprzedany' THEN cena_brutto ELSE 0 END), 0) as sprzedano_koszt,
                   SUM(CASE WHEN status = 'wystawiony' THEN 1 ELSE 0 END) as wystawione,
                   SUM(CASE WHEN status = 'magazyn' THEN 1 ELSE 0 END) as magazyn,
                   SUM(CASE WHEN status = 'sprzedany' THEN 1 ELSE 0 END) as sprzedane_cnt,
                   COALESCE(SUM(cena_brutto), 0) as zakup_brutto_suma,
                   COALESCE(SUM(cena_netto), 0) as zakup_netto_suma,
                   0 as sprzedano_offline_suma,
                   0 as przychod_offline_suma
            FROM produkty WHERE paleta_id = ?
        ''', (paleta_id,)).fetchone()

    # Pobierz rzeczywistą sprzedaż z tabeli sprzedaze (dla dokładniejszych danych)
    # LEFT JOIN przez oba ścieżki: bezpośredni produkt_id i przez oferta_id→oferty→produkt_id
    # (Allegro synced orders mają produkt_id=NULL ale mają oferta_id)
    sprzedaz_stats = conn.execute('''
        SELECT COALESCE(SUM(s.cena * s.ilosc), 0) as przychod,
               COALESCE(SUM(s.ilosc), 0) as szt_sprzedanych
        FROM sprzedaze s
        LEFT JOIN produkty pr  ON s.produkt_id = pr.id
        LEFT JOIN oferty o     ON s.oferta_id  = o.id
        LEFT JOIN produkty pr2 ON o.produkt_id = pr2.id
        WHERE COALESCE(pr.paleta_id, pr2.paleta_id) = ?
          AND COALESCE(s.status,'') NOT IN ('anulowana','anulowane','zwrot','')
    ''', (paleta_id,)).fetchone()

    if not paleta:
        return redirect('/palety')

    # Bezpieczne pobieranie ceny netto (kolumna może nie istnieć w starej bazie)
    cena_zakupu_netto = 0
    try:
        kolumny = [desc[0] for desc in conn.execute('PRAGMA table_info(palety)').fetchall()]
        if 'cena_zakupu_netto' in kolumny:
            val = conn.execute('SELECT cena_zakupu_netto FROM palety WHERE id = ?', (paleta_id,)).fetchone()
            cena_zakupu_netto = val[0] if val and val[0] else 0
    except:
        pass

    # AUTO-NAPRAWA: Jeśli cena_zakupu = 0, zapisz aktualną sumę (jednorazowo!)
    cena_zakupu = paleta['cena_zakupu'] or 0
    if cena_zakupu == 0:
        suma_netto = stats['zakup_netto_suma'] or 0
        suma_brutto = round(suma_netto * 1.23, 2)

        if suma_netto > 0:
            kolumny = [desc[0] for desc in conn.execute('PRAGMA table_info(palety)').fetchall()]
            if 'cena_zakupu_netto' in kolumny:
                conn.execute('''
                    UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ? WHERE id = ?
                ''', (suma_brutto, suma_netto, paleta_id))
            else:
                conn.execute('UPDATE palety SET cena_zakupu = ? WHERE id = ?', (suma_brutto, paleta_id))
            conn.commit()
            cena_zakupu = suma_brutto
            cena_zakupu_netto = suma_netto
            print(f"[PAID] Auto-naprawiono cenę zakupu palety #{paleta_id}: {suma_netto:.2f} netto | {suma_brutto:.2f} brutto")

    # Przychód offline ze sprzedaze (nowe rekordy po migracji)
    przychod_offline_sprzedaze = conn.execute('''
        SELECT COALESCE(SUM(s.cena * s.ilosc), 0)
        FROM sprzedaze s
        LEFT JOIN produkty pr  ON s.produkt_id = pr.id
        LEFT JOIN oferty o     ON s.oferta_id  = o.id
        LEFT JOIN produkty pr2 ON o.produkt_id = pr2.id
        WHERE COALESCE(pr.paleta_id, pr2.paleta_id) = ? AND s.kupujacy = 'offline'
          AND COALESCE(s.status,'') NOT IN ('anulowana','anulowane','zwrot','')
    ''', (paleta_id,)).fetchone()[0] or 0

    # Przychód offline ze starych danych - TYLKO dla produktów bez rekordu w sprzedaze
    przychod_offline_stare = conn.execute('''
        SELECT COALESCE(SUM(przychod_offline), 0)
        FROM produkty
        WHERE paleta_id = ? AND przychod_offline > 0
          AND NOT EXISTS (
              SELECT 1 FROM sprzedaze s
              WHERE s.produkt_id = produkty.id
                AND s.kupujacy = 'offline'
                AND s.cena > 0
          )
    ''', (paleta_id,)).fetchone()[0] or 0

    # Przychód z Allegro (sprzedaze bez offline) - oba ścieżki: bezpośredni + przez oferta_id
    przychod_allegro_db = conn.execute('''
        SELECT COALESCE(SUM(s.cena * s.ilosc), 0)
        FROM sprzedaze s
        LEFT JOIN produkty pr  ON s.produkt_id = pr.id
        LEFT JOIN oferty o     ON s.oferta_id  = o.id
        LEFT JOIN produkty pr2 ON o.produkt_id = pr2.id
        WHERE COALESCE(pr.paleta_id, pr2.paleta_id) = ?
          AND COALESCE(s.status,'') NOT IN ('anulowana','anulowane','zwrot','')
    ''', (paleta_id,)).fetchone()[0] or 0


    # cena_zakupu w bazie = BRUTTO
    koszt_palety_brutto = cena_zakupu
    koszt_palety_netto = cena_zakupu_netto if cena_zakupu_netto > 0 else round(cena_zakupu / 1.23, 2)

    # STAŁY koszt jednostkowy (NETTO/szt)
    try:
        _kj_netto = float(paleta['koszt_jednostkowy'] or 0)
    except:
        _kj_netto = 0
    if _kj_netto == 0 and koszt_palety_netto > 0:
        _total = (stats['sztuki'] or 0) + (sprzedaz_stats['szt_sprzedanych'] or 0)
        if _total > 0:
            _kj_netto = round(koszt_palety_netto / _total, 2)
            try:
                conn.execute('UPDATE palety SET koszt_jednostkowy = ? WHERE id = ?', (_kj_netto, paleta_id))
                conn.commit()
                print(f"[PAID] Auto-set koszt_jednostkowy palety #{paleta_id}: {_kj_netto:.2f} zł/szt netto")
            except:
                pass
    koszt_jednostkowy_netto = _kj_netto
    koszt_jednostkowy_brutto = round(_kj_netto * 1.23, 2) if _kj_netto > 0 else 0

    # Rzeczywiste dane sprzedaży
    sprzedano_szt_db = sprzedaz_stats['szt_sprzedanych'] or 0
    sprzedane_produkty = stats['sprzedane_produkty'] or 0
    try:
        sprzedano_offline = stats['sprzedano_offline_suma'] or 0
    except:
        sprzedano_offline = 0
    try:
        przychod_offline = stats['przychod_offline_suma'] or 0
    except:
        przychod_offline = 0

    print(f"[BAR_CHART] STATS paleta #{paleta_id}:")
    print(f"   - sprzedano_szt_db (tabela sprzedaze): {sprzedano_szt_db}")
    print(f"   - sprzedane_produkty (status=sprzedany bez offline): {sprzedane_produkty}")
    print(f"   - sprzedano_offline (suma): {sprzedano_offline}")
    print(f"   - przychod_offline (suma): {przychod_offline}")

    # Suma wszystkich źródeł sprzedaży
    offline_w_sprzedaze = conn.execute('''
        SELECT COALESCE(SUM(s.ilosc),0) FROM sprzedaze s
        LEFT JOIN produkty pr  ON s.produkt_id=pr.id
        LEFT JOIN oferty o     ON s.oferta_id=o.id
        LEFT JOIN produkty pr2 ON o.produkt_id=pr2.id
        WHERE COALESCE(pr.paleta_id, pr2.paleta_id)=? AND s.kupujacy='offline'
        AND COALESCE(s.status,'') NOT IN ('anulowana','anulowane','zwrot','')
    ''', (paleta_id,)).fetchone()[0] or 0
    offline_bez_sprzedaze = max(0, sprzedano_offline - offline_w_sprzedaze)
    sprzedano_szt = sprzedano_szt_db + offline_bez_sprzedaze

    # FALLBACK: jeśli ilosc_sztuk (oryginalna ilość przy imporcie) jest znana,
    # sprzedanych = ilosc_sztuk - aktualny_stan (obsługuje przypadek gdy "-" nie zapisuje sprzedaze)
    _ilosc_sztuk_orig = 0
    try:
        _ilosc_sztuk_orig = int(paleta['ilosc_sztuk'] or 0)
    except:
        pass
    _ilosc_aktualna = stats['sztuki'] or 0
    if _ilosc_sztuk_orig > 0 and (_ilosc_sztuk_orig - _ilosc_aktualna) > sprzedano_szt:
        # Więcej sprzedanych wg różnicy stanu niż wg sprzedaze — użyj różnicy
        sprzedano_szt = _ilosc_sztuk_orig - _ilosc_aktualna
        print(f"   - FALLBACK: ilosc_sztuk={_ilosc_sztuk_orig} - aktualna={_ilosc_aktualna} = {sprzedano_szt} sprzedanych")
    print(f"   - WYNIK sprzedano_szt = {sprzedano_szt_db} (sprzedaze) + {offline_bez_sprzedaze} (offline bez sprzedaze) = {sprzedano_szt}")

    # Przychód
    przychod_z_produktow = stats['sprzedano_wartosc'] or 0
    przychod_z_sprzedazy = sprzedaz_stats['przychod'] or 0

    przychod_rzeczywisty = przychod_allegro_db + przychod_offline_sprzedaze + przychod_offline_stare
    przychod_z_sprzedazy = przychod_allegro_db + przychod_offline_sprzedaze

    # FALLBACK PRZYCHÓD: jeśli sprzedano więcej niż jest w tabeli sprzedaze (np. kliknięcia "-"),
    # oszacuj brakujący przychód na podstawie średniej ceny allegro produktów
    _tracked_szt = sprzedano_szt_db + offline_bez_sprzedaze
    _untracked_szt = sprzedano_szt - _tracked_szt
    if _untracked_szt > 0:
        _avg_cena = conn.execute(
            'SELECT AVG(cena_allegro) FROM produkty WHERE paleta_id = ? AND cena_allegro > 0',
            (paleta_id,)
        ).fetchone()[0] or 0
        if _avg_cena > 0:
            _extra_przychod = _untracked_szt * _avg_cena
            przychod_rzeczywisty += _extra_przychod
            print(f"   - FALLBACK PRZYCHÓD: {_untracked_szt} nieśledzonych × {_avg_cena:.2f} zł = +{_extra_przychod:.2f} zł")

    print(f"[BAR_CHART] PRZYCHOD: z_produktow={przychod_z_produktow}, z_sprzedazy={przychod_z_sprzedazy}, offline={przychod_offline}, SUMA={przychod_rzeczywisty}")

    # Koszt sprzedanych
    wszystkie_szt = (stats['sztuki'] or 0) + sprzedano_szt
    if koszt_jednostkowy_brutto > 0:
        koszt_sprzedanych = sprzedano_szt * koszt_jednostkowy_brutto
    elif wszystkie_szt > 0 and koszt_palety_brutto > 0:
        koszt_sprzedanych = (sprzedano_szt / wszystkie_szt) * koszt_palety_brutto
    else:
        koszt_sprzedanych = 0

    zysk_rzeczywisty = przychod_rzeczywisty - koszt_sprzedanych

    # DEBUG
    print(f"[INVENTORY_2] PRODUKTY na palecie #{paleta_id}:")
    for p in produkty:
        try:
            offline_szt = p['sprzedano_offline'] or 0
        except:
            offline_szt = 0
        try:
            offline_przychod = p['przychod_offline'] or 0
        except:
            offline_przychod = 0
        nazwa = (p['nazwa'] or '')[:30]
        print(f"   - ID:{p['id']} | {nazwa} | status={p['status']} | ilosc={p['ilosc']} | offline_szt={offline_szt} | offline_przychod={offline_przychod}")

    # Zlicz produkty bez zdjęć (do przycisku scrapuj)
    bez_zdjec = conn.execute('''
        SELECT COUNT(*) FROM produkty
        WHERE paleta_id = ? AND (zdjecie_url IS NULL OR zdjecie_url = '')
        AND asin IS NOT NULL AND asin != '' AND asin != 'nan'
    ''', (paleta_id,)).fetchone()[0]

    produkty_html = ''
    for p in produkty:
        try:
            p_offline_szt = p['sprzedano_offline'] or 0
        except:
            p_offline_szt = 0

        if p['status'] == 'sprzedany':
            status_color = 'var(--green)'
            status_icon = '<span class=material-symbols-outlined>check_circle</span>'
            status_text = 'SPRZEDANY'
        elif p['status'] == 'wystawiony':
            status_color = 'var(--blue)'
            status_icon = '●'
            status_text = 'WYSTAWIONY'
        elif p['status'] == 'magazyn':
            status_color = 'var(--yellow)'
            status_icon = '<span class=material-symbols-outlined>inventory_2</span>'
            status_text = 'MAGAZYN'
        else:
            status_color = 'var(--text-muted)'
            status_icon = '●'
            status_text = 'NOWY'

        # Cena jednostkowa zakupu
        if koszt_jednostkowy_netto > 0:
            netto_szt = koszt_jednostkowy_netto
            brutto_szt = koszt_jednostkowy_brutto
            cena_glowna = f"{netto_szt:.2f} zł/szt netto (stała)"
            cena_dodatkowa = ""
        else:
            brutto_szt = 0
            netto_szt = 0
            cena_glowna = "brak - ustaw w edycji palety"
            cena_dodatkowa = ""

        stan_opcje = ''
        stany = ['Nowy', 'Nowy w otwartym opakowaniu', 'Używany', 'Uszkodzony', 'Odnowiony']
        for s in stany:
            sel = 'selected' if (p['stan'] or 'Nowy') == s else ''
            stan_opcje += f'<option value="{s}" {sel}>{s}</option>'

        status_opcje = ''
        statusy = [('magazyn','📦 Magazyn'),('wystawiony','🛒 Wystawiony'),('sprzedany','💰 Sprzedany'),('uszkodzony','⚠️ Uszkodzony'),('zwrot','↩️ Zwrot')]
        for sv, sl in statusy:
            sel = 'selected' if (p['status'] or 'magazyn') == sv else ''
            status_opcje += f'<option value="{sv}" {sel}>{sl}</option>'

        img_url = p['zdjecie_url'] or ''
        # Fallback: jeśli lokalna ścieżka nie istnieje, użyj Amazon CDN
        if img_url and img_url.startswith('/static/downloads/'):
            import os as _os
            if not _os.path.exists(img_url.lstrip('/')):
                asin_clean = (p['asin'] or '').strip().upper()
                if asin_clean and len(asin_clean) >= 10:
                    img_url = f'https://m.media-amazon.com/images/I/{asin_clean}._AC_SL1500_.jpg'
                else:
                    img_url = ''
        img_html = ''
        if img_url:
            img_html = f'<img src="{img_url}" style="width:50px;height:50px;object-fit:contain;border-radius:8px;background:#fff;flex-shrink:0" onerror="this.style.display=\'none\'" loading="lazy">'

        # EAN: traktuj "N/A", "nan", puste jako brak — pokaż ASIN zamiast
        ean_display = p['ean'] or ''
        if ean_display.upper() in ('N/A', 'NAN', 'NONE', ''):
            ean_display = ''
        identyfikator = ean_display or p['asin'] or '—'

        # PIR status colors
        _sl_clr = {'sprzedany':'#beee00','wystawiony':'#8ff5ff','magazyn':'#eab308'}.get(p['status'] or 'magazyn','#64748b')
        _sl_bg  = {'sprzedany':'rgba(190,238,0,0.08)','wystawiony':'rgba(143,245,255,0.08)','magazyn':'rgba(234,179,8,0.08)'}.get(p['status'] or 'magazyn','rgba(100,116,139,0.06)')

        produkty_html += f'''
        <div style="background:var(--bg-card);border:1px solid var(--border);border-left:3px solid {_sl_clr};padding:12px 12px 10px;margin-bottom:8px;transition:border-color 0.2s" data-produkt-id="{p['id']}" data-ilosc="{p['ilosc']}">
            <!-- TOP ROW: image + name + price -->
            <div style="display:flex;align-items:flex-start;gap:10px;margin-bottom:10px">
                {img_html}
                <div style="flex:1;min-width:0">
                    <a href="/magazyn/produkt/{p['id']}" style="font-size:0.84rem;font-weight:700;color:var(--text);text-decoration:none;display:block;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;line-height:1.3">{p['nazwa'][:50] if p['nazwa'] else '—'}</a>
                    <div style="display:flex;align-items:center;gap:6px;margin-top:4px;flex-wrap:wrap">
                        <span style="font-size:0.62rem;color:#64748b;font-family:monospace">{identyfikator}</span>
                        <span style="font-size:0.62rem;padding:1px 7px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;color:{_sl_clr};background:{_sl_bg}">{status_text}</span>
                        <span style="font-size:0.62rem;color:#64748b">{p['lokalizacja'] or 'regal —'}</span>
                    </div>
                </div>
                <div style="text-align:right;flex-shrink:0;font-family:'Space Grotesk',sans-serif">
                    <div style="font-size:1rem;font-weight:800;color:{_sl_clr}">{int(p['cena_allegro'] or 0)} zł</div>
                    <div style="display:flex;align-items:center;gap:4px;justify-content:flex-end;margin-top:4px">
                        <button onclick="szybkaMinus({p['id']},{p['ilosc']},{int(p['cena_allegro'] or 0)})" style="background:rgba(239,68,68,0.15);border:1px solid rgba(239,68,68,0.3);color:#ef4444;width:22px;height:22px;font-size:0.8rem;cursor:pointer;padding:0;font-weight:700" {'disabled' if (p['ilosc'] or 0) == 0 else ''}>-</button>
                        <span style="font-family:'Space Grotesk',sans-serif;font-weight:800;font-size:0.9rem;min-width:20px;text-align:center" id="ilosc-{p['id']}">{p['ilosc']}</span>
                        <button onclick="szybkaPlus({p['id']},{p['ilosc']})" style="background:rgba(190,238,0,0.12);border:1px solid rgba(190,238,0,0.3);color:#beee00;width:22px;height:22px;font-size:0.8rem;cursor:pointer;padding:0;font-weight:700">+</button>
                        <span style="font-size:0.62rem;color:#64748b">szt</span>
                    </div>
                    <div class="sztuki-dots"></div>
                </div>
            </div>

            <!-- SELECTS ROW -->
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px">
                <div>
                    <div style="font-size:0.58rem;color:#64748b;margin-bottom:3px;text-transform:uppercase;letter-spacing:0.8px;font-weight:600">Stan</div>
                    <select onchange="zapiszPole({p['id']}, 'stan', this.value, this)"
                        style="width:100%;background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.08);color:var(--text);padding:5px 6px;font-size:0.72rem">
                        {stan_opcje}
                    </select>
                </div>
                <div>
                    <div style="font-size:0.58rem;color:#64748b;margin-bottom:3px;text-transform:uppercase;letter-spacing:0.8px;font-weight:600">Status</div>
                    <select onchange="zapiszPole({p['id']}, 'status', this.value, this)"
                        style="width:100%;background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.08);color:var(--text);padding:5px 6px;font-size:0.72rem">
                        {status_opcje}
                    </select>
                </div>
            </div>

            <!-- QUICK ACTIONS -->
            <div style="display:flex;gap:4px">
                <button type="button"
                   onclick="document.getElementById('korektaProduktId').value='{p['id']}';document.getElementById('korektaIlosc').value={p['ilosc'] or 0};document.getElementById('maxIlosc').value={p['ilosc'] or 0};document.getElementById('sprzedajIlosc').value=1;document.getElementById('sprzedajIlosc').max={p['ilosc'] or 0};document.getElementById('sprzedajCena').value='{int(p['cena_allegro'] or p['cena_brutto'] or 0)}';document.getElementById('offlineSzt').value='{int(p_offline_szt)}';var cs=document.getElementById('cofnijOfflineSection');if({int(p_offline_szt)}>0){{cs.style.display='block';document.getElementById('offlineInfo').textContent='{int(p_offline_szt)} szt.';document.getElementById('cofnijIlosc').value=1;document.getElementById('cofnijIlosc').max={int(p_offline_szt)}}}else{{cs.style.display='none'}};document.getElementById('modalKorekta').style.display='block'"
                   style="padding:5px 8px;background:rgba(249,115,22,0.12);border:1px solid rgba(249,115,22,0.25);color:#f97316;font-size:0.62rem;font-weight:700;cursor:pointer;flex:1;text-transform:uppercase;letter-spacing:0.5px">
                    <span class=material-symbols-outlined style="font-size:0.7rem;vertical-align:middle">edit</span> Korekta
                </button>
                <a href="/magazyn/produkt/{p['id']}/edytuj"
                   style="padding:5px 8px;background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);color:#8ff5ff;text-decoration:none;font-size:0.62rem;font-weight:700;text-align:center;flex:1;text-transform:uppercase;letter-spacing:0.5px">
                    <span class=material-symbols-outlined style="font-size:0.7rem;vertical-align:middle">edit_note</span> Edytuj
                </a>
                <button onclick="pokazMenu(event, {p['id']}, {p['ilosc']}, '{p['nazwa'][:30].replace(chr(39), chr(96)).replace(chr(34), chr(96))}', this)"
                   style="padding:5px 8px;background:rgba(100,116,139,0.1);border:1px solid rgba(100,116,139,0.2);color:#94a3b8;font-size:0.62rem;font-weight:700;cursor:pointer;flex:1;text-transform:uppercase;letter-spacing:0.5px">
                    ⋯ Akcje
                </button>
            </div>
        </div>
        '''

    if not produkty_html:
        produkty_html = '<div style="text-align:center;color:var(--text-muted);padding:20px">Brak produktów. Importuj Excel!</div>'

    # ROI
    zysk_potencjalny = stats['wartosc'] - koszt_palety_brutto
    roi = (zysk_potencjalny / koszt_palety_brutto * 100) if koszt_palety_brutto > 0 else 0

    # Bezpieczne pobieranie regału
    try:
        regal_palety = paleta['regal'] if paleta['regal'] else ''
    except (KeyError, TypeError):
        regal_palety = ''

    content = f'''
    <div class="header">
        <h1><span class=material-symbols-outlined>inventory_2</span> {paleta['nazwa'] or f"Paleta #{paleta['id']}"}</h1>
        <small><span class="dostawca-name">{paleta['dostawca']}</span> • {paleta['data_zakupu']}</small>
        {f'<div style="margin-top:6px;font-size:0.85rem;color:var(--purple)"><span class=material-symbols-outlined>pin_drop</span> Regal: {regal_palety}</div>' if regal_palety else ''}
    </div>

    <!-- GŁÓWNE STATYSTYKI SPRZEDAŻY -->
    <div class="sale-banner">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
            <div class="sale-banner-label"><span class=material-symbols-outlined>bar_chart</span> SPRZEDAŻ Z PALETY</div>
            <div style="font-size:1.8rem;font-weight:800;color:#fff">{sprzedano_szt} <span style="font-size:0.9rem;color:#6ee7b7">/ {(stats['sztuki'] or 0) + sprzedano_szt} szt</span></div>
        </div>
        <div class="sale-grid">
            <div class="sale-cell">
                <div class="sale-cell-val" style="color:var(--green)">{przychod_rzeczywisty:.0f} zł</div>
                <div class="sale-cell-lbl">PRZYCHÓD</div>
            </div>
            <div class="sale-cell">
                <div class="sale-cell-val" style="color:var(--red)">-{koszt_sprzedanych:.0f} zł</div>
                <div class="sale-cell-lbl">KOSZT</div>
            </div>
            <div class="sale-cell">
                <div class="sale-cell-val" style="color:{'var(--green)' if zysk_rzeczywisty >= 0 else 'var(--red)'}">{zysk_rzeczywisty:+.0f} zł</div>
                <div class="sale-cell-lbl">ZYSK</div>
            </div>
        </div>
    </div>

    <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:15px">
        <div class="paleta-stats-box">
            <div class="paleta-stats-num" style="color:var(--orange)">{koszt_palety_netto:.0f} zł</div>
            <div class="paleta-stats-label">KOSZT NETTO (STAŁY)</div>
        </div>
        <div class="paleta-stats-box">
            <div class="paleta-stats-num" style="color:var(--red)">{koszt_palety_brutto:.0f} zł</div>
            <div class="paleta-stats-label">KOSZT BRUTTO (STAŁY)</div>
        </div>
        <div class="paleta-stats-box">
            <div class="paleta-stats-num" style="color:var(--green)">{stats['wartosc']:.0f} zł</div>
            <div class="paleta-stats-label">WARTOŚĆ ALLEGRO</div>
        </div>
        <div class="paleta-stats-box">
            <div class="paleta-stats-num" style="color:var(--blue)">{stats['cnt']} <span style="font-size:0.8rem;color:var(--text-muted)">({stats['sztuki']} szt)</span></div>
            <div class="paleta-stats-label">PRODUKTÓW</div>
        </div>
        <div class="paleta-stats-box">
            <div class="paleta-stats-num" style="color:var(--green)">{sprzedano_szt}</div>
            <div class="paleta-stats-label">SPRZEDANYCH</div>
        </div>
    </div>

    <div class="stat-row" style="margin-bottom:15px">
        <div class="stat-box" style="background:var(--green-soft);border-color:rgba(34,197,94,0.3)">
            <div class="stat-val green"><span class=material-symbols-outlined>check_circle</span> {stats['wystawione'] or 0}</div>
            <div class="stat-lbl">WYSTAWIONE</div>
        </div>
        <div class="stat-box" style="background:var(--yellow-soft);border-color:rgba(234,179,8,0.3)">
            <div class="stat-val" style="color:var(--yellow)"><span class=material-symbols-outlined>inventory_2</span> {stats['magazyn'] or 0}</div>
            <div class="stat-lbl">W MAGAZYNIE</div>
        </div>
        <div class="stat-box" style="background:var(--red-soft);border-color:rgba(239,68,68,0.3)">
            <div class="stat-val red"><span class=material-symbols-outlined>bar_chart</span> {roi:.1f}%</div>
            <div class="stat-lbl">ROI</div>
        </div>
    </div>

    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:10px">
        <a href="/palety/{paleta_id}/mass-edit" class="btn btn-purple" style="text-decoration:none"><span class=material-symbols-outlined>edit</span> MASOWE WYSTAWIANIE</a>
        <a href="/magazyn/import?paleta_id={paleta_id}" class="btn" style="background:var(--blue);text-decoration:none"><span class=material-symbols-outlined>download</span> IMPORTUJ EXCEL</a>
        <a href="/palety/{paleta_id}/edit" class="btn btn-warning" style="text-decoration:none"><span class=material-symbols-outlined>settings</span> EDYTUJ PALETE</a>
    </div>
    <button type="button" id="scrape-btn" data-paleta="{paleta_id}" class="btn" style="width:100%;background:linear-gradient(135deg,#8b5cf6,#6d28d9);margin-bottom:15px;font-size:0.85rem;cursor:pointer"><span class=material-symbols-outlined>photo_camera</span> SCRAPUJ ZDJECIA ({bez_zdjec} produktow bez zdjec)</button>
    ''' + '''<script>
    document.getElementById('scrape-btn').onclick = function() {
        var btn = this;
        btn.disabled = true;
        btn.style.opacity = '0.7';
        btn.innerHTML = '[HOURGLASS_TOP] Scrapuje zdjecia...';
        fetch('/palety/' + btn.getAttribute('data-paleta') + '/scrape-images', {method: 'POST'})
        .then(function(r) { return r.json(); })
        .then(function(d) {
            if (d.ok) {
                btn.innerHTML = '<span class=material-symbols-outlined>check_circle</span> Scraping ' + d.count + ' produktow w tle!';
                btn.style.background = '#22c55e';
                btn.style.opacity = '1';
            } else {
                btn.innerHTML = '<span class=material-symbols-outlined>cancel</span> ' + (d.error || 'Blad');
                btn.style.background = '#ef4444';
                btn.style.opacity = '1';
            }
            setTimeout(function() { location.reload(); }, 3000);
        })
        .catch(function(e) {
            btn.innerHTML = '<span class=material-symbols-outlined>cancel</span> ' + e;
            btn.disabled = false;
            btn.style.opacity = '1';
        });
    };
    </script>''' + f'''

    <!-- PRZEKAZ ZYSK NA CEL -->
    ''' + ('''
    <form action="/goal/add-contribution" method="POST" style="background:var(--green-soft);border:1px solid rgba(34,197,94,0.3);border-radius:12px;padding:15px;margin-bottom:15px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
            <div>
                <div style="font-weight:600;color:var(--green);font-size:1.05rem">&#x1F697; Przekaz zysk na Hyundaia i30 N</div>
                <div style="font-size:0.75rem;color:var(--text-muted);margin-top:3px">Potencjalny zysk: ''' + str(int(zysk_potencjalny)) + ''' PLN</div>
            </div>
        </div>
        <div style="display:flex;gap:10px;align-items:center">
            <input type="hidden" name="paleta_id" value="''' + str(paleta_id) + '''">
            <input type="hidden" name="description" value="Zysk z palety ''' + str(paleta['nazwa'] or paleta_id) + '''">
            <input type="number" name="amount" placeholder="Kwota PLN" required min="1" step="1"
                   value="''' + str(max(0, int(zysk_potencjalny))) + '''"
                   class="form-control" style="flex:1">
            <button type="submit" class="btn btn-success" style="width:auto;padding:12px 24px;white-space:nowrap;margin:0">&#x1F4B0; PRZEKAZ</button>
        </div>
    </form>
    ''' if zysk_potencjalny > 0 else '') + '''

    <div class="section-title">PRODUKTY (''' + str(stats['cnt']) + ''')</div>

    ''' + produkty_html + '''

    <form method="POST" action="/palety/''' + str(paleta_id) + '''/delete" style="margin-top:20px" onsubmit="return confirm('<span class=material-symbols-outlined>warning</span> UWAGA!\\n\\nTo usunie tę paletę i wszystkie jej produkty (''' + str(stats['cnt']) + ''' szt.)\\n\\nNa pewno kontynuować?')">
        <input type="hidden" name="csrf_token" value="''' + generate_csrf() + '''">
        <button type="submit" class="btn btn-danger">
            <span class=material-symbols-outlined>delete</span> USUŃ PALETĘ
        </button>
    </form>

    <a href="/palety" style="display:block;text-align:center;color:var(--text-muted);text-decoration:none;margin-top:15px">← Powrót do palet</a>

    <!-- MODAL KOREKTY ILOŚCI -->
    <div id="modalKorekta" onclick="if(event.target===this)this.style.display='none'" class="modal-overlay" style="display:none">
        <div class="modal-box" style="max-width:400px;margin:50px auto">
            <h3 style="margin:0 0 15px"><span class=material-symbols-outlined>edit</span> Korekta produktu</h3>

            <!-- KOREKTA ILOŚCI -->
            <form method="POST" action="/sprzedaze/korekta-ilosci">
                <input type="hidden" name="csrf_token" value="''' + generate_csrf() + '''">
                <input type="hidden" name="produkt_id" id="korektaProduktId" value="">
                <div class="form-group">
                    <label>Zmień ilość na:</label>
                    <input type="number" name="nowa_ilosc" id="korektaIlosc" min="0" class="form-control">
                </div>
                <div style="display:flex;gap:10px;margin-bottom:15px">
                    <button type="submit" class="btn" style="flex:1;background:var(--blue);margin:0"><span class=material-symbols-outlined>save</span> Zapisz ilość</button>
                    <button type="button" onclick="document.getElementById('modalKorekta').style.display='none'" class="btn btn-secondary" style="flex:0;width:auto;padding:12px 16px;margin:0">✕</button>
                </div>
            </form>

            <div style="border-top:1px solid var(--border);padding-top:15px;margin-top:10px">
                <label style="display:block;font-size:0.8rem;color:var(--orange);margin-bottom:8px"><span class=material-symbols-outlined>inventory_2</span> Sprzedaż offline (bez statystyk Allegro):</label>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px">
                    <div>
                        <label style="font-size:0.7rem;color:var(--text-muted)">Ile szt.:</label>
                        <input type="number" id="sprzedajIlosc" min="1" value="1" class="form-control" style="text-align:center;border-color:var(--orange)">
                    </div>
                    <div>
                        <label style="font-size:0.7rem;color:var(--text-muted)">Cena sprzedaży (zł):</label>
                        <input type="number" id="sprzedajCena" min="0.01" step="0.01" required placeholder="Wpisz cenę zł" class="form-control" style="text-align:center;border-color:var(--orange)">
                    </div>
                </div>
                <button onclick="oznaczSprzedany()" class="btn btn-warning" style="margin:0"><span class=material-symbols-outlined>inventory_2</span> Sprzedaj offline</button>
                <div style="font-size:0.65rem;color:var(--text-muted);margin-top:6px;text-align:center">
                    Dolicza do przychodu palety, ale NIE do statystyk sprzedaży Allegro
                </div>
            </div>

            <!-- COFNIJ OFFLINE -->
            <div id="cofnijOfflineSection" style="display:none;border-top:1px solid var(--border);padding-top:15px;margin-top:15px">
                <label style="display:block;font-size:0.8rem;color:var(--red);margin-bottom:8px"><span class=material-symbols-outlined>sync</span> Cofnij sprzedaż offline:</label>
                <div style="display:flex;gap:10px;align-items:center;margin-bottom:10px">
                    <span style="font-size:0.75rem;color:var(--text-muted)">Sprzedano offline:</span>
                    <span id="offlineInfo" style="color:var(--orange);font-weight:600">0 szt.</span>
                </div>
                <div style="display:flex;gap:10px">
                    <input type="number" id="cofnijIlosc" min="1" value="1" class="form-control" style="flex:1;text-align:center;border-color:var(--red)">
                    <button onclick="cofnijOffline()" class="btn btn-danger" style="width:auto;padding:12px 20px;margin:0"><span class=material-symbols-outlined>sync</span> Cofnij</button>
                </div>
            </div>

            <!-- COFNIJ SPRZEDAŻ -->
            <div style="border-top:1px solid var(--border);padding-top:15px;margin-top:15px">
                <label style="display:block;font-size:0.8rem;color:var(--red);margin-bottom:8px"><span class=material-symbols-outlined>sync</span> Cofnij sprzedaż (przywróć do magazynu):</label>
                <button onclick="cofnijSprzedaz()" class="btn btn-danger" style="margin:0"><span class=material-symbols-outlined>sync</span> Cofnij sprzedaż</button>
                <div style="font-size:0.65rem;color:var(--text-muted);margin-top:6px;text-align:center">
                    Cofa sprzedaż, przywraca ilość i zmienia status produktu na magazyn
                </div>
            </div>

            <input type="hidden" id="maxIlosc">
            <input type="hidden" id="offlineSzt">
        </div>
    </div>

    <!-- MODAL: ROZBIJ NA SZTUKI -->
    <div id="modalRozbij" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.8);z-index:1000;overflow-y:auto;padding:20px">
      <div style="background:var(--bg-card);border-radius:var(--radius);padding:20px;max-width:440px;margin:0 auto">
        <div style="font-size:1.2rem;font-weight:700;margin-bottom:4px"><span class=material-symbols-outlined>target</span> Rozbij stan na sztuki</div>
        <div id="rozbijNazwa" style="color:var(--text-secondary);font-size:0.85rem;margin-bottom:15px"></div>
        <div style="background:var(--bg);border-radius:10px;padding:12px;margin-bottom:15px">
          <div style="display:flex;justify-content:space-between;margin-bottom:4px">
            <span style="color:var(--text-secondary)">Łącznie sztuk:</span>
            <span id="rozbijLacznie" style="font-weight:700"></span>
          </div>
          <div style="display:flex;justify-content:space-between">
            <span style="color:var(--text-secondary)">Suma wpisanych:</span>
            <span id="rozbijSuma" style="font-weight:700;color:var(--green)"></span>
          </div>
        </div>
        <div id="rozbijStany"></div>
        <div style="color:var(--text-secondary);font-size:0.75rem;margin:12px 0 8px">Szybkie ustawienie:</div>
        <div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:15px">
          <button onclick="rozbijSzybko('Nowy')" style="padding:6px 12px;background:var(--green-soft);border:1px solid var(--green);border-radius:8px;color:var(--green);font-size:0.78rem;cursor:pointer">● Wszystko nowe</button>
          <button onclick="rozbijSzybko('Powystawowy')" style="padding:6px 12px;background:var(--blue-soft);border:1px solid var(--blue);border-radius:8px;color:var(--blue);font-size:0.78rem;cursor:pointer">● Powystawowe</button>
          <button onclick="rozbijSzybko('Używany')" style="padding:6px 12px;background:var(--yellow-soft);border:1px solid var(--yellow);border-radius:8px;color:var(--yellow);font-size:0.78rem;cursor:pointer">● Używane</button>
          <button onclick="rozbijSzybko('Uszkodzony')" style="padding:6px 12px;background:var(--red-soft);border:1px solid var(--red);border-radius:8px;color:var(--red);font-size:0.78rem;cursor:pointer"><span class=material-symbols-outlined>fiber_manual_record</span> Uszkodzone</button>
        </div>
        <div style="display:flex;gap:8px">
          <button onclick="rozbijWyczysc()" class="btn btn-secondary" style="flex:1;margin:0">Wyczyść</button>
          <button onclick="zamknijRozbij()" class="btn" style="flex:1;background:var(--text-muted);margin:0">Anuluj</button>
          <button onclick="zapiszRozbij()" class="btn btn-success" style="flex:1;margin:0;color:#000;font-weight:700">✓ Zapisz</button>
        </div>
      </div>
    </div>

    <!-- MODAL: DO NAPRAWY -->
    <div id="modalNaprawa" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.8);z-index:1000;overflow-y:auto;padding:20px">
      <div style="background:var(--bg-card);border-radius:var(--radius);padding:20px;max-width:440px;margin:0 auto">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">
          <div style="font-size:1.2rem;font-weight:700"><span class=material-symbols-outlined>build</span> Do naprawy</div>
          <button onclick="zamknijNaprawa()" style="background:none;border:none;color:var(--text-secondary);font-size:1.3rem;cursor:pointer">✕</button>
        </div>
        <div id="naprawaNazwa" style="color:var(--text-secondary);font-size:0.85rem;margin-bottom:15px"></div>
        <div id="naprawaLista"></div>
        <button onclick="zamknijNaprawa()" class="btn" style="background:var(--text-muted);margin-top:10px">Zamknij</button>
      </div>
    </div>

    <!-- MENU KONTEKSTOWE -->
    <div id="menuKontekst" style="display:none;position:fixed;z-index:2000;background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:8px;min-width:220px;box-shadow:var(--shadow-lg)">
      <div id="menuNaglowek" style="color:var(--text-muted);font-size:0.7rem;font-weight:600;padding:4px 10px;margin-bottom:4px"></div>
      <div id="menuStatusy"></div>
      <div style="color:var(--text-muted);font-size:0.7rem;font-weight:600;padding:4px 10px;margin:4px 0;border-top:1px solid var(--border);padding-top:8px">INNE AKCJE</div>
      <div id="menuInne"></div>
    </div>
    '''

    szczegoly_js = '''
    function pokazKorekta(produktId, aktualnaIlosc, cena, offlineSzt) {
        document.getElementById('korektaProduktId').value = produktId;
        document.getElementById('korektaIlosc').value = aktualnaIlosc;
        document.getElementById('maxIlosc').value = aktualnaIlosc;
        document.getElementById('sprzedajIlosc').value = 1;
        document.getElementById('sprzedajIlosc').max = aktualnaIlosc;
        document.getElementById('sprzedajCena').value = cena || '';
        document.getElementById('offlineSzt').value = offlineSzt || 0;
        const cofnijSection = document.getElementById('cofnijOfflineSection');
        if (offlineSzt && offlineSzt > 0) {
            cofnijSection.style.display = 'block';
            document.getElementById('offlineInfo').textContent = offlineSzt + ' szt.';
            document.getElementById('cofnijIlosc').value = 1;
            document.getElementById('cofnijIlosc').max = offlineSzt;
        } else {
            cofnijSection.style.display = 'none';
        }
        document.getElementById('modalKorekta').style.display = 'block';
    }

    document.addEventListener('click', function(e) {
        const btn = e.target.closest('.btn-korekta');
        if (btn) {
            e.preventDefault();
            pokazKorekta(
                btn.dataset.pid,
                parseInt(btn.dataset.ilosc),
                parseInt(btn.dataset.cena),
                parseInt(btn.dataset.offline)
            );
        }
    });

    function zamknijModal() {
        document.getElementById('modalKorekta').style.display = 'none';
    }

    function zapiszPole(produktId, pole, wartosc, el) {
        const fd = new FormData();
        fd.append('pole', pole);
        fd.append('wartosc', wartosc);
        fetch('/produkt/' + produktId + '/szybka-edycja', {method: 'POST', body: fd})
            .then(r => r.json())
            .then(d => {
                if (d.ok) {
                    el.style.border = '1px solid var(--green)';
                    if (pole === 'status' || d.reload) {
                        setTimeout(() => location.reload(), 400);
                    } else {
                        setTimeout(() => el.style.border = '1px solid var(--border)', 1200);
                    }
                } else {
                    el.style.border = '1px solid var(--red)';
                    alert('Błąd: ' + d.msg);
                }
            })
            .catch(() => { el.style.border = '1px solid var(--red)'; });
    }

    function cofnijOffline() {
        const ilosc = document.getElementById('cofnijIlosc').value;
        const maxOffline = document.getElementById('offlineSzt').value;

        if (parseInt(ilosc) > parseInt(maxOffline)) {
            alert('Nie możesz cofnąć więcej niż sprzedano offline (' + maxOffline + ' szt.)');
            return;
        }

        if (!confirm('Cofnąć ' + ilosc + ' szt. ze sprzedaży offline?\\n\\n(Produkty wrócą do magazynu)')) return;

        const produktId = document.getElementById('korektaProduktId').value;
        const form = document.createElement('form');
        form.method = 'POST';
        form.action = '/produkt/cofnij-offline/' + produktId;
        const inp = document.createElement('input');
        inp.type = 'hidden'; inp.name = 'ilosc'; inp.value = ilosc;
        form.appendChild(inp);
        document.body.appendChild(form);
        form.submit();
    }

    function cofnijSprzedaz() {
        const produktId = document.getElementById('korektaProduktId').value;
        if (!confirm('Cofnąć sprzedaż tego produktu?\\n\\nProdukt wróci do magazynu, sprzedaż zostanie oznaczona jako zwrot.')) return;
        const form = document.createElement('form');
        form.method = 'POST';
        form.action = '/produkt/cofnij-sprzedaz/' + produktId;
        document.body.appendChild(form);
        form.submit();
    }

    function zapiszKorekta() {
        const form = document.createElement('form');
        form.method = 'POST';
        form.action = '/sprzedaze/korekta-ilosci';

        const produktId = document.createElement('input');
        produktId.name = 'produkt_id';
        produktId.value = document.getElementById('korektaProduktId').value;
        form.appendChild(produktId);

        const ilosc = document.createElement('input');
        ilosc.name = 'nowa_ilosc';
        ilosc.value = document.getElementById('korektaIlosc').value;
        form.appendChild(ilosc);

        document.body.appendChild(form);
        form.submit();
    }

    function oznaczSprzedany() {
        const ilosc = document.getElementById('sprzedajIlosc').value;
        const cena = document.getElementById('sprzedajCena').value || 0;
        const maxIlosc = document.getElementById('maxIlosc').value;

        if (parseInt(ilosc) > parseInt(maxIlosc)) {
            alert('Nie możesz sprzedać więcej niż masz w magazynie (' + maxIlosc + ' szt.)');
            return;
        }

        const przychod = (parseFloat(cena) * parseInt(ilosc)).toFixed(2);
        if (!cena || parseFloat(cena) <= 0) {
            alert('Podaj cenę sprzedaży (zł) — pole nie może być puste ani zerowe.');
            document.getElementById('sprzedajCena').focus();
            return;
        }
        if (!confirm('Sprzedaż offline:\\n\\n' + ilosc + ' szt. × ' + cena + ' zł = ' + przychod + ' zł\\n\\n(Doliczy do przychodu palety)')) return;

        const produktId = document.getElementById('korektaProduktId').value;

        const cenaFixed = String(cena).replace(',', '.');
        console.log('OFFLINE SALE:', produktId, 'ilosc=' + ilosc, 'cena=' + cenaFixed);
        if (parseFloat(cena) <= 0) {
            if (!confirm('Cena wynosi 0 zł - czy na pewno chcesz sprzedać za darmo?')) return;
        }
        const form = document.createElement('form');
        form.method = 'POST';
        form.action = '/produkt/oznacz-sprzedany/' + produktId;
        const inpIlosc = document.createElement('input');
        inpIlosc.type = 'hidden'; inpIlosc.name = 'ilosc'; inpIlosc.value = ilosc;
        form.appendChild(inpIlosc);
        const inpCena = document.createElement('input');
        inpCena.type = 'hidden'; inpCena.name = 'cena'; inpCena.value = cenaFixed;
        form.appendChild(inpCena);
        document.body.appendChild(form);
        form.submit();
    }

    // Zamknij modal klikając poza nim
    document.getElementById('modalKorekta').addEventListener('click', function(e) {
        if (e.target === this) zamknijModal();
    });

    function szybkaMinus(produktId, aktIlosc, cena) {
        if (aktIlosc <= 0) { alert('Brak sztuk do odjęcia'); return; }
        const nowaIlosc = aktIlosc - 1;
        if (!confirm('Odjąć 1 szt? (' + aktIlosc + ' → ' + nowaIlosc + ')')) return;
        _submitKorekta(produktId, nowaIlosc);
    }
    function szybkaPlus(produktId, aktIlosc) {
        _submitKorekta(produktId, aktIlosc + 1);
    }
    function _submitKorekta(produktId, nowaIlosc) {
        const f = document.createElement('form');
        f.method = 'POST'; f.action = '/sprzedaze/korekta-ilosci';
        f.innerHTML = '<input name="produkt_id" value="'+produktId+'"><input name="nowa_ilosc" value="'+nowaIlosc+'">';
        document.body.appendChild(f); f.submit();
    }

    let _rozbijId = null, _rozbijIlosc = 0;
    let _naprawaId = null;
    let _menuId = null;

    const STANY_KOLORY = {
      'Nowy': '#22c55e', 'Powystawowy': '#3b82f6',
      'Używany': '#eab308', 'Uszkodzony': '#ef4444', 'Odnowiony': '#8b5cf6'
    };

    // ---- ROZBIJ NA SZTUKI ----
    function pokazRozbij(produktId, ilosc, nazwa) {
        _rozbijId = produktId; _rozbijIlosc = ilosc;
        document.getElementById('rozbijNazwa').textContent = nazwa;
        document.getElementById('rozbijLacznie').textContent = ilosc;
        fetch('/api/sztuki/' + produktId)
          .then(r => r.json()).then(d => {
            const istniejace = {};
            (d.sztuki || []).forEach(s => { istniejace[s.stan] = (istniejace[s.stan]||0)+1; });
            renderRozbijStany(istniejace);
          }).catch(() => renderRozbijStany({}));
        document.getElementById('modalRozbij').style.display = 'block';
    }
    function renderRozbijStany(wartosci) {
        const stany = ['Nowy','Powystawowy','Używany','Uszkodzony'];
        let html = '';
        stany.forEach(s => {
            const kolor = STANY_KOLORY[s];
            const val = wartosci[s] || 0;
            html += '<div style="display:flex;align-items:center;gap:12px;background:'+kolor+'11;border:1px solid '+kolor+'44;border-radius:10px;padding:12px;margin-bottom:8px">' +
              '<div style="width:14px;height:14px;border-radius:50%;background:'+kolor+';flex-shrink:0"></div>' +
              '<div style="flex:1;font-weight:600">'+s+'</div>' +
              '<button onclick="zmienjRozbij(\''+s+'\',-1)" style="width:36px;height:36px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px;color:var(--text);font-size:1.1rem;cursor:pointer">−</button>' +
              '<input type="number" id="rozbij_'+s+'" value="'+val+'" min="0" max="'+_rozbijIlosc+'"' +
              ' style="width:60px;text-align:center;background:var(--bg);border:1px solid var(--border);border-radius:8px;color:var(--text);padding:6px;font-size:1rem"' +
              ' oninput="aktualizujSume()">' +
              '<button onclick="zmienjRozbij(\''+s+'\',1)" style="width:36px;height:36px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px;color:var(--text);font-size:1.1rem;cursor:pointer">+</button>' +
            '</div>';
        });
        document.getElementById('rozbijStany').innerHTML = html;
        aktualizujSume();
    }
    function zmienjRozbij(stan, delta) {
        const el = document.getElementById('rozbij_' + stan);
        el.value = Math.max(0, parseInt(el.value||0) + delta);
        aktualizujSume();
    }
    function aktualizujSume() {
        const stany = ['Nowy','Powystawowy','Używany','Uszkodzony'];
        let suma = 0;
        stany.forEach(s => { suma += parseInt(document.getElementById('rozbij_'+s)?.value||0); });
        const el = document.getElementById('rozbijSuma');
        el.textContent = suma + ' / ' + _rozbijIlosc;
        el.style.color = suma === _rozbijIlosc ? 'var(--green)' : 'var(--red)';
    }
    function rozbijSzybko(stan) {
        ['Nowy','Powystawowy','Używany','Uszkodzony'].forEach(s => {
            const el = document.getElementById('rozbij_'+s);
            if(el) el.value = s === stan ? _rozbijIlosc : 0;
        });
        aktualizujSume();
    }
    function rozbijWyczysc() {
        ['Nowy','Powystawowy','Używany','Uszkodzony'].forEach(s => {
            const el = document.getElementById('rozbij_'+s);
            if(el) el.value = 0;
        });
        aktualizujSume();
    }
    function zamknijRozbij() { document.getElementById('modalRozbij').style.display='none'; }
    function zapiszRozbij() {
        const stany = ['Nowy','Powystawowy','Używany','Uszkodzony'];
        let suma = 0, podzial = {};
        stany.forEach(s => {
            const v = parseInt(document.getElementById('rozbij_'+s)?.value||0);
            if(v > 0) { podzial[s] = v; suma += v; }
        });
        if(suma !== _rozbijIlosc) { alert('Suma musi wynosić ' + _rozbijIlosc + ' sztuk!'); return; }
        fetch('/api/sztuki/' + _rozbijId + '/rozbij', {
            method:'POST', headers:{'Content-Type':'application/json'},
            body: JSON.stringify({podzial})
        }).then(r=>r.json()).then(d => {
            if(d.ok) { zamknijRozbij(); location.reload(); }
        });
    }

    // ---- NAPRAWA ----
    function pokazNaprawa(produktId, nazwa, ilosc) {
        _naprawaId = produktId;
        document.getElementById('naprawaNazwa').textContent = nazwa + ' — ' + ilosc + ' szt.';
        document.getElementById('naprawaLista').innerHTML = '<div style="color:var(--text-muted);text-align:center;padding:20px">Ładowanie...</div>';
        document.getElementById('modalNaprawa').style.display = 'block';
        fetch('/api/sztuki/' + produktId).then(r=>r.json()).then(d => {
            renderNaprawaLista(d.sztuki || [], ilosc);
        });
    }
    function renderNaprawaLista(sztuki, ilosc) {
        const pelna = [];
        for(let i=1; i<=ilosc; i++) {
            pelna.push(sztuki.find(s=>s.numer===i) || {id:null, numer:i, stan:'Nowy', status:'magazyn', opis_naprawy:''});
        }
        let html = '';
        pelna.forEach(s => {
            if(s.status === 'naprawa') {
                html += '<div style="background:var(--yellow-soft);border:1px solid rgba(245,158,11,0.3);border-radius:10px;padding:12px;margin-bottom:8px">' +
                  '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">' +
                    '<div style="font-weight:700;color:var(--orange)"><span class=material-symbols-outlined>build</span> szt. '+s.numer+' <span style="font-size:0.7rem">DO NAPRAWY</span></div>' +
                    '<div style="display:flex;gap:6px">' +
                      '<button onclick="edytujNaprawa('+s.id+', \''+((s.opis_naprawy||'').replace(/'/g,"\\'"))+'\')" style="padding:4px 10px;background:var(--purple);border:none;border-radius:6px;color:#fff;font-size:0.72rem;cursor:pointer"><span class=material-symbols-outlined>edit</span> Edytuj</button>' +
                      '<button onclick="cofnijNaprawa('+s.id+')" style="padding:4px 10px;background:var(--red-soft);border:1px solid var(--red);border-radius:6px;color:var(--red);font-size:0.72rem;cursor:pointer">↩ Cofnij</button>' +
                    '</div>' +
                  '</div>' +
                  '<div style="background:var(--bg-card);border-radius:6px;padding:8px;font-size:0.8rem"><span class=material-symbols-outlined>edit_note</span> '+(s.opis_naprawy || '—')+'</div>' +
                  (s.data_naprawy ? '<div style="font-size:0.7rem;color:var(--text-muted);margin-top:4px">'+s.data_naprawy+'</div>' : '') +
                '</div>';
            } else {
                html += '<div style="background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:12px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center">' +
                  '<div style="display:flex;align-items:center;gap:10px">' +
                    '<div style="width:12px;height:12px;border-radius:3px;background:'+(STANY_KOLORY[s.stan]||'var(--text-muted)')+'"></div>' +
                    '<span style="font-weight:600">szt. '+s.numer+'</span>' +
                    '<span style="font-size:0.72rem;color:var(--text-muted)">'+s.stan+'</span>' +
                  '</div>' +
                  '<button onclick="dodajNaprawa('+(s.id || 0)+', '+s.numer+', '+_naprawaId+')" style="padding:6px 14px;background:var(--orange);border:none;border-radius:8px;color:#000;font-size:0.75rem;font-weight:700;cursor:pointer">+ Do naprawy</button>' +
                '</div>';
            }
        });
        document.getElementById('naprawaLista').innerHTML = html;
    }
    function dodajNaprawa(sztukiId, numer, produktId) {
        const opis = prompt('Opis usterki dla szt. ' + numer + ':');
        if(opis === null) return;
        const doSave = (id) => {
            fetch('/api/sztuki/jednostka/' + id + '/naprawa', {
                method:'POST', headers:{'Content-Type':'application/json'},
                body: JSON.stringify({opis})
            }).then(r=>r.json()).then(() => {
                fetch('/api/sztuki/' + produktId).then(r=>r.json()).then(d => {
                    const ilosc = parseInt(document.getElementById('naprawaNazwa').textContent.match(/\\d+ szt/)[0]);
                    renderNaprawaLista(d.sztuki||[], ilosc);
                    location.reload();
                });
            });
        };
        if(sztukiId > 0) { doSave(sztukiId); }
        else {
            fetch('/api/sztuki/' + produktId + '/rozbij', {
                method:'POST', headers:{'Content-Type':'application/json'},
                body: JSON.stringify({podzial:{'Nowy': parseInt(document.getElementById('naprawaNazwa').textContent.match(/\\d+/)[0])}})
            }).then(r=>r.json()).then(() => {
                fetch('/api/sztuki/' + produktId).then(r=>r.json()).then(d => {
                    const szt = (d.sztuki||[]).find(s=>s.numer===numer);
                    if(szt) doSave(szt.id);
                });
            });
        }
    }
    function edytujNaprawa(id, opisCurrent) {
        const opis = prompt('Edytuj opis naprawy:', opisCurrent);
        if(opis === null) return;
        fetch('/api/sztuki/jednostka/' + id + '/naprawa', {
            method:'POST', headers:{'Content-Type':'application/json'},
            body: JSON.stringify({opis})
        }).then(() => location.reload());
    }
    function cofnijNaprawa(id) {
        fetch('/api/sztuki/jednostka/' + id + '/naprawa', {
            method:'POST', headers:{'Content-Type':'application/json'},
            body: JSON.stringify({cofnij: true})
        }).then(() => location.reload());
    }
    function zamknijNaprawa() { document.getElementById('modalNaprawa').style.display='none'; }

    // ---- MENU KONTEKSTOWE ----
    function pokazMenu(evt, produktId, ilosc, nazwa) {
        evt.stopPropagation();
        _menuId = produktId;
        const menu = document.getElementById('menuKontekst');
        document.getElementById('menuNaglowek').textContent = 'ZMIEŃ STATUS (dostępne: ' + ilosc + '/' + ilosc + ')';
        document.getElementById('menuStatusy').innerHTML =
            '<div onclick="menuStatus(\'sprzedany\')" class="menu-item"><span class=material-symbols-outlined>check_circle</span> Sprzedane</div>' +
            '<div onclick="menuStatus(\'sprzedany_uszkodzony\')" class="menu-item"><span class=material-symbols-outlined>warning</span> Sprzedane uszkodzone</div>' +
            '<div onclick="menuNaprawyModal('+produktId+', \''+nazwa.replace(/'/g,"\\'")+'\', '+ilosc+')" class="menu-item"><span class=material-symbols-outlined>build</span> Do naprawy...</div>' +
            '<div onclick="menuStatus(\'wyrzucenie\')" class="menu-item"><span class=material-symbols-outlined>delete</span> Do wyrzucenia</div>' +
            '<div onclick="menuStatus(\'zwrot\')" class="menu-item">[UNDO] Oddane (zwrot)</div>';
        document.getElementById('menuInne').innerHTML =
            '<div onclick="pokazRozbij('+produktId+', '+ilosc+', \''+nazwa.replace(/'/g,"\\'")+'\'); zamknijMenu()" class="menu-item"><span class=material-symbols-outlined>target</span> Rozbij na sztuki</div>' +
            '<a href="/magazyn/produkt/'+produktId+'/edytuj" class="menu-item" style="text-decoration:none;display:block;color:var(--text)"><span class=material-symbols-outlined>edit</span> Edytuj produkt</a>';
        const rect = evt.target.getBoundingClientRect();
        menu.style.display = 'block';
        menu.style.top = (rect.bottom + window.scrollY + 4) + 'px';
        menu.style.left = Math.min(rect.left, window.innerWidth - 240) + 'px';
    }
    function menuStatus(status) {
        if(!_menuId) return;
        zapiszPole(_menuId, 'status', status, document.createElement('span'));
        zamknijMenu();
        setTimeout(() => location.reload(), 300);
    }
    function menuNaprawyModal(id, nazwa, ilosc) {
        zamknijMenu();
        pokazNaprawa(id, nazwa, ilosc);
    }
    function zamknijMenu() { document.getElementById('menuKontekst').style.display='none'; }
    document.addEventListener('click', zamknijMenu);

    // Kropki stanów na kartach
    const KOLORY_STAN = {'Nowy':'#22c55e','Powystawowy':'#3b82f6','Używany':'#eab308','Uszkodzony':'#ef4444','Odnowiony':'#8b5cf6'};
    document.querySelectorAll('[data-produkt-id]').forEach(el => {
        const pid = el.dataset.produktId;
        const ilosc = parseInt(el.dataset.ilosc || 0);
        if(ilosc < 1) return;
        fetch('/api/sztuki/' + pid).then(r=>r.json()).then(d => {
            if(!d.sztuki || d.sztuki.length === 0) return;
            const counts = {};
            const naprawy = d.sztuki.filter(s => s.status === 'naprawa').length;
            d.sztuki.forEach(s => { counts[s.stan] = (counts[s.stan]||0)+1; });
            let html = '<div style="display:flex;gap:4px;flex-wrap:wrap;margin-top:4px">';
            Object.entries(counts).forEach(([k,v]) => {
                const kolor = KOLORY_STAN[k] || '#64748b';
                html += '<span style="background:'+kolor+'33;border:1px solid '+kolor+';color:'+kolor+';border-radius:20px;padding:1px 7px;font-size:0.62rem;font-weight:700">●'+k.slice(0,3)+' '+v+'</span>';
            });
            if(naprawy > 0) {
                html += '<span style="background:#f9730333;border:1px solid #f97316;color:#f97316;border-radius:20px;padding:1px 7px;font-size:0.62rem;font-weight:700"><span class=material-symbols-outlined>build</span>'+naprawy+'</span>';
            }
            html += '</div>';
            const dotsEl = el.querySelector('.sztuki-dots');
            if(dotsEl) dotsEl.innerHTML = html;
        }).catch(()=>{});
    });
    '''


    return render(content, f'Paleta {paleta["nazwa"] or paleta_id}', extra_js=szczegoly_js)

@palety_bp.route('/palety/<int:paleta_id>/scrape-images', methods=['POST'])
def paleta_scrape_images(paleta_id):
    """Scrapuj zdjęcia dla produktów bez zdjęć na danej palecie"""
    from modules.database import get_db
    try:
        conn = get_db()
        # Debug: pokaż WSZYSTKIE produkty na palecie i ich ASIN-y
        all_prods = conn.execute('SELECT id, nazwa, asin, ean, zdjecie_url FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchall()
        print(f"[SEARCH] SCRAPE DEBUG paleta #{paleta_id}: {len(all_prods)} produktów")
        for p in all_prods:
            print(f"   ID:{p['id']} | ASIN:{p['asin']!r} | EAN:{p['ean']!r} | img:{(p['zdjecie_url'] or '')[:40]!r} | {(p['nazwa'] or '')[:40]}")

        asins_rows = conn.execute('''
            SELECT DISTINCT asin FROM produkty
            WHERE paleta_id = ?
            AND asin IS NOT NULL AND asin != '' AND asin != 'nan'
            AND (zdjecie_url IS NULL OR zdjecie_url = '')
        ''', (paleta_id,)).fetchall()
        asins = [r['asin'] for r in asins_rows if r['asin'] and len(r['asin']) >= 5]
        print(f"[SEARCH] SCRAPE: znaleziono {len(asins)} ASIN-ów do scrapowania: {asins}")

        if not asins:
            # Sprawdź czy może wszystkie już mają zdjęcia
            with_img = conn.execute('SELECT COUNT(*) FROM produkty WHERE paleta_id = ? AND zdjecie_url IS NOT NULL AND zdjecie_url != ""', (paleta_id,)).fetchone()[0]
            without_asin = conn.execute('SELECT COUNT(*) FROM produkty WHERE paleta_id = ? AND (asin IS NULL OR asin = "" OR asin = "nan")', (paleta_id,)).fetchone()[0]
            return jsonify({'ok': False, 'error': f'Brak ASIN bez zdjęć (z_img:{with_img}, bez_asin:{without_asin})'})

        from modules.paletomat import auto_process_products
        auto_process_products(asins)
        return jsonify({'ok': True, 'count': len(asins), 'asins': asins})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)[:200]})


@palety_bp.route('/produkt/<int:produkt_id>/szybka-edycja', methods=['POST'])
def produkt_szybka_edycja(produkt_id):
    """Szybka inline zmiana pola produktu (stan, status) z palety"""
    from modules.database import get_db

    pole = request.form.get('pole', '').strip()
    wartosc = request.form.get('wartosc', '').strip()

    # Dozwolone pola do edycji inline
    DOZWOLONE = {'stan', 'status', 'lokalizacja', 'cena_allegro'}
    if pole not in DOZWOLONE:
        return jsonify({'ok': False, 'msg': 'Niedozwolone pole'}), 400

    conn = get_db()
    p = conn.execute('SELECT id, ilosc, status, nazwa, cena_allegro, cena_brutto FROM produkty WHERE id = ?', (produkt_id,)).fetchone()
    if not p:
        return jsonify({'ok': False, 'msg': 'Nie znaleziono'}), 404

    # Zabezpieczenie integralności: zmiana statusu na 'sprzedany' → zeruj ilosc + dodaj sprzedaz
    if pole == 'status' and wartosc == 'sprzedany' and (p['ilosc'] or 0) > 0:
        ilosc_sprzedana = p['ilosc'] or 1
        cena = float(p['cena_allegro'] or p['cena_brutto'] or 0)
        conn.execute('UPDATE produkty SET status = ?, ilosc = 0 WHERE id = ?', (wartosc, produkt_id))
        # Utwórz rekord sprzedaży żeby itemy nie "znikały" ze statystyk palety
        if cena > 0:
            from datetime import datetime
            conn.execute(
                '''INSERT INTO sprzedaze (produkt_id, nazwa, cena, ilosc, status, data_sprzedazy, kupujacy, notified)
                   VALUES (?, ?, ?, ?, 'sprzedana', ?, 'inline', 1)''',
                (produkt_id, p['nazwa'] or f'Produkt #{produkt_id}', cena, ilosc_sprzedana,
                 datetime.now().strftime('%Y-%m-%dT%H:%M:%S'))
            )
        conn.commit()
        return jsonify({'ok': True, 'msg': f'Sprzedano {ilosc_sprzedana} szt.', 'reload': True})

    # Zabezpieczenie: zmiana statusu z 'sprzedany' na inny → jeśli ilosc=0, to ostrzeżenie
    if pole == 'status' and p['status'] == 'sprzedany' and wartosc != 'sprzedany' and (p['ilosc'] or 0) == 0:
        return jsonify({'ok': False, 'msg': 'Produkt ma ilość 0 — najpierw skoryguj ilość'}), 400

    conn.execute('UPDATE produkty SET ' + pole + ' = ? WHERE id = ?', (wartosc, produkt_id))
    conn.commit()
    return jsonify({'ok': True})

@palety_bp.route('/palety/<int:paleta_id>/delete', methods=['POST'])
def paleta_delete(paleta_id):
    """Usuwa pojedynczą paletę i wszystkie jej produkty"""
    from modules.database import get_db

    conn = get_db()

    # Pobierz ASIN-y produktów do usunięcia ze scraped
    asiny = conn.execute('SELECT asin FROM produkty WHERE paleta_id = ? AND asin IS NOT NULL', (paleta_id,)).fetchall()
    asiny_list = [row[0] for row in asiny if row[0]]

    # Usuń produkty z palety ze scraped (Paletomat)
    scraped_cnt = 0
    if asiny_list:
        placeholders = ','.join(['?' for _ in asiny_list])
        scraped_cnt = conn.execute('DELETE FROM scraped WHERE asin IN (' + placeholders + ')', asiny_list).rowcount

    # Usuń produkty z palety
    produkty_cnt = conn.execute('DELETE FROM produkty WHERE paleta_id = ?', (paleta_id,)).rowcount

    # Usuń paletę
    conn.execute('DELETE FROM palety WHERE id = ?', (paleta_id,))
    conn.commit()

    content = f'''
    <div style="text-align:center;padding:60px 20px">
        <div style="font-size:3rem;margin-bottom:20px"><span class=material-symbols-outlined>check_circle</span></div>
        <div style="font-size:1.2rem">Paleta usunięta!</div>
        <div style="color:var(--text-muted);margin-top:10px">
            Usunięto {produkty_cnt} produktów{f' i {scraped_cnt} z Palatomatu' if scraped_cnt > 0 else ''}
        </div>
        <a href="/palety" class="btn btn-primary" style="display:inline-block;width:auto;margin-top:20px;padding:12px 24px">Powrót do palet</a>
    </div>
    <script>setTimeout(function(){{ window.location='/palety'; }}, 2000);</script>
    '''
    return render(content, 'Paleta usunięta')
