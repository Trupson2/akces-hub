"""
Paletomat module - scraper Amazon + generator ofert Allegro
"""

import os
import re
import json
import sqlite3
import threading
import time
from datetime import datetime, timedelta
from flask import Blueprint, render_template, render_template_string, request, redirect, jsonify, Response, current_app, session, flash
from flask_wtf.csrf import generate_csrf
from concurrent.futures import ThreadPoolExecutor, as_completed

from .database import get_db, query_db, execute_db, get_config, set_config
from .utils import get_amazon_image_url, oblicz_cene_allegro, generuj_opis_ai, generuj_opis_html_pro, scrape_amazon_product, optimize_title_seo, translate_product_name, generuj_gpsr_info
from .inventory_utils import SmartQuantityParser
from .title_generator_ai import generate_allegro_title_ai

paletomat_bp = Blueprint('paletomat', __name__)


def _extract_ean_from_specs(specs):
    """Wyciaga EAN/GTIN/UPC z slownika product_specs Amazonu.

    Amazon w roznych krajach uzywa roznych etykiet: 'EAN', 'GTIN', 'UPC',
    'Numer EAN', 'EAN/UPC', 'Code-barres', 'Codigo de barras' itd.
    Wartosc czasem ma kilka kodow oddzielonych przecinkiem - bierzemy pierwszy
    pasujacy do formatu cyfrowego (8-14 cyfr).
    """
    if not specs or not isinstance(specs, dict):
        return None
    # Slowa kluczowe ktore sygnalizuja ze pole zawiera EAN/GTIN/UPC
    EAN_KEYS = ('ean', 'gtin', 'upc', 'kod kreskowy', 'kod paskowy',
                'codigo de barras', 'code-barres', 'codice a barre',
                'streckkod', 'barcode')
    for key, val in specs.items():
        if not key or not val:
            continue
        kl = str(key).lower().strip().rstrip(':')
        if not any(k in kl for k in EAN_KEYS):
            continue
        # Wyciagnij pierwszy ciag 8-14 cyfr z wartosci (czesto sa kropki/spacje)
        for token in re.findall(r'\d[\d\s\-]{6,15}\d', str(val)):
            digits = re.sub(r'\D', '', token)
            if 8 <= len(digits) <= 14:
                return digits
    return None


def _resolve_initial_name(conn, asin, ean=None):
    """Zwraca nazwe dla nowo wstawianego produktu.

    Jezeli ASIN byl juz raz scrapeowany (jest w tabeli `scraped` z poprawna nazwa)
    - uzyj tej nazwy od razu, zeby uniknac placeholdera "Produkt B0...".
    W przeciwnym razie zwroc placeholder, ktory zostanie podmieniony pozniej
    przez auto-sync na widoku palety albo przycisk "Popraw nazwy".
    """
    if not asin:
        return f'Produkt {ean or "?"}'
    try:
        r = conn.execute(
            "SELECT nazwa FROM scraped WHERE UPPER(asin) = UPPER(?) "
            "AND nazwa IS NOT NULL AND nazwa != '' AND nazwa NOT LIKE 'Produkt %' "
            "LIMIT 1",
            (asin,)
        ).fetchone()
        if r and r['nazwa']:
            return r['nazwa']
    except Exception:
        pass
    return f'Produkt {asin}'


def _ensure_local_images(wszystkie_zdjecia, asin, zdjecie_url=''):
    """
    Sprawdza czy lokalne ścieżki zdjęć istnieją.
    Jeśli nie, re-downloaduje z CDN lub Amazona.
    Zwraca: (lista_zdjec, log_messages)
    """
    logs = []
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    import requests as _req
    _dl_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8',
        'Referer': 'https://www.google.com/',
    }

    # --- Sprawdź istniejące pliki/URL-e ---
    if wszystkie_zdjecia:
        first_img = wszystkie_zdjecia[0] if wszystkie_zdjecia else ''

        # Już URL-e - zwróć TYLKO jeśli mamy 2+ (jedno to za mało, spróbuj więcej)
        if first_img.startswith('http') and len(wszystkie_zdjecia) >= 2:
            return wszystkie_zdjecia, logs

        # Lokalne ścieżki - sprawdź czy pliki istnieją i nie są puste
        if not first_img.startswith('http'):
            existing = []
            for img in wszystkie_zdjecia:
                norm = os.path.normpath(img)
                if os.path.exists(norm) and os.path.getsize(norm) > 500:
                    existing.append(norm)
                else:
                    abs_path = os.path.normpath(os.path.join(base_dir, img))
                    if os.path.exists(abs_path) and os.path.getsize(abs_path) > 500:
                        existing.append(abs_path)

            if existing and len(existing) >= 8:
                return existing, logs
            # Mamy trochę lokalnych ale mniej niż 8 - spróbuj dobrać z scraped

    # --- Pliki nie istnieją lub brak listy - pobierz ---

    # KROK 1: Sprawdź cached CDN URLs w tabeli scraped
    cached_downloaded = []
    if asin:
        try:
            _conn = get_db()
            _scraped = _conn.execute(
                'SELECT wszystkie_zdjecia FROM scraped WHERE asin = ?', (asin,)
            ).fetchone()
            if _scraped and _scraped['wszystkie_zdjecia']:
                cached_urls = json.loads(_scraped['wszystkie_zdjecia'])
                if isinstance(cached_urls, list) and cached_urls:
                    cdn_urls = [u for u in cached_urls if isinstance(u, str) and u.startswith('http')]
                    if cdn_urls:
                        # Pobierz WSZYSTKIE dostępne (max 8) - zawsze, nie limituj do 2+
                        logs.append((f'<span class=material-symbols-outlined>photo_camera</span> Znaleziono {len(cdn_urls)} zdj w cache (scraped)', '#8b5cf6'))
                        asin_dir = os.path.join(base_dir, 'static', 'downloads', str(asin))
                        os.makedirs(asin_dir, exist_ok=True)
                        for i, url in enumerate(cdn_urls[:8], 1):
                            fpath = os.path.join(asin_dir, f"image_{i}.jpg")
                            # Pomij jeśli plik już istnieje i jest OK (oszczędność czasu)
                            if os.path.exists(fpath) and os.path.getsize(fpath) > 1000:
                                cached_downloaded.append(fpath)
                                continue
                            try:
                                resp = _req.get(url, headers=_dl_headers, timeout=15)
                                if resp.status_code == 200 and len(resp.content) > 1000:
                                    with open(fpath, 'wb') as fw:
                                        fw.write(resp.content)
                                    cached_downloaded.append(fpath)
                            except:
                                pass
                        if cached_downloaded:
                            logs.append((f'<span class=material-symbols-outlined>check_circle</span> Pobrano {len(cached_downloaded)}/{len(cdn_urls[:8])} zdj z cache CDN', '#22c55e'))
                            return cached_downloaded, logs
                        else:
                            logs.append(('<span class=material-symbols-outlined>photo_camera</span> Uzywam URL-i CDN bezposrednio', '#3b82f6'))
                            return cdn_urls[:8], logs
        except:
            pass

    # KROK 2: Pobierz główne zdjęcie z CDN (zabezpieczenie - przynajmniej 1 zdjęcie)
    cdn_fallback_path = None
    if zdjecie_url and zdjecie_url.startswith('http'):
        try:
            asin_dir = os.path.join(base_dir, 'static', 'downloads', str(asin or 'misc'))
            os.makedirs(asin_dir, exist_ok=True)
            resp = _req.get(zdjecie_url, headers=_dl_headers, timeout=15)
            if resp.status_code == 200 and len(resp.content) > 1000:
                cdn_fallback_path = os.path.join(asin_dir, "image_1.jpg")
                with open(cdn_fallback_path, 'wb') as fw:
                    fw.write(resp.content)
                logs.append((f'<span class=material-symbols-outlined>check_circle</span> Pobrano główne zdjęcie z CDN ({len(resp.content)//1024} KB)', '#22c55e'))
        except:
            pass

    if not asin:
        if cdn_fallback_path:
            return [cdn_fallback_path], logs
        if zdjecie_url and zdjecie_url.startswith('http'):
            return [zdjecie_url], [('<span class=material-symbols-outlined>photo_camera</span> Użyję URL bezpośrednio (bez lokalnego cache)', '#3b82f6')]
        return [], [('<span class=material-symbols-outlined>cancel</span> Lokalne zdjęcia nie istnieją, brak ASIN', '#ef4444')]

    # KROK 3: Scrapuj Amazon po WIĘCEJ zdjęć (amazon.pl priorytet)
    logs.append(('🌍 Scrapuję Amazon po więcej zdjęć (amazon.pl priorytet)...', '#eab308'))
    try:
        from .utils import scrape_amazon_product as _scrape
        amazon_data = _scrape(asin)
        if amazon_data and amazon_data.get('all_images'):
            amazon_urls = amazon_data['all_images']

            # Zapisz CDN URL-e do scraped table (cache na przyszłość)
            try:
                _conn = get_db()
                _conn.execute(
                    'UPDATE scraped SET wszystkie_zdjecia = ? WHERE asin = ?',
                    (json.dumps(amazon_urls[:8]), asin)
                )
                _conn.commit()
                logs.append((f'<span class=material-symbols-outlined>save</span> Zapisano {len(amazon_urls[:8])} URL-i CDN do cache', '#8b5cf6'))
            except:
                pass

            asin_dir = os.path.join(base_dir, 'static', 'downloads', str(asin))
            os.makedirs(asin_dir, exist_ok=True)

            nowe = []
            for i, url in enumerate(amazon_urls[:8], 1):
                try:
                    resp = _req.get(url, headers=_dl_headers, timeout=15)
                    if resp.status_code == 200 and len(resp.content) > 1000:
                        fpath = os.path.join(asin_dir, f"image_{i}.jpg")
                        with open(fpath, 'wb') as fw:
                            fw.write(resp.content)
                        nowe.append(fpath)
                except:
                    pass

            if nowe:
                logs.append((f'<span class=material-symbols-outlined>check_circle</span> Pobrano {len(nowe)} zdjęć z Amazon', '#22c55e'))
                return nowe, logs
            else:
                logs.append((f'<span class=material-symbols-outlined>photo_camera</span> Użyję {len(amazon_urls[:8])} URL-i bezpośrednio', '#3b82f6'))
                return amazon_urls[:8], logs
        else:
            # Scraping nie zwrócił zdjęć - fallback
            if cdn_fallback_path:
                return [cdn_fallback_path], logs
            if zdjecie_url and zdjecie_url.startswith('http'):
                return [zdjecie_url], logs
    except Exception as e:
        logs.append((f'<span class=material-symbols-outlined>cancel</span> Scrape error: {str(e)[:50]}', '#ef4444'))

    # Fallback - przynajmniej główne zdjęcie
    if cdn_fallback_path:
        return [cdn_fallback_path], logs
    if zdjecie_url and zdjecie_url.startswith('http'):
        return [zdjecie_url], [('<span class=material-symbols-outlined>photo_camera</span> Fallback na główne zdjęcie URL', '#eab308')]
    return [], logs


def auto_kategoryzuj(nazwa):
    """Automatycznie przypisz kategorię na podstawie nazwy produktu"""
    nazwa_lower = (nazwa or '').lower()
    
    # EV / Ładowarki samochodowe (priorytet!)
    if any(word in nazwa_lower for word in ['wallbox', 'ev ', ' ev', 'evse', 'ev charger', 'type 2', 'type2', 'type-2',
        'ccs', 'chademo', 'tesla', 'charging station', 'stacja ładowania', 'ładowarka samochod', 'ładowarka ev',
        'electric vehicle', 'elektromobil', 'green cell ev', 'juice booster', 'go-e', 'easee', 'zappi',
        'mennekes', 'j1772', 'nema', '11kw', '22kw', '7kw', '3.6kw', '32a', '16a']):
        return 'ev_ladowarki'
    
    # <span class=material-symbols-outlined>photo_camera</span> Foto / Video / Streaming
    if any(word in nazwa_lower for word in ['softbox', 'ring light', 'lampa pierścieniowa', 'pierścieniowa', 
        'tło fotograficzne', 'tlo fotograficzne', 'backdrop', 'greenscreen', 'green screen',
        'gopro', 'insta360', 'dji', 'osmo', 'action cam', 'kamera sportowa',
        'smallrig', 'cage', 'rig ', 'stabilizator', 'steadycam', 'follow focus',
        'mikrofon', 'microphone', 'lavalier', 'shotgun mic', 'rode', 'boya',
        'teleprompter', 'prompter', 'stojak na tło', 'statyw oświetleniowy', 'light stand',
        'reflektor', 'panel led', 'oświetlenie fotograficzne', 'oświetlenie studyjne',
        'transmisja', 'streaming', 'capture card', 'elgato', 'cam link']):
        return 'foto_video'
    
    # <span class=material-symbols-outlined>print</span> Druk 3D
    if any(word in nazwa_lower for word in ['filament', 'pla', 'abs', 'petg', 'tpu', 'drukarka 3d', '3d printer',
        'druk 3d', 'nozzle', 'dysza', 'hotend', 'extruder', 'bed ', 'stół grzewczy', 'creality', 'ender',
        'prusa', 'anycubic', 'elegoo', 'resin', 'żywica', 'szpula']):
        return 'druk3d'
    
    # [VIDE] Smart Home / Monitoring
    if any(word in nazwa_lower for word in ['kamera ip', 'kamera wifi', 'kamera wlan', 'monitoring', 'cctv',
        'ezviz', 'hikvision', 'dahua', 'reolink', 'imou', 'tapo', 'arlo',
        'smart home', 'smarthome', 'inteligentny dom', 'czujnik ruchu', 'pir',
        'wideodomofon', 'domofon', 'dzwonek wifi', 'ring doorbell', 'alarm']):
        return 'smart_home'
    
    # [DIRE] Motoryzacja (rozszerzona - kamera cofania!)
    if any(word in nazwa_lower for word in ['samochod', 'samochód', 'auto ', ' auto', 'car ', 'obd', 'diagnosty',
        'koło', 'opona', 'silnik', 'akumulator', 'kamera cofania', 'cofania', 'backup camera', 'reversing',
        'dash cam', 'dashcam', 'rejestrator jazdy', 'wideorejestrator', 'parkowania', 'czujnik parkowania',
        'cb radio', 'nawigacja gps', 'uchwyt samochodowy', 'ładowarka samochodowa', 'car charger']):
        return 'motoryzacja'
    
    # AGD małe
    if any(word in nazwa_lower for word in ['mikser', 'blender', 'toster', 'czajnik', 'kettle', 'odkurzacz', 'vacuum',
        'żelazko', 'iron', 'suszarka', 'dryer', 'golarki', 'shaver', 'depilator', 'maszynka', 'robot kuchenny',
        'ekspres', 'coffee', 'frytkownica', 'air fryer', 'grill', 'opiekacz', 'mikrofala', 'microwave',
        'robot sprzątający', 'roomba', 'roborock']):
        return 'agd_male'
    
    # AGD duże
    if any(word in nazwa_lower for word in ['lodówka', 'fridge', 'pralka', 'washing', 'zmywarka', 'dishwasher',
        'piekarnik', 'oven', 'kuchenka', 'cooker', 'klimatyzator', 'air condition', 'freezer', 'zamrażar']):
        return 'agd_duze'
    
    # Komputery / IT
    if any(word in nazwa_lower for word in ['laptop', 'notebook', 'komputer', 'computer', 'pc ', ' pc', 'monitor',
        'klawiatura', 'keyboard', 'myszka', 'mouse', 'drukarka', 'printer', 'skaner', 'scanner', 'ssd', 'hdd',
        'ram ', ' ram', 'procesor', 'cpu', 'gpu', 'grafika', 'płyta główna', 'motherboard', 'nas ', 'server']):
        return 'komputery'
    
    # Telefony / Smartfony
    if any(word in nazwa_lower for word in ['smartfon', 'smartphone', 'iphone', 'samsung galaxy', 'xiaomi', 'redmi',
        'huawei', 'oppo', 'realme', 'oneplus', 'google pixel', 'telefon', 'mobile phone', 'cell phone']):
        return 'telefony'
    
    # Akcesoria elektroniczne (rozszerzona - statywy, selfie)
    if any(word in nazwa_lower for word in ['ładowarka', 'charger', 'kabel', 'cable', 'słuchawki', 'headphone', 'earbuds',
        'powerbank', 'power bank', 'bluetooth', 'adapter', 'przejściówka', 'hub usb', 'dock', 'stacja dokująca',
        'etui', 'case', 'folia', 'szkło', 'uchwyt', 'holder', 'statyw', 'tripod', 'gimbal', 'selfi', 'selfie',
        'monopod', 'stick', 'okular', 'obiektyw', 'lens', 'filtr']):
        return 'akcesoria'
    
    # RTV / Audio-Video
    if any(word in nazwa_lower for word in ['telewizor', 'tv ', ' tv', 'soundbar', 'głośnik', 'speaker', 'kino domowe',
        'projektor', 'projector', 'odtwarzacz', 'player', 'amplituner', 'wzmacniacz', 'subwoofer', 'radio',
        'dvd', 'blu-ray', 'chromecast', 'fire stick', 'apple tv', 'roku']):
        return 'rtv'
    
    # Gaming
    if any(word in nazwa_lower for word in ['playstation', 'ps4', 'ps5', 'xbox', 'nintendo', 'switch', 'konsola',
        'gamepad', 'kontroler', 'joystick', 'gaming', 'gra ', ' gra', 'vr ', 'oculus', 'quest', 'pad perkusyjny']):
        return 'gaming'
    
    # Narzędzia
    if any(word in nazwa_lower for word in ['wiertarka', 'drill', 'wkrętarka', 'screwdriver', 'szlifierka', 'grinder',
        'piła', 'saw', 'młotek', 'hammer', 'klucz', 'wrench', 'zestaw narzędzi', 'tool kit', 'kompresor',
        'spawarka', 'welder', 'lutownica', 'multimetr', 'poziomica']):
        return 'narzedzia'
    
    # Dom i ogród
    if any(word in nazwa_lower for word in ['meble', 'furniture', 'ogród', 'garden', 
        'dywan', 'carpet', 'zasłon', 'curtain', 'doniczk', 'plant', 'sofa', 'krzesło', 'chair', 'stół', 'table',
        'kosiarka', 'mower', 'podkaszarka', 'trimmer', 'wąż ogrodowy', 'grill ogrodowy', 'parasol']):
        return 'dom_ogrod'
    
    # Sport / Fitness (rozszerzona - walkingpad, bieżnia)
    if any(word in nazwa_lower for word in ['rower', 'bike', 'bicycle', 'fahrrad', 'hulajnoga', 'scooter', 'roller',
        'rolki', 'skate', 'siłownia', 'gym',
        'hantle', 'dumbbell', 'bieżnia', 'treadmill', 'orbitrek', 'elliptical', 'rowerek', 'mata', 'yoga', 'fitness',
        'walkingpad', 'walking pad', 'stepper', 'kettlebell', 'gryf', 'sztanga', 'ćwiczeni', 'trampolin',
        'namiot', 'tent', 'śpiwór', 'sleeping bag', 'kajakow', 'kayak', 'wędka', 'fishing']):
        return 'sport'

    # Zabawki / Dzieci
    if any(word in nazwa_lower for word in ['zabawka', 'toy', 'spielzeug', 'klocki', 'lego', 'lalka', 'doll', 'pluszak',
        'gra planszowa', 'board game', 'brettspiel',
        'puzzle', 'samochodzik', 'kolejka', 'dziecięc', 'child', 'baby', 'wózek', 'fotelik', 'kindersitz',
        'kojec', 'łóżeczko', 'smoczek', 'pieluchy', 'pampers', 'bobas']):
        return 'zabawki'

    # Moda
    if any(word in nazwa_lower for word in ['buty', 'shoes', 'schuhe', 'ubrani', 'cloth', 'kleidung', 'koszul', 'shirt',
        'hemd', 'spodni', 'pants', 'hose',
        'sukienk', 'dress', 'kleid', 'kurtk', 'jacket', 'jacke', 'bluza', 'sweater', 'pullover',
        'czapk', 'hat', 'mütze', 'torebk', 'bag', 'tasche', 'plecak', 'rucksack',
        'zegarek', 'watch', 'uhr', 'biżuteri', 'jewelry', 'schmuck', 'okulary', 'glasses', 'brille']):
        return 'moda'

    # Zdrowie / Uroda
    if any(word in nazwa_lower for word in ['masażer', 'massager', 'massagegerät', 'ciśnieniomierz', 'termometr',
        'inhalator', 'pulsoksymetr',
        'szczoteczka', 'toothbrush', 'zahnbürste', 'suszarka do włosów', 'hair dryer', 'haartrockner',
        'prostownica', 'straightener', 'glätteisen', 'lokówka', 'curling', 'lockenstab',
        'trymer', 'golarka', 'shaver', 'rasierer', 'depilator', 'epilator',
        'waga łazienkowa', 'bathroom scale', 'personenwaage']):
        return 'zdrowie'

    # Zwierzęta
    if any(word in nazwa_lower for word in ['karma', 'pet food', 'tierfutter', 'obroża', 'collar', 'halsband',
        'smycz', 'leash', 'leine', 'klatka', 'cage', 'käfig',
        'akwarium', 'aquarium', 'terrarium', 'legowisko', 'pet bed', 'hundebett',
        'kuweta', 'litter box', 'transporter', 'drapak', 'scratching', 'kratzbaum']):
        return 'zwierzeta'

    # Biuro
    if any(word in nazwa_lower for word in ['krzesło biurowe', 'office chair', 'bürostuhl', 'biurko', 'desk',
        'schreibtisch', 'niszczarka', 'shredder', 'fotel biurowy', 'laminator', 'bindownica']):
        return 'biuro'

    # Łazienka
    if any(word in nazwa_lower for word in ['bateria łazienkowa', 'faucet', 'wasserhahn', 'prysznic', 'shower', 'dusche',
        'lustro', 'mirror', 'spiegel', 'umywalka', 'sink', 'waschbecken', 'wanna', 'bathtub', 'badewanne']):
        return 'lazienka'

    return 'inne'


# Stan scrapera
_scraper_running = False
_processing_queue = []

# [AGRI] KOMBAJN MODE: Ustawienia równoległego przetwarzania
# AUTO-DETECTION sprzętu dla optymalnej wydajności (opcjonalne)

def detect_optimal_workers():
    """
    Pi = 2 workery (mniej CPU, mniej grzania)
    PC = 5 workerów
    """
    import platform
    if platform.machine().startswith('a') or 'arm' in platform.machine().lower():
        return 2  # Raspberry Pi — oszczędzaj CPU
    return 5  # PC

# MAX_WORKERS: liczba równoległych zapytań do Amazon
MAX_WORKERS = detect_optimal_workers()
PROGRESS = {'current': 0, 'total': 0, 'errors': 0}

def process_single_product(asin, position, total, preferred_domain=None):
    """
    Przetwarza pojedynczy produkt (dla ThreadPoolExecutor).

    Args:
        asin: Kod ASIN produktu
        position: Numer produktu w kolejce
        total: Całkowita liczba produktów
        preferred_domain: np. 'co.uk', 'de', 'com'

    Returns:
        (asin, success, error_msg)
    """
    try:
        print(f"\n{'='*50}")
        print(f"[AGRI] [{position}/{total}] Processing: {asin}")
        print(f"{'='*50}")

        # Pobierz dane z Amazona
        amazon_data = scrape_amazon_product(asin, preferred_domain=preferred_domain)
        if not amazon_data:
            print(f"⚠️ Could not scrape: {asin}")
            return (asin, False, "Scraping failed")
        
        nazwa = amazon_data.get('title', '') or f'Produkt {asin}'
        # Przetłumacz nazwę na polski
        nazwa = translate_product_name(nazwa, use_ai=True)
        zdjecie_url = amazon_data.get('image_url', '') or get_amazon_image_url(asin)
        wszystkie_zdjecia = amazon_data.get('all_images', []) or [zdjecie_url]
        cena_amazon = amazon_data.get('price', 0) or 0
        bullet_points = amazon_data.get('bullet_points', [])
        product_specs = amazon_data.get('product_specs', {})
        # Auto-kategoryzacja na podstawie nazwy produktu
        kategoria = auto_kategoryzuj(nazwa)
        
        print(f"[CHECK_CIRCLE] Scraped: {nazwa[:50]}...")
        print(f"[PHOTO_CAMERA] Images: {len(wszystkie_zdjecia)}")
        print(f"[EDIT_NOTE] Features: {len(bullet_points)}")

        # <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> NATYCHMIAST zapisz nazwę do bazy (żeby nie było "Produkt B0...")
        # Wyciagnij EAN ze specyfikacji Amazonu (jesli jest)
        extracted_ean = _extract_ean_from_specs(product_specs)
        try:
            conn = get_db()
            # Aktualizuj produkty z placeholder nazwami
            conn.execute('UPDATE produkty SET nazwa=?, kategoria=?, zdjecie_url=? WHERE asin=? AND (nazwa LIKE "Produkt %" OR nazwa LIKE "%" || asin || "%")',
                (nazwa, kategoria, zdjecie_url, asin))
            # Aktualizuj zdjecie_url dla produktów BEZ ZDJĘĆ (np. po ręcznym imporcie)
            conn.execute('UPDATE produkty SET zdjecie_url=? WHERE asin=? AND (zdjecie_url IS NULL OR zdjecie_url = "")',
                (zdjecie_url, asin))
            conn.execute('UPDATE scraped SET nazwa=?, kategoria=? WHERE asin=?',
                (nazwa, kategoria, asin))
            # Zapisz EAN do produktow ktore go nie maja (NULL/puste/N/A)
            if extracted_ean:
                conn.execute(
                    "UPDATE produkty SET ean=? "
                    "WHERE asin=? AND (ean IS NULL OR ean='' OR UPPER(ean) IN ('N/A','NAN','NONE'))",
                    (extracted_ean, asin)
                )
                # Cache do scraped (kolumna ean dodana w pozniejszej migracji - moze nie istniec)
                try:
                    conn.execute(
                        "UPDATE scraped SET ean=? WHERE asin=? AND (ean IS NULL OR ean='')",
                        (extracted_ean, asin)
                    )
                except Exception:
                    pass
                print(f"[BARCODE] EAN ze specs: {extracted_ean}")
            conn.commit()
            print(f"[EDIT_NOTE] Nazwa zapisana od razu: {nazwa[:50]}")
        except Exception as e:
            print(f"⚠️ Szybki zapis nazwy: {e}")

        # <span class=material-symbols-outlined>download</span> POBIERZ WSZYSTKIE ZDJĘCIA LOKALNIE - NOWA ORGANIZACJA KATALOGÓW
        lokalne_zdjecia = []
        print(f"[DOWNLOAD] Pobieram {len(wszystkie_zdjecia)} zdjęć lokalnie...")
        
        # Stwórz katalog dla ASIN
        import os
        asin_dir = os.path.join('static', 'downloads', str(asin))
        os.makedirs(asin_dir, exist_ok=True)
        
        _img_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8',
            'Referer': 'https://www.google.com/',
        }
        for idx, img_url in enumerate(wszystkie_zdjecia[:8], 1):  # Max 8, numeruj od 1
            try:
                # Pobierz zdjęcie
                import requests
                response = requests.get(img_url, headers=_img_headers, timeout=15)
                if response.status_code == 200 and len(response.content) > 500:
                    # Zapisz lokalnie w NOWEJ STRUKTURZE: static/downloads/{asin}/image_N.jpg
                    local_filename = os.path.join(asin_dir, f"image_{idx}.jpg")
                    with open(local_filename, 'wb') as f:
                        f.write(response.content)
                    lokalne_zdjecia.append(local_filename)
                    print(f"   ✓ [{idx}/{len(wszystkie_zdjecia[:8])}] Pobrano: {local_filename}")
                else:
                    print(f"   ✗ [{idx}/{len(wszystkie_zdjecia[:8])}] HTTP {response.status_code}")
            except Exception as e:
                print(f"   ✗ [{idx}/{len(wszystkie_zdjecia[:8])}] Error: {str(e)[:50]}")
        
        print(f"[CHECK_CIRCLE] Pobrano {len(lokalne_zdjecia)}/{len(wszystkie_zdjecia[:8])} zdjęć lokalnie")

        # Zachowaj CDN URL-e do zapisu w scraped (do ponownego pobrania w przyszłości)
        cdn_urls_for_cache = [u for u in wszystkie_zdjecia[:8] if isinstance(u, str) and u.startswith('http')]

        # Użyj lokalnych ścieżek zamiast URL (do aktualnego opisu)
        wszystkie_zdjecia = lokalne_zdjecia if lokalne_zdjecia else [zdjecie_url]
        # Aktualizuj zdjecie_url na lokalną ścieżkę (CDN URLs wygasają)
        if lokalne_zdjecia:
            zdjecie_url = '/' + lokalne_zdjecia[0]  # /static/downloads/ASIN/image_1.jpg

        # === HYBRID ENHANCE: Wyczyszczone oryginały (1-4) + AI (5-8) ===
        # Pomiń na Pi (za ciężkie - image processing + Gemini API crashuje system)
        import platform as _plat
        _is_pi = _plat.machine().startswith('a') or 'arm' in _plat.machine().lower()
        try:
            from .image_enhancer import enhance_single, prepare_original_photo, GEMINI_AVAILABLE as _ENH_SCRAPE
            from .image_enhancer import HYBRID_ORIGINAL_SLOTS, HYBRID_AI_TEMPLATES
            from .image_cleaner import clean_image_from_bytes

            if _ENH_SCRAPE and lokalne_zdjecia and not _is_pi:
                print(f"[AUTO_AWESOME] [{position}/{total}] HYBRID: oryginały + AI dla {asin}...")
                _enh_dir = os.path.join('static', 'enhanced', str(asin))
                os.makedirs(_enh_dir, exist_ok=True)
                _enh_ok = []

                # --- KROK 1: Wyczyść i zapisz oryginały na sloty 1-4 ---
                _orig_count = 0
                for _oi, _slot_name in enumerate(HYBRID_ORIGINAL_SLOTS):
                    if _oi >= len(lokalne_zdjecia):
                        break
                    try:
                        with open(lokalne_zdjecia[_oi], 'rb') as _bf:
                            _raw = _bf.read()
                        # Cleaner usuwa wszystko: loga, tekst, infografiki
                        _cb, _cm, _ce = clean_image_from_bytes(_raw)
                        _clean = _cb if _cb else _raw
                        _prep, _perr = prepare_original_photo(_clean)
                        if _prep:
                            _epath = os.path.join(_enh_dir, f'{_slot_name}.jpg')
                            with open(_epath, 'wb') as _sf:
                                _sf.write(_prep)
                            _enh_ok.append(_epath)
                            _orig_count += 1
                            print(f"   [PHOTO_CAMERA] [{_orig_count}] {_slot_name} — wyczyszczony oryginał [CHECK_CIRCLE]")
                        else:
                            print(f"   ⚠️ {_slot_name}: {str(_perr)[:40]}")
                    except Exception as _oe:
                        print(f"   [CANCEL] {_slot_name}: {str(_oe)[:40]}")

                # Brakujące sloty → AI
                _missing = HYBRID_ORIGINAL_SLOTS[_orig_count:]
                _slot_to_tpl = {'mini': 1, 'det': 3, 'zest': 4, 'kat2': 5}

                # Baza do AI = wyczyszczony oryginał #1
                with open(lokalne_zdjecia[0], 'rb') as _bf:
                    _raw_base = _bf.read()
                _cb_base, _, _ = clean_image_from_bytes(_raw_base)
                _base_for_ai = _cb_base if _cb_base else _raw_base

                for _ms in _missing:
                    _tid = _slot_to_tpl.get(_ms, 1)
                    try:
                        _ed, _em, _ee = enhance_single(_base_for_ai, _tid, nazwa[:60])
                        if _ed:
                            from PIL import Image as _PImg
                            from io import BytesIO as _BIO_s
                            _eimg = _PImg.open(_BIO_s(_ed)).convert('RGB')
                            _epath = os.path.join(_enh_dir, f'{_ms}.jpg')
                            if max(_eimg.width, _eimg.height) < 2560:
                                _ur = 2560 / max(_eimg.width, _eimg.height)
                                _eimg = _eimg.resize((int(_eimg.width * _ur), int(_eimg.height * _ur)), _PImg.LANCZOS)
                            _eimg.save(_epath, 'JPEG', quality=95)
                            _enh_ok.append(_epath)
                            print(f"   [SMART_TOY] {_ms} — AI [CHECK_CIRCLE]")
                    except Exception:
                        pass
                    time.sleep(0.5)

                # --- KROK 2: AI generuje wymiary, uzycie, lifestyle, skala ---
                for _tid, _tname in HYBRID_AI_TEMPLATES:
                    try:
                        _ed, _em, _ee = enhance_single(_base_for_ai, _tid, nazwa[:60])
                        if _ed:
                            from PIL import Image as _PImg
                            from io import BytesIO as _BIO_s
                            _eimg = _PImg.open(_BIO_s(_ed)).convert('RGB')
                            _epath = os.path.join(_enh_dir, f'{_tname}.jpg')
                            if max(_eimg.width, _eimg.height) < 2560:
                                _ur = 2560 / max(_eimg.width, _eimg.height)
                                _eimg = _eimg.resize((int(_eimg.width * _ur), int(_eimg.height * _ur)), _PImg.LANCZOS)
                            _eimg.save(_epath, 'JPEG', quality=95)
                            _enh_ok.append(_epath)
                            print(f"   [SMART_TOY] {_tname} — AI [CHECK_CIRCLE]")
                    except Exception:
                        pass
                    time.sleep(0.5)

                if _enh_ok:
                    print(f"   [CHECK_CIRCLE] {len(_enh_ok)}/8 zdjęć ({_orig_count} oryg + {len(_enh_ok)-_orig_count} AI)")
                    wszystkie_zdjecia = _enh_ok[:8]
                else:
                    print(f"   ⚠️ Enhance nie powiódł się, oryginalne zdjęcia")
        except Exception as _enhx_s:
            print(f"   ⚠️ Enhance error: {str(_enhx_s)[:60]}")
        
        # <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> TURBO: Generuj tytuł SEO używając AI (Gemini)
        # Import klucza z gemini_config.py (jak smart_importer)
        try:
            from gemini_config import GEMINI_API_KEY as gemini_key
            has_gemini = bool(gemini_key and gemini_key != 'WKLEJ_TUTAJ_SWOJ_KLUCZ')
        except ImportError:
            # Fallback na config z DB (stary sposób)
            gemini_key = get_config('gemini_api_key', '')
            has_gemini = bool(gemini_key)
        
        if has_gemini:
            print(f"[SMART_TOY] [AI TITLE] Generuję tytuł przez Gemini...")
            product_data_for_title = {
                'nazwa': nazwa,
                'bullet_points': bullet_points,
                'kategoria': kategoria,
                'asin': asin
            }
            tytul_seo = generate_allegro_title_ai(product_data_for_title, gemini_key, max_length=75)
            if tytul_seo and len(tytul_seo) > 5:
                print(f"   [CHECK_CIRCLE] [SUCCESS] AI Title: {tytul_seo} ({len(tytul_seo)} znaków)")
            else:
                print(f"   ✗ [FAILED] AI nie wygenerował - używam fallback")
                tytul_seo = optimize_title_seo(nazwa, 75)
                print(f"   [EDIT_NOTE] Title (fallback): {tytul_seo}")
        else:
            # Fallback na starą metodę jeśli brak klucza
            print(f"⚠️  [NO API KEY] Brak klucza Gemini - używam fallback")
            tytul_seo = optimize_title_seo(nazwa, 75)
            print(f"   [EDIT_NOTE] Title (fallback): {tytul_seo}")
        
        # <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> TURBO: Generuj opis HTML (NOWA WERSJA - z bullet points + ASIN!)
        try:
            opis_html, opis_plain = generuj_opis_html_pro(nazwa, wszystkie_zdjecia, kategoria, bullet_points, gemini_key=gemini_key, asin=asin)
            print(f"[DESC] Description: {len(opis_html)} chars")
        except Exception as e:
            print(f"⚠️ Description generation failed: {e}")
            # Fallback - prosty opis
            opis_html = f"<p>{nazwa}</p>"
            if bullet_points:
                for bp in bullet_points[:5]:
                    opis_html += f"<p><span class=material-symbols-outlined>check_circle</span> {bp}</p>"
            if asin:
                opis_html += f'<p><i><span class=material-symbols-outlined>bookmark</span> Kod produktu (ASIN): {asin}</i></p>'
            opis_plain = nazwa
        
        # [SHIELD] GPSR: Generuj informacje bezpieczeństwa
        try:
            gpsr = generuj_gpsr_info(nazwa, kategoria, product_specs=product_specs)
            if gpsr:
                print(f"[SHIELD] GPSR: {len(gpsr)} znaków wygenerowane")
            else:
                print(f"   ℹ  GPSR: brak (produkt nie wymaga)")
                gpsr = ""
        except Exception as e:
            print(f"⚠️ GPSR generation failed: {e}")
            gpsr = ""
        
        # <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> KOMBAJN: Zapisz do bazy z retry logic
        def save_to_db():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''UPDATE scraped SET
                nazwa=?, zdjecie_url=?, wszystkie_zdjecia=?, cena_amazon=?,
                tytul_seo=?, opis_html=?, status='gotowy', kategoria=?, bullet_points=?, gpsr=?, product_specs=?
                WHERE asin=?''',
                (nazwa, zdjecie_url, json.dumps(cdn_urls_for_cache or wszystkie_zdjecia), cena_amazon,
                 tytul_seo, opis_html, kategoria, json.dumps(bullet_points), gpsr, json.dumps(product_specs), asin))
            
            # Zaktualizuj w magazynie — nadpisz nazwę jeśli nowa z Amazona jest dłuższa/lepsza
            produkty_do_aktualizacji = cursor.execute(
                'SELECT id, nazwa FROM produkty WHERE asin=? ORDER BY data_dodania ASC',
                (asin,)
            ).fetchall()
            for row in produkty_do_aktualizacji:
                stara = row['nazwa'] or ''
                # Aktualizuj nazwę jeśli: placeholder LUB nowa jest dłuższa (lepsza z Amazona)
                # NIE nadpisuj śmieciowymi nazwami typu "Amazon.com: ..."
                nazwa_ok = nazwa and len(nazwa) > 15 and not nazwa.lower().startswith('amazon')
                nowa_nazwa = nazwa if (nazwa_ok and (len(nazwa) > len(stara) or stara.startswith('Produkt'))) else stara
                cursor.execute('UPDATE produkty SET nazwa=?, zdjecie_url=?, kategoria=?, images=? WHERE id=?',
                    (nowa_nazwa, zdjecie_url, kategoria, json.dumps(wszystkie_zdjecia), row['id']))
            
            conn.commit()
            return True
        
        # Wykonaj z retry (3 próby)
        for retry in range(3):
            try:
                save_to_db()
                break
            except sqlite3.OperationalError as e:
                if 'locked' in str(e) and retry < 2:
                    print(f"⚠️ DB locked, retry {retry+1}/3...")
                    time.sleep(0.5 * (retry + 1))
                else:
                    raise
        
        print(f"[SAVE] Saved to database")
        print(f"[CHECK_CIRCLE] Completed: {asin}")
        
        # Update progress
        PROGRESS['current'] += 1
        
        return (asin, True, None)
        
    except Exception as e:
        print(f"[CANCEL] Error processing {asin}: {e}")
        import traceback
        traceback.print_exc()
        
        PROGRESS['current'] += 1
        PROGRESS['errors'] += 1
        
        return (asin, False, str(e))

def auto_process_products(asins, preferred_domain=None):
    """
    <span class=material-symbols-outlined>agriculture</span> KOMBAJN MODE: Automatycznie przetwarza produkty RÓWNOLEGLE.
    Zamiast 1 produkt na raz, robi 5 naraz = 5x SZYBCIEJ!
    preferred_domain: np. 'co.uk', 'de', 'com' - próbuje tę domenę jako pierwszą
    """
    global _processing_queue, _scraper_running, PROGRESS
    _processing_queue.extend(asins)

    _pref_domain = preferred_domain  # closure

    def process_in_background():
        global _processing_queue, _scraper_running, PROGRESS
        _scraper_running = True

        total = len(_processing_queue)
        PROGRESS = {'current': 0, 'total': total, 'errors': 0}

        print(f"\n{'='*70}")
        print(f"[AGRI] KOMBAJN MODE: Processing {total} products with {MAX_WORKERS} workers")
        if _pref_domain:
            print(f"   Preferred domain: amazon.{_pref_domain}")
        print(f"{'='*70}\n")

        start_time = time.time()

        # <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> RÓWNOLEGŁE PRZETWARZANIE - TO JEST KOMBAJN!
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Uruchom wszystkie taski
            futures = {
                executor.submit(process_single_product, asin, i+1, total, preferred_domain=_pref_domain): asin
                for i, asin in enumerate(_processing_queue)
            }
            
            # Zbieraj wyniki w miarę ukończenia
            for future in as_completed(futures):
                asin = futures[future]
                try:
                    result_asin, success, error = future.result()
                    if success:
                        print(f"[CHECK_CIRCLE] {result_asin} OK")
                    else:
                        print(f"[CANCEL] {result_asin} FAILED: {error}")
                except Exception as e:
                    print(f"[CANCEL] {asin} EXCEPTION: {e}")
        
        _processing_queue.clear()
        _scraper_running = False
        
        elapsed = time.time() - start_time
        success_count = PROGRESS['current'] - PROGRESS['errors']

        print(f"\n{'='*70}")
        print(f"[CELEBRATION] KOMBAJN COMPLETE!")
        print(f"[CHECK_CIRCLE] Success: {success_count}/{total}")
        print(f"[CANCEL] Errors: {PROGRESS['errors']}/{total}")
        print(f"[TIME]  Time: {elapsed:.1f}s ({elapsed/total:.1f}s per product)")
        print(f"<span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Speed: {total/elapsed*60:.1f} products/min")
        print(f"{'='*70}\n")

        # <span class=material-symbols-outlined>smart_toy</span> AUTO-ENHANCE: po scrapowaniu automatycznie generuj zdjęcia AI
        # Pomiń na Pi (za mało RAM/CPU, crashuje system)
        import platform
        is_pi = platform.machine().startswith('a') or 'arm' in platform.machine().lower()
        if success_count > 0 and not is_pi:
            _auto_start_enhance_after_scrape()
        elif is_pi:
            print("⚠️ Pi detected — pomijam auto-enhance (za ciężkie)")

    # Uruchom w osobnym wątku
    thread = threading.Thread(target=process_in_background, daemon=True)
    thread.start()


def _auto_start_enhance_after_scrape():
    """
    <span class=material-symbols-outlined>smart_toy</span> Auto-start generowania zdjęć AI po zakończeniu scrapingu.
    Odpala _bg_enhance_worker z force=True w osobnym wątku.
    Czeka 5s żeby scraper zdążył zapisać wszystko do bazy.
    """
    global _bg_enhance_status
    import time as _t

    # Nie startuj jeśli już działa
    if _bg_enhance_status.get('running'):
        print("[Auto-Enhance] ⏭ Pomijam — generowanie już działa w tle")
        return

    print("[Auto-Enhance] ⏳ Czekam 5s przed startem generowania zdjęć...")
    _t.sleep(5)

    # Sprawdź jeszcze raz
    if _bg_enhance_status.get('running'):
        print("[Auto-Enhance] ⏭ Pomijam — generowanie uruchomione w międzyczasie")
        return

    _bg_enhance_status = {
        'running': True, 'progress': 0, 'current': 0, 'total': 0,
        'done': 0, 'errors': 0, 'cost': 0.0, 'log': [], 'finished': False,
        'started_at': _t.time(), 'last_update': _t.time()
    }

    print("[Auto-Enhance] <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Automatyczny start generowania zdjęć AI!")

    try:
        from flask import current_app
        app = current_app._get_current_object()
    except RuntimeError:
        # Jeśli nie ma kontekstu Flask (np. z wątku) — importuj app
        import sys, os
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app import app

    t = threading.Thread(target=_bg_enhance_worker, args=(app, True), daemon=True)
    t.start()


# ============================================================
# FUNKCJE POMOCNICZE
# ============================================================
_pal_stats_cache = {'data': None, 'time': 0}

def get_stats():
    """Zwraca statystyki Paletomatu (cached 30s)"""
    import time as _time
    now = _time.time()
    if _pal_stats_cache['data'] and (now - _pal_stats_cache['time']) < 30:
        return _pal_stats_cache['data']
    conn = get_db()
    stats = {
        'scraped': conn.execute('SELECT COUNT(*) FROM scraped').fetchone()[0],
        'aktywne': conn.execute('SELECT COUNT(*) FROM oferty WHERE status="aktywna"').fetchone()[0],
        'sprzedane': conn.execute('SELECT COUNT(*) FROM sprzedaze').fetchone()[0],
        'przychod': round(conn.execute('SELECT COALESCE(SUM(CASE WHEN status != "zwrot" THEN cena*ilosc ELSE 0 END), 0) FROM sprzedaze').fetchone()[0]),
    }
    _pal_stats_cache['data'] = stats
    _pal_stats_cache['time'] = _time.time()
    return stats

def scraper_status():
    """Czy scraper jest uruchomiony"""
    global _scraper_running
    return _scraper_running

# ============================================================
# SZABLONY
# ============================================================
CSS = '''<style>
:root{--bg:#06060f;--bg-card:rgba(15,15,30,0.65);--bg-sidebar:#0a0a16;--border:rgba(255,255,255,0.08);--border-light:rgba(255,255,255,0.04);--text:#e2e8f0;--text-muted:#64748b;--accent:#8ff5ff;--accent2:#ff6b9b;--accent-soft:rgba(143,245,255,0.12);--green:#beee00;--green-soft:rgba(190,238,0,0.12);--yellow:#eab308;--yellow-soft:rgba(234,179,8,0.12);--red:#ef4444;--red-soft:rgba(239,68,68,0.12);--blue:#8ff5ff;--blue-soft:rgba(143,245,255,0.12);--cyan:#8ff5ff;--shadow:0 1px 3px rgba(0,0,0,0.2);--shadow-md:0 4px 12px rgba(0,0,0,0.25);--shadow-lg:0 8px 24px rgba(0,0,0,0.3);--radius:16px;--radius-sm:10px}
[data-theme="light"]{--bg:#f0f2f5;--bg-card:#ffffff;--bg-sidebar:#0c0f1a;--border:rgba(0,0,0,0.08);--border-light:rgba(0,0,0,0.04);--text:#1e293b;--text-muted:#64748b;--accent-soft:rgba(143,245,255,0.08);--green-soft:rgba(190,238,0,0.08);--yellow-soft:rgba(234,179,8,0.08);--red-soft:rgba(239,68,68,0.08);--blue-soft:rgba(143,245,255,0.08);--shadow:0 1px 3px rgba(0,0,0,0.06);--shadow-md:0 4px 12px rgba(0,0,0,0.08);--shadow-lg:0 8px 24px rgba(0,0,0,0.12)}
[data-theme="dark"]{--bg:#06060f;--bg-card:rgba(15,15,30,0.65);--bg-sidebar:#0a0a16;--border:rgba(255,255,255,0.08);--border-light:rgba(255,255,255,0.04);--text:#e2e8f0;--text-muted:#64748b;--accent-soft:rgba(143,245,255,0.12);--green-soft:rgba(190,238,0,0.12);--yellow-soft:rgba(234,179,8,0.12);--red-soft:rgba(239,68,68,0.12);--blue-soft:rgba(143,245,255,0.12);--shadow:0 1px 3px rgba(0,0,0,0.2);--shadow-md:0 4px 12px rgba(0,0,0,0.25);--shadow-lg:0 8px 24px rgba(0,0,0,0.3)}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter',-apple-system,system-ui,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}

/* Sidebar */
.app-layout{display:flex;min-height:100vh}
.sidebar{width:250px;background:var(--bg-sidebar);position:fixed;top:0;left:0;bottom:0;display:flex;flex-direction:column;z-index:100;transition:transform 0.3s}
.sidebar-brand{padding:24px 20px 20px;display:flex;align-items:center;gap:12px}
.sidebar-brand-icon{width:36px;height:36px;background:linear-gradient(135deg,#8ff5ff,#beee00);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1.1rem;color:#fff;font-weight:800}
.sidebar-brand-text h1{font-size:0.95rem;color:#fff;font-weight:700;letter-spacing:-0.3px}
.sidebar-brand-text small{font-size:0.62rem;color:rgba(255,255,255,0.4);display:block;margin-top:1px}
.sidebar-nav{flex:1;padding:8px 12px;overflow-y:auto}
.sidebar-section{font-size:0.6rem;text-transform:uppercase;letter-spacing:1.8px;color:rgba(255,255,255,0.25);padding:20px 12px 8px;font-weight:600}
.sidebar-link{display:flex;align-items:center;gap:12px;padding:10px 12px;border-radius:var(--radius-sm);text-decoration:none;color:rgba(255,255,255,0.5);font-size:0.82rem;font-weight:500;transition:all 0.2s;margin-bottom:2px}
.sidebar-link:hover{background:rgba(255,255,255,0.06);color:rgba(255,255,255,0.85)}
.sidebar-link.active{background:linear-gradient(135deg,#8ff5ff,#ff6b9b);color:#fff;font-weight:600;box-shadow:0 4px 12px rgba(143,245,255,0.3)}
.sidebar-link .sl-icon{font-size:1rem;width:22px;text-align:center}
.sidebar-link .sl-badge{margin-left:auto;background:rgba(255,255,255,0.15);color:#fff;font-size:0.58rem;padding:2px 8px;border-radius:10px;font-weight:700}
.sidebar-link .sl-dot{margin-left:auto;width:8px;height:8px;border-radius:50%;background:var(--green);animation:pulse 2s infinite}
.sidebar-bottom{padding:12px;border-top:1px solid rgba(255,255,255,0.08)}
.sidebar-bottom a{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:8px;text-decoration:none;color:rgba(255,255,255,0.4);font-size:0.78rem;transition:all 0.2s}
.sidebar-bottom a:hover{background:rgba(255,255,255,0.06);color:rgba(255,255,255,0.8)}

/* Main */
.main{margin-left:250px;flex:1;min-height:100vh}
.topbar{display:flex;align-items:center;justify-content:space-between;padding:16px 32px;background:var(--bg-card);border-bottom:1px solid var(--border);box-shadow:var(--shadow)}
.topbar-title{font-size:1.05rem;font-weight:700;color:var(--text)}
.topbar-actions{display:flex;gap:10px;align-items:center}
.c{padding:28px 32px;max-width:1200px}

/* Mobile menu */
.menu-toggle{display:none;position:fixed;top:14px;left:14px;z-index:200;background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius-sm);width:42px;height:42px;align-items:center;justify-content:center;cursor:pointer;font-size:1.2rem;box-shadow:var(--shadow-md)}
.sidebar-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.5);z-index:99;backdrop-filter:blur(2px)}

/* Dashboard KPI cards */
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
.kpi-card{background:rgba(15,15,30,0.65);backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.08);border-radius:var(--radius);padding:20px;box-shadow:var(--shadow);transition:all 0.25s;position:relative;overflow:hidden}
.kpi-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.kpi-card::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--radius) var(--radius) 0 0}
.kpi-card.purple::after{background:linear-gradient(90deg,#8ff5ff,#ff6b9b)}
.kpi-card.green::after{background:linear-gradient(90deg,#beee00,#8ff5ff)}
.kpi-card.blue::after{background:linear-gradient(90deg,#8ff5ff,#ff6b9b)}
.kpi-card.orange::after{background:linear-gradient(90deg,#ff6b9b,#8ff5ff)}
.kpi-icon{width:42px;height:42px;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;margin-bottom:14px}
.kpi-card.purple .kpi-icon{background:rgba(143,245,255,0.1)}
.kpi-card.green .kpi-icon{background:rgba(190,238,0,0.1)}
.kpi-card.blue .kpi-icon{background:rgba(143,245,255,0.08)}
.kpi-card.orange .kpi-icon{background:rgba(255,107,155,0.1)}
.kpi-value{font-size:1.6rem;font-weight:800;letter-spacing:-0.5px;color:var(--text);line-height:1;font-family:'Space Grotesk','Inter',sans-serif}
.kpi-label{font-size:0.7rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.5px;margin-top:6px;font-weight:500}
.kpi-change{display:inline-flex;align-items:center;gap:3px;font-size:0.68rem;font-weight:600;margin-top:8px;padding:2px 8px;border-radius:20px}
.kpi-change.up{background:var(--green-soft);color:var(--green)}
.kpi-change.down{background:var(--red-soft);color:var(--red)}

/* Dashboard grid */
.dash-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}
.dash-grid-3{display:grid;grid-template-columns:2fr 1fr;gap:20px;margin-bottom:24px}

/* Cards */
.card{background:rgba(15,15,30,0.65);backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.08);border-radius:var(--radius);padding:22px;margin-bottom:20px;box-shadow:var(--shadow);transition:all 0.25s}
.card:hover{box-shadow:var(--shadow-md);border-color:rgba(143,245,255,0.15)}
.card-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}
.card-title{font-weight:700;font-size:0.92rem;color:var(--text)}
.card-subtitle{font-size:0.72rem;color:var(--text-muted)}
.card-action{font-size:0.75rem;color:#8ff5ff;text-decoration:none;font-weight:600;transition:color 0.2s}
.card-action:hover{color:#ff6b9b}

/* Old compat */
.hdr{padding:0 0 20px;margin-bottom:20px;border-bottom:1px solid var(--border)}
.hdr h1{font-size:1.2rem;color:var(--text);font-weight:700}
.hdr small{color:var(--text-muted);font-size:0.75rem}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px}
.stat{background:rgba(15,15,30,0.65);backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.08);border-radius:var(--radius);padding:20px;text-align:center;transition:all 0.25s;box-shadow:var(--shadow)}
.stat:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.stat-v{font-size:1.5rem;font-weight:800;color:#8ff5ff;letter-spacing:-0.5px}
.stat-v.green{color:var(--green)}
.stat-l{font-size:0.68rem;color:var(--text-muted);text-transform:uppercase;margin-top:4px;letter-spacing:0.5px}
.status{display:flex;align-items:center;justify-content:space-between;padding:16px 20px;border-radius:var(--radius);margin-bottom:16px;transition:all 0.25s;box-shadow:var(--shadow)}
.status.on{background:rgba(190,238,0,0.06);border:1px solid rgba(190,238,0,0.2)}
.status.off{background:var(--bg-card);border:1px solid var(--border)}
.status-info{display:flex;align-items:center;gap:10px;font-weight:500}
.status-dot{width:10px;height:10px;border-radius:50%}
.status-dot.on{background:#beee00;box-shadow:0 0 10px rgba(190,238,0,0.4);animation:pulse 2s infinite}
.status-dot.off{background:var(--text-muted)}
@keyframes pulse{0%,100%{opacity:1;box-shadow:0 0 10px rgba(190,238,0,0.4)}50%{opacity:0.6;box-shadow:0 0 4px rgba(190,238,0,0.2)}}
@keyframes fadeIn{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
@keyframes slideIn{from{opacity:0;transform:translateX(-10px)}to{opacity:1;transform:translateX(0)}}

/* Buttons */
.btn{display:block;width:100%;padding:12px;font-size:0.92rem;font-weight:600;text-align:center;text-decoration:none;border:none;border-radius:12px;cursor:pointer;margin-bottom:8px;color:#fff;transition:all 0.2s}
.btn:hover{transform:translateY(-1px);box-shadow:var(--shadow-md)}
.btn-p{background:rgba(143,245,255,0.12);border:1px solid rgba(143,245,255,0.3);color:#8ff5ff}
.btn-ok{background:rgba(190,238,0,0.12);border:1px solid rgba(190,238,0,0.3);color:#beee00}
.btn-2{background:var(--bg-card);border:1px solid var(--border);color:var(--text)}
.btn-sm{padding:8px 16px;font-size:0.82rem;width:auto;display:inline-block;border-radius:8px}

/* Module links */
.module{display:flex;align-items:center;background:rgba(15,15,30,0.65);backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.08);border-radius:var(--radius);padding:16px 18px;margin-bottom:10px;text-decoration:none;color:var(--text);transition:all 0.25s;box-shadow:var(--shadow);animation:fadeIn 0.4s ease both}
.module:hover{border-color:rgba(143,245,255,0.2);transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.module-icon{font-size:1.4rem;margin-right:14px;width:42px;height:42px;display:flex;align-items:center;justify-content:center;background:rgba(143,245,255,0.08);border-radius:12px}
.module-info{flex:1}
.module-name{font-weight:700;font-size:0.92rem}
.module-desc{font-size:0.73rem;color:var(--text-muted);margin-top:3px}
.module-arrow{color:var(--text-muted);font-size:1.2rem;transition:transform 0.2s}
.module:hover .module-arrow{transform:translateX(3px);color:#8ff5ff}

/* Items */
.item{display:flex;align-items:center;background:rgba(15,15,30,0.65);backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:12px;margin-bottom:8px;text-decoration:none;color:var(--text);transition:all 0.2s;box-shadow:var(--shadow)}
.item:hover{border-color:rgba(143,245,255,0.15);box-shadow:var(--shadow-md)}
.item img{width:48px;height:48px;object-fit:contain;background:#fff;border-radius:8px;margin-right:12px}
.item-dot{width:8px;height:8px;border-radius:50%;margin-right:10px;flex-shrink:0}
.item-dot.green{background:#beee00;box-shadow:0 0 6px rgba(190,238,0,0.3)}
.item-dot.yellow{background:var(--yellow);box-shadow:0 0 6px rgba(234,179,8,0.3)}
.item-info{flex:1;min-width:0}
.item-name{font-weight:600;font-size:0.85rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.item-meta{font-size:0.7rem;color:var(--text-muted);margin-top:2px}
.item-right{text-align:right;margin-left:8px}
.item-price{font-weight:700;color:#beee00}
.item-margin{font-size:0.7rem;color:#8ff5ff}

/* Forms */
.form-group{margin-bottom:12px}
.form-group label{display:block;font-size:0.75rem;color:var(--text-muted);margin-bottom:4px;font-weight:600;text-transform:uppercase;letter-spacing:0.3px}
.form-ctrl{width:100%;padding:11px 14px;background:var(--bg);border:1px solid var(--border);border-radius:var(--radius-sm);color:var(--text);font-size:0.92rem;transition:all 0.2s}
.form-ctrl:focus{border-color:#8ff5ff;outline:none;box-shadow:0 0 0 3px rgba(143,245,255,0.1)}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.section{color:var(--text);font-weight:700;font-size:0.82rem;margin:24px 0 12px;display:flex;align-items:center;gap:8px;text-transform:uppercase;letter-spacing:0.5px}
.alert{padding:14px 18px;border-radius:12px;margin-bottom:14px;text-align:center;font-size:0.88rem;font-weight:500;box-shadow:var(--shadow)}
.alert-ok{background:rgba(190,238,0,0.08);border:1px solid rgba(190,238,0,0.15);color:#beee00}
.alert-warn{background:var(--yellow-soft);border:1px solid rgba(234,179,8,0.15);color:var(--yellow)}
.back{display:block;text-align:center;color:var(--text-muted);text-decoration:none;padding:14px;font-size:0.85rem;transition:color 0.2s}
.back:hover{color:#8ff5ff}
.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:0.68rem;font-weight:700}
.badge-ok{background:var(--green-soft);color:var(--green)}
.badge-warn{background:var(--yellow-soft);color:var(--yellow)}

/* Activity list */
.activity-item{display:flex;align-items:flex-start;gap:12px;padding:12px 0;border-bottom:1px solid var(--border-light)}
.activity-item:last-child{border-bottom:none}
.activity-icon{width:34px;height:34px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:0.9rem;flex-shrink:0}
.activity-icon.green{background:var(--green-soft)}
.activity-icon.blue{background:var(--blue-soft)}
.activity-icon.purple{background:var(--accent-soft)}
.activity-icon.orange{background:var(--yellow-soft)}
.activity-text{flex:1;min-width:0}
.activity-title{font-weight:600;font-size:0.82rem}
.activity-time{font-size:0.7rem;color:var(--text-muted);margin-top:2px}

/* Quick action buttons */
.quick-actions{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.qa-btn{display:flex;align-items:center;gap:10px;padding:14px 16px;background:var(--bg);border:1px solid var(--border);border-radius:var(--radius-sm);text-decoration:none;color:var(--text);font-size:0.82rem;font-weight:600;transition:all 0.2s}
.qa-btn:hover{border-color:rgba(143,245,255,0.3);background:rgba(143,245,255,0.08);transform:translateY(-1px)}
.qa-btn .qa-icon{width:36px;height:36px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1rem}

/* Hide old nav */
.nav{display:none}

/* Theme toggle */
.theme-toggle{background:var(--bg);border:1px solid var(--border);border-radius:var(--radius-sm);width:38px;height:38px;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:1.1rem;transition:all 0.3s}
.theme-toggle:hover{transform:scale(1.08);border-color:#8ff5ff}

/* Responsive */
@media(max-width:900px){
    .sidebar{transform:translateX(-100%)}
    .sidebar.open{transform:translateX(0)}
    .sidebar-overlay.open{display:block}
    .menu-toggle{display:flex}
    .main{margin-left:0}
    .topbar{padding-left:60px}
    .c{padding:16px}
    .kpi-grid{grid-template-columns:repeat(2,1fr)}
    .dash-grid,.dash-grid-3{grid-template-columns:1fr}
}
@media(max-width:600px){
    .stats{grid-template-columns:1fr}
    .form-row{grid-template-columns:1fr}
    .kpi-grid{grid-template-columns:1fr}
    .quick-actions{grid-template-columns:1fr}
    .c{padding:12px}
}
</style>'''

BASE = '''<!DOCTYPE html><html lang="pl" data-theme="dark"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Paletomat</title><meta name="theme-color" content="#0a0a0f">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="mobile-web-app-capable" content="yes">
<meta name="application-name" content="Akces Hub">
<link rel="manifest" href="/manifest.json">
<link rel="apple-touch-icon" href="/static/icon-192.png">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap">
<link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap" rel="stylesheet">
<style>.mi{font-family:'Material Symbols Outlined';font-weight:normal;font-style:normal;font-size:inherit;line-height:1;letter-spacing:normal;text-transform:none;display:inline-block;white-space:nowrap;word-wrap:normal;direction:ltr;-webkit-font-smoothing:antialiased;vertical-align:middle}</style>
<script>
const saved = localStorage.getItem('theme');
if(saved) document.documentElement.setAttribute('data-theme', saved);
// PWA: rejestracja service workera
if('serviceWorker' in navigator){
    navigator.serviceWorker.getRegistrations().then(function(regs){
        regs.forEach(function(r){
            if(r.scope.includes('/static')){r.unregister();}
            else{r.update();}
        });
    });
    navigator.serviceWorker.register('/sw.js?v=15', {scope: '/'}).then(function(reg){
        reg.update();
        reg.addEventListener('updatefound',function(){
            var nw=reg.installing;
            if(nw) nw.addEventListener('statechange',function(){
                if(nw.state==='installed' && navigator.serviceWorker.controller){
                    nw.postMessage({action:'skipWaiting'});
                }
            });
        });
    }).catch(function(e){console.log('SW fail',e);});
    var refreshing=false;
    navigator.serviceWorker.addEventListener('controllerchange',function(){
        if(!refreshing){refreshing=true; location.reload();}
    });
}
</script>''' + CSS + '''<link rel="stylesheet" href="/static/kiosk.css"></head><body>
<script>if(localStorage.getItem('kiosk_mode')==='1')document.body.classList.add('kiosk');</script>

<!-- Mobile menu toggle -->
<div class="menu-toggle" onclick="document.querySelector('.sidebar').classList.toggle('open');document.querySelector('.sidebar-overlay').classList.toggle('open')">☰</div>
<div class="sidebar-overlay" onclick="document.querySelector('.sidebar').classList.remove('open');this.classList.remove('open')"></div>

<div class="app-layout">
<!-- Sidebar -->
<aside class="sidebar">
    <div class="sidebar-brand">
        <div class="sidebar-brand-icon">P</div>
        <div class="sidebar-brand-text">
            <h1>Paletomat</h1>
            <small>v32.0 · <span style="color:__PLAN_COLOR__;font-weight:700">__PLAN_NAME__</span></small>
        </div>
    </div>
    <nav class="sidebar-nav">
        <div class="sidebar-section">Glowne</div>
        <a href="/paletomat" class="sidebar-link active">
            <span class="sl-icon"><span class=material-symbols-outlined>bar_chart</span></span>Dashboard
        </a>
        <a href="/paletomat/scraper" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>language</span></span>Amazon Scraper
        </a>
        <a href="/paletomat/generator" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>label</span></span>Generator ofert
        </a>

        <div class="sidebar-section">Sprzedaz</div>
        <a href="/paletomat/oferty" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>edit_note</span></span>Moje oferty
        </a>
        <a href="/paletomat/monitoring" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>trending_up</span></span>Monitoring
        </a>
        <a href="/telegram/live" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>paid</span></span>Sprzedaz LIVE
        </a>

        <div class="sidebar-section">Narzedzia</div>
        <a href="/palety/bulk-import" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>inventory_2</span></span>Bulk import
        </a>
        <a href="/analytics/profit" class="sidebar-link">
            <span class="sl-icon">💹</span>Profit Analyzer
        </a>
        <a href="/analityka/koszty-allegro" class="sidebar-link">
            <span class="sl-icon"><span class=material-symbols-outlined>paid</span></span>Koszty Allegro
        </a>
    </nav>
    <div class="sidebar-bottom">
        <a href="/narzedzia"><span class="sl-icon"><span class=material-symbols-outlined>bolt</span></span>Narzedzia</a>
        <a href="/"><span class="sl-icon"><span class=material-symbols-outlined>home</span></span>Strona glowna</a>
    </div>
</aside>

<!-- Main content -->
<div class="main">
<div class="topbar">
    <div class="topbar-title">Dashboard</div>
    <div class="topbar-actions">
        <div class="theme-toggle" onclick="toggleTheme()" title="Zmien motyw">
            <span id="theme-icon"><span class=material-symbols-outlined>dark_mode</span></span>
        </div>
    </div>
</div>
<div class="c">{content}</div>
</div>
</div>

<script>
function toggleTheme(){
    const html = document.documentElement;
    const current = html.getAttribute('data-theme');
    const next = current === 'dark' ? 'light' : 'dark';
    html.setAttribute('data-theme', next);
    localStorage.setItem('theme', next);
    document.getElementById('theme-icon').textContent = next === 'dark' ? '' : '☀';
    document.querySelector('meta[name="theme-color"]').content = next === 'dark' ? '#06060f' : '#f1f5f9';
}
const theme = document.documentElement.getAttribute('data-theme');
document.getElementById('theme-icon').textContent = theme === 'dark' ? '' : '☀';

// Highlight active sidebar link
const path = window.location.pathname;
document.querySelectorAll('.sidebar-link').forEach(function(link) {
    link.classList.remove('active');
    if (link.getAttribute('href') === path) link.classList.add('active');
});
</script>
</body></html>'''

_cached_plan = {'name': None, 'color': None, 'ts': 0}

def _get_plan_display():
    """Return (plan_name, plan_color) with 60s cache."""
    import time as _time
    now = _time.time()
    if _cached_plan['name'] and now - _cached_plan['ts'] < 60:
        return _cached_plan['name'], _cached_plan['color']
    try:
        from .license import check_license
        _valid, _plan, _msg = check_license()
        plan_name = (_plan or 'free').upper()
    except Exception:
        plan_name = 'FREE'
    plan_colors = {'STARTER': '#64748b', 'PRO': '#8ff5ff', 'BUSINESS': '#beee00', 'ENTERPRISE': '#f59e0b', 'FREE': '#ef4444'}
    plan_color = plan_colors.get(plan_name, '#64748b')
    _cached_plan.update(name=plan_name, color=plan_color, ts=now)
    return plan_name, plan_color

def render(content):
    plan_name, plan_color = _get_plan_display()
    html = BASE.replace('{content}', content)
    html = html.replace('__PLAN_NAME__', plan_name)
    html = html.replace('__PLAN_COLOR__', plan_color)
    return html

# ============================================================
# ROUTES
# ============================================================
@paletomat_bp.route('/')
def index():
    s = get_stats()
    is_running = scraper_status()
    
    # Sprawdź status kolejki auto-processingu
    global _processing_queue, _scraper_running
    queue_len = len(_processing_queue)
    auto_running = _scraper_running
    
    # Pobierz ostatnio scrapowane
    conn = get_db()
    scraped = conn.execute('SELECT * FROM scraped ORDER BY data_scrape DESC LIMIT 5').fetchall()
    oferty = conn.execute('SELECT * FROM oferty WHERE status="aktywna" ORDER BY data_aktualizacji DESC LIMIT 5').fetchall()
    
    status_class = 'on' if is_running else 'off'
    
    # Badge auto-processingu
    auto_badge = ''
    if auto_running or queue_len > 0:
        auto_status_class = 'on' if auto_running else 'off'
        auto_text = f"<span class=material-symbols-outlined>sync</span> Przetwarzanie w tle: {queue_len} produktów" if auto_running else f"⏸ Kolejka: {queue_len} produktów"
        auto_badge = f'''
        <div class="status {auto_status_class}" id="auto-status" style="margin-top:10px">
            <div class="status-info">
                <div class="status-dot {auto_status_class}"></div>
                <span id="auto-text">{auto_text}</span>
            </div>
        </div>'''
    
    # Dodatkowe dane do dashboardu (combined query)
    _sp = conn.execute('''
        SELECT
            COALESCE(SUM(CASE WHEN data_sprzedazy >= date('now', '-7 days') THEN cena * ilosc END), 0),
            COALESCE(SUM(cena * ilosc), 0)
        FROM sprzedaze
        WHERE data_sprzedazy >= date('now', '-30 days')
          AND status NOT IN ('zwrot','anulowane','anulowana')
    ''').fetchone()
    sprzedaz_7d, sprzedaz_30d = _sp[0], _sp[1]
    nowe_7d = conn.execute("SELECT COUNT(*) FROM scraped WHERE data_scrape >= date('now', '-7 days')").fetchone()[0]

    # Dane do wykresu — produkty dodane dziennie (ostatnie 30 dni)
    chart_rows = conn.execute('''
        SELECT date(data_scrape) as dzien, COUNT(*) as cnt
        FROM scraped
        WHERE data_scrape >= date('now', '-30 days')
        GROUP BY date(data_scrape)
        ORDER BY dzien
    ''').fetchall()
    # Buduj dane wykresu (uzupełnij brakujące dni zerami)
    _today = datetime.now().date()
    _chart_map = {r['dzien']: r['cnt'] for r in chart_rows}
    chart_labels = []
    chart_data = []
    chart_cumulative = []
    _running_total = s['scraped'] - sum(r['cnt'] for r in chart_rows)  # start = total - last 30d
    for i in range(30):
        d = (_today - timedelta(days=29 - i)).strftime('%Y-%m-%d')
        cnt = _chart_map.get(d, 0)
        _running_total += cnt
        chart_labels.append(d[5:])  # MM-DD
        chart_data.append(cnt)
        chart_cumulative.append(_running_total)

    # Statystyki do mini kart pod wykresem
    w_magazynie = conn.execute("SELECT COUNT(*) FROM produkty WHERE status IN ('magazyn','wystawiony') AND ilosc > 0").fetchone()[0]
    sprzedane = conn.execute("SELECT COUNT(*) FROM sprzedaze WHERE status NOT IN ('zwrot','anulowane','anulowana') AND data_sprzedazy >= date('now', '-30 days')").fetchone()[0]
    zalegajace = conn.execute("SELECT COUNT(*) FROM produkty WHERE status IN ('magazyn','wystawiony') AND date(data_dodania) < date('now', '-30 days') AND ilosc > 0").fetchone()[0]

    # Ostatnie sprzedaże
    ostatnie_sprzedaze = conn.execute('''
        SELECT s.cena, s.ilosc, s.data_sprzedazy,
               COALESCE(p.nazwa, s.nazwa, 'Produkt') as nazwa
        FROM sprzedaze s LEFT JOIN produkty p ON s.produkt_id = p.id
        WHERE s.status NOT IN ('zwrot','anulowane','anulowana')
        ORDER BY s.data_sprzedazy DESC LIMIT 6
    ''').fetchall()

    return render_template('paletomat_dashboard.html',
        stats=s,
        is_running=is_running,
        auto_running=auto_running,
        queue_len=queue_len,
        sprzedaz_7d=sprzedaz_7d,
        sprzedaz_30d=sprzedaz_30d,
        nowe_7d=nowe_7d,
        chart_labels=chart_labels,
        chart_data=chart_data,
        chart_cumulative=chart_cumulative,
        w_magazynie=w_magazynie,
        sprzedane=sprzedane,
        zalegajace=zalegajace,
        ostatnie_sprzedaze=ostatnie_sprzedaze,
        scraped=scraped,
        get_amazon_image_url=get_amazon_image_url,
    )

@paletomat_bp.route('/scraper')
def scraper():
    # Pobierz listę palet z magazynu
    conn = get_db()
    # Pobierz palety które:
    # 1. Mają produkty ALBO
    # 2. Są nowe (dodane w ostatnich 30 dniach)
    palety = conn.execute('''
        SELECT DISTINCT p.id, p.nazwa, p.dostawca
        FROM palety p
        LEFT JOIN produkty pr ON p.id = pr.paleta_id
        WHERE pr.id IS NOT NULL
           OR p.data_dodania >= date('now', '-30 days')
        ORDER BY p.data_dodania DESC
    ''').fetchall()

    # === METRYKI SCRAPERA ===
    try:
        total_produkty = conn.execute('SELECT COUNT(*) as c FROM produkty').fetchone()['c']
        new_today = conn.execute("SELECT COUNT(*) as c FROM produkty WHERE date(data_dodania)=date('now')").fetchone()['c']
        new_week = conn.execute("SELECT COUNT(*) as c FROM produkty WHERE data_dodania >= date('now','-7 days')").fetchone()['c']
        total_palety = conn.execute('SELECT COUNT(*) as c FROM palety').fetchone()['c']
        with_asin = conn.execute("SELECT COUNT(*) as c FROM produkty WHERE asin IS NOT NULL AND asin != ''").fetchone()['c']
        with_photo = conn.execute("SELECT COUNT(*) as c FROM produkty WHERE zdjecie_url IS NOT NULL AND zdjecie_url != ''").fetchone()['c']
        asin_pct = round(with_asin / total_produkty * 100) if total_produkty else 0
        photo_pct = round(with_photo / total_produkty * 100) if total_produkty else 0
    except:
        total_produkty = new_today = new_week = total_palety = asin_pct = photo_pct = 0

    # === OSTATNIO DODANE PRODUKTY ===
    try:
        ostatnio_rows = conn.execute('''
            SELECT pr.id, pr.nazwa, pr.asin, pr.ilosc, pr.stan,
                   COALESCE(p.nazwa, '') as paleta_nazwa,
                   COALESCE(p.dostawca, '') as dostawca,
                   pr.data_dodania
            FROM produkty pr
            LEFT JOIN palety p ON pr.paleta_id = p.id
            WHERE pr.data_dodania IS NOT NULL
            ORDER BY pr.data_dodania DESC LIMIT 12
        ''').fetchall()
        ostatnio_dodane = [{
            'id': r['id'],
            'nazwa': (r['nazwa'] or 'Brak nazwy')[:45],
            'asin': r['asin'] or '—',
            'ilosc': r['ilosc'] or 1,
            'stan': r['stan'] or 'Nowy',
            'paleta': (r['paleta_nazwa'] or '—')[:22],
            'dostawca': (r['dostawca'] or '—')[:16],
            'data': str(r['data_dodania'] or '')[:10],
        } for r in ostatnio_rows]
    except:
        ostatnio_dodane = []

    # Dropdown palet (pełny - dla formularza ASIN)
    palety_options = '<option value="">-- Bez palety --</option>'
    palety_options += '<option value="new">➕ Nowa paleta...</option>'
    for p in palety:
        palety_options += f'<option value="{p["id"]}">{p["nazwa"]} ({p["dostawca"] or "brak dostawcy"})</option>'

    # Dropdown palet (tylko lista - dla formularza FILE, bez "Bez palety" i "Nowa")
    palety_options_clean = ''
    for p in palety:
        palety_options_clean += f'<option value="{p["id"]}">{p["nazwa"]} ({p["dostawca"] or "brak dostawcy"})</option>'

    # Build recently discovered table rows
    disc_rows = ''
    for r in ostatnio_dodane:
        stan_color = {'Nowy':'#beee00','Używany':'#eab308','Uszkodzony':'#ef4444'}.get(r['stan'],'#8ff5ff')
        disc_rows += (
            f'<tr style="border-left:2px solid rgba(143,245,255,0.1)">'
            f'<td style="padding:10px 8px 10px 16px;font-size:0.75rem;color:#64748b;white-space:nowrap">{r["data"]}</td>'
            f'<td style="max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-weight:600;font-size:0.82rem">{r["nazwa"]}</td>'
            f'<td style="font-size:0.72rem;color:#8ff5ff;font-family:monospace">{r["asin"]}</td>'
            f'<td style="font-size:0.75rem;color:#64748b">{r["ilosc"]} szt</td>'
            f'<td style="font-size:0.72rem;color:#64748b">{r["paleta"]}</td>'
            f'<td style="padding-right:16px"><span style="font-size:0.65rem;font-weight:700;color:{stan_color};background:{stan_color}18;padding:2px 8px;text-transform:uppercase;letter-spacing:0.5px">{r["stan"]}</span></td>'
            f'</tr>'
        )

    # Pre-build discovered table to avoid backslash-in-f-string SyntaxError
    _th = '<th style="font-size:0.6rem;text-transform:uppercase;letter-spacing:1px;color:#64748b;padding:8px;text-align:left;border-bottom:1px solid rgba(255,255,255,0.06);font-weight:600">'
    _disc_table = (
        '<table style="width:100%;border-collapse:collapse"><thead><tr>'
        + '<th style="font-size:0.6rem;text-transform:uppercase;letter-spacing:1px;color:#64748b;padding:8px 8px 8px 16px;text-align:left;border-bottom:1px solid rgba(255,255,255,0.06);font-weight:600">Data</th>'
        + _th + 'Produkt</th>'
        + _th + 'ASIN</th>'
        + _th + 'Ilosc</th>'
        + _th + 'Paleta</th>'
        + '<th style="font-size:0.6rem;text-transform:uppercase;letter-spacing:1px;color:#64748b;padding:8px 16px 8px 8px;text-align:left;border-bottom:1px solid rgba(255,255,255,0.06);font-weight:600">Stan</th>'
        + '</tr></thead><tbody>'
        + (disc_rows if disc_rows else '<tr><td colspan="6" style="padding:24px;text-align:center;color:#64748b;font-size:0.82rem">Brak produktow</td></tr>')
        + '</tbody></table>'
    ) if ostatnio_dodane else '<div style="padding:24px;text-align:center;color:#64748b;font-size:0.82rem">Brak produktow w bazie</div>'

    html = f'''
    <!-- Page Header -->
    <div style="display:flex;align-items:flex-end;justify-content:space-between;margin-bottom:24px;border-left:3px solid #8ff5ff;padding-left:20px">
        <div>
            <h2 style="font-family:'Space Grotesk',sans-serif;font-size:2.2rem;font-weight:800;color:#f9f5f8;letter-spacing:-0.03em;margin:0;text-transform:uppercase">SCRAPER_HUB</h2>
            <div style="font-size:0.72rem;color:#adaaad;letter-spacing:0.1em;text-transform:uppercase;font-weight:600;margin-top:4px">Process_Node: Alpha-9 | Status: <span style="color:#beee00">OPERATIONAL</span> | Total_Items: <span style="color:#8ff5ff">{total_produkty}</span></div>
        </div>
        <div style="text-align:right;font-family:'Space Grotesk',sans-serif">
            <div style="font-size:1.5rem;font-weight:800;color:#beee00">+{new_today}</div>
            <div style="font-size:0.6rem;color:#64748b;text-transform:uppercase;letter-spacing:1px">Dzisiaj</div>
        </div>
    </div>

    <!-- METRICS GAUGES -->
    <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:24px">
        <div style="background:rgba(15,15,30,0.6);border-left:3px solid #8ff5ff;padding:16px 14px">
            <div style="font-size:0.58rem;text-transform:uppercase;letter-spacing:1.2px;color:#64748b;font-weight:600;margin-bottom:8px">Total Items</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800;color:#8ff5ff">{total_produkty}</div>
        </div>
        <div style="background:rgba(15,15,30,0.6);border-left:3px solid #beee00;padding:16px 14px">
            <div style="font-size:0.58rem;text-transform:uppercase;letter-spacing:1.2px;color:#64748b;font-weight:600;margin-bottom:8px">Nowe (7 dni)</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800;color:#beee00">+{new_week}</div>
        </div>
        <div style="background:rgba(15,15,30,0.6);border-left:3px solid rgba(143,245,255,0.3);padding:16px 14px">
            <div style="font-size:0.58rem;text-transform:uppercase;letter-spacing:1.2px;color:#64748b;font-weight:600;margin-bottom:8px">Palety</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800">{total_palety}</div>
        </div>
        <div style="background:rgba(15,15,30,0.6);border-left:3px solid rgba(143,245,255,0.3);padding:16px 14px">
            <div style="font-size:0.58rem;text-transform:uppercase;letter-spacing:1.2px;color:#64748b;font-weight:600;margin-bottom:8px">ASIN Coverage</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800;color:{"#beee00" if asin_pct>=80 else "#eab308" if asin_pct>=50 else "#ef4444"}">{asin_pct}%</div>
        </div>
        <div style="background:rgba(15,15,30,0.6);border-left:3px solid rgba(255,107,155,0.3);padding:16px 14px">
            <div style="font-size:0.58rem;text-transform:uppercase;letter-spacing:1.2px;color:#64748b;font-weight:600;margin-bottom:8px">Photo Coverage</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:800;color:{"#beee00" if photo_pct>=80 else "#eab308" if photo_pct>=50 else "#ff6b9b"}">{photo_pct}%</div>
        </div>
    </div>

    <!-- Action Buttons Row -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px">
        <button onclick="document.getElementById('form-asin').scrollIntoView({{behavior:'smooth'}})" style="background:#262528;border:1px solid rgba(72,71,74,0.2);padding:16px;cursor:pointer;text-align:left;transition:background 0.2s" onmouseover="this.style.background='#2c2c2f'" onmouseout="this.style.background='#262528'">
            <span style="display:block;font-size:0.6rem;color:#adaaad;text-transform:uppercase;margin-bottom:4px;font-weight:600">SCRAPER</span>
            <span style="display:block;font-family:'Space Grotesk',sans-serif;font-size:1rem;color:#8ff5ff;font-weight:700">ASIN_SCAN</span>
        </button>
        <button onclick="document.getElementById('form-file').scrollIntoView({{behavior:'smooth'}})" style="background:#262528;border:1px solid rgba(72,71,74,0.2);padding:16px;cursor:pointer;text-align:left;transition:background 0.2s" onmouseover="this.style.background='#2c2c2f'" onmouseout="this.style.background='#262528'">
            <span style="display:block;font-size:0.6rem;color:#adaaad;text-transform:uppercase;margin-bottom:4px;font-weight:600">IMPORT</span>
            <span style="display:block;font-family:'Space Grotesk',sans-serif;font-size:1rem;color:#f9f5f8;font-weight:700">FILE_LOAD</span>
        </button>
        <button onclick="document.getElementById('form-miglo').scrollIntoView({{behavior:'smooth'}})" style="background:#262528;border:1px solid rgba(72,71,74,0.2);padding:16px;cursor:pointer;text-align:left;transition:background 0.2s" onmouseover="this.style.background='#2c2c2f'" onmouseout="this.style.background='#262528'">
            <span style="display:block;font-size:0.6rem;color:#adaaad;text-transform:uppercase;margin-bottom:4px;font-weight:600">MIGLO</span>
            <span style="display:block;font-family:'Space Grotesk',sans-serif;font-size:1rem;color:#f9f5f8;font-weight:700">MANUAL_IN</span>
        </button>
        <a href="/paletomat/generator" style="background:#262528;border:1px solid rgba(72,71,74,0.2);padding:16px;text-decoration:none;display:block;transition:background 0.2s" onmouseover="this.style.background='#2c2c2f'" onmouseout="this.style.background='#262528'">
            <span style="display:block;font-size:0.6rem;color:#adaaad;text-transform:uppercase;margin-bottom:4px;font-weight:600">OFFERS</span>
            <span style="display:block;font-family:'Space Grotesk',sans-serif;font-size:1rem;color:#ff6b9b;font-weight:700">SYNC_HUB</span>
        </a>
    </div>

    <!-- FORMULARZ 1: SCRAPE PO ASIN -->
    <div id="form-asin" style="background:#1f1f22;border:1px solid rgba(72,71,74,0.15);padding:24px;margin-bottom:16px">
        <div style="font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;margin-bottom:16px;display:flex;align-items:center;gap:8px">
            <span class=material-symbols-outlined style="color:#8ff5ff">search</span> SCRAPE PO ASIN
        </div>
        <form action="/paletomat/scraper/asin" method="POST">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px;padding:12px;background:#131315;border:1px solid rgba(255,107,155,0.15);border-radius:6px">
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">inventory_2</span> Paleta</label>
                    <select name="paleta_id" class="form-ctrl" onchange="this.form.nowa_paleta_nazwa.style.display = this.value === 'new' ? 'block' : 'none'">
                        <option value="">-- Bez palety --</option>
                        <option value="new">+ Nowa paleta...</option>
                        {palety_options_clean}
                    </select>
                </div>
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">factory</span> Dostawca</label>
                    <select name="dostawca" class="form-ctrl">
                        <option value="">-- Wybierz --</option>
                        <option value="Jobalots">Jobalots</option>
                        <option value="Warrington">Warrington</option>
                        <option value="Miglo">Miglo</option>
                        <option value="Inny">Inny</option>
                    </select>
                </div>
            </div>
            <input type="text" name="nowa_paleta_nazwa" class="form-ctrl" placeholder="Nazwa nowej palety" style="display:none;margin-bottom:12px">
            <div class="form-group">
                <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">ASIN-y (format: ASIN lub ASIN,ilość)</label>
                <input type="text" name="asins" class="form-ctrl" placeholder="B0CFQBBT7G B088ZQ6B64,3 B0ABC12345,2">
                <div style="font-size:0.62rem;color:#767577;margin-top:4px">Miglo: ASIN,ilość rozdzielone spacjami</div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Domena Amazon</label>
                    <select name="domain" class="form-ctrl">
                        <option value="de">Amazon.de</option>
                        <option value="co.uk">Amazon.co.uk</option>
                        <option value="com">Amazon.com</option>
                        <option value="pl">Amazon.pl</option>
                    </select>
                </div>
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">paid</span> Cena jedn. (opcjonalnie)</label>
                    <input type="number" step="0.01" name="cena_jednostkowa" class="form-ctrl" placeholder="np. 25.50">
                </div>
            </div>
            <button type="submit" class="btn" style="width:100%;padding:14px;background:#8ff5ff;color:#005d63;font-weight:700;font-size:0.78rem;letter-spacing:0.1em;text-transform:uppercase;display:flex;align-items:center;justify-content:center;gap:6px;box-shadow:0 0 15px rgba(143,245,255,0.3)"><span class=material-symbols-outlined>search</span> SCRAPUJ ASIN-y</button>
        </form>
    </div>

    <!-- FORMULARZ 2: SCRAPE Z PLIKU -->
    <div id="form-file" style="background:#1f1f22;border:1px solid rgba(72,71,74,0.15);padding:24px;margin-bottom:16px">
        <div style="font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;margin-bottom:16px;display:flex;align-items:center;gap:8px">
            <span class=material-symbols-outlined style="color:#8ff5ff">folder</span> SCRAPE Z PLIKU
            <span style="font-size:0.6rem;color:#767577;font-weight:400;margin-left:4px">(JOBALOTS / WARRINGTON)</span>
        </div>
        <form action="/paletomat/scraper/file" method="POST" enctype="multipart/form-data">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px;padding:12px;background:#131315;border:1px solid rgba(143,245,255,0.15);border-radius:6px">
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">inventory_2</span> Paleta</label>
                    <select name="paleta_id" class="form-ctrl" onchange="this.form.nowa_paleta_nazwa.style.display = this.value === 'new' ? 'block' : 'none'">
                        <option value="">-- Bez palety --</option>
                        <option value="new">+ Nowa paleta...</option>
                        {palety_options_clean}
                    </select>
                </div>
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">factory</span> Dostawca</label>
                    <select name="dostawca" class="form-ctrl">
                        <option value="">-- Wybierz --</option>
                        <option value="Jobalots">Jobalots</option>
                        <option value="Warrington">Warrington</option>
                        <option value="Miglo">Miglo</option>
                        <option value="Inny">Inny</option>
                    </select>
                </div>
            </div>
            <input type="text" name="nowa_paleta_nazwa" class="form-ctrl" placeholder="Nazwa nowej palety" style="display:none;margin-bottom:12px">
            <div class="form-group">
                <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">description</span> Plik Excel (.xlsx) lub CSV/TXT</label>
                <input type="file" name="file" class="form-ctrl" accept=".txt,.csv,.xlsx,.xls" required>
                <div style="font-size:0.62rem;color:#767577;margin-top:4px">Automatycznie wykrywa: ASIN, cenę, ilość, EAN, zdjęcia</div>
            </div>
            <div class="form-group">
                <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">paid</span> Cena jednostkowa brutto (opcjonalnie)</label>
                <input type="number" step="0.01" name="cena_jednostkowa" class="form-ctrl" placeholder="Nadpisuje cenę z pliku">
            </div>
            <button type="submit" class="btn" style="width:100%;padding:14px;background:rgba(143,245,255,0.1);border:1px solid rgba(143,245,255,0.3);color:#8ff5ff;font-weight:700;font-size:0.78rem;letter-spacing:0.1em;text-transform:uppercase;display:flex;align-items:center;justify-content:center;gap:6px"><span class=material-symbols-outlined>upload</span> WGRAJ I DODAJ DO MAGAZYNU</button>
        </form>
    </div>

    <!-- FORMULARZ 3: IMPORT MIGLO -->
    <div id="form-miglo" style="background:#1f1f22;border:1px solid rgba(245,158,11,0.2);padding:24px;margin-bottom:16px">
        <div style="font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;margin-bottom:16px;display:flex;align-items:center;gap:8px">
            <span class=material-symbols-outlined style="color:#f59e0b">list_alt</span> IMPORT MIGLO
            <span style="font-size:0.6rem;color:#767577;font-weight:400;margin-left:4px">(licytacje)</span>
        </div>
        <form action="/paletomat/scraper/miglo" method="POST">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px;padding:12px;background:#131315;border:1px solid rgba(245,158,11,0.15);border-radius:6px">
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">inventory_2</span> Paleta</label>
                    <select name="paleta_id" class="form-ctrl" onchange="this.form.nowa_paleta_nazwa.style.display = this.value === 'new' ? 'block' : 'none'">
                        <option value="">-- Bez palety --</option>
                        <option value="new">+ Nowa paleta...</option>
                        {palety_options_clean}
                    </select>
                </div>
                <div class="form-group" style="margin-bottom:0">
                    <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700"><span class=material-symbols-outlined style="font-size:0.85rem">factory</span> Dostawca</label>
                    <select name="dostawca" class="form-ctrl">
                        <option value="Miglo" selected>Miglo</option>
                        <option value="Jobalots">Jobalots</option>
                        <option value="Warrington">Warrington</option>
                        <option value="Inny">Inny</option>
                    </select>
                </div>
            </div>
            <input type="text" name="nowa_paleta_nazwa" class="form-ctrl" placeholder="Nazwa nowej palety" style="display:none;margin-bottom:12px">
            <div class="form-group">
                <label style="font-size:0.68rem;color:#adaaad;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Dane z tabeli Miglo</label>
                <textarea name="miglo_data" class="form-ctrl" rows="6" placeholder="B0IL6PLPDI	5	HOME_IMPROVEMENT	1647,56	221,89
B0IN7ENHO6	4	CAMERA	588,45	79,25
..." style="font-family:monospace;font-size:0.78rem"></textarea>
                <div style="font-size:0.62rem;color:#767577;margin-top:4px">Skopiuj tabelę z Miglo (ASIN | Ilość | Kategoria | Cena | Cena netto)</div>
            </div>
            <button type="submit" class="btn" style="width:100%;padding:14px;background:rgba(245,158,11,0.15);border:1px solid rgba(245,158,11,0.3);color:#f59e0b;font-weight:700;font-size:0.78rem;letter-spacing:0.1em;text-transform:uppercase;display:flex;align-items:center;justify-content:center;gap:6px"><span class=material-symbols-outlined>download</span> IMPORTUJ Z MIGLO</button>
        </form>
    </div>

    <!-- DISCOVERED ITEMS TABLE -->
    <div style="margin-top:28px">
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:14px">
            <div style="width:2px;height:14px;background:#beee00;flex-shrink:0"></div>
            <span style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#64748b">Recently Discovered &mdash; ostatnio dodane</span>
        </div>
        <div style="background:rgba(15,15,30,0.6);overflow:hidden">
            {_disc_table}
        </div>
    </div>

    <div style="text-align:center;margin-top:24px"><a href="/paletomat" style="font-size:0.82rem;color:#adaaad;text-decoration:none;font-weight:600;letter-spacing:0.05em">&larr; Powrót</a></div>
    '''
    return render(html)

@paletomat_bp.route('/scraper/miglo', methods=['POST'])
def scraper_miglo():
    """Import ręczny z danych Miglo - parsuje tabelę z ASIN, ilość, cena netto"""
    miglo_data = request.form.get('miglo_data', '')
    paleta_id_raw = request.form.get('paleta_id', '').strip()
    dostawca = request.form.get('dostawca', 'Miglo')
    nowa_paleta_nazwa = request.form.get('nowa_paleta_nazwa', '').strip()
    
    print(f"[DOWNLOAD] [MIGLO] paleta_id='{paleta_id_raw}', dostawca='{dostawca}', nowa_paleta='{nowa_paleta_nazwa}'")
    
    # Parsuj paleta_id - może być: '', 'new', lub liczba
    paleta_id = None
    if paleta_id_raw and paleta_id_raw.lower() not in ('new', 'nowy'):
        try:
            paleta_id = int(paleta_id_raw)
        except ValueError:
            print(f"⚠️ Nieprawidłowe ID palety: {paleta_id_raw}, używam None")
            paleta_id = None
    
    if not miglo_data.strip():
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie wklejono danych</div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
    
    # Parsuj dane - format: ASIN | Ilość | (cokolwiek) | Cena netto
    asin_data = {}  # {asin: {'qty': int, 'netto': float}}
    
    lines = miglo_data.strip().split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Rozbij na kolumny (tab, |, wiele spacji)
        parts = re.split(r'\t+|\|+|\s{2,}', line)
        parts = [p.strip() for p in parts if p.strip()]
        
        if len(parts) < 2:
            continue
        
        # Szukaj ASIN-a (B0...) - 8-10 znaków po B0
        asin = None
        for p in parts:
            match = re.search(r'[Bb]0[A-Za-z0-9]{8,10}', p)
            if match:
                asin = match.group().upper()
                break
        
        if not asin:
            continue
        
        # Szukaj ilości (pierwsza liczba całkowita po ASIN)
        qty = 1
        for p in parts[1:]:
            clean = p.replace(',', '.').replace(' ', '')
            if clean.isdigit():
                qty = int(clean)
                break
            # Spróbuj jako float i zaokrąglij
            try:
                val = float(clean)
                if val == int(val) and 1 <= val <= 1000:
                    qty = int(val)
                    break
            except:
                pass
        
        # Szukaj ceny netto (ostatnia liczba zmiennoprzecinkowa)
        cena_netto = 0
        for p in reversed(parts):
            clean = p.replace(',', '.').replace(' ', '').replace('zł', '').replace('PLN', '')
            try:
                val = float(clean)
                if val > 0:
                    cena_netto = val
                    break
            except:
                pass
        
        if asin and cena_netto > 0:
            # Cena w tabeli Miglo to cena JEDNOSTKOWA (per sztuka)
            asin_data[asin] = {
                'qty': qty,
                'netto_jednostkowa': cena_netto,
                'netto_lacznie': cena_netto * qty,
                'brutto_jednostkowa': round(cena_netto * 1.23, 2),
                'brutto_lacznie': round(cena_netto * qty * 1.23, 2)
            }
            print(f"📦 Miglo: {asin} - {qty}szt × {cena_netto:.2f} netto = {cena_netto * qty:.2f} netto łącznie")
    
    if not asin_data:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie znaleziono prawidłowych danych.<br><br><small>Format: ASIN | Ilość | ... | Cena netto</small></div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
    
    # Utwórz nową paletę jeśli trzeba
    conn = get_db()
    
    if nowa_paleta_nazwa and not paleta_id:
        cursor = conn.execute(
            'INSERT INTO palety (nazwa, dostawca, data_zakupu, cena_zakupu) VALUES (?, ?, date("now"), 0)',
            (nowa_paleta_nazwa, dostawca)
        )
        paleta_id = cursor.lastrowid
        print(f"📦 Utworzono nową paletę: {nowa_paleta_nazwa} (ID: {paleta_id})")
    
    # paleta_id jest już int lub None - nie trzeba konwertować ponownie
    
    # Pobierz nazwę palety
    paleta_nazwa = ""
    if paleta_id:
        p = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (paleta_id,)).fetchone()
        if p:
            paleta_nazwa = p['nazwa']
    
    # Dodaj produkty do bazy
    added = 0
    total_netto = 0
    total_brutto = 0
    total_qty = 0
    
    for asin, data in asin_data.items():
        qty = data['qty']
        # Cena netto jednostkowa prosto z Miglo, brutto = netto × 1.23
        cena_netto = data['netto_jednostkowa']
        cena_brutto = round(cena_netto * 1.23, 2)
        
        try:
            # Dodaj do scraped
            conn.execute('''INSERT OR IGNORE INTO scraped (asin, status, zdjecie_url) 
                VALUES (?, 'nowy', ?)''', (asin, get_amazon_image_url(asin)))
            
            # Oblicz cenę Allegro automatycznie (mnożnik 2.5x od brutto)
            cena_allegro = round(cena_brutto * 2.5, 2) if cena_brutto else 0
            
            # Dodaj do produkty (magazyn)
            # Szukaj produktu TYLKO w tej samej palecie (nie nadpisuj produktu z innej palety!)
            existing = conn.execute('SELECT id, ean FROM produkty WHERE asin = ? AND paleta_id = ?', (asin, paleta_id)).fetchone()
            if not existing:
                _initial_name = _resolve_initial_name(conn, asin)
                conn.execute('''INSERT INTO produkty (asin, nazwa, ilosc, cena_netto, cena_brutto, cena_allegro, paleta_id, paleta, dostawca, zdjecie_url, stan)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (asin, _initial_name, qty, cena_netto, cena_brutto, cena_allegro, paleta_id, paleta_nazwa, dostawca, get_amazon_image_url(asin), 'Nowy'))
            else:
                # Aktualizuj - tylko w tej palecie - SUMUJ ilość (ten sam ASIN może być w wielu wierszach)
                conn.execute('UPDATE produkty SET cena_netto = ?, cena_brutto = ?, cena_allegro = ?, ilosc = ilosc + ?, paleta = ?, dostawca = ? WHERE id = ?',
                    (cena_netto, cena_brutto, cena_allegro, qty, paleta_nazwa, dostawca, existing['id']))
            
            total_netto += data['netto_lacznie']
            total_brutto += data['brutto_lacznie']
            total_qty += qty
            added += 1
        except Exception as e:
            print(f"[CANCEL] Błąd dodawania {asin}: {e}")
    
    # Zaktualizuj ilość produktów i sztuk w palecie
    if paleta_id:
        count = conn.execute('SELECT COUNT(*) as cnt FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()['cnt']
        # AKUMULUJ cena_zakupu — NIE nadpisuj, dodaj do istniejącej (żeby sprzedaż nie zmniejszała kosztu palety)
        stara_cena = conn.execute('SELECT COALESCE(cena_zakupu, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
        nowa_cena = round(stara_cena + total_brutto, 2)
        stary_netto = conn.execute('SELECT COALESCE(cena_zakupu_netto, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
        nowy_netto = round(stary_netto + total_netto, 2)
        try:
            conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ?, ilosc_produktow = ?, ilosc_sztuk = COALESCE(ilosc_sztuk, 0) + ? WHERE id = ?', (nowa_cena, nowy_netto, count, total_qty, paleta_id))
        except:
            conn.execute('UPDATE palety SET cena_zakupu = ?, ilosc_produktow = ? WHERE id = ?', (nowa_cena, count, paleta_id))
    
    conn.commit()
    
    # Uruchom auto-przetwarzanie w tle (z wybraną domeną)
    auto_process_products(list(asin_data.keys()), preferred_domain=domain if domain != 'de' else None)

    paleta_info = f'<br><span class=material-symbols-outlined>inventory_2</span> Paleta: <b>{paleta_nazwa}</b> ({dostawca})' if paleta_nazwa else ''
    
    return render(f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> IMPORT MIGLO</h1></div>
        <div class="alert alert-ok">
            Zaimportowano <b>{added}</b> produktów{paleta_info}
        </div>
        <div class="card" style="padding:15px">
            <div style="font-weight:600;margin-bottom:10px"><span class=material-symbols-outlined>bar_chart</span> Podsumowanie importu:</div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                <div>Produktów: <b>{added}</b></div>
                <div>Sztuk łącznie: <b>{total_qty}</b></div>
                <div>Netto: <span style="color:#8ff5ff"><b>{total_netto:.2f} PLN</b></span></div>
                <div>Brutto: <span style="color:#beee00"><b>{total_brutto:.2f} PLN</b></span></div>
            </div>
        </div>
        <div class="alert" style="background:#1a1a2e;font-size:0.85rem">
            <span class=material-symbols-outlined>check_circle</span> Produkty dodane do <a href="/magazyn" style="color:#8ff5ff">Magazynu</a><br>
            <span class=material-symbols-outlined>sync</span> <b>Auto-przetwarzanie w tle</b> - tytuły i opisy generują się automatycznie
        </div>
        <a href="/paletomat/scraper" class="btn btn-p"><span class=material-symbols-outlined>search</span> Dodaj więcej</a>
        <a href="/magazyn" class="btn btn-ok"><span class=material-symbols-outlined>inventory_2</span> Zobacz Magazyn</a>
        <a href="/paletomat" class="back">← Powrót</a>
    ''')

@paletomat_bp.route('/scraper/asin', methods=['POST'])
def scraper_asin():
    asins_raw = request.form.get('asins', '')
    domain = request.form.get('domain', 'de')
    paleta_id = request.form.get('paleta_id', '')
    dostawca = request.form.get('dostawca', '')
    nowa_paleta_nazwa = request.form.get('nowa_paleta_nazwa', '').strip()
    cena_jednostkowa_raw = request.form.get('cena_jednostkowa', '').strip()
    
    # Parsuj cenę jednostkową
    cena_brutto = 0
    cena_netto = 0
    if cena_jednostkowa_raw:
        try:
            cena_brutto = float(cena_jednostkowa_raw.replace(',', '.'))
            cena_netto = round(cena_brutto / 1.23, 2)
        except:
            pass
    
    # Parsuj ASIN-y z opcjonalną ilością (format: ASIN,ilość lub ASIN)
    # Miglo daje format: B0XXXXXXXXX,2 (ASIN + ilość po przecinku)
    asin_qty_map = {}  # {asin: qty}
    raw_items = [a.strip() for a in asins_raw.replace('\n', ' ').replace(';', ' ').split() if a.strip()]
    
    for item in raw_items:
        # Sprawdź czy to format "ASIN,ilość"
        if ',' in item:
            parts = item.split(',')
            asin_part = parts[0].strip().upper()
            qty_part = parts[1].strip() if len(parts) > 1 else '1'
            try:
                qty = int(qty_part)
            except:
                qty = 1
        else:
            asin_part = item.upper()
            qty = 1
        
        # Sprawdź czy to prawidłowy ASIN (8-10 znaków po B0)
        if re.match(r'^B0[A-Z0-9]{8,10}$', asin_part):
            if asin_part in asin_qty_map:
                asin_qty_map[asin_part] += qty
            else:
                asin_qty_map[asin_part] = qty
    
    asins = list(asin_qty_map.keys())
    
    if not asins:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-warn">Nie podano prawidłowych ASIN-ów</div>
            <a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>
        ''')
    
    conn = get_db()
    
    # Jeśli nowa paleta - utwórz ją
    if paleta_id == 'new' and nowa_paleta_nazwa:
        cursor = conn.execute(
            'INSERT INTO palety (nazwa, dostawca) VALUES (?, ?)', 
            (nowa_paleta_nazwa, dostawca)
        )
        paleta_id = cursor.lastrowid
        conn.commit()
    elif paleta_id and paleta_id != 'new':
        paleta_id = int(paleta_id)
    else:
        paleta_id = None
    
    # Pobierz nazwę palety do wyświetlenia
    paleta_nazwa = ""
    if paleta_id:
        p = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (paleta_id,)).fetchone()
        if p:
            paleta_nazwa = p['nazwa']
    
    # Dodaj do bazy scraped + produkty
    added = 0
    total_value = 0
    total_netto = 0
    total_qty = 0
    for asin in asins:
        qty = asin_qty_map.get(asin, 1)
        try:
            # WAŻNE: cena_brutto i cena_netto to ceny JEDNOSTKOWE (za sztukę)
            # Musimy zapisać CAŁKOWITĄ cenę zakupu = cena_jednostkowa * ilość
            cena_brutto_total = cena_brutto * qty
            cena_netto_total = cena_netto * qty
            # Cena Allegro = cena JEDNOSTKOWA * 2.5 (za 1 sztukę!)
            cena_allegro = round(cena_brutto * 2.5, 2) if cena_brutto else 0
            
            # Dodaj do scraped
            conn.execute('''INSERT OR IGNORE INTO scraped (asin, status, zdjecie_url) 
                VALUES (?, 'nowy', ?)''', (asin, get_amazon_image_url(asin)))
            
            # Dodaj do produkty (magazyn) - sprawdź duplikaty
            # 1. Szukaj w tej samej palecie
            existing = conn.execute('SELECT id FROM produkty WHERE asin = ? AND paleta_id = ?', (asin, paleta_id)).fetchone()
            if not existing:
                    # Nowy produkt w tej palecie — INSERT
                    _initial_name = _resolve_initial_name(conn, asin)
                    conn.execute('''INSERT INTO produkty (asin, nazwa, ilosc, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta, dostawca, zdjecie_url, stan)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (asin, _initial_name, qty, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta_nazwa, dostawca, get_amazon_image_url(asin), 'Nowy'))
            else:
                # Zaktualizuj paletę, dostawcę, cenę i ilość - WHERE id = ? (tylko ten konkretny rekord)
                if cena_brutto > 0:
                    conn.execute('UPDATE produkty SET paleta = ?, dostawca = ?, cena_brutto = ?, cena_netto = ?, cena_allegro = ?, ilosc = ilosc + ? WHERE id = ?',
                        (paleta_nazwa, dostawca, cena_brutto, cena_netto, cena_allegro, qty, existing['id']))
                elif paleta_id:
                    conn.execute('UPDATE produkty SET paleta = ?, dostawca = ?, ilosc = ilosc + ? WHERE id = ?',
                        (paleta_nazwa, dostawca, qty, existing['id']))
            
            total_value += cena_brutto_total
            total_netto += cena_netto_total
            total_qty += qty
            added += 1
        except:
            pass

    # Zaktualizuj ilość produktów i sztuk w palecie
    if paleta_id:
        count = conn.execute('SELECT COUNT(*) as cnt FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()['cnt']
        # AKUMULUJ cena_zakupu — NIE nadpisuj, dodaj do istniejącej (żeby sprzedaż nie zmniejszała kosztu palety)
        stara_cena = conn.execute('SELECT COALESCE(cena_zakupu, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
        nowa_cena = round(stara_cena + total_value, 2)
        stary_netto = conn.execute('SELECT COALESCE(cena_zakupu_netto, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
        nowy_netto = round(stary_netto + total_netto, 2)
        try:
            conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ?, ilosc_produktow = ?, ilosc_sztuk = COALESCE(ilosc_sztuk, 0) + ? WHERE id = ?', (nowa_cena, nowy_netto, count, total_qty, paleta_id))
        except:
            conn.execute('UPDATE palety SET cena_zakupu = ?, ilosc_produktow = ? WHERE id = ?', (nowa_cena, count, paleta_id))
    
    conn.commit()
    
    # Uruchom auto-przetwarzanie w tle (pobieranie tytułów i generowanie opisów)
    auto_process_products(list(asins))
    
    paleta_info = f' → Paleta: <b>{paleta_nazwa}</b> ({dostawca})' if paleta_nazwa else ''
    cena_info = f'<br><span class=material-symbols-outlined>paid</span> Cena jednostkowa: <b>{cena_brutto:.2f} zł</b> (netto: {cena_netto:.2f} zł)' if cena_brutto > 0 else ''
    
    return render(f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> DODANO</h1></div>
        <div class="alert alert-ok">Dodano {added} ASIN-ów do kolejki{paleta_info}{cena_info}</div>
        <div class="alert" style="background:#1a1a2e;font-size:0.85rem">
            <span class=material-symbols-outlined>check_circle</span> Produkty dodane do <a href="/magazyn" style="color:#8ff5ff">Magazynu</a><br>
            <span class=material-symbols-outlined>sync</span> <b>Auto-przetwarzanie w tle</b> - tytuły i opisy generują się automatycznie
        </div>
        <a href="/paletomat/scraper" class="btn btn-p"><span class=material-symbols-outlined>search</span> Dodaj więcej</a>
        <a href="/paletomat/generator" class="btn btn-ok"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Zobacz postęp</a>
        <a href="/paletomat" class="back">← Powrót do Paletomat</a>
    ''')

@paletomat_bp.route('/scraper/file', methods=['POST'])
def scraper_file():
    """Import ASIN-ów z pliku (Excel lub CSV/TXT)"""
    
    print("="*60)
    print("<span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> [SCRAPER FILE] START")
    print("="*60)
    
    # DEBUG: Pokaż WSZYSTKO co przyszło z formularza
    print(f"[LIST_ALT] Wszystkie pola formularza: {dict(request.form)}")
    print(f"[FOLD] Pliki: {list(request.files.keys())}")
    
    if 'file' not in request.files:
        print("[CANCEL] Brak pliku w request!")
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie wybrano pliku</div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
    
    file = request.files['file']
    if file.filename == '':
        print("[CANCEL] Pusta nazwa pliku!")
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie wybrano pliku</div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
    
    print(f"[DESC] Plik: {file.filename}")
    
    # Pobierz dane palety
    paleta_id = request.form.get('paleta_id', '')
    dostawca = request.form.get('dostawca', '')
    nowa_paleta_nazwa = request.form.get('nowa_paleta_nazwa', '').strip()
    
    # DEBUG: Pokaż co przyszło z formularza
    print(f"[DOWNLOAD] [FORM DATA]")
    print(f"   paleta_id = '{paleta_id}'")
    print(f"   dostawca = '{dostawca}'")
    print(f"   nowa_paleta_nazwa = '{nowa_paleta_nazwa}'")
    
    # Pobierz cenę jednostkową (jako fallback jeśli nie ma w Excelu)
    cena_jednostkowa_raw = request.form.get('cena_jednostkowa', '').strip()
    manual_cena_brutto = 0
    manual_cena_netto = 0
    if cena_jednostkowa_raw:
        try:
            manual_cena_brutto = float(cena_jednostkowa_raw.replace(',', '.'))
            manual_cena_netto = round(manual_cena_brutto / 1.23, 2)
        except:
            pass
    
    filename = file.filename.lower()
    asins = set()
    asin_prices = {}
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Import z Excela - z wykrywaniem cen
            import openpyxl
            import tempfile
            
            # Zapisz do pliku tymczasowego
            tmp_path = os.path.join(tempfile.gettempdir(), f'paletomat_{os.getpid()}.xlsx')
            file.save(tmp_path)
            
            try:
                try:
                    wb = openpyxl.load_workbook(tmp_path, read_only=True)
                except ValueError:
                    # Uszkodzone style/theme XML (np. pliki Warrington) — napraw i spróbuj ponownie
                    print("⚠️ Uszkodzone style w Excelu, naprawiam...")
                    import zipfile
                    repaired_path = tmp_path.replace('.xlsx', '_repaired.xlsx')
                    # Minimalny styles.xml z 10 pustymi stylami (wystarczy dla większości plików)
                    min_styles = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="1"><font><sz val="11"/></font></fonts><fills count="2"><fill><patternFill/></fill><fill><patternFill patternType="gray125"/></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="10"><xf/><xf/><xf/><xf/><xf/><xf/><xf/><xf/><xf/><xf/></cellXfs></styleSheet>'
                    with zipfile.ZipFile(tmp_path, 'r') as zin, zipfile.ZipFile(repaired_path, 'w') as zout:
                        for item in zin.namelist():
                            if item == 'xl/styles.xml':
                                zout.writestr(item, min_styles)
                            elif item == 'xl/theme/theme1.xml':
                                continue  # Pomiń uszkodzony theme
                            else:
                                zout.writestr(item, zin.read(item))
                    wb = openpyxl.load_workbook(repaired_path)
                    try:
                        os.remove(repaired_path)
                    except:
                        pass
                ws = wb.active
                
                # Znajdź nagłówki - szukaj pierwszego wiersza z niepustymi komórkami
                headers = []
                col_asin = -1
                col_price = -1
                col_qty = -1
                price_is_netto = False
                header_row_found = False
                rows_checked = 0
                
                for row in ws.iter_rows(values_only=True):
                    rows_checked += 1
                    
                    if not header_row_found:
                        # Szukamy wiersza z nagłówkami (max 10 pierwszych wierszy)
                        if rows_checked > 10:
                            print(f"⚠️ Nie znaleziono nagłówków w pierwszych 10 wierszach")
                            header_row_found = True  # Kontynuuj bez nagłówków
                            continue
                        
                        if not row:
                            continue
                        
                        # Sprawdź czy wiersz ma niepuste komórki
                        non_empty = [c for c in row if c is not None and str(c).strip()]
                        if len(non_empty) < 2:
                            # Za mało danych - to nie nagłówki
                            continue
                        
                        # Sprawdź czy to wygląda jak nagłówki (tekst, nie same liczby)
                        text_cells = [c for c in non_empty if not str(c).replace('.', '').replace(',', '').isdigit()]
                        if len(text_cells) < 2:
                            # Same liczby - to dane, nie nagłówki
                            continue
                        
                        # Mamy potencjalne nagłówki!
                        header_row_found = True
                        headers = [str(c).lower().strip() if c else '' for c in row]
                        print(f"[LIST_ALT] Nagłówki Excel (wiersz {rows_checked}): {[h for h in headers if h]}")
                        
                        col_unit_price = -1  # NAJWYŻSZY PRIORYTET: Cena jednostkowa sprzedaży
                        col_netto = -1       # WYSOKI: Cena sprzedaży netto
                        col_cost = -1        # ŚREDNI: Unit Cost / Cost
                        col_rrp = -1         # NAJNIŻSZY: RRP / Retail Price (UNIKAĆ!)
                        col_ean = -1         # Kolumna EAN
                        col_images = -1      # NOWE: Kolumna ze zdjęciami
                        
                        for i, h in enumerate(headers):
                            h_clean = h.replace(' ', '').replace('_', '').replace('-', '').replace('ó', 'o').replace('ś', 's').replace('ć', 'c')
                            h_orig = h.lower()
                            
                            # Kolumna EAN
                            if col_ean == -1 and any(x in h_clean for x in ['ean', 'barcode', 'kodkreskowy', 'gtin']):
                                col_ean = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę EAN: {i} ({h})")
                            
                            # NOWE: Kolumna ze zdjęciami
                            if col_images == -1 and any(x in h_clean for x in ['zdjec', 'image', 'images', 'photo', 'photos', 'link', 'links', 'url', 'urls']):
                                col_images = i
                                print(f"[PHOTO_CAMERA] Znaleziono kolumnę ZDJĘCIA: {i} ({h})")
                            
                            # Kolumna ASIN - PRIORYTET dla dokładnego "asin", potem inne
                            # Unikaj "product sku" - to nie jest ASIN!
                            if h_clean == 'asin':
                                col_asin = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę ASIN (dokładne): {i} ({h})")
                            elif col_asin == -1 and 'product' not in h_clean and any(x in h_clean for x in ['sku', 'kod2', 'code', 'artikelnummer', 'article']):
                                col_asin = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę ASIN (alternatywne): {i} ({h})")
                            
                            # UNIKAJ kolumn z cenami rynkowymi!
                            if any(x in h_orig for x in ['regularn', 'rynkow', 'rrp', 'retail', 'msrp']):
                                if 'jednostkow' not in h_orig:  # Ale nie unikaj "jednostkowa"
                                    print(f"⚠️ Pomijam kolumnę rynkową: {i} ({h})")
                                    continue
                            
                            # NAJWYŻSZY PRIORYTET: Cena jednostkowa sprzedaży (per sztuka!)
                            if col_unit_price == -1 and 'jednostkow' in h_orig and any(x in h_orig for x in ['sprzeda', 'cena']):
                                col_unit_price = i
                                print(f"[EMOJ] Znaleziono kolumnę CENA JEDNOSTKOWA: {i} ({h})")
                            
                            # WYSOKI PRIORYTET: Cena sprzedaży netto
                            if col_netto == -1 and 'sprzeda' in h_orig and 'netto' in h_orig:
                                col_netto = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę NETTO SPRZEDAŻY: {i} ({h})")
                            
                            # ŚREDNI PRIORYTET: Unit Cost, Cost, Cena zakupu
                            if col_cost == -1 and any(x in h_clean for x in ['unitcost', 'cenazakupu', 'koszt', 'einkaufspreis']):
                                col_cost = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę KOSZT: {i} ({h})")
                            
                            # NISKI PRIORYTET: Cena sprzedaży (może być łączna, nie jednostkowa)
                            if col_price == -1 and 'sprzeda' in h_orig and 'jednostkow' not in h_orig and 'netto' not in h_orig:
                                col_price = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę CENA SPRZEDAŻY: {i} ({h})")
                            
                            # Kolumna ilości - rozszerzone wzorce
                            if col_qty == -1 and any(x in h_clean for x in ['ilosc', 'ilość', 'qty', 'quantity', 'sztuk', 'szt', 'pcs', 'pieces', 'count', 'menge', 'anzahl', 'stueck', 'stück']):
                                col_qty = i
                                print(f"[CHECK_CIRCLE] Znaleziono kolumnę ILOŚĆ: {i} ({h})")
                        
                        # Wybierz najlepszą kolumnę ceny (priorytet: jednostkowa > netto > cost > sprzedaży)
                        price_is_netto = False
                        if col_unit_price >= 0:
                            col_price = col_unit_price
                            price_is_netto = True
                            print(f"[EMOJ] Używam kolumny JEDNOSTKOWA jako cena: {col_price} (netto ×1.23 → brutto)")
                        elif col_netto >= 0:
                            col_price = col_netto
                            price_is_netto = True
                            print(f"[PAID] Używam kolumny NETTO jako cena: {col_price} (×1.23 → brutto)")
                        elif col_cost >= 0:
                            col_price = col_cost
                            print(f"[PAID] Używam kolumny KOSZT jako cena: {col_price}")
                        # col_price już może być ustawiony na "cena sprzedaży"
                        
                        print(f"[BAR_CHART] Wykryte kolumny: ASIN={col_asin}, EAN={col_ean}, CENA={col_price}, ILOŚĆ={col_qty}")
                        continue
                    
                    if not row:
                        continue
                    
                    # Szukaj ASIN-ów w tym wierszu
                    found_asins = []
                    
                    # Szukaj ASIN-ów TYLKO w dedykowanej kolumnie (nie w całym wierszu!)
                    # Kolumny Image mogą zawierać ASIN-y w URL-ach co powoduje fałszywe wykrycia
                    if col_asin >= 0 and col_asin < len(row) and row[col_asin]:
                        cell_str = str(row[col_asin]).strip()
                        # Szukaj ASIN-ów (case-insensitive, 8-10 znaków po B0)
                        found_asins = re.findall(r'[Bb]0[A-Za-z0-9]{8,10}', cell_str)
                        # Konwertuj do uppercase
                        found_asins = [a.upper() for a in found_asins]
                        if found_asins:
                            print(f"   [SEARCH] Kolumna ASIN[{col_asin}]: znaleziono {found_asins}")
                    
                    # Jeśli nie znaleziono i nie ma dedykowanej kolumny, szukaj w kolumnach tekstowych (NIE w Image!)
                    if not found_asins and col_asin == -1:
                        for i, cell in enumerate(row):
                            # Pomiń kolumny które mogą zawierać URL-e (Image, zdjęcia, link)
                            header = headers[i].lower() if i < len(headers) and headers[i] else ''
                            if any(x in header for x in ['image', 'img', 'photo', 'zdjęci', 'zdjeci', 'link', 'url', 'http']):
                                continue
                            if cell:
                                cell_str = str(cell).strip()
                                # Pomiń jeśli to URL
                                if cell_str.startswith('http'):
                                    continue
                                # Szukaj ASIN-ów
                                matches = re.findall(r'[Bb]0[A-Za-z0-9]{8,10}', cell_str)
                                found_asins.extend([m.upper() for m in matches])
                        if found_asins:
                            # Usuń duplikaty
                            found_asins = list(dict.fromkeys(found_asins))
                            print(f"   [SEARCH] Szukanie w wierszu: znaleziono {found_asins}")
                    
                    # Pobierz EAN z wiersza (jeśli kolumna istnieje)
                    ean_value = ''
                    if col_ean >= 0 and col_ean < len(row) and row[col_ean]:
                        ean_raw = str(row[col_ean]).strip()
                        # Usuń .0 z floatów Excel
                        if ean_raw.endswith('.0'):
                            ean_raw = ean_raw[:-2]
                        ean_raw = ean_raw.replace('.0', '').replace(' ', '')
                        # Sprawdź czy wygląda jak EAN (tylko cyfry, 8-14 znaków)
                        if ean_raw.isdigit() and 8 <= len(ean_raw) <= 14:
                            ean_value = ean_raw
                    
                    # NOWE: Pobierz ZDJĘCIA z wiersza (jeśli kolumna istnieje)
                    images_list = []
                    if col_images >= 0 and col_images < len(row) and row[col_images]:
                        images_raw = str(row[col_images]).strip()
                        if images_raw:
                            # Rozdziel linki po przecinkach, średnikach lub spacjach
                            # Obsługuje formaty: "url1, url2, url3" lub "url1; url2" lub "url1 url2"
                            separators = [',', ';', '\n']
                            for sep in separators:
                                if sep in images_raw:
                                    images_list = [url.strip() for url in images_raw.split(sep) if url.strip()]
                                    break
                            # Jeśli nie ma separatorów, to może być jeden link
                            if not images_list:
                                images_list = [images_raw]
                            
                            # Filtruj tylko prawidłowe URL-e (http/https)
                            images_list = [url for url in images_list if url.startswith('http')]
                            
                            if images_list:
                                print(f"   [PHOTO_CAMERA] {len(images_list)} zdjęć: {images_list[0][:40]}... {f'(+{len(images_list)-1})' if len(images_list) > 1 else ''}")
                    
                    # Dodaj znalezione ASIN-y z cenami (sumuj ilości dla powtórzonych ASIN-ów)
                    for asin in found_asins:
                        # Pobierz cenę
                        price = 0
                        if col_price >= 0 and col_price < len(row) and row[col_price]:
                            try:
                                price_val = str(row[col_price]).replace(',', '.').replace(' ', '')
                                price_val = re.sub(r'[^\d.]', '', price_val)
                                price = float(price_val) if price_val else 0
                            except:
                                pass

                        # Pobierz ilość - użyj SmartQuantityParser
                        qty = 1
                        if col_qty >= 0 and col_qty < len(row) and row[col_qty]:
                            raw_qty = row[col_qty]
                            result = SmartQuantityParser.parse(raw_qty)
                            qty = result.value
                            print(f"   {asin}: RAW[{col_qty}]='{raw_qty}' -> qty={qty}, cena={price}")

                        if asin not in asins:
                            asins.add(asin)
                            asin_prices[asin] = {'price': price, 'qty': qty, 'ean': ean_value, 'images': images_list}
                        else:
                            # Ten sam ASIN w kolejnym wierszu — sumuj ilości
                            asin_prices[asin]['qty'] += qty
                            print(f"   {asin}: +{qty} szt (łącznie: {asin_prices[asin]['qty']})")
                
                wb.close()
            finally:
                try:
                    os.remove(tmp_path)
                except:
                    pass
        
        else:
            # Import z CSV/TXT (bez cen - tylko ASIN-y)
            raw_data = file.read()
            content = None
            
            for encoding in ['utf-8-sig', 'utf-8', 'cp1250', 'latin-1', 'iso-8859-2']:
                try:
                    content = raw_data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie można odczytać pliku</div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
            
            # Szukaj ASIN-ów (case insensitive, 8-10 znaków po B0)
            found = re.findall(r'[Bb]0[A-Za-z0-9]{8,10}', content)
            print(f"[DESC] CSV/TXT: znaleziono {len(found)} ASIN-ów")
            for asin in found:
                asin = asin.upper()
                asins.add(asin)
                asin_prices[asin] = {'price': 0, 'qty': 1, 'ean': '', 'images': []}
        
        if not asins:
            return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">Nie znaleziono ASIN-ów w pliku.<br><br><small style="color:#64748b">ASIN to kod Amazon zaczynający się od B0, np. B0CFQBBT7G</small></div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')
        
        # Dodaj do bazy
        conn = get_db()
        
        # Jeśli nowa paleta - utwórz ją
        if paleta_id == 'new' and nowa_paleta_nazwa:
            cursor = conn.execute(
                'INSERT INTO palety (nazwa, dostawca) VALUES (?, ?)', 
                (nowa_paleta_nazwa, dostawca)
            )
            paleta_id = cursor.lastrowid
            conn.commit()
            print(f"📦 Utworzono NOWĄ paletę: {nowa_paleta_nazwa} (ID: {paleta_id})")
        elif paleta_id and paleta_id != 'new':
            paleta_id = int(paleta_id)
            print(f"📦 Używam istniejącej palety ID: {paleta_id}")
        else:
            paleta_id = None
            print(f"⚠️ BRAK palety - produkty bez przypisania!")
        
        # Pobierz nazwę palety
        paleta_nazwa = ""
        if paleta_id:
            p = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (paleta_id,)).fetchone()
            if p:
                paleta_nazwa = p['nazwa']
                print(f"📦 Paleta nazwa: {paleta_nazwa}")
        
        added = 0
        updated = 0
        total_brutto = 0
        total_netto = 0
        
        print(f"[SYNC] Przetwarzam {len(asins)} ASIN-ów...")
        print(f"   Paleta ID: {paleta_id}, Nazwa: {paleta_nazwa}, Dostawca: {dostawca}")
        
        for asin in asins:
            try:
                # Pobierz cenę, ilość, EAN i ZDJĘCIA z pliku
                price_data = asin_prices.get(asin, {'price': 0, 'qty': 1, 'ean': '', 'images': []})
                cena_z_pliku = price_data['price']
                qty = price_data['qty']
                ean = price_data.get('ean', '')
                images_from_file = price_data.get('images', [])
                
                # Użyj ręcznej ceny jako fallback jeśli nie ma w pliku
                if cena_z_pliku == 0 and manual_cena_brutto > 0:
                    cena_z_pliku = manual_cena_brutto
                
                # === PRZELICZANIE NETTO → BRUTTO ===
                # Jeśli kolumna w Excelu to "Cena sprzedaży netto" → cena jest netto, mnóż ×1.23
                # Jeśli inna kolumna (np. jobalots) → cena już jest brutto
                if cena_z_pliku > 0 and price_is_netto:
                    cena_netto = cena_z_pliku
                    cena_brutto = round(cena_z_pliku * 1.23, 2)
                    print(f"[PAID] Netto z pliku: {cena_z_pliku:.2f} × 1.23 = {cena_brutto:.2f} brutto")
                else:
                    # Cena z pliku jest brutto (jobalots, inne)
                    cena_brutto = cena_z_pliku
                    cena_netto = round(cena_brutto / 1.23, 2) if cena_brutto > 0 else 0
                
                # Dodaj do sumy
                total_brutto += cena_brutto * qty
                total_netto += cena_netto * qty
                
                # WAŻNE: Zapisujemy CAŁKOWITĄ cenę zakupu = cena_jednostkowa * ilość
                cena_brutto_total = cena_brutto * qty
                cena_netto_total = cena_netto * qty
                # Cena Allegro = cena JEDNOSTKOWA * 2.5 (za 1 sztukę!)
                cena_allegro = round(cena_brutto * 2.5, 2) if cena_brutto else 0
                
                # Przygotuj zdjęcia do zapisu (JSON)
                images_json = ''
                if images_from_file:
                    images_json = json.dumps(images_from_file)
                    print(f"[PHOTO_CAMERA] {asin}: zapisuję {len(images_from_file)} zdjęć do bazy")
                
                # Dodaj do scraped
                conn.execute('''INSERT OR IGNORE INTO scraped (asin, status, zdjecie_url) 
                    VALUES (?, 'nowy', ?)''', (asin, images_from_file[0] if images_from_file else get_amazon_image_url(asin)))
                
                # Aktualizuj zdjęcia w scraped jeśli istnieją (może nie być kolumny images)
                if images_json:
                    try:
                        conn.execute('''UPDATE scraped SET images = ? WHERE asin = ?''', (images_json, asin))
                    except Exception:
                        pass  # Kolumna images może nie istnieć w starszych bazach
                
                # Dodaj do produkty (magazyn)
                # Szukaj produktu TYLKO w tej samej palecie (nie nadpisuj produktu z innej palety!)
                existing = conn.execute('SELECT id, ean FROM produkty WHERE asin = ? AND paleta_id = ?', (asin, paleta_id)).fetchone()
                if not existing:
                    print(f"➕ Nowy produkt: {asin}, paleta_id={paleta_id}, ean={ean or 'brak'}, ilość={qty}, cena_jedn={cena_brutto}, cena_allegro={cena_allegro}, zdjęć={len(images_from_file)}")
                    _initial_name = _resolve_initial_name(conn, asin, ean)
                    try:
                        # Próbuj z kolumną images
                        # WAŻNE: zapisujemy ceny JEDNOSTKOWE (cena_brutto, cena_netto), nie całkowite!
                        conn.execute('''INSERT INTO produkty (asin, ean, nazwa, ilosc, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta, dostawca, zdjecie_url, images)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (asin, ean, _initial_name, qty, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta_nazwa, dostawca,
                             images_from_file[0] if images_from_file else get_amazon_image_url(asin), images_json))
                    except Exception:
                        # Fallback bez kolumny images (starsza baza)
                        conn.execute('''INSERT INTO produkty (asin, ean, nazwa, ilosc, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta, dostawca, zdjecie_url)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (asin, ean, _initial_name, qty, cena_brutto, cena_netto, cena_allegro, paleta_id, paleta_nazwa, dostawca,
                             images_from_file[0] if images_from_file else get_amazon_image_url(asin)))
                    added += 1
                else:
                    # Aktualizuj istniejący produkt
                    existing_ean = existing['ean'] or ''
                    new_ean = ean if ean and not existing_ean else existing_ean
                    print(f"[SYNC] Aktualizacja: {asin}, paleta_id={paleta_id}, ean={new_ean or 'brak'}, ilość={qty}, cena_jedn={cena_brutto}, cena_allegro={cena_allegro}, zdjęć={len(images_from_file)}")
                    
                    # ZAWSZE aktualizuj paletę i dostawcę jeśli podane
                    update_fields = []
                    update_values = []
                    
                    # EAN - tylko jeśli jest nowy i nie było w bazie
                    if ean and not existing_ean:
                        update_fields.append('ean = ?')
                        update_values.append(ean)
                    
                    # Cena - zapisuj JEDNOSTKOWĄ (nie całkowitą!)
                    if cena_brutto > 0:
                        update_fields.extend(['cena_brutto = ?', 'cena_netto = ?', 'cena_allegro = ?'])
                        update_values.extend([cena_brutto, cena_netto, cena_allegro])
                    
                    # Ilość - SUMUJ (ten sam ASIN może być rozbity na wiele wierszy w Excelu)
                    update_fields.append('ilosc = ilosc + ?')
                    update_values.append(qty)
                    
                    # Paleta - zawsze jeśli podana
                    if paleta_id:
                        update_fields.extend(['paleta_id = ?', 'paleta = ?'])
                        update_values.extend([paleta_id, paleta_nazwa])
                    
                    # Dostawca - zawsze jeśli podany
                    if dostawca:
                        update_fields.append('dostawca = ?')
                        update_values.append(dostawca)
                    
                    # Zdjęcia - jeśli są (tylko zdjecie_url, images opcjonalnie)
                    if images_from_file:
                        update_fields.append('zdjecie_url = ?')
                        update_values.append(images_from_file[0])
                    
                    # Wykonaj UPDATE
                    if update_fields:
                        update_values.append(asin)  # WHERE asin = ?
                        sql = "UPDATE produkty SET " + ', '.join(update_fields) + " WHERE id = ?"  # noqa: B608
                        # Zmień ostatni param z asin na id (WHERE id = ? jest bezpieczne - tylko konkretny rekord)
                        update_values[-1] = existing['id']
                        print(f"   SQL: {sql}")
                        print(f"   VALUES: {update_values}")
                        try:
                            conn.execute(sql, update_values)
                            # Spróbuj też zaktualizować images jeśli kolumna istnieje
                            if images_from_file and images_json:
                                try:
                                    conn.execute('UPDATE produkty SET images = ? WHERE id = ?', (images_json, existing['id']))
                                except Exception:
                                    pass  # Kolumna images może nie istnieć
                        except Exception as e:
                            print(f"   [CANCEL] UPDATE error: {e}")
                        updated += 1
                    else:
                        print(f"   ⚠️ Brak pól do aktualizacji!")
                
            except Exception as e:
                print(f"[CANCEL] BŁĄD przy {asin}: {e}")
                import traceback
                traceback.print_exc()
        
        # Zaktualizuj ilość produktów i sztuk w palecie
        if paleta_id:
            count = conn.execute('SELECT COUNT(*) as cnt FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()['cnt']
            # AKUMULUJ cena_zakupu — NIE nadpisuj, dodaj do istniejącej
            stara_cena = conn.execute('SELECT COALESCE(cena_zakupu, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
            nowa_cena = round(stara_cena + total_brutto, 2)
            stary_netto = conn.execute('SELECT COALESCE(cena_zakupu_netto, 0) FROM palety WHERE id = ?', (paleta_id,)).fetchone()[0]
            nowy_netto = round(stary_netto + total_netto, 2)
            try:
                conn.execute('UPDATE palety SET cena_zakupu = ?, cena_zakupu_netto = ?, ilosc_produktow = ?, ilosc_sztuk = COALESCE(ilosc_sztuk, 0) + ? WHERE id = ?', (nowa_cena, nowy_netto, count, total_qty, paleta_id))
                print(f"[BAR_CHART] Paleta {paleta_id}: {count} produktów, {total_qty} sztuk dodanych, koszt: {nowa_cena:.2f} zł")
            except:
                conn.execute('UPDATE palety SET cena_zakupu = ?, ilosc_produktow = ? WHERE id = ?', (nowa_cena, count, paleta_id))
                print(f"[BAR_CHART] Paleta {paleta_id}: {count} produktów")
        
        conn.commit()
        print(f"[CHECK_CIRCLE] COMMIT wykonany - {added} nowych, {updated} zaktualizowanych")
        
        # Uruchom auto-przetwarzanie w tle (pobieranie tytułów i generowanie opisów)
        auto_process_products(list(asins))
        
        paleta_info = f'<br><span class=material-symbols-outlined>inventory_2</span> Paleta: <b>{paleta_nazwa}</b> ({dostawca})' if paleta_nazwa else ''
        
        # Oblicz sumaryczną ilość
        total_qty = sum(p['qty'] for p in asin_prices.values())
        print(f"[BAR_CHART] DEBUG total_qty: {total_qty}, asin_prices count: {len(asin_prices)}")
        print(f"[BAR_CHART] DEBUG asin_prices: {[(k, v['qty']) for k, v in asin_prices.items()]}")
        
        # Info o wartości
        value_info = ''
        if total_brutto > 0 or total_qty > len(asins):
            value_info = f'''<br><br>
            <span class=material-symbols-outlined>bar_chart</span> <b>Podsumowanie importu:</b><br>
            Produktów: <b>{len(asins)}</b> | Sztuk łącznie: <b>{total_qty}</b><br>'''
            if total_brutto > 0:
                value_info += f'''Brutto: <span style="color:#beee00">{total_brutto:.2f} PLN</span> |
                Netto: <span style="color:#8ff5ff">{total_netto:.2f} PLN</span>'''
        
        # Komunikat z uwzględnieniem nowych i zaktualizowanych
        status_msg = []
        if added > 0:
            status_msg.append(f"dodano <b>{added}</b> nowych")
        if updated > 0:
            status_msg.append(f"zaktualizowano <b>{updated}</b>")
        status_text = ", ".join(status_msg) if status_msg else "brak zmian"
        
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> IMPORT ZAKOŃCZONY</h1></div>
            <div class="alert alert-ok">Znaleziono {len(asins)} ASIN-ów: {status_text}{paleta_info}{value_info}</div>
            <div class="alert" style="background:#1a1a2e;font-size:0.85rem">
                <span class=material-symbols-outlined>check_circle</span> Produkty dodane do <a href="/magazyn" style="color:#8ff5ff">Magazynu</a><br>
                <span class=material-symbols-outlined>sync</span> <b>Auto-przetwarzanie w tle</b> - tytuły i opisy generują się automatycznie
            </div>
            <a href="/paletomat/scraper" class="btn btn-p"><span class=material-symbols-outlined>search</span> Dodaj więcej</a>
            <a href="/paletomat/generator" class="btn btn-ok"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Zobacz postęp</a>
            <a href="/paletomat" class="back">← Powrót</a>
        ''')
        
    except Exception as e:
        return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-warn">{str(e)}</div><a href="/paletomat/scraper" class="btn btn-p">← Powrót</a>')

@paletomat_bp.route('/scraper/toggle', methods=['POST'])
def scraper_toggle():
    global _scraper_running
    _scraper_running = not _scraper_running
    return redirect('/paletomat')

@paletomat_bp.route('/rescrape/<int:product_id>', methods=['POST'])
def rescrape_product(product_id):
    """Re-scrapuje produkt z Amazona (pobiera nazwę, zdjęcia, cechy)"""
    conn = get_db()
    row = conn.execute('SELECT asin FROM produkty WHERE id = ?', (product_id,)).fetchone()
    if not row or not row['asin']:
        flash('Brak ASIN dla tego produktu', 'error')
        return redirect(request.referrer or '/paletomat')

    asin = row['asin']
    data = scrape_amazon_product(asin)

    if data and data.get('title') and not data['title'].startswith('Produkt '):
        conn.execute('UPDATE produkty SET nazwa = ? WHERE id = ?', (data['title'], product_id))
        conn.execute('UPDATE scraped SET nazwa = ? WHERE asin = ?', (data['title'], asin))
        conn.commit()
        flash(f'Zaktualizowano: {data["title"][:60]}', 'success')
    else:
        flash(f'Nie udało się pobrać nazwy dla {asin}', 'error')

    return redirect(request.referrer or '/paletomat')


# ═══════════════════════════════════════════════════════════════
# REMOTE SCRAPER API — laptop scrapuje Amazon, wysyła wyniki tu
# ═══════════════════════════════════════════════════════════════

@paletomat_bp.route('/api/scraper/asins-needed', methods=['GET'])
def api_scraper_asins_needed():
    """Zwraca ASINy które potrzebują scrapowania (brak nazwy lub fallback)"""
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT p.asin FROM produkty p
        LEFT JOIN scraped s ON UPPER(p.asin) = UPPER(s.asin)
        WHERE p.asin IS NOT NULL AND p.asin != ''
        AND (
            s.nazwa IS NULL OR s.nazwa = '' OR LENGTH(s.nazwa) < 20
            OR p.nazwa IS NULL OR p.nazwa = '' OR LENGTH(p.nazwa) < 20
        )
    """).fetchall()
    asins = [r['asin'] for r in rows]
    return jsonify({'asins': asins, 'count': len(asins)})


@paletomat_bp.route('/api/scraper/update', methods=['POST'])
def api_scraper_update():
    """Przyjmuje dane z remote scrapera i aktualizuje DB"""
    data = request.get_json()
    if not data or 'asin' not in data or 'title' not in data:
        return jsonify({'error': 'Wymagane: asin, title'}), 400

    asin = data['asin'].strip().upper()
    title = data['title'].strip()
    bullet_points = data.get('bullet_points', [])
    all_images = data.get('all_images', [])
    price = data.get('price', 0)
    category = data.get('category', '')
    product_specs = data.get('product_specs', {})

    if not title or len(title) < 10:
        return jsonify({'error': 'Title za krótki'}), 400

    conn = get_db()

    # Aktualizuj scraped
    existing = conn.execute('SELECT id FROM scraped WHERE UPPER(asin) = ?', (asin,)).fetchone()
    if existing:
        conn.execute('''UPDATE scraped SET nazwa=?, bullet_points=?, wszystkie_zdjecia=?,
                       cena_amazon=?, kategoria=?, product_specs=? WHERE id=?''',
                    (title, json.dumps(bullet_points) if bullet_points else '',
                     json.dumps(all_images) if all_images else '',
                     price, category, json.dumps(product_specs) if product_specs else '',
                     existing['id']))
    else:
        conn.execute('''INSERT INTO scraped (asin, nazwa, bullet_points, wszystkie_zdjecia,
                       cena_amazon, kategoria, product_specs, status)
                       VALUES (?, ?, ?, ?, ?, ?, ?, 'nowy')''',
                    (asin, title, json.dumps(bullet_points) if bullet_points else '',
                     json.dumps(all_images) if all_images else '',
                     price, category, json.dumps(product_specs) if product_specs else ''))

    # Aktualizuj produkty.nazwa jeśli obecna jest krótka
    conn.execute('''UPDATE produkty SET nazwa = ?
                   WHERE UPPER(asin) = ? AND (nazwa IS NULL OR nazwa = '' OR LENGTH(nazwa) < LENGTH(?))''',
                (title, asin, title))

    # Auto meta_title — generuj SEO tytuł od razu
    from modules.smart_importer import _optimize_amazon_title
    meta_title = _optimize_amazon_title(title, 75)
    if meta_title and len(meta_title) >= 20:
        conn.execute('UPDATE produkty SET meta_title = ? WHERE UPPER(asin) = ? AND (meta_title IS NULL OR LENGTH(meta_title) < 30)',
                    (meta_title, asin))

    conn.commit()
    return jsonify({'ok': True, 'asin': asin, 'title': title[:60], 'meta_title': meta_title})


@paletomat_bp.route('/rescrape-all-fallback', methods=['POST'])
def rescrape_all_fallback():
    """Synchronizuje nazwy produktów z tabeli scraped → produkty (bez live scraping)"""
    try:
        conn = get_db()

        # Krok 1: produkty z fallback nazwą ale mające poprawną nazwę w scraped
        synced = conn.execute("""
            UPDATE produkty
            SET nazwa = (SELECT s.nazwa FROM scraped s WHERE UPPER(s.asin) = UPPER(produkty.asin))
            WHERE asin IS NOT NULL AND asin != ''
            AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')
            AND EXISTS (
                SELECT 1 FROM scraped s
                WHERE UPPER(s.asin) = UPPER(produkty.asin)
                AND s.nazwa IS NOT NULL AND s.nazwa != '' AND s.nazwa NOT LIKE 'Produkt %'
            )
        """).rowcount

        # Krok 2: scraped z fallback nazwą ale produkty mające poprawną nazwę
        synced2 = conn.execute("""
            UPDATE scraped
            SET nazwa = (SELECT p.nazwa FROM produkty p WHERE UPPER(p.asin) = UPPER(scraped.asin))
            WHERE asin IS NOT NULL AND asin != ''
            AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')
            AND EXISTS (
                SELECT 1 FROM produkty p
                WHERE UPPER(p.asin) = UPPER(scraped.asin)
                AND p.nazwa IS NOT NULL AND p.nazwa != '' AND p.nazwa NOT LIKE 'Produkt %'
            )
        """).rowcount

        conn.commit()

        total = synced + synced2
        if total > 0:
            flash(f'Zsynchronizowano nazwy dla {total} produktów', 'success')
        else:
            flash('Brak produktów do synchronizacji (nazwy już pobrane)', 'info')
    except Exception as e:
        flash(f'Błąd synchronizacji: {str(e)}', 'error')

    return redirect(request.referrer or '/paletomat')


@paletomat_bp.route('/api/rescrape-names-stream/paleta/<int:paleta_id>', methods=['POST'])
def api_rescrape_names_paleta_stream(paleta_id):
    """Streamuje proces podmiany placeholder nazw 'Produkt B0...' na realne nazwy z Amazon.

    Strategia 2-fazowa:
      1) Sync z tabeli `scraped` (jeden UPDATE) - tani, dziala dla ASIN-ow
         juz raz scrapeowanych gdzie indziej.
      2) Dla ASIN-ow ktorych nie ma w `scraped` - wywolaj scrape_amazon_product
         i zapisz do produkty + scraped. Po kazdym produkcie wysyla event SSE.
    """
    def generate():
        conn = get_db()

        # Faza 1: sync z scraped (zlap to co juz mamy)
        synced = 0
        try:
            cur = conn.execute('''
                UPDATE produkty
                SET nazwa = (SELECT s.nazwa FROM scraped s WHERE UPPER(s.asin) = UPPER(produkty.asin))
                WHERE paleta_id = ?
                  AND asin IS NOT NULL AND asin != ''
                  AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')
                  AND EXISTS (
                      SELECT 1 FROM scraped s
                      WHERE UPPER(s.asin) = UPPER(produkty.asin)
                        AND s.nazwa IS NOT NULL AND s.nazwa != '' AND s.nazwa NOT LIKE 'Produkt %'
                  )
            ''', (paleta_id,))
            synced = cur.rowcount or 0
            conn.commit()
        except Exception as e:
            yield f"data: {json.dumps({'type':'error','message':f'Sync z scraped: {e}'}, ensure_ascii=False)}\n\n"

        if synced > 0:
            yield f"data: {json.dumps({'type':'sync','count':synced}, ensure_ascii=False)}\n\n"

        # Faza 2: produkty z placeholder + ASIN-em ktorych dalej nie ma w scraped
        produkty = conn.execute('''
            SELECT id, asin, nazwa
            FROM produkty
            WHERE paleta_id = ?
              AND asin IS NOT NULL AND asin != ''
              AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')
        ''', (paleta_id,)).fetchall()

        total = len(produkty)
        stats = {'type': 'done', 'total': total, 'synced': synced, 'scraped': 0, 'errors': 0, 'skipped': 0}
        processed = 0

        for p in produkty:
            processed += 1
            asin = (p['asin'] or '').strip().upper()
            nazwa_old = p['nazwa'] or ''

            ev = {'type': 'progress', 'current': processed, 'total': total, 'asin': asin, 'old_name': nazwa_old[:50]}

            # Walidacja formatu ASIN-a Amazon (musi byc B0 + 8-10 alfanum)
            if not re.match(r'^B0[A-Z0-9]{8,10}$', asin):
                stats['skipped'] += 1
                ev['source'] = 'invalid_asin'
                ev['new_name'] = None
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                continue

            try:
                amazon_data = scrape_amazon_product(asin)
                if not amazon_data or not amazon_data.get('title'):
                    stats['errors'] += 1
                    ev['source'] = 'amazon_blocked'
                    ev['new_name'] = None
                    yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                    continue

                nowa_nazwa = translate_product_name(amazon_data['title'], use_ai=True)

                # Zapisz do produkty (tylko placeholdery dla tego ASIN-a w tej palecie)
                conn.execute('''
                    UPDATE produkty SET nazwa = ?
                    WHERE paleta_id = ? AND UPPER(asin) = ?
                      AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')
                ''', (nowa_nazwa, paleta_id, asin))

                # Zapisz tez do scraped zeby inne palety mogly skorzystac
                conn.execute('''
                    INSERT INTO scraped (asin, nazwa, status) VALUES (?, ?, 'nowy')
                    ON CONFLICT(asin) DO UPDATE SET nazwa = excluded.nazwa
                ''', (asin, nowa_nazwa))
                conn.commit()

                stats['scraped'] += 1
                ev['source'] = 'amazon'
                ev['new_name'] = nowa_nazwa[:60]
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

            except Exception as e:
                stats['errors'] += 1
                ev['source'] = 'error'
                ev['new_name'] = None
                ev['error'] = str(e)[:80]
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

        yield f"data: {json.dumps(stats, ensure_ascii=False)}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'}
    )


@paletomat_bp.route('/generator')
def generator():
    # Pobierz zescrapowane produkty gotowe do generowania
    conn = get_db()
    products = conn.execute('SELECT * FROM scraped WHERE status IN ("nowy", "gotowy") ORDER BY data_scrape DESC LIMIT 150').fetchall()
    wystawione = conn.execute('SELECT COUNT(*) as cnt FROM scraped WHERE status = "wystawiony"').fetchone()['cnt']
    
    # Policz produkty z pustymi nazwami (wymagają przetworzenia)
    needs_processing = conn.execute('''SELECT COUNT(*) as cnt FROM scraped 
        WHERE status IN ("nowy", "gotowy") AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')''').fetchone()['cnt']
    
    
    # Sprawdź status Allegro
    from .allegro_api import is_authenticated
    from .database import get_config
    allegro_ok = is_authenticated()
    shipping_ok = bool(get_config('allegro_shipping_id', ''))
    
    html = f'''
    <!-- Page Header -->
    <div style="display:flex;flex-direction:column;gap:4px;margin-bottom:24px">
        <h2 style="font-family:'Space Grotesk',sans-serif;font-size:2.2rem;font-weight:800;color:#f9f5f8;letter-spacing:-0.03em;margin:0">OFFER_ENGINE</h2>
        <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap">
            <span style="padding:4px 10px;background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);color:#8ff5ff;font-size:0.6rem;font-weight:700;letter-spacing:0.15em;text-transform:uppercase">Status: {'Optimal' if allegro_ok else 'Disconnected'}</span>
            <span style="padding:4px 10px;background:rgba(255,107,155,0.08);border:1px solid rgba(255,107,155,0.2);color:#ff6b9b;font-size:0.6rem;font-weight:700;letter-spacing:0.15em;text-transform:uppercase">Scrapers: Active</span>
        </div>
    </div>

    <!-- Stats Row -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px">
        <div style="background:#131315;padding:20px;border-left:3px solid #8ff5ff;box-shadow:0 0 15px rgba(143,245,255,0.1)">
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#adaaad;text-transform:uppercase;margin-bottom:8px">EST_ROI</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:2.2rem;font-weight:800;color:#8ff5ff">{len(products)}</div>
            <div style="font-size:0.65rem;color:#767577;margin-top:4px">produktów w kolejce</div>
        </div>
        <div style="background:#131315;padding:20px;border-left:3px solid #ff6b9b;box-shadow:0 0 15px rgba(255,107,155,0.1)">
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#adaaad;text-transform:uppercase;margin-bottom:8px">ACTIVE_OFFERS</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:2.2rem;font-weight:800;color:#ff6b9b">{wystawione}</div>
            <div style="font-size:0.65rem;color:#767577;margin-top:4px">wystawionych</div>
        </div>
    </div>
    '''

    # GEN_PARAMETERS panel - left side controls
    html += '<div style="display:grid;grid-template-columns:1fr 2fr;gap:24px;align-items:start">'

    # Left: Settings & Actions
    html += f'''
    <div style="background:#1f1f22;border:1px solid rgba(72,71,74,0.15);padding:24px">
        <h3 style="font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;margin-bottom:20px;display:flex;align-items:center;gap:8px">
            <span class=material-symbols-outlined style="color:#8ff5ff">settings_suggest</span> GEN_PARAMETERS
        </h3>

        <div style="display:flex;flex-direction:column;gap:12px;margin-bottom:24px">
            <div style="display:flex;align-items:center;justify-content:space-between;padding:8px 0">
                <div>
                    <div style="font-size:0.78rem;font-weight:700;color:#f9f5f8">AUTO_PRICING</div>
                    <div style="font-size:0.65rem;color:#adaaad">AI Market Calibration</div>
                </div>
                <div style="width:36px;height:20px;background:rgba(143,245,255,0.2);border-radius:10px;position:relative">
                    <div style="width:16px;height:16px;background:#8ff5ff;border-radius:50%;position:absolute;top:2px;right:2px;box-shadow:0 0 8px #8ff5ff"></div>
                </div>
            </div>
            <div style="display:flex;align-items:center;justify-content:space-between;padding:8px 0">
                <div>
                    <div style="font-size:0.78rem;font-weight:700;color:#f9f5f8">SYNC_ALLEGRO</div>
                    <div style="font-size:0.65rem;color:#adaaad">{'Connected' if allegro_ok else 'Not connected'}</div>
                </div>
                <div style="width:36px;height:20px;background:{'rgba(143,245,255,0.2)' if allegro_ok else '#262528;border:1px solid rgba(72,71,74,0.3)'};border-radius:10px;position:relative">
                    <div style="width:16px;height:16px;background:{'#8ff5ff' if allegro_ok else '#767577'};border-radius:50%;position:absolute;top:2px;{'right' if allegro_ok else 'left'}:2px;{'box-shadow:0 0 8px #8ff5ff' if allegro_ok else ''}"></div>
                </div>
            </div>
        </div>

        <div style="border-top:1px solid rgba(72,71,74,0.15);padding-top:20px;display:flex;flex-direction:column;gap:8px">
    '''

    if allegro_ok and shipping_ok and products:
        html += f'''<a href="/paletomat/generator/mass-create" class="btn btn-ok" style="width:100%;padding:14px;text-align:center;display:flex;align-items:center;justify-content:center;gap:6px;font-size:0.72rem;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;box-shadow:0 0 15px rgba(143,245,255,0.3)"
            onclick="return confirm('Wystawić {len(products)} produktów na Allegro?')">
            <span class=material-symbols-outlined>bolt</span> RUN_BATCH_GENERATION</a>'''
    elif products:
        html += '<div style="width:100%;padding:14px;text-align:center;background:#262528;color:#767577;font-size:0.72rem;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;opacity:0.5;cursor:not-allowed"><span class=material-symbols-outlined>bolt</span> RUN_BATCH (połącz Allegro)</div>'

    if needs_processing > 0:
        html += f'''<a href="/paletomat/generator/reprocess" class="btn btn-2" style="width:100%;padding:12px;text-align:center;display:flex;align-items:center;justify-content:center;gap:6px;font-size:0.72rem"
            onclick="return confirm('Przetworzyć {needs_processing} produktów?')">
            <span class=material-symbols-outlined>sync</span> PRZETWORZ ({needs_processing})</a>'''

    if wystawione > 0:
        html += f'''<a href="/paletomat/generator/cleanup" class="btn btn-warn" style="width:100%;padding:12px;text-align:center;display:flex;align-items:center;justify-content:center;gap:6px;font-size:0.72rem"
            onclick="return confirm('Usunąć {wystawione} wystawionych szkiców?')">
            <span class=material-symbols-outlined>delete</span> USUŃ WYSTAWIONE ({wystawione})</a>'''

    html += f'<a href="/paletomat/generator/enhance-existing" class="btn" style="width:100%;padding:12px;text-align:center;display:flex;align-items:center;justify-content:center;gap:6px;font-size:0.72rem;background:rgba(245,158,11,0.15);border:1px solid rgba(245,158,11,0.25);color:#f59e0b"><span class=material-symbols-outlined>auto_awesome</span> GENERUJ ZDJĘCIA AI</a>'

    # Przycisk re-scrapuj fallback nazwy
    fallback_cnt = conn.execute("""
        SELECT COUNT(DISTINCT p.asin) as cnt FROM produkty p
        LEFT JOIN scraped s ON UPPER(p.asin) = UPPER(s.asin)
        WHERE p.asin IS NOT NULL AND p.asin != ''
        AND (p.nazwa LIKE 'Produkt %' OR s.nazwa LIKE 'Produkt %' OR s.nazwa IS NULL OR s.nazwa = '')
    """).fetchone()['cnt']
    if fallback_cnt > 0:
        html += f'''<form method="POST" action="/paletomat/rescrape-all-fallback" style="width:100%" id="rescrape-form"
            onsubmit="var btn=this.querySelector('button');btn.disabled=true;btn.innerHTML='<span>⟳</span>&nbsp;Synchronizuję...';return true;">
            <input type="hidden" name="csrf_token" value="{generate_csrf()}">
            <button type="submit" class="btn" style="width:100%;padding:12px;background:rgba(143,245,255,0.1);border:1px solid rgba(143,245,255,0.2);color:#8ff5ff;display:flex;align-items:center;justify-content:center;gap:6px;font-size:0.72rem">
                <span class=material-symbols-outlined>sync</span> POBIERZ NAZWY ({fallback_cnt})
            </button>
        </form>'''

    html += '''
        </div>
    </div>
    '''

    # Right: Scraped Queue
    html += f'''
    <div>
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
            <h3 style="font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;display:flex;align-items:center;gap:8px">
                <span class=material-symbols-outlined style="color:#ff6b9b">dataset</span> SCRAPED_QUEUED
                <span style="color:rgba(173,170,173,0.4);margin-left:4px">({len(products)})</span>
            </h3>
        </div>
        <div style="display:flex;flex-direction:column;gap:12px">
    '''

    if products:
        for p in products:
            nazwa_display = (p['nazwa'] or f"Produkt {p['asin']}")[:45]
            _p_status = p['status'] if p['status'] else 'nowy'
            _ready = _p_status in ('nowy', 'gotowy')
            _bcolor = '#beee00' if _ready else '#ff6b9b'

            html += f'''
            <div style="background:#131315;display:flex;align-items:center;border-left:3px solid {_bcolor};transition:background 0.2s"
                 onmouseover="this.style.background='#19191c'" onmouseout="this.style.background='#131315'">
                <div style="width:80px;height:72px;background:#262528;flex-shrink:0;overflow:hidden;{'filter:grayscale(1);' if not _ready else ''}">
                    <img src="{get_amazon_image_url(p['asin'])}" style="width:100%;height:100%;object-fit:cover;opacity:0.7" onerror="this.style.display='none'">
                </div>
                <div style="flex:1;padding:12px 16px">
                    <div style="font-weight:700;font-size:0.82rem;letter-spacing:0.03em;color:#f9f5f8;text-transform:uppercase;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:300px">{nazwa_display}</div>
                    <div style="display:flex;align-items:center;gap:10px;margin-top:4px">
                        <span style="font-size:0.6rem;color:#adaaad;font-family:monospace">ASIN: {p['asin']}</span>
                        <span style="font-size:0.55rem;{'color:#8ff5ff;background:rgba(143,245,255,0.08)' if _ready else 'color:#ff6b9b;background:rgba(255,107,155,0.08)'};padding:1px 6px;font-weight:700;text-transform:uppercase">{'READY' if _ready else 'MAPPING'}</span>
                    </div>
                </div>
                <div style="padding:12px 16px;background:rgba(38,37,40,0.5)">
                    <a href="/paletomat/generator/{p['asin']}" style="padding:8px 18px;background:transparent;border:1px solid rgba(143,245,255,0.3);color:#8ff5ff;font-size:0.6rem;font-weight:800;letter-spacing:0.1em;text-transform:uppercase;text-decoration:none;display:inline-block;transition:all 0.2s"
                       onmouseover="this.style.background='#8ff5ff';this.style.color='#005d63'" onmouseout="this.style.background='transparent';this.style.color='#8ff5ff'">
                        GENERATE_OFFER</a>
                </div>
            </div>'''
    else:
        html += '<div style="padding:40px;text-align:center;background:#131315;border:1px dashed rgba(255,107,155,0.2)"><span class=material-symbols-outlined style="font-size:2rem;color:#ff6b9b;margin-bottom:8px;display:block">inbox</span><div style="font-size:0.82rem;color:#adaaad">Brak produktów do wystawienia. Najpierw użyj scrapera.</div></div>'

    html += '''
        </div>
    </div>
    </div>
    '''

    html += '<div style="text-align:center;margin-top:24px"><a href="/paletomat" style="font-size:0.82rem;color:#adaaad;text-decoration:none;font-weight:600;letter-spacing:0.05em">&larr; Powrót</a></div>'
    return render(html)

def _render_stan_fields(grupy_stanow, stan_magazyn):
    """Renderuje pola stanu - grupowo jeśli są sztuki, pojedynczo jeśli nie"""
    stan_colors = {
        'Nowy': ('#beee00', '●'),
        'Powystawowy': ('#8ff5ff', '●'),
        'Używany': ('#eab308', '●'),
        'Uszkodzony': ('#ef4444', '<span class=material-symbols-outlined>fiber_manual_record</span>'),
        'Odnowiony': ('#ff6b9b', '●'),
    }
    
    if grupy_stanow:
        # Wiele grup stanów - osobna oferta per stan
        html = '''<div style="background:rgba(190,238,0,0.08);border:1px solid #beee0044;border-radius:10px;padding:14px;margin-bottom:12px">
            <div style="font-size:0.8rem;color:#22c55e;font-weight:700;margin-bottom:10px">
                <span class=material-symbols-outlined>inventory_2</span> Produkt rozdzielony na stany — zostaną wystawione <b>osobne oferty</b> per stan, zgrupowane ilościowo
            </div>'''
        for g in grupy_stanow:
            stan = g['stan']
            ile = g['ilosc']
            color, icon = stan_colors.get(stan, ('#94a3b8', '⬜'))
            html += f'''<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;background:#0a0a0f;border-radius:8px;padding:10px;border:1px solid {color}33">
                <span style="font-size:1.1rem">{icon}</span>
                <span style="font-weight:600;color:{color};flex:1">{stan}</span>
                <span style="font-size:0.85rem;color:#94a3b8">ilość:</span>
                <input type="hidden" name="stan_grupa_{stan}" value="{stan}">
                <input type="number" name="ilosc_grupa_{stan}" value="{ile}" min="1" max="{ile}"
                    style="width:70px;padding:6px;background:#1e1e2e;border:1px solid {color}66;border-radius:6px;color:#fff;text-align:center;font-weight:700">
                <span style="font-size:0.75rem;color:#64748b">/ {ile} szt</span>
            </div>'''
        html += '<div style="font-size:0.7rem;color:#64748b;margin-top:6px"><span class=material-symbols-outlined>lightbulb</span> Jedna oferta na Allegro = jeden stan. System wystawi tyle ofert ile masz grup.</div>'
        html += '</div>'
        # Hidden single stan field for non-grouped fallback
        html += f'<input type="hidden" name="stan" value="{grupy_stanow[0]["stan"] if grupy_stanow else stan_magazyn}">'
        return html
    else:
        # Brak podziału na sztuki - zwykły select
        opts = ''
        for stan_val, (color, icon) in stan_colors.items():
            sel = 'selected' if stan_val == stan_magazyn else ''
            opts += f'<option value="{stan_val}" {sel}>{icon} {stan_val}</option>'
        return f'''<div class="form-group">
            <label><span class=material-symbols-outlined>inventory_2</span> Stan produktu</label>
            <select name="stan" class="form-ctrl" style="padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                {opts}
            </select>
            <div style="font-size:0.65rem;color:#64748b;margin-top:2px">Stan z magazynu: <b style="color:#fff">{stan_magazyn}</b></div>
        </div>'''


@paletomat_bp.route('/generator/mass-create')
def generator_mass_create():
    """Strona masowego wystawiania z progressem"""
    from .allegro_api import is_authenticated
    from .database import get_config
    
    # Sprawdź wymagania PRZED startem
    allegro_ok = is_authenticated()
    shipping_id = get_config('allegro_shipping_id', '')
    
    if not allegro_ok:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nie jesteś zalogowany do Allegro!</div>
            <p style="color:#94a3b8;margin:15px 0">Musisz najpierw połączyć konto Allegro żeby wystawiać oferty.</p>
            <a href="/allegro" class="btn btn-ok">[KEY] Połącz z Allegro</a>
            <a href="/paletomat/generator" class="back">← Powrót</a>
        ''')
    
    if not shipping_id:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Brak cennika wysyłki!</div>
            <p style="color:#94a3b8;margin:15px 0">Musisz wybrać cennik wysyłki w ustawieniach Allegro.</p>
            <a href="/allegro/config" class="btn btn-ok"><span class=material-symbols-outlined>settings</span> Ustawienia Allegro</a>
            <a href="/paletomat/generator" class="back">← Powrót</a>
        ''')
    
    conn = get_db()
    products = conn.execute('SELECT * FROM scraped WHERE status IN ("nowy", "gotowy") ORDER BY data_scrape DESC LIMIT 150').fetchall()
    
    count = len(products)
    
    if count == 0:
        return render('''
            <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> MASOWE WYSTAWIANIE</h1></div>
            <div class="alert alert-warn">Brak produktów do wystawienia</div>
            <a href="/paletomat/generator" class="back">← Powrót</a>
        ''')
    
    html = f'''
    <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> MASOWE WYSTAWIANIE</h1><small>{count} produktów</small></div>
    
    <div class="alert alert-ok" style="font-size:0.85rem">
        <span class=material-symbols-outlined>check_circle</span> Allegro połączone | <span class=material-symbols-outlined>check_circle</span> Cennik wysyłki OK | <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Startujemy automatycznie...
    </div>
    
    <div class="card" style="text-align:center;padding:30px">
        <div id="progress-icon" style="font-size:3rem;margin-bottom:15px">
            <div style="display:inline-block;width:48px;height:48px;border:4px solid rgba(143,245,255,0.2);border-top-color:#8ff5ff;border-radius:50%;animation:spin 1s linear infinite"></div>
        </div>
        <div id="progress-text" style="font-size:1.2rem;font-weight:600;margin-bottom:10px">Laczenie...</div>
        <div style="background:#1e1e2e;border-radius:10px;height:20px;overflow:hidden;margin-bottom:10px">
            <div id="progress-bar" style="background:linear-gradient(90deg,#8ff5ff,#ff6b9b);height:100%;width:0%;transition:width 0.3s"></div>
        </div>
        <div id="progress-count" style="font-size:0.85rem;color:#64748b">0 / {count}</div>
        <div id="timer" style="font-size:1.4rem;font-weight:700;color:#8ff5ff;margin-top:8px;font-variant-numeric:tabular-nums">00:00</div>
        <div id="timer-avg" style="font-size:0.8rem;color:#94a3b8;margin-top:2px"></div>
    </div>

    <style>@keyframes spin {{ 0% {{ transform:rotate(0deg) }} 100% {{ transform:rotate(360deg) }} }}</style>

    <div id="connection-error" class="alert alert-err" style="display:none;margin-bottom:15px">
        <span class=material-symbols-outlined>cancel</span> Błąd połączenia z serwerem!<br>
        <small>Sprawdzam ponownie za <span id="retry-countdown">5</span> sekund...</small>
    </div>
    
    <div id="log" class="card" style="max-height:400px;overflow-y:auto;font-family:monospace;font-size:0.8rem">
        <div style="color:#64748b;padding:4px 0"><span class=material-symbols-outlined>sync</span> Inicjalizacja...</div>
    </div>
    
    <div id="done-buttons" style="display:none;margin-top:15px">
        <a href="/paletomat/generator" class="btn btn-p">← Powrót do listy</a>
        <a href="/allegro/moje-oferty" class="btn btn-ok" style="margin-left:8px"><span class=material-symbols-outlined>list_alt</span> Zobacz oferty</a>
    </div>
    
    <script>
    (function() {{
        const log = document.getElementById('log');
        const bar = document.getElementById('progress-bar');
        const text = document.getElementById('progress-text');
        const count = document.getElementById('progress-count');
        const icon = document.getElementById('progress-icon');
        const connectionError = document.getElementById('connection-error');
        const retryCountdown = document.getElementById('retry-countdown');
        const doneButtons = document.getElementById('done-buttons');
        const _timerEl = document.getElementById('timer');
        const _timerAvg = document.getElementById('timer-avg');
        const _totalProducts = {count};

        // TIMER
        const _startTime = Date.now();
        const _timerInterval = setInterval(() => {{
            const elapsed = Math.floor((Date.now() - _startTime) / 1000);
            const m = String(Math.floor(elapsed / 60)).padStart(2, '0');
            const s = String(elapsed % 60).padStart(2, '0');
            _timerEl.textContent = m + ':' + s;
        }}, 1000);
        function _stopTimer() {{
            clearInterval(_timerInterval);
            const elapsed = ((Date.now() - _startTime) / 1000).toFixed(1);
            _timerEl.textContent = elapsed + 's';
            _timerEl.style.color = '#22c55e';
        }}
        function _updateAvg() {{
            const done = sukces + bledy;
            if (done > 0) {{
                const elapsed = (Date.now() - _startTime) / 1000;
                const avg = (elapsed / done).toFixed(1);
                const remaining = _totalProducts - done;
                const eta = Math.ceil(avg * remaining);
                const etaM = Math.floor(eta / 60);
                const etaS = eta % 60;
                _timerAvg.textContent = avg + 's/oferta | ETA: ~' + (etaM > 0 ? etaM + 'min ' : '') + etaS + 's';
            }}
        }}

        let sukces = 0, bledy = 0;
        let evtSource = null;
        let retryCount = 0;
        const MAX_RETRIES = 5;
        let retryTimer = null;
        
        function addLog(message, color = '#94a3b8') {{
            log.innerHTML += `<div style="color:${{color}};padding:4px 0">${{message}}</div>`;
            log.scrollTop = log.scrollHeight;
        }}
        
        function showConnectionError(show, retryIn = 5) {{
            connectionError.style.display = show ? 'block' : 'none';
            if (show) {{
                retryCountdown.textContent = retryIn;
            }}
        }}
        
        function connectToStream() {{
            if (retryCount >= MAX_RETRIES) {{
                icon.textContent = '';
                text.textContent = 'Nie udało się połączyć';
                addLog('<span class=material-symbols-outlined>cancel</span> Przekroczono limit prób połączenia. Sprawdź czy backend działa i ngrok jest uruchomiony!', '#ef4444');
                showConnectionError(false);
                doneButtons.style.display = 'flex';
                doneButtons.style.gap = '10px';
                return;
            }}
            
            retryCount++;
            addLog(`<span class=material-symbols-outlined>sync</span> Próba połączenia ${{retryCount}}/${{MAX_RETRIES}}...`, '#3b82f6');
            
            try {{
                if (evtSource) {{
                    evtSource.close();
                }}
                
                evtSource = new EventSource('/paletomat/generator/mass-create-stream');
                
                let connectionTimeout = setTimeout(() => {{
                    addLog('<span class=material-symbols-outlined>warning</span> Timeout połączenia (10s) - próbuję ponownie...', '#f59e0b');
                    if (evtSource) {{
                        evtSource.close();
                    }}
                    retryWithDelay(5);
                }}, 10000);
                
                evtSource.onopen = function() {{
                    clearTimeout(connectionTimeout);
                    addLog('<span class=material-symbols-outlined>check_circle</span> Połączono z serwerem!', '#22c55e');
                    showConnectionError(false);
                    retryCount = 0;
                }};
                
                evtSource.onmessage = function(e) {{
                    clearTimeout(connectionTimeout);
                    
                    try {{
                        const data = JSON.parse(e.data);
                        
                        if (data.type === 'start') {{
                            addLog('<span class=material-symbols-outlined style="font-size:1rem;vertical-align:middle">rocket_launch</span> Start wystawiania ' + data.total + ' produktów...', '#3b82f6');
                            icon.textContent = '';
                        }}
                        else if (data.type === 'progress') {{
                            bar.style.width = data.percent + '%';
                            count.textContent = data.current + ' / ' + data.total;
                            text.textContent = 'Wystawianie: ' + (data.asin || '');
                        }}
                        else if (data.type === 'success') {{
                            sukces++;
                            const zdjeciaInfo = data.zdjecia ? ` (${{data.zdjecia}} zdjec)` : '';
                            addLog('<span class=material-symbols-outlined>check_circle</span> ' + data.asin + ': ' + data.title + zdjeciaInfo, '#22c55e');
                            _updateAvg();
                        }}
                        else if (data.type === 'error') {{
                            bledy++;
                            addLog('<span class=material-symbols-outlined>cancel</span> ' + data.asin + ': ' + data.error, '#ef4444');
                            _updateAvg();
                        }}
                        else if (data.type === 'log') {{
                            const logColor = data.color || '#94a3b8';
                            addLog(data.message, logColor);
                        }}
                        else if (data.type === 'done') {{
                            _stopTimer();
                            icon.innerHTML = sukces > bledy ? '<span class=material-symbols-outlined>check_circle</span>' : '<span class=material-symbols-outlined>warning</span>';
                            const totalTime = ((Date.now() - _startTime) / 1000).toFixed(1);
                            const avg = sukces > 0 ? ((Date.now() - _startTime) / 1000 / sukces).toFixed(1) : '0';
                            text.textContent = 'Gotowe: ' + sukces + ' OK, ' + bledy + ' bledow';
                            _timerAvg.textContent = 'Lacznie: ' + totalTime + 's | Srednia: ' + avg + 's/oferta';
                            bar.style.width = '100%';
                            addLog('<span class=material-symbols-outlined>check_circle</span> GOTOWE! Sukces: ' + sukces + ', Bledy: ' + bledy, '#22c55e');
                            doneButtons.style.display = 'flex';
                            doneButtons.style.gap = '10px';

                            if (evtSource) {{
                                evtSource.close();
                            }}
                        }}
                    }} catch (parseError) {{
                        console.error('JSON parse error:', parseError, e.data);
                        addLog('<span class=material-symbols-outlined>warning</span> Błąd parsowania danych: ' + parseError.message, '#f59e0b');
                    }}
                }};
                
                evtSource.onerror = function(err) {{
                    clearTimeout(connectionTimeout);
                    console.error('EventSource error:', err);
                    
                    if (evtSource.readyState === EventSource.CLOSED) {{
                        addLog('<span class=material-symbols-outlined>cancel</span> Połączenie zamknięte przez serwer', '#ef4444');
                    }} else if (evtSource.readyState === EventSource.CONNECTING) {{
                        addLog('<span class=material-symbols-outlined>sync</span> Ponowne łączenie...', '#f59e0b');
                    }} else {{
                        addLog('<span class=material-symbols-outlined>cancel</span> Błąd połączenia: ' + (err.message || 'Unknown error'), '#ef4444');
                    }}
                    
                    if (evtSource) {{
                        evtSource.close();
                    }}
                    
                    retryWithDelay(5);
                }};
                
            }} catch (error) {{
                console.error('Connection error:', error);
                addLog('<span class=material-symbols-outlined>cancel</span> Błąd inicjalizacji: ' + error.message, '#ef4444');
                retryWithDelay(5);
            }}
        }}
        
        function retryWithDelay(seconds) {{
            if (retryCount >= MAX_RETRIES) {{
                icon.textContent = '';
                text.textContent = 'Nie udało się połączyć';
                addLog('<span class=material-symbols-outlined>cancel</span> Osiągnięto limit prób. Sprawdź logi serwera i upewnij się że backend działa!', '#ef4444');
                doneButtons.style.display = 'flex';
                doneButtons.style.gap = '10px';
                return;
            }}
            
            showConnectionError(true, seconds);
            icon.textContent = '';
            text.textContent = 'Ponowne łączenie...';
            
            let countdown = seconds;
            retryTimer = setInterval(() => {{
                countdown--;
                retryCountdown.textContent = countdown;
                
                if (countdown <= 0) {{
                    clearInterval(retryTimer);
                    showConnectionError(false);
                    connectToStream();
                }}
            }}, 1000);
        }}
        
        addLog('<span class=material-symbols-outlined>sync</span> Inicjalizacja połączenia...', '#3b82f6');
        connectToStream();
        
        window.addEventListener('beforeunload', function() {{
            if (evtSource) {{
                evtSource.close();
            }}
            if (retryTimer) {{
                clearInterval(retryTimer);
            }}
        }});
    }})();
    </script>
    '''
    return render(html)



# ═══════════════════════════════════════════════════════════════════════════
# MASOWE WYSTAWIANIE Z PALETY - Adrian's custom feature v3.1.0
# ═══════════════════════════════════════════════════════════════════════════

@paletomat_bp.route('/generator/mass-create-from-paleta')
def generator_mass_create_from_paleta():
    """Masowe wystawianie produktów z magazynu"""
    from .allegro_api import is_authenticated
    from .database import get_config
    
    paleta_id = request.args.get('paleta_id', type=int)
    ids_str = request.args.get('ids', '')
    
    if not ids_str:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nie wybrano żadnych produktów!</div>
            <a href="/palety" class="back">← Powrót</a>
        ''')
    
    try:
        product_ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
    except:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nieprawidłowe ID produktów!</div>
            <a href="/palety" class="back">← Powrót</a>
        ''')
    
    allegro_ok = is_authenticated()
    shipping_id = get_config('allegro_shipping_id', '')

    if not allegro_ok:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nie jesteś zalogowany do Allegro!</div>
            <a href="/allegro" class="btn btn-ok">[KEY] Połącz z Allegro</a>
            <a href="/palety" class="back">← Powrót</a>
        ''')

    if not shipping_id:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Brak cennika wysyłki!</div>
            <a href="/allegro/config" class="btn btn-ok"><span class=material-symbols-outlined>settings</span> Ustawienia Allegro</a>
            <a href="/palety" class="back">← Powrót</a>
        ''')

    # Pobierz wszystkie cenniki z Allegro (do dropdownu wyboru przed wystawianiem)
    shipping_options_html = ''
    try:
        from .allegro_api import get_shipping_rates
        rates, _ = get_shipping_rates()
        if rates and 'shippingRates' in rates:
            for rate in rates['shippingRates']:
                sel = 'selected' if rate['id'] == shipping_id else ''
                shipping_options_html += f'<option value="{rate["id"]}" {sel}>{rate["name"]}</option>'
    except Exception as e:
        print(f'[mass-create] Blad pobierania cennikow: {e}')

    count = len(product_ids)

    # Per-product shipping_map z mass-edit (jesli user wybral rozne cenniki per wiersz)
    # Format URL: ?shipping_map=<JSON {"pid": "uuid", ...}>
    # PRE-RENDER do JS variable + info HTML zeby uniknac nested f-string (Python 3.11 fail)
    _ship_map_raw_in = request.args.get('shipping_map', '').strip()
    _ship_map_in = {}
    if _ship_map_raw_in:
        try:
            _parsed = json.loads(_ship_map_raw_in)
            _ship_map_in = {str(k): v for k, v in _parsed.items() if v}
        except Exception as _e:
            print(f'[mass-create page] shipping_map parse fail: {_e}')
    ship_map_json_for_js = json.dumps(_ship_map_in) if _ship_map_in else 'null'
    ship_map_info_html = ''
    if _ship_map_in:
        ship_map_info_html = (
            '<div style="padding:10px 14px;margin-bottom:12px;background:rgba(143,245,255,0.06);'
            'border-left:3px solid #8ff5ff;border-radius:8px;font-size:0.82rem;color:#8ff5ff">'
            '<span class="material-symbols-outlined" style="font-size:1rem;vertical-align:middle">tune</span> '
            f'<b>Per-produkt cennik z mass-edit aktywny</b> ({len(_ship_map_in)} produktów). '
            'Wybór poniżej nadpisze TYLKO produkty bez indywidualnego cennika.'
            '</div>'
        )

    html = f'''
    <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> MASOWE WYSTAWIANIE Z PALETY</h1><small>{count} produktów</small></div>

    <div class="alert alert-ok" style="font-size:0.85rem">
        <span class=material-symbols-outlined>check_circle</span> Allegro połączone | <span class=material-symbols-outlined>check_circle</span> Cennik wysyłki OK | <span class=material-symbols-outlined>check_circle</span> Produkty wybrane
    </div>

    {ship_map_info_html}

    <!-- WYBOR CENNIKA WYSYLKI (per batch) -->
    <div class="card" id="shipping-picker" style="padding:20px;margin-bottom:15px;border:2px solid rgba(143,245,255,0.2)">
        <div style="display:flex;flex-wrap:wrap;align-items:center;gap:14px">
            <div style="flex:1;min-width:280px">
                <label style="display:block;font-size:0.78rem;color:#94a3b8;margin-bottom:6px;text-transform:uppercase;letter-spacing:1px">
                    <span class=material-symbols-outlined style='font-size:1rem;vertical-align:middle'>local_shipping</span>
                    Cennik wysyłki dla tej partii ({count} produktów)
                </label>
                {'<select id="shipping-select" class="form-control" style="width:100%">' + shipping_options_html + '</select>' if shipping_options_html else '<input type="text" id="shipping-select-input" class="form-control" value="' + shipping_id + '" placeholder="UUID cennika">'}
                <div style="font-size:0.72rem;color:#64748b;margin-top:6px">
                    Wybierz cennik pasujacy do gabarytu (np. maly cennik dla paczkomatu, duzy dla kuriera).
                    Default = ustawiony w /allegro.
                </div>
            </div>
            <button id="start-btn" class="btn btn-ok" style="white-space:nowrap;padding:14px 26px;font-size:0.95rem;font-weight:700">
                <span class='material-symbols-outlined' style='font-size:1.1rem;vertical-align:middle'>rocket_launch</span> Wystaw ({count})
            </button>
        </div>
    </div>
    
    <div class="card" style="text-align:center;padding:30px">
        <div id="progress-icon" style="font-size:3rem;margin-bottom:15px">
            <div style="display:inline-block;width:48px;height:48px;border:4px solid rgba(143,245,255,0.2);border-top-color:#8ff5ff;border-radius:50%;animation:spin 1s linear infinite"></div>
        </div>
        <div id="progress-text" style="font-size:1.2rem;font-weight:600;margin-bottom:10px">Laczenie...</div>
        <div style="background:#1e1e2e;border-radius:10px;height:20px;overflow:hidden;margin-bottom:10px">
            <div id="progress-bar" style="background:linear-gradient(90deg,#8ff5ff,#ff6b9b);height:100%;width:0%;transition:width 0.3s"></div>
        </div>
        <div id="progress-count" style="font-size:0.85rem;color:#64748b">0 / {count}</div>
        <div id="timer" style="font-size:1.4rem;font-weight:700;color:#8ff5ff;margin-top:8px;font-variant-numeric:tabular-nums">00:00</div>
        <div id="timer-avg" style="font-size:0.8rem;color:#94a3b8;margin-top:2px"></div>
    </div>

    <style>@keyframes spin {{ 0% {{ transform:rotate(0deg) }} 100% {{ transform:rotate(360deg) }} }}</style>

    <div id="log" class="card" style="max-height:400px;overflow-y:auto;font-family:monospace;font-size:0.8rem">
        <div style="color:#64748b;padding:4px 0"><span class=material-symbols-outlined>sync</span> Inicjalizacja...</div>
    </div>

    <div id="done-buttons" style="display:none;margin-top:15px">
        <a href="/palety/{paleta_id}" class="btn btn-p">← Powrot do palety</a>
        <a href="/allegro/moje-oferty" class="btn btn-ok" style="margin-left:8px"><span class=material-symbols-outlined>list_alt</span> Zobacz oferty na Allegro</a>
    </div>

    <script>
    // TIMER
    const _startTime = Date.now();
    const _timerEl = document.getElementById('timer');
    const _timerAvg = document.getElementById('timer-avg');
    const _totalProducts = {count};
    const _timerInterval = setInterval(() => {{
        const elapsed = Math.floor((Date.now() - _startTime) / 1000);
        const m = String(Math.floor(elapsed / 60)).padStart(2, '0');
        const s = String(elapsed % 60).padStart(2, '0');
        _timerEl.textContent = m + ':' + s;
    }}, 1000);
    function _stopTimer() {{
        clearInterval(_timerInterval);
        const elapsed = ((Date.now() - _startTime) / 1000).toFixed(1);
        _timerEl.textContent = elapsed + 's';
        _timerEl.style.color = '#22c55e';
    }}
    function _updateAvg() {{
        const done = sukces + bledy;
        if (done > 0) {{
            const elapsed = (Date.now() - _startTime) / 1000;
            const avg = (elapsed / done).toFixed(1);
            const remaining = _totalProducts - done;
            const eta = Math.ceil(avg * remaining);
            const etaM = Math.floor(eta / 60);
            const etaS = eta % 60;
            _timerAvg.textContent = avg + 's/oferta | ETA: ~' + (etaM > 0 ? etaM + 'min ' : '') + etaS + 's';
        }}
    }}

    // Per-product shipping_map z mass-edit (lub null jesli brak)
    const PRESET_SHIPPING_MAP = {ship_map_json_for_js};

    // Odpal stream dopiero po klik Start (uzytkownik wybiera cennik wczesniej)
    let evtSource = null;
    document.getElementById('start-btn').onclick = function() {{
        const selEl = document.getElementById('shipping-select') || document.getElementById('shipping-select-input');
        const shipId = selEl ? selEl.value.trim() : '';
        let url = '/paletomat/generator/mass-create-from-paleta-stream?ids={ids_str}' +
                  (shipId ? '&shipping_id=' + encodeURIComponent(shipId) : '');
        // Doklej per-produkt cennik jesli byl z mass-edit
        if (PRESET_SHIPPING_MAP && Object.keys(PRESET_SHIPPING_MAP).length > 0) {{
            url += '&shipping_map=' + encodeURIComponent(JSON.stringify(PRESET_SHIPPING_MAP));
        }}
        document.getElementById('shipping-picker').style.opacity = '0.5';
        document.getElementById('shipping-picker').style.pointerEvents = 'none';
        this.disabled = true;
        this.innerHTML = '<span class=material-symbols-outlined>hourglass_top</span> Wystawiam...';
        evtSource = new EventSource(url);
        attachStreamHandlers();
    }};

    function attachStreamHandlers() {{
    const log = document.getElementById('log');
    const bar = document.getElementById('progress-bar');
    const text = document.getElementById('progress-text');
    const count = document.getElementById('progress-count');
    const icon = document.getElementById('progress-icon');
    let sukces = 0, bledy = 0;

    evtSource.onmessage = function(e) {{
        const data = JSON.parse(e.data);

        if (data.type === 'start') {{
            log.innerHTML += '<div style="color:#3b82f6;padding:4px 0"><span class=material-symbols-outlined style="font-size:1rem;vertical-align:middle">rocket_launch</span> Start wystawiania ' + data.total + ' produktow z magazynu...</div>';
        }}
        else if (data.type === 'progress') {{
            bar.style.width = data.percent + '%';
            count.textContent = data.current + ' / ' + data.total;
            text.textContent = 'Wystawianie: ' + (data.title || '').substring(0, 40);
        }}
        else if (data.type === 'success') {{
            sukces++;
            log.innerHTML += '<div style="color:#22c55e;padding:4px 0"><span class=material-symbols-outlined>check_circle</span> ' + data.title + (data.price ? ' (' + data.price + ' zl)' : (data.message ? ' — ' + data.message : '')) + '</div>';
            log.scrollTop = log.scrollHeight;
            _updateAvg();
        }}
        else if (data.type === 'error') {{
            bledy++;
            log.innerHTML += '<div style="color:#ef4444;padding:4px 0"><span class=material-symbols-outlined>cancel</span> ' + (data.title || 'Produkt') + ': ' + data.error + '</div>';
            log.scrollTop = log.scrollHeight;
            _updateAvg();
        }}
        else if (data.type === 'log') {{
            const color = data.color || '#94a3b8';
            log.innerHTML += '<div style="color:' + color + ';padding:4px 0">' + data.message + '</div>';
            log.scrollTop = log.scrollHeight;
        }}
        else if (data.type === 'done') {{
            evtSource.close();
            _stopTimer();
            bar.style.width = '100%';
            icon.innerHTML = (sukces > 0 ? '<span class=material-symbols-outlined>check_circle</span>' : '<span class=material-symbols-outlined>cancel</span>');
            const totalTime = ((Date.now() - _startTime) / 1000).toFixed(1);
            const avg = sukces > 0 ? ((Date.now() - _startTime) / 1000 / sukces).toFixed(1) : '0';
            text.textContent = 'Gotowe! ' + sukces + ' wystawiono, ' + bledy + ' bledow';
            _timerAvg.textContent = 'Lacznie: ' + totalTime + 's | Srednia: ' + avg + 's/oferta';
            document.getElementById('done-buttons').style.display = 'flex';
            document.getElementById('done-buttons').style.gap = '10px';
        }}
    }};

    evtSource.onerror = function() {{
        evtSource.close();
        _stopTimer();
        _timerEl.style.color = '#ef4444';
        icon.innerHTML = '<span class=material-symbols-outlined>cancel</span>';
        text.textContent = 'Blad polaczenia ze streamem';
        log.innerHTML += '<div style="color:#ef4444;padding:4px 0"><span class=material-symbols-outlined>cancel</span> Utracono polaczenie</div>';
        document.getElementById('done-buttons').style.display = 'flex';
    }};
    }}  // end attachStreamHandlers
    </script>
    '''
    return render(html)


@paletomat_bp.route('/generator/mass-create-from-paleta-stream')
def generator_mass_create_from_paleta_stream():
    """SSE stream dla masowego wystawiania produktów z magazynu"""
    from .allegro_api import create_offer, is_authenticated, upload_image_to_allegro, search_categories
    from .database import get_config
    import time
    import sqlite3
    
   # --- POPRAWIONY FRAGMENT (Wklej w miejsce starego) ---

    # 1. Pobieramy dane tutaj (NAD funkcją generate), póki mamy dostęp do requestu
    ids_str = request.args.get('ids', '')
    force_new = request.args.get('force', '0') == '1'
    # Override cennika:
    # - shipping_id = jeden cennik dla wszystkich (legacy / bulk)
    # - shipping_map = JSON {pid: ship_id, ...} per-product (nowy)
    shipping_override = request.args.get('shipping_id', '').strip() or None
    shipping_map_raw = request.args.get('shipping_map', '').strip()
    shipping_map = {}
    if shipping_map_raw:
        try:
            shipping_map = json.loads(shipping_map_raw)
            # Klucze jako str (product_id)
            shipping_map = {str(k): v for k, v in shipping_map.items() if v}
        except Exception as _e:
            print(f'[mass-create] shipping_map parse fail: {_e}')

    def generate():
        # Wewnątrz funkcji już NIE wywołujemy request.args.get
        # Korzystamy ze zmiennej 'ids_str' pobranej wyżej
        yield ": ping\n\n"  # Potwierdza że stream działa (Waitress flush)

        if not ids_str:
            yield "data: " + json.dumps({'type': 'error', 'title': 'System', 'error': 'Brak ID produktów'}) + "\n\n"
            yield "data: " + json.dumps({'type': 'done'}) + "\n\n"
            return

        try:
            # Tu używamy zmiennej z góry
            product_ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
        except:
            yield "data: " + json.dumps({'type': 'error', 'title': 'System', 'error': 'Nieprawidłowe ID'}) + "\n\n"
            yield "data: " + json.dumps({'type': 'done'}) + "\n\n"
            return

        if not is_authenticated():
            yield "data: " + json.dumps({'type': 'error', 'title': 'System', 'error': 'Nie zalogowany do Allegro'}) + "\n\n"
            yield "data: " + json.dumps({'type': 'done'}) + "\n\n"
            return
            
        # ... dalsza część kodu bez zmian ...

        try:
            conn = get_db()
            placeholders = ','.join('?' * len(product_ids))
            # Najpierw - zlicz pominiete "dla siebie" zeby zaraportowac uzytkownikowi
            try:
                _skipped_dla_siebie = conn.execute(
                    "SELECT id, nazwa, COALESCE(powod_zatrzymania,'') AS powod FROM produkty "
                    "WHERE id IN (" + placeholders + ") AND COALESCE(dla_siebie,0) = 1",
                    product_ids).fetchall()
            except Exception:
                _skipped_dla_siebie = []
            if _skipped_dla_siebie:
                _names = ', '.join((dict(r).get('nazwa') or '?')[:30] for r in _skipped_dla_siebie[:5])
                _more = f' +{len(_skipped_dla_siebie)-5}' if len(_skipped_dla_siebie) > 5 else ''
                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>lock</span> Pominieto {len(_skipped_dla_siebie)} produktow zatrzymanych dla siebie: {_names}{_more}', 'color': '#ef4444'}) + "\n\n"

            products = conn.execute(
                "SELECT * FROM produkty WHERE id IN (" + placeholders + ") "
                "AND status NOT IN ('usuniety', 'sprzedany') "
                "AND COALESCE(dla_siebie, 0) = 0 "
                "ORDER BY id",
                product_ids).fetchall()
            total = len(products)
        except Exception as _dberr:
            yield "data: " + json.dumps({'type': 'error', 'title': 'System', 'error': f'Błąd bazy danych: {str(_dberr)[:80]}'}) + "\n\n"
            yield "data: " + json.dumps({'type': 'done'}) + "\n\n"
            return

        if total == 0:
            yield "data: " + json.dumps({'type': 'error', 'title': 'System', 'error': 'Brak produktów do wystawienia (wszystkie sprzedane lub zatrzymane dla siebie)'}) + "\n\n"
            yield "data: " + json.dumps({'type': 'done'}) + "\n\n"
            return
        
        # === SYNC: Synchronizuj statusy ofert z Allegro przed wystawianiem ===
        try:
            from .allegro_api import sync_offers_status
            yield "data: " + json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>sync</span> Synchronizacja ofert z Allegro...', 'color': '#3b82f6'}) + "\n\n"
            sync_result = sync_offers_status()
            if sync_result and not sync_result.get('error'):
                _synced = sync_result.get('updated', 0)
                _ended = sync_result.get('ended', 0)
                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> Sync: {sync_result.get("active", 0)} aktywnych, {_ended} zakończonych, {_synced} zaktualizowanych', 'color': '#22c55e'}) + "\n\n"
            else:
                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Sync: {sync_result.get("error", "?")}', 'color': '#f59e0b'}) + "\n\n"
        except Exception as e:
            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Sync error: {str(e)[:50]}', 'color': '#f59e0b'}) + "\n\n"

        # === BATCH DEDUP: grupuj produkty z tym samym ASIN — sumuj ilości ===
        _seen_asins = {}  # asin -> index w products
        _skip_ids = set()  # product IDs do pominięcia (duplikaty)
        for _di, _dp in enumerate(products):
            _dp = dict(_dp)
            _asin_key = (_dp.get('asin') or '').strip().upper()
            if _asin_key and len(_asin_key) >= 5 and _asin_key not in ('N/A', 'NAN', 'NONE'):
                if _asin_key in _seen_asins:
                    # Duplikat ASIN — sumuj ilość do pierwszego, pomiń ten
                    _first_idx = _seen_asins[_asin_key]
                    _first_p = dict(products[_first_idx])
                    _add_qty = int(_dp.get('ilosc', 0) or 0)
                    if _add_qty > 0:
                        _first_p['ilosc'] = int(_first_p.get('ilosc', 0) or 0) + _add_qty
                        products[_first_idx] = _first_p
                    _skip_ids.add(_dp['id'])
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>merge</span> Batch dedup: {_dp["nazwa"][:30]} (id={_dp["id"]}) → zsumowano z id={_first_p["id"]} (ASIN: {_asin_key})', 'color': '#f59e0b'}) + "\n\n"
                else:
                    _seen_asins[_asin_key] = _di

        # Filtruj duplikaty
        products = [p for p in products if dict(p)['id'] not in _skip_ids]
        total = len(products)

        # Tracking ofert utworzonych w tym batchu (asin -> allegro_id)
        _batch_created = {}

        yield "data: " + json.dumps({'type': 'start', 'total': total}) + "\n\n"
        time.sleep(0.5)

        for i, p in enumerate(products):
            p = dict(p)
            product_id = p['id']
            nazwa = p['nazwa']

            yield "data: " + json.dumps({'type': 'progress', 'current': i+1, 'total': total, 'percent': int((i+1)/total*100), 'title': nazwa[:40]}) + "\n\n"

            try:
                cena = float(p['cena_allegro']) if p['cena_allegro'] else 99.99
                ilosc = int(p['ilosc']) if p['ilosc'] else 1
                ean = p.get('ean') or None
                asin = p.get('asin') or None
                zdjecie_url = p.get('zdjecie_url') or ''
                kategoria = p.get('kategoria', 'inne') or 'inne'

                # === BATCH DEDUP: sprawdź czy ten ASIN już został wystawiony w tym batchu ===
                _asin_upper = (asin or '').strip().upper()
                if _asin_upper and _asin_upper in _batch_created:
                    _prev_id = _batch_created[_asin_upper]
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>block</span> ASIN {_asin_upper} już wystawiony w tym batchu (oferta #{_prev_id}) — pomijam duplikat', 'color': '#f59e0b'}) + "\n\n"
                    yield "data: " + json.dumps({'type': 'success', 'title': nazwa[:40], 'message': f'Pominięto (duplikat ASIN w batchu)'}) + "\n\n"
                    time.sleep(0.3)
                    continue

                # === DEDUPLIKACJA: sprawdź czy produkt ma już aktywną ofertę ===
                if force_new:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>bolt</span> Force: pomijam deduplikację', 'color': '#f59e0b'}) + "\n\n"
                    existing_offer = None  # skip dedup
                else:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>search</span> Dedup: id={product_id}, ean={ean}, asin={asin}', 'color': '#6366f1'}) + "\n\n"

                    existing_offer = None
                    _nazwa_lower = (nazwa or '').lower()[:40]

                # Pomocnicza: sprawdź czy nazwa oferty pasuje do produktu
                def _nazwa_match(offer_tytul):
                    """Porównaj nazwy — żeby nie łączyć różnych produktów"""
                    if not offer_tytul or not _nazwa_lower:
                        return True  # brak danych = pozwól
                    t = offer_tytul.lower()[:60]
                    # Wyciągnij słowa > 2 znaki (pomijaj "na", "do", "ze" itp.)
                    # Ignoruj generyczne słowa które pasują do wielu produktów
                    _ignore = {'uniwersalne', 'uniwersalny', 'premium', 'zestaw', 'komplet', 'sztuk', 'szt', 'nowy', 'nowe', 'nowa'}
                    words_p = [w for w in _nazwa_lower.split() if len(w) > 2 and w not in _ignore][:5]
                    words_o = [w for w in t.split() if len(w) > 2 and w not in _ignore][:5]
                    if not words_p or not words_o:
                        return True
                    matching = sum(1 for w in words_p if any(w in wo or wo in w for wo in words_o))
                    # Wymagaj >= 50% słów z produktu (minimum 2)
                    threshold = max(2, len(words_p) // 2 + 1)
                    result = matching >= threshold
                    if not result:
                        print(f"      [BLOC] _nazwa_match FAIL: prod={words_p} vs oferta={words_o}, match={matching}/{threshold}")
                    return result

                # 1. Szukaj po produkt_id + weryfikacja nazwy
                existing_offer = None
                _pid_candidate = conn.execute('''
                    SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status
                    FROM oferty o
                    WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published')
                    AND o.allegro_id IS NOT NULL AND o.allegro_id != ''
                    AND o.produkt_id = ?
                    LIMIT 1
                ''', (product_id,)).fetchone()
                if _pid_candidate and _nazwa_match(_pid_candidate['tytul']):
                    existing_offer = _pid_candidate
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>search</span> Match po produkt_id={product_id}', 'color': '#6366f1'}) + "\n\n"
                elif _pid_candidate:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>block</span> produkt_id={product_id} znaleziony ale nazwa nie pasuje: "{_pid_candidate["tytul"][:40]}"', 'color': '#f59e0b'}) + "\n\n"
                    # Odlinkuj błędne powiązanie
                    conn.execute('UPDATE oferty SET produkt_id = NULL WHERE id = ?', (_pid_candidate['id'],))
                    conn.commit()
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>build</span> Auto-odlinkowano błędne powiązanie oferty', 'color': '#f59e0b'}) + "\n\n"

                # Stan produktu — kluczowy dla dedup (różny stan = osobna oferta Allegro)
                _current_stan = (p.get('stan') or 'Nowy').strip()

                # 2. Szukaj po ASIN (najwiarygodniejszy identyfikator) + pasujący stan
                if asin and not existing_offer:
                    all_prod_ids = [r['id'] for r in conn.execute('SELECT id FROM produkty WHERE asin = ?', (asin,)).fetchall()]
                    if all_prod_ids:
                        ph = ','.join('?' * len(all_prod_ids))
                        candidates = conn.execute(
                            "SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status, p2.stan as prod_stan "
                            "FROM oferty o "
                            "LEFT JOIN produkty p2 ON o.produkt_id = p2.id "
                            "WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published') "
                            "AND o.allegro_id IS NOT NULL AND o.allegro_id != '' "
                            "AND o.produkt_id IN (" + ph + ")",
                            all_prod_ids).fetchall()
                        for c in candidates:
                            if not _nazwa_match(c['tytul']):
                                continue
                            # Sprawdź czy stan się zgadza — różny stan = osobna oferta
                            _offer_stan = (c['prod_stan'] or 'Nowy').strip()
                            if _offer_stan != _current_stan:
                                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>inventory_2</span> ASIN {asin}: oferta ma stan "{_offer_stan}" != "{_current_stan}" — tworzę osobną ofertę', 'color': '#8b5cf6'}) + "\n\n"
                                continue
                            existing_offer = c
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>search</span> Match po ASIN {asin} + stan "{_current_stan}": oferta="{c["tytul"][:40]}"', 'color': '#6366f1'}) + "\n\n"
                            break
                        if not existing_offer:
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>search</span> ASIN {asin}: brak pasującej oferty (stan={_current_stan}) — tworzę nową', 'color': '#6366f1'}) + "\n\n"

                # 3. Szukaj po EAN (TYLKO jeśli produkt NIE MA ASIN — EAN bez ASIN jest ryzykowny)
                if ean and not existing_offer and not asin:
                    all_prod_ids = [r['id'] for r in conn.execute('SELECT id FROM produkty WHERE ean = ?', (ean,)).fetchall()]
                    if all_prod_ids:
                        ph = ','.join('?' * len(all_prod_ids))
                        candidates = conn.execute(
                            "SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status, p2.stan as prod_stan "
                            "FROM oferty o "
                            "LEFT JOIN produkty p2 ON o.produkt_id = p2.id "
                            "WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published') "
                            "AND o.allegro_id IS NOT NULL AND o.allegro_id != '' "
                            "AND o.produkt_id IN (" + ph + ")",
                            all_prod_ids).fetchall()
                        for c in candidates:
                            if not _nazwa_match(c['tytul']):
                                continue
                            _offer_stan = (c['prod_stan'] or 'Nowy').strip()
                            if _offer_stan != _current_stan:
                                continue  # różny stan = osobna oferta
                            existing_offer = c
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>search</span> Match po EAN {ean} + stan "{_current_stan}": oferta="{c["tytul"][:40]}"', 'color': '#6366f1'}) + "\n\n"
                            break

                if force_new:
                    existing_offer = None  # wymuś nowe oferty

                if existing_offer:
                    ex = dict(existing_offer)
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>inventory_2</span> Znaleziono aktywną ofertę: {ex["allegro_id"]} ({ex["tytul"][:30]}), szt: {ex["ilosc"]}', 'color': '#f59e0b'}) + "\n\n"
                    try:
                        from .allegro_api import update_offer_stock
                        obecna_ilosc = ex.get('ilosc', 0) or 0
                        dodaj = int(p.get('ilosc', 0) or 1)
                        # UWAGA: nie pomijamy nawet jeśli oferta ma już tyle szt —
                        # mogą być 3 osobne rekordy po 1 szt każdy, każdy musi dodać swoje qty
                        new_qty = obecna_ilosc + dodaj
                        result, error = update_offer_stock(ex['allegro_id'], new_qty)
                        if error and str(error).startswith('OFFER_NOT_EXISTS:'):
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>delete</span> Oferta #{ex["allegro_id"]} nie istnieje — zakończona w DB, tworzę nową', 'color': '#f59e0b'}) + "\n\n"
                        elif error:
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Błąd dodawania do oferty: {error}', 'color': '#f59e0b'}) + "\n\n"
                        else:
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> +{dodaj} szt → oferta #{ex["allegro_id"]} (było {obecna_ilosc}, teraz {new_qty})', 'color': '#22c55e'}) + "\n\n"
                            yield "data: " + json.dumps({'type': 'success', 'title': nazwa[:40], 'message': f'Dodano {dodaj} szt do istniejącej oferty (łącznie {new_qty})'}) + "\n\n"
                            time.sleep(0.3)
                            continue
                    except Exception as e:
                        yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Błąd dedup: {str(e)[:50]}', 'color': '#f59e0b'}) + "\n\n"
                else:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'[ADD_CIRCLE] Brak aktywnej oferty — tworzę nową', 'color': '#3b82f6'}) + "\n\n"

                # === POPRAWKA: POBIERZ WSZYSTKIE ZDJĘCIA, NIE TYLKO GŁÓWNE ===
                wszystkie_zdjecia = []
                
                # 1. Sprawdź kolumnę 'images' w produkty (JSON z listą zdjęć)
                try:
                    saved_images = p.get('images', '') or ''
                    if saved_images:
                        wszystkie_zdjecia = json.loads(saved_images)
                        if not isinstance(wszystkie_zdjecia, list):
                            wszystkie_zdjecia = []
                except:
                    pass
                
                # 2. FALLBACK: Pobierz z tabeli scraped (kolumna wszystkie_zdjecia)
                if not wszystkie_zdjecia and asin:
                    try:
                        scraped_row = conn.execute('SELECT wszystkie_zdjecia FROM scraped WHERE asin = ?', (asin,)).fetchone()
                        if scraped_row and scraped_row['wszystkie_zdjecia']:
                            wszystkie_zdjecia = json.loads(scraped_row['wszystkie_zdjecia'])
                            if not isinstance(wszystkie_zdjecia, list):
                                wszystkie_zdjecia = []
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> Pobrano {len(wszystkie_zdjecia)} zdjęć z cache scraped', 'color': '#8b5cf6'}) + "\n\n"
                    except Exception as e:
                        pass
                
                # 3. Fallback: użyj głównego zdjęcia
                if not wszystkie_zdjecia and zdjecie_url:
                    wszystkie_zdjecia = [zdjecie_url]
                
                # 3.5 SPRAWDŹ: czy lokalne pliki istnieją, re-download jeśli nie
                wszystkie_zdjecia, img_logs = _ensure_local_images(wszystkie_zdjecia, asin, zdjecie_url)
                for msg, color in img_logs:
                    yield "data: " + json.dumps({'type': 'log', 'message': msg, 'color': color}) + "\n\n"
                
                # === ZBIERZ ZDJĘCIA: enhanced (walidowane) + downloaded + oryginalne ===
                _base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                _enh_dir = os.path.join(_base_dir, 'static', 'enhanced', str(asin or product_id))
                _dl_dir = os.path.join(_base_dir, 'static', 'downloads', str(asin or product_id))

                # Zbierz WSZYSTKIE dostępne lokalne zdjęcia (enhanced + downloads)
                _local_images = []
                _enh_count = 0
                _dl_count = 0

                # 1. Enhanced images (walidowane - plik musi istnieć i >5KB)
                if os.path.isdir(_enh_dir):
                    for f in sorted(os.listdir(_enh_dir)):
                        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                            fp = os.path.join(_enh_dir, f)
                            sz = os.path.getsize(fp) if os.path.exists(fp) else 0
                            if sz > 5000:  # min 5KB - odrzuć puste/uszkodzone
                                _local_images.append(fp)
                                _enh_count += 1
                                print(f"    [IMG] Enhanced: {f} ({sz} bytes) ✓")
                            else:
                                print(f"    [IMG] Enhanced SKIP: {f} ({sz} bytes - za mały)")

                # 2. Downloaded images (uzupełnij do 8)
                if os.path.isdir(_dl_dir) and len(_local_images) < 8:
                    for f in sorted(os.listdir(_dl_dir)):
                        if len(_local_images) >= 8:
                            break
                        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                            fp = os.path.join(_dl_dir, f)
                            sz = os.path.getsize(fp) if os.path.exists(fp) else 0
                            if sz > 5000 and fp not in _local_images:
                                _local_images.append(fp)
                                _dl_count += 1

                if _local_images:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>folder</span> Znaleziono {len(_local_images)} lokalnych zdjęć ({_enh_count} enhanced + {_dl_count} downloaded)', 'color': '#3b82f6'}) + "\n\n"
                    wszystkie_zdjecia = _local_images[:8]

                # Backup oryginalnych URL-i (z scraped) na wypadek fallbacku
                _original_urls = [u for u in (wszystkie_zdjecia if wszystkie_zdjecia else []) if isinstance(u, str) and u.startswith('http')]

                # 3. Upload WSZYSTKICH zdjęć do Allegro (max 8)
                # Używamy wątku + keepalive żeby SSE stream się nie zerwał
                zdjecia_urls = []
                if wszystkie_zdjecia:
                    _imgs_to_upload = wszystkie_zdjecia[:8]
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> {len(_imgs_to_upload)} zdjęć, uploaduję...', 'color': '#3b82f6'}) + "\n\n"

                    import threading, queue
                    _upload_q = queue.Queue()

                    def _upload_worker():
                        for idx, img_url in enumerate(_imgs_to_upload, 1):
                            try:
                                _is_local = not str(img_url).startswith('http')
                                _exists = os.path.exists(img_url) if _is_local else True
                                _size = os.path.getsize(img_url) if (_is_local and _exists) else 0
                                print(f"    [Upload {idx}] Path: {img_url}")
                                print(f"    [Upload {idx}] Local={_is_local} Exists={_exists} Size={_size}")
                                allegro_url = upload_image_to_allegro(img_url, asin=asin)
                                _upload_q.put(('ok', idx, allegro_url, img_url))
                            except Exception as e:
                                import traceback
                                traceback.print_exc()
                                _upload_q.put(('error', idx, str(e)[:80], img_url))
                        _upload_q.put(('done', 0, None, None))

                    _upload_thread = threading.Thread(target=_upload_worker, daemon=True)
                    _upload_thread.start()

                    _upload_done = False
                    _failed_paths = []
                    while not _upload_done:
                        try:
                            status, idx, result, src_path = _upload_q.get(timeout=3)
                            if status == 'done':
                                _upload_done = True
                            elif status == 'ok' and result:
                                zdjecia_urls.append(result)
                                yield "data: " + json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>check_circle</span> [{idx}/{len(_imgs_to_upload)}] Uploadowano', 'color': '#22c55e'}) + "\n\n"
                            elif status == 'ok':
                                _failed_paths.append(src_path)
                                yield "data: " + json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> [{idx}] Upload failed (sprawdź logi serwera)', 'color': '#ef4444'}) + "\n\n"
                            else:
                                _failed_paths.append(src_path)
                                yield "data: " + json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> [{idx}] {result}', 'color': '#ef4444'}) + "\n\n"
                        except queue.Empty:
                            yield ": keepalive\n\n"

                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> Uploadowano {len(zdjecia_urls)}/{len(_imgs_to_upload)} zdjęć', 'color': '#22c55e' if zdjecia_urls else '#ef4444'}) + "\n\n"

                    # === FALLBACK: jeśli 0 zdjęć - spróbuj oryginalne URL-e z CDN ===
                    if not zdjecia_urls and _original_urls:
                        yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>sync</span> Próbuję {len(_original_urls)} URL-i z CDN...', 'color': '#f59e0b'}) + "\n\n"
                        for _ui, _uurl in enumerate(_original_urls[:8], 1):
                            try:
                                _ur = upload_image_to_allegro(_uurl, asin=asin)
                                if _ur:
                                    zdjecia_urls.append(_ur)
                                    yield "data: " + json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>check_circle</span> [CDN {_ui}] OK', 'color': '#22c55e'}) + "\n\n"
                                else:
                                    yield "data: " + json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> [CDN {_ui}] Nie przeszło', 'color': '#ef4444'}) + "\n\n"
                            except:
                                pass
                            yield ": keepalive\n\n"
                        if zdjecia_urls:
                            yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> CDN fallback: {len(zdjecia_urls)} zdjęć', 'color': '#22c55e'}) + "\n\n"
                
                from .utils import optimize_title_seo, generuj_opis_html_pro, generuj_gpsr_info
                
                # Pobierz klucz Gemini API
                gemini_key = get_config('gemini_api_key', '')
                
                # Pobierz bullet points z produkty
                bullet_points_raw = p.get('bullet_points') or ''
                bullet_points = []
                if bullet_points_raw:
                    try:
                        import json as json_module
                        bullet_points = json_module.loads(bullet_points_raw) if isinstance(bullet_points_raw, str) else bullet_points_raw
                        if not isinstance(bullet_points, list):
                            bullet_points = []
                    except:
                        pass
                
                # FALLBACK: Pobierz bullet_points z tabeli scraped jeśli brak w produkty
                if not bullet_points and asin:
                    try:
                        scraped_row = conn.execute('SELECT bullet_points FROM scraped WHERE asin = ?', (asin,)).fetchone()
                        if scraped_row and scraped_row['bullet_points']:
                            bp_raw = scraped_row['bullet_points']
                            import json as json_module
                            bullet_points = json_module.loads(bp_raw) if isinstance(bp_raw, str) else bp_raw
                            if not isinstance(bullet_points, list):
                                bullet_points = []
                            if bullet_points:
                                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>edit_note</span> Pobrano {len(bullet_points)} cech z cache scraped', 'color': '#8b5cf6'}) + "\n\n"
                    except Exception as e:
                        pass
                
                # Użyj meta_title jeśli istnieje i jest sensowny, w przeciwnym razie pełna nazwa
                if p.get('meta_title') and len(p['meta_title']) >= 30:
                    tytul_seo = p['meta_title'][:75]
                else:
                    tytul_seo = optimize_title_seo(nazwa, 75)

                # Kod magazynowy (do użytku wewnętrznego, NIE do tytułu/opisu Allegro)
                _km = p.get('kod_magazynowy') or f"MAG-{product_id:05d}"

                # Auto-wykryj kategorię Allegro po POLSKIM tytule (nie angielskim!)
                kategoria_id = None
                try:
                    _cat_search = tytul_seo[:50] if tytul_seo else nazwa[:50]
                    cat_result, cat_error = search_categories(_cat_search)
                    if cat_result and cat_result.get('matchingCategories'):
                        kategoria_id = cat_result['matchingCategories'][0].get('id')
                        print(f"[FOLD] Kategoria z polskiego tytułu '{_cat_search}': {kategoria_id}")
                except:
                    pass

                # Pobierz specyfikację produktu (potrzebna do GPSR)
                product_specs = {}
                if asin:
                    try:
                        _specs_row = conn.execute('SELECT product_specs FROM scraped WHERE asin = ?', (asin,)).fetchone()
                        if _specs_row and _specs_row['product_specs']:
                            product_specs = json.loads(_specs_row['product_specs'])
                            if product_specs:
                                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>list_alt</span> Specyfikacja: {len(product_specs)} parametrów', 'color': '#8b5cf6'}) + "\n\n"
                    except:
                        pass

                # Generuj opis + GPSR RÓWNOLEGLE (oba to Gemini calls)
                yield "data: " + json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>bolt</span> Generuję opis + GPSR równolegle...', 'color': '#3b82f6'}) + "\n\n"
                _zdjecia_do_opisu = zdjecia_urls if zdjecia_urls else (wszystkie_zdjecia if wszystkie_zdjecia else ([zdjecie_url] if zdjecie_url else []))

                # Helper: uruchom funkcję w wątku z keepalive co 3s
                def _run_with_keepalive(func, *args, **kwargs):
                    """Uruchamia blokującą funkcję w wątku, zwraca wynik"""
                    _result_q = queue.Queue()
                    def _worker():
                        try:
                            r = func(*args, **kwargs)
                            _result_q.put(('ok', r))
                        except Exception as e:
                            _result_q.put(('error', e))
                    t = threading.Thread(target=_worker, daemon=True)
                    t.start()
                    while True:
                        try:
                            status, val = _result_q.get(timeout=3)
                            if status == 'ok':
                                return val
                            raise val
                        except queue.Empty:
                            yield ": keepalive\n\n"

                # Generuj opis w tle z keepalive
                opis_html = ''
                gpsr = ''
                from concurrent.futures import ThreadPoolExecutor, as_completed
                with ThreadPoolExecutor(max_workers=2) as executor:
                    f_opis = executor.submit(generuj_opis_html_pro,
                        nazwa, _zdjecia_do_opisu,
                        kategoria, bullet_points, gemini_key=gemini_key, asin=asin,
                        kod_magazynowy=_km
                    )
                    f_gpsr = executor.submit(generuj_gpsr_info, nazwa, kategoria, product_specs=product_specs)

                    # Keepalive podczas oczekiwania na Gemini
                    _futures = {f_opis: 'opis', f_gpsr: 'gpsr'}
                    _done_count = 0
                    while _done_count < 2:
                        _newly_done = [f for f in _futures if f.done() and _futures[f] != 'got']
                        for f in _newly_done:
                            _done_count += 1
                            _futures[f] = 'got'
                        if _done_count < 2:
                            yield ": keepalive\n\n"
                            time.sleep(3)

                    try:
                        _opis_result = f_opis.result()
                        opis_html = _opis_result[0] if isinstance(_opis_result, tuple) else _opis_result
                    except Exception as e:
                        print(f"   [CANCEL] Opis error: {e}")
                    try:
                        gpsr = f_gpsr.result()
                    except Exception as e:
                        print(f"   [CANCEL] GPSR error: {e}")

                if gpsr:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>shield</span> GPSR: {len(gpsr)} znaków', 'color': '#22c55e'}) + "\n\n"

                # Logi diagnostyczne
                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>folder</span> Kategoria: {kategoria_id or "auto-detect"}', 'color': '#6366f1'}) + "\n\n"
                if ean:
                    yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>bar_chart</span> EAN: {ean}', 'color': '#6366f1'}) + "\n\n"
                yield "data: " + json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>smart_toy</span> Tworzę ofertę + parametry AI...', 'color': '#3b82f6'}) + "\n\n"

                # Wybor cennika dla TEGO produktu:
                # 1. Per-product z mapy (najwazniejsze - klient wybral indywidualnie)
                # 2. Bulk override (klient ustawil ten sam dla wszystkich)
                # 3. None -> create_offer uzyje config default
                _ship_for_product = shipping_map.get(str(product_id)) or shipping_override

                # create_offer w wątku z keepalive
                _offer_q = queue.Queue()
                def _create_offer_worker():
                    try:
                        r = create_offer(
                            nazwa=tytul_seo, opis=opis_html, cena=cena,
                            zdjecia_urls=zdjecia_urls if zdjecia_urls else None,
                            kategoria_id=kategoria_id, ilosc=ilosc,
                            ean=ean, asin=asin, gpsr=gpsr,
                            product_specs=product_specs, bullet_points=bullet_points,
                            kod_magazynowy=_km,
                            shipping_id_override=_ship_for_product,
                        )
                        _offer_q.put(r)
                    except Exception as e:
                        _offer_q.put((None, str(e)))

                _offer_t = threading.Thread(target=_create_offer_worker, daemon=True)
                _offer_t.start()
                while True:
                    try:
                        result, error = _offer_q.get(timeout=3)
                        break
                    except queue.Empty:
                        yield ": keepalive\n\n"

                if error:
                    yield "data: " + json.dumps({'type': 'error', 'title': nazwa[:40], 'error': error[:80]}) + "\n\n"
                else:
                    offer_id = result.get('id') if result else None

                    # Zapisz do batch tracking żeby nie tworzyć duplikatów w ramach tego batcha
                    if offer_id and _asin_upper:
                        _batch_created[_asin_upper] = offer_id

                    max_db_retries = 5
                    for db_attempt in range(max_db_retries):
                        try:
                            # create_offer tworzy ofertę jako INACTIVE (szkic) na Allegro —
                            # dopóki nie zostanie opublikowana, produkt NIE jest "wystawiony"
                            # i oferta ma status 'draft' (widoczna na /magazyn/do-wystawienia)

                            # === KLUCZOWE: Zapisz link oferta→produkt do tabeli oferty ===
                            # Bez tego sync zamówień nie powiąże sprzedaży z produktem!
                            if offer_id:
                                conn.execute('''INSERT OR REPLACE INTO oferty
                                    (tytul, opis, cena, ilosc, status, allegro_id, produkt_id, data_aktualizacji)
                                    VALUES (?, ?, ?, ?, 'draft', ?, ?, datetime('now'))''',
                                    (tytul_seo[:100], '', cena, ilosc, str(offer_id), product_id))

                            try:
                                from .inventory_utils import add_historia
                                add_historia(product_id, 'szkic_utworzony', f'Utworzono szkic na Allegro za {cena:,.0f} zł', {'allegro_id': offer_id, 'cena': cena})
                            except:
                                pass

                            # Commit po każdym produkcie - żeby dedup widział nowe oferty
                            conn.commit()

                            break

                        except sqlite3.OperationalError as e:
                            if 'database is locked' in str(e) and db_attempt < max_db_retries - 1:
                                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Baza zablokowana, retry {db_attempt+1}...', 'color': '#f59e0b'}) + "\n\n"
                                time.sleep(1 * (db_attempt + 1))
                                continue
                            else:
                                yield "data: " + json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> Błąd bazy: {str(e)[:40]}', 'color': '#ef4444'}) + "\n\n"
                                break

                    yield "data: " + json.dumps({'type': 'success', 'title': nazwa[:40], 'price': int(cena)}) + "\n\n"
                
                time.sleep(1.5)
                
            except Exception as e:
                yield "data: " + json.dumps({'type': 'error', 'title': nazwa[:40], 'error': str(e)[:80]}) + "\n\n"
        
        try:
            conn.commit()
        except:
            pass

        yield "data: " + json.dumps({'type': 'done'}) + "\n\n"

    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })

@paletomat_bp.route('/generator/mass-create-stream')
def generator_mass_create_stream():
    """SSE stream dla masowego wystawiania"""
    from .allegro_api import create_offer, is_authenticated, upload_image_to_allegro, get_category_parameters, search_categories
    from .database import get_config
    import time
    
    def generate():
        yield ": ping\n\n"  # Potwierdza że stream działa (Waitress flush)

        # Pobierz klucz Gemini API
        gemini_key = get_config('gemini_api_key', '')

        # Sprawdź autoryzację
        if not is_authenticated():
            yield f"data: {json.dumps({'type': 'error', 'asin': '-', 'error': 'Nie zalogowany do Allegro'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return

        shipping_id = get_config('allegro_shipping_id', '')
        if not shipping_id:
            yield f"data: {json.dumps({'type': 'error', 'asin': '-', 'error': 'Brak cennika wysyłki'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return

        try:
            conn = get_db()
            products = conn.execute('SELECT * FROM scraped WHERE status NOT IN ("usuniety", "blad") ORDER BY data_scrape DESC LIMIT 150').fetchall()
            total = len(products)
        except Exception as _dberr:
            yield f"data: {json.dumps({'type': 'error', 'asin': '-', 'error': f'Błąd bazy: {str(_dberr)[:80]}'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return

        if total == 0:
            yield f"data: {json.dumps({'type': 'error', 'asin': '-', 'error': 'Brak produktów'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return
        
        # Start
        yield f"data: {json.dumps({'type': 'start', 'total': total})}\n\n"
        
        for i, p in enumerate(products):
            p = dict(p)  # Konwertuj Row na dict
            asin = p['asin']
            
            # Progress
            yield f"data: {json.dumps({'type': 'progress', 'current': i+1, 'total': total, 'percent': int((i+1)/total*100), 'asin': asin})}\n\n"
            
            try:
                # Pobierz dane
                nazwa = p['nazwa'] or f'Produkt {asin}'
                
                # Scrapuj jeśli brak nazwy
                if nazwa == f'Produkt {asin}' or len(nazwa) < 10:
                    amazon_data = scrape_amazon_product(asin)
                    if amazon_data and amazon_data.get('title'):
                        nazwa = amazon_data['title']
                
                # === DUPLIKAT CHECK ===
                _existing_offer = conn.execute(
                    "SELECT id FROM oferty WHERE allegro_id IS NOT NULL AND allegro_id != '' AND produkt_id IN (SELECT id FROM produkty WHERE asin = ?) AND status = 'aktywna'",
                    (asin,)
                ).fetchone()
                if _existing_offer:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'⏭ {asin}: Oferta już istnieje — pomijam'})}\n\n"
                    continue

                # === ILOŚĆ, EAN I ID Z MAGAZYNU ===
                # SUMA ilości ze wszystkich produktów z tym ASIN
                _sum_row = conn.execute(
                    'SELECT SUM(ilosc) as total, MIN(id) as first_id FROM produkty WHERE (asin = ? OR ean = ?) AND ilosc > 0',
                    (asin, asin)
                ).fetchone()
                magazyn_produkt = conn.execute(
                    'SELECT id, ilosc, ean FROM produkty WHERE (asin = ? OR ean = ?) AND ilosc > 0 ORDER BY ilosc DESC LIMIT 1',
                    (asin, asin)
                ).fetchone()
                if magazyn_produkt:
                    magazyn_produkt = dict(magazyn_produkt)
                _produkt_id = magazyn_produkt['id'] if magazyn_produkt else None
                ilosc = int(_sum_row['total']) if _sum_row and _sum_row['total'] else (magazyn_produkt['ilosc'] if magazyn_produkt else 1)
                ean = magazyn_produkt.get('ean') if magazyn_produkt and magazyn_produkt.get('ean') else None
                # Fallback EAN ze scraped jeśli brak w produkty
                if not ean:
                    _ean_row = conn.execute('SELECT ean FROM scraped WHERE asin = ? AND ean IS NOT NULL AND ean != ""', (asin,)).fetchone()
                    if _ean_row:
                        ean = _ean_row['ean']
                
                # Zdjęcia - FIXED: dodano szczegółowe logi
                wszystkie_zdjecia = []
                try:
                    saved = p.get('wszystkie_zdjecia', '') or ''
                    if saved:
                        wszystkie_zdjecia = json.loads(saved)
                        if not isinstance(wszystkie_zdjecia, list):
                            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> {asin}: wszystkie_zdjecia nie jest listą (typ: {type(wszystkie_zdjecia).__name__})'})}\n\n"
                            wszystkie_zdjecia = []
                        else:
                            # DIAGNOSTYKA: Pokaż wszystkie URL-e
                            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> {asin}: Znaleziono {len(wszystkie_zdjecia)} zdjęć w bazie'})}\n\n"
                            if wszystkie_zdjecia:
                                for idx, url in enumerate(wszystkie_zdjecia[:3], 1):
                                    yield f"data: {json.dumps({'type': 'log', 'message': f'   [{idx}] {url[:70]}...'})}\n\n"
                                if len(wszystkie_zdjecia) > 3:
                                    yield f"data: {json.dumps({'type': 'log', 'message': f'   ... i {len(wszystkie_zdjecia)-3} więcej'})}\n\n"
                except json.JSONDecodeError as e:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> {asin}: Błąd JSON w wszystkie_zdjecia: {str(e)[:50]}'})}\n\n"
                    yield f"data: {json.dumps({'type': 'log', 'message': f'   Raw: {saved[:100]}'})}\n\n"
                except Exception as e:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> {asin}: Błąd parsowania zdjęć: {str(e)[:50]}'})}\n\n"
                
                # Fallback na pojedyncze zdjęcie TYLKO gdy lista jest pusta
                if not wszystkie_zdjecia:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> {asin}: Brak zdjęć w bazie, używam fallback...'})}\n\n"
                    img_url = p.get('zdjecie_url') or get_amazon_image_url(asin)
                    if img_url:
                        wszystkie_zdjecia = [img_url]
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> {asin}: Fallback - 1 zdjęcie z URL'})}\n\n"
                
                # Bullet points (cechy produktu)
                bullet_points = []
                try:
                    saved_bp = p.get('bullet_points', '') or ''
                    if saved_bp:
                        bullet_points = json.loads(saved_bp)
                except:
                    pass
                
                # FALLBACK: Pobierz bullet_points z tabeli scraped
                if not bullet_points and asin:
                    try:
                        scraped_row = conn.execute('SELECT bullet_points FROM scraped WHERE asin = ?', (asin,)).fetchone()
                        if scraped_row and scraped_row['bullet_points']:
                            bullet_points = json.loads(scraped_row['bullet_points'])
                            if bullet_points:
                                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>edit_note</span> {asin}: Pobrano {len(bullet_points)} cech z cache'})}\n\n"
                    except:
                        pass
                
                # Kategoria
                kategoria = p.get('kategoria', 'inne') or 'inne'
                
                # SPRAWDŹ: czy lokalne pliki istnieją, re-download jeśli nie
                zdjecie_url_fallback = p.get('zdjecie_url') or ''
                wszystkie_zdjecia, img_logs = _ensure_local_images(wszystkie_zdjecia, asin, zdjecie_url_fallback)
                for msg, color in img_logs:
                    yield f"data: {json.dumps({'type': 'log', 'message': msg, 'color': color})}\n\n"
                
                # === UŻYJ ENHANCED ZDJĘĆ (pre-generated przez scraper) ===
                _enh_dir_chk = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced', str(asin))
                if os.path.isdir(_enh_dir_chk):
                    _enh_fs = sorted([os.path.join(_enh_dir_chk, f) for f in os.listdir(_enh_dir_chk) if f.endswith('.jpg')])
                    if _enh_fs:
                        wszystkie_zdjecia = _enh_fs[:8]
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>auto_awesome</span> {asin}: Użyto {len(_enh_fs)} zdjęć AI (pre-generated)'})}\n\n"

                # UPLOAD ZDJĘĆ DO ALLEGRO
                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>upload</span> {asin}: Rozpoczynam upload {len(wszystkie_zdjecia[:8])} zdjęć...'})}\n\n"

                uploaded_urls = []
                for idx, img_url in enumerate(wszystkie_zdjecia[:8], 1):  # Max 8 zdjęć
                    try:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'   [{idx}/{len(wszystkie_zdjecia[:8])}] Uploaduję...'})}\n\n"
                        allegro_url = upload_image_to_allegro(img_url, asin=asin)
                        if allegro_url:
                            uploaded_urls.append(allegro_url)
                            yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>check_circle</span> [{idx}] Sukces'})}\n\n"
                        else:
                            yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> [{idx}] Brak URL z Allegro'})}\n\n"
                    except Exception as upload_err:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> [{idx}] Błąd: {str(upload_err)[:40]}'})}\n\n"
                
                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> {asin}: Uploadowano {len(uploaded_urls)}/{len(wszystkie_zdjecia[:8])} zdjęć'})}\n\n"
                
                # Generuj tytuł i opis (NOWA WERSJA - z bullet points + ASIN!)
                tytul_seo = optimize_title_seo(nazwa, 75)

                # === SUGEROWANA KATEGORIA (po polskim tytule!) ===
                kategoria_id = None
                try:
                    _cat_search = tytul_seo[:50] if tytul_seo else nazwa[:50]
                    cat_result, cat_error = search_categories(_cat_search)
                    if cat_result and cat_result.get('matchingCategories'):
                        kategoria_id = cat_result['matchingCategories'][0].get('id')
                        print(f"[FOLD] Kategoria z polskiego tytułu '{_cat_search}': {kategoria_id}")
                except:
                    pass
                # Kod magazynowy (do użytku wewnętrznego, NIE do tytułu/opisu Allegro)
                _km = p.get('kod_magazynowy') or f"MAG-{product_id:05d}"
                # Pobierz specyfikację produktu (potrzebna do GPSR)
                product_specs = {}
                if asin:
                    try:
                        _specs_row = conn.execute('SELECT product_specs FROM scraped WHERE asin = ?', (asin,)).fetchone()
                        if _specs_row and _specs_row['product_specs']:
                            product_specs = json.loads(_specs_row['product_specs'])
                            if product_specs:
                                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>list_alt</span> Specyfikacja: {len(product_specs)} parametrów', 'color': '#8b5cf6'})}\n\n"
                    except:
                        pass

                # Generuj opis + GPSR RÓWNOLEGLE
                from concurrent.futures import ThreadPoolExecutor
                with ThreadPoolExecutor(max_workers=2) as executor:
                    f_opis = executor.submit(generuj_opis_html_pro, nazwa, uploaded_urls if uploaded_urls else wszystkie_zdjecia, kategoria, bullet_points, gemini_key=gemini_key, asin=asin)
                    f_gpsr = executor.submit(generuj_gpsr_info, nazwa, kategoria, product_specs=product_specs)
                    opis_html, _ = f_opis.result()
                    gpsr = f_gpsr.result()

                # Cena
                cena_amazon = p.get('cena_amazon') or 0
                if cena_amazon and float(cena_amazon) > 0:
                    wynik = oblicz_cene_allegro(float(cena_amazon) * 4.35, 40, 'inne')
                    cena = float(wynik.get('cena_sugerowana', 99.99))
                else:
                    cena = 99.99

                # WYSTAW NA ALLEGRO z uploadowanymi URL-ami
                result, error = create_offer(
                    nazwa=tytul_seo,
                    opis=opis_html,
                    cena=cena,
                    zdjecia_urls=uploaded_urls if uploaded_urls else None,
                    kategoria_id=kategoria_id,
                    ilosc=ilosc,
                    ean=ean,
                    asin=asin,
                    gpsr=gpsr,
                    product_specs=product_specs,
                    bullet_points=bullet_points,
                    kod_magazynowy=_km
                )

                if error:
                    yield f"data: {json.dumps({'type': 'error', 'asin': asin, 'error': error[:80]})}\n\n"
                else:
                    # Dodaj info o liczbie zdjęć
                    yield f"data: {json.dumps({'type': 'success', 'asin': asin, 'title': tytul_seo[:50], 'ilosc': ilosc, 'zdjecia': len(uploaded_urls)})}\n\n"
                    
                    # Zapisz link oferta→produkt (oferta tworzona jako szkic — INACTIVE na Allegro)
                    _offer_id = result.get('id') if result else None
                    if _offer_id:
                        try:
                            if _produkt_id:
                                conn.execute('''INSERT OR REPLACE INTO oferty
                                    (tytul, opis, cena, ilosc, status, allegro_id, produkt_id, data_aktualizacji)
                                    VALUES (?, ?, ?, ?, 'draft', ?, ?, datetime('now'))''',
                                    (tytul_seo[:100], '', cena, ilosc, str(_offer_id), _produkt_id))
                            else:
                                conn.execute('''INSERT OR REPLACE INTO oferty
                                    (tytul, cena, ilosc, status, allegro_id, data_aktualizacji)
                                    VALUES (?, ?, ?, 'draft', ?, datetime('now'))''',
                                    (tytul_seo[:100], cena, ilosc, str(_offer_id)))
                        except:
                            pass

                    # Oznacz jako wystawione
                    max_db_retries = 5
                    for db_attempt in range(max_db_retries):
                        try:
                            conn.execute('UPDATE scraped SET status="wystawiony" WHERE asin=?', (asin,))
                            
                            # Commit co 10 produktów
                            if (i + 1) % 10 == 0:
                                conn.commit()
                                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>save</span> Zapisano {i+1}/{total} produktów'})}\n\n"
                                time.sleep(0.5)  # Opóźnienie przed ponownym otwarciem
                                conn = get_db()  # Nowe połączenie
                            
                            break  # Sukces - wyjdź z retry loop
                            
                        except sqlite3.OperationalError as e:
                            if 'database is locked' in str(e) and db_attempt < max_db_retries - 1:
                                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Baza zablokowana, retry {db_attempt+1}/{max_db_retries}...', 'color': '#f59e0b'})}\n\n"
                                time.sleep(1 * (db_attempt + 1))  # Exponential backoff: 1s, 2s, 3s, 4s, 5s
                                continue
                            else:
                                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> Błąd bazy dla {asin}: {str(e)[:40]}', 'color': '#ef4444'})}\n\n"
                                break  # Kontynuuj z następnym produktem
                    
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'asin': asin, 'error': str(e)[:80]})}\n\n"
            
            time.sleep(1)  # 1 sekunda między ofertami żeby nie przeciążyć API
        
        # Finalny commit (dla pozostałych produktów)
        try:
            conn.commit()
            yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>save</span> Zapisano wszystkie zmiany do bazy'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Błąd zapisu: {str(e)[:50]}', 'color': '#ef4444'})}\n\n"
        
        yield f"data: {json.dumps({'type': 'done'})}\n\n"
    
    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })

@paletomat_bp.route('/generator/cleanup')
def generator_cleanup():
    """Usuwa wystawione szkice z bazy i resetuje gotowe do regeneracji"""
    conn = get_db()
    wystawione = conn.execute('SELECT COUNT(*) as cnt FROM scraped WHERE status = "wystawiony"').fetchone()['cnt']
    conn.execute('DELETE FROM scraped WHERE status = "wystawiony"')
    # Resetuj gotowe — żeby przy ponownym wystawieniu regenerowały opis
    gotowe = conn.execute('SELECT COUNT(*) as cnt FROM scraped WHERE status = "gotowy"').fetchone()['cnt']
    conn.execute("UPDATE scraped SET opis_html = '', gpsr = '', tytul_seo = '', status = 'nowy' WHERE status = 'gotowy'")
    conn.commit()

    msg = f'Usunięto {wystawione} wystawionych szkiców!'
    if gotowe > 0:
        msg += f' Zresetowano {gotowe} gotowych do regeneracji opisów.'

    return render(f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>delete</span> CZYSZCZENIE</h1></div>
        <div class="alert alert-ok">{msg}</div>
        <a href="/paletomat/generator" class="btn btn-p">← Powrót do listy</a>
    ''')

@paletomat_bp.route('/generator/reprocess')
def generator_reprocess():
    """Ponownie przetwarza produkty z pustymi nazwami"""
    conn = get_db()
    
    # Znajdź produkty wymagające przetworzenia
    products = conn.execute('''SELECT asin FROM scraped 
        WHERE status IN ("nowy", "gotowy") 
        AND (nazwa IS NULL OR nazwa = '' OR nazwa LIKE 'Produkt %')''').fetchall()
    
    
    asins = [p['asin'] for p in products]
    
    if not asins:
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>sync</span> PRZETWARZANIE</h1></div>
            <div class="alert alert-warn">Brak produktów do przetworzenia!</div>
            <a href="/paletomat/generator" class="btn btn-p">← Powrót do listy</a>
        ''')
    
    # Uruchom auto-przetwarzanie
    auto_process_products(asins)
    
    return render(f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>sync</span> PRZETWARZANIE</h1></div>
        <div class="alert alert-ok">
            Uruchomiono przetwarzanie {len(asins)} produktów!<br><br>
            <small>Proces działa w tle. Odśwież stronę generatora za chwilę aby zobaczyć efekty.</small>
        </div>
        <a href="/paletomat/generator" class="btn btn-p">← Powrót do listy</a>
        <a href="/paletomat" class="btn btn-2"><span class=material-symbols-outlined>home</span> Strona główna (zobacz status)</a>
    ''')


@paletomat_bp.route('/generator/enhance-existing')
def generator_enhance_existing():
    """Strona z przyciskiem do generowania zdjęć AI dla istniejących produktów"""
    conn = get_db()

    # Policz produkty z/bez enhanced
    total = conn.execute("SELECT COUNT(*) as c FROM produkty WHERE status NOT IN ('usuniety','sprzedany') AND zdjecie_url IS NOT NULL AND zdjecie_url != ''").fetchone()['c']
    already = 0
    _enh_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced')
    if os.path.isdir(_enh_base):
        already = len([d for d in os.listdir(_enh_base) if os.path.isdir(os.path.join(_enh_base, d))])
    todo = max(0, total - already)

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>auto_awesome</span> GENERUJ ZDJĘCIA AI</h1><small>Dla produktów w magazynie</small></div>

    <div class="card" style="padding:20px">
        <div style="display:flex;gap:20px;justify-content:center;margin-bottom:20px">
            <div style="text-align:center"><div style="font-size:2rem;font-weight:700;color:#3b82f6">{total}</div><div style="font-size:0.8rem;color:#64748b">Produktów</div></div>
            <div style="text-align:center"><div style="font-size:2rem;font-weight:700;color:#22c55e">{already}</div><div style="font-size:0.8rem;color:#64748b">Ma zdjęcia AI</div></div>
            <div style="text-align:center"><div style="font-size:2rem;font-weight:700;color:#f59e0b">{todo}</div><div style="font-size:0.8rem;color:#64748b">Do wygenerowania</div></div>
        </div>
        <div style="font-size:0.8rem;color:#64748b;text-align:center;margin-bottom:15px">
            ~60 sek/produkt × {todo} = ~{todo} min
        </div>
        <button onclick="startEnhance(false)" id="btnStart" class="btn btn-ok" style="width:100%"><span class=material-symbols-outlined>auto_awesome</span> GENERUJ ZDJĘCIA AI ({todo} produktów)</button>
        <button onclick="if(confirm('Usunąć wszystkie wygenerowane zdjęcia i zacząć od nowa?'))startEnhance(true)" class="btn" style="width:100%;margin-top:8px;background:#ef4444;color:#fff;font-size:0.85rem"><span class=material-symbols-outlined>sync</span> WYCZYŚĆ I GENERUJ OD NOWA (wszystkie {total})</button>
        <div style="margin-top:12px;border-top:1px solid #334155;padding-top:12px">
            <div style="font-size:0.75rem;color:#94a3b8;margin-bottom:8px;text-align:center"><span class=material-symbols-outlined>dark_mode</span> Tryb nocny — generuj w tle (możesz zamknąć przeglądarkę)</div>
            <button onclick="startBgEnhance(false)" id="btnBg" class="btn" style="width:100%;background:#7c3aed;color:#fff"><span class=material-symbols-outlined>dark_mode</span> GENERUJ W TLE ({todo} produktów)</button>
            <button onclick="if(confirm('Usunąć wszystko i generować od nowa w tle?'))startBgEnhance(true)" class="btn" style="width:100%;margin-top:6px;background:#581c87;color:#fff;font-size:0.8rem"><span class=material-symbols-outlined>dark_mode</span><span class=material-symbols-outlined>sync</span> WYCZYŚĆ + GENERUJ W TLE (wszystkie {total})</button>
        </div>
    </div>

    <style>
    @keyframes spin {{ 0%{{transform:rotate(0deg)}} 100%{{transform:rotate(360deg)}} }}
    .spinner {{ display:inline-block;width:20px;height:20px;border:3px solid #334155;border-top:3px solid #f59e0b;border-radius:50%;animation:spin 0.8s linear infinite;vertical-align:middle;margin-right:8px }}
    @keyframes pulse {{ 0%,100%{{opacity:1}} 50%{{opacity:0.5}} }}
    .pulsing {{ animation:pulse 1.5s ease-in-out infinite }}
    </style>

    <div id="progressCard" class="card" style="display:none;padding:20px;text-align:center">
        <div style="margin-bottom:10px">
            <span id="spinnerEl" class="spinner"></span>
            <span id="progressText" style="font-size:1.1rem;font-weight:600">Startuje...</span>
        </div>
        <div style="background:#1e1e2e;border-radius:10px;height:16px;overflow:hidden;margin-bottom:10px">
            <div id="progressBar" style="background:linear-gradient(90deg,#f59e0b,#22c55e);height:100%;width:0%;transition:width 0.3s"></div>
        </div>
        <div style="display:flex;justify-content:space-between;font-size:0.8rem;color:#94a3b8">
            <div><span class=material-symbols-outlined>timer</span> <span id="elapsed">0:00</span></div>
            <div id="etaText"></div>
            <div><span class=material-symbols-outlined>check_circle</span> <span id="doneCount">0</span>/{todo}</div>
        </div>
    </div>

    <div id="log" class="card" style="display:none;max-height:400px;overflow-y:auto;font-family:monospace;font-size:0.7rem;padding:10px"></div>

    <script>
    let _startTime = null;
    let _timerInterval = null;
    let _doneN = 0;
    let _totalN = {todo};

    function fmtTime(s) {{
        const h = Math.floor(s/3600);
        const m = Math.floor((s%3600)/60);
        const sec = Math.floor(s%60);
        if (h > 0) return h+':'+String(m).padStart(2,'0')+':'+String(sec).padStart(2,'0');
        return m+':'+String(sec).padStart(2,'0');
    }}

    function updateTimer() {{
        if (!_startTime) return;
        const elapsed = (Date.now() - _startTime) / 1000;
        document.getElementById('elapsed').textContent = fmtTime(elapsed);
        if (_doneN > 0) {{
            const perItem = elapsed / _doneN;
            const remaining = perItem * (_totalN - _doneN);
            document.getElementById('etaText').textContent = ' Zostało ~' + fmtTime(remaining);
        }}
    }}

    function startEnhance(force) {{
        document.getElementById('btnStart').disabled = true;
        document.getElementById('btnStart').innerHTML = '<span class="spinner" style="width:14px;height:14px;border-width:2px;margin-right:6px"></span>Generuję...';
        document.getElementById('btnStart').classList.add('pulsing');
        document.getElementById('progressCard').style.display = 'block';
        document.getElementById('log').style.display = 'block';
        _startTime = Date.now();
        _timerInterval = setInterval(updateTimer, 1000);

        const url = '/paletomat/generator/enhance-existing-stream' + (force ? '?force=1' : '');
        const evtSource = new EventSource(url);
        evtSource.onmessage = function(e) {{
            const data = JSON.parse(e.data);
            if (data.type === 'progress') {{
                document.getElementById('progressText').textContent = data.message;
                document.getElementById('progressBar').style.width = data.percent + '%';
                if (data.current) {{
                    _doneN = data.current;
                    document.getElementById('doneCount').textContent = _doneN;
                }}
            }} else if (data.type === 'log') {{
                const div = document.createElement('div');
                div.style.padding = '2px 0';
                div.style.color = data.color || '#94a3b8';
                div.textContent = data.message;
                document.getElementById('log').appendChild(div);
                document.getElementById('log').scrollTop = document.getElementById('log').scrollHeight;
            }} else if (data.type === 'done') {{
                clearInterval(_timerInterval);
                const totalTime = fmtTime((Date.now() - _startTime) / 1000);
                document.getElementById('spinnerEl').style.display = 'none';
                document.getElementById('btnStart').textContent = ' Gotowe!';
                document.getElementById('btnStart').style.background = '#22c55e';
                document.getElementById('btnStart').classList.remove('pulsing');
                document.getElementById('progressBar').style.width = '100%';
                document.getElementById('progressText').textContent = (data.message || 'Zakończono!') + ' (' + totalTime + ')';
                document.getElementById('etaText').textContent = ' Czas: ' + totalTime;
                evtSource.close();
            }}
        }};
    }}

    function startBgEnhance(force) {{
        document.getElementById('btnBg').disabled = true;
        document.getElementById('btnBg').textContent = ' Uruchamiam...';
        fetch('/paletomat/generator/enhance-bg-start' + (force ? '?force=1' : ''))
            .then(r => r.json())
            .then(d => {{
                if (d.ok) {{
                    document.getElementById('btnBg').textContent = ' Działa w tle!';
                    document.getElementById('btnBg').style.background = '#22c55e';
                    document.getElementById('progressCard').style.display = 'block';
                    document.getElementById('log').style.display = 'block';
                    _startTime = Date.now();
                    _timerInterval = setInterval(updateTimer, 1000);
                    // Polluj status co 5 sek
                    _bgPoll = setInterval(pollBgStatus, 5000);
                    pollBgStatus();
                }} else {{
                    document.getElementById('btnBg').textContent = ' ' + (d.error || 'Błąd');
                    document.getElementById('btnBg').style.background = '#ef4444';
                    // Pokaż przycisk reset
                    if (!document.getElementById('btnReset')) {{
                        const resetBtn = document.createElement('button');
                        resetBtn.id = 'btnReset';
                        resetBtn.className = 'btn';
                        resetBtn.style.cssText = 'width:100%;margin-top:6px;background:#dc2626;color:#fff';
                        resetBtn.textContent = ' RESETUJ STATUS (odblokuj)';
                        resetBtn.onclick = function() {{
                            fetch('/paletomat/generator/enhance-bg-reset')
                                .then(r => r.json())
                                .then(d2 => {{ if(d2.ok) location.reload(); }});
                        }};
                        document.getElementById('btnBg').parentNode.appendChild(resetBtn);
                    }}
                }}
            }});
    }}

    let _bgPoll = null;
    function pollBgStatus() {{
        fetch('/paletomat/generator/enhance-bg-status')
            .then(r => r.json())
            .then(d => {{
                document.getElementById('progressText').textContent = d.current + '/' + d.total + ' produktów';
                document.getElementById('progressBar').style.width = d.progress + '%';
                _doneN = d.done;
                document.getElementById('doneCount').textContent = d.done;
                // Logi
                const logEl = document.getElementById('log');
                logEl.innerHTML = '';
                (d.log_last || []).forEach(msg => {{
                    const div = document.createElement('div');
                    div.style.padding = '2px 0';
                    div.style.color = '#94a3b8';
                    div.textContent = msg;
                    logEl.appendChild(div);
                }});
                logEl.scrollTop = logEl.scrollHeight;
                // Timer
                if (d.elapsed) {{
                    document.getElementById('elapsed').textContent = fmtTime(d.elapsed);
                    if (d.done > 0 && d.total > 0) {{
                        const perItem = d.elapsed / d.done;
                        const remaining = perItem * (d.total - d.done);
                        document.getElementById('etaText').textContent = ' Zostało ~' + fmtTime(remaining);
                    }}
                }}
                if (d.finished) {{
                    clearInterval(_bgPoll);
                    clearInterval(_timerInterval);
                    document.getElementById('spinnerEl').style.display = 'none';
                    document.getElementById('progressBar').style.width = '100%';
                    document.getElementById('progressText').textContent = ' Gotowe! ' + d.done + ' produktów | $' + d.cost.toFixed(2);
                }}
            }});
    }}
    </script>

    <a href="/paletomat/generator/enhance-gallery" class="back" style="display:block;text-align:center;margin-top:15px"><span class=material-symbols-outlined>image</span> Przeglądaj wygenerowane zdjęcia</a>
    <a href="/paletomat/generator" class="back" style="display:block;text-align:center;margin-top:10px">← Powrót</a>
    '''
    return render(html)


@paletomat_bp.route('/generator/enhance-gallery')
def generator_enhance_gallery():
    """Galeria wygenerowanych zdjęć AI — przeglądanie po produktach"""
    conn = get_db()
    _enh_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced')

    products = []
    if os.path.isdir(_enh_base):
        for asin_dir in sorted(os.listdir(_enh_base), reverse=True):
            full_dir = os.path.join(_enh_base, asin_dir)
            if not os.path.isdir(full_dir):
                continue
            # Filtruj — tylko pliki > 1KB (odrzuć broken/puste)
            _all_imgs = [f for f in os.listdir(full_dir)
                        if (f.endswith('.jpg') or f.endswith('.png'))
                        and os.path.getsize(os.path.join(full_dir, f)) > 1024]
            # Sortuj w logicznej kolejności (nie alfabetycznie!)
            _img_order = ['mini', 'det', 'zest', 'kat2', 'wym', 'uzycie', 'life']
            imgs = sorted(_all_imgs, key=lambda f: (
                _img_order.index(f.rsplit('.', 1)[0]) if f.rsplit('.', 1)[0] in _img_order else 99
            ))
            if not imgs:
                continue

            # Nazwa produktu z DB
            row = conn.execute("SELECT nazwa FROM produkty WHERE asin = ? LIMIT 1", (asin_dir,)).fetchone()
            nazwa = row['nazwa'][:50] if row else asin_dir

            products.append({
                'asin': asin_dir,
                'nazwa': nazwa,
                'imgs': imgs,
                'count': len(imgs),
            })

    # Paginacja — 10 na stronę
    page = request.args.get('page', 1, type=int)
    per_page = 10
    total_pages = max(1, (len(products) + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    page_products = products[(page-1)*per_page : page*per_page]

    # Template nazwy plików
    tpl_labels = {'mini': '<span class=material-symbols-outlined>photo_camera</span> Miniaturka', 'wym': '<span class=material-symbols-outlined>smart_toy</span> Wymiary', 'det': '<span class=material-symbols-outlined>photo_camera</span> Detale', 'zest': '<span class=material-symbols-outlined>photo_camera</span> Zestaw',
                  'kat2': '<span class=material-symbols-outlined>photo_camera</span> Drugi kąt', 'uzycie': '<span class=material-symbols-outlined>smart_toy</span> W użyciu', 'life': '<span class=material-symbols-outlined>smart_toy</span> Lifestyle'}

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>image</span> GALERIA ZDJĘĆ AI</h1><small>{len(products)} produktów z wygenerowanymi zdjęciami</small></div>

    <div style="text-align:center;margin-bottom:15px">
        <span style="color:#94a3b8">Strona {page}/{total_pages}</span>
        {"" if page <= 1 else f' <a href="?page={page-1}" class="btn" style="font-size:0.8rem;padding:4px 12px">← Poprzednia</a>'}
        {"" if page >= total_pages else f' <a href="?page={page+1}" class="btn" style="font-size:0.8rem;padding:4px 12px">Następna →</a>'}
    </div>
    '''

    for p in page_products:
        html += f'''
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="font-weight:600;margin-bottom:10px;font-size:0.9rem">
            <span style="color:#f59e0b">{p["asin"]}</span> — {p["nazwa"]}
            <span style="color:#64748b;font-size:0.75rem">({p["count"]} zdjęć)</span>
        </div>
        <div style="display:flex;gap:8px;overflow-x:auto;padding-bottom:8px">'''

        import hashlib as _hl
        for img_name in p['imgs']:
            base = img_name.rsplit('.', 1)[0]
            label = tpl_labels.get(base, base)
            # Cache-bust: użyj rozmiaru pliku jako query param
            _fpath = os.path.join(_enh_base, p["asin"], img_name)
            _cb = str(os.path.getsize(_fpath))[:6] if os.path.exists(_fpath) else '0'
            html += f'''
            <div style="flex-shrink:0;text-align:center">
                <a href="/static/enhanced/{p["asin"]}/{img_name}?v={_cb}" target="_blank">
                    <img src="/paletomat/generator/thumb/{p["asin"]}/{img_name}?v={_cb}" style="width:120px;height:120px;object-fit:cover;border-radius:8px;border:1px solid #334155" loading="lazy" onerror="this.parentElement.parentElement.style.display='none'">
                </a>
                <div style="font-size:0.65rem;color:#94a3b8;margin-top:3px">{label}</div>
            </div>'''

        html += '''
        </div>
    </div>'''

    html += f'''
    <div style="text-align:center;margin:15px 0">
        {"" if page <= 1 else f'<a href="?page={page-1}" class="btn" style="font-size:0.8rem;padding:4px 12px">← Poprzednia</a>'}
        {"" if page >= total_pages else f'<a href="?page={page+1}" class="btn" style="font-size:0.8rem;padding:4px 12px">Następna →</a>'}
    </div>
    <a href="/paletomat/generator/enhance-existing" class="back" style="display:block;text-align:center">← Generuj zdjęcia</a>
    <a href="/paletomat/generator" class="back" style="display:block;text-align:center;margin-top:8px">← Generator</a>
    '''
    return render(html)


@paletomat_bp.route('/generator/thumb/<asin>/<filename>')
def generator_thumb(asin, filename):
    """Serwuje miniaturki zdjęć enhanced — 200px zamiast pełnych 2560px"""
    from flask import send_file
    import re
    # Sanitize inputs against path traversal
    if not re.match(r'^[A-Za-z0-9_-]+$', asin) or '..' in filename or '/' in filename or '\\' in filename:
        from flask import abort
        abort(400)
    _enh_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced')
    full_path = os.path.join(_enh_base, asin, filename)
    # Verify resolved path stays within _enh_base
    if not os.path.realpath(full_path).startswith(os.path.realpath(_enh_base)):
        from flask import abort
        abort(403)
    if not os.path.exists(full_path) or os.path.getsize(full_path) < 1024:
        from flask import abort
        abort(404)

    # Generuj miniaturkę w pamięci
    try:
        from PIL import Image as _ThI
        from io import BytesIO as _ThB
        img = _ThI.open(full_path)
        img.thumbnail((200, 200), _ThI.LANCZOS)
        buf = _ThB()
        img.save(buf, 'JPEG', quality=75)
        buf.seek(0)
        return send_file(buf, mimetype='image/jpeg', max_age=60)
    except Exception:
        return send_file(full_path, mimetype='image/jpeg')


@paletomat_bp.route('/generator/enhance-existing-stream')
def generator_enhance_existing_stream():
    """SSE stream — generuje zdjęcia AI dla produktów w magazynie które ich nie mają"""
    import time

    force = request.args.get('force', '0') == '1'

    def generate():
        conn = get_db()

        # Pobierz produkty które mają zdjęcie
        products = conn.execute('''
            SELECT id, asin, ean, nazwa, zdjecie_url, images
            FROM produkty
            WHERE status NOT IN ('usuniety', 'sprzedany')
            AND zdjecie_url IS NOT NULL AND zdjecie_url != ''
            ORDER BY data_dodania DESC
        ''').fetchall()

        _enh_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced')

        # Force mode — usuń stare foldery
        if force and os.path.isdir(_enh_base):
            import shutil
            _old_count = len([d for d in os.listdir(_enh_base) if os.path.isdir(os.path.join(_enh_base, d))])
            shutil.rmtree(_enh_base)
            os.makedirs(_enh_base, exist_ok=True)
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>delete</span> Usunięto {_old_count} starych folderów — generuję od nowa', 'color': '#ef4444'})}\n\n"

        # Filtruj — tylko te bez enhanced (lub wszystkie jeśli force)
        todo = []
        for p in products:
            p = dict(p)
            _key = p.get('asin') or str(p['id'])
            _dir = os.path.join(_enh_base, str(_key))
            if not os.path.isdir(_dir) or len([f for f in os.listdir(_dir) if f.endswith('.jpg')]) < 4:
                todo.append(p)

        if not todo:
            yield f"data: {json.dumps({'type': 'done', 'message': 'Wszystkie produkty mają już zdjęcia AI!'})}\n\n"
            return

        yield f"data: {json.dumps({'type': 'progress', 'percent': 0, 'message': f'0/{len(todo)} produktów...'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': f'Znaleziono {len(todo)} produktów bez zdjęć AI', 'color': '#f59e0b'})}\n\n"

        try:
            from .image_enhancer import enhance_single, prepare_original_photo, GEMINI_AVAILABLE
            from .image_enhancer import HYBRID_ORIGINAL_SLOTS, HYBRID_AI_TEMPLATES
            from .image_cleaner import clean_image_from_url, clean_image_from_bytes
        except Exception as _ie:
            yield f"data: {json.dumps({'type': 'log', 'message': f'Błąd importu: {_ie}', 'color': '#ef4444'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'message': 'Błąd!'})}\n\n"
            return

        if not GEMINI_AVAILABLE:
            yield f"data: {json.dumps({'type': 'done', 'message': 'Gemini API niedostępne!'})}\n\n"
            return

        yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>photo_camera</span> Tryb HYBRID: oryginały (1-4) + AI (5-8)', 'color': '#8b5cf6'})}\n\n"

        ok_count = 0
        _total_cost = 0.0
        for i, p in enumerate(todo):
            _key = p.get('asin') or str(p['id'])
            nazwa = p.get('nazwa', '')[:60]
            pct = int((i / len(todo)) * 100)

            yield f"data: {json.dumps({'type': 'progress', 'percent': pct, 'current': i, 'message': f'{i+1}/{len(todo)}: {nazwa[:30]}...'})}\n\n"
            yield f"data: {json.dumps({'type': 'log', 'message': f'[{i+1}/{len(todo)}] {_key}: {nazwa[:40]}', 'color': '#3b82f6'})}\n\n"

            # Znajdź WSZYSTKIE zdjęcia (nie tylko pierwsze!)
            _all_images = []
            try:
                # Spróbuj z static/downloads/{asin}/ — tu są lokalne pliki
                _dl_dir = os.path.join('static', 'downloads', str(_key))
                if os.path.isdir(_dl_dir):
                    _all_images = sorted([os.path.join(_dl_dir, f) for f in os.listdir(_dl_dir) if f.endswith('.jpg')])

                # Spróbuj z images JSON
                if not _all_images and p.get('images'):
                    try:
                        _imgs_json = json.loads(p['images']) if isinstance(p['images'], str) else p['images']
                        _all_images = [u for u in _imgs_json if isinstance(u, str)]
                    except:
                        pass

                # Fallback na zdjecie_url
                if not _all_images and p.get('zdjecie_url'):
                    _all_images = [p['zdjecie_url']]

                if not _all_images:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>warning</span> Brak zdjęcia', 'color': '#f59e0b'})}\n\n"
                    continue

            except Exception as _ex:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> Błąd pobierania: {str(_ex)[:40]}', 'color': '#ef4444'})}\n\n"
                continue

            # --- KROK 1: Załaduj zdjęcia i wyczyść oryginały ---
            _enh_dir = os.path.join(_enh_base, str(_key))
            os.makedirs(_enh_dir, exist_ok=True)
            _ok = 0
            _orig_count = 0
            _slot_to_template = {'mini': 1, 'det': 3, 'zest': 4, 'kat2': 5}

            # Załaduj wszystkie zdjęcia jako bytes
            _loaded_images = []
            for _src in _all_images:
                try:
                    if isinstance(_src, str) and _src.startswith('http'):
                        import requests as _rq
                        _raw = _rq.get(_src, timeout=20).content
                    elif isinstance(_src, str) and os.path.exists(_src):
                        with open(_src, 'rb') as _f:
                            _raw = _f.read()
                    else:
                        _abs = os.path.abspath(_src)
                        if os.path.exists(_abs):
                            with open(_abs, 'rb') as _f:
                                _raw = _f.read()
                        else:
                            continue

                    # Walidacja
                    from PIL import Image as _PIv
                    from io import BytesIO as _BIv
                    _PIv.open(_BIv(_raw)).verify()
                    _loaded_images.append(_raw)
                except Exception:
                    pass

            if not _loaded_images:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>warning</span> Brak prawidłowych zdjęć', 'color': '#f59e0b'})}\n\n"
                continue

            # Wyczyść i zapisz oryginały na sloty 1-4 (cleaner usunie tekst/loga/infografiki)
            for _oi, _slot_name in enumerate(HYBRID_ORIGINAL_SLOTS):
                if _oi >= len(_loaded_images):
                    break
                try:
                    _cb, _cm, _ce = clean_image_from_bytes(_loaded_images[_oi])
                    _clean = _cb if _cb else _loaded_images[_oi]
                    _prep, _perr = prepare_original_photo(_clean)
                    if _prep:
                        _epath = os.path.join(_enh_dir, f'{_slot_name}.jpg')
                        with open(_epath, 'wb') as _sf:
                            _sf.write(_prep)
                        _ok += 1
                        _orig_count += 1
                except Exception:
                    pass

            if _orig_count > 0:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>photo_camera</span> {_orig_count} oryginałów wyczyszczonych', 'color': '#7c3aed'})}\n\n"

            # Brakujące sloty oryginałów → AI
            _missing_orig = HYBRID_ORIGINAL_SLOTS[_orig_count:]

            # Baza do AI = wyczyszczony oryginał #1
            try:
                _cb0, _, _ = clean_image_from_bytes(_loaded_images[0])
                _base_for_ai = _cb0 if _cb0 else _loaded_images[0]

                from PIL import Image as _PIv2
                from io import BytesIO as _BIv2
                _PIv2.open(_BIv2(_base_for_ai)).verify()
            except Exception:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>warning</span> Nieprawidłowy obrazek bazowy — pomijam', 'color': '#f59e0b'})}\n\n"
                continue

            # AI dla brakujących oryginałów
            for _ms in _missing_orig:
                _tid = _slot_to_template.get(_ms, 1)
                try:
                    _ed, _em, _ee = enhance_single(_base_for_ai, _tid, nazwa)
                    if _ed:
                        from PIL import Image as _PI
                        from io import BytesIO as _BI
                        _img = _PI.open(_BI(_ed)).convert('RGB')
                        if max(_img.width, _img.height) < 2560:
                            _ratio = 2560 / max(_img.width, _img.height)
                            _img = _img.resize((int(_img.width * _ratio), int(_img.height * _ratio)), _PI.LANCZOS)
                        _img.save(os.path.join(_enh_dir, f'{_ms}.jpg'), 'JPEG', quality=95)
                        _ok += 1
                except Exception:
                    pass
                time.sleep(0.5)

            # --- KROK 2: AI sloty 5-8 ---
            for _tid, _tname in HYBRID_AI_TEMPLATES:
                try:
                    _ed, _em, _ee = enhance_single(_base_for_ai, _tid, nazwa)
                    if _ed:
                        from PIL import Image as _PI
                        from io import BytesIO as _BI
                        _img = _PI.open(_BI(_ed)).convert('RGB')
                        if max(_img.width, _img.height) < 2560:
                            _ratio = 2560 / max(_img.width, _img.height)
                            _img = _img.resize((int(_img.width * _ratio), int(_img.height * _ratio)), _PI.LANCZOS)
                        _img.save(os.path.join(_enh_dir, f'{_tname}.jpg'), 'JPEG', quality=95)
                        _ok += 1
                    else:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> {_tname}: {str(_ee)[:30]}', 'color': '#ef4444'})}\n\n"
                except Exception as _te:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> {_tname}: {str(_te)[:30]}', 'color': '#ef4444'})}\n\n"
                time.sleep(0.5)

            _ai_count = _ok - _orig_count
            if _ok > 0:
                ok_count += 1
                _prod_cost = _ai_count * 0.001 + (_orig_count * 0.001 if _orig_count > 0 else 0)  # AI + clean
                _total_cost += _prod_cost
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>check_circle</span> {_ok}/8 zdjęć ({_orig_count} oryg + {_ai_count} AI) ${_prod_cost:.3f}', 'color': '#22c55e'})}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>warning</span> Żadne zdjęcie nie wygenerowane', 'color': '#f59e0b'})}\n\n"

            time.sleep(1)  # Cooldown między produktami

        # Loguj łączny koszt do monitor_stats
        try:
            conn.execute('''INSERT INTO monitor_stats (event_type, model, prompt_tokens, completion_tokens, cost_usd, extra, timestamp)
                VALUES ('gemini', 'enhance_batch', 0, 0, ?, ?, datetime('now'))''',
                (_total_cost, json.dumps({'products': ok_count, 'images': ok_count * 8})))
            conn.commit()
        except Exception:
            pass

        yield f"data: {json.dumps({'type': 'done', 'message': f'Gotowe! {ok_count}/{len(todo)} produktów | Koszt: ${_total_cost:.2f}'})}\n\n"

    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })


# === BACKGROUND ENHANCE TASK ===
# Generuje zdjęcia w tle na Pi — nie wymaga otwartej przeglądarki
_bg_enhance_status = {
    'running': False,
    'progress': 0,
    'current': 0,
    'total': 0,
    'done': 0,
    'errors': 0,
    'cost': 0.0,
    'log': [],
    'finished': False,
    'started_at': None,
}

def _bg_enhance_worker(app, force=False):
    """Background thread do generowania zdjęć — działa nawet po zamknięciu przeglądarki"""
    global _bg_enhance_status
    import time as _t

    with app.app_context():
        conn = get_db()
        _bg_enhance_status['log'] = []
        _bg_enhance_status['finished'] = False
        _bg_enhance_status['started_at'] = _t.time()

        def _log(msg):
            _bg_enhance_status['log'].append(msg)
            if len(_bg_enhance_status['log']) > 200:
                _bg_enhance_status['log'] = _bg_enhance_status['log'][-100:]
            print(f"[BG-ENHANCE] {msg}")

        try:
            from .image_enhancer import prepare_original_photo
            from .image_enhancer import HYBRID_ORIGINAL_SLOTS
            from .image_cleaner import clean_image_from_bytes, REMBG_AVAILABLE
        except Exception as _ie:
            _log(f"<span class=material-symbols-outlined>cancel</span> Import error: {_ie}")
            _bg_enhance_status['running'] = False
            _bg_enhance_status['finished'] = True
            return

        _vps_url = ''
        try:
            from .database import get_config as _gc
            _vps_url = _gc('rembg_vps_url', '')
        except Exception:
            pass

        if not REMBG_AVAILABLE and not _vps_url:
            _log("<span class=material-symbols-outlined>cancel</span> rembg niedostępny i VPS nie skonfigurowany!")
            _bg_enhance_status['running'] = False
            _bg_enhance_status['finished'] = True
            return

        if _vps_url:
            _log(f"<span class=material-symbols-outlined>language</span> Tryb VPS: {_vps_url}")
        else:
            _log("<span class=material-symbols-outlined>computer</span> Tryb lokalny rembg (uwaga: grzeje Pi)")

        products = conn.execute('''
            SELECT id, asin, ean, nazwa, zdjecie_url, images
            FROM produkty
            WHERE status NOT IN ('usuniety', 'sprzedany')
            AND zdjecie_url IS NOT NULL AND zdjecie_url != ''
            ORDER BY data_dodania DESC
        ''').fetchall()

        _enh_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced')

        if force and os.path.isdir(_enh_base):
            import shutil
            _old = len([d for d in os.listdir(_enh_base) if os.path.isdir(os.path.join(_enh_base, d))])
            shutil.rmtree(_enh_base)
            os.makedirs(_enh_base, exist_ok=True)
            _log(f"<span class=material-symbols-outlined>delete</span> Usunięto {_old} starych folderów")

        todo = []
        for p in products:
            p = dict(p)
            _key = p.get('asin') or str(p['id'])
            _dir = os.path.join(_enh_base, str(_key))
            if not os.path.isdir(_dir) or len([f for f in os.listdir(_dir) if f.endswith('.jpg')]) < 1:
                todo.append(p)

        _bg_enhance_status['total'] = len(todo)

        if not todo:
            _log("<span class=material-symbols-outlined>check_circle</span> Wszystkie produkty mają już wyczyszczone zdjęcia!")
            _bg_enhance_status['running'] = False
            _bg_enhance_status['finished'] = True
            return

        _log(f"<span class=material-symbols-outlined>photo_camera</span> Start: {len(todo)} produktów do przetworzenia")

        ok_count = 0

        for i, p in enumerate(todo):
            if not _bg_enhance_status['running']:
                _log("⏹ Zatrzymano ręcznie")
                break

            _key = p.get('asin') or str(p['id'])
            nazwa = p.get('nazwa', '')[:60]
            _bg_enhance_status['current'] = i + 1
            _bg_enhance_status['progress'] = int(((i + 1) / len(todo)) * 100)
            _bg_enhance_status['last_update'] = _t.time()

            _log(f"[{i+1}/{len(todo)}] {_key}: {nazwa[:40]}")

            # Znajdź zdjęcia
            _all_images = []
            _dl_dir = os.path.join('static', 'downloads', str(_key))
            if os.path.isdir(_dl_dir):
                _all_images = sorted([os.path.join(_dl_dir, f) for f in os.listdir(_dl_dir) if f.endswith('.jpg')])
            if not _all_images and p.get('images'):
                try:
                    _imgs = json.loads(p['images']) if isinstance(p['images'], str) else p['images']
                    _all_images = [u for u in _imgs if isinstance(u, str)]
                except:
                    pass
            if not _all_images and p.get('zdjecie_url'):
                _all_images = [p['zdjecie_url']]
            if not _all_images:
                continue

            # Załaduj bytes
            _loaded = []
            for _src in _all_images:
                try:
                    if isinstance(_src, str) and _src.startswith('http'):
                        import requests as _rq
                        _raw = _rq.get(_src, timeout=20).content
                    elif os.path.exists(_src):
                        with open(_src, 'rb') as _f:
                            _raw = _f.read()
                    elif os.path.exists(os.path.abspath(_src)):
                        with open(os.path.abspath(_src), 'rb') as _f:
                            _raw = _f.read()
                    else:
                        continue
                    from PIL import Image as _PIv
                    from io import BytesIO as _BIv
                    _PIv.open(_BIv(_raw)).verify()
                    _loaded.append(_raw)
                except:
                    pass

            if not _loaded:
                continue

            _enh_dir = os.path.join(_enh_base, str(_key))
            os.makedirs(_enh_dir, exist_ok=True)
            _ok = 0

            # Czyść tło rembg na wszystkich oryginalnych zdjęciach (max 4 sloty)
            _slots = HYBRID_ORIGINAL_SLOTS  # ['mini', 'det', 'zest', 'kat2']
            for _oi, _slot in enumerate(_slots):
                if _oi >= len(_loaded):
                    break
                try:
                    _cb, _, _err = clean_image_from_bytes(_loaded[_oi])
                    _clean = _cb if _cb else _loaded[_oi]
                    _prep, _ = prepare_original_photo(_clean)
                    if _prep:
                        with open(os.path.join(_enh_dir, f'{_slot}.jpg'), 'wb') as _sf:
                            _sf.write(_prep)
                        _ok += 1
                except:
                    pass
                if not _vps_url:
                    _t.sleep(2)  # Cooldown rembg lokalne — anti-overheat na Pi

            if _ok > 0:
                ok_count += 1
                _bg_enhance_status['done'] = ok_count
                _log(f"   <span class=material-symbols-outlined>check_circle</span> {_ok} zdjęć (rembg, darmowe)")
            else:
                _bg_enhance_status['errors'] += 1

            if not _vps_url:
                _t.sleep(5)  # Cooldown między produktami (tylko lokalne)
            else:
                _t.sleep(1)  # Minimalny cooldown VPS

        # Loguj statystyki (rembg = darmowe, koszt $0)
        try:
            conn.execute('''INSERT INTO monitor_stats (event_type, model, prompt_tokens, completion_tokens, cost_usd, extra, timestamp)
                VALUES ('rembg', 'enhance_bg_batch', 0, 0, 0, ?, datetime('now'))''',
                (json.dumps({'products': ok_count, 'total': len(todo)}),))
            conn.commit()
        except:
            pass

        _log(f"<span class=material-symbols-outlined>flag</span> GOTOWE! {ok_count}/{len(todo)} produktów (rembg, darmowe)")
        _bg_enhance_status['running'] = False
        _bg_enhance_status['finished'] = True


@paletomat_bp.route('/generator/enhance-bg-reset')
def enhance_bg_reset():
    """Wymuszony reset statusu — gdyby się zawiesił"""
    global _bg_enhance_status
    _bg_enhance_status = {
        'running': False, 'progress': 0, 'current': 0, 'total': 0,
        'done': 0, 'errors': 0, 'cost': 0.0, 'log': [], 'finished': False,
        'started_at': None
    }
    return jsonify({'ok': True, 'message': 'Status zresetowany!'})


@paletomat_bp.route('/generator/enhance-bg-start')
def enhance_bg_start():
    """Startuje generowanie w tle — można zamknąć przeglądarkę"""
    global _bg_enhance_status
    import threading
    import time as _t

    # Auto-reset: jeśli "running" ale ostatnia aktywność >5 min temu = zombie
    if _bg_enhance_status['running']:
        started = _bg_enhance_status.get('started_at') or 0
        last_update = _bg_enhance_status.get('last_update') or started
        if _t.time() - max(started, last_update) > 300:  # 5 min bez aktywności
            print("[BG-Enhance] Auto-reset — worker zombie (>5 min bez aktywności)")
            _bg_enhance_status['running'] = False

    if _bg_enhance_status['running']:
        return jsonify({'ok': False, 'error': 'Już działa!'})

    force = request.args.get('force', '0') == '1'
    _bg_enhance_status = {
        'running': True, 'progress': 0, 'current': 0, 'total': 0,
        'done': 0, 'errors': 0, 'cost': 0.0, 'log': [], 'finished': False,
        'started_at': None
    }

    from flask import current_app
    app = current_app._get_current_object()
    t = threading.Thread(target=_bg_enhance_worker, args=(app, force), daemon=True)
    t.start()

    return jsonify({'ok': True, 'message': 'Generowanie uruchomione w tle!'})


@paletomat_bp.route('/generator/enhance-bg-stop')
def enhance_bg_stop():
    """Zatrzymuje generowanie w tle"""
    global _bg_enhance_status
    _bg_enhance_status['running'] = False
    return jsonify({'ok': True, 'message': 'Zatrzymywanie...'})


@paletomat_bp.route('/generator/enhance-bg-status')
def enhance_bg_status():
    """Status generowania w tle — JSON"""
    import time as _t
    elapsed = 0
    if _bg_enhance_status.get('started_at'):
        elapsed = int(_t.time() - _bg_enhance_status['started_at'])
    return jsonify({
        **_bg_enhance_status,
        'elapsed': elapsed,
        'log_last': _bg_enhance_status['log'][-20:] if _bg_enhance_status['log'] else [],
    })


@paletomat_bp.route('/generator/from-magazyn/<int:product_id>')
def generator_from_magazyn(product_id):
    """Wystawianie produktu bezpośrednio z magazynu - bez potrzeby scrapowania"""
    conn = get_db()

    # Pobierz produkt z magazynu
    produkt = conn.execute('SELECT * FROM produkty WHERE id=?', (product_id,)).fetchone()

    if not produkt:
        return redirect('/magazyn')

    p = dict(produkt)

    # Ustal identyfikator (ASIN lub EAN lub ID)
    asin = p.get('asin', '') or ''
    ean = p.get('ean', '') or ''

    if asin and asin not in ('N/A', 'None', ''):
        identyfikator = asin
    elif ean and ean not in ('N/A', 'None', ''):
        identyfikator = ean
    else:
        identyfikator = f"MAG{product_id}"

    # === SPRAWDŹ CZY PRODUKT JUŻ MA AKTYWNĄ OFERTĘ NA ALLEGRO ===
    force_new = request.args.get('force_new', '0') == '1'
    if not force_new:
        existing_offer = None

        # 1. Szukaj po ASIN (najdokładniejsze)
        if asin and asin not in ('N/A', 'None', '') and not existing_offer:
            all_prod_ids = [r['id'] for r in conn.execute('SELECT id FROM produkty WHERE asin = ?', (asin,)).fetchall()]
            if all_prod_ids:
                ph = ','.join('?' * len(all_prod_ids))
                existing_offer = conn.execute(
                    "SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status, o.cena "
                    "FROM oferty o "
                    "WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published') "
                    "AND o.allegro_id IS NOT NULL AND o.allegro_id != '' "
                    "AND o.produkt_id IN (" + ph + ") "
                    "LIMIT 1",
                    all_prod_ids).fetchone()
                if existing_offer:
                    print(f"[DEDUP-WYSTAW] Match po ASIN {asin}: oferta={existing_offer['allegro_id']}")

        # 2. Szukaj po EAN (TYLKO jeśli produkt NIE MA ASIN — EAN jest mniej wiarygodny)
        if ean and ean not in ('N/A', 'None', '') and not existing_offer and not (asin and asin not in ('N/A', 'None', '')):
            all_prod_ids = [r['id'] for r in conn.execute('SELECT id FROM produkty WHERE ean = ?', (ean,)).fetchall()]
            if all_prod_ids:
                ph = ','.join('?' * len(all_prod_ids))
                existing_offer = conn.execute(
                    "SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status, o.cena "
                    "FROM oferty o "
                    "WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published') "
                    "AND o.allegro_id IS NOT NULL AND o.allegro_id != '' "
                    "AND o.produkt_id IN (" + ph + ") "
                    "LIMIT 1",
                    all_prod_ids).fetchone()
                if existing_offer:
                    print(f"[DEDUP-WYSTAW] Match po EAN {ean}: oferta={existing_offer['allegro_id']}")

        # 3. Szukaj po produkt_id + weryfikacja nazwy
        if not existing_offer:
            _pid_candidate = conn.execute('''
                SELECT o.id, o.allegro_id, o.tytul, o.ilosc, o.status, o.cena
                FROM oferty o
                WHERE o.status IN ('active','ACTIVE','aktywna','wystawiona','published')
                AND o.allegro_id IS NOT NULL AND o.allegro_id != ''
                AND o.produkt_id = ?
                LIMIT 1
            ''', (product_id,)).fetchone()
            if _pid_candidate:
                # Weryfikuj nazwę — żeby nie łączyć różnych produktów
                _nazwa_lower = (p.get('nazwa', '') or '').lower()[:40]
                _ignore = {'uniwersalne', 'uniwersalny', 'premium', 'zestaw', 'komplet'}
                _words_p = [w for w in _nazwa_lower.split() if len(w) > 2 and w not in _ignore][:5]
                _words_o = [w for w in (_pid_candidate['tytul'] or '').lower()[:60].split() if len(w) > 2 and w not in _ignore][:5]
                _matching = sum(1 for w in _words_p if any(w in wo or wo in w for wo in _words_o)) if _words_p and _words_o else 0
                _threshold = max(2, len(_words_p) // 2 + 1) if _words_p else 2
                if _matching >= _threshold or not _words_p:
                    existing_offer = _pid_candidate
                    print(f"[DEDUP-WYSTAW] Match po produkt_id={product_id}: oferta={existing_offer['allegro_id']}")
                else:
                    print(f"[DEDUP-WYSTAW] produkt_id={product_id} ODRZUCONY - nazwa nie pasuje: '{_pid_candidate['tytul'][:40]}' vs '{_nazwa_lower}'")
                    # Odlinkuj błędne powiązanie
                    conn.execute('UPDATE oferty SET produkt_id = NULL WHERE id = ?', (_pid_candidate['id'],))
                    conn.commit()

        # 4. FALLBACK: Sprawdź bezpośrednio na Allegro API (łapie oferty z Sales Center)
        if not existing_offer:
            from .allegro_api import find_active_offer_on_allegro
            api_match = find_active_offer_on_allegro(asin=asin, ean=ean, nazwa=p.get('nazwa', ''))
            if api_match:
                existing_offer = {
                    'id': None,
                    'allegro_id': api_match['allegro_id'],
                    'tytul': api_match['tytul'],
                    'ilosc': api_match['ilosc'],
                    'status': 'aktywna',
                    'cena': api_match['cena']
                }
                print(f"[DEDUP-WYSTAW] Match z Allegro API: {api_match['allegro_id']}")

        if existing_offer:
            o = dict(existing_offer)
            # Automatycznie dodaj sztuki z magazynu do istniejącej oferty
            from .allegro_api import update_offer_stock
            magazyn_ilosc = int(p.get('ilosc', 0) or 0)
            obecna_ilosc = o.get('ilosc', 0) or 0
            dodaj = max(min(magazyn_ilosc, magazyn_ilosc), 1)  # nie więcej niż stan magazynowy
            new_qty = obecna_ilosc + dodaj

            result, error = update_offer_stock(o['allegro_id'], new_qty)

            if error:
                # Fallback: pokaż stronę z formularzem ręcznym
                return render_template_string(TEMPLATE_ALREADY_LISTED,
                                              produkt=p, oferta=o, product_id=product_id,
                                              identyfikator=identyfikator,
                                              error_msg=f"Blad auto-dodawania: {error}")

            return render_template_string(TEMPLATE_STOCK_UPDATED,
                                          produkt=p, oferta=o, product_id=product_id,
                                          dodaj=dodaj, new_qty=new_qty)

    # Sprawdz czy juz istnieje w scraped
    existing = conn.execute('SELECT asin FROM scraped WHERE asin=?', (identyfikator,)).fetchone()

    if not existing:
        try:
            conn.execute('''
                INSERT INTO scraped (asin, nazwa, zdjecie_url, cena_amazon, status, data_scrape, ean)
                VALUES (?, ?, ?, ?, 'nowy', datetime('now'), ?)
            ''', (
                identyfikator,
                p.get('nazwa', f'Produkt {identyfikator}'),
                p.get('zdjecie_url', ''),
                p.get('cena_brutto', 0) or 0,
                ean
            ))
        except:
            conn.execute('''
                INSERT INTO scraped (asin, nazwa, zdjecie_url, cena_amazon, status, data_scrape)
                VALUES (?, ?, ?, ?, 'nowy', datetime('now'))
            ''', (
                identyfikator,
                p.get('nazwa', f'Produkt {identyfikator}'),
                p.get('zdjecie_url', ''),
                p.get('cena_brutto', 0) or 0
            ))
        conn.commit()
    else:
        conn.execute('''
            UPDATE scraped SET nazwa=?, zdjecie_url=?, status='nowy'
            WHERE asin=?
        ''', (
            p.get('nazwa', f'Produkt {identyfikator}'),
            p.get('zdjecie_url', ''),
            identyfikator
        ))
        conn.commit()

    # Przekieruj do generatora
    return redirect(f'/paletomat/generator/{identyfikator}?produkt_id={product_id}')


# === TEMPLATE: Produkt już wystawiony ===
# === TEMPLATE: Automatycznie dodano sztuki ===
TEMPLATE_STOCK_UPDATED = '''<!DOCTYPE html><html><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dodano sztuki</title>
<link rel="stylesheet" href="/static/styles.css">
<meta http-equiv="refresh" content="3;url=/magazyn/produkt/{{ produkt.get('kod_magazynowy','') or product_id }}">
</head><body class="{{ 'kiosk' if request.cookies.get('kiosk') else '' }}">
<div class="container" style="max-width:600px;margin:auto;padding:20px">
  <div class="card" style="padding:20px;border-left:4px solid var(--green,#22c55e)">
    <h2 style="color:var(--green,#22c55e)">Dodano {{ dodaj }} szt do oferty</h2>
    <p><b>{{ oferta.get('tytul','')[:60] }}</b></p>
    <p>Bylo: {{ oferta.get('ilosc',0) }} szt → teraz: <b>{{ new_qty }} szt</b></p>
    <p style="font-size:0.85rem;opacity:0.7">Przekierowanie za 3s...</p>
    <a href="/magazyn/produkt/{{ produkt.get('kod_magazynowy','') or product_id }}" class="btn btn-p">← Powrot</a>
  </div>
</div>
</body></html>'''


# === TEMPLATE: Fallback — ręczne dodawanie (gdy auto-dodanie się nie udało) ===
TEMPLATE_ALREADY_LISTED = '''<!DOCTYPE html><html><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Produkt juz wystawiony</title>
<link rel="stylesheet" href="/static/styles.css">
</head><body class="{{ 'kiosk' if request.cookies.get('kiosk') else '' }}">
<div class="container" style="max-width:700px;margin:auto;padding:20px">
  <div class="card" style="padding:20px;margin-bottom:16px">
    <h2 style="margin:0 0 12px;color:var(--accent)"><span class=material-symbols-outlined>warning</span> Produkt juz wystawiony na Allegro</h2>
    {% if error_msg %}<div style="background:var(--red-bg,#3b1111);color:var(--red,#ef4444);padding:8px 12px;border-radius:8px;margin-bottom:8px;font-size:0.9rem">{{ error_msg }}</div>{% endif %}
    <p style="margin:0 0 8px"><b>{{ produkt.get('nazwa','')[:60] }}</b></p>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:0.9rem;margin-bottom:12px">
      <div>EAN: <b>{{ produkt.get('ean','—') }}</b></div>
      <div>ASIN: <b>{{ produkt.get('asin','—') }}</b></div>
      <div>W magazynie: <b>{{ produkt.get('ilosc',0) }} szt</b></div>
      <div>Kod: <b>{{ produkt.get('kod_magazynowy','—') }}</b></div>
    </div>
  </div>

  <div class="card" style="padding:20px;margin-bottom:16px;border-left:4px solid var(--green,#22c55e)">
    <h3 style="margin:0 0 8px">Aktywna oferta Allegro</h3>
    <p style="margin:0 0 4px"><b>{{ oferta.get('tytul','')[:70] }}</b></p>
    <div style="display:flex;gap:16px;font-size:0.9rem;margin-bottom:12px">
      <span>Cena: <b>{{ "%.2f"|format(oferta.get('cena',0)) }} zl</b></span>
      <span>Ilosc: <b>{{ oferta.get('ilosc',0) }} szt</b></span>
      <span>Status: <b>{{ oferta.get('status','?') }}</b></span>
    </div>

    <form method="POST" action="/paletomat/generator/add-stock/{{ oferta.get('id') }}"
          style="display:flex;gap:8px;align-items:end;flex-wrap:wrap">
      <input type="hidden" name="csrf_token" value="{{ csrf_token() }}">
      <input type="hidden" name="product_id" value="{{ product_id }}">
      <div style="flex:1;min-width:120px">
        <label style="font-size:0.8rem;display:block;margin-bottom:4px">Dodaj sztuki:</label>
        <input type="number" name="dodaj_sztuki" value="1" min="1" max="999"
               class="form-ctrl" style="width:100%">
      </div>
      <button type="submit" class="btn btn-p" style="min-height:48px;flex:2">
        <span class=material-symbols-outlined>add</span> Dodaj sztuki do oferty
      </button>
    </form>
  </div>

  <div style="display:flex;gap:8px;flex-wrap:wrap">
    <a href="/paletomat/generator/from-magazyn/{{ product_id }}?force_new=1"
       class="btn btn-2" style="flex:1;text-align:center">
      [ADD_CIRCLE] Wystaw jako nowa oferte
    </a>
    <a href="/magazyn/produkt/{{ produkt.get('kod_magazynowy','') or product_id }}"
       class="btn" style="flex:1;text-align:center">
      ← Powrot do produktu
    </a>
  </div>
</div>
</body></html>'''


@paletomat_bp.route('/generator/add-stock/<int:oferta_id>', methods=['POST'])
def generator_add_stock(oferta_id):
    """Dodaje sztuki do istniejącej oferty Allegro zamiast tworzenia nowej"""
    from .allegro_api import update_offer_stock

    conn = get_db()
    oferta = conn.execute('SELECT id, allegro_id, ilosc, tytul FROM oferty WHERE id = ?',
                          (oferta_id,)).fetchone()

    if not oferta:
        return redirect('/paletomat/generator')

    o = dict(oferta)
    dodaj = int(request.form.get('dodaj_sztuki', 1))
    product_id = request.form.get('product_id', '')

    if dodaj < 1:
        dodaj = 1

    new_qty = (o.get('ilosc', 0) or 0) + dodaj

    result, error = update_offer_stock(o['allegro_id'], new_qty)

    if error:
        return render_template_string('''<!DOCTYPE html><html><head>
            <meta charset="utf-8"><link rel="stylesheet" href="/static/styles.css">
            </head><body><div class="container" style="max-width:600px;margin:auto;padding:20px">
            <div class="card" style="padding:20px">
              <h2 style="color:var(--red)">Blad aktualizacji</h2>
              <p>{{ error }}</p>
              <a href="/magazyn" class="btn">← Magazyn</a>
            </div></div></body></html>''', error=error)

    # Sukces
    return render_template_string('''<!DOCTYPE html><html><head>
        <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
        <link rel="stylesheet" href="/static/styles.css">
        <meta http-equiv="refresh" content="3;url=/magazyn/produkt/{{ kod }}">
        </head><body><div class="container" style="max-width:600px;margin:auto;padding:20px">
        <div class="card" style="padding:20px;border-left:4px solid var(--green,#22c55e)">
          <h2 style="color:var(--green,#22c55e)">Zaktualizowano ilosc</h2>
          <p><b>{{ tytul[:60] }}</b></p>
          <p>Dodano <b>{{ dodaj }} szt</b> → nowa ilosc: <b>{{ new_qty }} szt</b></p>
          <p style="font-size:0.85rem;opacity:0.7">Przekierowanie za 3s...</p>
          <a href="/magazyn" class="btn btn-p">← Magazyn</a>
        </div></div></body></html>''',
        tytul=o.get('tytul', ''), dodaj=dodaj, new_qty=new_qty,
        kod=request.form.get('product_id', ''))


@paletomat_bp.route('/generator/<asin>')
def generator_detail(asin):
    conn = get_db()
    p = conn.execute('SELECT * FROM scraped WHERE asin=?', (asin,)).fetchone()
    
    if not p:
        return redirect('/paletomat/generator')
    
    p = dict(p)  # Konwertuj Row na dict
    
    # Scrapuj z Amazona jeśli brak nazwy
    nazwa = p['nazwa'] or ''
    zdjecie_url = p['zdjecie_url'] or get_amazon_image_url(asin)
    cena_amazon = p['cena_amazon'] or 0
    kategoria = p.get('kategoria', 'inne') or 'inne'
    wszystkie_zdjecia = []
    bullet_points = []
    
    # Pobierz zapisane zdjęcia z bazy
    try:
        saved_images = p.get('wszystkie_zdjecia', '') or ''
        if saved_images:
            wszystkie_zdjecia = json.loads(saved_images)
    except:
        pass
    
    # Pobierz zapisane bullet points z bazy
    try:
        saved_bp = p.get('bullet_points', '') or ''
        if saved_bp:
            bullet_points = json.loads(saved_bp)
    except:
        pass
    
    amazon_domain = None
    if not nazwa or nazwa == f'Produkt {asin}':
        amazon_data = scrape_amazon_product(asin)
        if amazon_data:
            nazwa = amazon_data.get('title', '') or f'Produkt {asin}'
            # Przetłumacz nazwę na polski
            nazwa = translate_product_name(nazwa, use_ai=True)
            if amazon_data.get('image_url'):
                zdjecie_url = amazon_data['image_url']
            if amazon_data.get('all_images'):
                wszystkie_zdjecia = amazon_data['all_images']
            if amazon_data.get('price') and amazon_data['price'] > 0:
                cena_amazon = amazon_data['price']
            if amazon_data.get('bullet_points'):
                bullet_points = amazon_data['bullet_points']
            if amazon_data.get('category'):
                kategoria = amazon_data['category'] or 'inne'
            product_specs = amazon_data.get('product_specs', {})
            amazon_domain = amazon_data.get('domain')

            # Zapisz do bazy (z wszystkimi danymi)
            with get_db() as conn:
                conn.execute('''UPDATE scraped SET nazwa=?, zdjecie_url=?, wszystkie_zdjecia=?,
                               cena_amazon=?, bullet_points=?, kategoria=?, product_specs=? WHERE asin=?''',
                            (nazwa, zdjecie_url, json.dumps(wszystkie_zdjecia), cena_amazon,
                             json.dumps(bullet_points), kategoria, json.dumps(product_specs), asin))
                conn.commit()

    # Jeśli brak zdjęć w bazie, użyj głównego
    if not wszystkie_zdjecia:
        wszystkie_zdjecia = [zdjecie_url] if zdjecie_url else [get_amazon_image_url(asin)]

    # Pobierz klucz Gemini API
    from .database import get_config
    gemini_key = get_config('gemini_api_key', '')

    # Tytuł SEO: 1) z bazy (już wygenerowany) 2) AI 3) fallback optimize
    tytul_seo = (p.get('tytul_seo') or '').strip()
    if len(tytul_seo) < 20 and gemini_key:
        try:
            _product_data = {
                'nazwa': nazwa,
                'bullet_points': bullet_points,
                'kategoria': kategoria,
                'asin': asin
            }
            _ai_title = generate_allegro_title_ai(_product_data, gemini_key, max_length=75)
            if _ai_title and len(_ai_title) >= 20:
                tytul_seo = _ai_title
        except:
            pass
    if len(tytul_seo) < 5:
        tytul_seo = optimize_title_seo(nazwa, 75)

    # Oblicz sugerowaną cenę — konwersja walut wg domeny Amazon (kurs NBP)
    if cena_amazon and cena_amazon > 0:
        from .magazynier import _amazon_price_to_pln
        cena_pln = _amazon_price_to_pln(cena_amazon, amazon_domain)
        wynik = oblicz_cene_allegro(cena_pln, 40, kategoria)
    else:
        wynik = {'cena_sugerowana': '99.99'}
    
    # Generuj profesjonalny opis HTML z layoutem zdjęć (NOWA WERSJA - z bullet points + ASIN!)
    opis_html, opis_plain = generuj_opis_html_pro(nazwa, wszystkie_zdjecia, kategoria, bullet_points, gemini_key=gemini_key, asin=asin)
    
    # Generuj informacje o bezpieczeństwie GPSR (NOWA WERSJA - z kategorią!)
    gpsr_info = generuj_gpsr_info(nazwa, kategoria)
    
    # Sprawdź czy Allegro jest połączone i czy jest cennik
    from .allegro_api import is_authenticated, search_categories
    from .database import get_config
    allegro_ok = is_authenticated()
    shipping_ok = bool(get_config('allegro_shipping_id', ''))
    
    # Pobierz ilość i EAN z magazynu (jeśli produkt jest w magazynie)
    # Jeśli przyszliśmy z konkretnego produktu (/generator/from-magazyn/ID) - użyj ID
    conn = get_db()
    produkt_id_z_url = request.args.get('produkt_id', type=int)
    if produkt_id_z_url:
        magazyn_produkt = conn.execute(
            'SELECT id, ilosc, ean, stan FROM produkty WHERE id = ?',
            (produkt_id_z_url,)
        ).fetchone()
    else:
        magazyn_produkt = conn.execute(
            'SELECT id, ilosc, ean, stan FROM produkty WHERE (asin = ? OR ean = ?) AND ilosc > 0 ORDER BY ilosc DESC LIMIT 1',
            (asin, asin)
        ).fetchone()
    ilosc_magazyn = magazyn_produkt['ilosc'] if magazyn_produkt else 1
    ean_magazyn = magazyn_produkt['ean'] if magazyn_produkt and magazyn_produkt['ean'] else ''
    stan_magazyn = magazyn_produkt['stan'] if magazyn_produkt and magazyn_produkt['stan'] else 'Nowy'
    
    # Sprawdź czy produkt ma rozdzielone sztuki wg stanu
    grupy_stanow = []  # [{'stan': 'Nowy', 'ilosc': 3}, {'stan': 'Używany', 'ilosc': 2}]
    if magazyn_produkt:
        prod_id = magazyn_produkt['id']
        sztuki_rows = conn.execute(
            'SELECT stan, COUNT(*) as ile FROM sztuki WHERE produkt_id=? AND status="magazyn" GROUP BY stan ORDER BY stan',
            (prod_id,)
        ).fetchall()
        if sztuki_rows:
            grupy_stanow = [{'stan': r['stan'], 'ilosc': r['ile']} for r in sztuki_rows]
    
    # Pobierz sugerowaną kategorię z Allegro
    detected_cat = '258682'  # Domyślna: Inne
    detected_cat_name = 'Inne (auto)'
    try:
        cat_result, cat_error = search_categories(nazwa[:50])
        if cat_result and cat_result.get('matchingCategories'):
            first_cat = cat_result['matchingCategories'][0]
            detected_cat = first_cat.get('id', '258682')
            detected_cat_name = first_cat.get('name', 'Sugerowana') + ' (auto)'
    except:
        pass
    
    # Pobierz listę produktów z magazynu (dla lewego panelu)
    magazyn_products = conn.execute(
        'SELECT id, nazwa, ilosc, cena_allegro, zdjecie_url, kod_magazynowy, stan FROM produkty WHERE ilosc > 0 ORDER BY data_dodania DESC LIMIT 50'
    ).fetchall()
    magazyn_products = [dict(mp) for mp in magazyn_products]
    current_product_id = produkt_id_z_url or (magazyn_produkt['id'] if magazyn_produkt else 0)
    
    # Status Allegro
    allegro_status = ''
    if not allegro_ok:
        allegro_status = '<div class="alert alert-err" style="font-size:0.8rem;padding:8px"><span class=material-symbols-outlined>cancel</span> Nie połączono z Allegro → <a href="/allegro" style="color:#ef4444">Połącz</a></div>'
    elif not shipping_ok:
        allegro_status = '<div class="alert alert-err" style="font-size:0.8rem;padding:8px"><span class=material-symbols-outlined>cancel</span> Brak cennika wysyłki → <a href="/allegro/config" style="color:#ef4444">Wybierz cennik</a></div>'
    
    # Przygotuj JSON zdjęć
    import html as html_lib
    zdjecia_json = json.dumps(wszystkie_zdjecia)
    opis_html_escaped = html_lib.escape(opis_html)

    # Render stan fields HTML
    stan_fields_html = _render_stan_fields(grupy_stanow, stan_magazyn)

    return render_template('paletomat_generator.html',
        asin=asin,
        product=p,
        zdjecie_url=zdjecie_url,
        images=wszystkie_zdjecia,
        cena_amazon=cena_amazon,
        tytul_seo=tytul_seo,
        cena_sugerowana=wynik['cena_sugerowana'],
        detected_cat=detected_cat,
        detected_cat_name=detected_cat_name,
        ilosc_magazyn=ilosc_magazyn,
        ean_magazyn=ean_magazyn,
        stan_magazyn=stan_magazyn,
        grupy_stanow=grupy_stanow,
        stan_fields_html=stan_fields_html,
        opis_html=opis_html,
        opis_html_escaped=opis_html_escaped,
        gpsr_info=gpsr_info,
        zdjecia_json=zdjecia_json,
        allegro_ok=allegro_ok,
        shipping_ok=shipping_ok,
        magazyn_products=magazyn_products,
        current_product_id=current_product_id,
        product_count=len(magazyn_products),
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        current_user=session.get('username')
    )
    

@paletomat_bp.route('/generator/<asin>/create', methods=['POST'])
def generator_create(asin):
    tytul = request.form.get('tytul', '')[:75]
    cena = float(request.form.get('cena_allegro', 0) or 0)
    ilosc = int(request.form.get('ilosc', 1) or 1)
    opis = request.form.get('opis', '')
    kategoria = request.form.get('kategoria', '')
    ean = request.form.get('ean', '').strip()
    gpsr = request.form.get('gpsr', '').strip()
    # Auto-generuj GPSR jeśli nie podany
    if not gpsr:
        from .utils import generuj_gpsr_info
        gpsr = generuj_gpsr_info(tytul or request.form.get('nazwa', ''), kategoria or '')
    action = request.form.get('action', 'draft')
    enhance_photos_flag = '1' if request.form.get('enhance_photos') == '1' else '0'
    produkt_id_param = request.form.get('produkt_id', type=int)

    # Blokada: produkt zatrzymany "dla siebie" nie moze isc na Allegro
    if produkt_id_param:
        try:
            _conn = get_db()
            _row = _conn.execute(
                "SELECT nazwa, COALESCE(dla_siebie,0) AS ds, COALESCE(powod_zatrzymania,'') AS powod "
                "FROM produkty WHERE id = ?", (produkt_id_param,)
            ).fetchone()
            if _row and int(_row['ds']) == 1:
                _powod = (_row['powod'] or '').strip()
                _msg = f'Produkt "{(_row["nazwa"] or "?")[:60]}" jest zatrzymany dla siebie' + (f' ("{_powod[:80]}")' if _powod else '') + '. Zwolnij go na stronie produktu zeby wystawic.'
                return render(f'''
                    <div class="hdr"><h1><span class=material-symbols-outlined>lock</span> ZABLOKOWANE</h1></div>
                    <div class="alert alert-err">⛔ {_msg}</div>
                    <a href="/magazyn/produkty" class="back">← Powrót do produktów</a>
                ''')
        except Exception:
            pass

    # Pobierz wszystkie zdjęcia z JSON
    zdjecia_json = request.form.get('zdjecia', '[]')
    try:
        zdjecia = json.loads(zdjecia_json)
        if not isinstance(zdjecia, list):
            zdjecia = []
    except:
        zdjecia = []
    
    # Fallback na pojedyncze zdjęcie
    if not zdjecia:
        zdjecie = request.form.get('zdjecie', '')
        if zdjecie:
            zdjecia = [zdjecie]
    
    if action == 'allegro':
        # Pokaż stronę z progressem
        import html as html_lib
        
        zdjecia_json_safe = html_lib.escape(json.dumps(zdjecia))
        opis_escaped = opis.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$')
        gpsr_escaped = gpsr.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$').replace("'", "\\'")
        
        html = f'''
        <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> WYSTAWIANIE</h1><small>{tytul[:40]}...</small></div>

        <div class="card" style="text-align:center;padding:30px">
            <div id="progress-icon" style="font-size:3rem;margin-bottom:15px">
                <div class="spinner-ring" style="display:inline-block;width:48px;height:48px;border:4px solid rgba(143,245,255,0.2);border-top-color:#8ff5ff;border-radius:50%;animation:spin 1s linear infinite"></div>
            </div>
            <div id="progress-text" style="font-size:1.1rem;font-weight:600;margin-bottom:10px">Przygotowywanie...</div>
            <div style="background:#1e1e2e;border-radius:10px;height:16px;overflow:hidden;margin-bottom:10px">
                <div id="progress-bar" style="background:linear-gradient(90deg,#8ff5ff,#ff6b9b);height:100%;width:0%;transition:width 0.3s"></div>
            </div>
            <div id="progress-detail" style="font-size:0.8rem;color:#64748b"></div>
            <div id="timer" style="font-size:1.4rem;font-weight:700;color:#8ff5ff;margin-top:8px;font-variant-numeric:tabular-nums">00:00</div>
        </div>

        <style>@keyframes spin {{ 0% {{ transform:rotate(0deg) }} 100% {{ transform:rotate(360deg) }} }}</style>

        <div id="log" class="card" style="max-height:300px;overflow-y:auto;font-family:monospace;font-size:0.75rem">
            <div style="color:#64748b;padding:4px 0"><span class=material-symbols-outlined>sync</span> Inicjalizacja...</div>
        </div>
        
        <div id="result" style="display:none"></div>
        
        <!-- Hidden data storage -->
        <input type="hidden" id="zdjecia-data" value="{zdjecia_json_safe}">
        
        <script>
        // TIMER
        const _startTime = Date.now();
        const _timerEl = document.getElementById('timer');
        const _timerInterval = setInterval(() => {{
            const elapsed = Math.floor((Date.now() - _startTime) / 1000);
            const m = String(Math.floor(elapsed / 60)).padStart(2, '0');
            const s = String(elapsed % 60).padStart(2, '0');
            _timerEl.textContent = m + ':' + s;
        }}, 1000);
        function _stopTimer() {{
            clearInterval(_timerInterval);
            const elapsed = ((Date.now() - _startTime) / 1000).toFixed(1);
            _timerEl.textContent = elapsed + 's';
            _timerEl.style.color = '#22c55e';
        }}

        // Parse JSON from hidden input - SAFE!
        const zdjeciaRaw = document.getElementById('zdjecia-data').value;
        const zdjecia = JSON.parse(zdjeciaRaw);
        
        const params = new URLSearchParams();
        params.append('tytul', `{tytul}`);
        params.append('cena', '{cena}');
        params.append('ilosc', '{ilosc}');
        params.append('opis', `{opis_escaped}`);
        params.append('kategoria', '{kategoria}');
        params.append('zdjecia', JSON.stringify(zdjecia));  // Proper JSON encoding
        params.append('ean', '{ean}');
        params.append('gpsr', `{gpsr_escaped}`);
        params.append('produkt_id', '{produkt_id_param or ''}');
        params.append('enhance_photos', '{enhance_photos_flag}');
        // Stan — albo z selecta (brak podzialu) albo z grup sztuk
        const stanSelect = document.querySelector("select[name='stan']");
        if (stanSelect) {{
            params.append("stan", stanSelect.value || "Nowy");
        }}
        // Grupy stanow (gdy produkt rozdzielony na sztuki)
        document.querySelectorAll("input[name^='ilosc_grupa_']").forEach(el => {{
            const stanKey = el.name.replace("ilosc_grupa_", "");
            const ile = parseInt(el.value) || 0;
            if (ile > 0) {{
                params.append("ilosc_grupa_" + stanKey, ile);
                params.append("stan_grupa_" + stanKey, stanKey);
            }}
        }});
        
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: Zdjęcia z hidden input:', zdjeciaRaw);
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: Zdjęcia parsed:', zdjecia);
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: Liczba zdjęć:', zdjecia.length);
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: Pierwsze 3 URL-e:', zdjecia.slice(0, 3));
        
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: opis length:', params.get('opis') ? params.get('opis').length : 0);
        console.log('<span class=material-symbols-outlined>search</span> DEBUG: gpsr length:', params.get('gpsr') ? params.get('gpsr').length : 0);
        fetch('/paletomat/generator/{asin}/create-stream', {{method: 'POST', headers: {{'Content-Type': 'application/x-www-form-urlencoded'}}, body: params.toString()}})
            .then(response => response.body.getReader())
            .then(reader => {{
                const decoder = new TextDecoder();
                function read() {{
                    reader.read().then(({{done, value}}) => {{
                        if (done) return;
                        const text = decoder.decode(value);
                        const lines = text.split('\\n');
                        lines.forEach(line => {{
                            if (line.startsWith('data: ')) {{
                                try {{
                                    const data = JSON.parse(line.substring(6));
                                    handleEvent(data);
                                }} catch(e) {{}}
                            }}
                        }});
                        read();
                    }});
                }}
                read();
            }});
        
        function handleEvent(data) {{
            const bar = document.getElementById('progress-bar');
            const text = document.getElementById('progress-text');
            const icon = document.getElementById('progress-icon');
            const detail = document.getElementById('progress-detail');
            const log = document.getElementById('log');
            const result = document.getElementById('result');
            
            if (data.type === 'progress') {{
                bar.style.width = data.percent + '%';
                text.textContent = data.message;
                if (data.detail) detail.textContent = data.detail;
            }}
            else if (data.type === 'log') {{
                log.innerHTML += '<div style="color:' + (data.color || '#64748b') + ';padding:2px 0">' + data.message + '</div>';
                log.scrollTop = log.scrollHeight;
            }}
            else if (data.type === 'success') {{
                _stopTimer();
                bar.style.width = '100%';
                bar.style.background = '#22c55e';
                icon.innerHTML = '<span class=material-symbols-outlined>check_circle</span>';
                text.textContent = 'WYSTAWIONO!';
                text.style.color = '#22c55e';
                result.innerHTML = `
                    <div class="alert alert-ok" style="margin-top:15px">
                        <b>[CELEBRATION] OFERTA UTWORZONA!</b><br>
                        <small>ID: ${{data.offer_id}}</small>
                    </div>
                    <div class="alert alert-warn" style="font-size:0.85rem"><span class=material-symbols-outlined>warning</span> Oferta jest NIEAKTYWNA - kliknij AKTYWUJ</div>
                    <div style="display:flex;gap:10px;margin-top:10px">
                        <a href="/paletomat/oferty/${{data.offer_id}}/publish" class="btn btn-ok" style="flex:1"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> AKTYWUJ</a>
                        <a href="/paletomat/generator" class="btn btn-2" style="flex:1"><span class=material-symbols-outlined>label</span> Generuj więcej</a>
                    </div>
                `;
                result.style.display = 'block';
            }}
            else if (data.type === 'missing_location') {{
                // Produkt nie ma lokalizacji - pokaż UI przypisywania
                result.innerHTML = `
                    <div class="alert alert-warn" style="margin-top:15px">
                        <b><span class=material-symbols-outlined>pin_drop</span> PRZYPISZ LOKALIZACJĘ</b><br>
                        <small>Produkt nie ma przypisanej lokalizacji w magazynie</small>
                    </div>
                    <div style="background:rgba(255,255,255,0.05);padding:15px;border-radius:8px;margin-top:10px">
                        <input type="text" id="assign-location-input" placeholder="Wpisz lokalizację (np. A11, B2...)" 
                            style="width:100%;padding:12px;background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.2);border-radius:8px;color:white;font-size:14px;margin-bottom:10px">
                        <button onclick="assignLocationAndPrint(${{data.produkt_id}}, '${{data.offer_id}}')" 
                            class="btn btn-ok" style="width:100%">
                            <span class=material-symbols-outlined>label</span> Przypisz i wydrukuj etykietę
                        </button>
                        <div id="assign-result" style="margin-top:10px;font-size:13px"></div>
                    </div>
                    <div style="display:flex;gap:10px;margin-top:10px">
                        <a href="/paletomat/oferty/${{data.offer_id}}/publish" class="btn btn-2" style="flex:1"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> Aktywuj ofertę</a>
                        <a href="/magazynier" class="btn btn-2" style="flex:1"><span class=material-symbols-outlined>inventory_2</span> Magazynier</a>
                    </div>
                `;
                result.style.display = 'block';
            }}
            else if (data.type === 'error') {{
                _stopTimer();
                bar.style.width = '100%';
                bar.style.background = '#ef4444';
                icon.innerHTML = '<span class=material-symbols-outlined>cancel</span>';
                text.textContent = 'Błąd';
                _timerEl.style.color = '#ef4444';
                result.innerHTML = `
                    <div class="alert alert-err" style="margin-top:15px">${{data.error}}</div>
                    <a href="/paletomat/generator/{asin}" class="btn btn-p">← Spróbuj ponownie</a>
                `;
                result.style.display = 'block';
            }}
        }}
        
        // Funkcja do przypisania lokalizacji i drukowania
        function assignLocationAndPrint(produktId, offerId) {{
            const input = document.getElementById('assign-location-input');
            const resultDiv = document.getElementById('assign-result');
            const lokalizacja = input.value.trim().toUpperCase();
            
            if (!lokalizacja) {{
                resultDiv.innerHTML = '<div style="color:#ef4444"><span class=material-symbols-outlined>warning</span> Wpisz lokalizację!</div>';
                return;
            }}
            
            resultDiv.innerHTML = '<div style="color:#3b82f6"><span class=material-symbols-outlined>sync</span> Przypisuję lokalizację i drukuję...</div>';
            
            fetch('/paletomat/api/assign-location-and-print', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{
                    produkt_id: produktId,
                    lokalizacja: lokalizacja,
                    offer_id: offerId
                }})
            }})
            .then(r => r.json())
            .then(data => {{
                if (data.success) {{
                    if (data.printed) {{
                        resultDiv.innerHTML = '<div style="color:#22c55e"><span class=material-symbols-outlined>check_circle</span> Lokalizacja przypisana: ' + lokalizacja + '<br><span class=material-symbols-outlined>print</span> Etykieta wydrukowana na Niimbot B1!</div>';
                    }} else {{
                        resultDiv.innerHTML = '<div style="color:#f59e0b"><span class=material-symbols-outlined>check_circle</span> Lokalizacja przypisana: ' + lokalizacja + '<br><span class=material-symbols-outlined>warning</span> Drukarka niedostępna - wydrukuj ręcznie</div>';
                    }}
                }} else {{
                    resultDiv.innerHTML = '<div style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> Błąd: ' + (data.error || 'Nieznany błąd') + '</div>';
                }}
            }})
            .catch(err => {{
                resultDiv.innerHTML = '<div style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> Błąd połączenia: ' + err.message + '</div>';
            }});
        }}
        </script>
        
        <a href="/paletomat/generator" class="back" style="margin-top:15px">← Powrót</a>
        '''
        return render(html)
    
    else:
        # Zapisz jako szkic
        conn = get_db()
        conn.execute('''INSERT INTO oferty (tytul, opis, cena, ilosc, status, data_wystawienia) 
            VALUES (?, ?, ?, ?, 'draft', CURRENT_TIMESTAMP)''', (tytul, opis, cena, ilosc))
        conn.execute('UPDATE scraped SET status="wystawiony" WHERE asin=?', (asin,))
        conn.commit()
        
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> ZAPISANO</h1></div>
            <div class="alert alert-ok">Oferta zapisana jako szkic</div>
            <a href="/paletomat/oferty" class="btn btn-p"><span class=material-symbols-outlined>edit_note</span> Moje oferty</a>
            <a href="/paletomat/generator" class="btn btn-2"><span class=material-symbols-outlined>label</span> Generuj więcej</a>
            <a href="/paletomat" class="back">← Powrót</a>
        ''')

@paletomat_bp.route('/generator/<asin>/create-stream', methods=['GET', 'POST'])
def generator_create_stream(asin):
    """SSE stream dla pojedynczego wystawiania z progressem"""
    from .allegro_api import create_offer, is_authenticated, upload_image_to_allegro
    import time

    # Obsługa POST (duże dane) i GET (wsteczna kompatybilność)
    _src = request.form if request.method == 'POST' else request.args

    tytul = _src.get('tytul', '')[:75]
    cena = float(_src.get('cena', 0) or 0)
    ilosc = int(_src.get('ilosc', 1) or 1)
    opis = _src.get('opis', '')
    kategoria = _src.get('kategoria', '')
    ean = _src.get('ean', '').strip() or None
    gpsr = _src.get('gpsr', '').strip() or None

    print(f"[LIST_ALT] CREATE-STREAM [{asin}]: method={request.method}, opis={len(opis)} chars, gpsr={len(gpsr) if gpsr else 0} chars, tytul={tytul[:40]}")

    # Auto-generuj opis + GPSR RÓWNOLEGLE (oba to Gemini API calls)
    _need_opis = not opis or len(opis) < 50
    _need_gpsr = not gpsr

    if _need_opis and _need_gpsr:
        from concurrent.futures import ThreadPoolExecutor
        from .utils import generuj_opis_html_pro, generuj_gpsr_info
        from .database import get_config
        _gemini_key = get_config('gemini_api_key', '')
        _nazwa_gpsr = tytul or _src.get('nazwa', '')
        _kat_gpsr = kategoria or ''
        print(f"[BOLT] Generuję opis + GPSR równolegle...")

        with ThreadPoolExecutor(max_workers=2) as executor:
            f_opis = executor.submit(generuj_opis_html_pro, tytul or asin, [], kategoria, gemini_key=_gemini_key, asin=asin)
            f_gpsr = executor.submit(generuj_gpsr_info, _nazwa_gpsr, _kat_gpsr)
            opis, _ = f_opis.result()
            gpsr = f_gpsr.result()
        print(f"[EDIT_NOTE] Opis: {len(opis)} chars | GPSR: {len(gpsr) if gpsr else 0} chars")
    elif _need_opis:
        from .utils import generuj_opis_html_pro
        from .database import get_config
        print(f"⚠️ Opis pusty/za krótki ({len(opis)} chars) -> generuję automatycznie")
        _gemini_key = get_config('gemini_api_key', '')
        opis, _ = generuj_opis_html_pro(tytul or asin, [], kategoria, gemini_key=_gemini_key, asin=asin)
        print(f"[EDIT_NOTE] Wygenerowany opis: {len(opis)} chars")
    elif _need_gpsr:
        from .utils import generuj_gpsr_info
        _nazwa_gpsr = tytul or _src.get('nazwa', '')
        _kat_gpsr = kategoria or ''
        gpsr = generuj_gpsr_info(_nazwa_gpsr, _kat_gpsr)
    mark_as_published = _src.get('mark_as_published', '0') == '1'  # Czy oznaczać jako wystawione?
    enhance_photos = _src.get('enhance_photos', '0') == '1'  # Czy podrasować zdjęcia AI?
    stan = _src.get('stan', 'Nowy').strip() or 'Nowy'
    # Przekazany produkt_id — używamy do precyzyjnego update'u (unikamy pomylenia produktów z tym samym ASIN)
    produkt_id_param = _src.get('produkt_id', type=int)

    # Grupy stanów (gdy produkt rozdzielony na sztuki) — format: stan_grupa_Nowy=Nowy&ilosc_grupa_Nowy=5
    stan_grupy = []  # [{'stan': 'Nowy', 'ilosc': 5}, {'stan': 'Używany', 'ilosc': 2}]
    for key in _src:
        if key.startswith('ilosc_grupa_'):
            stan_key = key.replace('ilosc_grupa_', '')
            ile = int(_src.get(key, 0) or 0)
            if ile > 0:
                stan_grupy.append({'stan': stan_key, 'ilosc': ile})

    # Dopisz kod magazynowy do tytułu
    if produkt_id_param:
        _km_row = get_db().execute('SELECT kod_magazynowy FROM produkty WHERE id = ?', (produkt_id_param,)).fetchone()
        _km = _km_row['kod_magazynowy'] if _km_row and _km_row['kod_magazynowy'] else f"MAG-{produkt_id_param:05d}"
    else:
        _km_row = get_db().execute('SELECT id, kod_magazynowy FROM produkty WHERE (asin=? OR ean=?) LIMIT 1', (asin, asin)).fetchone()
        _km = _km_row['kod_magazynowy'] if _km_row and _km_row['kod_magazynowy'] else None
    # _km używany tylko wewnętrznie (etykiety, QR), NIE w tytule/opisie Allegro

    # Jeśli nie podano EAN w formularzu, spróbuj pobrać z magazynu
    if not ean:
        conn = get_db()
        if produkt_id_param:
            magazyn_produkt = conn.execute('SELECT ean FROM produkty WHERE id = ?', (produkt_id_param,)).fetchone()
        else:
            magazyn_produkt = conn.execute(
                'SELECT ean FROM produkty WHERE (asin = ? OR ean = ?) LIMIT 1',
                (asin, asin)
            ).fetchone()
        if magazyn_produkt and magazyn_produkt['ean']:
            ean = magazyn_produkt['ean']
    
    # FIXED: Poprawne parsowanie URL-encoded JSON + DEBUG
    from urllib.parse import unquote
    
    # DEBUG: Pokaż surowe dane
    zdjecia_raw = _src.get('zdjecia', '[]')
    print(f"\n{'='*70}")
    print(f"[SEARCH] DEBUG PARSOWANIA ZDJĘĆ:")
    print(f"   RAW (pierwsze 200 znaków): {zdjecia_raw[:200]}")
    print(f"   RAW (długość): {len(zdjecia_raw)} znaków")
    
    try:
        # Dekoduj URL encoding
        zdjecia_decoded = unquote(zdjecia_raw)
        print(f"   DECODED: {zdjecia_decoded[:200]}")
        
        # Parsuj JSON
        zdjecia = json.loads(zdjecia_decoded)
        
        if not isinstance(zdjecia, list):
            print(f"   [CANCEL] BŁĄD: Nie jest listą! Typ: {type(zdjecia)}")
            zdjecia = []
        else:
            print(f"   [CHECK_CIRCLE] SUKCES: {len(zdjecia)} zdjęć")
            for i, img in enumerate(zdjecia[:3]):
                print(f"      [{i}] {img[:80] if isinstance(img, str) else img}")
                
    except json.JSONDecodeError as e:
        print(f"   [CANCEL] BŁĄD JSON: {e}")
        print(f"   Próba parsowania: {zdjecia_decoded[:100]}")
        zdjecia = []
    except Exception as e:
        print(f"   [CANCEL] BŁĄD OGÓLNY: {e}")
        zdjecia = []
    
    print(f"{'='*70}\n")
    
    # ===================================================================
    # POBRANIE BASE_URL PRZED GENERATOREM
    # (żeby uniknąć "Working outside of request context")
    # ===================================================================
    try:
        base_url = request.url_root.rstrip('/')
    except (RuntimeError, AttributeError):
        # Fallback 1: Spróbuj z configu
        base_url = get_config('base_url')
        if not base_url:
            # Fallback 2: Użyj localhost
            base_url = 'http://localhost:5000'
    
    # Walidacja base_url (usuń trailing slash)
    base_url = base_url.rstrip('/')
    
    def generate(base_url_arg):
        # Sprawdź autoryzację
        yield f"data: {json.dumps({'type': 'progress', 'percent': 10, 'message': 'Sprawdzam połączenie...'})}\n\n"
        time.sleep(0.2)
        
        if not is_authenticated():
            yield f"data: {json.dumps({'type': 'error', 'error': 'Nie zalogowany do Allegro'})}\n\n"
            return
        
        yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>check_circle</span> Allegro OK', 'color': '#22c55e'})}\n\n"
        
        # Info o zdjęciach
        yield f"data: {json.dumps({'type': 'progress', 'percent': 20, 'message': f'Przygotowuję {len(zdjecia)} zdjęć...'})}\n\n"
        
        # Pobierz zdjecie_url z produktu (dla CDN fallback)
        _zdjecie_url_cdn = ''
        try:
            _conn = get_db()
            if produkt_id_param:
                _p_row = _conn.execute('SELECT zdjecie_url FROM produkty WHERE id = ?', (produkt_id_param,)).fetchone()
            else:
                _p_row = _conn.execute('SELECT zdjecie_url FROM produkty WHERE asin = ? LIMIT 1', (asin,)).fetchone()
            if _p_row and _p_row['zdjecie_url']:
                _zdjecie_url_cdn = _p_row['zdjecie_url']
        except:
            pass

        # SPRAWDŹ: czy lokalne pliki istnieją, re-download jeśli nie
        zdjecia_do_uploadu, img_logs = _ensure_local_images(zdjecia, asin, _zdjecie_url_cdn)
        for msg, color in img_logs:
            yield f"data: {json.dumps({'type': 'log', 'message': msg, 'color': color})}\n\n"
        if not zdjecia_do_uploadu:
            zdjecia_do_uploadu = zdjecia  # fallback na oryginalne
        
        # Normalizuj ścieżki zdjęć
        processed = []
        for img_path in zdjecia_do_uploadu[:8]:
            if img_path and isinstance(img_path, str):
                # Jeśli to lokalna ścieżka (zaczyna się od paletomat/ lub images/)
                if img_path.startswith(('paletomat/', 'images/', 'static/')):
                    # LOKALNY PLIK - konwertuj na ABSOLUTNĄ ścieżkę
                    import os as os_module
                    abs_path = os_module.path.abspath(img_path)
                    
                    # Sprawdź czy plik istnieje
                    if os_module.path.exists(abs_path):
                        processed.append(abs_path)
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>folder</span> Lokalny plik: {img_path}', 'color': '#8b5cf6'})}\n\n"
                    else:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Plik nie istnieje: {img_path}', 'color': '#f59e0b'})}\n\n"
                        
                # Jeśli to już URL
                elif img_path.startswith('http'):
                    # Amazon URL - zostaw bez zmian
                    # upload_image_to_allegro pobierze, zapisze i uploaduje
                    if 'media-amazon.com' in img_path:
                        # Optymalizuj Amazon URL do full size
                        img_path = re.sub(r'\._[A-Z0-9_,]+_\.', '._AC_SL1500_.', img_path)
                    processed.append(img_path)
                    yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>language</span> URL: {img_path[:50]}...', 'color': '#3b82f6'})}\n\n"
                else:
                    # Fallback - spróbuj jako absolutna ścieżka
                    import os as os_module
                    abs_path = os_module.path.abspath(img_path)
                    if os_module.path.exists(abs_path):
                        processed.append(abs_path)
                    else:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Nieznana ścieżka: {img_path}', 'color': '#f59e0b'})}\n\n"
        
        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> {len(processed)} zdjęć do uploadu', 'color': '#3b82f6'})}\n\n"
        
        # === UŻYJ ENHANCED ZDJĘĆ (pre-generated przez scraper) ===
        _enh_dir_s = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'enhanced', str(asin))
        if os.path.isdir(_enh_dir_s):
            _enh_fs_s = sorted([os.path.join(_enh_dir_s, f) for f in os.listdir(_enh_dir_s) if f.endswith('.jpg')])
            if _enh_fs_s:
                processed = _enh_fs_s[:8]
                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>auto_awesome</span> Użyto {len(_enh_fs_s)} zdjęć AI (pre-generated)', 'color': '#f59e0b'})}\n\n"
        # UPLOAD ZDJĘĆ DO ALLEGRO - PRZED create_offer!
        yield f"data: {json.dumps({'type': 'progress', 'percent': 55, 'message': 'Uploaduję zdjęcia...'})}\n\n"
        
        uploaded_urls = []
        for i, url in enumerate(processed[:8]):
            # Progress bar dla każdego zdjęcia
            upload_percent = 55 + int((i / len(processed)) * 20)
            yield f"data: {json.dumps({'type': 'progress', 'percent': upload_percent, 'message': f'Upload {i+1}/{len(processed)}...'})}\n\n"
            
            # Log z URL
            short_url = url[:50] + '...' if len(url) > 50 else url
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>upload</span> [{i+1}/{len(processed)}] {short_url}', 'color': '#64748b'})}\n\n"
            
            try:
                # Przekaż ASIN do upload_image_to_allegro
                allegro_url = upload_image_to_allegro(url, asin=asin)
                if allegro_url:
                    uploaded_urls.append(allegro_url)
                    yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>check_circle</span> Uploadowano pomyślnie', 'color': '#22c55e'})}\n\n"
                else:
                    yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> Upload się nie powiódł', 'color': '#ef4444'})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'log', 'message': f'   <span class=material-symbols-outlined>cancel</span> Błąd: {str(e)[:50]}', 'color': '#ef4444'})}\n\n"
            
            # Małe opóźnienie żeby user widział postęp
            time.sleep(0.3)
        
        yield f"data: {json.dumps({'type': 'progress', 'percent': 65, 'message': f'Zdjęcia gotowe ({len(uploaded_urls)}/{len(processed)})'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>photo_camera</span> Uploadowano {len(uploaded_urls)}/{len(processed)} zdjęć', 'color': '#3b82f6'})}\n\n"
        
        if not uploaded_urls:
            yield f"data: {json.dumps({'type': 'error', 'error': 'Nie udało się uploadować żadnego zdjęcia'})}\n\n"
            return
        
        yield f"data: {json.dumps({'type': 'progress', 'percent': 70, 'message': 'Tworzę ofertę...'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>upload</span> Wysyłam do Allegro...', 'color': '#64748b'})}\n\n"
        
        kat_msg = f'<span class=material-symbols-outlined>folder</span> Kategoria: {kategoria}'
        ean_msg = f'<span class=material-symbols-outlined>bar_chart</span> EAN: {ean}' if ean else '<span class=material-symbols-outlined>bar_chart</span> EAN: brak'
        asin_msg = f'<span class=material-symbols-outlined>label</span> ASIN: {asin}'
        gpsr_msg = f'<span class=material-symbols-outlined>shield</span> GPSR: {len(gpsr) if gpsr else 0} znaków'
        yield f"data: {json.dumps({'type': 'log', 'message': kat_msg, 'color': '#3b82f6'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': ean_msg, 'color': '#3b82f6'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': asin_msg, 'color': '#3b82f6'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': gpsr_msg, 'color': '#3b82f6'})}\n\n"
        
        # === TWORZENIE OFERTY (lub OFERT jeśli grupowanie po stanie) ===
        yield f"data: {json.dumps({'type': 'progress', 'percent': 75, 'message': 'Wysyłam zapytanie do API...'})}\n\n"
        
        # Jeśli są grupy stanów → tworzymy osobną ofertę per stan z odpowiednią ilością
        oferty_do_stworzenia = []
        if stan_grupy:
            for g in stan_grupy:
                oferty_do_stworzenia.append({'stan': g['stan'], 'ilosc': g['ilosc']})
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>inventory_2</span> Tworzę {len(oferty_do_stworzenia)} oferty pogrupowane po stanie...', 'color': '#8b5cf6'})}\n\n"
        else:
            oferty_do_stworzenia.append({'stan': stan, 'ilosc': ilosc})
        
        wszystkie_offer_ids = []
        for oferta_info in oferty_do_stworzenia:
            o_stan = oferta_info['stan']
            o_ilosc = oferta_info['ilosc']
            # Nie dodawaj stanu do tytulu - zmniejsza konwersje na Allegro
            o_tytul = tytul
            
            if len(oferty_do_stworzenia) > 1:
                yield f"data: {json.dumps({'type': 'log', 'message': f'  → Tworzę ofertę: {o_stan} x{o_ilosc}', 'color': '#94a3b8'})}\n\n"
            
            # Pobierz bullet_points z bazy
            _bp_for_offer = []
            if asin:
                try:
                    _bp_row = get_db().execute('SELECT bullet_points FROM scraped WHERE asin = ?', (asin,)).fetchone()
                    if _bp_row and _bp_row['bullet_points']:
                        _bp_for_offer = json.loads(_bp_row['bullet_points'])
                        if not isinstance(_bp_for_offer, list):
                            _bp_for_offer = []
                except:
                    pass
            result, error = create_offer(o_tytul, opis, cena, uploaded_urls, kategoria_id=kategoria, ilosc=o_ilosc, ean=ean, asin=asin, gpsr=gpsr, stan=o_stan, bullet_points=_bp_for_offer, kod_magazynowy=_km)
            if error:
                yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> Błąd dla {o_stan}: {error}', 'color': '#ef4444'})}\n\n"
                if len(oferty_do_stworzenia) == 1:
                    yield f"data: {json.dumps({'type': 'error', 'error': error})}\n\n"
                    return
                continue
            offer_id = result.get('id', '')
            wszystkie_offer_ids.append(offer_id)
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> Oferta {o_stan} x{o_ilosc} → ID: {offer_id}', 'color': '#22c55e'})}\n\n"
        
        # Dla kompatybilności z resztą kodu
        offer_id = wszystkie_offer_ids[0] if wszystkie_offer_ids else ''
        result = {'id': offer_id}
        error = None if wszystkie_offer_ids else 'Brak udanych ofert'
        
        if not wszystkie_offer_ids:
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>cancel</span> BŁĄD API: {error}', 'color': '#ef4444'})}\n\n"
            yield f"data: {json.dumps({'type': 'error', 'error': error})}\n\n"
            return
        
        # Sukces API!
        yield f"data: {json.dumps({'type': 'progress', 'percent': 85, 'message': f'Utworzono {len(wszystkie_offer_ids)} ofert!'})}\n\n"
        
        # === ZAPIS DO BAZY ===
        yield f"data: {json.dumps({'type': 'progress', 'percent': 90, 'message': 'Zapisuję do bazy...'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>save</span> Aktualizuję bazę danych...', 'color': '#64748b'})}\n\n"
        
        try:
            # Retry logic dla zapisu do bazy
            import time as time_mod
            max_retries = 3
            
            for attempt in range(max_retries):
                try:
                    conn = get_db()
                    
                    # Status zależy od checkboxa
                    new_status = "wystawiony" if mark_as_published else "nowy"
                    
                    # Podstawowe zapisy
                    conn.execute('UPDATE scraped SET status=? WHERE asin=?', (new_status, asin))
                    # Znajdź produkt_id — MUSI być konkretny ID żeby webhook wiedział co sprzedano
                    produkt_id = None
                    if produkt_id_param:
                        produkt_id = produkt_id_param
                    elif ean or asin:
                        p_row = conn.execute(
                            'SELECT id FROM produkty WHERE (ean = ? OR asin = ?) AND status != "sprzedany" ORDER BY data_dodania DESC LIMIT 1',
                            (ean or '', asin or '')
                        ).fetchone()
                        if p_row:
                            produkt_id = p_row['id']
                    
                    conn.execute('''INSERT OR REPLACE INTO oferty (tytul, opis, cena, ilosc, status, allegro_id, produkt_id, data_wystawienia) 
                        VALUES (?, ?, ?, ?, 'wystawiona', ?, ?, CURRENT_TIMESTAMP)''', 
                        (tytul, opis, cena, ilosc, offer_id, produkt_id))
                    
                    # Zaktualizuj produkt w magazynie używając produkt_id ustalonego wyżej
                    if produkt_id:
                        conn.execute('UPDATE produkty SET cena_allegro = ?, status = ? WHERE id = ?', (cena, new_status, produkt_id))
                    
                    # Commit i zamknij PRZED add_historia
                    conn.commit()
                    
                    # Teraz dodaj historię (otworzy własne połączenie)
                    if produkt_id:
                        from .database import add_historia
                        time_mod.sleep(0.1)  # Małe opóźnienie dla pewności
                        
                        if mark_as_published:
                            add_historia(produkt_id, 'wystawiono', f'Wystawiono na Allegro za {cena:,.0f} zł', {'allegro_id': offer_id, 'cena': cena})
                            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>inventory_2</span> Zaktualizowano magazyn: WYSTAWIONY (ID: {produkt_id})', 'color': '#22c55e'})}\n\n"
                        else:
                            add_historia(produkt_id, 'edytowano', f'Utworzono ofertę Allegro za {cena:,.0f} zł (bez zmiany statusu)', {'allegro_id': offer_id, 'cena': cena})
                            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>inventory_2</span> Zaktualizowano magazyn: NIE ZMIENIONO statusu (ID: {produkt_id})', 'color': '#3b82f6'})}\n\n"
                    
                    yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>save</span> Zapisano do bazy pomyślnie', 'color': '#22c55e'})}\n\n"
                    break  # Sukces - wyjdź z pętli retry
                    
                except sqlite3.OperationalError as e:
                    if 'database is locked' in str(e) and attempt < max_retries - 1:
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Baza zablokowana, retry {attempt+1}/{max_retries}...', 'color': '#f59e0b'})}\n\n"
                        time_mod.sleep(1 * (attempt + 1))  # Exponential backoff
                        continue
                    else:
                        raise
        
        except Exception as e:
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Błąd zapisu do bazy: {str(e)[:50]}', 'color': '#f59e0b'})}\n\n"
            # Kontynuuj mimo błędu - oferta jest na Allegro
        
        yield f"data: {json.dumps({'type': 'progress', 'percent': 100, 'message': 'Gotowe!'})}\n\n"
        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>check_circle</span> Oferta utworzona: {offer_id}', 'color': '#22c55e'})}\n\n"
        
        # ============================================================
        # INTEGRACJA: PALETOMAT → MAGAZYNIER + NIIMBOT
        # ============================================================
        try:
            from .paletomat_magazynier_integration import trigger_auto_workflow, get_locations_for_select
            
            # Sprawdź czy produkt ma już przypisaną lokalizację
            if produkt_id:
                conn = get_db()
                product = conn.execute('SELECT lokalizacja FROM produkty WHERE id = ?', (produkt_id,)).fetchone()
                
                existing_location = product['lokalizacja'] if product else None
                
                if existing_location and existing_location.strip():
                    # Ma lokalizację - wydrukuj etykietę
                    yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>label</span>  Lokalizacja: {existing_location}', 'color': '#8b5cf6'})}\n\n"
                    yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>print</span>  Automatyczne drukowanie etykiety...', 'color': '#8b5cf6'})}\n\n"
                    
                    # Wywołaj workflow (auto_print=True - AUTOMATYCZNE DRUKOWANIE!)
                    result = trigger_auto_workflow(
                        produkt_id=produkt_id,
                        offer_id=offer_id,
                        lokalizacja=existing_location,
                        auto_print=True  # <span class=material-symbols-outlined>print</span> AUTO-DRUKUJ ETYKIETĘ!
                    )
                    
                    if result['success']:
                        if result.get('printed'):
                            yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>check_circle</span> Etykieta wydrukowana na Niimbot B1!', 'color': '#22c55e'})}\n\n"
                        else:
                            yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>check_circle</span> Integracja zapisana (drukarka niedostępna)', 'color': '#3b82f6'})}\n\n"
                            yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>lightbulb</span> Wydrukuj ręcznie: python quick_print.py {produkt_id} {existing_location}', 'color': '#64748b'})}\n\n"
                    else:
                        error_msg = result.get('error', 'Nieznany błąd')
                        yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Błąd drukowania: {error_msg[:50]}', 'color': '#f59e0b'})}\n\n"
                else:
                    # Brak lokalizacji - poinformuj użytkownika + daj link
                    yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>pin_drop</span> Produkt nie ma przypisanej lokalizacji', 'color': '#f59e0b'})}\n\n"
                    yield f"data: {json.dumps({'type': 'log', 'message': '<span class=material-symbols-outlined>lightbulb</span> Przypisz lokalizację aby wydrukować etykietę', 'color': '#64748b'})}\n\n"
                    
                    # Przekaż produkt_id do frontendu żeby mógł pokazać UI przypisywania
                    yield f"data: {json.dumps({'type': 'missing_location', 'produkt_id': produkt_id, 'offer_id': offer_id})}\n\n"
        except Exception as e:
            # Nie przerywaj procesu jeśli integracja nie zadziała
            yield f"data: {json.dumps({'type': 'log', 'message': f'<span class=material-symbols-outlined>warning</span> Integracja Magazynier: {str(e)[:50]}', 'color': '#f59e0b'})}\n\n"
        
        yield f"data: {json.dumps({'type': 'log', 'message': '[CELEBRATION] ZAKOŃCZONO POMYŚLNIE!', 'color': '#22c55e'})}\n\n"
        yield f"data: {json.dumps({'type': 'success', 'offer_id': offer_id})}\n\n"
    
    # Wywołaj generator z przekazanym base_url
    return Response(generate(base_url), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })


@paletomat_bp.route('/oferty/<offer_id>/publish')
def publish_allegro_offer(offer_id):
    """Aktywuje ofertę na Allegro"""
    from .allegro_api import publish_offer
    
    result, error = publish_offer(offer_id)
    
    if error:
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">{error}</div>
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')
    
    # Zaktualizuj status w lokalnej bazie + ustaw produkt jako wystawiony
    conn = get_db()
    conn.execute('UPDATE oferty SET status="aktywna", data_wystawienia=datetime(\'now\') WHERE allegro_id=?', (offer_id,))
    _ofr = conn.execute('SELECT produkt_id FROM oferty WHERE allegro_id=?', (offer_id,)).fetchone()
    if _ofr and _ofr['produkt_id']:
        conn.execute('UPDATE produkty SET status="wystawiony" WHERE id=?', (_ofr['produkt_id'],))
    conn.commit()
    
    return render('''
        <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> AKTYWOWANO!</h1></div>
        <div class="alert alert-ok">Oferta jest teraz aktywna na Allegro!</div>
        <a href="/paletomat/oferty" class="btn btn-p"><span class=material-symbols-outlined>edit_note</span> Moje oferty</a>
        <a href="/paletomat" class="back">← Powrót</a>
    ''')

@paletomat_bp.route('/ustawienia', methods=['GET', 'POST'])
def ustawienia():
    """Ustawienia Paletomatu (Gemini API, domyślna marża itp.)"""
    if request.method == 'POST':
        set_config('gemini_api_key', request.form.get('gemini_key', '').strip())
        set_config('paletomat_marza', request.form.get('marza', '40'))
        set_config('paletomat_kurs_eur', request.form.get('kurs_eur', '4.35'))
        return redirect('/paletomat/ustawienia?saved=1')
    
    gemini_key = get_config('gemini_api_key', '')
    marza = get_config('paletomat_marza', '40')
    kurs_eur = get_config('paletomat_kurs_eur', '4.35')
    saved = request.args.get('saved')
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>settings</span> USTAWIENIA</h1><small>Paletomat</small></div>
    '''
    
    if saved:
        html += '<div class="alert alert-ok">Zapisano!</div>'
    
    html += f'''
    <form method="POST">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <div class="card">
            <div class="card-title"><span class=material-symbols-outlined>smart_toy</span> Gemini API (opisy AI)</div>
            <div class="form-group">
                <label>API Key</label>
                <input type="password" name="gemini_key" class="form-ctrl" value="{gemini_key}" placeholder="AIza...">
            </div>
            <div style="font-size:0.75rem;color:#64748b">
                Pobierz klucz z <a href="https://aistudio.google.com/app/apikey" target="_blank" style="color:#8b5cf6">Google AI Studio</a> (darmowy)
            </div>
        </div>
        
        <div class="card">
            <div class="card-title"><span class=material-symbols-outlined>paid</span> Ceny</div>
            <div class="form-row">
                <div class="form-group">
                    <label>Domyślna marża (%)</label>
                    <input type="number" name="marza" class="form-ctrl" value="{marza}">
                </div>
                <div class="form-group">
                    <label>Kurs EUR/PLN</label>
                    <input type="number" step="0.01" name="kurs_eur" class="form-ctrl" value="{kurs_eur}">
                </div>
            </div>
        </div>
        
        <button type="submit" class="btn btn-ok"><span class=material-symbols-outlined>save</span> ZAPISZ</button>
    </form>
    
    <a href="/paletomat" class="back">← Powrót</a>
    '''
    return render(html)

@paletomat_bp.route('/oferty')
def oferty():
    conn = get_db()
    all_oferty = conn.execute('SELECT * FROM oferty ORDER BY data_aktualizacji DESC').fetchall()
    drafts_count = conn.execute('SELECT COUNT(*) FROM oferty WHERE status="draft"').fetchone()[0]
    aktywne_count = conn.execute('SELECT COUNT(*) FROM oferty WHERE status="aktywna"').fetchone()[0]
    
    # Sprawdz czy Allegro polaczone
    from .allegro_api import is_authenticated
    allegro_ok = is_authenticated()
    
    html = f'''<div class="hdr"><h1><span class=material-symbols-outlined>edit_note</span> MOJE OFERTY</h1><small>{len(all_oferty)} ofert</small></div>'''
    
    # Przycisk synchronizacji (jeśli Allegro połączone)
    if allegro_ok:
        html += f'''
        <a href="/paletomat/oferty/sync" class="btn btn-2" style="margin-bottom:10px;display:flex;align-items:center;justify-content:center;gap:8px">
            <span><span class=material-symbols-outlined>sync</span></span>
            <span>SYNCHRONIZUJ Z ALLEGRO</span>
            <span style="font-size:0.7rem;opacity:0.7">(szkice: {drafts_count} | aktywne: {aktywne_count})</span>
        </a>
        '''
    
    # Przycisk masowego wystawiania
    if drafts_count > 0:
        html += f'''
        <div class="card" style="background:linear-gradient(135deg,rgba(139,92,246,0.2),rgba(88,28,135,0.2));border-color:rgba(139,92,246,0.3);margin-bottom:15px">
            <div style="font-weight:600;margin-bottom:8px"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> {drafts_count} szkiców do wystawienia</div>
            <div style="display:flex;gap:8px;flex-wrap:wrap">
        '''
        if allegro_ok:
            html += '<a href="/paletomat/oferty/publish-all" class="btn btn-ok" style="flex:1">WYSTAW NA ALLEGRO</a>'
        else:
            html += '<a href="/allegro" class="btn btn-2" style="flex:1">POŁĄCZ ALLEGRO</a>'
        html += '''
                <a href="/paletomat/oferty/export-csv" class="btn btn-2" style="flex:1"><span class=material-symbols-outlined>download</span> EKSPORT CSV</a>
            </div>
        </div>
        '''
    
    for o in all_oferty:
        if o['status'] == 'aktywna':
            status_badge = 'background:#22c55e'
            status_text = 'aktywna'
        elif o['status'] == 'wystawiona':
            status_badge = 'background:#3b82f6'
            status_text = 'wystawiona'
        else:
            status_badge = 'background:#64748b'
            status_text = 'szkic'
        
        html += f'''<a href="/paletomat/oferta/{o['id']}" class="item" style="text-decoration:none;color:#fff">
            <div class="item-info">
                <div class="item-name">{o['tytul'][:35]}...</div>
                <div class="item-meta">
                    <span style="padding:2px 8px;border-radius:4px;font-size:0.65rem;{status_badge}">{status_text}</span>
                    | {o['ilosc']} szt
                </div>
            </div>
            <div class="item-right">
                <div class="item-price">{o['cena']:.2f} zł</div>
            </div>
        </a>'''
    
    if not all_oferty:
        html += '<div class="alert alert-warn">Brak ofert. Użyj generatora aby utworzyć pierwszą.</div>'
    
    html += '''
    <a href="/paletomat/generator" class="btn btn-p" style="margin-top:15px"><span class=material-symbols-outlined>label</span> GENERUJ NOWE</a>
    <a href="/paletomat" class="back">← Powrót</a>
    '''
    return render(html)

@paletomat_bp.route('/oferty/export-csv')
def export_csv():
    """Eksportuje szkice do CSV"""
    import csv
    import io
    
    conn = get_db()
    oferty = conn.execute('SELECT * FROM oferty WHERE status="draft"').fetchall()
    
    if not oferty:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined style=color:#3b82f6>info</span> INFO</h1></div>
            <div class="alert alert-warn">Brak szkiców do eksportu</div>
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Nagłówki
    writer.writerow(['Tytuł', 'Opis', 'Cena', 'Ilość', 'Stan'])
    
    for o in oferty:
        writer.writerow([
            o['tytul'],
            o['opis'].replace('\n', ' ').replace('\r', ''),
            f"{o['cena']:.2f}",
            o['ilosc'],
            'Nowy'
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=oferty_szkice.csv'}
    )

@paletomat_bp.route('/oferty/sync')
def sync_oferty():
    """Synchronizuje statusy ofert z Allegro API"""
    from .allegro_api import sync_offers_status, is_authenticated
    
    if not is_authenticated():
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-warn">Nie jesteś połączony z Allegro. Połącz konto w ustawieniach.</div>
            <a href="/allegro" class="btn btn-p"><span class=material-symbols-outlined>link</span> POŁĄCZ ALLEGRO</a>
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')
    
    # Wywołaj synchronizację
    stats = sync_offers_status()
    
    if 'error' in stats:
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD SYNCHRONIZACJI</h1></div>
            <div class="alert alert-warn">{stats['error']}</div>
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')
    
    # Wyświetl wyniki
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> SYNCHRONIZACJA ZAKOŃCZONA</h1></div>
    
    <div class="card">
        <div class="card-title"><span class=material-symbols-outlined>bar_chart</span> Statystyki</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:15px">
            <div>
                <div style="font-size:2rem;font-weight:700;color:var(--accent)">{stats['total']}</div>
                <div style="font-size:0.8rem;color:var(--text-muted)">Wszystkich ofert</div>
            </div>
            <div>
                <div style="font-size:2rem;font-weight:700;color:#64748b">{stats['draft']}</div>
                <div style="font-size:0.8rem;color:var(--text-muted)">Szkice</div>
            </div>
            <div>
                <div style="font-size:2rem;font-weight:700;color:var(--green)">{stats['active']}</div>
                <div style="font-size:0.8rem;color:var(--text-muted)">Aktywne</div>
            </div>
            <div>
                <div style="font-size:2rem;font-weight:700;color:#ef4444">{stats['ended']}</div>
                <div style="font-size:0.8rem;color:var(--text-muted)">Zakończone</div>
            </div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-title"><span class=material-symbols-outlined>sync</span> Zmiany</div>
        <div style="padding:10px 0">
            <div style="display:flex;justify-content:space-between;padding:8px 0">
                <span>Zaktualizowano statusów:</span>
                <span style="font-weight:700;color:var(--accent)">{stats['updated']}</span>
            </div>
            <div style="display:flex;justify-content:space-between;padding:8px 0">
                <span>Nowe oferty:</span>
                <span style="font-weight:700;color:var(--green)">{stats['new']}</span>
            </div>
        </div>
    </div>
    
    <div class="alert alert-ok">
        Statusy ofert zostały zsynchronizowane z Allegro! Możesz teraz zobaczyć aktualne stany wszystkich ofert.
    </div>
    
    <a href="/paletomat/oferty" class="btn btn-p"><span class=material-symbols-outlined>edit_note</span> ZOBACZ OFERTY</a>
    <a href="/paletomat" class="back">← Powrót</a>
    '''
    
    return render(html)

@paletomat_bp.route('/oferta/<int:oferta_id>')
def oferta_detail(oferta_id):
    """Szczegoly oferty z opcja wystawienia/edycji"""
    conn = get_db()
    o = conn.execute('SELECT * FROM oferty WHERE id=?', (oferta_id,)).fetchone()
    
    if not o:
        return redirect('/paletomat/oferty')
    
    from .allegro_api import is_authenticated
    allegro_ok = is_authenticated()
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>edit_note</span> OFERTA</h1><small>#{o['id']}</small></div>
    
    <div class="card">
        <div style="font-weight:600;font-size:1.1rem;margin-bottom:10px">{o['tytul']}</div>
        <div style="font-size:1.3rem;font-weight:700;color:#22c55e;margin-bottom:15px">{o['cena']:.2f} zł</div>
        <div style="background:#0a0a0f;border-radius:8px;padding:12px;font-size:0.85rem;max-height:200px;overflow-y:auto;white-space:pre-wrap">{o['opis'][:500]}...</div>
    </div>
    '''
    
    if o['status'] == 'draft':
        if allegro_ok:
            html += f'<a href="/paletomat/oferta/{o["id"]}/publish" class="btn btn-ok"><span class=material-symbols-outlined>shopping_cart</span> WYSTAW NA ALLEGRO</a>'
        else:
            html += '<div class="alert alert-warn"><a href="/allegro" style="color:#eab308">Połącz Allegro</a> żeby wystawiać</div>'
        html += f'<a href="/paletomat/oferta/{o["id"]}/delete" class="btn btn-2" style="color:#ef4444"><span class=material-symbols-outlined>delete</span> USUŃ</a>'
    elif o['status'] == 'wystawiona' and o['allegro_id']:
        html += f'<a href="/paletomat/oferty/{o["allegro_id"]}/publish" class="btn btn-ok"><span class="material-symbols-outlined" style="font-size:1rem;vertical-align:middle">rocket_launch</span> AKTYWUJ NA ALLEGRO</a>'
    
    html += '<a href="/paletomat/oferty" class="back">← Powrót</a>'
    return render(html)

@paletomat_bp.route('/oferta/<int:oferta_id>/publish')
def publish_single_draft(oferta_id):
    """Wystawia pojedynczy szkic na Allegro"""
    from .allegro_api import create_offer, is_authenticated
    
    if not is_authenticated():
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nie jesteś zalogowany do Allegro</div>
            <a href="/allegro" class="btn btn-allegro">[KEY] Zaloguj</a>
        ''')
    
    conn = get_db()
    o = conn.execute('SELECT * FROM oferty WHERE id=?', (oferta_id,)).fetchone()

    if not o:
        return redirect('/paletomat/oferty')

    # Blokada: jesli powiazany produkt jest "dla siebie" - nie wystawiaj
    if o['produkt_id']:
        _p = conn.execute(
            "SELECT nazwa, COALESCE(dla_siebie,0) AS ds, COALESCE(powod_zatrzymania,'') AS powod "
            "FROM produkty WHERE id = ?", (o['produkt_id'],)
        ).fetchone()
        if _p and int(_p['ds']) == 1:
            _powod = (_p['powod'] or '').strip()
            return render(f'''
                <div class="hdr"><h1><span class=material-symbols-outlined>lock</span> ZABLOKOWANE</h1></div>
                <div class="alert alert-err">⛔ Produkt "{(_p['nazwa'] or '?')[:60]}" jest zatrzymany dla siebie{(' (' + _powod[:80] + ')') if _powod else ''}. Zwolnij go zeby wystawic.</div>
                <a href="/paletomat/oferty" class="back">← Powrót</a>
            ''')

    # Auto-generuj GPSR
    from .utils import generuj_gpsr_info
    gpsr = generuj_gpsr_info(o['tytul'] or '', o.get('kategoria', '') or '')

    # Wystaw na Allegro
    result, error = create_offer(o['tytul'], o['opis'], o['cena'], ilosc=o['ilosc'], gpsr=gpsr)
    
    if error:
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">{error}</div>
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')
    
    allegro_id = result.get('id', '')
    conn.execute('UPDATE oferty SET status="wystawiona", allegro_id=? WHERE id=?', (allegro_id, oferta_id))
    conn.commit()
    
    return render(f'''
        <div class="hdr"><h1>[CELEBRATION] WYSTAWIONO!</h1></div>
        <div class="alert alert-ok">Oferta wystawiona na Allegro!<br><small>ID: {allegro_id}</small></div>
        <a href="/paletomat/oferty/{allegro_id}/publish" class="btn btn-ok"><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> AKTYWUJ</a>
        <a href="/paletomat/oferty" class="btn btn-2"><span class=material-symbols-outlined>edit_note</span> Wróć do ofert</a>
    ''')

@paletomat_bp.route('/oferta/<int:oferta_id>/delete')
def delete_draft(oferta_id):
    """Usuwa szkic"""
    conn = get_db()
    conn.execute('DELETE FROM oferty WHERE id=? AND status="draft"', (oferta_id,))
    conn.commit()
    return redirect('/paletomat/oferty')

@paletomat_bp.route('/oferty/publish-all')
def publish_all_drafts():
    """Wystawia wszystkie szkice na Allegro"""
    from .allegro_api import create_offer, is_authenticated
    
    if not is_authenticated():
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Nie jesteś zalogowany do Allegro</div>
            <a href="/allegro" class="btn btn-allegro">[KEY] Zaloguj</a>
        ''')
    
    conn = get_db()
    drafts = conn.execute('''
        SELECT o.* FROM oferty o
        LEFT JOIN produkty p ON p.id = o.produkt_id
        WHERE o.status="draft"
          AND COALESCE(p.dla_siebie, 0) = 0
    ''').fetchall()
    skipped_dla_siebie = conn.execute('''
        SELECT o.tytul, COALESCE(p.powod_zatrzymania,'') AS powod
        FROM oferty o JOIN produkty p ON p.id = o.produkt_id
        WHERE o.status="draft" AND COALESCE(p.dla_siebie, 0) = 1
    ''').fetchall()

    if not drafts:
        skipped_html = ''
        if skipped_dla_siebie:
            skipped_html = f'<div class="alert alert-warn" style="margin-top:10px"><span class=material-symbols-outlined>lock</span> Pominieto {len(skipped_dla_siebie)} szkicow zatrzymanych dla siebie</div>'
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined style=color:#3b82f6>info</span> INFO</h1></div>
            <div class="alert alert-warn">Brak szkicow do wystawienia</div>
            {skipped_html}
            <a href="/paletomat/oferty" class="back">← Powrót</a>
        ''')

    success = 0
    errors = []
    
    from .utils import generuj_gpsr_info

    for draft in drafts:
        gpsr = generuj_gpsr_info(draft['tytul'] or '', draft.get('kategoria', '') or '')
        result, error = create_offer(draft['tytul'], draft['opis'], draft['cena'], ilosc=draft['ilosc'], gpsr=gpsr)
        
        if error:
            errors.append(f"{draft['tytul'][:30]}: {error}")
        else:
            allegro_id = result.get('id', '')
            conn.execute('UPDATE oferty SET status="wystawiona", allegro_id=? WHERE id=?', (allegro_id, draft['id']))
            success += 1
    
    conn.commit()
    
    html = f'''
    <div class="hdr"><h1><span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> WYSTAWIONO!</h1></div>
    <div class="alert alert-ok">Wystawiono {success} z {len(drafts)} ofert</div>
    '''
    
    if errors:
        html += f'<div class="alert alert-err" style="font-size:0.8rem">Błędy ({len(errors)}):<br>{"<br>".join(errors[:5])}</div>'

    if skipped_dla_siebie:
        _list = '<br>'.join(f'• {(dict(r).get("tytul") or "?")[:50]}' + (f' <span style="color:#94a3b8">("{(dict(r).get("powod") or "")[:40]}")</span>' if dict(r).get("powod") else '') for r in skipped_dla_siebie[:10])
        _more = f'<br><em>...i jeszcze {len(skipped_dla_siebie)-10}</em>' if len(skipped_dla_siebie) > 10 else ''
        html += f'<div class="alert alert-warn" style="font-size:0.8rem;border-left:3px solid #ef4444"><span class=material-symbols-outlined>lock</span> Pominieto {len(skipped_dla_siebie)} szkicow zatrzymanych dla siebie:<br>{_list}{_more}</div>'

    html += '''
    <div class="alert alert-warn" style="font-size:0.85rem"><span class=material-symbols-outlined>warning</span> Oferty są NIEAKTYWNE - aktywuj je w panelu Allegro lub pojedynczo</div>
    <a href="/paletomat/oferty" class="btn btn-p"><span class=material-symbols-outlined>edit_note</span> Moje oferty</a>
    <a href="/paletomat" class="back">← Powrót</a>
    '''
    return render(html)

@paletomat_bp.route('/monitoring')
def monitoring():
    s = get_stats()
    
    conn = get_db()
    sprzedaze = conn.execute('SELECT * FROM sprzedaze ORDER BY data_sprzedazy DESC LIMIT 10').fetchall()
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>bar_chart</span> MONITORING</h1><small>Sprzedaż i statystyki</small></div>
    
    <div class="stats">
        <div class="stat"><div class="stat-v green">{s['sprzedane']}</div><div class="stat-l">Sprzedanych</div></div>
        <div class="stat"><div class="stat-v green">{s['przychod']:,.0f} zł</div><div class="stat-l">Przychód</div></div>
        <div class="stat"><div class="stat-v">{s['aktywne']}</div><div class="stat-l">Aktywnych</div></div>
    </div>
    
    <div class="section"><span class=material-symbols-outlined>paid</span> OSTATNIE SPRZEDAŻE</div>
    '''
    
    for sp in sprzedaze:
        html += f'''<div class="item">
            <div class="item-dot green"></div>
            <div class="item-info">
                <div class="item-name">Sprzedaż #{sp['id']}</div>
                <div class="item-meta">{sp['data_sprzedazy']}</div>
            </div>
            <div class="item-price">{sp['cena']:.2f} zł</div>
        </div>'''
    
    if not sprzedaze:
        html += '<div style="text-align:center;color:#64748b;padding:20px">Brak sprzedaży</div>'
    
    html += '<a href="/paletomat" class="back">← Powrót</a>'
    return render(html)

# ============================================================
# API
# ============================================================
@paletomat_bp.route('/api/stats')
def api_stats():
    return jsonify(get_stats())

@paletomat_bp.route('/api/queue-status')
def api_queue_status():
    """Zwraca status kolejki auto-processingu + progress kombajnu"""
    global _processing_queue, _scraper_running, PROGRESS
    return jsonify({
        'running': _scraper_running,
        'queue_length': len(_processing_queue),
        'queue': list(_processing_queue),
        'progress': PROGRESS,  # 🌾 NOWE: pokazuje postęp
        'workers': MAX_WORKERS  # <span class=material-symbols-outlined>agriculture</span> NOWE: ile równolegle
    })




@paletomat_bp.route('/api/assign-location-and-print', methods=['POST'])
def api_assign_location_and_print():
    """
    API endpoint: Przypisz lokalizację do produktu i od razu wydrukuj etykietę
    Używane po wystawieniu oferty gdy produkt nie ma lokalizacji
    """
    try:
        data = request.get_json()
        produkt_id = data.get('produkt_id')
        lokalizacja = data.get('lokalizacja', '').strip().upper()
        offer_id = data.get('offer_id')
        
        if not produkt_id or not lokalizacja:
            return jsonify({
                'success': False,
                'error': 'Brak wymaganych danych (produkt_id, lokalizacja)'
            }), 400
        
        # Zapisz lokalizację w magazynie
        conn = get_db()
        conn.execute('UPDATE produkty SET lokalizacja = ? WHERE id = ?', (lokalizacja, produkt_id))
        conn.commit()
        
        # Trigger workflow - auto-drukuj
        from .paletomat_magazynier_integration import trigger_auto_workflow
        
        result = trigger_auto_workflow(
            produkt_id=produkt_id,
            offer_id=offer_id or f'manual_{produkt_id}',
            lokalizacja=lokalizacja,
            auto_print=True  # AUTO-DRUKUJ!
        )
        
        if result['success']:
            return jsonify({
                'success': True,
                'printed': result.get('printed', False),
                'message': 'Lokalizacja przypisana i etykieta wydrukowana!' if result.get('printed') else 'Lokalizacja przypisana (drukarka niedostępna)'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Nieznany błąd')
            }), 500
            
    except Exception as e:
        print(f"[CANCEL] Error in assign_location_and_print: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# IMAGE CLEANER - usuwanie overlayow ze zdjec
# ============================================================

@paletomat_bp.route('/api/clean-image', methods=['POST'])
def api_clean_image():
    """Czysci zdjecie z overlayow (strzalki, napisy, wymiary)"""
    from .image_cleaner import clean_image_from_url, GEMINI_AVAILABLE

    if not GEMINI_AVAILABLE:
        return jsonify({'success': False, 'error': 'Gemini API niedostepne'}), 500

    data = request.get_json()
    image_url = data.get('url', '')

    if not image_url:
        return jsonify({'success': False, 'error': 'Brak URL zdjecia'}), 400

    img_bytes, mime_type, error = clean_image_from_url(image_url)

    if error:
        return jsonify({'success': False, 'error': error}), 500

    # Zapisz oczyszczone zdjecie i zwroc URL
    import hashlib
    import base64
    fname = hashlib.md5(image_url.encode(), usedforsecurity=False).hexdigest()[:12] + '_clean.jpg'

    static_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'cleaned')
    os.makedirs(static_dir, exist_ok=True)

    # Konwertuj na JPEG
    from PIL import Image as PILImage
    from io import BytesIO as BIO
    cleaned_img = PILImage.open(BIO(img_bytes)).convert('RGB')
    out_path = os.path.join(static_dir, fname)
    cleaned_img.save(out_path, 'JPEG', quality=92)

    return jsonify({
        'success': True,
        'cleaned_url': f'/static/cleaned/{fname}',
        'original_url': image_url
    })


@paletomat_bp.route('/api/enhance-image', methods=['POST'])
def api_enhance_image():
    """Generuje pojedyncze zdjecie wg szablonu Allegro (1-8)"""
    from .image_enhancer import enhance_single, GEMINI_AVAILABLE

    if not GEMINI_AVAILABLE:
        return jsonify({'success': False, 'error': 'Gemini API niedostepne'}), 500

    data = request.get_json()
    image_url = data.get('url', '')
    template_id = data.get('template_id', 1)
    product_name = data.get('product_name', '')

    if not image_url:
        return jsonify({'success': False, 'error': 'Brak URL zdjecia'}), 400

    # Pobierz zdjecie (moze byc lokalne /static/cleaned/ lub zdalne)
    import hashlib
    try:
        if image_url.startswith('/static/'):
            # Lokalne zdjecie (np. po czyszczeniu)
            local_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                image_url.lstrip('/')
            )
            with open(local_path, 'rb') as f:
                img_bytes = f.read()
        else:
            import requests as req_lib
            resp = req_lib.get(image_url, timeout=30)
            resp.raise_for_status()
            img_bytes = resp.content
    except Exception as e:
        return jsonify({'success': False, 'error': f'Blad pobierania: {e}'}), 500

    # Generuj
    result_bytes, mime_type, error = enhance_single(
        img_bytes, int(template_id), product_name
    )

    if error:
        return jsonify({'success': False, 'error': error}), 500

    # Zapisz
    fname_hash = hashlib.md5(image_url.encode(), usedforsecurity=False).hexdigest()[:8]
    fname = f'{fname_hash}_t{template_id}.jpg'

    static_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'static', 'enhanced'
    )
    os.makedirs(static_dir, exist_ok=True)

    from PIL import Image as PILImage
    from io import BytesIO as BIO
    enhanced_img = PILImage.open(BIO(result_bytes)).convert('RGB')
    out_path = os.path.join(static_dir, fname)
    enhanced_img.save(out_path, 'JPEG', quality=92)

    return jsonify({
        'success': True,
        'enhanced_url': f'/static/enhanced/{fname}',
        'template_id': template_id
    })


@paletomat_bp.route('/api/fix-names', methods=['POST'])
def api_fix_names():
    """Naprawia nazwy produktów — re-scrapuje tytuły z Amazona dla produktów z ASIN w nazwie lub placeholder"""
    conn = get_db()
    # Znajdź produkty z podejrzanymi nazwami
    bad_products = conn.execute('''
        SELECT id, nazwa, asin, ean FROM produkty
        WHERE asin IS NOT NULL AND asin != ''
        AND (
            nazwa LIKE '%' || asin || '%'
            OR nazwa LIKE 'Produkt %'
            OR nazwa IS NULL
            OR nazwa = ''
            OR LENGTH(nazwa) < 10
        )
    ''').fetchall()

    if not bad_products:
        return jsonify({'success': True, 'msg': 'Wszystkie nazwy OK', 'fixed': 0})

    fixed = 0
    errors = []
    for p in bad_products:
        asin = p['asin']
        if not asin:
            continue
        try:
            amazon_data = scrape_amazon_product(asin)
            if amazon_data and amazon_data.get('title'):
                nazwa = amazon_data['title']
                nazwa = translate_product_name(nazwa, use_ai=True)
                if nazwa and len(nazwa) > 15 and not nazwa.lower().startswith('amazon'):
                    conn.execute('UPDATE produkty SET nazwa=? WHERE id=?', (nazwa, p['id']))
                    # Zaktualizuj też scraped
                    conn.execute('UPDATE scraped SET nazwa=? WHERE asin=?', (nazwa, asin))
                    fixed += 1
                else:
                    errors.append(f"{asin}: nazwa za krótka lub zła ({nazwa[:30] if nazwa else 'None'})")
            else:
                errors.append(f"{asin}: scraping failed")
            time.sleep(2)  # Nie spamuj Amazona
        except Exception as e:
            errors.append(f"{asin}: {str(e)[:50]}")

    conn.commit()
    return jsonify({
        'success': True,
        'msg': f'Naprawiono {fixed}/{len(bad_products)} nazw',
        'fixed': fixed,
        'total_bad': len(bad_products),
        'errors': errors[:10]
    })


def _search_allegro_prices(ean=None, nazwa=None):
    """Szuka cen konkurencji na Allegro po EAN lub nazwie produktu"""
    from .allegro_api import allegro_request
    prices = []

    # 1. Szukaj po EAN (najdokładniejsze)
    if ean and len(str(ean)) >= 8:
        result, err = allegro_request('GET', '/offers/listing', params={
            'phrase': str(ean),
            'limit': 10,
            'sort': 'price_asc'
        })
        if result and not err:
            items = result.get('items', {})
            for group in ['promoted', 'regular']:
                for item in items.get(group, []):
                    try:
                        p = item.get('sellingMode', {}).get('price', {})
                        cena = float(p.get('amount', 0))
                        if cena > 0:
                            prices.append(cena)
                    except (ValueError, TypeError):
                        pass
            if prices:
                return prices

    # 2. Szukaj po nazwie (fallback)
    if nazwa and len(nazwa) > 10:
        # Weź pierwsze 5 słów nazwy (bez zbędnych)
        slowa = [w for w in nazwa.split() if len(w) > 2][:5]
        fraza = ' '.join(slowa)
        result, err = allegro_request('GET', '/offers/listing', params={
            'phrase': fraza,
            'limit': 10,
            'sort': 'price_asc'
        })
        if result and not err:
            items = result.get('items', {})
            for group in ['promoted', 'regular']:
                for item in items.get(group, []):
                    try:
                        p = item.get('sellingMode', {}).get('price', {})
                        cena = float(p.get('amount', 0))
                        if cena > 0:
                            prices.append(cena)
                    except (ValueError, TypeError):
                        pass

    return prices


@paletomat_bp.route('/api/auto-price', methods=['GET', 'POST'])
def api_auto_price():
    """Autowycena — sprawdza ceny konkurencji na Allegro i ustawia taniej"""
    conn = get_db()
    if request.method == 'POST':
        data = request.get_json() or {}
    else:
        data = {}

    podciecie = float(data.get('podciecie', '2'))  # O ile % taniej niż konkurencja
    min_cena = float(data.get('min_cena', '19.99'))
    marza_fallback = float(data.get('marza', get_config('paletomat_marza', '40')))

    # Kurs EUR na fallback
    try:
        from .magazynier import _get_nbp_rate
        kurs_eur = _get_nbp_rate('EUR')
    except Exception:
        kurs_eur = float(get_config('paletomat_kurs_eur', '4.35'))

    produkty = conn.execute('''
        SELECT p.id, p.asin, p.ean, p.nazwa, p.cena_allegro, p.kategoria,
               s.cena_amazon
        FROM produkty p
        LEFT JOIN scraped s ON p.asin = s.asin
        WHERE p.status IN ('magazyn', 'wystawiony')
        AND p.asin IS NOT NULL AND p.asin != ''
    ''').fetchall()

    updated = 0
    skipped = 0
    details = []

    for p in produkty:
        stara_cena = float(p['cena_allegro'] or 0)
        cena_amazon = float(p['cena_amazon'] or 0)
        ean = p['ean'] or ''
        nazwa = p['nazwa'] or ''
        kategoria = p['kategoria'] or 'inne'

        # Szukaj cen na Allegro
        try:
            allegro_prices = _search_allegro_prices(ean=ean, nazwa=nazwa)
        except Exception:
            allegro_prices = []

        if allegro_prices:
            # Najtańsza oferta na Allegro
            min_allegro = min(allegro_prices)
            avg_allegro = sum(allegro_prices) / len(allegro_prices)

            # Ustaw cenę: najtańsza - X%
            nowa_cena = round(min_allegro * (1 - podciecie / 100), 2)

            # Zaokrągl do X.99
            nowa_cena = int(nowa_cena) + 0.99 if nowa_cena > 10 else round(nowa_cena, 2)

            zrodlo = f'allegro (min:{min_allegro:.0f}, avg:{avg_allegro:.0f}, ofert:{len(allegro_prices)})'
        elif cena_amazon > 0:
            # Cena Amazon jest już w PLN — obniż o 10-20%
            # Losowy rabat 10-20% żeby ceny nie były identyczne
            import random
            rabat = random.uniform(0.10, 0.20)
            nowa_cena = round(cena_amazon * (1 - rabat), 2)

            # Zaokrągl do X.99
            nowa_cena = int(nowa_cena) + 0.99 if nowa_cena > 10 else round(nowa_cena, 2)

            zrodlo = f'amazon-{int(rabat*100)}%'
        else:
            skipped += 1
            continue

        # Minimalna cena
        if nowa_cena < min_cena:
            nowa_cena = min_cena

        # Aktualizuj jeśli zmiana >1 zł lub brak ceny
        if abs(nowa_cena - stara_cena) > 1 or stara_cena == 0:
            conn.execute('UPDATE produkty SET cena_allegro=? WHERE id=?', (nowa_cena, p['id']))
            updated += 1
            details.append({
                'nazwa': nazwa[:40],
                'stara': round(stara_cena, 2),
                'nowa': nowa_cena,
                'zrodlo': zrodlo
            })

        # Nie spamuj API
        if allegro_prices or not cena_amazon:
            time.sleep(0.5)

    conn.commit()
    return jsonify({
        'success': True,
        'msg': f'Zaktualizowano {updated} cen (pominięto {skipped})',
        'updated': updated,
        'skipped': skipped,
        'podciecie_procent': podciecie,
        'details': details[:30]
    })