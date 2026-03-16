"""
MAIL IMPORT - Automatyczny import palet z emaili
=================================================
Monitoruje skrzynkę IMAP, szuka maili z Excelem od skonfigurowanych nadawców,
auto-tworzy paletę z kolejnym numerem i importuje produkty.

Konfiguracja: /ustawienia/mail-import
"""

import imaplib
import email
import os
import re
import hashlib
import tempfile
import logging
import json
import threading
import time
from datetime import datetime, timedelta
from email.header import decode_header
from flask import Blueprint, request, redirect, flash, jsonify, session

mail_import_bp = Blueprint('mail_import', __name__)


# ============================================================
# HELPERS
# ============================================================

def _get_config():
    """Pobiera konfigurację mail import z DB"""
    from modules.database import get_config
    return {
        'enabled': get_config('mail_import_enabled', '0') == '1',
        'imap_server': get_config('mail_import_imap_server', 'imap.gmail.com'),
        'imap_port': int(get_config('mail_import_imap_port', '993')),
        'email': get_config('mail_import_email', ''),
        'password': get_config('mail_import_password', ''),
        'sender_filter': get_config('mail_import_sender_filter', ''),  # email dziadka
        'check_interval': int(get_config('mail_import_check_interval', '15')),  # minuty
        'auto_import': get_config('mail_import_auto_import', '1') == '1',
        'default_dostawca': get_config('mail_import_default_dostawca', 'Warrington'),
    }


def _get_next_paleta_number():
    """Zwraca następny numer palety (#26, #27, etc.)"""
    from modules.database import get_db
    conn = get_db()
    rows = conn.execute("SELECT nazwa FROM palety").fetchall()

    max_num = 0
    for row in rows:
        nazwa = row['nazwa'] or ''
        match = re.match(r'#(\d+)', nazwa)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num

    return max_num + 1


def _get_attachment_hash(data):
    """Generuje hash załącznika do wykrywania duplikatów"""
    return hashlib.sha256(data).hexdigest()[:16]


def _is_already_imported(file_hash):
    """Sprawdza czy plik z tym hashem już został zaimportowany"""
    from modules.database import get_db
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM mail_import_log WHERE file_hash = ?", (file_hash,)
    ).fetchone()
    return row is not None


def _log_import(file_hash, filename, paleta_id, products_count, sender, subject, status='success', error=''):
    """Zapisuje log importu do DB"""
    from modules.database import get_db
    conn = get_db()
    local_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT INTO mail_import_log (file_hash, filename, paleta_id, products_count,
                                      sender, subject, status, error, data_importu)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (file_hash, filename, paleta_id, products_count, sender, subject, status, error, local_now))
    conn.commit()


def _decode_header_value(value):
    """Dekoduje nagłówek email (obsługuje UTF-8, ISO etc.)"""
    if not value:
        return ''
    decoded_parts = decode_header(value)
    result = ''
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            result += part.decode(charset or 'utf-8', errors='replace')
        else:
            result += part
    return result


def _render_flash_messages():
    """Renderuje flash messages jako HTML"""
    from flask import get_flashed_messages
    messages = get_flashed_messages(with_categories=True)
    if not messages:
        return ''
    html = ''
    for category, message in messages:
        color = '#22c55e' if category == 'success' else '#ef4444' if category == 'error' else '#3b82f6'
        html += f'<div style="padding:12px;margin-bottom:12px;background:rgba({",".join(str(int(color[i:i+2],16)) for i in (1,3,5))},0.15);border:1px solid {color};border-radius:8px;color:{color};font-size:0.85rem;word-break:break-all">{message}</div>'
    return html


def _send_telegram_alert(message):
    """Wysyła powiadomienie Telegram o imporcie"""
    try:
        from modules.telegram_bot import send_telegram
        send_telegram(message, silent=False)
    except Exception as e:
        logging.warning(f"[MailImport] Telegram error: {e}")


# ============================================================
# INIT DB TABLE
# ============================================================

def init_mail_import_db():
    """Tworzy tabelę logów mail importu"""
    from modules.database import get_db
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS mail_import_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_hash TEXT UNIQUE,
            filename TEXT DEFAULT '',
            paleta_id INTEGER,
            products_count INTEGER DEFAULT 0,
            sender TEXT DEFAULT '',
            subject TEXT DEFAULT '',
            status TEXT DEFAULT 'success',
            error TEXT DEFAULT '',
            data_importu TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()


# ============================================================
# CORE: CHECK MAILBOX
# ============================================================

def check_mailbox(manual=False):
    """
    Sprawdza skrzynkę IMAP, szuka nowych maili z Excelem.

    Returns:
        dict: {"checked": int, "imported": int, "skipped": int, "errors": list}
    """
    config = _get_config()
    result = {"checked": 0, "imported": 0, "skipped": 0, "errors": [], "details": []}

    logging.warning(f"[MailImport] check_mailbox(manual={manual}) — email={config['email']}, sender={config['sender_filter']}, enabled={config['enabled']}")

    if not manual and not config['enabled']:
        result["errors"].append("Auto-import wyłączony")
        return result

    if not config['email'] or not config['password']:
        result["errors"].append("Brak konfiguracji email (adres/hasło)")
        return result

    if not config['sender_filter']:
        result["errors"].append("Brak filtra nadawcy (od kogo szukać maili)")
        return result

    try:
        # Połącz z IMAP
        logging.warning(f"[MailImport] Łączę z {config['imap_server']}:{config['imap_port']}...")
        mail = imaplib.IMAP4_SSL(config['imap_server'], config['imap_port'])
        mail.login(config['email'], config['password'])
        mail.select('INBOX')
        logging.info("[MailImport] IMAP login OK")

        # Szukaj maili od nadawcy z ostatnich 30 dni
        since_date = (datetime.now() - timedelta(days=30)).strftime('%d-%b-%Y')
        sender = config['sender_filter']

        # Szukaj maili od nadawcy
        search_criteria = f'(FROM "{sender}" SINCE {since_date})'
        logging.warning(f"[MailImport] IMAP search: {search_criteria}")
        status, messages = mail.search(None, search_criteria)

        if status != 'OK':
            result["errors"].append(f"IMAP search error: {status}")
            mail.logout()
            return result

        message_ids = messages[0].split()
        result["checked"] = len(message_ids)
        result["details"].append(f"Znaleziono {len(message_ids)} maili od {sender}")
        logging.warning(f"[MailImport] Znaleziono {len(message_ids)} maili od {sender}")

        for msg_id in message_ids:
            try:
                status, msg_data = mail.fetch(msg_id, '(RFC822)')
                if status != 'OK':
                    continue

                msg = email.message_from_bytes(msg_data[0][1])
                subject = _decode_header_value(msg['Subject'])
                from_addr = _decode_header_value(msg['From'])

                # Szukaj załączników Excel
                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition") or '')
                    if 'attachment' not in content_disposition:
                        continue

                    filename = part.get_filename()
                    if not filename:
                        continue

                    filename = _decode_header_value(filename)

                    # Tylko pliki Excel
                    if not filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                        continue

                    # Pobierz dane załącznika
                    attachment_data = part.get_payload(decode=True)
                    if not attachment_data:
                        continue

                    # Sprawdź duplikat
                    file_hash = _get_attachment_hash(attachment_data)

                    if _is_already_imported(file_hash):
                        result["skipped"] += 1
                        result["details"].append(f"⏭️ Pominięto (już zaimportowany): {filename}")
                        continue

                    # Zapisz do pliku tymczasowego i importuj
                    import_result = _import_attachment(
                        attachment_data, filename, file_hash,
                        from_addr, subject, config
                    )

                    if import_result.get('success'):
                        result["imported"] += 1
                        result["details"].append(
                            f"✅ Zaimportowano: {filename} → Paleta #{import_result['paleta_numer']} "
                            f"({import_result['products_count']} produktów)"
                        )
                    else:
                        result["errors"].append(
                            f"❌ Błąd importu {filename}: {import_result.get('error', 'nieznany')}"
                        )

            except Exception as e:
                result["errors"].append(f"Błąd przetwarzania maila: {str(e)}")

        mail.logout()

    except imaplib.IMAP4.error as e:
        result["errors"].append(f"IMAP error: {str(e)}")
    except Exception as e:
        result["errors"].append(f"Błąd połączenia: {str(e)}")

    return result


def _import_attachment(data, filename, file_hash, sender, subject, config):
    """Importuje pojedynczy załącznik Excel jako nową paletę"""
    from modules.database import get_db, add_paleta
    from modules.smart_importer import smart_import_excel

    result = {"success": False, "paleta_numer": 0, "products_count": 0, "error": ""}

    tmp_path = None
    try:
        # Zapisz do pliku tymczasowego
        tmp_path = os.path.join(tempfile.gettempdir(), f'mail_import_{file_hash}.xlsx')
        with open(tmp_path, 'wb') as f:
            f.write(data)

        # Następny numer palety
        next_num = _get_next_paleta_number()

        # Dostawca z konfiguracji lub auto-detect
        dostawca = config.get('default_dostawca', 'Warrington')

        # Tymczasowa nazwa — zostanie zaktualizowana po imporcie przez Gemini
        paleta_nazwa = f"#{next_num} Import"

        # Utwórz paletę
        paleta_id = add_paleta(
            nazwa=paleta_nazwa,
            dostawca=dostawca,
            cena_zakupu=0,  # Będzie uzupełnione przez smart_import
            data_zakupu=datetime.now().strftime('%Y-%m-%d'),
            notatki=f"Auto-import z maila: {subject}"
        )

        if not paleta_id:
            result["error"] = "Nie udało się utworzyć palety"
            _log_import(file_hash, filename, None, 0, sender, subject, 'error', result["error"])
            return result

        # === DEBUG: sprawdź plik przed importem ===
        debug_info = []
        try:
            import openpyxl
            wb_check = openpyxl.load_workbook(tmp_path, data_only=True)
            ws_check = wb_check.active
            debug_info.append(f"Arkusz: '{ws_check.title}', wierszy: {ws_check.max_row}, kolumn: {ws_check.max_column}")
            # Pokaż pierwsze 3 wiersze
            for r_idx in range(1, min(4, ws_check.max_row + 1)):
                row_vals = [str(cell.value or '')[:30] for cell in ws_check[r_idx]]
                debug_info.append(f"Wiersz {r_idx}: {row_vals}")
            wb_check.close()
        except Exception as dbg_e:
            debug_info.append(f"Debug read error: {dbg_e}")

        # Smart import — przepuść przez istniejący parser
        logging.warning(f"[MailImport] Uruchamiam smart_import_excel: {tmp_path}, filename={filename}, paleta_id={paleta_id}, vendor={dostawca}")
        import_result = smart_import_excel(
            file_path=tmp_path,
            filename=filename,
            paleta_id=paleta_id,
            manual_vendor=dostawca
        )

        products_count = import_result.get('products_imported', 0)
        debug_info.append(f"smart_import: products={products_count}, success={import_result.get('success')}")
        if import_result.get('errors'):
            debug_info.append(f"smart_errors: {import_result['errors']}")
        if import_result.get('details'):
            for d in import_result['details']:
                debug_info.append(f"smart_detail: {d}")

        # Jeśli smart_import nie zadziałał, spróbuj bezpośrednio import_excel_manifest
        if products_count == 0:
            from modules.inventory_utils import import_excel_manifest
            direct_result = import_excel_manifest(
                file_path=tmp_path,
                dostawca=dostawca,
                paleta_id=paleta_id,
                force_insert=True  # Zawsze wstaw na nową paletę
            )
            products_count = direct_result.get('added', 0) + direct_result.get('updated', 0)
            debug_info.append(f"direct_import: added={direct_result.get('added')}, updated={direct_result.get('updated')}")
            if direct_result.get('errors'):
                debug_info.append(f"direct_errors: {direct_result['errors']}")
            if direct_result.get('details'):
                for d in direct_result['details']:
                    debug_info.append(f"direct_detail: {d}")

        # Pobierz nazwy produktów i wygeneruj nazwę palety przez Gemini
        conn = get_db()
        product_rows = conn.execute(
            'SELECT nazwa FROM produkty WHERE paleta_id = ?', (paleta_id,)
        ).fetchall()
        product_names = [r['nazwa'] for r in product_rows if r['nazwa']]

        paleta_name = _generate_paleta_name_ai(product_names)
        paleta_nazwa = f"#{next_num} {paleta_name}"

        # Aktualizuj paletę z prawdziwą nazwą i ilością produktów
        conn.execute(
            'UPDATE palety SET nazwa = ?, ilosc_produktow = ? WHERE id = ?',
            (paleta_nazwa, products_count, paleta_id)
        )
        conn.commit()

        result["success"] = True
        result["paleta_numer"] = next_num
        result["paleta_id"] = paleta_id
        result["products_count"] = products_count

        # Log z debug info — zawsze zapisuj szczegóły
        error_debug = '\n'.join(debug_info) if debug_info else ''
        _log_import(file_hash, filename, paleta_id, products_count, sender, subject, 'success', error_debug)

        # Telegram alert
        _send_telegram_alert(
            f"📬 <b>AUTO-IMPORT Z MAILA</b>\n\n"
            f"📦 Paleta: <b>{paleta_nazwa}</b>\n"
            f"📋 Produktów: <b>{products_count}</b>\n"
            f"📎 Plik: {filename}\n"
            f"✉️ Od: {sender}\n"
            f"📝 Temat: {subject}\n\n"
            f"⏰ {datetime.now():%H:%M:%S}"
        )

        logging.warning(f"[MailImport] ✅ Zaimportowano: {paleta_nazwa} ({products_count} produktów)")
        for di in debug_info:
            logging.warning(f"[MailImport] DEBUG: {di}")

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        result["error"] = str(e)
        _log_import(file_hash, filename, None, 0, sender, subject, 'error', f"{e}\n{tb}")
        logging.warning(f"[MailImport] ❌ Błąd: {e}\n{tb}")

    finally:
        # Nie usuwaj pliku jeśli 0 produktów (do debugowania)
        if tmp_path and os.path.exists(tmp_path):
            if result.get("products_count", 0) > 0:
                try:
                    os.remove(tmp_path)
                except:
                    pass
            else:
                logging.warning(f"[MailImport] Plik zachowany do debugowania: {tmp_path}")

    return result


def _generate_paleta_name_ai(product_names):
    """Generuje krótką nazwę palety przez Gemini AI na podstawie listy produktów.

    Styl: 'Myszki i Klawiatury', 'Adaptery Bluetooth', 'Fotele Skórzane'
    """
    if not product_names:
        return "Mix"

    try:
        from google import genai
        from gemini_config import GEMINI_API_KEY

        if not GEMINI_API_KEY or GEMINI_API_KEY == 'WKLEJ_TUTAJ_SWOJ_KLUCZ':
            return _fallback_paleta_name(product_names)

        client = genai.Client(api_key=GEMINI_API_KEY)

        # Max 20 produktów do promptu
        sample = product_names[:20]
        products_text = '\n'.join(f'- {name}' for name in sample)

        prompt = f"""Na podstawie listy produktów z palety, wygeneruj KRÓTKĄ nazwę palety (2-3 słowa po polsku).

PRODUKTY:
{products_text}

ZASADY:
1. Max 3 słowa, po polsku
2. Opisz główną kategorię produktów
3. Jeśli mix — wymień 2 główne typy
4. Styl: "Myszki i Klawiatury", "Adaptery Bluetooth", "Fotele Skórzane", "Drukarki 3D", "Poduszki Ortopedyczne"
5. Bez numerów, bez "#", bez cudzysłowów

PRZYKŁADY:
- 10x mysz bezprzewodowa, 5x klawiatura → "Myszki i Klawiatury"
- 20x adapter bluetooth → "Adaptery Bluetooth"
- 5x fotel skórzany, 3x krzesło biurowe → "Fotele i Krzesła"

Wygeneruj TYLKO nazwę, bez komentarzy:"""

        response = client.models.generate_content(
            model='gemini-2.0-flash',
            contents=prompt
        )

        if hasattr(response, 'text') and response.text:
            name = response.text.strip().strip('"\'').strip()
            # Cleanup
            name = re.sub(r'^#\d+\s*', '', name)  # Usuń ewentualny #XX
            if 3 <= len(name) <= 50:
                logging.warning(f"[MailImport] 🤖 Gemini nazwa palety: {name}")
                return name

    except Exception as e:
        logging.warning(f"[MailImport] Gemini name error: {e}")

    return _fallback_paleta_name(product_names)


def _fallback_paleta_name(product_names):
    """Fallback — generuje nazwę bez AI, bierze najczęstsze słowo kluczowe"""
    if not product_names:
        return "Mix"

    # Policz najczęstsze słowa (>3 znaki) z nazw produktów
    word_count = {}
    skip_words = {'the', 'and', 'for', 'with', 'from', 'pack', 'set', 'pcs', 'szt',
                  'new', 'pro', 'max', 'mini', 'ultra', 'plus', 'edition', 'version'}
    for name in product_names:
        words = re.findall(r'[a-zA-ZąćęłńóśźżĄĆĘŁŃÓŚŹŻ]{4,}', name)
        for w in words:
            w_lower = w.lower()
            if w_lower not in skip_words:
                word_count[w_lower] = word_count.get(w_lower, 0) + 1

    if word_count:
        top = sorted(word_count.items(), key=lambda x: -x[1])[:2]
        return ' '.join(w.capitalize() for w, _ in top)

    return "Mix"


# ============================================================
# SCHEDULER - BACKGROUND THREAD
# ============================================================

_scheduler_thread = None
_scheduler_running = False


def start_mail_import_scheduler():
    """Uruchamia scheduler sprawdzający pocztę co X minut"""
    global _scheduler_thread, _scheduler_running

    if _scheduler_running:
        return

    config = _get_config()
    if not config['enabled']:
        print("[MailImport] Scheduler wyłączony (mail_import_enabled=0)")
        return

    _scheduler_running = True
    _scheduler_thread = threading.Thread(target=_scheduler_loop, daemon=True)
    _scheduler_thread.start()
    logging.warning(f"[MailImport] ✅ Scheduler uruchomiony (co {config['check_interval']} min)")


def stop_mail_import_scheduler():
    """Zatrzymuje scheduler"""
    global _scheduler_running
    _scheduler_running = False
    print("[MailImport] Scheduler zatrzymany")


def _scheduler_loop():
    """Główna pętla schedulera"""
    global _scheduler_running

    # Poczekaj 60s na start systemu
    time.sleep(60)

    while _scheduler_running:
        try:
            config = _get_config()
            if not config['enabled']:
                _scheduler_running = False
                break

            logging.warning(f"[MailImport] Sprawdzam pocztę... ({datetime.now():%H:%M})")
            result = check_mailbox()

            if result['imported'] > 0:
                logging.warning(f"[MailImport] Zaimportowano {result['imported']} palet")
            if result['errors']:
                for err in result['errors']:
                    logging.warning(f"[MailImport] ⚠️ {err}")

            # Czekaj X minut
            interval = config.get('check_interval', 15) * 60
            time.sleep(interval)

        except Exception as e:
            logging.warning(f"[MailImport] Scheduler error: {e}")
            time.sleep(300)  # 5 min przy błędzie


# ============================================================
# ROUTES - KONFIGURACJA I MANUAL TRIGGER
# ============================================================

@mail_import_bp.route('/ustawienia/mail-import')
def mail_import_config():
    """Strona konfiguracji auto-importu z maili"""
    from modules.database import get_db
    from modules.shared import CSS

    config = _get_config()

    # Historia importów
    conn = get_db()
    logs = conn.execute('''
        SELECT * FROM mail_import_log
        ORDER BY data_importu DESC LIMIT 20
    ''').fetchall()

    logs_html = ''
    for log in logs:
        status_icon = '✅' if log['status'] == 'success' and log['products_count'] > 0 else '⚠️' if log['status'] == 'success' else '❌'
        error_text = log['error'] or ''
        # Pokaż debug info jeśli jest
        error_html = ''
        if error_text:
            error_color = '#ef4444' if log['status'] == 'error' else '#f59e0b'
            error_html = f'''<details style="margin-top:4px">
                <summary style="color:{error_color};cursor:pointer">Szczegóły</summary>
                <pre style="color:{error_color};font-size:0.75rem;white-space:pre-wrap;margin-top:4px">{error_text}</pre>
            </details>'''
        logs_html += f'''
        <div style="padding:10px;border-bottom:1px solid #2a2a3a;font-size:0.85rem">
            <div>{status_icon} <b>{log['filename']}</b></div>
            <div style="color:#64748b;margin-top:4px">
                Paleta ID: {log['paleta_id'] or '-'} |
                Produktów: {log['products_count']} |
                {log['data_importu']}
            </div>
            {error_html}
        </div>
        '''

    if not logs:
        logs_html = '<div style="padding:20px;text-align:center;color:#64748b">Brak importów</div>'

    flash_html = _render_flash_messages()

    html = CSS + f'''
    <div class="container">
        <div class="header">
            <h1>📬 AUTO-IMPORT Z MAILI</h1>
            <small>Automatyczne tworzenie palet z załączników Excel</small>
        </div>

        {flash_html}

        <form action="/ustawienia/mail-import/save" method="POST">
            <div class="card" style="padding:15px">
                <div style="font-weight:600;margin-bottom:15px">📧 Konfiguracja IMAP</div>

                <div style="display:flex;align-items:center;gap:10px;margin-bottom:15px">
                    <label style="font-size:0.9rem">Włączony:</label>
                    <input type="checkbox" name="enabled" value="1"
                           {'checked' if config['enabled'] else ''}
                           style="width:20px;height:20px">
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Serwer IMAP</label>
                    <input type="text" name="imap_server" value="{config['imap_server']}"
                           placeholder="imap.gmail.com"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Port</label>
                    <input type="number" name="imap_port" value="{config['imap_port']}"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Twój email</label>
                    <input type="email" name="email" value="{config['email']}"
                           placeholder="twoj@gmail.com"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Hasło (App Password dla Gmail)</label>
                    <input type="password" name="password" value="{config['password']}"
                           placeholder="xxxx xxxx xxxx xxxx"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                    <div style="font-size:0.75rem;color:#64748b;margin-top:4px">
                        Gmail: Konto → Bezpieczeństwo → Hasła do aplikacji
                    </div>
                </div>
            </div>

            <div class="card" style="padding:15px;margin-top:12px">
                <div style="font-weight:600;margin-bottom:15px">🎯 Filtr nadawcy</div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Email nadawcy (od kogo szukać maili z Excelem)</label>
                    <input type="email" name="sender_filter" value="{config['sender_filter']}"
                           placeholder="dziadek@email.com"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Domyślny dostawca</label>
                    <select name="default_dostawca" class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                        <option value="Warrington" {'selected' if config['default_dostawca'] == 'Warrington' else ''}>Warrington</option>
                        <option value="Jobalots" {'selected' if config['default_dostawca'] == 'Jobalots' else ''}>Jobalots</option>
                        <option value="Miglo" {'selected' if config['default_dostawca'] == 'Miglo' else ''}>Miglo</option>
                    </select>
                </div>

                <div style="margin-bottom:10px">
                    <label style="font-size:0.85rem;color:#94a3b8">Sprawdzaj co (minuty)</label>
                    <input type="number" name="check_interval" value="{config['check_interval']}" min="5" max="120"
                           class="form-ctrl" style="padding:10px;width:100%;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:8px;color:#fff">
                </div>
            </div>

            <div style="display:flex;gap:10px;margin-top:15px">
                <button type="submit" class="btn" style="flex:1;padding:12px;background:#6366f1;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600">
                    💾 Zapisz konfigurację
                </button>
            </div>
        </form>

        <div style="display:flex;gap:10px;margin-top:12px">
            <a href="/ustawienia/mail-import/check" class="btn"
               style="flex:1;padding:12px;background:#22c55e;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;text-align:center;text-decoration:none;display:block">
                📬 Sprawdź pocztę teraz
            </a>
            <a href="/ustawienia/mail-import/test" class="btn"
               style="flex:1;padding:12px;background:#3b82f6;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;text-align:center;text-decoration:none;display:block">
                🔌 Test połączenia
            </a>
        </div>

        <div class="card" style="padding:15px;margin-top:12px">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:15px">
                <div style="font-weight:600">📋 Historia importów</div>
                <a href="/ustawienia/mail-import/clear-log" onclick="return confirm('Wyczyścić historię? Pliki będą mogły być ponownie zaimportowane.')"
                   style="font-size:0.75rem;padding:4px 10px;background:#ef4444;color:#fff;border-radius:6px;text-decoration:none">🗑️ Wyczyść</a>
            </div>
            {logs_html}
        </div>

        <div class="card" style="padding:15px;margin-top:12px;background:rgba(99,102,241,0.05)">
            <div style="font-weight:600;margin-bottom:10px">💡 Jak to działa?</div>
            <div style="font-size:0.85rem;color:#94a3b8;line-height:1.6">
                1. System sprawdza pocztę co {config['check_interval']} minut<br>
                2. Szuka maili od <b>{config['sender_filter'] or '(nie ustawiony)'}</b> z załącznikiem Excel<br>
                3. Pobiera Excel → sprawdza czy już nie był importowany (hash)<br>
                4. Tworzy paletę z kolejnym numerem (np. #26)<br>
                5. Importuje produkty przez Smart Importer (ten sam co ręczny import)<br>
                6. Wysyła powiadomienie na Telegram
            </div>
        </div>

        <a href="/ustawienia" class="back" style="display:block;margin-top:15px">← Ustawienia</a>
    </div>
    '''

    return html


@mail_import_bp.route('/ustawienia/mail-import/save', methods=['POST'])
def mail_import_save():
    """Zapisuje konfigurację"""
    from modules.database import set_config

    set_config('mail_import_enabled', '1' if request.form.get('enabled') else '0')
    set_config('mail_import_imap_server', request.form.get('imap_server', 'imap.gmail.com'))
    set_config('mail_import_imap_port', request.form.get('imap_port', '993'))
    set_config('mail_import_email', request.form.get('email', ''))
    set_config('mail_import_password', request.form.get('password', ''))
    set_config('mail_import_sender_filter', request.form.get('sender_filter', ''))
    set_config('mail_import_default_dostawca', request.form.get('default_dostawca', 'Warrington'))
    set_config('mail_import_check_interval', request.form.get('check_interval', '15'))

    # Restart scheduler jeśli włączony
    enabled = request.form.get('enabled')
    if enabled:
        stop_mail_import_scheduler()
        start_mail_import_scheduler()
    else:
        stop_mail_import_scheduler()

    flash('Konfiguracja mail import zapisana!', 'success')
    return redirect('/ustawienia/mail-import')


@mail_import_bp.route('/ustawienia/mail-import/test')
def mail_import_test():
    """Test połączenia IMAP"""
    config = _get_config()

    if not config['email'] or not config['password']:
        flash('Najpierw skonfiguruj email i hasło!', 'error')
        return redirect('/ustawienia/mail-import')

    try:
        mail = imaplib.IMAP4_SSL(config['imap_server'], config['imap_port'])
        mail.login(config['email'], config['password'])

        # Policz maile od nadawcy
        mail.select('INBOX')
        count = 0
        if config['sender_filter']:
            since = (datetime.now() - timedelta(days=30)).strftime('%d-%b-%Y')
            status, messages = mail.search(None, f'(FROM "{config["sender_filter"]}" SINCE {since})')
            if status == 'OK':
                count = len(messages[0].split()) if messages[0] else 0

        mail.logout()
        flash(f'✅ Połączenie OK! Znaleziono {count} maili od {config["sender_filter"]} (ostatnie 30 dni)', 'success')

    except imaplib.IMAP4.error as e:
        flash(f'❌ Błąd IMAP: {str(e)}', 'error')
    except Exception as e:
        flash(f'❌ Błąd: {str(e)}', 'error')

    return redirect('/ustawienia/mail-import')


@mail_import_bp.route('/ustawienia/mail-import/check')
def mail_import_check_now():
    """Ręczne sprawdzenie poczty"""
    try:
        result = check_mailbox(manual=True)
    except Exception as e:
        flash(f"❌ CRASH: {str(e)}", 'error')
        return redirect('/ustawienia/mail-import')

    # Debug info w flash
    debug_parts = [f"checked={result['checked']}", f"imported={result['imported']}", f"skipped={result['skipped']}"]
    if result.get('details'):
        debug_parts.extend(result['details'])
    if result.get('errors'):
        debug_parts.extend(result['errors'])
    debug_msg = ' | '.join(debug_parts)

    if result['imported'] > 0:
        flash(f"✅ Zaimportowano {result['imported']} palet! (pominięto: {result['skipped']}) | {debug_msg}", 'success')
    elif result['skipped'] > 0:
        flash(f"⏭️ Wszystkie pliki już zaimportowane ({result['skipped']} pominięto) | {debug_msg}", 'info')
    elif result['errors']:
        flash(f"❌ {debug_msg}", 'error')
    else:
        flash(f"📭 Brak nowych maili z Excelem | {debug_msg}", 'info')

    return redirect('/ustawienia/mail-import')


@mail_import_bp.route('/ustawienia/mail-import/clear-log')
def mail_import_clear_log():
    """Czyści historię importów (pozwala ponownie zaimportować pliki)"""
    from modules.database import get_db
    conn = get_db()
    conn.execute('DELETE FROM mail_import_log')
    conn.commit()
    flash('🗑️ Historia importów wyczyszczona — pliki mogą być ponownie zaimportowane', 'success')
    return redirect('/ustawienia/mail-import')


@mail_import_bp.route('/api/mail-import/status')
def mail_import_status():
    """API: status schedulera i ostatni import"""
    from modules.database import get_db

    config = _get_config()
    conn = get_db()
    last = conn.execute(
        'SELECT * FROM mail_import_log ORDER BY data_importu DESC LIMIT 1'
    ).fetchone()

    return jsonify({
        "enabled": config['enabled'],
        "scheduler_running": _scheduler_running,
        "sender_filter": config['sender_filter'],
        "check_interval": config['check_interval'],
        "last_import": {
            "filename": last['filename'] if last else None,
            "status": last['status'] if last else None,
            "date": last['data_importu'] if last else None,
            "products": last['products_count'] if last else 0,
        } if last else None
    })
