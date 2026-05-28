"""
Inventory Utils - Inteligentny parser importu Excel i zarządzanie stanami
=========================================================================

Moduł zawiera:
1. SmartQuantityParser - inteligentne wykrywanie i parsowanie ilości z Excel
2. Funkcje aktualizacji stanów magazynowych po sprzedaży
3. Ulepszony import manifestów z różnych formatów dostawców
"""

import re
import os
import tempfile
from datetime import datetime
from typing import Optional, Tuple, List, Dict, Any, Union
from dataclasses import dataclass

from .database import get_db, execute_db, query_db


# ============================================================
# INTELIGENTNY PARSER ILOŚCI
# ============================================================

@dataclass
class QuantityParseResult:
    """Wynik parsowania ilości"""
    value: int
    original: str
    confidence: float  # 0.0 - 1.0
    method: str  # metoda użyta do parsowania


class SmartQuantityParser:
    """
    Inteligentny parser ilości z różnych formatów Excel.
    
    Obsługuje formaty:
    - "5 szt."
    - "5szt"
    - "5 sztuk"
    - "5 pcs"
    - "5x"
    - "x5"
    - "5.0" (Excel float)
    - "5,0" (European decimal)
    - "5 units"
    - "qty: 5"
    - "Stan: 5"
    - Puste/None -> 1
    """
    
    # Wzorce do usuwania z wartości
    CLEANUP_PATTERNS = [
        r'\s*szt\.?\s*',      # szt. / szt
        r'\s*sztuk[ia]?\s*',  # sztuk / sztuki / sztuka
        r'\s*pcs\.?\s*',      # pcs / pcs.
        r'\s*units?\s*',      # unit / units
        r'\s*pieces?\s*',     # piece / pieces
        r'\s*items?\s*',      # item / items
        r'\s*x\s*$',          # trailing x (np. "5x")
        r'^\s*x\s*',          # leading x (np. "x5")
        r'\s*qty:?\s*',       # qty: / qty
        r'\s*ilość:?\s*',     # ilość:
        r'\s*stan:?\s*',      # stan:
        r'\s*count:?\s*',     # count:
    ]
    
    # Nazwy kolumn z ilością (case insensitive)
    QUANTITY_COLUMN_NAMES = [
        # Polski
        'ilość', 'ilosc', 'ilośc', 'iloć',
        'sztuk', 'sztuki', 'szt',
        'stan', 'stany',
        'na stanie', 'na_stanie',
        'dostępne', 'dostepne',
        'zapas', 'zapasy',
        'magazyn', 'w magazynie',
        
        # Angielski
        'qty', 'quantity', 'quantities',
        'count', 'counts',
        'stock', 'in stock', 'instock',
        'available', 'avail',
        'units', 'unit',
        'pcs', 'pieces',
        'amount', 'amounts',
        'inventory',
        
        # Skróty
        'il', 'il.', 'sz', 'q',
    ]
    
    @classmethod
    def detect_quantity_column(cls, headers: List[str]) -> Optional[int]:
        """
        Wykrywa kolumnę z ilością na podstawie nagłówków.
        
        Args:
            headers: Lista nagłówków kolumn
            
        Returns:
            Indeks kolumny lub None
        """
        if not headers:
            return None
            
        # Normalizuj nagłówki
        normalized = [cls._normalize_header(h) for h in headers]
        
        # Szukaj dopasowania
        for i, header in enumerate(normalized):
            for pattern in cls.QUANTITY_COLUMN_NAMES:
                # Dokładne dopasowanie
                if header == pattern:
                    return i
                # Częściowe dopasowanie (nagłówek zawiera wzorzec)
                if pattern in header and len(pattern) >= 3:
                    return i
                    
        return None
    
    @classmethod
    def _normalize_header(cls, header: Any) -> str:
        """Normalizuje nagłówek do porównania"""
        if header is None:
            return ""
        text = str(header).lower().strip()
        # Usuń znaki specjalne
        text = re.sub(r'[^a-ząćęłńóśźż0-9\s]', '', text)
        # Normalizuj spacje
        text = re.sub(r'\s+', ' ', text)
        return text
    
    @classmethod
    def parse(cls, value: Any) -> QuantityParseResult:
        """
        Parsuje wartość ilości z różnych formatów.
        
        Args:
            value: Wartość do sparsowania (str, int, float, None)
            
        Returns:
            QuantityParseResult z wartością int
        """
        original = str(value) if value is not None else ""
        
        # None, puste, N/A -> 1
        if value is None or str(value).strip() == "":
            return QuantityParseResult(1, original, 0.5, "default")
            
        if str(value).upper().strip() in ('N/A', 'NA', 'NONE', '-', '—', '–'):
            return QuantityParseResult(1, original, 0.3, "na_default")
        
        # Konwertuj do stringa
        text = str(value).strip()
        
        # Już jest int/float
        if isinstance(value, (int, float)):
            try:
                result = int(value)
                if result <= 0:
                    result = 1
                return QuantityParseResult(result, original, 1.0, "numeric")
            except (ValueError, OverflowError):
                pass
        
        # Spróbuj bezpośredniej konwersji
        try:
            # Zamień przecinek na kropkę (European decimal)
            clean = text.replace(',', '.')
            result = int(float(clean))
            if result <= 0:
                result = 1
            return QuantityParseResult(result, original, 0.95, "direct")
        except (ValueError, OverflowError):
            pass
        
        # Wyczyść tekst z sufiksów
        cleaned = text
        for pattern in cls.CLEANUP_PATTERNS:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
        cleaned = cleaned.strip()
        
        # Spróbuj ponownie po czyszczeniu
        if cleaned:
            try:
                clean = cleaned.replace(',', '.')
                # Usuń kropki (tysiące w polskim formacie)
                if '.' in clean and clean.count('.') > 1:
                    clean = clean.replace('.', '')
                result = int(float(clean))
                if result <= 0:
                    result = 1
                return QuantityParseResult(result, original, 0.9, "cleaned")
            except (ValueError, OverflowError):
                pass
        
        # Szukaj liczby w tekście
        numbers = re.findall(r'\d+', text)
        if numbers:
            try:
                # Weź pierwszą znalezioną liczbę
                result = int(numbers[0])
                if result <= 0:
                    result = 1
                return QuantityParseResult(result, original, 0.7, "extracted")
            except (ValueError, OverflowError):
                pass
        
        # Fallback - domyślnie 1
        return QuantityParseResult(1, original, 0.2, "fallback")
    
    @classmethod
    def parse_batch(cls, values: List[Any]) -> List[QuantityParseResult]:
        """Parsuje listę wartości"""
        return [cls.parse(v) for v in values]


# ============================================================
# IMPORT EXCEL Z INTELIGENTNYM PARSEREM
# ============================================================

def import_excel_manifest(
    file_path: str = None,
    file_obj = None,
    dostawca: str = "",
    paleta_id: int = None,
    update_existing: bool = True,
    force_insert: bool = False
) -> Dict[str, Any]:
    """
    Importuje manifest produktów z pliku Excel z inteligentnym parserem ilości.
    
    Args:
        file_path: Ścieżka do pliku Excel
        file_obj: Lub obiekt pliku (np. z Flask request.files)
        dostawca: Nazwa dostawcy (auto-wykrywana jeśli puste)
        paleta_id: ID palety do przypisania
        update_existing: Czy aktualizować istniejące produkty
        
    Returns:
        {"success": bool, "added": int, "updated": int, "errors": list, "details": list}
    """
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "added": 0,
            "updated": 0,
            "errors": ["Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl"],
            "details": []
        }
    
    result = {
        "success": False,
        "added": 0,
        "updated": 0,
        "errors": [],
        "details": [],
        "paleta_id": None,  # ID utworzonej palety
        "quantity_stats": {
            "total_parsed": 0,
            "high_confidence": 0,
            "low_confidence": 0,
            "methods": {}
        }
    }
    
    # Załaduj plik
    tmp_path = None
    try:
        if file_obj:
            # Zapisz do pliku tymczasowego
            tmp_path = os.path.join(tempfile.gettempdir(), f'import_{os.getpid()}.xlsx')
            file_obj.save(tmp_path)
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
        ws = wb.active
        
    except Exception as e:
        result["errors"].append(f"Błąd odczytu pliku: {str(e)}")
        return result
    
    try:
        # ========================================
        # 1. ZNAJDŹ WIERSZ Z NAGŁÓWKAMI
        # ========================================
        header_row = 1
        header_keywords = ['EAN', 'ASIN', 'KOD', 'SKU', 'NAZWA', 'NAME', 'PRODUCT', 'TITLE']
        
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = [str(cell.value or '').strip().upper() for cell in ws[row_idx]]
            row_text = ' '.join(row_values)
            
            if any(keyword in row_text for keyword in header_keywords):
                header_row = row_idx
                break
        
        # Pobierz nagłówki
        headers_raw = [cell.value for cell in ws[header_row]]
        headers = [str(h or '').strip() for h in headers_raw]
        headers_upper = [h.upper() for h in headers]
        
        # ========================================
        # 2. AUTO-WYKRYJ DOSTAWCĘ
        # ========================================
        if not dostawca:
            dostawca = _detect_supplier_from_headers(headers)
            if dostawca:
                result["details"].append(f"Auto-wykryto dostawcę: {dostawca}")
        
        # ========================================
        # 3. MAPOWANIE KOLUMN
        # ========================================
        col_map = {
            'ean': None,
            'asin': None,
            'nazwa': None,
            'ilosc': None,
            'cena': None,
            'lokalizacja': None,
            'stan': None,  # Stan fizyczny (nowy/używany)
        }
        
        # Mapowanie nazw kolumn -> klucz
        column_mappings = {
            'ean': ['EAN', 'KOD', 'BARCODE', 'KOD KRESKOWY', 'CODE', 'SKU', 'PRODUCT CODE', 'KOD 2'],
            'asin': ['ASIN', 'AMAZON ID', 'AMAZON'],
            'nazwa': ['NAZWA', 'NAME', 'TITLE', 'TYTUŁ', 'PRODUCT', 'OPIS', 'DESCRIPTION', 'PRODUKT'],
            'ilosc': SmartQuantityParser.QUANTITY_COLUMN_NAMES,  # Użyj rozszerzonej listy
            # UWAGA: cena będzie wykrywana osobno - potrzebujemy ceny ZAKUPU, nie rynkowej!
            'cena': [],  # Wykrywane osobno
            'lokalizacja': ['LOKALIZACJA', 'LOCATION', 'LOC', 'MIEJSCE', 'REGAŁ', 'REGAL', 'SHELF', 'POSITION'],
            'stan': ['STAN', 'KONDYCJA', 'CONDITION', 'STATUS', 'GRADE', 'QUALITY'],
        }
        
        for i, h in enumerate(headers_upper):
            h_clean = re.sub(r'[^A-ZĄĆĘŁŃÓŚŹŻ0-9\s]', '', h).strip()
            
            for key, patterns in column_mappings.items():
                if col_map[key] is None:
                    for pattern in patterns:
                        pattern_upper = pattern.upper()
                        if pattern_upper in h_clean or h_clean in pattern_upper:
                            col_map[key] = i
                            break
        
        # ========================================
        # INTELIGENTNE WYKRYWANIE KOLUMNY CENY ZAKUPU (per sztuka!)
        # ========================================
        # Szukamy kolumny z ceną JEDNOSTKOWĄ zakupu, nie łączną!
        # Priorytet: "JEDNOSTKOWA" > "SPRZEDAŻY NETTO" > "UNIT COST" > inne
        # Unikamy: "RYNKOWA", "REGULARNA", "RRP", "BRUTTO" (to ceny rynkowe)
        
        cena_zakupu_priority = [
            # Najwyższy priorytet - cena jednostkowa
            ('CENA JEDNOSTKOWA SPRZEDAŻY', 110),
            ('CENA JEDNOSTKOWA SPRZEDAZY', 110),
            ('UNIT COST', 105),
            ('CENA JEDNOSTKOWA', 100),
            ('UNIT PRICE', 95),
            # Ceny które mogą być jednostkowe
            ('CENA SPRZEDAŻY NETTO', 90),
            ('CENA SPRZEDAZY NETTO', 90),
            ('CENA ZAKUPU', 85),
            ('PURCHASE PRICE', 85),
            ('COST', 70),
            ('KOSZT', 70),
            ('NETTO', 60),
            # Najniższy - może być łączna
            ('CENA SPRZEDAŻY', 50),
            ('CENA SPRZEDAZY', 50),
        ]
        
        # Kolumny do UNIKANIA (ceny rynkowe)
        cena_unikaj = ['RYNKOWA', 'REGULARNA', 'RRP', 'RETAIL', 'MSRP', 'LIST PRICE']
        
        best_cena_col = None
        best_cena_score = 0
        
        for i, h in enumerate(headers):
            h_upper = h.upper().strip()
            
            # Sprawdź czy to kolumna z ceną
            if not any(x in h_upper for x in ['CENA', 'PRICE', 'COST', 'KOSZT', 'NETTO']):
                continue
            
            # Sprawdź czy to cena do UNIKANIA (rynkowa)
            if any(x in h_upper for x in cena_unikaj):
                continue
            
            # Oblicz score
            score = 10  # Bazowy score dla każdej kolumny z "cena"
            
            for pattern, priority in cena_zakupu_priority:
                if pattern in h_upper:
                    score = max(score, priority)
                    break
            
            if score > best_cena_score:
                best_cena_score = score
                best_cena_col = i
        
        if best_cena_col is not None:
            col_map['cena'] = best_cena_col
            result["details"].append(f"Wykryto kolumnę ceny zakupu: '{headers[best_cena_col]}' (score: {best_cena_score})")
        else:
            # Fallback - szukaj czegokolwiek z "cena" lub "price"
            for i, h in enumerate(headers_upper):
                if any(x in h for x in ['CENA', 'PRICE', 'COST']):
                    col_map['cena'] = i
                    result["details"].append(f"Fallback kolumna ceny: '{headers[i]}'")
                    break
        
        # Inteligentne wykrywanie kolumny ilości (fallback)
        if col_map['ilosc'] is None:
            detected_qty_col = SmartQuantityParser.detect_quantity_column(headers)
            if detected_qty_col is not None:
                col_map['ilosc'] = detected_qty_col
                result["details"].append(f"Wykryto kolumnę ilości: '{headers[detected_qty_col]}'")
        
        result["details"].append(f"Mapowanie kolumn: {col_map}")
        
        # ========================================
        # 4. AUTOMATYCZNE TWORZENIE PALETY (jeśli nie podano)
        # ========================================
        if paleta_id is None and dostawca:
            # Stwórz nową paletę automatycznie
            from datetime import datetime
            dzis = datetime.now().strftime('%Y-%m-%d')
            nazwa_palety = f"{dostawca} {dzis}"
            
            conn_temp = get_db()
            cursor = conn_temp.execute('''
                INSERT INTO palety (nazwa, dostawca, data_zakupu, ilosc_produktow)
                VALUES (?, ?, ?, 0)
            ''', (nazwa_palety, dostawca, dzis))
            paleta_id = cursor.lastrowid
            conn_temp.commit()

            result["paleta_id"] = paleta_id  # Zapisz ID palety w wyniku
            result["details"].append(f"<span class=material-symbols-outlined style=color:#22c55e>check_circle</span> Utworzono paletę: {nazwa_palety} (ID: {paleta_id})")
        elif paleta_id:
            result["paleta_id"] = paleta_id  # Użyto istniejącej palety
        
        # ========================================
        # 5. IMPORTUJ PRODUKTY
        # ========================================
        conn = get_db()
        
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
                
            try:
                # --- EAN ---
                ean_idx = col_map.get('ean')
                ean = ""
                if ean_idx is not None and ean_idx < len(row):
                    ean = str(row[ean_idx] or '').strip()
                    # Wyczyść z .0 (Excel float)
                    if ean.endswith('.0'):
                        ean = ean[:-2]
                    ean = ean.replace('.0', '').replace(' ', '')
                    
                # --- ASIN ---
                asin_idx = col_map.get('asin')
                asin = ""
                if asin_idx is not None and asin_idx < len(row):
                    asin = str(row[asin_idx] or '').strip()
                    if asin.endswith('.0'):
                        asin = asin[:-2]
                
                # Pomiń wiersze bez kodu
                if not ean and not asin:
                    continue
                if ean.upper() in ('NONE', 'NAN', 'N/A', ''):
                    if not asin or asin.upper() in ('NONE', 'NAN', 'N/A'):
                        continue
                    ean = asin  # Użyj ASIN jako identyfikatora
                    
                # --- NAZWA ---
                nazwa_idx = col_map.get('nazwa')
                nazwa = ean  # Domyślnie kod
                if nazwa_idx is not None and nazwa_idx < len(row):
                    nazwa_val = row[nazwa_idx]
                    if nazwa_val and str(nazwa_val).strip().upper() not in ('NONE', 'NAN', 'N/A'):
                        nazwa = str(nazwa_val).strip()[:500]  # Ogranicz długość
                
                # --- ILOŚĆ (inteligentny parser) ---
                ilosc = 1
                ilosc_idx = col_map.get('ilosc')
                if ilosc_idx is not None and ilosc_idx < len(row):
                    qty_result = SmartQuantityParser.parse(row[ilosc_idx])
                    ilosc = qty_result.value
                    
                    # Statystyki
                    result["quantity_stats"]["total_parsed"] += 1
                    if qty_result.confidence >= 0.8:
                        result["quantity_stats"]["high_confidence"] += 1
                    else:
                        result["quantity_stats"]["low_confidence"] += 1
                    
                    method = qty_result.method
                    result["quantity_stats"]["methods"][method] = \
                        result["quantity_stats"]["methods"].get(method, 0) + 1
                        
                    # Log dla niskiej pewności
                    if qty_result.confidence < 0.5:
                        result["details"].append(
                            f"<span class=material-symbols-outlined>warning</span> Niska pewność ilości dla '{ean}': "
                            f"'{qty_result.original}' -> {ilosc} (metoda: {method})"
                        )
                
                # --- CENA (netto z Excela, brutto = netto × 1.23) ---
                cena_netto = 0.0
                cena_brutto = 0.0
                cena_idx = col_map.get('cena')
                if cena_idx is not None and cena_idx < len(row):
                    cena_val = row[cena_idx]
                    if cena_val:
                        try:
                            cena_netto = float(str(cena_val).replace(',', '.').replace(' ', '').replace('zł', ''))
                            cena_brutto = round(cena_netto * 1.23, 2)  # VAT 23%
                        except (ValueError, TypeError):
                            pass
                
                # --- LOKALIZACJA ---
                lokalizacja = ""
                lok_idx = col_map.get('lokalizacja')
                if lok_idx is not None and lok_idx < len(row):
                    lok_val = row[lok_idx]
                    if lok_val and str(lok_val).strip().upper() not in ('NONE', 'NAN', 'N/A'):
                        lokalizacja = str(lok_val).strip()[:50]
                
                # --- STAN FIZYCZNY ---
                stan = "Nowy"
                stan_idx = col_map.get('stan')
                if stan_idx is not None and stan_idx < len(row):
                    stan_val = str(row[stan_idx] or '').strip().lower()
                    if any(x in stan_val for x in ['używany', 'uzywany', 'used', 'refurb', 'b-stock']):
                        stan = "Używany"
                    elif any(x in stan_val for x in ['uszkodzony', 'damaged', 'broken']):
                        stan = "Uszkodzony"
                
                # --- ZAPIS DO BAZY ---
                if force_insert:
                    # Tryb mail-import: zawsze wstaw nowy (na nową paletę)
                    conn.execute('''
                        INSERT INTO produkty
                        (ean, asin, nazwa, ilosc, cena_netto, cena_brutto, lokalizacja, stan, dostawca, paleta_id, data_dodania)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (ean, asin, nazwa, ilosc, cena_netto, cena_brutto, lokalizacja, stan, dostawca, paleta_id, datetime.now().isoformat()))
                    result["added"] += 1
                else:
                    existing = conn.execute(
                        'SELECT id, ilosc FROM produkty WHERE ean = ? OR (asin = ? AND asin != "")',
                        (ean, asin)
                    ).fetchone()

                    if existing:
                        if update_existing:
                            # Aktualizuj istniejący (dodaj ilość)
                            new_qty = existing['ilosc'] + ilosc
                            conn.execute('''
                                UPDATE produkty SET
                                    nazwa = COALESCE(NULLIF(?, ''), nazwa),
                                    ilosc = ?,
                                    cena_netto = CASE WHEN ? > 0 THEN ? ELSE cena_netto END,
                                    cena_brutto = CASE WHEN ? > 0 THEN ? ELSE cena_brutto END,
                                    lokalizacja = COALESCE(NULLIF(?, ''), lokalizacja),
                                    stan = ?,
                                    dostawca = COALESCE(NULLIF(?, ''), dostawca),
                                    paleta_id = COALESCE(?, paleta_id)
                                WHERE id = ?
                            ''', (nazwa, new_qty, cena_netto, cena_netto, cena_brutto, cena_brutto,
                                  lokalizacja, stan, dostawca, paleta_id, existing['id']))
                            result["updated"] += 1
                    else:
                        # Wstaw nowy
                        conn.execute('''
                            INSERT INTO produkty
                            (ean, asin, nazwa, ilosc, cena_netto, cena_brutto, lokalizacja, stan, dostawca, paleta_id, data_dodania)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (ean, asin, nazwa, ilosc, cena_netto, cena_brutto, lokalizacja, stan, dostawca, paleta_id, datetime.now().isoformat()))
                        result["added"] += 1
                    
            except Exception as e:
                result["errors"].append(f"Błąd wiersza (EAN: {ean}): {str(e)}")
        
        # Aktualizuj liczbę produktów w palecie
        if paleta_id:
            count = conn.execute('SELECT COUNT(*) FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchone()[0]
            conn.execute('UPDATE palety SET ilosc_produktow = ? WHERE id = ?', (count, paleta_id))
            result["details"].append(f"<span class=material-symbols-outlined>inventory_2</span> Zaktualizowano paletę: {count} produktów")
                
        conn.commit()

        result["success"] = True
        result["details"].append(f"Zaimportowano: {result['added']} nowych, {result['updated']} zaktualizowanych")
        
    except Exception as e:
        result["errors"].append(f"Błąd importu: {str(e)}")
        import traceback
        result["errors"].append(traceback.format_exc())
        
    finally:
        # Cleanup
        try:
            wb.close()
        except:
            pass
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass
                
    return result


def _detect_supplier_from_headers(headers: List[str]) -> str:
    """Wykrywa dostawcę na podstawie charakterystycznych nagłówków"""
    text = ' '.join(headers).upper()
    
    supplier_patterns = {
        'Jobalots': ['JOBALOTS', 'JOB LOTS'],
        'Warrington': ['WARRINGTON', 'WARR'],
        'Miglo': ['MIGLO', 'MGL'],
        'Amazon Returns': ['AMAZON', 'AMZN', 'LPN'],
        'Customer Returns': ['CUSTOMER RETURN', 'RETURNS'],
        'Lidl': ['LIDL'],
        'Kaufland': ['KAUFLAND'],
        'MediaMarkt': ['MEDIA MARKT', 'MEDIAMARKT'],
    }
    
    for supplier, patterns in supplier_patterns.items():
        if any(p in text for p in patterns):
            return supplier
            
    return ""


# ============================================================
# AKTUALIZACJA STANÓW MAGAZYNOWYCH PO SPRZEDAŻY
# ============================================================

def update_stock_on_sale(
    allegro_offer_id: str = None,
    ean: str = None,
    asin: str = None,
    product_id: int = None,
    quantity_sold: int = 1
) -> Dict[str, Any]:
    """
    Aktualizuje stan magazynowy po sprzedaży.
    
    Args:
        allegro_offer_id: ID oferty Allegro
        ean: Kod EAN produktu
        asin: Kod ASIN produktu
        product_id: ID produktu w bazie
        quantity_sold: Ilość sprzedana (domyślnie 1)
        
    Returns:
        {"success": bool, "product_id": int, "old_stock": int, "new_stock": int, "message": str}
    """
    result = {
        "success": False,
        "product_id": None,
        "old_stock": 0,
        "new_stock": 0,
        "message": ""
    }
    
    conn = get_db()
    
    try:
        # Znajdź produkt
        product = None
        
        if product_id:
            product = conn.execute('SELECT * FROM produkty WHERE id = ?', (product_id,)).fetchone()
            
        elif allegro_offer_id:
            # Szukaj przez powiązanie z ofertą Allegro
            oferta = conn.execute('SELECT produkt_id FROM oferty WHERE allegro_id = ?', (allegro_offer_id,)).fetchone()
            if oferta and oferta['produkt_id']:
                product = conn.execute('SELECT * FROM produkty WHERE id = ?', (oferta['produkt_id'],)).fetchone()
                
        elif ean:
            product = conn.execute('SELECT * FROM produkty WHERE ean = ?', (ean,)).fetchone()
            
        elif asin:
            product = conn.execute('SELECT * FROM produkty WHERE asin = ?', (asin,)).fetchone()
        
        if not product:
            result["message"] = "Nie znaleziono produktu w bazie"
            return result

        # Jesli oryginalny produkt jest pusty, sprobuj alternatywnego z tym samym
        # ASIN/EAN i ilosc > 0 (FIFO). Inaczej stan z innych palet by nie schodzil.
        if (product['ilosc'] or 0) <= 0:
            _alt = None
            _asin = product['asin'] if 'asin' in product.keys() else ''
            _ean = product['ean'] if 'ean' in product.keys() else ''
            if _asin and len(_asin) >= 5:
                _alt = conn.execute(
                    "SELECT * FROM produkty WHERE asin = ? AND ilosc > 0 AND id != ? ORDER BY data_dodania ASC LIMIT 1",
                    (_asin, product['id'])
                ).fetchone()
            if not _alt and _ean and len(_ean) >= 8:
                _alt = conn.execute(
                    "SELECT * FROM produkty WHERE ean = ? AND ilosc > 0 AND id != ? ORDER BY data_dodania ASC LIMIT 1",
                    (_ean, product['id'])
                ).fetchone()
            if _alt:
                print(f"[INVE] Swap: produkt {product['id']} pusty -> uzywam {_alt['id']} (stan {_alt['ilosc']})")
                product = _alt
                # Przepnij oferte na alternatywny produkt zeby kolejne sprzedaze trafialy w wlasciwe
                if allegro_offer_id:
                    conn.execute('UPDATE oferty SET produkt_id = ? WHERE allegro_id = ?',
                               (product['id'], allegro_offer_id))

        result["product_id"] = product['id']
        result["old_stock"] = product['ilosc']

        # Oblicz nowy stan
        new_stock = max(0, product['ilosc'] - quantity_sold)
        result["new_stock"] = new_stock

        # RETRY: 4 proby exp backoff 1s/2s/4s/8s gdy 'database is locked'
        # (auto-sync Allegro / inny watek moze trzymac WRITE lock)
        import sqlite3 as _sqlite3
        import time as _time
        _ok = False
        for _att in range(4):
            try:
                # Aktualizuj w bazie
                conn.execute('''
                    UPDATE produkty SET
                        ilosc = ?,
                        status = CASE WHEN ? = 0 THEN 'sprzedany' ELSE status END,
                        data_sprzedazy = CASE WHEN ? = 0 THEN ? ELSE data_sprzedazy END
                    WHERE id = ?
                ''', (new_stock, new_stock, new_stock, datetime.now().isoformat(), product['id']))
                conn.commit()
                _ok = True
                break
            except _sqlite3.OperationalError as _e:
                if 'database is locked' in str(_e).lower() and _att < 3:
                    _w = 2 ** _att  # 1, 2, 4, 8
                    print(f"[RETRY] update_stock_on_sale locked produkt {product['id']} (proba {_att+1}/4), sleep {_w}s")
                    _time.sleep(_w)
                    continue
                raise

        if not _ok:
            result["message"] = "Blad: database locked po 4 probach"
            return result

        result["success"] = True
        result["message"] = f"Zaktualizowano stan: {result['old_stock']} -> {new_stock}"
        
        # Log
        print(f"[INVE] Stock update: {product['nazwa'][:30]} ({result['old_stock']} -> {new_stock})")
        
    except Exception as e:
        result["message"] = f"Błąd aktualizacji: {str(e)}"
        import traceback
        traceback.print_exc()
        
    finally:
        pass

    return result


def sync_stock_from_allegro_order(order_data: dict) -> List[Dict[str, Any]]:
    """
    Synchronizuje stany magazynowe na podstawie zamówienia Allegro.
    
    Args:
        order_data: Dane zamówienia z API Allegro (checkout form)
        
    Returns:
        Lista wyników aktualizacji dla każdego produktu
    """
    results = []
    
    line_items = order_data.get('lineItems', [])
    
    for item in line_items:
        offer = item.get('offer', {})
        offer_id = offer.get('id')
        quantity = item.get('quantity', 1)
        
        if offer_id:
            result = update_stock_on_sale(
                allegro_offer_id=offer_id,
                quantity_sold=quantity
            )
            results.append({
                "offer_id": offer_id,
                "offer_name": offer.get('name', 'Unknown'),
                **result
            })
            
    return results


# ============================================================
# ROZSZERZONA SYNCHRONIZACJA ZAMÓWIEŃ Z AKTUALIZACJĄ STANÓW
# ============================================================

def sync_orders_with_stock(today_only: bool = True) -> Dict[str, Any]:
    """
    Synchronizuje zamówienia z Allegro i automatycznie aktualizuje stany.
    
    Rozszerza standardową funkcję sync_orders o logikę aktualizacji magazynu.
    
    Returns:
        {"synced": int, "stock_updates": list, "errors": list}
    """
    from .allegro_api import get_orders, allegro_request
    from .telegram_bot import alert_sprzedaz
    from datetime import datetime, date
    
    result = {
        "synced": 0,
        "notified": 0,
        "stock_updates": [],
        "errors": []
    }
    
    # Pobierz zamówienia
    from_date = None
    if today_only:
        from_date = date.today().strftime('%Y-%m-%dT00:00:00Z')
        
    orders_data, error = get_orders('READY_FOR_PROCESSING', from_date=from_date)
    
    if error:
        result["errors"].append(f"Błąd pobierania zamówień: {error}")
        return result
        
    if not orders_data or 'checkoutForms' not in orders_data:
        return result
    
    conn = get_db()
    
    for order in orders_data['checkoutForms']:
        order_id = order['id']
        
        # Sprawdź czy już zsynchronizowane
        existing = conn.execute('SELECT id FROM sprzedaze WHERE allegro_order_id = ?', (order_id,)).fetchone()
        if existing:
            continue
            
        for item in order.get('lineItems', []):
            try:
                offer = item.get('offer', {})
                nazwa = offer.get('name', 'Produkt')[:50]
                cena = float(item['price']['amount'])
                kupujacy = order.get('buyer', {}).get('login', 'Nieznany')
                ilosc = item.get('quantity', 1)
                offer_id = offer.get('id', '')
                
                # Znajdź oferta_id i produkt_id
                oferta_row = conn.execute('SELECT id, produkt_id FROM oferty WHERE allegro_id = ?', (offer_id,)).fetchone()
                oferta_db_id = oferta_row['id'] if oferta_row else None
                produkt_id = oferta_row['produkt_id'] if oferta_row else None

                # Data zamówienia z Allegro
                order_date_raw = order.get('boughtAt') or order.get('updatedAt') or ''
                try:
                    dt_str = order_date_raw.replace('Z', '+00:00')
                    dt = datetime.fromisoformat(dt_str)
                    dt_local = dt.astimezone().replace(tzinfo=None)
                    order_date = dt_local.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    order_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Zapisz sprzedaż
                # OR IGNORE: UNIQUE INDEX na (allegro_order_id, nazwa) blokuje duplikaty
                cur = conn.execute('''
                    INSERT OR IGNORE INTO sprzedaze
                    (allegro_order_id, oferta_id, produkt_id, cena, ilosc, kupujacy, status, data_sprzedazy, nazwa, notified)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                ''', (order_id, oferta_db_id, produkt_id, cena, ilosc, kupujacy, 'nowa', order_date, nazwa))
                if cur.rowcount == 0:
                    continue  # duplikat, juz zapisany przez inny path

                result["synced"] += 1
                
                # ========================================
                # AKTUALIZACJA STANÓW MAGAZYNOWYCH
                # ========================================
                stock_result = update_stock_on_sale(
                    allegro_offer_id=offer_id,
                    quantity_sold=ilosc
                )
                
                result["stock_updates"].append({
                    "product": nazwa,
                    "quantity_sold": ilosc,
                    **stock_result
                })
                
                # Powiadomienie Telegram
                try:
                    alert_sprzedaz(nazwa, cena, kupujacy)
                    result["notified"] += 1
                except Exception as e:
                    result["errors"].append(f"Telegram error: {e}")
                    
            except Exception as e:
                result["errors"].append(f"Błąd przetwarzania zamówienia: {e}")
                
    conn.commit()

    print(f"[OK] Zsync: {result['synced']}, Stock updates: {len(result['stock_updates'])}")
    return result


# ============================================================
# TESTY
# ============================================================

if __name__ == "__main__":
    print("[SCIE] Test SmartQuantityParser")
    
    test_values = [
        "5",
        "5.0",
        "5,0",
        "5 szt.",
        "5szt",
        "5 sztuk",
        "5 pcs",
        "5x",
        "qty: 5",
        "Stan: 5",
        None,
        "",
        "N/A",
        "abc",
        "10 units available",
        "  12  ",
        "5.00 szt.",
    ]
    
    for val in test_values:
        result = SmartQuantityParser.parse(val)
        status = "<span class=material-symbols-outlined style=color:#22c55e>check_circle</span>" if result.confidence >= 0.7 else "<span class=material-symbols-outlined>warning</span>"
        print(f"  {status} '{val}' -> {result.value} (conf: {result.confidence:.2f}, method: {result.method})")
