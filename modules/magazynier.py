"""
Magazynier module - zarządzanie stanami magazynowymi
"""

import os
import io
import csv
import json
import tempfile
from datetime import datetime
from flask import Blueprint, render_template, render_template_string, request, redirect, Response, url_for, session, current_app, jsonify
from flask_wtf.csrf import generate_csrf

from .database import get_db, query_db, execute_db, get_config, get_config_cached
from .utils import get_amazon_image_url, get_product_image, oblicz_cene_allegro, is_code, DOSTAWCY, detect_supplier, parse_price, ALLEGRO_PROWIZJE

magazynier_bp = Blueprint('magazynier', __name__)

# ============================================================
# FUNKCJE POMOCNICZE
# ============================================================
def get_product_code(p):
    """Zwraca kod_magazynowy (MAG-XXXXX) do użycia w URL. Fallback: ID."""
    if hasattr(p, 'keys'):
        p = dict(p)
    kod = p.get('kod_magazynowy', '') or ''
    if kod:
        return kod
    pid = p.get('id')
    if pid:
        return str(pid)
    return ''

def get_produkt_by_code(conn, code):
    """Szukaj produktu po: kod_magazynowy (MAG-xxx), ID (krótki numer), EAN/ASIN."""
    c = str(code).strip()
    # 1. kod_magazynowy (MAG-00123)
    if c.upper().startswith('MAG-'):
        r = conn.execute('SELECT * FROM produkty WHERE kod_magazynowy=?', (c.upper(),)).fetchone()
        if r:
            return r
    # 2. Krótki numer (<=6 cyfr) → szukaj po ID
    if c.isdigit() and len(c) <= 6:
        r = conn.execute('SELECT * FROM produkty WHERE id=?', (int(c),)).fetchone()
        if r:
            return r
    # 3. EAN/ASIN — może się powtarzać, bierzemy z największą ilością
    return conn.execute(
        "SELECT * FROM produkty WHERE (ean=? OR asin=?) AND status != 'sprzedany' ORDER BY ilosc DESC",
        (c, c)).fetchone()


_PLACEHOLDER_IMG_SM = 'data:image/svg+xml,%3Csvg xmlns=%27http://www.w3.org/2000/svg%27 width=%2790%27 height=%2790%27%3E%3Crect fill=%27%23262528%27 width=%2790%27 height=%2790%27 rx=%278%27/%3E%3Ctext x=%2745%27 y=%2752%27 fill=%27%23767577%27 text-anchor=%27middle%27 font-size=%2728%27%3E%F0%9F%93%A6%3C/text%3E%3C/svg%3E'

def _resolve_product_image(p, size='sm'):
    """Rozwiąż URL zdjęcia produktu z pełnym fallback chain.

    Kolejność: zdjecie_url → images JSON → ASIN folder → placeholder.
    size: 'sm' (lista 90px) lub 'lg' (detal).
    """
    img = p.get('zdjecie_url') or '' if hasattr(p, 'get') else (p['zdjecie_url'] if 'zdjecie_url' in p.keys() else '')

    # Waliduj lokalne ścieżki
    if img and img.startswith('/static/downloads/'):
        if not os.path.exists(img.lstrip('/')):
            img = ''

    # Fallback: kolumna images (JSON array lokalnych plików)
    if not img:
        _images_raw = p.get('images', '[]') if hasattr(p, 'get') else (p['images'] if 'images' in p.keys() else '[]')
        if _images_raw:
            try:
                _imgs = json.loads(_images_raw) if isinstance(_images_raw, str) else _images_raw
                if _imgs and len(_imgs) > 0:
                    first_img = _imgs[0]
                    if first_img.startswith('static/'):
                        img = '/' + first_img
                    else:
                        img = first_img
            except Exception:
                pass

    # Fallback: szukaj lokalnego pliku po ASIN
    if not img:
        _asin = p.get('asin', '') if hasattr(p, 'get') else (p['asin'] if 'asin' in p.keys() else '')
        if _asin:
            local_path = f"static/downloads/{_asin}/image_1.jpg"
            if os.path.exists(local_path):
                img = '/' + local_path

    # Ostateczny placeholder
    if not img:
        if size == 'lg':
            img = 'data:image/svg+xml,%3Csvg xmlns=%27http://www.w3.org/2000/svg%27 width=%27400%27 height=%27180%27%3E%3Crect fill=%27%2312121a%27 width=%27400%27 height=%27180%27/%3E%3Ctext x=%27200%27 y=%2795%27 fill=%27%23555%27 text-anchor=%27middle%27 font-size=%2718%27%3EBRAK ZDJECIA%3C/text%3E%3C/svg%3E'
        else:
            img = _PLACEHOLDER_IMG_SM

    return img


def _format_stan_label(stan_przyjecia, klasa_jakosci):
    """Formatuj stan na etykietę: A/Nowy, B/Powystawowy itp."""
    stan = (stan_przyjecia or '').strip()
    klasa = (klasa_jakosci or '').strip().upper()
    if not stan or stan == 'nieoceniony':
        return 'NIEOCENIONY'
    if klasa:
        return f"{klasa}/{stan}"
    return stan


def _paleta_koszt_szt(conn, paleta_id):
    """Koszt brutto/szt = paleta.cena_zakupu (brutto) / łączna ilość sztuk."""
    if not paleta_id:
        return 0
    row = conn.execute('''
        SELECT pal.cena_zakupu,
            COALESCE((SELECT SUM(pr.ilosc) FROM produkty pr WHERE pr.paleta_id = pal.id), 0)
            + COALESCE((SELECT SUM(pr.sprzedano_offline) FROM produkty pr WHERE pr.paleta_id = pal.id), 0)
            + COALESCE((SELECT SUM(sp.ilosc) FROM sprzedaze sp
                JOIN produkty pp ON sp.produkt_id = pp.id
                WHERE pp.paleta_id = pal.id
                AND sp.status NOT IN ('zwrot','anulowane','anulowana')), 0)
            as ilosc_sztuk
        FROM palety pal WHERE pal.id = ?
    ''', (paleta_id,)).fetchone()
    if row and row['cena_zakupu'] and row['ilosc_sztuk'] and row['ilosc_sztuk'] > 0:
        return round(row['cena_zakupu'] / row['ilosc_sztuk'], 2)
    return 0


_mag_stats_cache = {'data': None, 'time': 0}

def get_stats():
    """Zwraca statystyki magazynu (cached 30s)"""
    import time as _time
    now = _time.time()
    if _mag_stats_cache['data'] and (now - _mag_stats_cache['time']) < 30:
        return _mag_stats_cache['data']
    conn = get_db()
    # Filtr: status IN ('magazyn','wystawiony') — spójnie z KPI dashboard
    _w_statuses = ('magazyn', 'wystawiony')
    stats = {
        'produkty': conn.execute('SELECT COUNT(*) FROM produkty WHERE status IN (?,?)', _w_statuses).fetchone()[0],
        'sztuki': conn.execute('SELECT COALESCE(SUM(ilosc),0) FROM produkty WHERE status IN (?,?)', _w_statuses).fetchone()[0],
        'palety': conn.execute('SELECT COUNT(DISTINCT paleta) FROM produkty WHERE paleta!="" AND status IN (?,?)', _w_statuses).fetchone()[0],
        'dostawcy': conn.execute('SELECT COUNT(DISTINCT dostawca) FROM produkty WHERE dostawca!="" AND status IN (?,?)', _w_statuses).fetchone()[0],
        # Wartość zakupu = avg koszt/szt z palet × sztuki w magazynie
        'wartosc_zakupu': 0,  # obliczone poniżej
        'wartosc_netto': 0,   # obliczone poniżej
        'wartosc_allegro': round(conn.execute('SELECT COALESCE(SUM(cena_allegro*ilosc),0) FROM produkty WHERE status IN (?,?)', _w_statuses).fetchone()[0] or 0, 2),
    }
    # Średni koszt/szt — identycznie jak w KPI dashboard (analytics.py)
    # total_items = aktualne w magazynie + sprzedane przez sprzedaze
    # Zoptymalizowane: JOIN zamiast correlated subqueries
    avg = conn.execute('''
        SELECT SUM(pal.cena_zakupu) as total_koszt,
        COALESCE(SUM(mag.mag_szt), 0) + COALESCE(SUM(spr.spr_szt), 0) as total_items
        FROM palety pal
        LEFT JOIN (
            SELECT paleta_id, SUM(CASE WHEN status NOT IN ('sprzedany','wyslany') THEN ilosc ELSE 0 END) as mag_szt
            FROM produkty GROUP BY paleta_id
        ) mag ON mag.paleta_id = pal.id
        LEFT JOIN (
            SELECT pp.paleta_id, SUM(sp.ilosc) as spr_szt
            FROM sprzedaze sp
            JOIN produkty pp ON sp.produkt_id = pp.id
            WHERE sp.status NOT IN ('zwrot','anulowane','anulowana')
            GROUP BY pp.paleta_id
        ) spr ON spr.paleta_id = pal.id
        WHERE pal.cena_zakupu > 0
    ''').fetchone()
    total_koszt = avg['total_koszt'] or 0
    total_items = avg['total_items'] or 1
    avg_cost = total_koszt / total_items if total_items > 0 else 0
    brutto = avg_cost * stats['sztuki']
    stats['wartosc_zakupu'] = round(brutto, 2)
    stats['wartosc_netto'] = round(brutto / 1.23, 2)
    _mag_stats_cache['data'] = stats
    _mag_stats_cache['time'] = _time.time()
    return stats

# ============================================================
# SZABLONY
# ============================================================
_MAGAZYNIER_CSS = '''
/* ═══════════════════════════════════════════════════════════════
   AKCES HUB · Magazynier Module — Cyberpunk Design System
   Colors: Cyan #8ff5ff · Pink #ff6b9b · Lime #beee00
   Font: Space Grotesk · Icons: Material Symbols
   ═══════════════════════════════════════════════════════════════ */

/* --- Header --- */
.hdr{text-align:center;padding:20px 0;border-bottom:1px solid rgba(143,245,255,0.12);margin-bottom:15px}
.hdr h1{font-size:1.5rem;color:#8ff5ff;font-family:'Space Grotesk',sans-serif;font-weight:700;text-shadow:0 0 20px rgba(143,245,255,0.35);letter-spacing:0.05em;display:flex;align-items:center;justify-content:center;gap:10px}
.hdr h1 .material-symbols-outlined{font-size:1.6rem;color:#ff6b9b}
.hdr small{color:var(--text-muted);font-size:0.8rem}

/* --- Stats --- */
.stat-v{font-size:1.4rem;font-weight:700;color:#8ff5ff;font-family:'Space Grotesk',sans-serif}
.stat-v.green{color:#beee00}
.stat-l{font-size:0.7rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.08em;margin-top:4px}

/* --- Buttons --- */
.btn-p{background:rgba(143,245,255,0.10);border:1px solid rgba(143,245,255,0.25);color:#8ff5ff;transition:all 0.2s}
.btn-p:hover{background:rgba(143,245,255,0.18);box-shadow:0 0 12px rgba(143,245,255,0.15)}
.btn-ok{background:rgba(190,238,0,0.12);border:1px solid rgba(190,238,0,0.25);color:#beee00}
.btn-ok:hover{background:rgba(190,238,0,0.20)}
.btn-2{background:var(--bg);border:1px solid var(--border);color:var(--text)}
.btn-warn{background:var(--yellow);color:#000}
.btn-err{background:var(--red)}

/* --- Search --- */
.search{backdrop-filter:blur(20px);background:rgba(10,10,22,0.70);border:1px solid rgba(143,245,255,0.08);border-radius:14px;padding:14px;margin-bottom:15px}
.search form{display:flex;gap:10px}
.search input{flex:1;padding:14px;background:rgba(14,14,20,0.8);border:1px solid rgba(143,245,255,0.10);border-radius:10px;color:var(--text);font-size:1rem;transition:all 0.3s;font-family:'Space Grotesk',sans-serif}
.search input:focus{outline:none;border-color:#8ff5ff;box-shadow:0 0 16px rgba(143,245,255,0.12)}
.search button{padding:14px 20px;background:rgba(143,245,255,0.10);border:1px solid rgba(143,245,255,0.25);border-radius:10px;color:#8ff5ff;font-size:1.2rem;cursor:pointer;transition:all 0.2s}
.search button:hover{transform:scale(1.05);box-shadow:0 0 12px rgba(143,245,255,0.2)}

/* --- Product Grid --- */
.items-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
.item{display:flex;align-items:center;backdrop-filter:blur(20px);background:rgba(10,10,22,0.70);border:1px solid rgba(143,245,255,0.06);border-radius:12px;padding:12px;margin-bottom:8px;text-decoration:none;color:var(--text);transition:all 0.3s;box-shadow:0 2px 8px rgba(0,0,0,0.3)}
.item:hover{border-color:#8ff5ff;transform:translateX(4px);box-shadow:0 0 16px rgba(143,245,255,0.08)}
.item img{width:50px;height:50px;object-fit:contain;background:rgba(255,255,255,0.95);border-radius:8px;margin-right:12px}
.item-info{flex:1;min-width:0}
.item-name{font-weight:600;font-size:0.9rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-family:'Space Grotesk',sans-serif}
.item-meta{font-size:0.75rem;color:var(--text-muted)}
.item-right{text-align:right;margin-left:10px}
.item-qty{font-size:1.2rem;font-weight:700;color:#8ff5ff;font-family:'Space Grotesk',sans-serif}
.item-price{font-size:0.75rem;color:#beee00}

/* --- Product Card --- */
.card-img{width:100%;max-height:250px;object-fit:contain;background:rgba(255,255,255,0.95);padding:10px;border-radius:14px 14px 0 0}
.card-body{padding:15px}
.card-name{font-size:1.2rem;font-weight:700;margin-bottom:14px;font-family:'Space Grotesk',sans-serif;line-height:1.3}

/* --- Location Panel --- */
.loc{background:rgba(10,10,22,0.6);border:2px solid rgba(143,245,255,0.20);border-radius:12px;padding:14px;margin-bottom:14px;position:relative;overflow:hidden}
.loc::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#8ff5ff,#ff6b9b,#beee00);opacity:0.6}
.loc-title{font-size:0.72rem;color:#8ff5ff;text-transform:uppercase;margin-bottom:10px;font-weight:600;letter-spacing:0.1em;display:flex;align-items:center;gap:6px}
.loc-title .material-symbols-outlined{font-size:1rem}
.loc-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;text-align:center}
.loc-v{font-size:1.1rem;font-weight:700;color:#beee00;font-family:'Space Grotesk',sans-serif}
.loc-l{font-size:0.65rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.05em}

/* --- Detail Grid --- */
.det-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:14px}
.det{background:rgba(10,10,22,0.5);padding:12px;border-radius:10px;border:1px solid rgba(255,255,255,0.04);transition:all 0.25s}
.det:hover{background:rgba(143,245,255,0.04);border-color:rgba(143,245,255,0.10)}
.det-l{font-size:0.68rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.05em;display:flex;align-items:center;gap:4px}
.det-l .material-symbols-outlined{font-size:0.85rem;color:#8ff5ff}
.det-v{font-size:0.95rem;font-weight:600;margin-top:3px;font-family:'Space Grotesk',sans-serif}
.det-v.green{color:#beee00}
.badge-ok{background:rgba(190,238,0,0.10);color:#beee00;padding:3px 8px;border-radius:6px;font-weight:700}
.badge-err{background:var(--red-soft);color:var(--red);padding:3px 8px;border-radius:6px;font-weight:700}

/* --- Forms --- */
.form-ctrl,.form-input{width:100%;padding:12px;background:rgba(14,14,20,0.8);border:1px solid rgba(143,245,255,0.10);border-radius:10px;color:var(--text);font-size:1rem;transition:all 0.3s;font-family:'Space Grotesk',sans-serif}
.form-ctrl:focus,.form-input:focus{outline:none;border-color:#8ff5ff;box-shadow:0 0 16px rgba(143,245,255,0.12)}
.form-row-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}

/* --- Alerts --- */
.alert-ok{background:rgba(190,238,0,0.08);border:1px solid rgba(190,238,0,0.25);color:#beee00;padding:12px 16px;border-radius:10px;margin-bottom:12px;font-size:0.9rem;display:flex;align-items:center;gap:8px}
.alert-warn{background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.25);color:var(--yellow);padding:12px 16px;border-radius:10px;margin-bottom:12px;font-size:0.9rem;display:flex;align-items:center;gap:8px}
.alert-err{background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.25);color:var(--red);padding:12px 16px;border-radius:10px;margin-bottom:12px;font-size:0.9rem;display:flex;align-items:center;gap:8px}

/* --- Sections --- */
.section{color:#8ff5ff;font-weight:600;font-size:0.9rem;margin:18px 0 12px;display:flex;align-items:center;gap:8px;font-family:'Space Grotesk',sans-serif;letter-spacing:0.05em}

/* --- Quick Actions --- */
.quick-actions{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px}
.quick-btn{padding:10px 5px;font-size:0.75rem;border-radius:10px;border:none;cursor:pointer;font-weight:600;color:#fff;transition:all 0.25s;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:5px;text-align:center}
.quick-btn .material-symbols-outlined{font-size:1rem}
.quick-btn:hover{transform:translateY(-2px);box-shadow:0 4px 16px rgba(0,0,0,0.4)}

/* --- Action Grid (Tool buttons) --- */
.act-grid{display:grid;gap:6px;margin-bottom:10px}
.act-grid-2{grid-template-columns:1fr 1fr}
.act-grid-4{grid-template-columns:repeat(4,1fr)}
.act-grid-3{grid-template-columns:repeat(3,1fr)}
.act-btn{display:flex;align-items:center;justify-content:center;gap:6px;padding:12px;border-radius:10px;font-weight:600;font-size:0.82rem;text-decoration:none;transition:all 0.25s;cursor:pointer;border:1px solid;font-family:'Space Grotesk',sans-serif}
.act-btn:hover{transform:translateY(-1px)}
.act-btn .material-symbols-outlined{font-size:1.1rem}
.act-btn-sm{display:flex;flex-direction:column;align-items:center;gap:4px;padding:10px 4px;border-radius:10px;font-size:0.65rem;text-decoration:none;text-align:center;transition:all 0.25s;cursor:pointer;border:1px solid;font-weight:500}
.act-btn-sm .material-symbols-outlined{font-size:1.2rem}
.act-btn-sm:hover{transform:translateY(-1px)}

/* --- Status Panel --- */
.status-panel{background:rgba(10,10,22,0.6);border:2px solid;border-radius:12px;padding:16px;margin-bottom:15px;position:relative;overflow:hidden}
.status-panel::after{content:'';position:absolute;top:0;right:0;width:60px;height:60px;border-radius:0 0 0 60px;opacity:0.06}
.status-label{font-size:0.72rem;color:var(--text-muted);margin-bottom:4px;text-transform:uppercase;letter-spacing:0.1em}
.status-value{font-size:1.3rem;font-weight:700;font-family:'Space Grotesk',sans-serif;display:flex;align-items:center;gap:8px}
.status-value .material-symbols-outlined{font-size:1.4rem}

/* --- Timeline / Historia --- */
.timeline{position:relative;padding-left:25px;margin:15px 0}
.timeline::before{content:'';position:absolute;left:8px;top:0;bottom:0;width:2px;background:linear-gradient(180deg,#8ff5ff,#ff6b9b,#beee00)}
.timeline-item{position:relative;padding:10px 0 10px 15px;border-bottom:1px solid rgba(255,255,255,0.04)}
.timeline-item:last-child{border-bottom:none}
.timeline-item::before{content:'';position:absolute;left:-21px;top:14px;width:12px;height:12px;border-radius:50%;background:#8ff5ff;border:2px solid var(--bg);box-shadow:0 0 8px rgba(143,245,255,0.3)}
.timeline-item.green::before{background:#beee00;box-shadow:0 0 8px rgba(190,238,0,0.3)}
.timeline-item.yellow::before{background:var(--yellow)}
.timeline-item.pink::before{background:#ff6b9b;box-shadow:0 0 8px rgba(255,107,155,0.3)}
.timeline-date{font-size:0.7rem;color:var(--text-muted)}
.timeline-text{font-size:0.85rem;margin-top:2px}

/* --- Historia Card --- */
.hist-card{backdrop-filter:blur(20px);background:rgba(10,10,22,0.70);border:1px solid rgba(255,107,155,0.15);border-radius:14px;margin-bottom:20px;overflow:hidden}
.hist-header{background:linear-gradient(135deg,rgba(255,107,155,0.12),rgba(143,245,255,0.06));border-bottom:1px solid rgba(255,107,155,0.15);padding:18px;display:flex;align-items:center;gap:12px}
.hist-header .material-symbols-outlined{font-size:1.5rem;color:#ff6b9b}
.hist-header h3{font-size:1.15rem;font-weight:700;color:#fff;font-family:'Space Grotesk',sans-serif;margin:0}
.hist-header small{font-size:0.82rem;color:rgba(255,255,255,0.65);margin-top:2px}
.hist-entry{padding:12px 16px;border-bottom:1px solid rgba(255,255,255,0.04);display:flex;justify-content:space-between;align-items:start;gap:12px;transition:background 0.2s}
.hist-entry:hover{background:rgba(143,245,255,0.03)}
.hist-entry:last-child{border-bottom:none}
.hist-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.hist-icon .material-symbols-outlined{font-size:1.1rem}
.hist-text{font-size:0.9rem;font-weight:500;color:#e2e8f0}
.hist-date{font-size:0.75rem;color:#64748b;white-space:nowrap}
.hist-tags{display:flex;flex-wrap:wrap;gap:4px;margin-top:6px}
.hist-tag{font-size:0.7rem;padding:2px 8px;border-radius:6px;background:rgba(255,107,155,0.10);color:#ff6b9b}
.hist-actions{display:flex;gap:6px;margin-left:8px;flex-shrink:0}
.hist-actions a{width:28px;height:28px;border-radius:6px;display:flex;align-items:center;justify-content:center;transition:all 0.2s;text-decoration:none}
.hist-actions a .material-symbols-outlined{font-size:0.95rem}
.hist-actions a:first-child{background:rgba(245,158,11,0.10);color:#fbbf24}
.hist-actions a:first-child:hover{background:rgba(245,158,11,0.20)}
.hist-actions a:last-child{background:rgba(239,68,68,0.10);color:#ef4444}
.hist-actions a:last-child:hover{background:rgba(239,68,68,0.20)}

/* --- Siblings Panel --- */
.siblings-panel{margin:15px;padding:16px;backdrop-filter:blur(20px);background:rgba(10,10,22,0.70);border:1px solid rgba(143,245,255,0.08);border-radius:14px}
.siblings-title{font-family:'Space Grotesk',sans-serif;font-weight:700;font-size:1rem;margin-bottom:12px;color:#8ff5ff;display:flex;align-items:center;gap:8px}
.siblings-title .material-symbols-outlined{font-size:1.2rem}
.sib-row{display:flex;align-items:center;gap:10px;padding:10px;margin-bottom:4px;background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.05);border-radius:10px;text-decoration:none;color:inherit;transition:all 0.25s}
.sib-row:hover{background:rgba(143,245,255,0.04)}
.sib-row.current{background:rgba(143,245,255,0.06);border-color:rgba(143,245,255,0.15)}
.sib-klasa{font-size:1.4rem;font-weight:800;font-family:'Space Grotesk',sans-serif;min-width:40px;text-align:center}

/* --- Modals --- */
.cyber-modal{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.85);backdrop-filter:blur(6px);z-index:1000;overflow-y:auto;padding:20px}
.cyber-modal-inner{background:rgba(15,15,30,0.95);backdrop-filter:blur(20px);border:1px solid rgba(143,245,255,0.10);border-radius:16px;padding:24px;max-width:480px;margin:40px auto;position:relative;overflow:hidden}
.cyber-modal-inner::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#8ff5ff,#ff6b9b,#beee00)}
.cyber-modal-title{font-size:1.2rem;font-weight:700;font-family:'Space Grotesk',sans-serif;display:flex;align-items:center;gap:8px;margin-bottom:4px}
.cyber-modal-close{background:none;border:none;color:#64748b;font-size:1.3rem;cursor:pointer;transition:color 0.2s;padding:4px}
.cyber-modal-close:hover{color:#ff6b9b}

/* --- Toast Notifications --- */
.toast-container{position:fixed;top:80px;right:20px;z-index:1000;display:flex;flex-direction:column;gap:10px;max-width:400px}
.toast{backdrop-filter:blur(20px);background:rgba(10,10,22,0.85);border:1px solid rgba(143,245,255,0.08);border-radius:12px;padding:16px;box-shadow:0 8px 32px rgba(0,0,0,0.4);display:flex;align-items:start;gap:12px;animation:slideInRight 0.3s ease-out;min-width:300px}
.toast .material-symbols-outlined{font-size:1.3rem;flex-shrink:0;margin-top:2px}
.toast.success{border-left:3px solid #beee00}.toast.success .material-symbols-outlined{color:#beee00}
.toast.error{border-left:3px solid var(--red)}.toast.error .material-symbols-outlined{color:var(--red)}
.toast.warning{border-left:3px solid var(--yellow)}.toast.warning .material-symbols-outlined{color:var(--yellow)}
.toast.info{border-left:3px solid #8ff5ff}.toast.info .material-symbols-outlined{color:#8ff5ff}
.toast-content{flex:1}
.toast-title{font-weight:600;margin-bottom:4px;font-size:0.95rem;font-family:'Space Grotesk',sans-serif}
.toast-message{font-size:0.85rem;color:var(--text-muted)}
.toast-close{cursor:pointer;color:var(--text-muted);font-size:1.2rem;flex-shrink:0;transition:color 0.2s}
.toast-close:hover{color:#ff6b9b}
.toast.removing{animation:slideOutRight 0.3s ease-in forwards}

/* --- Loading Overlay --- */
.loading-overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.7);backdrop-filter:blur(4px);z-index:999;display:flex;align-items:center;justify-content:center}
.loading-spinner{width:60px;height:60px;border:3px solid rgba(143,245,255,0.15);border-top:3px solid #8ff5ff;border-radius:50%;animation:spin 0.8s linear infinite}
.loading-text{color:#8ff5ff;margin-top:20px;font-size:1.1rem;font-weight:600;text-align:center;font-family:'Space Grotesk',sans-serif}

/* --- Animations --- */
@keyframes slideDown{from{opacity:0;transform:translateY(-20px)}to{opacity:1;transform:translateY(0)}}
@keyframes slideInRight{from{opacity:0;transform:translateX(100px)}to{opacity:1;transform:translateX(0)}}
@keyframes slideOutRight{to{opacity:0;transform:translateX(100px)}}
@keyframes spin{to{transform:rotate(360deg)}}
@keyframes neonPulse{0%,100%{box-shadow:0 0 8px rgba(143,245,255,0.2)}50%{box-shadow:0 0 20px rgba(143,245,255,0.4)}}
@keyframes gradientShift{0%{background-position:0% 50%}50%{background-position:100% 50%}100%{background-position:0% 50%}}

/* --- Responsive --- */
@media(min-width:1200px){.items-grid{grid-template-columns:repeat(3,1fr)}}
@media(max-width:900px){.toast-container{right:10px;max-width:320px}.toast{min-width:280px}}
@media(max-width:768px){.items-grid{grid-template-columns:1fr}.form-row{grid-template-columns:1fr}.quick-actions{grid-template-columns:repeat(2,1fr)}.act-grid-4{grid-template-columns:repeat(2,1fr)}.toast-container{top:70px;right:10px;left:10px;max-width:none}.toast{min-width:auto}}

/* === Legacy class overrides → Cyberpunk === */
.card{backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:18px;margin-bottom:14px;transition:all 0.2s}
.card:hover{border-color:rgba(143,245,255,0.15)}
.header{text-align:center;padding:20px 0;border-bottom:1px solid rgba(143,245,255,0.12);margin-bottom:18px}
.header h1,.header h2{font-family:'Space Grotesk',sans-serif;color:#8ff5ff;text-shadow:0 0 20px rgba(143,245,255,0.25);font-weight:800}
.form-control{width:100%;padding:10px 12px;background:rgba(14,14,20,0.8);border:1px solid rgba(143,245,255,0.10);border-radius:8px;color:var(--text);font-family:'Space Grotesk',sans-serif;font-size:0.9rem;transition:all 0.3s;box-sizing:border-box}
.form-control:focus{outline:none;border-color:#8ff5ff;box-shadow:0 0 16px rgba(143,245,255,0.12)}
'''

_MAGAZYNIER_JS = '''
// Toast Notifications System — Material Symbols
function showToast(title, message, type, duration){
    type=type||'info'; duration=duration===undefined?3000:duration;
    var container=document.getElementById('toast-container');
    if(!container){container=document.createElement('div');container.id='toast-container';container.className='toast-container';document.body.appendChild(container);}
    var toast=document.createElement('div');
    toast.className='toast '+type;
    var icons={success:'check_circle',error:'error',warning:'warning',info:'info'};
    var icon=icons[type]||'info';
    toast.innerHTML='<span class=material-symbols-outlined>'+icon+'</span><div class="toast-content"><div class="toast-title">'+title+'</div>'+(message?'<div class="toast-message">'+message+'</div>':'')+'</div><div class="toast-close" onclick="removeToast(this.parentElement)">&times;</div>';
    container.appendChild(toast);
    if(duration>0){setTimeout(function(){removeToast(toast)},duration);}
    return toast;
}
function removeToast(toast){toast.classList.add('removing');setTimeout(function(){toast.remove()},300);}

// Loading Overlay System
function showLoading(text){
    text=text||'Ladowanie...';
    if(document.getElementById('loading-overlay'))return;
    var overlay=document.createElement('div');
    overlay.id='loading-overlay';overlay.className='loading-overlay';
    overlay.innerHTML='<div style="text-align:center"><div class="loading-spinner"></div><div class="loading-text">'+text+'</div></div>';
    document.body.appendChild(overlay);
}
function hideLoading(){var o=document.getElementById('loading-overlay');if(o){o.style.animation='fadeOut 0.2s ease-out';setTimeout(function(){o.remove()},200);}}

// Auto-convert URL msg params to toasts
window.addEventListener('DOMContentLoaded', function(){
    var params=new URLSearchParams(window.location.search);
    var msg=params.get('msg');
    if(msg){
        var decoded=decodeURIComponent(msg.replace(/\\+/g,' '));
        var type=(decoded.indexOf('Blad')>=0||decoded.indexOf('\\u274c')>=0)?'error':'success';
        showToast('', decoded, type);
        var url=new URL(window.location);url.searchParams.delete('msg');
        window.history.replaceState({}, '', url);
    }
});
'''

_MAGAZYNIER_TEMPLATE = '''{% extends "base.html" %}
{% block page_title %}{{ page_title }}{% endblock %}
{% block content %}
<style>{{ magazynier_css|safe }}</style>
<div class="toast-container" id="toast-container"></div>
{{ content_html|safe }}
<script>{{ magazynier_js|safe }}</script>
{% endblock %}'''

def render(content, page_title='Magazynier'):
    return render_template_string(
        _MAGAZYNIER_TEMPLATE,
        content_html=content,
        page_title=page_title,
        magazynier_css=_MAGAZYNIER_CSS,
        magazynier_js=_MAGAZYNIER_JS,
        version=current_app.config.get('VERSION', ''),
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        current_user=session.get('username')
    )

# ============================================================
# ROUTES
# ============================================================
@magazynier_bp.route('/')
def index():
    s = get_stats()
    conn = get_db()
    products = conn.execute('SELECT * FROM produkty ORDER BY data_dodania DESC LIMIT 10').fetchall()
    # Convert Row objects to dicts for Jinja2 template access
    products = [dict(p) for p in products]
    # Rozwiąż obrazki z pełnym fallback chain (images JSON, ASIN folder, placeholder)
    for p in products:
        p['_resolved_img'] = _resolve_product_image(p, size='sm')
    return render_template(
        'magazyn_index.html',
        stats=s,
        products=products,
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        version=current_app.config.get('VERSION', ''),
        current_user=session.get('username')
    )

@magazynier_bp.route('/skaner')
def skaner():
    """Skaner kodów kreskowych z kamery"""
    html = '''
    <div class="hdr"><h1><span class=material-symbols-outlined>qr_code_scanner</span> SKANER KODÓW</h1><small>Zeskanuj EAN/ASIN</small></div>

    <div id="scanner-container" style="position:relative;width:100%;max-width:400px;margin:0 auto 15px">
        <video id="video" style="width:100%;border-radius:12px;background:#000" playsinline></video>
        <div id="scan-line" style="position:absolute;left:10%;right:10%;top:50%;height:2px;background:#beee00;box-shadow:0 0 10px #beee00;animation:scan 2s infinite"></div>
    </div>
    
    <style>
        @keyframes scan { 0%,100%{opacity:1} 50%{opacity:0.3} }
    </style>
    
    <div id="result" style="text-align:center;margin-bottom:15px">
        <div style="color:#64748b;font-size:0.9rem">Skieruj kamerę na kod kreskowy</div>
    </div>
    
    <div class="card" style="padding:15px">
        <div class="form-group">
            <label style="font-size:0.8rem;color:#64748b">Lub wpisz ręcznie:</label>
            <form action="/magazyn/szukaj" method="GET" id="manualForm" style="display:flex;gap:8px;margin-top:8px">
                <input type="text" name="q" id="manual-input" class="form-ctrl" placeholder="EAN / ASIN / MAG-kod..." style="flex:1;padding:12px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff" autofocus
                    inputmode="text" autocomplete="off">
                <button type="submit" class="btn btn-p" style="width:auto;padding:12px 20px;margin:0"><span class=material-symbols-outlined style=font-size:1.1rem>search</span></button>
            </form>
            <div id="manualHint" style="font-size:0.7rem;color:#64748b;margin-top:4px;display:none"></div>
        </div>
    </div>
    
    <div id="last-scanned" style="display:none" class="card">
        <div style="padding:15px">
            <div style="font-size:0.75rem;color:#64748b;margin-bottom:5px">Ostatnio zeskanowany:</div>
            <div id="last-code" style="font-size:1.2rem;font-weight:700;color:#beee00"></div>
        </div>
    </div>
    
    <a href="/magazyn" class="back">← Powrót</a>
    
    <script src="https://cdn.jsdelivr.net/npm/@zxing/library@0.19.1/umd/index.min.js" integrity="sha384-NyKHkzm0aj4yWFC3Hh4cp1VflBgCLfStVlAK6WJOdXAht/pj6RHbMcZgUj48rcAs" crossorigin="anonymous"></script>
    <script>
    (function() {
        const video = document.getElementById('video');
        const resultDiv = document.getElementById('result');
        const lastScanned = document.getElementById('last-scanned');
        const lastCode = document.getElementById('last-code');
        const manualInput = document.getElementById('manual-input');
        
        let lastResult = '';
        let lastTime = 0;
        
        // Inicjalizacja skanera
        const codeReader = new ZXing.BrowserMultiFormatReader();
        
        codeReader.listVideoInputDevices()
            .then(devices => {
                // Preferuj tylną kamerę
                let selectedDevice = devices[0];
                for (const device of devices) {
                    if (device.label.toLowerCase().includes('back') || 
                        device.label.toLowerCase().includes('rear') ||
                        device.label.toLowerCase().includes('environment')) {
                        selectedDevice = device;
                        break;
                    }
                }
                
                if (selectedDevice) {
                    startScanning(selectedDevice.deviceId);
                } else {
                    resultDiv.innerHTML = '<div style="color:#ef4444">Brak kamery</div>';
                }
            })
            .catch(err => {
                resultDiv.innerHTML = '<div style="color:#ef4444">Błąd dostępu do kamery: ' + err.message + '</div>';
            });
        
        function startScanning(deviceId) {
            codeReader.decodeFromVideoDevice(deviceId, 'video', (result, err) => {
                if (result) {
                    const code = result.getText();
                    const now = Date.now();
                    
                    // Zapobiegaj wielokrotnemu skanowaniu tego samego kodu
                    if (code !== lastResult || now - lastTime > 3000) {
                        lastResult = code;
                        lastTime = now;
                        
                        // Wibracja (jeśli dostępna) - mocna podwójna
                        if (navigator.vibrate) navigator.vibrate([50, 30, 100]);
                        
                        // Pokaż wynik
                        resultDiv.innerHTML = '<div style="color:#beee00;font-size:1.2rem;font-weight:700"><span class=material-symbols-outlined style=vertical-align:middle>check_circle</span> ' + code + '</div>';
                        lastScanned.style.display = 'block';
                        lastCode.textContent = code;
                        manualInput.value = code;
                        
                        // Przekieruj do produktu po 1.5s
                        setTimeout(() => {
                            window.location.href = '/magazyn/szukaj?q=' + encodeURIComponent(code);
                        }, 1500);
                    }
                }
            });
        }
        
        // Zatrzymaj skaner przy opuszczaniu strony
        window.addEventListener('beforeunload', () => {
            codeReader.reset();
        });

        // === AUTO-SEARCH: szukaj po wpisaniu kodu ręcznie ===
        let searchTimer = null;
        function triggerSearch() {
            clearTimeout(searchTimer);
            const val = manualInput.value.trim();
            if (val.length >= 3) {
                document.getElementById('manualForm').submit();
            }
        }

        // Debounce na input (1.2s żeby zdążyć wpisać)
        manualInput.addEventListener('input', function() {
            clearTimeout(searchTimer);
            const val = this.value.trim();
            const hint = document.getElementById('manualHint');
            if (val.length >= 3) {
                hint.style.display = 'block';
                hint.textContent = 'Szukam za chwilę...';
                hint.style.color = '#8ff5ff';
                searchTimer = setTimeout(triggerSearch, 1200);
            } else {
                hint.style.display = 'none';
            }
        });

        // Paste = szukaj od razu
        manualInput.addEventListener('paste', function() {
            clearTimeout(searchTimer);
            setTimeout(triggerSearch, 100);
        });

        // Enter = natychmiastowe szukanie (formularz obsłuży to natywnie)
        // Nie potrzeba preventDefault — form submit zadziała
    })();
    </script>
    '''
    return render(html)

@magazynier_bp.route('/produkty')
def produkty():
    # Pobierz parametry filtrowania i sortowania
    filter_status = request.args.get('status', '')
    filter_paleta = request.args.get('paleta', '')
    filter_dostawca = request.args.get('dostawca', '')
    sort_by = request.args.get('sort', 'data')  # data, cena, nazwa, ilosc
    sort_dir = request.args.get('dir', 'desc').upper()
    if sort_dir not in ('ASC', 'DESC'):
        sort_dir = 'DESC'
    search = request.args.get('search', '')
    msg = request.args.get('msg', '')
    
    conn = get_db()
    
    # Buduj query z filtrami
    query = 'SELECT * FROM produkty WHERE 1=1'
    params = []
    
    filter_stan = request.args.get('stan', '')
    filter_klasa = request.args.get('klasa', '')

    if filter_status == 'nieoceniony':
        query += " AND (stan_przyjecia = 'nieoceniony' OR stan_przyjecia IS NULL OR stan_przyjecia = '')"
    elif filter_status:
        query += ' AND status = ?'
        params.append(filter_status)

    if filter_stan:
        query += ' AND stan_przyjecia = ?'
        params.append(filter_stan)

    if filter_klasa:
        query += ' AND klasa_jakosci = ?'
        params.append(filter_klasa)

    filter_paleta_id = request.args.get('paleta_id', '')
    if filter_paleta_id:
        query += ' AND paleta_id = ?'
        params.append(int(filter_paleta_id))
    elif filter_paleta:
        query += ' AND (paleta = ? OR paleta_id IN (SELECT id FROM palety WHERE nazwa = ?))'
        params.extend([filter_paleta, filter_paleta])
    
    if filter_dostawca:
        query += ' AND dostawca = ?'
        params.append(filter_dostawca)
    
    if search:
        query += ' AND (nazwa LIKE ? OR ean LIKE ? OR asin LIKE ? OR kod_magazynowy LIKE ?)'
        search_term = f'%{search}%'
        params.extend([search_term, search_term, search_term, f'%{search.upper()}%'])
    
    # Dodaj sortowanie
    sort_columns = {
        'data': 'data_dodania',
        'cena': 'cena_allegro',
        'nazwa': 'nazwa',
        'ilosc': 'ilosc'
    }
    sort_col = sort_columns.get(sort_by, 'data_dodania')
    query += f' ORDER BY {sort_col} {sort_dir}'
    
    products = conn.execute(query, params).fetchall()

    # Pre-compute pallet cost/szt for profit (cena_zakupu_brutto / total_szt)
    _paleta_ids = set(p['paleta_id'] for p in products if p['paleta_id'])
    _koszt_cache = {}
    for pid in _paleta_ids:
        _koszt_cache[pid] = _paleta_koszt_szt(conn, pid)

    # Pobierz unikalne wartości dla filtrów
    palety = [r[0] for r in conn.execute('SELECT DISTINCT paleta FROM produkty WHERE paleta IS NOT NULL ORDER BY paleta').fetchall()]
    dostawcy = [r[0] for r in conn.execute('SELECT DISTINCT dostawca FROM produkty WHERE dostawca IS NOT NULL ORDER BY dostawca').fetchall()]
    
    # Policz statusy
    status_counts = {}
    conn = get_db()
    for status in ['nowy', 'wystawiony', 'sprzedany', 'wyslany', 'uszkodzony', 'zwrot']:
        count = conn.execute('SELECT COUNT(*) FROM produkty WHERE status = ? OR (status IS NULL AND ? = "nowy")', (status, status)).fetchone()[0]
        status_counts[status] = count
    nieocenione_cnt = conn.execute("SELECT COUNT(*) FROM produkty WHERE stan_przyjecia = 'nieoceniony' OR (stan_przyjecia IS NULL AND status = 'magazyn') OR stan_przyjecia = ''").fetchone()[0]

    # Stan przedmiotu counts
    _stan_counts = {}
    for _sn in ['Nowy','Jak nowy','Używany','Uszkodzony','Zniszczony','nieoceniony']:
        if _sn == 'nieoceniony':
            _stan_counts[_sn] = conn.execute("SELECT COUNT(*) FROM produkty WHERE stan_przyjecia = 'nieoceniony' OR stan_przyjecia IS NULL OR stan_przyjecia = ''").fetchone()[0]
        else:
            _stan_counts[_sn] = conn.execute("SELECT COUNT(*) FROM produkty WHERE stan_przyjecia = ?", (_sn,)).fetchone()[0]

    # Klasa jakości counts
    _klasa_counts = {}
    for _kl in ['A','A-','B','C','D']:
        _klasa_counts[_kl] = conn.execute("SELECT COUNT(*) FROM produkty WHERE klasa_jakosci = ?", (_kl,)).fetchone()[0]
    _klasa_counts['brak'] = conn.execute("SELECT COUNT(*) FROM produkty WHERE klasa_jakosci IS NULL OR klasa_jakosci = ''").fetchone()[0]

    html = f'''
    <!-- Page Header -->
    <div style="margin-bottom:24px">
        <div style="font-size:0.6rem;text-transform:uppercase;letter-spacing:0.15em;color:#8ff5ff;font-weight:700;font-family:'Space Grotesk',sans-serif;margin-bottom:4px">Produkty</div>
        <h2 style="font-family:'Space Grotesk',sans-serif;font-size:1.8rem;font-weight:800;color:#f9f5f8;letter-spacing:-0.02em;margin:0">Katalog Zasobów</h2>
    </div>
    '''

    if msg:
        html += f'<script>Toast.success("{msg}");</script>'

    # Category tabs (horizontal scroll)
    html += f'''
    <div style="display:flex;overflow-x:auto;gap:10px;padding-bottom:8px;margin-bottom:20px;scrollbar-width:none;-ms-overflow-style:none">
        <a href="/magazyn/produkty" style="flex-shrink:0;padding:8px 18px;border-radius:10px;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;text-decoration:none;white-space:nowrap;transition:all 0.2s;{'background:#8ff5ff;color:#005d63;box-shadow:0 0 15px rgba(143,245,255,0.3)' if not filter_status else 'background:#262528;color:#adaaad;border:1px solid rgba(72,71,74,0.2)'}">
            Wszystkie</a>
        <a href="/magazyn/produkty?status=nowy" style="flex-shrink:0;padding:8px 18px;border-radius:10px;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;text-decoration:none;white-space:nowrap;transition:all 0.2s;{'background:#8ff5ff;color:#005d63;box-shadow:0 0 15px rgba(143,245,255,0.3)' if filter_status == 'nowy' else 'background:#262528;color:#adaaad;border:1px solid rgba(72,71,74,0.2)'}">
            Magazyn ({status_counts['nowy']})</a>
        <a href="/magazyn/produkty?status=wystawiony" style="flex-shrink:0;padding:8px 18px;border-radius:10px;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;text-decoration:none;white-space:nowrap;transition:all 0.2s;{'background:#ff6b9b;color:#47001f;box-shadow:0 0 15px rgba(255,107,155,0.3)' if filter_status == 'wystawiony' else 'background:#262528;color:#adaaad;border:1px solid rgba(72,71,74,0.2)'}">
            Allegro ({status_counts['wystawiony']})</a>
        <a href="/magazyn/produkty?status=sprzedany" style="flex-shrink:0;padding:8px 18px;border-radius:10px;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;text-decoration:none;white-space:nowrap;transition:all 0.2s;{'background:#beee00;color:#3a4a00;box-shadow:0 0 15px rgba(190,238,0,0.3)' if filter_status == 'sprzedany' else 'background:#262528;color:#adaaad;border:1px solid rgba(72,71,74,0.2)'}">
            Sprzedane ({status_counts['sprzedany']})</a>
        <a href="/magazyn/produkty?status=nieoceniony" style="flex-shrink:0;padding:8px 18px;border-radius:10px;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;text-decoration:none;white-space:nowrap;transition:all 0.2s;{'background:#f59e0b;color:#422006;box-shadow:0 0 15px rgba(245,158,11,0.3)' if filter_status == 'nieoceniony' else 'background:#262528;color:#adaaad;border:1px solid rgba(72,71,74,0.2)'}">
            Nieocenione ({nieocenione_cnt})</a>
    </div>

    <!-- STAN PRZEDMIOTU + KLASA JAKOŚCI - compact bento grid -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px">
        <div style="background:#131315;border:1px solid rgba(72,71,74,0.15);border-radius:10px;padding:14px">
            <div style="font-size:0.62rem;color:#8ff5ff;font-weight:700;letter-spacing:0.12em;margin-bottom:10px;font-family:'Space Grotesk',sans-serif;text-transform:uppercase">Stan przedmiotu</div>
            <div style="display:flex;gap:6px;flex-wrap:wrap">
                <a href="/magazyn/produkty?stan=Nowy" style="flex:1;min-width:50px;text-align:center;padding:8px 4px;background:rgba(190,238,0,0.06);border:1px solid rgba(190,238,0,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#beee00;font-family:'Space Grotesk',sans-serif">{_stan_counts.get('Nowy',0)}</div>
                    <div style="font-size:0.55rem;color:#767577;margin-top:2px;text-transform:uppercase;letter-spacing:0.05em;font-weight:700">Nowy</div>
                </a>
                <a href="/magazyn/produkty?stan=Jak nowy" style="flex:1;min-width:50px;text-align:center;padding:8px 4px;background:rgba(143,245,255,0.06);border:1px solid rgba(143,245,255,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#8ff5ff;font-family:'Space Grotesk',sans-serif">{_stan_counts.get('Jak nowy',0)}</div>
                    <div style="font-size:0.55rem;color:#767577;margin-top:2px;text-transform:uppercase;letter-spacing:0.05em;font-weight:700">Jak nowy</div>
                </a>
                <a href="/magazyn/produkty?stan=Używany" style="flex:1;min-width:50px;text-align:center;padding:8px 4px;background:rgba(234,179,8,0.06);border:1px solid rgba(234,179,8,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#eab308;font-family:'Space Grotesk',sans-serif">{_stan_counts.get('Używany',0)}</div>
                    <div style="font-size:0.55rem;color:#767577;margin-top:2px;text-transform:uppercase;letter-spacing:0.05em;font-weight:700">Używany</div>
                </a>
                <a href="/magazyn/produkty?stan=Uszkodzony" style="flex:1;min-width:50px;text-align:center;padding:8px 4px;background:rgba(249,115,22,0.06);border:1px solid rgba(249,115,22,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#f97316;font-family:'Space Grotesk',sans-serif">{_stan_counts.get('Uszkodzony',0)}</div>
                    <div style="font-size:0.55rem;color:#767577;margin-top:2px;text-transform:uppercase;letter-spacing:0.05em;font-weight:700">Uszk.</div>
                </a>
                <a href="/magazyn/produkty?status=nieoceniony" style="flex:1;min-width:50px;text-align:center;padding:8px 4px;background:rgba(255,255,255,0.02);border:1px solid rgba(72,71,74,0.15);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#767577;font-family:'Space Grotesk',sans-serif">{_stan_counts.get('nieoceniony',0)}</div>
                    <div style="font-size:0.55rem;color:#767577;margin-top:2px;text-transform:uppercase;letter-spacing:0.05em;font-weight:700">Brak</div>
                </a>
            </div>
        </div>
        <div style="background:#131315;border:1px solid rgba(72,71,74,0.15);border-radius:10px;padding:14px">
            <div style="font-size:0.62rem;color:#ff6b9b;font-weight:700;letter-spacing:0.12em;margin-bottom:10px;font-family:'Space Grotesk',sans-serif;text-transform:uppercase">Klasa jakości</div>
            <div style="display:flex;gap:6px;flex-wrap:wrap">
                <a href="/magazyn/produkty?klasa=A" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(190,238,0,0.06);border:1px solid rgba(190,238,0,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#beee00;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('A',0)}</div>
                    <div style="font-size:0.6rem;color:#beee00;margin-top:2px;font-weight:700">A</div>
                </a>
                <a href="/magazyn/produkty?klasa=A-" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(143,245,255,0.06);border:1px solid rgba(143,245,255,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#8ff5ff;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('A-',0)}</div>
                    <div style="font-size:0.6rem;color:#8ff5ff;margin-top:2px;font-weight:700">A-</div>
                </a>
                <a href="/magazyn/produkty?klasa=B" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(234,179,8,0.06);border:1px solid rgba(234,179,8,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#eab308;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('B',0)}</div>
                    <div style="font-size:0.6rem;color:#eab308;margin-top:2px;font-weight:700">B</div>
                </a>
                <a href="/magazyn/produkty?klasa=C" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(249,115,22,0.06);border:1px solid rgba(249,115,22,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#f97316;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('C',0)}</div>
                    <div style="font-size:0.6rem;color:#f97316;margin-top:2px;font-weight:700">C</div>
                </a>
                <a href="/magazyn/produkty?klasa=D" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(239,68,68,0.06);border:1px solid rgba(239,68,68,0.12);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#ef4444;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('D',0)}</div>
                    <div style="font-size:0.6rem;color:#ef4444;margin-top:2px;font-weight:700">D</div>
                </a>
                <a href="/magazyn/produkty?status=nieoceniony" style="flex:1;min-width:42px;text-align:center;padding:8px 4px;background:rgba(255,255,255,0.02);border:1px solid rgba(72,71,74,0.15);border-radius:8px;text-decoration:none">
                    <div style="font-size:1.2rem;font-weight:800;color:#767577;font-family:'Space Grotesk',sans-serif">{_klasa_counts.get('brak',0)}</div>
                    <div style="font-size:0.6rem;color:#767577;margin-top:2px;font-weight:700">Brak</div>
                </a>
            </div>
        </div>
    </div>
    '''

    # Filtry i sortowanie - sleek inline bar
    html += f'''
    <div style="padding:12px 16px;margin-bottom:16px;background:#131315;border:1px solid rgba(72,71,74,0.15);border-radius:10px">
        <form method="GET" action="/magazyn/produkty" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
            <input type="hidden" name="status" value="{filter_status}">
            <input type="text" name="search" value="{search}" placeholder="Szukaj produktu..." class="form-input" style="flex:2;min-width:140px;font-size:0.82rem">
            <select name="paleta" class="form-input" style="flex:1;min-width:100px;font-size:0.82rem">
                <option value="">Paleta</option>
                {"".join([f'<option value="{p}" {"selected" if filter_paleta == p else ""}>{p}</option>' for p in palety])}
            </select>
            <select name="dostawca" class="form-input" style="flex:1;min-width:100px;font-size:0.82rem">
                <option value="">Dostawca</option>
                {"".join([f'<option value="{d}" {"selected" if filter_dostawca == d else ""}>{d}</option>' for d in dostawcy])}
            </select>
            <select name="sort" class="form-input" style="min-width:90px;font-size:0.82rem">
                <option value="data" {"selected" if sort_by == "data" else ""}>Data</option>
                <option value="cena" {"selected" if sort_by == "cena" else ""}>Cena</option>
                <option value="nazwa" {"selected" if sort_by == "nazwa" else ""}>Nazwa</option>
                <option value="ilosc" {"selected" if sort_by == "ilosc" else ""}>Ilość</option>
            </select>
            <select name="dir" class="form-input" style="min-width:85px;font-size:0.82rem">
                <option value="desc" {"selected" if sort_dir == "desc" else ""}>Malejąco</option>
                <option value="asc" {"selected" if sort_dir == "asc" else ""}>Rosnąco</option>
            </select>
            <button type="submit" class="btn btn-ok" style="display:flex;align-items:center;gap:4px;padding:8px 14px;font-size:0.78rem"><span class=material-symbols-outlined style=font-size:0.95rem>filter_list</span> Filtruj</button>
            <a href="/magazyn/produkty" class="btn" style="display:flex;align-items:center;gap:4px;padding:8px 14px;font-size:0.78rem"><span class=material-symbols-outlined style=font-size:0.95rem>clear_all</span> Wyczyść</a>
        </form>
    </div>
    '''

    # Masowa edycja - collapsible panel
    html += f'''
    <form id="mass-edit-form" method="POST" action="/magazyn/produkty/masowa-edycja">
        <details style="margin-bottom:16px">
            <summary style="padding:12px 16px;background:#131315;border:1px solid rgba(255,107,155,0.15);border-radius:10px;cursor:pointer;list-style:none;display:flex;align-items:center;gap:8px;color:#ff6b9b;font-weight:700;font-size:0.82rem;font-family:'Space Grotesk',sans-serif">
                <span class=material-symbols-outlined style=font-size:1.1rem>bolt</span> MASOWA EDYCJA
                <span id="selected-count" style="margin-left:auto;font-size:0.75rem;color:#ff6b9b">Zaznaczono: <span id="count">0</span></span>
            </summary>
            <div style="padding:16px;background:#131315;border:1px solid rgba(255,107,155,0.15);border-top:none;border-radius:0 0 10px 10px">
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">
                    <div>
                        <label style="display:block;color:#adaaad;font-size:0.68rem;margin-bottom:4px;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Status</label>
                        <select name="new_status" class="form-input" style="width:100%;font-size:0.82rem">
                            <option value="">-- bez zmiany --</option>
                            <option value="magazyn">Magazyn</option>
                            <option value="wystawiony">Wystawiony (Allegro)</option>
                            <option value="sprzedany">Sprzedany</option>
                            <option value="uszkodzony">Uszkodzony</option>
                            <option value="zwrot">Zwrot</option>
                        </select>
                    </div>
                    <div>
                        <label style="display:block;color:#adaaad;font-size:0.68rem;margin-bottom:4px;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Stan</label>
                        <select name="new_stan" class="form-input" style="width:100%;font-size:0.82rem">
                            <option value="">-- bez zmiany --</option>
                            <option value="Nowy">Nowy</option>
                            <option value="Nowy w otwartym opakowaniu">Nowy w otwartym opak.</option>
                            <option value="Używany">Używany</option>
                            <option value="Uszkodzony">Uszkodzony</option>
                            <option value="Odnowiony">Odnowiony</option>
                        </select>
                    </div>
                    <div>
                        <label style="display:block;color:#adaaad;font-size:0.68rem;margin-bottom:4px;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Lokalizacja</label>
                        <input type="text" name="new_lokalizacja" class="form-input" placeholder="np. A1, B2" style="width:100%;font-size:0.82rem">
                    </div>
                    <div>
                        <label style="display:block;color:#adaaad;font-size:0.68rem;margin-bottom:4px;letter-spacing:0.06em;text-transform:uppercase;font-weight:700">Cena Allegro (zł)</label>
                        <input type="number" name="new_cena_allegro" class="form-input" placeholder="puste = bez zmiany" step="0.01" min="0" style="width:100%;font-size:0.82rem">
                    </div>
                </div>
                <div style="display:flex;gap:8px;flex-wrap:wrap">
                    <button type="button" onclick="toggleAll()" class="btn" style="background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);color:#8ff5ff;flex:1;min-width:120px;display:flex;align-items:center;justify-content:center;gap:4px;padding:10px;font-size:0.78rem">
                        <span class=material-symbols-outlined style=font-size:0.95rem>check_box</span> Zaznacz wszystkie
                    </button>
                    <button type="submit" class="btn btn-ok" onclick="return confirm('Zastosować zmiany dla ' + document.getElementById('count').textContent + ' produktów?')" style="flex:1;min-width:120px;display:flex;align-items:center;justify-content:center;gap:4px;padding:10px;font-size:0.78rem">
                        <span class=material-symbols-outlined style=font-size:0.95rem>check_circle</span> Zastosuj
                    </button>
                    <button type="button" onclick="pokazBoxModal()" class="btn" style="background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.2);color:#f59e0b;flex:1;min-width:120px;display:flex;align-items:center;justify-content:center;gap:4px;padding:10px;font-size:0.78rem">
                        <span class=material-symbols-outlined style=font-size:0.95rem>inbox</span> Zgrupuj w box
                    </button>
                </div>
            </div>
        </details>
    '''
    
    # Lista produktów - card-based layout
    html += '<div style="display:flex;flex-direction:column;gap:12px">'

    for p in products:
        img = _resolve_product_image(p, size='sm')
        pcode = get_product_code(p)
        _km = p['kod_magazynowy'] if p['kod_magazynowy'] else f"#{p['id']}"
        _ean_clean = p['ean'] if p['ean'] and p['ean'].upper() not in ('N/A','NAN','NONE') else ''
        display_code = f"{_km}"

        # Zysk per item (koszt = paleta.cena_zakupu / szt, fallback to cena_brutto)
        _ca = float(p['cena_allegro'] or 0)
        _ks = _koszt_cache.get(p['paleta_id'], 0)
        if _ks <= 0:
            _ks = float(p['cena_brutto'] or 0)
        try:
            _kat = (p['kategoria'] or 'inne').lower()
        except (KeyError, IndexError):
            _kat = 'inne'
        _pr = ALLEGRO_PROWIZJE.get(_kat, 0.11)
        _zy = _ca - _ks - (_ca * _pr) if _ca > 0 and _ks > 0 else None

        product_status = p['status'] if p['status'] else 'nowy'

        # Border & label colors by klasa/status
        try:
            _klasa = p['klasa_jakosci'] or ''
        except (IndexError, KeyError):
            _klasa = ''
        _border_colors = {'A': '#beee00', 'A-': '#8ff5ff', 'B': '#eab308', 'C': '#f97316', 'D': '#ef4444'}
        _status_colors = {'nowy': '#8ff5ff', 'wystawiony': '#ff6b9b', 'sprzedany': '#beee00', 'wyslany': '#beee00', 'uszkodzony': '#ef4444', 'zwrot': '#f59e0b'}
        _bcolor = _border_colors.get(_klasa, _status_colors.get(product_status, '#48474a'))

        # Klasa label
        _klasa_label = f'<span style="font-size:0.6rem;color:{_bcolor};text-transform:uppercase;letter-spacing:0.1em;font-weight:700;font-family:\'Space Grotesk\',sans-serif">Klasa {_klasa}</span>' if _klasa else f'<span style="font-size:0.6rem;color:#767577;text-transform:uppercase;letter-spacing:0.1em;font-weight:700;font-family:\'Space Grotesk\',sans-serif">{product_status}</span>'

        # Stock indicator
        _qty = int(p['ilosc'] or 0)
        if _qty <= 0:
            _stock_dot = f'<span style="width:7px;height:7px;border-radius:50%;background:#ef4444;display:inline-block;box-shadow:0 0 5px #ef4444"></span>'
            _stock_text = 'Brak w magazynie'
        elif _qty <= 2:
            _stock_dot = f'<span style="width:7px;height:7px;border-radius:50%;background:#ff6b9b;display:inline-block;box-shadow:0 0 5px #ff6b9b"></span>'
            _stock_text = f'Ostatnie {_qty} szt.'
        else:
            _stock_dot = f'<span style="width:7px;height:7px;border-radius:50%;background:#beee00;display:inline-block;box-shadow:0 0 5px #beee00"></span>'
            _stock_text = f'Stan: {_qty} szt.'

        # Price color
        _price_color = '#8ff5ff' if _klasa in ('A','A-') else '#f9f5f8'
        _is_sold_out = _qty <= 0
        _opacity = 'opacity:0.5;' if _is_sold_out else ''

        html += f'''
        <div class="product-item" style="position:relative;{_opacity}"
             data-name="{p['nazwa'].lower()}" data-status="{product_status}"
             data-paleta="{p['paleta'] or ''}" data-dostawca="{p['dostawca'] or ''}">
            <input type="checkbox" name="product_ids" value="{p['id']}" class="product-checkbox"
                   style="position:absolute;left:10px;top:50%;transform:translateY(-50%);width:18px;height:18px;cursor:pointer;z-index:2;accent-color:#8ff5ff"
                   onchange="updateCount()">
            <a href="/magazyn/produkt/{pcode}" style="display:flex;gap:14px;background:#131315;padding:14px 14px 14px 40px;border-radius:10px;border-left:3px solid {_bcolor};text-decoration:none;color:inherit;transition:background 0.2s"
               onmouseover="this.style.background='#1f1f22'" onmouseout="this.style.background='#131315'">
                <div style="width:80px;height:80px;background:#262528;border-radius:8px;overflow:hidden;flex-shrink:0;{'filter:grayscale(1);' if _is_sold_out else ''}">
                    <img src="{img}" style="width:100%;height:100%;object-fit:cover" onerror="this.src='data:image/svg+xml,%3Csvg xmlns=%27http://www.w3.org/2000/svg%27 width=%2790%27 height=%2790%27%3E%3Crect fill=%27%23262528%27 width=%2790%27 height=%2790%27 rx=%278%27/%3E%3Ctext x=%2745%27 y=%2752%27 fill=%27%23767577%27 text-anchor=%27middle%27 font-size=%2728%27%3E%F0%9F%93%A6%3C/text%3E%3C/svg%3E'">
                </div>
                <div style="display:flex;flex-direction:column;justify-content:space-between;flex:1;min-width:0">
                    <div>
                        <div style="display:flex;justify-content:space-between;align-items:flex-start">
                            {_klasa_label}
                            <span style="font-size:0.7rem;font-weight:700;color:#adaaad">{display_code}</span>
                        </div>
                        <div style="font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;color:#f9f5f8;line-height:1.3;margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{p['nazwa'][:45]}</div>
                    </div>
                    <div style="display:flex;justify-content:space-between;align-items:flex-end">
                        <div>
                            <div style="font-size:1.35rem;font-weight:900;color:{_price_color};font-family:'Space Grotesk',sans-serif;letter-spacing:-0.02em">{_ca:.0f} zł</div>
                            <div style="display:flex;align-items:center;gap:5px;margin-top:2px">
                                {_stock_dot}
                                <span style="font-size:0.6rem;color:#adaaad;font-weight:700;text-transform:uppercase;letter-spacing:0.06em">{_stock_text}</span>
                            </div>
                        </div>
                        {('<div style="text-align:right"><div style="font-size:0.85rem;font-weight:800;color:' + ("#beee00" if _zy >= 0 else "#ef4444") + ';font-family:Space Grotesk,sans-serif">' + format(_zy, "+.0f") + ' zł</div><div style="font-size:0.55rem;color:#767577;text-transform:uppercase;letter-spacing:0.05em;font-weight:600">zysk/szt</div></div>') if _zy is not None else ''}
                    </div>
                </div>
            </a>
        </div>'''

    html += '</div>'
    
    html += '''
    </form>
    
    <script>
    let allChecked = false;
    
    function toggleAll() {
        allChecked = !allChecked;
        const checkboxes = document.querySelectorAll('.product-checkbox');
        checkboxes.forEach(cb => cb.checked = allChecked);
        updateCount();
    }
    
    function updateCount() {
        const checked = document.querySelectorAll('.product-checkbox:checked').length;
        document.getElementById('count').textContent = checked;
    }
    
    // Zapobiegaj przypadkowemu kliknięciu checkboxa zamiast linku
    document.querySelectorAll('.product-checkbox').forEach(cb => {
        cb.addEventListener('click', (e) => e.stopPropagation());
    });

    // === BOX GROUPING ===
    function pokazBoxModal() {
        const checked = document.querySelectorAll('.product-checkbox:checked');
        if (checked.length < 2) { alert('Zaznacz minimum 2 produkty'); return; }

        let listHtml = '';
        let ids = [];
        checked.forEach(cb => {
            ids.push(cb.value);
            const card = cb.closest('.prod-item') || cb.closest('a') || cb.parentElement;
            const name = card ? card.textContent.substring(0, 60).trim() : 'Produkt #' + cb.value;
            listHtml += '<div style="padding:6px 0;border-bottom:1px solid var(--border);font-size:0.82rem;color:var(--text-secondary)">' + name.substring(0, 50) + '</div>';
        });

        document.getElementById('boxProdukty').innerHTML = listHtml;
        document.getElementById('boxIds').value = JSON.stringify(ids);
        document.getElementById('boxNazwa').value = 'Box #' + (Math.floor(Math.random()*900)+100);
        document.getElementById('boxCena').value = '';
        document.getElementById('boxCount').textContent = ids.length;
        document.getElementById('modalBox').style.display = 'flex';
    }

    function zapiszBox() {
        const ids = JSON.parse(document.getElementById('boxIds').value);
        const nazwa = document.getElementById('boxNazwa').value.trim();
        const cena = parseFloat(document.getElementById('boxCena').value) || 0;
        const cena_sprzedazy = parseFloat(document.getElementById('boxCenaSprzedazy').value) || 0;

        if (!nazwa) { alert('Podaj nazwę boxa'); return; }
        if (cena <= 0) { alert('Podaj cenę zakupu'); return; }

        const btn = document.getElementById('boxSaveBtn');
        btn.disabled = true;
        btn.textContent = ' Tworzę...';

        fetch('/magazyn/api/utworz-box', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'ngrok-skip-browser-warning': '1'},
            body: JSON.stringify({product_ids: ids, nazwa: nazwa, cena_zakupu: cena, cena_sprzedazy: cena_sprzedazy})
        })
        .then(r => r.json())
        .then(d => {
            if (d.ok) {
                btn.textContent = ' Utworzono!';
                setTimeout(() => { window.location.href = '/magazyn/paleta-id/' + d.box_id; }, 800);
            } else {
                btn.textContent = ' ' + (d.error || 'Błąd');
                btn.disabled = false;
            }
        })
        .catch(e => { btn.textContent = ' ' + e.message; btn.disabled = false; });
    }
    </script>

    <!-- Modal: Zgrupuj w Box -->
    <div id="modalBox" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.7);z-index:999;align-items:center;justify-content:center">
        <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border-radius:var(--radius);padding:25px;max-width:450px;width:90%;max-height:80vh;overflow-y:auto;border:2px solid #f59e0b">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:15px">
                <h3 style="margin:0;color:#f59e0b"><span class=material-symbols-outlined>inbox</span> Nowy Box</h3>
                <button onclick="document.getElementById('modalBox').style.display='none'" style="background:none;border:none;color:var(--text-muted);font-size:1.3rem;cursor:pointer">&times;</button>
            </div>

            <div style="margin-bottom:12px;padding:10px;background:var(--bg);border-radius:var(--radius-sm)">
                <div style="font-size:0.8rem;color:var(--text-muted);margin-bottom:5px">Produkty (<span id="boxCount">0</span> szt.):</div>
                <div id="boxProdukty" style="max-height:150px;overflow-y:auto"></div>
            </div>

            <input type="hidden" id="boxIds" value="[]">

            <div style="margin-bottom:10px">
                <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px">Nazwa boxa</label>
                <input type="text" id="boxNazwa" class="form-input" placeholder="np. Box elektronika" style="width:100%">
            </div>

            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:15px">
                <div>
                    <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px"><span class=material-symbols-outlined>paid</span> Cena zakupu (zł)</label>
                    <input type="number" id="boxCena" class="form-input" placeholder="Koszt łączny" step="0.01" min="0" style="width:100%">
                </div>
                <div>
                    <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px"><span class=material-symbols-outlined>shopping_cart</span> Cena sprzedaży (zł)</label>
                    <input type="number" id="boxCenaSprzedazy" class="form-input" placeholder="Cena Allegro" step="0.01" min="0" style="width:100%">
                </div>
            </div>

            <button id="boxSaveBtn" onclick="zapiszBox()" class="btn" style="width:100%;background:#f59e0b;color:#000;font-weight:700;padding:12px">
                <span class=material-symbols-outlined>inbox</span> Utwórz Box
            </button>
        </div>
    </div>

    <div style="text-align:center;margin-top:24px"><a href="/magazyn" style="font-size:0.82rem;color:#adaaad;text-decoration:none;font-weight:600;letter-spacing:0.05em">&larr; Powrót</a></div>
    '''
    return render(html)

@magazynier_bp.route('/produkt/<path:code>')
def produkt(code):
    conn = get_db()
    # Szukaj po EAN, ASIN lub ID
    p = get_produkt_by_code(conn, code)

    # Pobierz historię produktu (50 ostatnich wpisów)
    historia = []
    if p:
        historia = conn.execute('''
            SELECT * FROM historia_produktu
            WHERE produkt_id = ?
            ORDER BY data DESC
            LIMIT 50
        ''', (p['id'],)).fetchall()
        historia = [dict(h) for h in historia]

    msg = request.args.get('msg', '')
    is_new = p is None

    if is_new:
        p = {
            'id': 0, 'ean': code, 'asin': '', 'nazwa': f'Nowy produkt {code}', 'ilosc': 0,
            'cena_netto': 0, 'cena_brutto': 0, 'cena_allegro': 0, 'stan': 'Nowy',
            'lokalizacja': '', 'paleta': '', 'dostawca': '',
            'kategoria': 'inne', 'zdjecie_url': get_amazon_image_url(code)
        }
    else:
        p = dict(p)
        # Pobierz nazwę palety z tabeli palety
        if p.get('paleta_id') and not p.get('paleta'):
            _pal = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (p['paleta_id'],)).fetchone()
            if _pal:
                p['paleta'] = _pal['nazwa']

    product_code = get_product_code(p) if not is_new else code
    img = _resolve_product_image(p, size='lg')

    # Koszt brutto/szt = własna cena produktu (jednostkowa z importu)
    # Fallback na średnią z palety tylko gdy produkt nie ma własnej ceny
    _koszt_brutto_szt = 0
    if p.get('paleta_id'):
        _koszt_brutto_szt = _paleta_koszt_szt(conn, p['paleta_id'])
    if _koszt_brutto_szt <= 0 and p.get('cena_brutto') and p['cena_brutto'] > 0:
        _koszt_brutto_szt = float(p['cena_brutto'])

    _klasa_map = {'A':'<span style="color:#beee00">A</span>','A-':'<span style="color:#8ff5ff">A-</span>','B':'<span style="color:#eab308">B</span>','C':'<span style="color:#f97316">C</span>','D':'<span style="color:#ef4444">D</span>'}
    _klasa_display = _klasa_map.get(p.get('klasa_jakosci','') or '', '—')

    # Sibling products — ten sam produkt (EAN/nazwa) z tej samej palety, rozbite na klasy
    _siblings = []
    _ean = p.get('ean', '') or ''
    _nazwa_base = p['nazwa'][:30]
    if p.get('paleta_id'):
        if _ean and len(_ean) >= 8:
            _siblings = conn.execute(
                "SELECT id, nazwa, ilosc, stan_przyjecia, stan, lokalizacja, kod_magazynowy FROM produkty WHERE paleta_id = ? AND ean = ? AND id != ? ORDER BY stan_przyjecia",
                (p['paleta_id'], _ean, p['id'])
            ).fetchall()
        if not _siblings:
            _siblings = conn.execute(
                "SELECT id, nazwa, ilosc, stan_przyjecia, stan, lokalizacja, kod_magazynowy FROM produkty WHERE paleta_id = ? AND nazwa = ? AND id != ? ORDER BY stan_przyjecia",
                (p['paleta_id'], p['nazwa'], p['id'])
            ).fetchall()

    # Convert siblings to dicts for template dot-access
    _siblings = [dict(s) for s in _siblings]

    # Display values
    ean_display = p.get('ean') or ''
    if ean_display.upper() in ('N/A', 'NAN', 'NONE', ''):
        ean_display = '—'
    asin_display = p.get('asin') or '—'

    # Zysk per item = cena_allegro - koszt_zakupu - prowizja
    _cena_al = float(p['cena_allegro'] or 0)
    _kat = (p.get('kategoria') or 'inne').lower()
    _prowizja_rate = ALLEGRO_PROWIZJE.get(_kat, 0.11)
    _prowizja_kwota = _cena_al * _prowizja_rate
    _zysk_szt = _cena_al - _koszt_brutto_szt - _prowizja_kwota if _cena_al > 0 and _koszt_brutto_szt > 0 else 0
    _zysk_color = '#beee00' if _zysk_szt >= 0 else '#ef4444'

    # Convert p to object-like dict for template dot-access
    class DotDict(dict):
        __getattr__ = dict.get
        __setattr__ = dict.__setitem__
    p_obj = DotDict(p)

    # Przetworzone zdjęcia (Studio Foto)
    processed_photos = []
    try:
        processed_photos = [dict(r) for r in conn.execute(
            "SELECT variant, path FROM processed_photos WHERE product_id=? ORDER BY variant",
            (p['id'],)
        ).fetchall()]
    except Exception:
        pass

    return render_template('produkt_detail.html',
        p=p_obj,
        historia=historia,
        is_new=is_new,
        msg=msg,
        product_code=product_code,
        img=img,
        ean_display=ean_display,
        asin_display=asin_display,
        koszt_brutto_szt=_koszt_brutto_szt,
        klasa_display=_klasa_display,
        siblings=_siblings,
        zysk_szt=_zysk_szt,
        zysk_color=_zysk_color,
        prowizja_rate=_prowizja_rate,
        prowizja_kwota=_prowizja_kwota,
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        current_user=session.get('username'),
        processed_photos=processed_photos,
    )

# -- Old inline HTML removed, now in templates/produkt_detail.html --

@magazynier_bp.route('/produkt/<path:code>/zmien-status', methods=['POST'])
def zmien_status_produktu(code):
    """Szybka zmiana statusu produktu"""
    from .database import add_historia
    
    new_status = request.form.get('new_status', '').strip()
    
    if not new_status:
        return redirect(f'/magazyn/produkt/{code}')
    
    conn = get_db()
    p = get_produkt_by_code(conn, code)
    
    if not p:
        return redirect('/magazyn')
    
    old_status = p['status'] or 'nowy'
    
    # Nazwy statusów
    status_names = {
        'nowy': 'Magazyn',
        'wystawiony': 'Allegro',
        'sprzedany': 'Sprzedany',
        'wyslany': 'Wysłany',
        'uszkodzony': 'Uszkodzony',
        'zwrot': 'Zwrot'
    }
    
    # Aktualizuj status
    conn.execute('UPDATE produkty SET status = ? WHERE id = ?', (new_status, p['id']))
    conn.commit()
    
    # Dodaj do historii
    opis = f"Zmiana statusu: {status_names.get(old_status, old_status)} → {status_names.get(new_status, new_status)}"
    add_historia(p['id'], 'edytowano', opis, {'stary_status': old_status, 'nowy_status': new_status})
    
    product_code = get_product_code(dict(p))
    return redirect(f'/magazyn/produkt/{product_code}?msg=Status+zmieniony+na+{status_names.get(new_status, new_status)}')

@magazynier_bp.route('/produkt/<path:code>/edytuj', methods=['GET', 'POST'])
def edytuj_produkt(code):
    from .database import add_historia

    # ── POST: zapisz zmiany ──────────────────────────────────
    if request.method == 'POST':
        try:
            d = {}
            for k in ['nazwa','lokalizacja','dostawca','zdjecie_url','stan','kategoria','ean','asin','klasa_jakosci']:
                d[k] = (request.form.get(k) or '').strip()

            # Obsluga customowego dostawcy
            if d.get('dostawca') == '__custom__':
                d['dostawca'] = (request.form.get('dostawca_custom') or '').strip()
                if d['dostawca']:
                    from modules.database import save_custom_dostawca
                    save_custom_dostawca(d['dostawca'])

            d['ilosc'] = int(request.form.get('ilosc', 0) or 0)

            cena_netto_szt  = float(request.form.get('cena_netto',  0) or 0)
            cena_brutto_szt = float(request.form.get('cena_brutto', 0) or 0)

            if cena_brutto_szt == 0 and cena_netto_szt > 0:
                cena_brutto_szt = round(cena_netto_szt * 1.23, 2)
            if cena_netto_szt == 0 and cena_brutto_szt > 0:
                cena_netto_szt = round(cena_brutto_szt / 1.23, 2)

            d['cena_netto']   = cena_netto_szt   # JEDNOSTKOWA (nie × ilosc)
            d['cena_brutto']  = cena_brutto_szt  # JEDNOSTKOWA (nie × ilosc)
            d['cena_allegro'] = float(request.form.get('cena_allegro', 0) or 0)

            conn = get_db()
            existing = get_produkt_by_code(conn, code)

            if existing:
                pid = existing['id']
                product_code = str(pid)

                # Paleta — obsłuż dropdown z ID lub nową paletę
                paleta_id_val = request.form.get('paleta_id_select', '')
                paleta_nazwa = ''
                if paleta_id_val == '__nowa__':
                    paleta_nazwa = (request.form.get('paleta_nowa') or '').strip()
                    if paleta_nazwa:
                        from .database import add_paleta
                        paleta_id = add_paleta(paleta_nazwa, d.get('dostawca', ''), 0)
                    else:
                        paleta_id = existing['paleta_id']
                elif paleta_id_val:
                    paleta_id = int(paleta_id_val)
                    row = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (paleta_id,)).fetchone()
                    paleta_nazwa = row['nazwa'] if row else ''
                else:
                    paleta_id = None
                    paleta_nazwa = ''

                conn.execute('''UPDATE produkty
                    SET ean=?,asin=?,nazwa=?,ilosc=?,stan=?,lokalizacja=?,
                        paleta=?,paleta_id=?,dostawca=?,zdjecie_url=?,
                        cena_netto=?,cena_brutto=?,cena_allegro=?,kategoria=?,klasa_jakosci=?
                    WHERE id=?''',
                    (d['ean'],d['asin'],d['nazwa'],d['ilosc'],d['stan'],d['lokalizacja'],
                     paleta_nazwa,paleta_id,d['dostawca'],d['zdjecie_url'],
                     d['cena_netto'],d['cena_brutto'],d['cena_allegro'],d['kategoria'],d['klasa_jakosci'],
                     pid))
                conn.commit()

                old_cena = existing['cena_allegro'] or 0
                if old_cena != d['cena_allegro']:
                    add_historia(pid, 'zmiana_ceny',
                        f'Cena: {old_cena:.0f} → {d["cena_allegro"]:.0f} zł',
                        {'stara_cena': old_cena, 'nowa_cena': d['cena_allegro']})
                else:
                    add_historia(pid, 'edytowano', 'Edytowano produkt', {'nazwa': d['nazwa']})

            else:
                cur = conn.execute('''INSERT INTO produkty
                    (ean,asin,nazwa,ilosc,stan,lokalizacja,paleta,dostawca,
                     zdjecie_url,cena_netto,cena_brutto,cena_allegro,kategoria)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (d['ean'],d['asin'],d['nazwa'],d['ilosc'],d['stan'],d['lokalizacja'],
                     d['paleta'],d['dostawca'],d['zdjecie_url'],
                     d['cena_netto'],d['cena_brutto'],d['cena_allegro'],d['kategoria']))
                pid = cur.lastrowid
                product_code = str(pid)
                conn.commit()
                add_historia(pid, 'dodano', f'Dodano do magazynu', {'paleta': d['paleta']})

            return redirect(f'/magazyn/produkt/{product_code}?msg=Zapisano!')

        except Exception as e:
            import traceback
            return f'<div style="padding:20px;color:#ef4444;background:#1e1e2e;font-family:monospace">' \
                   f'<h2><span class=material-symbols-outlined>cancel</span> Błąd zapisu</h2><pre>{traceback.format_exc()}</pre>' \
                   f'<a href="/magazyn">← Powrót</a></div>', 500

    # ── GET: wyświetl formularz ───────────────────────────────
    conn = get_db()
    p_row = get_produkt_by_code(conn, code)
    paleta_koszt_per_szt = 0

    if p_row:
        p = dict(p_row)
        product_code = str(p['id'])
        if p.get('paleta_id'):
            paleta_koszt_per_szt = _paleta_koszt_szt(conn, p['paleta_id'])
    else:
        ean  = code if not code.startswith('B0') else ''
        asin = code if code.startswith('B0') else ''
        p = {'id':0,'ean':ean,'asin':asin,'nazwa':'','ilosc':0,
             'cena_netto':0,'cena_brutto':0,'cena_allegro':0,
             'lokalizacja':'','paleta':'','paleta_id':None,
             'dostawca':'','kategoria':'inne',
             'zdjecie_url':get_amazon_image_url(code),'stan':'Nowy','status':'magazyn'}
        product_code = code


    # Jednostkowa cena do formularza — priorytet: koszt z palety > cena_brutto (Amazon)
    try:
        ilosc_p = p.get('ilosc') or 1
        cb = p.get('cena_brutto') or 0
        cn = p.get('cena_netto') or 0
        if paleta_koszt_per_szt > 0:
            _p_brutto_szt = round(paleta_koszt_per_szt, 2)
            _p_netto_szt  = round(_p_brutto_szt / 1.23, 2)
        elif cb > 0:
            _p_brutto_szt = round(cb, 2)
            _p_netto_szt  = round(cn, 2) if cn > 0 else round(_p_brutto_szt / 1.23, 2)
        else:
            _p_brutto_szt = 0
            _p_netto_szt  = 0
        if _p_netto_szt > _p_brutto_szt > 0:
            _p_netto_szt, _p_brutto_szt = _p_brutto_szt, _p_netto_szt
    except:
        _p_brutto_szt = 0
        _p_netto_szt  = 0

    from modules.database import get_dostawcy_list
    _dlist = get_dostawcy_list()
    dostawcy_options = ''.join([
        f'<option value="{d}" {"selected" if p.get("dostawca")==d else ""}>{d}</option>'
        for d in _dlist
    ])
    dostawcy_options += '<option value="__custom__">+ Dodaj nowego...</option>'

    conn2 = get_db()
    palety_lista = conn2.execute(
        'SELECT id, nazwa FROM palety ORDER BY id DESC'
    ).fetchall()

    current_paleta_id = p.get('paleta_id') or 0
    palety_options = '<option value="">-- Brak palety --</option>'
    for pr in palety_lista:
        sel = 'selected' if pr['id'] == current_paleta_id else ''
        palety_options += f'<option value="{pr["id"]}" {sel}>{pr["nazwa"]}</option>'

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>edit</span> EDYTUJ</h1></div>
    
    <form action="/magazyn/produkt/{product_code}/edytuj" method="POST">
    <div class="card" style="padding:15px">
        <div class="form-row">
            <div class="form-group"><label>EAN</label>
                <input type="text" name="ean" class="form-ctrl" value="{p.get('ean', '')}" placeholder="EAN-13">
            </div>
            <div class="form-group"><label>ASIN</label>
                <input type="text" name="asin" class="form-ctrl" value="{p.get('asin', '')}" placeholder="B0XXXXXXXX">
            </div>
        </div>
        
        <div class="form-group"><label>Nazwa</label>
            <input type="text" name="nazwa" class="form-ctrl" value="{p['nazwa']}" required>
        </div>
        
        <div class="form-row">
            <div class="form-group"><label>Ilość</label>
                <input type="number" name="ilosc" class="form-ctrl" value="{p['ilosc']}" min="0">
            </div>
        </div>
        <div class="form-row-3">
            <div class="form-group"><label><span class=material-symbols-outlined>paid</span> Netto/szt</label>
                <input type="number" step="0.01" name="cena_netto" class="form-ctrl" value="{_p_netto_szt:.2f}">
            </div>
            <div class="form-group"><label><span class=material-symbols-outlined>paid</span> Brutto/szt</label>
                <input type="number" step="0.01" name="cena_brutto" class="form-ctrl" value="{_p_brutto_szt:.2f}">
            </div>
            <div class="form-group"><label><span class=material-symbols-outlined>payments</span> Cena Allegro</label>
                <input type="number" step="0.01" name="cena_allegro" class="form-ctrl" value="{p['cena_allegro'] or 0}">
            </div>
        </div>
        
        <div class="form-row-3">
            <div class="form-group"><label>Kategoria</label>
                <select name="kategoria" class="form-ctrl">
                    <option value="ev_ladowarki" {"selected" if p.get('kategoria')=='ev_ladowarki' else ''}>⚡ Ładowarki EV</option>
                    <option value="foto_video" {"selected" if p.get('kategoria')=='foto_video' else ''}>📷 Foto/Video</option>
                    <option value="druk3d" {"selected" if p.get('kategoria')=='druk3d' else ''}>🖨 Druk 3D</option>
                    <option value="smart_home" {"selected" if p.get('kategoria')=='smart_home' else ''}>📹 Smart Home</option>
                    <option value="motoryzacja" {"selected" if p.get('kategoria')=='motoryzacja' else ''}>🚗 Motoryzacja</option>
                    <option value="optyka" {"selected" if p.get('kategoria')=='optyka' else ''}>🔭 Optyka</option>
                    <option value="rolnictwo" {"selected" if p.get('kategoria')=='rolnictwo' else''}>🌾 Rolnictwo</option>
                    <option value="dekoracje" {"selected" if p.get('kategoria')=='dekoracje' else ''}>🌿 Dekoracje</option>
                    <option value="oswietlenie" {"selected" if p.get('kategoria')=='oswietlenie' else ''}>💡 Oświetlenie</option>
                    <option value="kuchnia" {"selected" if p.get('kategoria')=='kuchnia' else ''}>🍽 Kuchnia</option>
                    <option value="budowa" {"selected" if p.get('kategoria')=='budowa' else ''}>🔧 Budowa</option>
                    <option value="biuro" {"selected" if p.get('kategoria')=='biuro' else ''}>💼 Biuro</option>
                    <option value="outdoor" {"selected" if p.get('kategoria')=='outdoor' else ''}>🏕 Outdoor</option>
                    <option value="rehabilitacja" {"selected" if p.get('kategoria')=='rehabilitacja' else ''}>♿ Rehabilitacja</option>
                    <option value="tekstylia" {"selected" if p.get('kategoria')=='tekstylia' else ''}>🛏 Tekstylia</option>
                    <option value="kosmetyki" {"selected" if p.get('kategoria')=='kosmetyki' else ''}>💄 Kosmetyki</option>
                    <option value="ksiazki" {"selected" if p.get('kategoria')=='ksiazki' else ''}>📚 Książki</option>
                    <option value="prezenty" {"selected" if p.get('kategoria')=='prezenty' else ''}>🎁 Prezenty</option>
                    <option value="bezpieczenstwo" {"selected" if p.get('kategoria')=='bezpieczenstwo' else ''}>🔒 Bezpieczeństwo</option>
                    <option value="bagaz" {"selected" if p.get('kategoria')=='bagaz' else ''}>🧳 Bagaż</option>
                    <option value="silownia" {"selected" if p.get('kategoria')=='silownia' else ''}>🏋 Siłownia</option>
                    <option value="rowery" {"selected" if p.get('kategoria')=='rowery' else ''}>🚲 Rowery</option>
                    <option value="hulajnogi" {"selected" if p.get('kategoria')=='hulajnogi' else ''}>🛴 Hulajnogi</option>
                    <option value="elektronika" {"selected" if p.get('kategoria')=='elektronika' else ''}>📷 Elektronika</option>
                    <option value="akcesoria" {"selected" if p.get('kategoria')=='akcesoria' else ''}>🔋 Akcesoria</option>
                    <option value="agd_male" {"selected" if p.get('kategoria')=='agd_male' else ''}>🔌 AGD małe</option>
                    <option value="agd_duze" {"selected" if p.get('kategoria')=='agd_duze' else ''}>🏠 AGD duże</option>
                    <option value="komputery" {"selected" if p.get('kategoria')=='komputery' else ''}>💻 Komputery</option>
                    <option value="telefony" {"selected" if p.get('kategoria')=='telefony' else ''}>[SMARTPHONE] Telefony</option>
                    <option value="rtv" {"selected" if p.get('kategoria')=='rtv' else ''}>📺 RTV/Audio</option>
                    <option value="gaming" {"selected" if p.get('kategoria')=='gaming' else ''}>🎮 Gaming</option>
                    <option value="narzedzia" {"selected" if p.get('kategoria')=='narzedzia' else ''}>🔨 Narzędzia</option>
                    <option value="dom_ogrod" {"selected" if p.get('kategoria')=='dom_ogrod' else ''}>🏡 Dom/Ogród</option>
                    <option value="sport" {"selected" if p.get('kategoria')=='sport' else ''}>⚽ Sport</option>
                    <option value="moda" {"selected" if p.get('kategoria')=='moda' else ''}>👔 Moda</option>
                    <option value="zabawki" {"selected" if p.get('kategoria')=='zabawki' else ''}>🧸 Zabawki</option>
                    <option value="zdrowie" {"selected" if p.get('kategoria')=='zdrowie' else ''}>💊 Zdrowie</option>
                    <option value="zwierzeta" {"selected" if p.get('kategoria')=='zwierzeta' else ''}>🐾 Zwierzęta</option>
                    <option value="muzyka" {"selected" if p.get('kategoria')=='muzyka' else ''}>🎵 Muzyka</option>
                    <option value="elektronarzedzia" {"selected" if p.get('kategoria')=='elektronarzedzia' else ''}>🔧 Elektronarzędzia</option>
                    <option value="hobby" {"selected" if p.get('kategoria')=='hobby' else ''}>🎨 Hobby</option>
                    <option value="niemowleta" {"selected" if p.get('kategoria')=='niemowleta' else ''}>👶 Niemowlęta</option>
                    <option value="car_audio" {"selected" if p.get('kategoria')=='car_audio' else ''}>🔊 Car Audio</option>
                    <option value="klimatyzacja" {"selected" if p.get('kategoria')=='klimatyzacja' else ''}>🌡 Klimatyzacja</option>
                    <option value="hydroponika" {"selected" if p.get('kategoria')=='hydroponika' else ''}>🪴 Hydroponika</option>
                    <option value="wedkarstwo" {"selected" if p.get('kategoria')=='wedkarstwo' else ''}>🎣 Wędkarstwo</option>
                    <option value="laboratorium" {"selected" if p.get('kategoria')=='laboratorium' else ''}>🔬 Laboratorium</option>
                    <option value="event" {"selected" if p.get('kategoria')=='event' else ''}>🏪 Event</option>
                    <option value="cb_radio" {"selected" if p.get('kategoria')=='cb_radio' else ''}>📡 CB/Radio</option>
                    <option value="inne" {"selected" if p.get('kategoria')=='inne' or not p.get('kategoria') else ''}>📦 Inne</option>
                </select>
            </div>
            <div class="form-group"><label>Stan</label>
                <select name="stan" class="form-ctrl" onchange="var m={{'Nowy':'A','Powystawowy':'A-','Używany':'B','Uszkodzony':'C','Odnowiony':'B'}};var k=m[this.value]||'';var sel=this.closest('form').querySelector('[name=klasa_jakosci]');if(sel)sel.value=k;">
                    <option {"selected" if p.get('stan')=='Nowy' else ''}>Nowy</option>
                    <option {"selected" if p.get('stan')=='Powystawowy' else ''}>Powystawowy</option>
                    <option {"selected" if p.get('stan')=='Używany' else ''}>Używany</option>
                    <option {"selected" if p.get('stan')=='Uszkodzony' else ''}>Uszkodzony</option>
                    <option {"selected" if p.get('stan')=='Odnowiony' else ''}>Odnowiony</option>
                </select>
            </div>
            <div class="form-group"><label>Klasa jakości</label>
                <select name="klasa_jakosci" class="form-ctrl" onchange="var m={{'A':'Nowy','A-':'Powystawowy','B':'Używany','C':'Uszkodzony','D':'Uszkodzony'}};var s=m[this.value]||'';var sel=this.closest('form').querySelector('[name=stan]');if(sel&&s)sel.value=s;">
                    <option value="" {"selected" if not p.get('klasa_jakosci') else ''}>— Brak —</option>
                    <option value="A" {"selected" if p.get('klasa_jakosci')=='A' else ''}>● A — Nowy / Fabryczny</option>
                    <option value="A-" {"selected" if p.get('klasa_jakosci')=='A-' else ''}>● A- — Otwarte opakowanie</option>
                    <option value="B" {"selected" if p.get('klasa_jakosci')=='B' else ''}>● B — Używany, dobry stan</option>
                    <option value="C" {"selected" if p.get('klasa_jakosci')=='C' else ''}>● C — Widoczne ślady użytk.</option>
                    <option value="D" {"selected" if p.get('klasa_jakosci')=='D' else ''}>● D — Uszkodzony / niekompletny</option>
                </select>
            </div>
            <div class="form-group"><label>Dostawca</label>
                <select name="dostawca" class="form-ctrl" onchange="if(this.value==='__custom__'){{this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}}else{{this.nextElementSibling.style.display='none'}}"><option value="">—</option>{dostawcy_options}</select>
                <input type="text" name="dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px" class="form-ctrl">
            </div>
        </div>
        
        <div class="form-row-3">
            <div class="form-group"><label>Regał</label>
                <input type="text" name="lokalizacja" class="form-ctrl" value="{p['lokalizacja'] or ''}">
            </div>
            <div class="form-group">
                <label>Paleta</label>
                <select name="paleta_id_select" class="form-ctrl" id="paleta-select" onchange="togglePaletaInput()">
                    {palety_options}
                    <option value="__nowa__">✨ Nowa paleta...</option>
                </select>
                <input type="text" name="paleta_nowa" id="paleta-nowa" class="form-ctrl" 
                       placeholder="Wpisz nazwę nowej palety" 
                       style="display:none;margin-top:8px">
            </div>
            <div class="form-group"></div>
        </div>
        
        <script>
        function togglePaletaInput() {{
            const select = document.getElementById('paleta-select');
            const input = document.getElementById('paleta-nowa');
            if (select.value === '__nowa__') {{
                input.style.display = 'block';
                input.required = true;
            }} else {{
                input.style.display = 'none';
                input.required = false;
                input.value = '';
            }}
        }}
        // Auto-kalkulacja netto <-> brutto (VAT 23%)
        (function() {{
            const fNetto  = document.querySelector('input[name="cena_netto"]');
            const fBrutto = document.querySelector('input[name="cena_brutto"]');
            if (!fNetto || !fBrutto) return;
            fBrutto.addEventListener('input', function() {{
                const b = parseFloat(this.value);
                if (!isNaN(b) && b > 0) fNetto.value = (b / 1.23).toFixed(2);
                else fNetto.value = '';
            }});
            fNetto.addEventListener('input', function() {{
                const n = parseFloat(this.value);
                if (!isNaN(n) && n > 0) fBrutto.value = (n * 1.23).toFixed(2);
                else fBrutto.value = '';
            }});
        }})();
        </script>
        
        <div class="form-group"><label>URL zdjęcia</label>
            <input type="text" name="zdjecie_url" class="form-ctrl" value="{p['zdjecie_url'] or ''}">
        </div>
    </div>
    
    <button type="submit" class="btn btn-ok"><span class=material-symbols-outlined>save</span> ZAPISZ</button>
    </form>
    
    <form action="/magazyn/produkt/{product_code}/usun" method="POST" onsubmit="return confirm('Na pewno usunąć?')">
        <button type="submit" class="btn btn-err"><span class=material-symbols-outlined>delete</span> USUŃ</button>
    </form>
    
    <a href="/magazyn/produkt/{product_code}" class="back">← Anuluj</a>
    '''
    return render(html)

@magazynier_bp.route('/produkt/<path:code>/sprzedaj')
def sprzedaj_produkt(code):
    """Quick action: -1 szt
    
    Logika:
    - Zmniejsza ilosc o 1 w tabeli produkty
    - Zwiększa sprzedano_offline o 1 (licznik sztuk)
    - Zwiększa przychod_offline o cena_allegro (przychód na palecie)
    - NIE dodaje nic do tabeli sprzedaze (dzienne statystyki bez zmian)
    """
    from .database import add_historia
    from datetime import datetime
    
    conn = get_db()
    p = get_produkt_by_code(conn, code)
    
    if p and int(p['ilosc'] or 0) > 0:
        new_qty = int(p['ilosc']) - 1
        old_status = p['status'] or 'magazyn'
        new_status = 'sprzedany' if new_qty == 0 else old_status
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Cena za sztukę — doliczy się do przychodu palety
        cena_szt = float(p['cena_allegro'] or p['cena_brutto'] or 0)

        # Atomowy UPDATE — bez race condition (inkrementacja w SQL)
        conn.execute('''UPDATE produkty
            SET ilosc = ilosc - 1,
                status = CASE WHEN ilosc - 1 = 0 THEN 'sprzedany' ELSE status END,
                sprzedano_offline = COALESCE(sprzedano_offline, 0) + 1,
                przychod_offline = COALESCE(przychod_offline, 0) + ?,
                data_sprzedazy = CASE WHEN ilosc - 1 = 0 THEN ? ELSE data_sprzedazy END
            WHERE id = ? AND ilosc > 0''',
            (cena_szt, now_str, p['id']))
        conn.commit()

        # Dodaj do sprzedaze_prywatne (żeby było w statystykach)
        try:
            conn.execute('''INSERT INTO sprzedaze_prywatne
                (produkt_id, cena, ilosc, data_sprzedazy, kanal)
                VALUES (?, ?, 1, ?, 'offline')''',
                (p['id'], cena_szt, now_str))
            conn.commit()
        except Exception:
            pass

        add_historia(p['id'], 'sprzedano',
            f'Sprzedaz offline -1 szt. Pozostalo: {new_qty}. +{cena_szt:.0f} zl',
            {'poprzednia_ilosc': int(p['ilosc']), 'nowa_ilosc': new_qty,
             'stary_status': old_status, 'nowy_status': new_status,
             'cena_sprzedazy': cena_szt})

        if cena_szt > 0:
            msg = f'<span class=material-symbols-outlined>check_circle</span> -1 szt. (+{cena_szt:.0f} zl) Pozostalo: {new_qty} szt'
        else:
            msg = f'<span class=material-symbols-outlined>check_circle</span> -1 szt. Pozostalo: {new_qty} szt'
    else:
        msg = '<span class=material-symbols-outlined>cancel</span> Brak na stanie!'

    product_code = get_product_code(p) if p else code
    return redirect(f'/magazyn/produkt/{product_code}?msg={msg}')

@magazynier_bp.route('/produkt/<path:code>/usun', methods=['POST'])
def usun_produkt(code):
    conn = get_db()
    
    # Pobierz ID produktu przed usunięciem
    p = get_produkt_by_code(conn, code)
    
    if p:
        product_id = p['id']
        
        # Usuń historię produktu
        conn.execute('DELETE FROM historia_produktu WHERE produkt_id=?', (product_id,))
        
        # Usuń produkt
        conn.execute('DELETE FROM produkty WHERE id=?', (product_id,))
        
        conn.commit()
    
    return redirect('/magazyn?msg=Produkt+usunięty')

@magazynier_bp.route('/produkt/<path:code>/opis')
def produkt_opis(code):
    """Generuje opis AI dla produktu"""
    from .utils import generuj_opis_ai
    
    conn = get_db()
    p = get_produkt_by_code(conn, code)
    
    if not p:
        return redirect('/magazyn')
    
    product_code = get_product_code(p)
    _ean_c = p['ean'] if p['ean'] and p['ean'].upper() not in ('N/A','NAN','NONE') else ''
    display_code = _ean_c or p['asin'] or f"#{p['id']}"
    opis = generuj_opis_ai(p['nazwa'], p['kategoria'] or 'inne')
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>auto_awesome</span> OPIS AI</h1><small>{display_code}</small></div>
    
    <div class="card" style="padding:15px">
        <div style="font-weight:600;margin-bottom:10px">{p['nazwa']}</div>
        <div style="background:#0a0a0f;border-radius:10px;padding:15px;white-space:pre-wrap;font-size:0.9rem;line-height:1.6;max-height:300px;overflow-y:auto">{opis}</div>
    </div>
    
    <button onclick="navigator.clipboard.writeText(document.querySelector('div[style*=pre-wrap]').innerText);this.innerText='<span class=material-symbols-outlined>check_circle</span> Skopiowano!';setTimeout(()=>this.innerText='<span class=material-symbols-outlined>list_alt</span> KOPIUJ DO SCHOWKA',2000)" class="btn btn-ok"><span class=material-symbols-outlined>list_alt</span> KOPIUJ DO SCHOWKA</button>
    
    <a href="/magazyn/produkt/{product_code}" class="back">← Powrót do produktu</a>
    '''
    return render(html)

@magazynier_bp.route('/szukaj')
def szukaj():
    q = request.args.get('q', '').strip()
    if not q:
        return redirect('/magazyn')
    
    # Obsługa kodów QR z prefiksem MAG: (format: MAG:B0B6NDX3SS)
    if q.upper().startswith('MAG:'):
        q = q[4:]  # Usuń prefiks "MAG:" - zostaje sam kod np. "B0B6NDX3SS"

    conn = get_db()

    # Szukaj po kodzie magazynowym (MAG-XXXXX)
    if q.upper().startswith('MAG-'):
        results = conn.execute('SELECT * FROM produkty WHERE kod_magazynowy=?', (q.upper(),)).fetchall()
    elif is_code(q):
        # Szukaj po EAN lub ASIN
        results = conn.execute('SELECT * FROM produkty WHERE ean=? OR asin=?', (q.upper(), q.upper())).fetchall()
    else:
        # Szukaj po nazwie lub kodzie magazynowym
        results = conn.execute('SELECT * FROM produkty WHERE nazwa LIKE ? OR kod_magazynowy LIKE ?', (f'%{q}%', f'%{q.upper()}%')).fetchall()
    
    if len(results) == 1:
        return redirect(f'/magazyn/produkt/{get_product_code(results[0])}')
    
    from markupsafe import escape as _esc
    html = f'''<div class="hdr"><h1><span class=material-symbols-outlined>search</span> WYNIKI</h1><small>"{_esc(q)}"</small></div>'''
    
    for r in results:
        img = _resolve_product_image(r, size='sm')
        pcode = get_product_code(r)
        display_code = r['ean'] or r['asin'] or f"#{r['id']}"
        html += f'''<a href="/magazyn/produkt/{pcode}" class="item">
            <img src="{img}" onerror="this.src='{_PLACEHOLDER_IMG_SM}'">
            <div class="item-info">
                <div class="item-name">{r['nazwa']}</div>
                <div class="item-meta">{display_code}</div>
            </div>
            <div class="item-qty">{r['ilosc']}</div>
        </a>'''
    
    if not results:
        html += '<div class="alert alert-warn">Brak wyników</div>'
        if is_code(q):
            html += f'<a href="/magazyn/produkt/{q}" class="btn btn-ok"><span class=material-symbols-outlined>add</span> DODAJ NOWY</a>'
    
    html += '<a href="/magazyn" class="back">← Powrót</a>'
    return render(html)

@magazynier_bp.route('/backup')
def backup_page():
    """Strona zarządzania backupami"""
    from modules.backup_manager import get_backups, create_backup, verify_backup
    
    backups = get_backups()
    
    html = '''
    <div class="hdr">
        <h1><span class=material-symbols-outlined>save</span> BACKUP & PRZYWRACANIE</h1>
        <small>Zarządzanie kopiami zapasowymi bazy danych</small>
    </div>
    
    <div class="card" style="padding:20px;margin-bottom:15px;background:rgba(190,238,0,0.1);border:2px solid #beee00">
        <div style="display:flex;align-items:center;gap:15px">
            <div style="font-size:2.5rem"><span class=material-symbols-outlined>save</span></div>
            <div style="flex:1">
                <div style="font-weight:600;font-size:1.1rem;margin-bottom:5px">Automatyczne backupy</div>
                <div style="font-size:0.9rem;opacity:0.8">System tworzy backup bazy co godzinę automatycznie</div>
                <div style="font-size:0.85rem;opacity:0.7;margin-top:3px">Przechowywane jest 24 ostatnie backupy (1 dzień)</div>
            </div>
        </div>
    </div>
    
    <div style="display:flex;gap:10px;margin-bottom:20px">
        <button onclick="createBackup()" class="btn btn-ok" style="flex:1">
            <span class=material-symbols-outlined>save</span> Utwórz backup teraz
        </button>
    </div>
    
    <!-- WGRYWANIE ZEWNĘTRZNEGO BACKUPU -->
    <div class="card" style="padding:20px;margin-bottom:20px;background:rgba(249,115,22,0.1);border:2px solid #f97316">
        <div style="font-weight:600;font-size:1.1rem;margin-bottom:10px;color:#f97316"><span class=material-symbols-outlined>upload</span> Wgraj zewnętrzny backup</div>
        <div style="font-size:0.85rem;color:#94a3b8;margin-bottom:15px">
            Możesz wgrać stary plik bazy danych (.db) z komputera
        </div>
        <form action="/magazyn/backup/upload" method="POST" enctype="multipart/form-data" id="uploadBackupForm">
            <div style="display:flex;gap:10px;align-items:center">
                <div style="flex:1;position:relative">
                    <input type="file" id="backupFile" name="backup_file" accept=".db" style="display:none" onchange="updateFileName()">
                    <div onclick="document.getElementById('backupFile').click()" 
                         style="padding:12px 15px;background:#0a0a0f;border:1px dashed #f97316;border-radius:8px;cursor:pointer;text-align:center;color:#94a3b8">
                        <span id="fileNameDisplay"><span class=material-symbols-outlined>folder</span> Kliknij aby wybrać plik .db</span>
                    </div>
                </div>
                <button type="submit" class="btn" style="background:#f97316;padding:12px 20px" id="uploadBtn" disabled>
                    [ARROW_UPWARD] Wgraj
                </button>
            </div>
        </form>
    </div>
    
    <div class="section"><span class=material-symbols-outlined>list_alt</span> DOSTĘPNE BACKUPY</div>
    '''
    
    if not backups:
        html += '''
        <div class="card" style="padding:30px;text-align:center">
            <div style="font-size:3rem;opacity:0.3;margin-bottom:10px"><span class=material-symbols-outlined>inventory_2</span></div>
            <div style="opacity:0.6">Brak backupów</div>
        </div>
        '''
    else:
        for backup in backups:
            # Weryfikacja backupu
            is_ok, status_msg = verify_backup(backup['filename'])
            status_icon = "<span class=material-symbols-outlined>check_circle</span>" if is_ok else "<span class=material-symbols-outlined>cancel</span>"
            status_color = "#beee00" if is_ok else "#ef4444"
            
            html += f'''
            <div class="card" style="padding:15px;margin-bottom:10px">
                <div style="display:flex;align-items:center;gap:15px">
                    <div style="font-size:2rem">{status_icon}</div>
                    <div style="flex:1">
                        <div style="font-weight:600;margin-bottom:3px">{backup['filename']}</div>
                        <div style="font-size:0.85rem;color:var(--text-dim)">
                            <span class=material-symbols-outlined>calendar_month</span> {backup['created_str']} | 
                            <span class=material-symbols-outlined>save</span> {backup['size_mb']:.2f} MB | 
                            <span style="color:{status_color}">{status_msg}</span>
                        </div>
                    </div>
                    <div style="display:flex;gap:8px">
                        <button onclick="restoreBackup('{backup['filename']}')" class="btn" style="background:var(--purple);padding:8px 15px;font-size:0.85rem">
                            [UNDO] Przywróć
                        </button>
                    </div>
                </div>
            </div>
            '''
    
    html += '''
    <a href="/magazyn" class="back">← Powrót</a>
    
    <script>
    function updateFileName() {
        var input = document.getElementById('backupFile');
        var display = document.getElementById('fileNameDisplay');
        var btn = document.getElementById('uploadBtn');
        
        if (input.files.length > 0) {
            var file = input.files[0];
            var sizeMB = (file.size / 1024 / 1024).toFixed(2);
            display.innerHTML = '<span class=material-symbols-outlined>folder</span> ' + file.name + ' (' + sizeMB + ' MB)';
            display.style.color = '#f97316';
            btn.disabled = false;
        } else {
            display.innerHTML = '<span class=material-symbols-outlined>folder</span> Kliknij aby wybrać plik .db';
            display.style.color = '#94a3b8';
            btn.disabled = true;
        }
    }
    
    document.getElementById('uploadBackupForm').onsubmit = function(e) {
        if (!confirm('UWAGA!\\n\\nTo zastąpi aktualną bazę danych wgranym plikiem.\\nAktualna baza zostanie najpierw zbackupowana.\\n\\nKontynuować?')) {
            e.preventDefault();
            return false;
        }
        document.getElementById('uploadBtn').disabled = true;
        document.getElementById('uploadBtn').innerHTML = '⏳ Wgrywanie...';
    };
    
    function createBackup() {
        if (!confirm('Utworzyć backup bazy danych?')) return;
        
        showLoading('Tworzenie backupu...');
        
        fetch('/api/backup/create', { method: 'POST' })
            .then(r => r.json())
            .then(data => {
                hideLoading();
                if (data.success) {
                    showToast('Sukces', 'Backup utworzony: ' + data.backup, 'success');
                    setTimeout(() => location.reload(), 1000);
                } else {
                    showToast('Błąd', data.error || 'Nieznany błąd', 'error');
                }
            })
            .catch(err => {
                hideLoading();
                showToast('Błąd', 'Błąd połączenia: ' + err, 'error');
            });
    }
    
    function restoreBackup(filename) {
        if (!confirm('UWAGA! To przywróci bazę danych z backupu.\\n\\nObecna baza zostanie zastąpiona.\\n\\nKontynuować?')) {
            return;
        }
        
        showLoading('Przywracanie backupu...');
        
        fetch('/api/backup/restore', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ filename: filename })
        })
            .then(r => {
                if (!r.ok) {
                    throw new Error('HTTP ' + r.status);
                }
                return r.json();
            })
            .then(data => {
                hideLoading();
                if (data.success) {
                    showToast('Sukces', data.message, 'success');
                    setTimeout(() => location.href = '/magazyn', 2000);
                } else {
                    showToast('Błąd', data.message || data.error || 'Nieznany błąd', 'error');
                }
            })
            .catch(err => {
                hideLoading();
                showToast('Błąd', 'Błąd połączenia: ' + err.message, 'error');
            });
    }
    </script>
    '''
    
    return render(html)

@magazynier_bp.route('/backup/upload', methods=['POST'])
def backup_upload():
    """Wgrywanie zewnętrznego backupu bazy danych"""
    import sqlite3
    
    if 'backup_file' not in request.files:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')
    
    file = request.files['backup_file']
    if file.filename == '':
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')
    
    if not file.filename.lower().endswith('.db'):
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Plik musi mieć rozszerzenie .db</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')
    
    try:
        # Ścieżki
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        backups_dir = os.path.join(base_dir, 'backups')
        
        # Utwórz folder backups jeśli nie istnieje
        os.makedirs(backups_dir, exist_ok=True)
        
        # Zapisz wgrany plik jako backup (z prefiksem uploaded_)
        backup_filename = f'uploaded_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        backup_path = os.path.join(backups_dir, backup_filename)
        file.save(backup_path)
        
        # Weryfikuj że to poprawna baza SQLite
        try:
            test_conn = sqlite3.connect(backup_path)
            # Sprawdź czy ma tabelę produkty
            tables = test_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
            table_names = [t[0] for t in tables]
            
            # Policz produkty
            count = 0
            if 'produkty' in table_names:
                count = test_conn.execute('SELECT COUNT(*) FROM produkty').fetchone()[0]
            
            test_conn.close()
            
            if 'produkty' not in table_names:
                os.remove(backup_path)
                return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Plik nie zawiera tabeli produkty - to nie jest baza {get_config_cached("brand_name", "Akces Hub")}</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')
            
        except sqlite3.DatabaseError as e:
            os.remove(backup_path)
            return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Plik nie jest poprawną bazą SQLite: {str(e)}</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')
        
        # Rozmiar pliku
        size_mb = os.path.getsize(backup_path) / 1024 / 1024
        
        html = f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> BACKUP WGRANY</h1></div>
        
        <div class="alert alert-ok" style="margin-bottom:15px">
            Plik został dodany do listy backupów!
        </div>
        
        <div class="card" style="padding:20px;text-align:center">
            <div style="font-size:3rem;margin-bottom:10px"><span class=material-symbols-outlined>inventory_2</span></div>
            <div style="font-weight:600;margin-bottom:5px">{backup_filename}</div>
            <div style="font-size:1.3rem;color:#beee00">{count} produktów</div>
            <div style="font-size:0.85rem;color:#64748b">{size_mb:.2f} MB</div>
        </div>
        
        <div class="card" style="padding:15px;margin-top:15px;background:rgba(249,115,22,0.1);border:1px solid #f97316">
            <div style="font-size:0.9rem;color:#f97316">
                <span class=material-symbols-outlined>warning</span> Aby aktywować ten backup, kliknij <strong>"[UNDO] Przywróć"</strong> przy nim na liście backupów
            </div>
        </div>
        
        <a href="/magazyn/backup" class="btn btn-ok" style="margin-top:20px"><span class=material-symbols-outlined>save</span> Przejdź do listy backupów</a>
        <a href="/magazyn/backup" class="back">← Powrót</a>
        '''
        return render(html)
        
    except Exception as e:
        return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">{str(e)}</div><a href="/magazyn/backup" class="btn btn-p">← Powrót</a>')

@magazynier_bp.route('/statystyki')
def statystyki():
    """Statystyki sprzedaży z wykresami"""
    import json
    conn = get_db()
    
    # FILTR: tylko opłacone (bez zwrotów i anulowanych)
    # STATUS_FILTER inlined in each query to avoid f-string SQL (B608)

    # Sprzedaż miesięcznie (bieżący rok) — Allegro + prywatne
    current_year = datetime.now().year
    miesieczne_raw = conn.execute('''
        SELECT strftime('%m', REPLACE(SUBSTR(data_sprzedazy,1,19), 'T', ' ')) as miesiac,
               COUNT(*) as ilosc,
               COALESCE(SUM(cena * ilosc + COALESCE(koszt_dostawy, 0)), 0) as suma
        FROM sprzedaze
        WHERE strftime('%Y', REPLACE(SUBSTR(data_sprzedazy,1,19), 'T', ' ')) = ?
          AND status NOT IN ('zwrot', 'anulowane', 'anulowana')
          AND data_sprzedazy IS NOT NULL AND data_sprzedazy != ''
        GROUP BY miesiac
        HAVING miesiac IS NOT NULL
        ORDER BY miesiac
    ''', (str(current_year),)).fetchall()

    # Dołącz sprzedaze_prywatne do sum miesięcznych
    miesieczne_dict = {}
    for row in miesieczne_raw:
        m = row['miesiac']
        miesieczne_dict[m] = {'miesiac': m, 'ilosc': row['ilosc'], 'suma': float(row['suma'])}
    try:
        pryw_miesiac = conn.execute('''
            SELECT strftime('%m', data) as miesiac,
                   COUNT(*) as ilosc,
                   COALESCE(SUM(kwota), 0) as suma
            FROM sprzedaze_prywatne
            WHERE strftime('%Y', data) = ?
            GROUP BY miesiac
        ''', (str(current_year),)).fetchall()
        for pm in pryw_miesiac:
            m = pm['miesiac']
            if m in miesieczne_dict:
                miesieczne_dict[m]['ilosc'] += pm['ilosc']
                miesieczne_dict[m]['suma'] += float(pm['suma'])
            else:
                miesieczne_dict[m] = {'miesiac': m, 'ilosc': pm['ilosc'], 'suma': float(pm['suma'])}
    except Exception:
        pass
    miesieczne = sorted(miesieczne_dict.values(), key=lambda x: x['miesiac'])

    # Sprzedaż dziennie dla każdego miesiąca (do drill-down) - z ilością zamówień!
    dzienne_all = conn.execute('''
        SELECT
            strftime('%m', REPLACE(SUBSTR(data_sprzedazy,1,19), 'T', ' ')) as miesiac,
            strftime('%d', REPLACE(SUBSTR(data_sprzedazy,1,19), 'T', ' ')) as dzien,
            COUNT(*) as cnt,
            COALESCE(SUM(cena * ilosc), 0) as suma
        FROM sprzedaze
        WHERE strftime('%Y', REPLACE(SUBSTR(data_sprzedazy,1,19), 'T', ' ')) = ?
          AND status NOT IN ('zwrot', 'anulowane', 'anulowana')
          AND data_sprzedazy IS NOT NULL
          AND data_sprzedazy != ''
        GROUP BY miesiac, dzien
        HAVING miesiac IS NOT NULL AND dzien IS NOT NULL
        ORDER BY miesiac, dzien
    ''', (str(current_year),)).fetchall()
    
    # Przygotuj dane dzienne per miesiąc (suma + ilość)
    dzienne_per_miesiac = {}
    dzienne_cnt_per_miesiac = {}
    for d in dzienne_all:
        m = int(d['miesiac'])
        if m not in dzienne_per_miesiac:
            dzienne_per_miesiac[m] = {}
            dzienne_cnt_per_miesiac[m] = {}
        dzien = int(d['dzien'])
        dzienne_per_miesiac[m][dzien] = float(d['suma'])
        dzienne_cnt_per_miesiac[m][dzien] = int(d['cnt'])

    # Dodaj sprzedaze_prywatne do dziennych danych (żeby luty i inne mies. były kompletne)
    try:
        pryw_dzienne = conn.execute('''
            SELECT strftime('%m', data) as miesiac,
                   CAST(strftime('%d', data) AS INTEGER) as dzien,
                   SUM(kwota) as suma, COUNT(*) as cnt
            FROM sprzedaze_prywatne
            WHERE strftime('%Y', data) = ?
            GROUP BY miesiac, dzien
        ''', (str(current_year),)).fetchall()
        for pd in pryw_dzienne:
            m = int(pd['miesiac'])
            dzien = int(pd['dzien'])
            if m not in dzienne_per_miesiac:
                dzienne_per_miesiac[m] = {}
                dzienne_cnt_per_miesiac[m] = {}
            dzienne_per_miesiac[m][dzien] = dzienne_per_miesiac[m].get(dzien, 0) + float(pd['suma'])
            dzienne_cnt_per_miesiac[m][dzien] = dzienne_cnt_per_miesiac[m].get(dzien, 0) + int(pd['cnt'])
    except Exception:
        pass
    
    # Sprzedaż rocznie (wszystkie lata) — z prywatnym (spójne z dashboardem)
    roczne = conn.execute('''
        SELECT rok, SUM(ilosc) as ilosc, SUM(suma) as suma FROM (
            SELECT strftime('%Y', data_sprzedazy) as rok,
                   COUNT(*) as ilosc,
                   COALESCE(SUM(cena * ilosc), 0) as suma
            FROM sprzedaze
            WHERE status NOT IN ('zwrot', 'anulowane', 'anulowana')
            GROUP BY rok
            UNION ALL
            SELECT strftime('%Y', data) as rok,
                   COUNT(*) as ilosc,
                   COALESCE(SUM(kwota), 0) as suma
            FROM sprzedaze_prywatne
            GROUP BY rok
        ) GROUP BY rok ORDER BY rok
    ''').fetchall()
    
    # Podsumowanie ogólne
    podsumowanie = conn.execute('''
        SELECT COUNT(*) as ilosc_transakcji,
               COALESCE(SUM(cena * ilosc), 0) as suma_total,
               COALESCE(AVG(cena * ilosc), 0) as srednia
        FROM sprzedaze
        WHERE status NOT IN ('zwrot', 'anulowane', 'anulowana')
    ''').fetchone()
    if podsumowanie is None:
        import sqlite3
        podsumowanie = sqlite3.Row.__new__(sqlite3.Row)
        podsumowanie = {'ilosc_transakcji': 0, 'suma_total': 0.0, 'srednia': 0.0}
    
    # Top 5 produktów (najczęściej sprzedawane) - bierze nazwę z sprzedaze
    top_produkty = conn.execute('''
        SELECT
            CASE
                WHEN s.nazwa IS NOT NULL AND s.nazwa != '' AND s.nazwa != 'Produkt' THEN SUBSTR(s.nazwa, 1, 50)
                WHEN o.tytul IS NOT NULL AND o.tytul != '' THEN SUBSTR(o.tytul, 1, 50)
                WHEN p.nazwa IS NOT NULL AND p.nazwa != '' THEN p.nazwa
                ELSE 'Produkt #' || s.id
            END as produkt_nazwa,
            COUNT(*) as sprzedane,
            SUM(s.cena * s.ilosc) as przychod
        FROM sprzedaze s
        LEFT JOIN oferty o ON s.oferta_id = o.id
        LEFT JOIN produkty p ON COALESCE(s.produkt_id, o.produkt_id) = p.id
        WHERE s.status NOT IN ('zwrot', 'anulowane', 'anulowana')
        GROUP BY produkt_nazwa
        ORDER BY sprzedane DESC
        LIMIT 5
    ''').fetchall()
    
    # Top dostawcy - pobiera z palety przez produkt lub ofertę
    top_dostawcy = conn.execute('''
        SELECT
            COALESCE(pal.dostawca, pal2.dostawca, 'Allegro') as dostawca_nazwa,
            COUNT(*) as sprzedane,
            SUM(s.cena * s.ilosc) as przychod
        FROM sprzedaze s
        LEFT JOIN produkty p ON s.produkt_id = p.id
        LEFT JOIN palety pal ON p.paleta_id = pal.id
        LEFT JOIN oferty o ON s.oferta_id = o.id
        LEFT JOIN produkty p2 ON o.produkt_id = p2.id
        LEFT JOIN palety pal2 ON p2.paleta_id = pal2.id
        WHERE s.status NOT IN ('zwrot', 'anulowane', 'anulowana')
        GROUP BY dostawca_nazwa
        ORDER BY przychod DESC
        LIMIT 5
    ''').fetchall()
    
    # Koszty miesięczne (bieżący rok)
    try:
        conn.execute('CREATE TABLE IF NOT EXISTS koszty (id INTEGER PRIMARY KEY AUTOINCREMENT, nazwa TEXT, kwota REAL, kategoria TEXT, data DATE, notatka TEXT)')
        koszty_msc_rows = conn.execute('''
            SELECT strftime('%m', data) as m, SUM(kwota) as suma
            FROM koszty WHERE strftime('%Y', data) = ?
            GROUP BY m
        ''', (str(current_year),)).fetchall()
        koszty_per_msc = {int(r['m']): float(r['suma']) for r in koszty_msc_rows}
        koszty_total_rok = sum(koszty_per_msc.values())
    except:
        koszty_per_msc = {}
        koszty_total_rok = 0
    
    # Sprzedaż prywatna miesięcznie (bieżący rok)
    try:
        conn.execute('CREATE TABLE IF NOT EXISTS sprzedaze_prywatne (id INTEGER PRIMARY KEY AUTOINCREMENT, opis TEXT, kwota REAL, data DATE, notatka TEXT)')
        pryw_msc_rows = conn.execute('''
            SELECT strftime('%m', data) as m, SUM(kwota) as suma
            FROM sprzedaze_prywatne WHERE strftime('%Y', data) = ?
            GROUP BY m
        ''', (str(current_year),)).fetchall()
        pryw_per_msc = {int(r['m']): float(r['suma']) for r in pryw_msc_rows}
        pryw_total_rok = sum(pryw_per_msc.values())
    except:
        pryw_per_msc = {}
        pryw_total_rok = 0

    # COGS per miesiąc — koszt SPRZEDANYCH produktów (nie kupionych palet)
    # koszt_per_szt = paleta.cena_zakupu / łączna_ilość_sztuk_z_palety
    try:
        cogs_msc_rows = conn.execute('''
            SELECT strftime('%m', s.data_sprzedazy) as m,
                COALESCE(SUM(
                    CASE
                        WHEN pal.cena_zakupu > 0 AND pal_total.total_szt > 0
                        THEN (pal.cena_zakupu / pal_total.total_szt) * s.ilosc
                        ELSE 0
                    END
                ), 0) as cogs
            FROM sprzedaze s
            LEFT JOIN produkty p ON s.produkt_id = p.id
            LEFT JOIN palety pal ON p.paleta_id = pal.id
            LEFT JOIN (
                SELECT pr.paleta_id,
                    COALESCE(SUM(pr.ilosc), 0)
                    + COALESCE(SUM(pr.sprzedano_offline), 0)
                    + COALESCE((
                        SELECT SUM(sp2.ilosc) FROM sprzedaze sp2
                        JOIN produkty pp2 ON sp2.produkt_id = pp2.id
                        WHERE pp2.paleta_id = pr.paleta_id
                        AND sp2.status NOT IN ('zwrot','anulowane','anulowana')
                    ), 0) as total_szt
                FROM produkty pr GROUP BY pr.paleta_id
            ) pal_total ON pal_total.paleta_id = pal.id
            WHERE strftime('%Y', s.data_sprzedazy) = ?
            AND s.status NOT IN ('zwrot', 'anulowane', 'anulowana')
            GROUP BY m
        ''', (str(current_year),)).fetchall()
        palety_per_msc = {int(r['m']): float(r['cogs'] or 0) for r in cogs_msc_rows}
        palety_total_rok = sum(palety_per_msc.values())
    except:
        palety_per_msc = {}
        palety_total_rok = 0

    # Ilość palet kupionych + kwota zakupu (do wyświetlania)
    try:
        palety_cnt_rows = conn.execute('''
            SELECT strftime('%m', data_zakupu) as m, COUNT(*) as cnt,
                   COALESCE(SUM(cena_zakupu), 0) as suma_zakupu
            FROM palety WHERE strftime('%Y', data_zakupu) = ?
            AND data_zakupu IS NOT NULL AND data_zakupu != ''
            GROUP BY m
        ''', (str(current_year),)).fetchall()
        palety_cnt_per_msc = {int(r['m']): int(r['cnt']) for r in palety_cnt_rows}
        palety_zakup_per_msc = {int(r['m']): float(r['suma_zakupu']) for r in palety_cnt_rows}
        palety_total_cnt_rok = sum(palety_cnt_per_msc.values())
        palety_zakup_total_rok = sum(palety_zakup_per_msc.values())
    except:
        palety_cnt_per_msc = {}
        palety_zakup_per_msc = {}
        palety_total_cnt_rok = 0
        palety_zakup_total_rok = 0

    # ROI per paleta - ta sama logika co /analityka (proporcjonalny koszt sprzedanych)
    try:
        palety_roi_rows = conn.execute('''
            SELECT
                p.id,
                p.nazwa,
                p.dostawca,
                COALESCE(p.cena_zakupu, 0) as koszt_palety,
                COALESCE(p.ilosc_produktow, 0) as ilosc_produktow,
                (SELECT COALESCE(SUM(ilosc), 0) FROM produkty WHERE paleta_id = p.id) as aktualna_ilosc,
                (SELECT COALESCE(SUM(CASE WHEN status = 'sprzedany' AND (sprzedano_offline IS NULL OR sprzedano_offline = 0) THEN cena_allegro ELSE 0 END), 0) FROM produkty WHERE paleta_id = p.id) as przychod_produkty,
                COALESCE((SELECT SUM(s.cena * s.ilosc) FROM sprzedaze s JOIN produkty pr ON s.produkt_id = pr.id WHERE pr.paleta_id = p.id AND s.status NOT IN ('anulowana', 'zwrot') AND (s.kupujacy IS NULL OR s.kupujacy != 'offline')), 0) as przychod_tabela,
                (SELECT COALESCE(SUM(przychod_offline), 0) FROM produkty WHERE paleta_id = p.id) as przychod_offline,
                (SELECT COALESCE(SUM(sprzedano_offline), 0) FROM produkty WHERE paleta_id = p.id) as sprzedano_offline_szt,
                (SELECT COUNT(*) FROM produkty WHERE paleta_id = p.id AND status = 'sprzedany' AND (sprzedano_offline IS NULL OR sprzedano_offline = 0)) as sprzedano_produkty,
                COALESCE((SELECT SUM(s.ilosc) FROM sprzedaze s JOIN produkty pr ON s.produkt_id = pr.id WHERE pr.paleta_id = p.id AND s.status NOT IN ('anulowana', 'zwrot') AND (s.kupujacy IS NULL OR s.kupujacy != 'offline')), 0) as sprzedano_tabela
            FROM palety p
            ORDER BY p.id
        ''').fetchall()
        palety_roi = []
        for r in palety_roi_rows:
            koszt_palety = float(r['koszt_palety'] or 0)
            if koszt_palety <= 0:
                continue
            # Przychód - tylko z realnych transakcji (tabela sprzedaze) + offline
            # NIE używamy cena_allegro z produktów bo to cena wywoławcza, nie rzeczywista sprzedaż
            przychod_tabela = float(r['przychod_tabela'] or 0)
            przychod_offline = float(r['przychod_offline'] or 0)
            przychod = przychod_tabela + przychod_offline
            # Sprzedane sztuki - z tabeli sprzedaze (realne transakcje) + offline
            sprzedano_allegro = int(r['sprzedano_tabela'] or 0)
            sprzedano_offline = int(r['sprzedano_offline_szt'] or 0)
            sprzedanych = sprzedano_allegro + sprzedano_offline
            aktualna_ilosc = int(r['aktualna_ilosc'] or 0)
            ilosc_produktow = int(r['ilosc_produktow'] or 0)
            # Total: użyj ilosc_produktow z palety jeśli sensowne (> sprzedanych)
            # Fallback na aktualna + sprzedanych gdy ilosc_produktow nie wpisane lub za małe
            if ilosc_produktow > 0 and ilosc_produktow >= sprzedanych:
                total = ilosc_produktow
            else:
                total = aktualna_ilosc + sprzedanych
            if sprzedanych == 0 or przychod == 0:
                continue
            # Proporcjonalny koszt = koszt_palety * (sprzedane / wszystkie)
            udzial = (sprzedanych / total) if total > 0 else 1
            koszt_sprzedanych = koszt_palety * udzial
            zysk_pal = przychod - koszt_sprzedanych
            roi_pal = (zysk_pal / koszt_sprzedanych * 100)
            palety_roi.append({
                'nazwa': (r['nazwa'] or r['dostawca'] or f"Paleta #{r['id']}"),
                'koszt': koszt_sprzedanych,
                'koszt_palety': koszt_palety,
                'przychod': przychod,
                'zysk': zysk_pal,
                'roi': roi_pal,
                'sprzedane': sprzedanych,
                'total': total
            })
    except Exception as _e:
        palety_roi = []

    # ===== HISTOGRAM CZASU SPRZEDAŻY (od wystawienia do sprzedaży) =====
    sell_time_histogram = [0, 0, 0, 0, 0, 0, 0, 0]
    sell_time_labels = ['≤1 dzień', '2-3 dni', '4-7 dni', '1-2 tyg', '2-4 tyg', '1-2 mies', '2-3 mies', '3+ mies']
    najszybciej_sprzedane = []
    try:
        sell_times_raw = conn.execute('''
            SELECT 
                COALESCE(NULLIF(p.nazwa,''), NULLIF(s.nazwa,''), CASE WHEN s.allegro_order_id IS NOT NULL THEN 'Zamówienie #' || SUBSTR(s.allegro_order_id,-6) ELSE 'Brak nazwy' END) as nazwa,
                s.cena,
                s.data_sprzedazy,
                COALESCE(o.data_wystawienia, p.data_dodania) as data_od
            FROM sprzedaze s
            LEFT JOIN produkty p ON s.produkt_id = p.id
            LEFT JOIN oferty o ON s.oferta_id = o.id
            WHERE s.status NOT IN ('zwrot', 'anulowane', 'anulowana')
              AND s.data_sprzedazy IS NOT NULL AND s.data_sprzedazy != ''
            ORDER BY s.data_sprzedazy DESC LIMIT 500
        ''').fetchall()
        for row in sell_times_raw:
            try:
                dt_sprzed = datetime.fromisoformat(str(row['data_sprzedazy'])[:19].replace('T', ' '))
                data_od_val = row['data_od']
                if not data_od_val:
                    continue  # brak daty dodania — pomijamy w histogramie
                dt_od = datetime.fromisoformat(str(data_od_val)[:19].replace('T', ' '))
                diff_hours = (dt_sprzed - dt_od).total_seconds() / 3600
                if diff_hours < 0:
                    continue
                diff_days = diff_hours / 24
                if diff_days <= 1:       sell_time_histogram[0] += 1
                elif diff_days <= 3:     sell_time_histogram[1] += 1
                elif diff_days <= 7:     sell_time_histogram[2] += 1
                elif diff_days <= 14:    sell_time_histogram[3] += 1
                elif diff_days <= 28:    sell_time_histogram[4] += 1
                elif diff_days <= 60:    sell_time_histogram[5] += 1
                elif diff_days <= 90:    sell_time_histogram[6] += 1
                else:                    sell_time_histogram[7] += 1
                nazwa = row['nazwa'] or 'Produkt'
                time_str = f"{diff_hours:.0f}h" if diff_hours < 24 else f"{diff_days:.1f} dni"
                najszybciej_sprzedane.append({'czas': time_str, 'czas_h': diff_hours,
                    'nazwa': nazwa[:55], 'cena': float(row['cena'] or 0)})
            except:
                continue
        najszybciej_sprzedane.sort(key=lambda x: x['czas_h'])
        # Deduplikacja — każdy produkt tylko raz (najszybszy czas)
        _seen_ns = set()
        _unique_ns = []
        for item in najszybciej_sprzedane:
            if item['nazwa'] not in _seen_ns:
                _seen_ns.add(item['nazwa'])
                _unique_ns.append(item)
                if len(_unique_ns) >= 10:
                    break
        najszybciej_sprzedane = _unique_ns
    except:
        pass

    # Prebuduj HTML histogramu (unika zagnieżdżonych f-stringów)
    if sum(sell_time_histogram) == 0:
        histogram_html = '<div style="color:#64748b;font-size:0.85rem;text-align:center;padding:20px">Brak danych — synchronizuj sprzedaż z Allegro</div>'
    else:
        najszybciej_rows = ''.join([
            f'<div style="display:flex;align-items:center;gap:10px;padding:6px 0;border-bottom:1px solid #1e1e2e">'
            f'<div style="font-size:0.85rem;font-weight:700;color:#beee00;min-width:60px">{item["czas"]}</div>'
            f'<div style="flex:1;font-size:0.8rem;color:#e2e8f0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{item["nazwa"]}</div>'
            f'<div style="font-size:0.8rem;color:#eab308;min-width:55px;text-align:right">{item["cena"]:.0f} zł</div>'
            f'</div>' for item in najszybciej_sprzedane
        ])
        histogram_html = (
            f'<canvas id="chartCzasSprzedazy" height="150" style="margin-bottom:15px"></canvas>'
            f'<div style="font-size:0.75rem;color:#beee00;font-weight:600;margin-bottom:8px"><span class=material-symbols-outlined>bolt</span> Najszybciej sprzedane (od wystawienia)</div>'
            f'{najszybciej_rows}'
        )

    
    # === HOURLY THROUGHPUT (sprzedaż po godzinie dnia) ===
    godzinowe = [0.0] * 24
    try:
        hour_rows = conn.execute('''
            SELECT CAST(strftime('%H', data_sprzedazy) AS INTEGER) as h,
                   SUM(cena * ilosc) as suma
            FROM sprzedaze
            WHERE status NOT IN ('zwrot','anulowane','anulowana')
              AND data_sprzedazy IS NOT NULL AND data_sprzedazy != ''
            GROUP BY h
        ''').fetchall()
        for row in hour_rows:
            if row['h'] is not None and 0 <= row['h'] < 24:
                godzinowe[row['h']] = round(float(row['suma'] or 0), 2)
    except:
        pass

    # === OSTATNIE TRANSAKCJE (transaction matrix) ===
    ostatnie_transakcje = []
    try:
        ostatnie_rows = conn.execute('''
            SELECT s.id, s.allegro_order_id,
                   COALESCE(NULLIF(p.nazwa,''), NULLIF(s.nazwa,''), 'Zamówienie') as nazwa,
                   s.cena, s.ilosc, (s.cena * s.ilosc) as wartosc,
                   s.status, s.data_sprzedazy, s.kupujacy,
                   COALESCE(p.kategoria, '') as kategoria
            FROM sprzedaze s
            LEFT JOIN produkty p ON s.produkt_id = p.id
            WHERE s.status NOT IN ('anulowane','anulowana')
              AND s.data_sprzedazy IS NOT NULL AND s.data_sprzedazy != ''
            ORDER BY s.data_sprzedazy DESC LIMIT 15
        ''').fetchall()
        for row in ostatnie_rows:
            ds = str(row['data_sprzedazy'] or '')[:16].replace('T',' ')
            ostatnie_transakcje.append({
                'id': row['id'],
                'nazwa': (row['nazwa'] or 'Produkt')[:50],
                'wartosc': float(row['wartosc'] or 0),
                'ilosc': int(row['ilosc'] or 1),
                'status': row['status'] or 'sprzedano',
                'data': ds,
                'kupujacy': (row['kupujacy'] or '—')[:20],
            })
    except:
        pass

    # Przygotuj dane do wykresów
    nazwy_miesiecy = ['Sty', 'Lut', 'Mar', 'Kwi', 'Maj', 'Cze', 'Lip', 'Sie', 'Wrz', 'Paź', 'Lis', 'Gru']
    # Przychód roku (filtr roku + offline) — do kalkulacji podatkowej i wykresu
    _przychod_rok_row = conn.execute('''
        SELECT COALESCE(SUM(cena * ilosc), 0) as suma
        FROM sprzedaze
        WHERE strftime('%Y', REPLACE(SUBSTR(data_sprzedazy,1,19),'T',' ')) = ?
          AND status NOT IN ('zwrot', 'anulowane', 'anulowana')
          AND (kupujacy IS NULL OR kupujacy != 'offline')
    ''', (str(current_year),)).fetchone()
    przychod_total = float(_przychod_rok_row['suma'] or 0) + pryw_total_rok
    koszty_total_lacznie = koszty_total_rok + palety_zakup_total_rok  # koszty operacyjne + REALNE zakupy palet (cashflow)
    zysk_rok = przychod_total - koszty_total_lacznie
    zysk_kolor = '#beee00' if zysk_rok >= 0 else '#ef4444'
    
    # === KALKULACJA PODATKOWA ===
    # VAT 23%, podatek liniowy 19%
    # Przychód = brutto (z VAT 23%) — zarówno Allegro jak i prywatne (faktury)
    # Koszty = brutto (z VAT 23%)
    # VAT ze sprzedaży
    vat_sprzedaz = przychod_total - (przychod_total / 1.23)
    # VAT z kosztów (odliczany)
    vat_koszty = koszty_total_lacznie - (koszty_total_lacznie / 1.23)
    # VAT do zapłaty = VAT ze sprzedaży - VAT z kosztów
    vat_do_zaplaty = vat_sprzedaz - vat_koszty
    # Przychód netto (bez VAT)
    przychod_netto = przychod_total / 1.23
    # Koszty netto (bez VAT) - odliczasz VAT z kosztów
    koszty_netto = koszty_total_lacznie / 1.23
    # Dochód do opodatkowania
    dochod = przychod_netto - koszty_netto
    # Podatek dochodowy 19%
    podatek = max(0, dochod * 0.19)
    # Prawdziwy zysk na rękę
    zysk_na_reke = dochod - podatek
    zysk_na_reke_kolor = '#beee00' if zysk_na_reke >= 0 else '#ef4444'
    dane_miesieczne = [0] * 12
    dane_miesieczne_cnt = [0] * 12  # ilość zamówień per miesiąc
    dane_koszty = [0] * 12          # koszty per miesiąc
    dane_prywatne = [0] * 12
    dane_palety = [0] * 12
    dane_palety_cnt = [0] * 12  # ilość palet kupionych per miesiąc
    for m in miesieczne:
        idx = int(m['miesiac']) - 1
        dane_miesieczne[idx] = float(m['suma'])
        dane_miesieczne_cnt[idx] = int(m['ilosc'])
    for m_int, suma in koszty_per_msc.items():
        dane_koszty[m_int - 1] = suma
    for m_int, suma in pryw_per_msc.items():
        dane_prywatne[m_int - 1] = suma
    for m_int, suma in palety_per_msc.items():
        dane_palety[m_int - 1] = suma
    for m_int, cnt in palety_cnt_per_msc.items():
        dane_palety_cnt[m_int - 1] = cnt
    dane_palety_zakup = [0] * 12
    for m_int, suma in palety_zakup_per_msc.items():
        dane_palety_zakup[m_int - 1] = suma
    
    dane_roczne_labels = [r['rok'] for r in roczne] if roczne else [str(current_year)]
    dane_roczne_values = [float(r['suma']) for r in roczne] if roczne else [0]
    
    # Przygotuj dane dzienne jako JSON (dla drill-down) - suma i ilość
    dzienne_json = {}
    dzienne_cnt_json = {}
    for m in range(1, 13):
        dni_w_miesiacu = 31 if m in [1,3,5,7,8,10,12] else (30 if m in [4,6,9,11] else 29)
        dzienne_json[m] = [dzienne_per_miesiac.get(m, {}).get(d, 0) for d in range(1, dni_w_miesiacu+1)]
        dzienne_cnt_json[m] = [dzienne_cnt_per_miesiac.get(m, {}).get(d, 0) for d in range(1, dni_w_miesiacu+1)]
    
    # Status SYPIE - DZISIAJ (nie miesiąc!)
    current_month = datetime.now().month
    today_day = datetime.now().day
    today_sales = dzienne_per_miesiac.get(current_month, {}).get(today_day, 0)
    today_cnt = dzienne_cnt_per_miesiac.get(current_month, {}).get(today_day, 0)
    
    # 5 POZIOMÓW SYPANIA (bazowane na kwocie dziennej)
    if today_sales >= 5000:
        status_text = "MEGA SYPIE!"
        status_icon = "local_fire_department"
        status_color = "#beee00"
    elif today_sales >= 3000:
        status_text = "SYPIE!"
        status_icon = "trending_up"
        status_color = "#beee00"
    elif today_sales >= 1500:
        status_text = "Calkiem niezle"
        status_icon = "trending_up"
        status_color = "#eab308"
    elif today_sales >= 500:
        status_text = "Sypie troche"
        status_icon = "speed"
        status_color = "#f97316"
    else:
        status_text = "NIE SYPIE"
        status_icon = "bedtime"
        status_color = "#ef4444"

    # === ROI calculations for template ===
    roi_total = roi_sredni = 0
    top3 = []
    worst3 = []
    if palety_roi:
        roi_total_koszt = sum(p['koszt'] for p in palety_roi)
        roi_total_przychod = sum(p['przychod'] for p in palety_roi)
        roi_total_zysk = roi_total_przychod - roi_total_koszt
        roi_total = (roi_total_zysk / roi_total_koszt * 100) if roi_total_koszt > 0 else 0
        roi_sredni = sum(p['roi'] for p in palety_roi) / len(palety_roi)
        sorted_desc = sorted(palety_roi, key=lambda x: x['roi'], reverse=True)
        sorted_asc = sorted(palety_roi, key=lambda x: x['roi'])
        top_n = min(5, max(1, len(palety_roi) // 2))
        top3 = sorted_desc[:top_n]
        top_names = {p['nazwa'] for p in top3}
        worst3 = [p for p in sorted_asc if p['nazwa'] not in top_names][:top_n]

    from flask import render_template
    return render_template('statystyki.html',
        current_year=current_year,
        today_sales=today_sales,
        today_cnt=today_cnt,
        status_text=status_text,
        status_icon=status_icon,
        status_color=status_color,
        podsumowanie=podsumowanie,
        przychod_total=przychod_total,
        przychod_netto=przychod_netto,
        koszty_total_lacznie=koszty_total_lacznie,
        koszty_netto=koszty_netto,
        palety_total_rok=palety_total_rok,
        palety_zakup_total_rok=palety_zakup_total_rok,
        pryw_total_rok=pryw_total_rok,
        zysk_rok=zysk_rok,
        zysk_kolor=zysk_kolor,
        dochod=dochod,
        vat_do_zaplaty=vat_do_zaplaty,
        vat_sprzedaz=vat_sprzedaz,
        vat_koszty=vat_koszty,
        podatek=podatek,
        zysk_na_reke=zysk_na_reke,
        zysk_na_reke_kolor=zysk_na_reke_kolor,
        histogram_html=histogram_html,
        palety_roi=palety_roi,
        roi_total=roi_total,
        roi_sredni=roi_sredni,
        top3=top3,
        worst3=worst3,
        top_produkty=top_produkty,
        top_dostawcy=top_dostawcy,
        roczne=roczne,
        dane_palety_cnt=[dane_palety_cnt[i] for i in range(12)],
        palety_total_cnt_rok=palety_total_cnt_rok,
        nazwy_miesiecy_json=json.dumps(nazwy_miesiecy),
        dane_miesieczne_json=json.dumps(dane_miesieczne),
        dane_miesieczne_cnt_json=json.dumps(dane_miesieczne_cnt),
        dane_koszty_json=json.dumps(dane_koszty),
        dane_prywatne_json=json.dumps(dane_prywatne),
        dane_palety_json=json.dumps(dane_palety),
        dane_palety_zakup_json=json.dumps(dane_palety_zakup),
        dzienne_json=json.dumps(dzienne_json),
        dzienne_cnt_json=json.dumps(dzienne_cnt_json),
        sell_time_labels_json=json.dumps(sell_time_labels),
        sell_time_histogram_json=json.dumps(sell_time_histogram),
        dane_roczne_labels_json=json.dumps(dane_roczne_labels),
        dane_roczne_values_json=json.dumps(dane_roczne_values),
        godzinowe_json=json.dumps(godzinowe),
        ostatnie_transakcje=ostatnie_transakcje,
        now=datetime.now(),
    )

    # === OLD INLINE HTML REMOVED - now using templates/statystyki.html ===
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>bar_chart</span> STATYSTYKI</h1><small>Sprzedaż i przychody (tylko opłacone)</small></div>
    <div style="text-align:right;margin-bottom:10px">
        <a href="/sync-historyczny" style="font-size:0.75rem;color:#64748b;text-decoration:none;background:#1e1e2e;padding:5px 10px;border-radius:6px"><span class=material-symbols-outlined>sync</span> Sync historyczny (poprzednie miesiące)</a>
    </div>
    
    <!-- STATUS SYPIE - DZISIAJ -->
    <div style="background:linear-gradient(135deg, {status_color}22, {status_color}11);border:2px solid {status_color};border-radius:16px;padding:20px;margin-bottom:20px;text-align:center">
        <div style="font-size:2.5rem;font-weight:700;color:{status_color}">{status_text}</div>
        <div style="font-size:0.9rem;color:#94a3b8;margin-top:5px">
            Dzisiaj ({datetime.now().strftime('%d.%m')}): <strong>{today_sales:.0f} zł</strong> | <strong>{today_cnt}</strong> zamówień
        </div>
    </div>
    
    <!-- Podsumowanie -->
    <div class="stats" style="margin-bottom:20px">
        <div class="stat">
            <div class="stat-v">{podsumowanie['ilosc_transakcji']}</div>
            <div class="stat-l">Transakcji</div>
        </div>
        <div class="stat">
            <div class="stat-v green">{przychod_total:.0f} zł</div>
            <div class="stat-l">Przychód ({current_year}){f' (w tym [HANDSHAKE] {pryw_total_rok:.0f} zł prywatne)' if pryw_total_rok > 0 else ''}</div>
        </div>
        <div class="stat">
            <div class="stat-v" style="color:#f43f5e">-{koszty_total_lacznie:.0f} zł</div>
            <div class="stat-l">Koszty ({current_year}) <span style="font-size:0.65rem;color:#64748b">(w tym <span class=material-symbols-outlined>inventory_2</span> {palety_total_rok:.0f} zł palety)</span> <a href="/magazyn/koszty" style="color:#64748b;font-size:0.7rem;margin-left:4px">+dodaj</a></div>
        </div>
        <div class="stat">
            <div class="stat-v" style="color:{zysk_kolor}">{zysk_rok:.0f} zł</div>
            <div class="stat-l">Zysk brutto (przed podatkiem)</div>
        </div>
    </div>
    
    <!-- HISTOGRAM CZASU SPRZEDAŻY -->
    <div class="card" style="padding:15px;margin-bottom:15px;border:1px solid #beee0044">
        <div style="font-weight:700;margin-bottom:12px;color:#beee00"><span class=material-symbols-outlined>timer</span> Czas sprzedaży (od wystawienia)</div>
        {histogram_html}
    </div>
    
    <!-- Kalkulacja podatkowa -->
    <div class="card" style="padding:15px;margin-bottom:15px;border:1px solid #ff6b9b44">
        <div style="font-weight:700;margin-bottom:12px;color:#a78bfa"><span class=material-symbols-outlined>receipt_long</span> Rozliczenie podatkowe ({current_year})</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:0.85rem">
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">Przychód brutto</div>
                <div style="font-weight:700;color:#8ff5ff">{przychod_total:.0f} zł</div>
            </div>
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">Przychód netto (bez VAT)</div>
                <div style="font-weight:700">{przychod_netto:.0f} zł</div>
            </div>
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">Koszty netto (bez VAT)</div>
                <div style="font-weight:700;color:#f43f5e">-{koszty_netto:.0f} zł</div>
            </div>
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">Dochód do opodatkowania</div>
                <div style="font-weight:700">{dochod:.0f} zł</div>
            </div>
            <div style="background:#1a1025;border:1px solid #ef444433;border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">VAT do zapłaty (23%)</div>
                <div style="font-weight:700;color:#ef4444">-{vat_do_zaplaty:.0f} zł</div>
                <div style="font-size:0.7rem;color:#475569;margin-top:2px">VAT sprzedaż {vat_sprzedaz:.0f} − VAT koszty {vat_koszty:.0f}</div>
            </div>
            <div style="background:#1a1025;border:1px solid #ef444433;border-radius:8px;padding:10px">
                <div style="color:#64748b;margin-bottom:4px">Podatek dochodowy (19%)</div>
                <div style="font-weight:700;color:#ef4444">-{podatek:.0f} zł</div>
                <div style="font-size:0.7rem;color:#475569;margin-top:2px">{dochod:.0f} × 19%</div>
            </div>
        </div>
        <div style="margin-top:12px;padding:12px;background:#0a1f12;border:2px solid {zysk_na_reke_kolor}55;border-radius:10px;display:flex;justify-content:space-between;align-items:center">
            <div style="color:#94a3b8;font-size:0.9rem"><span class=material-symbols-outlined>paid</span> Zysk na rękę (po VAT i podatku)</div>
            <div style="font-size:1.4rem;font-weight:700;color:{zysk_na_reke_kolor}">{zysk_na_reke:.0f} zł</div>
        </div>
        <div style="font-size:0.7rem;color:#475569;margin-top:8px;text-align:center"><span class=material-symbols-outlined>warning</span> Szacunkowe — skonsultuj z księgową. Nie uwzględnia ZUS, ulg i odpisów.</div>
    </div>
    '''

    # === RENTOWNOŚĆ PALET ROI ===
    html_roi = ''
    if palety_roi:
        roi_total_koszt = sum(p['koszt'] for p in palety_roi)
        roi_total_przychod = sum(p['przychod'] for p in palety_roi)
        roi_total_zysk = roi_total_przychod - roi_total_koszt
        roi_total = (roi_total_zysk / roi_total_koszt * 100) if roi_total_koszt > 0 else 0
        roi_sredni = sum(p['roi'] for p in palety_roi) / len(palety_roi)
        roi_total_kolor = '#beee00' if roi_total >= 0 else '#ef4444'
        roi_sr_kolor = '#beee00' if roi_sredni >= 0 else '#ef4444'
        sorted_desc = sorted(palety_roi, key=lambda x: x['roi'], reverse=True)
        sorted_asc = sorted(palety_roi, key=lambda x: x['roi'])
        # Pokaż max 5 najlepszych, ale nie więcej niż połowa
        top_n = min(5, max(1, len(palety_roi) // 2))
        top3 = sorted_desc[:top_n]
        top_names = {p['nazwa'] for p in top3}
        # Najgorsze - wyklucz te które już są w najlepszych
        worst3 = [p for p in sorted_asc if p['nazwa'] not in top_names][:top_n]
        worst_label = '<span class=material-symbols-outlined>trending_down</span> Najgorsze' if worst3 and worst3[0]['roi'] < 0 else '<span class=material-symbols-outlined>bar_chart</span> Najmniej rentowne'

        def _roi_row(p):
            kol = '#beee00' if p['roi'] >= 0 else '#ef4444'
            pct = min(100, max(0, abs(p['roi'])))
            sign = '+' if p['roi'] >= 0 else ''
            bar = f'<div style="height:5px;background:#1e1e2e;border-radius:3px;margin-top:4px"><div style="height:5px;width:{pct:.0f}%;background:{kol};border-radius:3px"></div></div>'
            return (f'<div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:10px;margin-bottom:6px">'
                    f'<div style="display:flex;justify-content:space-between;align-items:center">'
                    f'<div style="font-size:0.8rem;color:#e2e8f0;flex:1;margin-right:8px;overflow:hidden;white-space:nowrap;text-overflow:ellipsis">{p["nazwa"][:35]}</div>'
                    f'<div style="font-weight:700;color:{kol};white-space:nowrap">{sign}{p["roi"]:.0f}%</div>'
                    f'</div>'
                    f'<div style="font-size:0.7rem;color:#64748b;margin-top:2px">{p.get("sprzedane","?")}/{p.get("total","?")} szt. · koszt prop. {p["koszt"]:.0f} zł / {p.get("koszt_palety", p["koszt"]):.0f} zł · zysk {"+" if p["zysk"]>=0 else ""}{p["zysk"]:.0f} zł</div>'
                    f'{bar}'
                    f'</div>')

        top_html = ''.join(_roi_row(p) for p in top3)
        worst_html = ''.join(_roi_row(p) for p in worst3)

        html_roi = f'''<div class="card" style="padding:15px;margin-bottom:15px;border:1px solid #beee0033">
        <div style="font-weight:700;margin-bottom:12px;color:#beee00"><span class=material-symbols-outlined>inventory_2</span> Rentowność palet ({current_year})</div>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:14px">
            <div style="background:#0a1f12;border:1px solid {roi_total_kolor}44;border-radius:10px;padding:12px;text-align:center">
                <div style="font-size:1.6rem;font-weight:700;color:{roi_total_kolor}">{roi_total:.0f}%</div>
                <div style="font-size:0.7rem;color:#64748b;margin-top:2px">ROI całkowity</div>
            </div>
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:10px;padding:12px;text-align:center">
                <div style="font-size:1.6rem;font-weight:700;color:#a78bfa">{len(palety_roi)}</div>
                <div style="font-size:0.7rem;color:#64748b;margin-top:2px">Palety z danymi</div>
            </div>
            <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:10px;padding:12px;text-align:center">
                <div style="font-size:1.6rem;font-weight:700;color:{roi_sr_kolor}">{roi_sredni:.0f}%</div>
                <div style="font-size:0.7rem;color:#64748b;margin-top:2px">Średni ROI</div>
            </div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
            <div>
                <div style="font-size:0.75rem;color:#beee00;font-weight:600;margin-bottom:6px"><span class=material-symbols-outlined>emoji_events</span> Najlepsze</div>
                {top_html}
            </div>
            <div>
                <div style="font-size:0.75rem;color:#eab308;font-weight:600;margin-bottom:6px">{worst_label}</div>
                {worst_html}
            </div>
        </div>
        <div style="font-size:0.7rem;color:#475569;margin-top:10px;text-align:center">ROI = (Przychód − Koszt palety) ÷ Koszt × 100%  |  <a href="/analityka" style="color:#64748b"><span class=material-symbols-outlined>bar_chart</span> Szczegółowa analityka →</a></div>
    </div>'''

    html += html_roi

    # Buduj HTML palet per miesiąc (osobno, bo zagnieżdżone f-stringi w f-strings nie działają)
    nazwy_msc = ["Sty","Lut","Mar","Kwi","Maj","Cze","Lip","Sie","Wrz","Paź","Lis","Gru"]
    palety_cells = ''
    for i in range(12):
        cnt = dane_palety_cnt[i]
        kolor = '#8ff5ff' if cnt > 0 else '#2d2d48'
        palety_cells += f'<div style="background:#1e1e2e;border-radius:8px;padding:8px;text-align:center"><div style="font-size:0.65rem;color:#64748b">{nazwy_msc[i]}</div><div style="font-size:1.1rem;font-weight:700;color:{kolor}">{cnt}</div></div>'

    html += f'''
    <!-- Wykres miesięczny z drill-down -->
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
            <div id="chartTitle" style="font-weight:600"><span class=material-symbols-outlined>calendar_month</span> Sprzedaż miesięcznie ({current_year})</div>
            <button id="btnBack" onclick="showMonthlyView()" style="display:none;padding:5px 10px;background:#8ff5ff;border:none;border-radius:5px;color:#0e0e10;font-weight:700;cursor:pointer">← Miesiące</button>
        </div>
        <div style="font-size:0.75rem;color:#64748b;margin-bottom:10px"><span class=material-symbols-outlined>lightbulb</span> Kliknij na słupek miesiąca aby zobaczyć rozkład dzienny</div>
        <canvas id="chartMiesiace" height="200"></canvas>
        <div id="monthSummary" style="display:none;margin-top:12px"></div>
    </div>
    
    <!-- Wykres roczny -->
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="font-weight:600;margin-bottom:10px"><span class=material-symbols-outlined>trending_up</span> Sprzedaż rocznie</div>
        <canvas id="chartLata" height="150"></canvas>
    </div>

    <!-- Palety kupione per miesiąc -->
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="font-weight:600;margin-bottom:10px"><span class=material-symbols-outlined>inventory_2</span> Palety kupione ({current_year}) — łącznie {palety_total_cnt_rok} szt.</div>
        <div style="display:grid;grid-template-columns:repeat(6, 1fr);gap:6px">
            {palety_cells}
        </div>
    </div>

    <!-- Top produkty -->
    <div class="section"><span class=material-symbols-outlined>emoji_events</span> TOP 5 PRODUKTÓW</div>
    '''
    
    if top_produkty:
        for i, p in enumerate(top_produkty):
            html += f'''<div class="item">
                <div style="font-size:1.2rem;margin-right:10px;width:25px;text-align:center">{i+1}</div>
                <div class="item-info">
                    <div class="item-name">{(p['produkt_nazwa'] or 'Nieznany')[:35]}</div>
                    <div class="item-meta">Sprzedano: {p['sprzedane']}x</div>
                </div>
                <div class="item-right">
                    <div class="item-qty" style="color:#beee00">{p['przychod'] or 0:.0f} zł</div>
                </div>
            </div>'''
    else:
        html += '<div class="alert alert-warn">Brak danych o sprzedaży</div>'
    
    # Top dostawcy
    html += '<div class="section"><span class=material-symbols-outlined>local_shipping</span> TOP DOSTAWCY</div>'
    
    if top_dostawcy:
        for i, d in enumerate(top_dostawcy):
            html += f'''<div class="item">
                <div style="font-size:1.2rem;margin-right:10px;width:25px;text-align:center">{i+1}</div>
                <div class="item-info">
                    <div class="item-name dostawca-name">{d['dostawca_nazwa']}</div>
                    <div class="item-meta">Sprzedano: {d['sprzedane']}x</div>
                </div>
                <div class="item-right">
                    <div class="item-qty" style="color:#beee00">{d['przychod'] or 0:.0f} zł</div>
                </div>
            </div>'''
    else:
        html += '<div class="alert alert-warn">Brak danych o dostawcach</div>'
    
    html += f'''
    <a href="/magazyn" class="back">← Powrót</a>
    
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js" integrity="sha384-jb8JQMbMoBUzgWatfe6COACi2ljcDdZQ2OxczGA3bGNeWe+6DChMTBJemed7ZnvJ" crossorigin="anonymous"></script>
    <script>
    // Dane
    const nazwyMiesiecy = {json.dumps(nazwy_miesiecy)};
    const daneMiesieczne = {json.dumps(dane_miesieczne)};
    const daneMiesieczneCnt = {json.dumps(dane_miesieczne_cnt)};
    const daneKoszty = {json.dumps(dane_koszty)};
    const danePrywatne = {json.dumps(dane_prywatne)};
    const danePalety = {json.dumps(dane_palety)};
    const danePaletyZakup = {json.dumps(dane_palety_zakup)};
    const danePaletyCnt = {json.dumps([dane_palety_cnt[i] for i in range(12)])};
    // Łączne koszty = operacyjne + COGS (koszt sprzedanych produktów)
    const daneKosztyLacznie = daneKoszty.map((k, i) => k + danePalety[i]);
    // null dla miesięcy bez aktywności - linia nie spada do zera
    const daneZysk = daneMiesieczne.map((p, i) => {{
        const total = p + danePrywatne[i];
        if (total === 0 && daneKosztyLacznie[i] === 0) return null;
        return total - daneKosztyLacznie[i];
    }});

    // Histogram czasu sprzedaży
    const histLabels = {json.dumps(sell_time_labels)};
    const histData = {json.dumps(sell_time_histogram)};
    if (document.getElementById('chartCzasSprzedazy') && histData.some(v => v > 0)) {{
        new Chart(document.getElementById('chartCzasSprzedazy'), {{
            type: 'bar',
            data: {{
                labels: histLabels,
                datasets: [{{
                    label: 'Sprzedaży',
                    data: histData,
                    backgroundColor: histData.map((v, i) => i === 0 ? 'rgba(34,197,94,0.9)' : i <= 1 ? 'rgba(34,197,94,0.7)' : i <= 2 ? 'rgba(234,179,8,0.7)' : 'rgba(100,116,139,0.6)'),
                    borderRadius: 6
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{ legend: {{ display: false }} }},
                scales: {{
                    y: {{ beginAtZero: true, grid: {{ color: 'rgba(255,255,255,0.07)' }}, ticks: {{ color: '#64748b', stepSize: 1 }} }},
                    x: {{ grid: {{ display: false }}, ticks: {{ color: '#94a3b8', font: {{ size: 11 }} }} }}
                }}
            }}
        }});
    }}
    const daneDzienne = {json.dumps(dzienne_json)};
    const daneDzienneCnt = {json.dumps(dzienne_cnt_json)};
    
    let chartMiesiace = null;
    let currentView = 'monthly'; // monthly lub daily
    let currentMonth = 0;
    
    // Funkcja sprawdzająca czy dzień "sypał" (>= 5 zamówień i >= 300 zł)
    // Funkcja zwracająca kolor słupka według 5 progów (bazowane tylko na kwocie)
    function getBarColor(kwota) {{
        if (kwota >= 5000) return 'rgba(22, 163, 74, 0.9)';   // ciemny zielony - MEGA SYPIE
        if (kwota >= 3000) return 'rgba(34, 197, 94, 0.9)';   // zielony - SYPIE
        if (kwota >= 1500) return 'rgba(234, 179, 8, 0.9)';   // żółty - całkiem nieźle
        if (kwota >= 500) return 'rgba(249, 115, 22, 0.9)';   // pomarańczowy - sypie trochę
        return 'rgba(59, 130, 246, 0.6)';                      // niebieski - nie sypie
    }}
    
    // Funkcja sprawdzająca czy dzień "sypał" (>= 1500 zł)
    function czySypalo(kwota, cnt) {{
        return kwota >= 1500;
    }}
    
    // Inicjalizacja wykresu miesięcznego
    function initMonthlyChart() {{
        const ctx = document.getElementById('chartMiesiace');
        chartMiesiace = new Chart(ctx, {{
            type: 'bar',
            data: {{
                labels: nazwyMiesiecy,
                datasets: [{{
                    label: 'Przychód (zł)',
                    data: daneMiesieczne.map((v, i) => v + danePrywatne[i]),
                    backgroundColor: 'rgba(59, 130, 246, 0.6)',
                    borderColor: 'rgba(59, 130, 246, 1)',
                    borderWidth: 1,
                    borderRadius: 5,
                    order: 2
                }}, {{
                    label: 'Koszty (zł)',
                    data: daneKosztyLacznie,
                    backgroundColor: 'rgba(244, 63, 94, 0.5)',
                    borderColor: 'rgba(244, 63, 94, 1)',
                    borderWidth: 1,
                    borderRadius: 5,
                    order: 1
                }}, {{
                    label: 'Zakup palet (zł)',
                    data: danePaletyZakup,
                    backgroundColor: 'rgba(251, 191, 36, 0.5)',
                    borderColor: 'rgba(251, 191, 36, 1)',
                    borderWidth: 1,
                    borderRadius: 5,
                    order: 1
                }}, {{
                    label: 'Zysk netto',
                    data: daneZysk,
                    type: 'line',
                    borderColor: '#beee00',
                    backgroundColor: 'rgba(190,238,0,0.1)',
                    borderWidth: 2,
                    pointRadius: 4,
                    pointBackgroundColor: '#beee00',
                    tension: 0.3,
                    fill: false,
                    order: 0
                }}]
            }},
            options: {{
                responsive: true,
                onClick: function(event, elements, chart) {{
                    if (currentView === 'monthly' && elements && elements.length > 0) {{
                        const index = elements[0].index;
                        showDailyView(index + 1);
                    }}
                }},
                plugins: {{
                    legend: {{ display: true, labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }},
                    tooltip: {{
                        callbacks: {{
                            afterBody: function(items) {{
                                if (currentView !== 'monthly') return [];
                                const i = items[0].dataIndex;
                                const cnt = daneMiesieczneCnt[i];
                                const koszty = daneKosztyLacznie[i];
                                const pryw = danePrywatne[i];
                                const palety = danePalety[i];
                                const zysk = daneMiesieczne[i] + pryw - koszty;
                                const zakupPalet = danePaletyZakup[i];
                                const cntPalet = danePaletyCnt[i];
                                const lines = [cnt + ' zamowien'];
                                if (pryw > 0) lines.push('+ ' + pryw.toFixed(0) + ' zl prywatna');
                                if (zakupPalet > 0) lines.push('<span class=material-symbols-outlined>shopping_cart</span> ' + cntPalet + ' palet kupiono za ' + zakupPalet.toFixed(0) + ' zl');
                                if (palety > 0) lines.push('- ' + palety.toFixed(0) + ' zl COGS (sprzedanych)');
                                if (koszty > palety) lines.push('- ' + (koszty-palety).toFixed(0) + ' zl inne koszty');
                                lines.push('Zysk: ' + zysk.toFixed(0) + ' zl');
                                return lines;
                            }},
                            afterLabel: function(context) {{
                                if (currentView !== 'monthly') {{
                                    const cnt = daneDzienneCnt[currentMonth] ? daneDzienneCnt[currentMonth][context.dataIndex] : 0;
                                    const sypalo = czySypalo(context.raw, cnt);
                                    return cnt + ' zamowien ' + (sypalo ? 'SYPALO!' : '');
                                }}
                                return '';
                            }}
                        }}
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        grid: {{ color: 'rgba(255,255,255,0.1)' }},
                        ticks: {{ color: '#64748b' }}
                    }},
                    x: {{
                        grid: {{ display: false }},
                        ticks: {{ color: '#64748b' }}
                    }}
                }}
            }}
        }});
    }}
    
    // Pokaż widok dzienny
    function showDailyView(month) {{
        currentView = 'daily';
        currentMonth = month;
        
        const daysInMonth = daneDzienne[month] ? daneDzienne[month].length : 31;
        const dayLabels = Array.from({{length: daysInMonth}}, (_, i) => (i + 1).toString());
        const dailyData = daneDzienne[month] || [];
        
        const colors = dailyData.map(kwota => getBarColor(kwota));
        
        // Koszty miesięczne rozłożone równo na każdy dzień
        const kosztMiesieczny = daneKoszty[month - 1] || 0;
        const kosztDzienny = kosztMiesieczny / daysInMonth;
        // null dla dni bez sprzedaży żeby linia nie wisiała w powietrzu
        const dailyZysk = dailyData.map(p => p === 0 ? null : parseFloat((p - kosztDzienny).toFixed(2)));
        
        chartMiesiace.data.labels = dayLabels;
        chartMiesiace.data.datasets[0].data = dailyData;
        chartMiesiace.data.datasets[0].backgroundColor = colors;
        chartMiesiace.data.datasets[0].borderColor = colors.map(c => c.replace('0.9', '1').replace('0.6', '0.8'));
        // Koszty dzienne jako słupki
        chartMiesiace.data.datasets[1].data = new Array(daysInMonth).fill(parseFloat(kosztDzienny.toFixed(2)));
        chartMiesiace.data.datasets[1].hidden = kosztMiesieczny === 0;
        // Ukryj zakup palet w widoku dziennym
        chartMiesiace.data.datasets[2].data = new Array(daysInMonth).fill(0);
        chartMiesiace.data.datasets[2].hidden = true;
        // Zysk netto dzienny jako linia
        chartMiesiace.data.datasets[3].data = dailyZysk;
        chartMiesiace.data.datasets[3].hidden = false;
        chartMiesiace.update();
        
        document.getElementById('chartTitle').textContent = ' ' + nazwyMiesiecy[month-1] + ' {current_year} - rozkład dzienny';
        document.getElementById('btnBack').style.display = 'inline-block';
        
        // Podsumowanie miesiąca
        const przychod = (daneMiesieczne[month-1] || 0) + (danePrywatne[month-1] || 0);
        const koszty = daneKosztyLacznie[month-1] || 0;
        const kosztPalety = danePalety[month-1] || 0;
        const zysk = przychod - koszty;
        const zyskKolor = zysk >= 0 ? '#beee00' : '#ef4444';
        const cnt = daneMiesieczneCnt[month-1] || 0;
        document.getElementById('monthSummary').innerHTML = `
            <div style="display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap">
                <div style="flex:1;min-width:100px;background:#1e1e2e;border-radius:10px;padding:10px;text-align:center">
                    <div style="font-size:1.1rem;font-weight:700;color:#8ff5ff">${{przychod.toFixed(0)}} zł</div>
                    <div style="font-size:0.7rem;color:#64748b;margin-top:2px">Przychód</div>
                </div>
                <div style="flex:1;min-width:100px;background:#1e1e2e;border-radius:10px;padding:10px;text-align:center">
                    <div style="font-size:1.1rem;font-weight:700;color:#f43f5e">-${{koszty.toFixed(0)}} zł</div>
                    <div style="font-size:0.7rem;color:#64748b;margin-top:2px">Koszty (COGS${{kosztPalety > 0 ? ' ' + kosztPalety.toFixed(0) + ' zł' : ''}})</div>
                </div>
                <div style="flex:1;min-width:100px;background:#1e1e2e;border-radius:10px;padding:10px;text-align:center">
                    <div style="font-size:1.1rem;font-weight:700">${{cnt}}</div>
                    <div style="font-size:0.7rem;color:#64748b;margin-top:2px">Zamówień</div>
                </div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-top:8px">
                ${{(function() {{
                    const vatSprzedaz = przychod - (przychod / 1.23);
                    const vatKoszty = koszty - (koszty / 1.23);
                    const vatDoZaplaty = vatSprzedaz - vatKoszty;
                    const przychodNetto = przychod / 1.23;
                    const kosztaNetto = koszty / 1.23;
                    const dochod = przychodNetto - kosztaNetto;
                    const podatek = Math.max(0, dochod * 0.19);
                    const naReke = dochod - podatek;
                    const kolor = naReke >= 0 ? '#beee00' : '#ef4444';
                    return `
                    <div style="background:#1a1025;border:1px solid #ef444433;border-radius:8px;padding:8px;text-align:center">
                        <div style="font-size:0.95rem;font-weight:700;color:#ef4444">-${{vatDoZaplaty.toFixed(0)}} zł</div>
                        <div style="font-size:0.65rem;color:#64748b;margin-top:2px">VAT do zapłaty</div>
                    </div>
                    <div style="background:#1a1025;border:1px solid #ef444433;border-radius:8px;padding:8px;text-align:center">
                        <div style="font-size:0.95rem;font-weight:700;color:#ef4444">-${{podatek.toFixed(0)}} zł</div>
                        <div style="font-size:0.65rem;color:#64748b;margin-top:2px">Podatek 19%</div>
                    </div>
                    <div style="background:#0a1f12;border:2px solid ${{kolor}}55;border-radius:8px;padding:8px;text-align:center">
                        <div style="font-size:0.95rem;font-weight:700;color:${{kolor}}">${{naReke.toFixed(0)}} zł</div>
                        <div style="font-size:0.65rem;color:#64748b;margin-top:2px">Na rękę</div>
                    </div>`;
                }})()}}
            </div>`;
        document.getElementById('monthSummary').style.display = 'block';
    }}
    
    // Wróć do widoku miesięcznego
    function showMonthlyView() {{
        currentView = 'monthly';
        currentMonth = 0;
        
        chartMiesiace.data.labels = nazwyMiesiecy;
        chartMiesiace.data.datasets[0].data = daneMiesieczne.map((v, i) => v + danePrywatne[i]);
        chartMiesiace.data.datasets[0].backgroundColor = 'rgba(59, 130, 246, 0.6)';
        chartMiesiace.data.datasets[0].borderColor = 'rgba(59, 130, 246, 1)';
        chartMiesiace.data.datasets[1].data = daneKosztyLacznie;
        chartMiesiace.data.datasets[1].hidden = false;
        chartMiesiace.data.datasets[2].data = danePaletyZakup;
        chartMiesiace.data.datasets[2].hidden = false;
        chartMiesiace.data.datasets[3].data = daneZysk;
        chartMiesiace.data.datasets[3].hidden = false;
        chartMiesiace.data.datasets[3].spanGaps = false;
        chartMiesiace.update();
        
        document.getElementById('chartTitle').textContent = ' Sprzedaż miesięcznie ({current_year})';
        document.getElementById('btnBack').style.display = 'none';
        document.getElementById('monthSummary').style.display = 'none';
        document.getElementById('monthSummary').innerHTML = '';
    }}
    
    // Inicjalizacja
    initMonthlyChart();

    // Auto-otwórz miesiąc z URL (?miesiac=2026-03)
    const urlParams = new URLSearchParams(window.location.search);
    const paramMiesiac = urlParams.get('miesiac');
    if (paramMiesiac) {{
        const parts = paramMiesiac.split('-');
        if (parts.length === 2) {{
            const monthNum = parseInt(parts[1]);
            if (monthNum >= 1 && monthNum <= 12) {{
                setTimeout(() => showDailyView(monthNum), 500);
            }}
        }}
    }}
    
    // Wykres roczny
    new Chart(document.getElementById('chartLata'), {{
        type: 'bar',
        data: {{
            labels: {json.dumps(dane_roczne_labels)},
            datasets: [{{
                label: 'Przychód (zł)',
                data: {json.dumps(dane_roczne_values)},
                backgroundColor: 'rgba(34, 197, 94, 0.8)',
                borderColor: 'rgba(34, 197, 94, 1)',
                borderWidth: 1,
                borderRadius: 5
            }}]
        }},
        options: {{
            responsive: true,
            plugins: {{
                legend: {{ display: false }}
            }},
            scales: {{
                y: {{
                    beginAtZero: true,
                    grid: {{ color: 'rgba(255,255,255,0.1)' }},
                    ticks: {{ color: '#64748b' }}
                }},
                x: {{
                    grid: {{ display: false }},
                    ticks: {{ color: '#64748b' }}
                }}
            }}
        }}
    }});
    </script>
    '''
    return render(html)

@magazynier_bp.route('/palety')
def palety():
    from urllib.parse import quote
    conn = get_db()
    
    # Pobierz palety z tabeli palety + statystyki z produkty
    try:
        result = conn.execute('''
            SELECT p.id, p.nazwa, p.dostawca, p.data_zakupu, p.cena_zakupu, p.cena_zakupu_netto,
                   p.ilosc_sztuk, 0 as dostarczona, COALESCE(p.typ, 'paleta') as typ,
                   COUNT(pr.id) as cnt,
                   COALESCE(SUM(pr.ilosc), 0) as items,
                   COALESCE(SUM(pr.cena_allegro * pr.ilosc), 0) as wartosc_allegro
            FROM palety p
            LEFT JOIN produkty pr ON pr.paleta_id = p.id
            GROUP BY p.id
            ORDER BY p.id DESC
        ''').fetchall()
    except:
        result = conn.execute('''
            SELECT p.id, p.nazwa, p.dostawca, p.data_zakupu, p.cena_zakupu,
                   0 as cena_zakupu_netto, 0 as ilosc_sztuk, 0 as dostarczona, 'paleta' as typ,
                   COUNT(pr.id) as cnt,
                   COALESCE(SUM(pr.ilosc), 0) as items,
                   COALESCE(SUM(pr.cena_allegro * pr.ilosc), 0) as wartosc_allegro
            FROM palety p
            LEFT JOIN produkty pr ON pr.paleta_id = p.id
            GROUP BY p.id
            ORDER BY p.id DESC
        ''').fetchall()
    
    # Dodaj też produkty bez palety
    bez_palety = conn.execute('''
        SELECT COUNT(*) as cnt, COALESCE(SUM(ilosc), 0) as items,
               COALESCE(SUM(cena_brutto), 0) as wartosc_zakupu,
               COALESCE(SUM(cena_allegro * ilosc), 0) as wartosc_allegro
        FROM produkty WHERE paleta_id IS NULL OR paleta_id = 0
    ''').fetchone()
    
    # Pobierz dostarczona osobno (z auto-migracją jeśli kolumna nie istnieje)
    dostarczona_map = {}
    try:
        d_rows = conn.execute('SELECT id, COALESCE(dostarczona, 0) as dostarczona FROM palety').fetchall()
        dostarczona_map = {r['id']: r['dostarczona'] for r in d_rows}
    except:
        try:
            conn.execute('ALTER TABLE palety ADD COLUMN dostarczona INTEGER DEFAULT 0')
            conn.commit()
            print('[CHECK_CIRCLE] Dodano kolumnę dostarczona')
        except:
            pass

    
    total_palety = len(result)

    html = f'''
    <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap" rel="stylesheet">
    <style>
        .pl-label{{font-family:'Manrope',sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:0.2em;color:rgba(255,255,255,0.45)}}
        .pl-headline{{font-family:'Space Grotesk',sans-serif}}
        .pl-card{{background:rgba(25,25,28,0.7);backdrop-filter:blur(12px);padding:20px;border:1px solid rgba(255,255,255,0.05);margin-bottom:12px;transition:all 0.2s;position:relative}}
        .pl-card:hover{{background:rgba(31,31,34,0.9);box-shadow:0 0 15px rgba(0,0,0,0.3)}}
        .pl-badge{{display:inline-block;padding:2px 8px;font-size:9px;font-weight:900;text-transform:uppercase;letter-spacing:0.1em;font-family:'Manrope',sans-serif}}
        .pl-badge-box{{background:rgba(143,245,255,0.1);color:#8ff5ff;border:1px solid rgba(143,245,255,0.2)}}
        .pl-badge-paleta{{background:rgba(190,238,0,0.1);color:#beee00;border:1px solid rgba(190,238,0,0.2)}}
        .pl-status-ok{{border-color:rgba(190,238,0,0.4);color:#beee00;background:rgba(190,238,0,0.05);box-shadow:0 0 5px rgba(190,238,0,0.2)}}
        .pl-status-ship{{border-color:rgba(255,107,155,0.4);color:#ff6b9b;background:rgba(255,107,155,0.05);box-shadow:0 0 5px rgba(255,107,155,0.2)}}
        .pl-bar-left{{position:absolute;left:0;top:0;bottom:0;width:3px;border-radius:2px}}
        .pl-action{{flex-shrink:0;padding:6px 14px;border:1px solid rgba(255,255,255,0.1);background:transparent;color:rgba(255,255,255,0.7);font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;cursor:pointer;font-family:'Manrope',sans-serif;transition:all 0.15s;display:inline-flex;align-items:center;gap:4px;border-radius:20px}}
        .pl-action:hover{{transform:translateY(-1px)}}
        .pl-action-green{{border-color:rgba(190,238,0,0.3);color:#beee00}}.pl-action-green:hover{{background:rgba(190,238,0,0.1)}}
        .pl-action-amber{{border-color:rgba(255,173,194,0.3);color:#ffadc2}}.pl-action-amber:hover{{background:rgba(255,173,194,0.1)}}
        .pl-action-red{{border-color:rgba(255,113,108,0.3);color:#ff716c}}.pl-action-red:hover{{background:rgba(255,113,108,0.1)}}
        .pl-action-cyan{{border-color:rgba(143,245,255,0.3);color:#8ff5ff}}.pl-action-cyan:hover{{background:rgba(143,245,255,0.1)}}
    </style>

    <!-- Header -->
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:24px">
        <div style="display:flex;align-items:center;gap:12px">
            <div style="padding:8px;background:rgba(143,245,255,0.1);border:1px solid rgba(143,245,255,0.2)">
                <span class=material-symbols-outlined style=font-size:28px;color:#8ff5ff>inventory_2</span>
            </div>
            <div>
                <h1 class="pl-headline" style="font-size:clamp(1.8rem,4vw,2.5rem);font-weight:700;letter-spacing:-0.03em;color:#8ff5ff;margin:0;text-shadow:0 0 10px rgba(143,245,255,0.4)">PALETY</h1>
                <div class="pl-label">Inventory Hub</div>
            </div>
        </div>
        <div style="text-align:right">
            <div class="pl-label">Total</div>
            <div class="pl-headline" style="font-size:1.5rem;font-weight:700">{total_palety}</div>
        </div>
    </div>

    <!-- Actions -->
    <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px;align-items:center">
        <button onclick="selectAll()" class="pl-action pl-action-cyan">
            <span class=material-symbols-outlined style=font-size:14px>select_all</span> Wszystkie
        </button>
        <button onclick="selectNone()" class="pl-action" style="border-color:rgba(255,255,255,0.15)">
            <span class=material-symbols-outlined style=font-size:14px>deselect</span> Odznacz
        </button>
        <div style="flex:1"></div>
        <button onclick="massUpdate(1)" class="pl-action pl-action-green">
            <span class=material-symbols-outlined style=font-size:14px>task_alt</span> Dostarczone
        </button>
        <button onclick="massUpdate(0)" class="pl-action pl-action-amber">
            <span class=material-symbols-outlined style=font-size:14px>local_shipping</span> W drodze
        </button>
        <button onclick="pokazBoxPaletyModal()" class="pl-action pl-action-cyan">
            <span class=material-symbols-outlined style=font-size:14px>inventory</span> Zgrupuj w box
        </button>
        <button onclick="massDelete()" class="pl-action pl-action-red">
            <span class=material-symbols-outlined style=font-size:14px>delete</span> Usuń
        </button>
        <span id="selectedCount" style="font-size:11px;color:rgba(255,255,255,0.4);margin-left:4px">(0)</span>
    </div>

    <!-- Search -->
    <div style="position:relative;margin-bottom:20px">
        <span class=material-symbols-outlined style=position:absolute;left:14px;top:50%;transform:translateY(-50%);color:#8ff5ff;font-size:18px>search</span>
        <input type="text" id="paletaSearch" oninput="searchPalety()" placeholder="SZUKAJ PALETY LUB SKU..."
            style="width:100%;padding:14px 14px 14px 44px;background:rgba(38,37,40,0.8);border:none;border-bottom:1px solid rgba(143,245,255,0.15);color:#f9f5f8;font-size:12px;letter-spacing:0.15em;font-family:'Manrope',sans-serif;outline:none">
    </div>'''

    # Pobierz nazwy produktów per paleta (do wyszukiwania)
    _prod_names = {}
    try:
        _pn_rows = conn.execute('SELECT paleta_id, GROUP_CONCAT(nazwa, " ") as names FROM produkty WHERE paleta_id IS NOT NULL GROUP BY paleta_id').fetchall()
        _prod_names = {r['paleta_id']: (r['names'] or '') for r in _pn_rows}
    except:
        pass

    for p in result:
        link = f"/magazyn/paleta-id/{p['id']}"
        dostawca_info = f' • <span class="dostawca-name">{p["dostawca"]}</span>' if p['dostawca'] else ""
        data_info = f" • {p['data_zakupu']}" if p['data_zakupu'] else ""
        
        # Ilość sztuk: z palety (preferowane) lub z produktów
        try:
            sztuki = p['ilosc_sztuk'] if p['ilosc_sztuk'] and p['ilosc_sztuk'] > 0 else p['items']
        except:
            sztuki = p['items']
        
        # Kolory w zależności od stanu
        cnt_color = "#beee00" if p['cnt'] > 0 else "#ef4444"
        
        # Cena zakupu: cena_zakupu w bazie = BRUTTO z faktury
        zakup_brutto = p['cena_zakupu'] or 0
        
        dostarczona = dostarczona_map.get(p['id'], 0)
        dostarczona_label = '<span class=material-symbols-outlined>check_circle</span> Dostarczona' if dostarczona else '<span class=material-symbols-outlined>local_shipping</span> W drodze'
        dostarczona_color = '#beee00' if dostarczona else '#f59e0b'
        is_box = p['typ'] == 'box'
        bar_color = '#ff6b9b' if not dostarczona else '#beee00'
        status_class = 'pl-status-ok' if dostarczona else 'pl-status-ship'
        badge_class = 'pl-badge-box' if is_box else 'pl-badge-paleta'
        badge_text = 'BOX' if is_box else 'PALETA'
        icon = 'package_2' if is_box else 'inventory'
        wartosc = p['wartosc_allegro'] or 0
        dostawca_txt = p['dostawca'] or ''
        data_txt = p['data_zakupu'] or ''

        _pn = _prod_names.get(p['id'], '').replace('"', '&quot;')
        html += f'''<div class="pl-card" onclick="window.location='{link}'" style="cursor:pointer;padding-left:24px" data-products="{_pn}">
            <div class="pl-bar-left" style="background:{bar_color};box-shadow:0 0 10px {bar_color}80"></div>
            <div style="display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:12px">
                <div style="display:flex;gap:12px;align-items:flex-start;flex:1;min-width:0">
                    <input type="checkbox" class="paleta-cb" data-id="{p['id']}" onclick="event.stopPropagation();licz()"
                        style="width:18px;height:18px;cursor:pointer;accent-color:#8ff5ff;flex-shrink:0;margin-top:4px">
                    <div style="min-width:0">
                        <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
                            <span class=material-symbols-outlined style=font-size:18px;color:rgba(255,255,255,0.5)>{icon}</span>
                            <h3 class="pl-headline item-name" style="font-size:1.05rem;font-weight:700;color:#f9f5f8;margin:0;line-height:1.2">{p['nazwa']}</h3>
                        </div>
                        <div style="display:flex;align-items:center;gap:6px">
                            <span class="pl-badge {badge_class}">{badge_text}</span>
                            <span style="font-size:11px;color:rgba(255,255,255,0.4)">#{p['id']}</span>
                        </div>
                    </div>
                </div>
                <button onclick="event.stopPropagation();toggleDostarczona({p['id']}, this)" data-val="{dostarczona}"
                    style="padding:4px 10px;border:1px solid;font-size:10px;font-weight:700;text-transform:uppercase;cursor:pointer;white-space:nowrap;font-family:'Manrope',sans-serif;background:transparent" class="{status_class}">
                    {dostarczona_label}
                </button>
            </div>
            <div style="display:flex;align-items:flex-end;justify-content:space-between;border-top:1px solid rgba(255,255,255,0.05);padding-top:12px">
                <div>
                    <div class="item-meta" style="display:flex;align-items:center;gap:6px;font-size:10px;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px">
                        <span class=material-symbols-outlined style=font-size:12px>list_alt</span>
                        {p['cnt']} prod. | {sztuki} szt.
                    </div>
                    <div style="display:flex;align-items:center;gap:6px;font-size:10px;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.1em">
                        <span class=material-symbols-outlined style=font-size:12px>event</span>
                        {dostawca_txt}{' &bull; ' + data_txt if data_txt else ''}
                    </div>
                </div>
                <div style="text-align:right">
                    <div style="font-size:10px;color:rgba(255,255,255,0.4);margin-bottom:2px">ZAKUP: {zakup_brutto:.0f} ZŁ</div>
                    <div class="pl-headline" style="font-size:1.5rem;font-weight:700;color:#8ff5ff;letter-spacing:-0.03em;text-shadow:0 0 8px rgba(143,245,255,0.3)">{wartosc:.0f} ZŁ</div>
                </div>
            </div>
        </div>'''
    
    # Produkty bez palety
    if bez_palety['cnt'] > 0:
        html += f'''<div class="pl-card" onclick="window.location='/magazyn/paleta/brak'" style="cursor:pointer;padding-left:24px;border-color:rgba(255,107,155,0.2)">
            <div class="pl-bar-left" style="background:#ff716c;box-shadow:0 0 10px rgba(255,113,108,0.5)"></div>
            <div style="display:flex;align-items:center;justify-content:space-between">
                <div style="display:flex;align-items:center;gap:10px">
                    <span class=material-symbols-outlined style=font-size:22px;color:#ff716c>warning</span>
                    <div>
                        <div class="pl-headline item-name" style="font-size:1rem;font-weight:700;color:#ff716c">Bez palety</div>
                        <div style="font-size:10px;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.1em">{bez_palety['cnt']} prod. | {bez_palety['items']} szt.</div>
                    </div>
                </div>
                <div class="pl-headline" style="font-size:1.3rem;font-weight:700;color:#ff716c">{bez_palety['wartosc_allegro'] or 0:.0f} ZŁ</div>
            </div>
        </div>'''
    
    html += '''<script>
    var paletyCbs = document.getElementsByClassName("paleta-cb");
    var countEl = document.getElementById("selectedCount");

    function licz() {
        var n = 0;
        for (var i = 0; i < paletyCbs.length; i++) {
            if (paletyCbs[i].checked) n++;
        }
        countEl.innerText = "(" + n + " zaznaczonych)";
        return n;
    }

    // Podepnij onclick na kazdy checkbox
    for (var i = 0; i < paletyCbs.length; i++) {
        paletyCbs[i].onclick = function(e) {
            e.stopPropagation();
            licz();
        };
    }

    function selectAll() {
        for (var i = 0; i < paletyCbs.length; i++) paletyCbs[i].checked = true;
        licz();
    }
    function selectNone() {
        for (var i = 0; i < paletyCbs.length; i++) paletyCbs[i].checked = false;
        licz();
    }

    function getSelectedIds() {
        var ids = [];
        for (var i = 0; i < paletyCbs.length; i++) {
            if (paletyCbs[i].checked) ids.push(parseInt(paletyCbs[i].getAttribute("data-id")));
        }
        return ids;
    }

    function massUpdate(val) {
        var ids = getSelectedIds();
        if (!ids.length) { alert("Zaznacz najpierw palety"); return; }
        fetch("/magazyn/api/paleta-dostarczona-bulk", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({ids: ids, dostarczona: val})
        }).then(function(r){return r.json()}).then(function(d) {
            if (d.ok) location.reload();
        });
    }

    function toggleDostarczona(paletaId, btn) {
        var newVal = btn.getAttribute("data-val") == "1" ? 0 : 1;
        fetch("/magazyn/api/paleta-dostarczona/" + paletaId, {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({dostarczona: newVal})
        }).then(function(r){return r.json()}).then(function(d) {
            if (d.ok) {
                btn.setAttribute("data-val", newVal);
                if (newVal == 1) {
                    btn.innerText = "\\u2705 Dostarczona";
                    btn.style.borderColor = "#beee00";
                    btn.style.color = "#beee00";
                    btn.style.background = "#beee0022";
                } else {
                    btn.innerText = "\\ud83d\\ude9a W drodze";
                    btn.style.borderColor = "#f59e0b";
                    btn.style.color = "#f59e0b";
                    btn.style.background = "#f59e0b22";
                }
            }
        });
    }

    function massDelete() {
        var ids = getSelectedIds();
        if (!ids.length) { alert("Zaznacz najpierw palety"); return; }
        if (!confirm("Usunac " + ids.length + " palet i ich produkty? Tej operacji nie mozna cofnac!")) return;
        fetch("/magazyn/api/palety-usun", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({ids: ids})
        }).then(function(r){return r.json()}).then(function(d) {
            if (d.ok) location.reload();
            else alert("Blad: " + (d.error || "nieznany"));
        });
    }

    function searchPalety() {
        var q = document.getElementById("paletaSearch").value.toLowerCase();
        var items = document.getElementsByClassName("pl-card");
        for (var i = 0; i < items.length; i++) {
            var name = items[i].querySelector(".item-name");
            var meta = items[i].querySelector(".item-meta");
            var products = items[i].getAttribute("data-products") || "";
            var text = (name ? name.textContent : "") + " " + (meta ? meta.textContent : "") + " " + products;
            items[i].style.display = (!q || text.toLowerCase().indexOf(q) >= 0) ? "" : "none";
        }
    }

    function pokazBoxPaletyModal() {
        var ids = getSelectedIds();
        if (ids.length < 1) { alert("Zaznacz palety do zgrupowania w box"); return; }
        var names = [];
        ids.forEach(function(id) {
            var cb = document.querySelector('.paleta-cb[data-id="'+id+'"]');
            if (cb) {
                var item = cb.closest('.item');
                var name = item ? item.querySelector('.item-name') : null;
                names.push(name ? name.textContent.trim().substring(0, 40) : '#'+id);
            }
        });
        document.getElementById('boxPaletyList').innerHTML = names.map(function(n){return '<div style="padding:4px 0;border-bottom:1px solid var(--border);font-size:0.82rem">'+n+'</div>'}).join('');
        document.getElementById('boxPaletyIds').value = JSON.stringify(ids);
        document.getElementById('boxPaletyCount').textContent = ids.length;
        document.getElementById('boxPaletyNazwa').value = '';
        document.getElementById('boxPaletyCena').value = '';
        document.getElementById('modalBoxPalety').style.display = 'flex';
    }

    function zapiszBoxPalety() {
        var ids = JSON.parse(document.getElementById('boxPaletyIds').value);
        var nazwa = document.getElementById('boxPaletyNazwa').value.trim();
        var cena = parseFloat(document.getElementById('boxPaletyCena').value) || 0;
        var cenaZakupu = parseFloat(document.getElementById('boxPaletyCenaZakupu').value) || 0;
        if (!nazwa) { alert('Podaj nazwę boxa'); return; }
        var btn = document.getElementById('boxPaletySaveBtn');
        btn.disabled = true; btn.textContent = ' Tworzę...';
        fetch('/magazyn/api/zgrupuj-palety-box', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({paleta_ids: ids, nazwa: nazwa, cena_sprzedazy: cena, cena_zakupu: cenaZakupu})
        }).then(function(r){return r.json()}).then(function(d) {
            if (d.ok) {
                btn.textContent = ' Utworzono!';
                setTimeout(function(){ window.location.href = '/palety/' + d.box_id; }, 800);
            } else {
                btn.textContent = ' ' + (d.error||'Błąd'); btn.disabled = false;
            }
        }).catch(function(e){ btn.textContent = ' ' + e.message; btn.disabled = false; });
    }
    </script>

    <!-- Modal: Zgrupuj palety w Box -->
    <div id="modalBoxPalety" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.7);z-index:999;align-items:center;justify-content:center">
        <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border-radius:14px;padding:25px;max-width:450px;width:90%;max-height:80vh;overflow-y:auto;border:2px solid #f59e0b">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:15px">
                <h3 style="margin:0;color:#f59e0b"><span class=material-symbols-outlined>inbox</span> Zgrupuj w Box</h3>
                <button onclick="document.getElementById('modalBoxPalety').style.display='none'" style="background:none;border:none;color:var(--text-muted);font-size:1.3rem;cursor:pointer">&times;</button>
            </div>
            <div style="margin-bottom:12px;padding:10px;backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:10px">
                <div style="font-size:0.8rem;color:var(--text-muted);margin-bottom:5px">Palety (<span id="boxPaletyCount">0</span>):</div>
                <div id="boxPaletyList" style="max-height:120px;overflow-y:auto"></div>
            </div>
            <input type="hidden" id="boxPaletyIds" value="[]">
            <div style="margin-bottom:10px">
                <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px">Nazwa boxa</label>
                <input type="text" id="boxPaletyNazwa" placeholder="np. Box mix elektronika" style="width:100%;padding:10px;background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;color:#e2e8f0">
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:15px">
                <div>
                    <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px"><span class=material-symbols-outlined>paid</span> Cena zakupu (zł)</label>
                    <input type="number" id="boxPaletyCenaZakupu" placeholder="Ile zapłaciłeś" step="0.01" style="width:100%;padding:10px;background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;color:#e2e8f0">
                </div>
                <div>
                    <label style="display:block;color:var(--text-secondary);font-size:0.8rem;margin-bottom:4px"><span class=material-symbols-outlined>shopping_cart</span> Cena sprzedaży (zł)</label>
                    <input type="number" id="boxPaletyCena" placeholder="Cena na Allegro" step="0.01" style="width:100%;padding:10px;background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:8px;color:#e2e8f0">
                </div>
            </div>
            <button id="boxPaletySaveBtn" onclick="zapiszBoxPalety()" style="width:100%;padding:12px;background:#f59e0b;border:none;border-radius:8px;color:#000;font-weight:700;cursor:pointer;font-size:1rem">
                <span class=material-symbols-outlined>inbox</span> Utwórz Box
            </button>
        </div>
    </div>'''
    html += '<a href="/magazyn" class="back">← Powrót</a>'
    return render(html)

@magazynier_bp.route('/api/paleta-dostarczona-bulk', methods=['POST'])
def api_paleta_dostarczona_bulk():
    from flask import jsonify, request as req
    conn = get_db()
    try:
        data = req.get_json()
        ids = data.get('ids', [])
        val = int(data.get('dostarczona', 0))
        for pid in ids:
            conn.execute('UPDATE palety SET dostarczona = ? WHERE id = ?', (val, int(pid)))
        conn.commit()
        return jsonify({'ok': True, 'updated': len(ids)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@magazynier_bp.route('/api/paleta-edit/<int:paleta_id>', methods=['POST'])
def api_paleta_edit(paleta_id):
    """Edycja danych palety"""
    from flask import jsonify, request as req
    conn = get_db()
    try:
        data = req.get_json()
        nazwa = data.get('nazwa', '')
        cena_zakupu = float(data.get('cena_zakupu', 0) or 0)
        dostawca = data.get('dostawca', '')
        regal = data.get('regal', '')
        conn.execute('''
            UPDATE palety SET nazwa = ?, cena_zakupu = ?, dostawca = ?, regal = ?
            WHERE id = ?
        ''', (nazwa, cena_zakupu, dostawca, regal, paleta_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@magazynier_bp.route('/api/palety-usun', methods=['POST'])
def api_palety_usun():
    """Masowe usuwanie palet + ich produktów"""
    from flask import jsonify, request as req
    conn = get_db()
    try:
        data = req.get_json()
        ids = data.get('ids', [])
        for pid in ids:
            pid = int(pid)
            # Usuń produkty przypisane do palety
            conn.execute('DELETE FROM produkty WHERE paleta_id = ?', (pid,))
            # Usuń paletę
            conn.execute('DELETE FROM palety WHERE id = ?', (pid,))
        conn.commit()
        print(f"[DELETE] Usunięto {len(ids)} palet z produktami")
        return jsonify({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@magazynier_bp.route('/api/paleta-dostarczona/<int:paleta_id>', methods=['POST'])
def api_paleta_dostarczona(paleta_id):
    from flask import jsonify, request as req
    import json
    conn = get_db()
    try:
        data = req.get_json()
        val = int(data.get('dostarczona', 0))
        conn.execute('UPDATE palety SET dostarczona = ? WHERE id = ?', (val, paleta_id))
        conn.commit()
        return jsonify({'ok': True, 'dostarczona': val})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@magazynier_bp.route('/paleta-id/<int:paleta_id>')
def paleta_detail_by_id(paleta_id):
    """Szczegóły palety po ID (bezpieczniejsze niż po nazwie)"""
    from urllib.parse import quote
    conn = get_db()
    
    # Pobierz dane palety z ceną zakupu
    try:
        paleta_row = conn.execute('SELECT nazwa, dostawca, regal, ilosc_sztuk, cena_zakupu, cena_zakupu_netto, COALESCE(dostarczona, 0) as dostarczona FROM palety WHERE id = ?', (paleta_id,)).fetchone()
        ilosc_sztuk_paleta = paleta_row['ilosc_sztuk'] or 0 if paleta_row else 0
        cena_zakupu = paleta_row['cena_zakupu'] or 0 if paleta_row else 0
        try:
            cena_zakupu_netto = paleta_row['cena_zakupu_netto'] or 0 if paleta_row else 0
        except:
            cena_zakupu_netto = 0
        # cena_zakupu w bazie = BRUTTO z faktury
        # netto = brutto / 1.23
        if cena_zakupu > 0 and cena_zakupu_netto == 0:
            cena_zakupu_netto = round(cena_zakupu / 1.23, 2)
    except:
        try:
            paleta_row = conn.execute('SELECT nazwa, dostawca, regal, ilosc_sztuk, cena_zakupu FROM palety WHERE id = ?', (paleta_id,)).fetchone()
            ilosc_sztuk_paleta = paleta_row['ilosc_sztuk'] or 0 if paleta_row else 0
            cena_zakupu = paleta_row['cena_zakupu'] or 0 if paleta_row else 0
            cena_zakupu_netto = round(cena_zakupu / 1.23, 2)
        except:
            paleta_row = conn.execute('SELECT nazwa, dostawca, regal FROM palety WHERE id = ?', (paleta_id,)).fetchone()
            ilosc_sztuk_paleta = 0
            cena_zakupu = 0
            cena_zakupu_netto = 0
    
    if not paleta_row:
        return redirect('/magazyn/palety')
    
    nazwa_palety = paleta_row['nazwa']
    
    # Pobierz produkty
    products = conn.execute('SELECT * FROM produkty WHERE paleta_id = ?', (paleta_id,)).fetchall()

    # Mapa: produkt_id → status oferty na Allegro (realne dane z tabeli oferty)
    _oferty_status = {}
    try:
        _of_rows = conn.execute('''
            SELECT produkt_id, status, allegro_id
            FROM oferty
            WHERE produkt_id IN (SELECT id FROM produkty WHERE paleta_id = ?)
            ORDER BY CASE status WHEN 'aktywna' THEN 0 WHEN 'draft' THEN 1 ELSE 2 END
        ''', (paleta_id,)).fetchall()
        for r in _of_rows:
            # Zachowaj najlepszy status per produkt (aktywna > draft > zakonczona)
            if r['produkt_id'] not in _oferty_status:
                _oferty_status[r['produkt_id']] = {'status': r['status'], 'allegro_id': r['allegro_id']}
    except Exception:
        pass

    # Statystyki palety
    stats = conn.execute('''
        SELECT COUNT(*) as cnt, SUM(ilosc) as items, 
               SUM(CASE WHEN status IN ('wystawiony', 'szkic') THEN cena_allegro*ilosc ELSE 0 END) as allegro,
               SUM(cena_allegro * ilosc) as allegro_all,
               SUM(cena_brutto) as suma_produktow_brutto
        FROM produkty WHERE paleta_id = ?
    ''', (paleta_id,)).fetchone()
    
    # Cena zakupu palety
    # Fallback: jeśli paleta nie ma wpisanej ceny, użyj sumy produktów
    suma_prod_brutto = stats['suma_produktow_brutto'] or 0
    
    if cena_zakupu > 0:
        # Upewnij się że cena_zakupu to BRUTTO (większa liczba)
        # Jeśli cena_zakupu < suma_prod / 1.5 to prawdopodobnie jest netto
        if suma_prod_brutto > 0 and cena_zakupu < suma_prod_brutto / 1.5:
            # cena_zakupu wydaje się być netto
            netto = cena_zakupu
            brutto = round(cena_zakupu * 1.23, 2)
        else:
            brutto = cena_zakupu
            netto = round(cena_zakupu / 1.23, 2)
    elif suma_prod_brutto > 0:
        brutto = suma_prod_brutto
        netto = round(brutto / 1.23, 2)
    else:
        brutto = 0
        netto = 0
    
    # Ostateczne zabezpieczenie - netto zawsze < brutto
    if netto > brutto and brutto > 0:
        netto, brutto = brutto, netto
    
    allegro = stats['allegro_all'] or 0  # Potencjalny przychód ze WSZYSTKICH produktów
    zysk = allegro - brutto - (allegro * 0.11)
    
    # Użyj ilosc_sztuk z palety jeśli jest, w przeciwnym razie z produktów
    sztuki_display = ilosc_sztuk_paleta if ilosc_sztuk_paleta > 0 else (stats['items'] or 0)
    
    dostarczona_val = paleta_row['dostarczona'] if paleta_row and 'dostarczona' in paleta_row.keys() else 0
    dostarczona_label = '<span class=material-symbols-outlined>check_circle</span> Dostarczona' if dostarczona_val else '<span class=material-symbols-outlined>local_shipping</span> W drodze'
    dostarczona_color = '#beee00' if dostarczona_val else '#f59e0b'
    paleta_dostawca = paleta_row['dostawca'] if 'dostawca' in paleta_row.keys() else ''
    paleta_regal = paleta_row['regal'] if 'regal' in paleta_row.keys() else ''
    dostawca_badge = f' • <span class="dostawca-name" style="color:#8ff5ff">{paleta_dostawca}</span>' if paleta_dostawca else ''
    regal_badge = f' • <span class=material-symbols-outlined>pin_drop</span> {paleta_regal}' if paleta_regal else ''
    # Statystyka ofert Allegro per paleta
    _of_aktywne = sum(1 for pid, d in _oferty_status.items() if d['status'] == 'aktywna')
    _of_szkice = sum(1 for pid, d in _oferty_status.items() if d['status'] == 'draft')
    _of_brak = len(products) - len(_oferty_status)
    _of_summary = f' • <span style="color:#beee00">{_of_aktywne} aktywnych</span>' if _of_aktywne else ''
    _of_summary += f' <span style="color:#f59e0b">{_of_szkice} szkiców</span>' if _of_szkice else ''
    _of_summary += f' <span style="color:#ef4444">{_of_brak} niewystawionych</span>' if _of_brak > 0 else ''

    html = f'''<div class="hdr">
        <div><h1><span class=material-symbols-outlined>inventory_2</span> {nazwa_palety}</h1><small>{len(products)} prod. ({sztuki_display} szt.){dostawca_badge}{regal_badge}{_of_summary}</small></div>
        <div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:10px">
            <button id="btnDostarczona" onclick="toggleDostarczona({paleta_id}, this)"
                data-val="{dostarczona_val}"
                style="padding:8px 14px;border:2px solid {dostarczona_color};background:{dostarczona_color}22;color:{dostarczona_color};border-radius:10px;font-size:0.85rem;font-weight:600;cursor:pointer;white-space:nowrap">
                {dostarczona_label}
            </button>
            <a href="/magazyn/przyjecie/{paleta_id}" style="padding:8px 14px;border:2px solid #7c3aed;background:#7c3aed22;color:#7c3aed;border-radius:10px;font-size:0.85rem;font-weight:600;cursor:pointer;text-decoration:none;white-space:nowrap">
                <span class=material-symbols-outlined>list_alt</span> Przyjęcie
            </a>
            <a href="/magazyn/etykiety?paleta_id={paleta_id}" style="padding:8px 14px;border:2px solid #ff6b9b;background:#ff6b9b22;color:#ff6b9b;border-radius:10px;font-size:0.85rem;font-weight:600;cursor:pointer;text-decoration:none;white-space:nowrap">
                <span class=material-symbols-outlined>label</span> Etykiety
            </a>
            <button onclick="document.getElementById('editPaletaModal').style.display='flex'" style="padding:8px 14px;border:2px solid #f59e0b;background:#f59e0b22;color:#f59e0b;border-radius:10px;font-size:0.85rem;font-weight:600;cursor:pointer;white-space:nowrap">
                <span class=material-symbols-outlined>edit</span> Edytuj
            </button>
            <a href="/magazyn/paleta-etykieta/{paleta_id}" style="padding:8px 14px;border:2px solid #22c55e;background:#22c55e22;color:#22c55e;border-radius:10px;font-size:0.85rem;font-weight:600;cursor:pointer;text-decoration:none;white-space:nowrap">
                <span class=material-symbols-outlined>print</span> Drukuj etykietę
            </a>
        </div>
    </div>

    <!-- Modal edycji palety -->
    <div id="editPaletaModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.8);z-index:99999;align-items:center;justify-content:center;padding:20px" onclick="if(event.target===this)this.style.display='none'">
        <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:16px;padding:24px;width:100%;max-width:450px">
            <h3 style="margin:0 0 16px;font-size:1.1rem"><span class=material-symbols-outlined>edit</span> Edytuj paletę #{paleta_id}</h3>
            <form id="editPaletaForm" onsubmit="savePaleta(event)">
                <div style="margin-bottom:12px">
                    <label style="font-size:0.8rem;color:#64748b;display:block;margin-bottom:4px">Nazwa</label>
                    <input type="text" id="ep_nazwa" value="{nazwa_palety}" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#e2e8f0;font-size:0.9rem">
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">
                    <div>
                        <label style="font-size:0.8rem;color:#64748b;display:block;margin-bottom:4px"><span class=material-symbols-outlined>paid</span> Cena zakupu brutto (zł)</label>
                        <input type="number" step="0.01" id="ep_cena" value="{brutto:.2f}" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#e2e8f0;font-size:0.9rem">
                    </div>
                    <div>
                        <label style="font-size:0.8rem;color:#64748b;display:block;margin-bottom:4px">Dostawca</label>
                        <input type="text" id="ep_dostawca" value="{paleta_dostawca}" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#e2e8f0;font-size:0.9rem">
                    </div>
                </div>
                <div style="margin-bottom:16px">
                    <label style="font-size:0.8rem;color:#64748b;display:block;margin-bottom:4px">Regał / lokalizacja</label>
                    <input type="text" id="ep_regal" value="{paleta_regal}" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#e2e8f0;font-size:0.9rem">
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                    <button type="button" onclick="document.getElementById('editPaletaModal').style.display='none'" style="padding:12px;background:#1e293b;border:1px solid #334155;border-radius:10px;color:#94a3b8;cursor:pointer;font-size:0.9rem">Anuluj</button>
                    <button type="submit" style="padding:12px;background:#f59e0b;border:none;border-radius:10px;color:#000;font-weight:700;cursor:pointer;font-size:0.9rem"><span class=material-symbols-outlined>save</span> Zapisz</button>
                </div>
            </form>
        </div>
    </div>
    <script>
    function savePaleta(e) {{
        e.preventDefault();
        fetch('/magazyn/api/paleta-edit/{paleta_id}', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{
                nazwa: document.getElementById('ep_nazwa').value,
                cena_zakupu: parseFloat(document.getElementById('ep_cena').value) || 0,
                dostawca: document.getElementById('ep_dostawca').value,
                regal: document.getElementById('ep_regal').value
            }})
        }}).then(function(r){{return r.json()}}).then(function(d){{
            if(d.ok) location.reload();
            else alert('Blad: '+(d.error||''));
        }});
    }}
    </script>
    <script>
    function toggleDostarczona(id, btn) {{
        const newVal = btn.dataset.val == '1' ? 0 : 1;
        fetch('/magazyn/api/paleta-dostarczona/' + id, {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{dostarczona: newVal}})
        }}).then(r => r.json()).then(d => {{
            if (d.ok) {{
                btn.dataset.val = newVal;
                if (newVal == 1) {{
                    btn.textContent = ' Dostarczona';
                    btn.style.borderColor = '#beee00';
                    btn.style.color = '#beee00';
                    btn.style.background = '#beee0022';
                }} else {{
                    btn.textContent = ' W drodze';
                    btn.style.borderColor = '#f59e0b';
                    btn.style.color = '#f59e0b';
                    btn.style.background = '#f59e0b22';
                }}
            }}
        }});
    }}
    </script>
    
    <div class="stats" style="margin-bottom:15px">
        <div class="stat">
            <div class="stat-v">{stats['cnt'] or 0}</div>
            <div class="stat-l">PRODUKTÓW</div>
        </div>
        <div class="stat">
            <div class="stat-v">{netto:.0f} zł</div>
            <div class="stat-l">ZAKUP NETTO</div>
        </div>
        <div class="stat" style="border:2px solid #8ff5ff;border-radius:12px">
            <div class="stat-v">{brutto:.0f} zł</div>
            <div class="stat-l">ZAKUP BRUTTO</div>
        </div>
        <div class="stat">
            <div class="stat-v green">{allegro:.0f} zł</div>
            <div class="stat-l">ALLEGRO (suma)</div>
        </div>
        <div class="stat">
            <div class="stat-v" style="color:{('#beee00' if zysk > 0 else '#ef4444')}">{zysk:.0f} zł</div>
            <div class="stat-l">ZYSK</div>
        </div>
    </div>
    <div style="display:flex;gap:8px;margin-bottom:15px;flex-wrap:wrap">
        <button onclick="filterPaleta('all', this)" class="pal-fbtn active" style="padding:6px 14px;border:1px solid rgba(255,255,255,0.15);background:rgba(143,245,255,0.08);color:#8ff5ff;font-size:0.75rem;font-weight:700;cursor:pointer;font-family:'Space Grotesk',sans-serif">Wszystkie ({len(products)})</button>
        <button onclick="filterPaleta('aktywna', this)" class="pal-fbtn" style="padding:6px 14px;border:1px solid rgba(190,238,0,0.2);background:transparent;color:#beee00;font-size:0.75rem;font-weight:700;cursor:pointer;font-family:'Space Grotesk',sans-serif">Aktywne ({_of_aktywne})</button>
        <button onclick="filterPaleta('draft', this)" class="pal-fbtn" style="padding:6px 14px;border:1px solid rgba(245,158,11,0.2);background:transparent;color:#f59e0b;font-size:0.75rem;font-weight:700;cursor:pointer;font-family:'Space Grotesk',sans-serif">Szkice ({_of_szkice})</button>
        <button onclick="filterPaleta('brak', this)" class="pal-fbtn" style="padding:6px 14px;border:1px solid rgba(239,68,68,0.2);background:transparent;color:#ef4444;font-size:0.75rem;font-weight:700;cursor:pointer;font-family:'Space Grotesk',sans-serif">Niewystawione ({_of_brak})</button>
        <button onclick="filterPaleta('magazyn', this)" class="pal-fbtn" style="padding:6px 14px;border:1px solid rgba(143,245,255,0.2);background:transparent;color:#8ff5ff;font-size:0.75rem;font-weight:700;cursor:pointer;font-family:'Space Grotesk',sans-serif">W magazynie ({sum(1 for p in products if (p['ilosc'] or 0) > 0)})</button>
    </div>
    '''
    
    # Przyciski akcji na palecie
    html += '<div style="display:flex;gap:10px;margin-bottom:15px;flex-wrap:wrap">'
    html += f'<a href="/palety/{paleta_id}/mass-edit" class="btn" style="background:var(--purple);flex:1"><span class=material-symbols-outlined>shopping_cart</span> Wystaw bezpośrednio</a>'
    html += f'<a href="/magazyn/paleta-id/{paleta_id}/to-paletomat" class="btn btn-ok" style="flex:1"><span class=material-symbols-outlined>sync</span> PALETOMAT (scrapuj)</a>'
    html += f'<button onclick="autoWycenaPaleta({paleta_id})" class="btn" style="background:#f59e0b;flex:1"><span class=material-symbols-outlined>paid</span> Auto-wycena</button>'
    html += '</div>'
    
    # Script dla Auto-wyceny (streamowane)
    html += '''
    <script>
    async function autoWycenaPaleta(paletaId) {
        const btn = event.target;
        btn.disabled = true;
        btn.innerHTML = '⏳ Pobieranie cen... 0%';

        // Stwórz progress div
        let progressDiv = document.getElementById('autowycena-progress');
        if (!progressDiv) {
            progressDiv = document.createElement('div');
            progressDiv.id = 'autowycena-progress';
            progressDiv.style.cssText = 'background:#1a1a2e;border:1px solid #f59e0b;border-radius:8px;padding:12px;margin:10px 0;font-size:13px;max-height:300px;overflow-y:auto;display:none';
            btn.parentNode.after(progressDiv);
        }
        progressDiv.style.display = 'block';
        progressDiv.innerHTML = '<b><span class=material-symbols-outlined>sync</span> Auto-wycena startuje...</b><br>';

        try {
            const resp = await fetch('/magazyn/api/autowycena-stream/paleta/' + paletaId, {method: 'POST'});
            const reader = resp.body.getReader();
            const decoder = new TextDecoder();
            let buffer = '';
            let stats = {updated:0, total:0, from_amazon:0, from_estimate:0, titles_optimized:0, errors:0};

            while (true) {
                const {done, value} = await reader.read();
                if (done) break;

                buffer += decoder.decode(value, {stream: true});
                const lines = buffer.split('\\n');
                buffer = lines.pop();

                for (const line of lines) {
                    if (!line.startsWith('data: ')) continue;
                    try {
                        const ev = JSON.parse(line.slice(6));
                        if (ev.type === 'progress') {
                            const pct = Math.round(ev.current / ev.total * 100);
                            btn.innerHTML = '⏳ ' + pct + '% (' + ev.current + '/' + ev.total + ')';
                            let color = ev.source === 'amazon' ? '#10b981' : ev.source === 'estimate' ? '#f59e0b' : '#ef4444';
                            progressDiv.innerHTML += '<span style="color:' + color + '">• ' + ev.name + ' → ' + (ev.price ? ev.price + ' zł' : 'brak ceny') + ' [' + (ev.source||'—') + ']</span><br>';
                            progressDiv.scrollTop = progressDiv.scrollHeight;
                        } else if (ev.type === 'done') {
                            stats = ev;
                        } else if (ev.type === 'error') {
                            progressDiv.innerHTML += '<span style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> ' + ev.message + '</span><br>';
                        }
                    } catch(e) {}
                }
            }

            progressDiv.innerHTML += '<br><b style="color:#10b981"><span class=material-symbols-outlined>check_circle</span> Gotowe! Wycenione: ' + stats.updated + '/' + stats.total + ', Tytuły: ' + stats.titles_optimized + ', Błędy: ' + stats.errors + '</b>';

            if (stats.updated > 0) {
                setTimeout(() => location.reload(), 2000);
            }
        } catch (e) {
            progressDiv.innerHTML += '<br><b style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> Błąd: ' + e.message + '</b>';
        }

        btn.disabled = false;
        btn.innerHTML = '<span class=material-symbols-outlined>paid</span> Auto-wycena';
    }
    </script>
    '''
    
    html += '<div style="font-size:0.62rem;color:#8ff5ff;font-weight:700;letter-spacing:0.12em;font-family:\'Space Grotesk\',sans-serif;text-transform:uppercase;margin-bottom:12px;margin-top:20px">Produkty</div>'
    html += '<div style="display:flex;flex-direction:column;gap:10px">'

    for p in products:
        img = _resolve_product_image(p, size='sm')
        pcode = get_product_code(p)
        _ean_c = p['ean'] if p['ean'] and p['ean'].upper() not in ('N/A','NAN','NONE') else ''
        display_code = _ean_c or p['asin'] or f"#{p['id']}"
        _qty = int(p['ilosc'] or 0)
        _ca = float(p['cena_allegro'] or 0)

        # Stock indicator
        if _qty <= 0:
            _stock_dot = '<span style="width:7px;height:7px;border-radius:50%;background:#ef4444;display:inline-block;box-shadow:0 0 5px #ef4444"></span>'
            _stock_text = 'Brak'
        elif _qty <= 2:
            _stock_dot = '<span style="width:7px;height:7px;border-radius:50%;background:#ff6b9b;display:inline-block;box-shadow:0 0 5px #ff6b9b"></span>'
            _stock_text = f'{_qty} szt'
        else:
            _stock_dot = '<span style="width:7px;height:7px;border-radius:50%;background:#beee00;display:inline-block;box-shadow:0 0 5px #beee00"></span>'
            _stock_text = f'{_qty} szt'

        try:
            _klasa = p['klasa_jakosci'] or ''
        except (KeyError, IndexError):
            _klasa = ''
        _border_colors = {'A': '#beee00', 'A-': '#8ff5ff', 'B': '#eab308', 'C': '#f97316', 'D': '#ef4444'}
        _bcolor = _border_colors.get(_klasa, '#48474a')
        _opacity = 'opacity:0.5;' if _qty <= 0 else ''

        # Status oferty na Allegro (realne dane)
        _of = _oferty_status.get(p['id'])
        if _of and _of['status'] == 'aktywna':
            _of_badge = '<span style="font-size:0.58rem;padding:1px 6px;background:rgba(190,238,0,0.12);border:1px solid rgba(190,238,0,0.25);color:#beee00;font-weight:700;letter-spacing:0.3px">AKTYWNA</span>'
        elif _of and _of['status'] == 'draft':
            _of_badge = '<span style="font-size:0.58rem;padding:1px 6px;background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.25);color:#f59e0b;font-weight:700;letter-spacing:0.3px">SZKIC</span>'
        elif _of:
            _of_badge = '<span style="font-size:0.58rem;padding:1px 6px;background:rgba(100,116,139,0.12);border:1px solid rgba(100,116,139,0.25);color:#64748b;font-weight:700;letter-spacing:0.3px">ZAKOŃCZONA</span>'
        else:
            _of_badge = '<span style="font-size:0.58rem;padding:1px 6px;background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.15);color:#ef4444;font-weight:700;letter-spacing:0.3px">NIE WYSTAWIONY</span>'

        _of_data = _oferty_status.get(p['id'], {}).get('status', 'brak')
        _mag_data = 'tak' if _qty > 0 else 'nie'

        html += f'''
        <a href="/magazyn/produkt/{pcode}" data-oferta="{_of_data}" data-magazyn="{_mag_data}" class="pal-prod" style="display:flex;gap:14px;background:#131315;padding:14px;border-radius:10px;border-left:3px solid {_bcolor};text-decoration:none;color:inherit;transition:background 0.2s;{_opacity}"
           onmouseover="this.style.background='#1f1f22'" onmouseout="this.style.background='#131315'">
            <div style="width:64px;height:64px;background:#262528;border-radius:8px;overflow:hidden;flex-shrink:0">
                <img src="{img}" style="width:100%;height:100%;object-fit:cover" onerror="this.src='data:image/svg+xml,%3Csvg xmlns=%27http://www.w3.org/2000/svg%27 width=%2790%27 height=%2790%27%3E%3Crect fill=%27%23262528%27 width=%2790%27 height=%2790%27 rx=%278%27/%3E%3Ctext x=%2745%27 y=%2752%27 fill=%27%23767577%27 text-anchor=%27middle%27 font-size=%2728%27%3E%F0%9F%93%A6%3C/text%3E%3C/svg%3E'">
            </div>
            <div style="display:flex;flex-direction:column;justify-content:center;flex:1;min-width:0">
                <div style="font-family:'Space Grotesk',sans-serif;font-size:0.88rem;font-weight:700;color:#f9f5f8;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{p['nazwa'][:40]}</div>
                <div style="display:flex;align-items:center;gap:8px;margin-top:4px;flex-wrap:wrap">
                    <span style="font-size:0.65rem;color:#adaaad;font-family:monospace">{display_code}</span>
                    {_stock_dot}
                    <span style="font-size:0.6rem;color:#adaaad;font-weight:600">{_stock_text}</span>
                    {_of_badge}
                </div>
            </div>
            <div style="display:flex;align-items:center;flex-shrink:0">
                <span style="font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:800;color:#8ff5ff">{_ca:.0f} zł</span>
            </div>
        </a>'''

    html += '</div>'

    if not products:
        html += '<div style="text-align:center;padding:40px;color:#767577;font-size:0.82rem"><span class=material-symbols-outlined style="font-size:2rem;display:block;margin-bottom:8px">inbox</span>Brak produktów</div>'

    html += '<div style="text-align:center;margin-top:24px"><a href="/magazyn/palety" style="font-size:0.82rem;color:#adaaad;text-decoration:none;font-weight:600;letter-spacing:0.05em">&larr; Powrót</a></div>'

    html += '''
    <script>
    function filterPaleta(type, btn) {
        var items = document.querySelectorAll('.pal-prod');
        items.forEach(function(el) {
            if (type === 'all') { el.style.display = ''; return; }
            if (type === 'magazyn') { el.style.display = el.dataset.magazyn === 'tak' ? '' : 'none'; return; }
            if (type === 'brak') { el.style.display = el.dataset.oferta === 'brak' ? '' : 'none'; return; }
            el.style.display = el.dataset.oferta === type ? '' : 'none';
        });
        document.querySelectorAll('.pal-fbtn').forEach(function(b) {
            b.style.background = 'transparent';
        });
        btn.style.background = 'rgba(143,245,255,0.08)';
    }
    </script>
    '''
    return render(html)


@magazynier_bp.route('/paleta-id/<int:paleta_id>/to-paletomat')
def paleta_to_paletomat_by_id(paleta_id):
    """Przenosi produkty z palety do Paletomat (scrapuje)"""
    conn = get_db()
    
    # Pobierz produkty z palety które mają ASIN
    products = conn.execute('''
        SELECT id, asin, ean, nazwa, cena_brutto, zdjecie_url 
        FROM produkty 
        WHERE paleta_id = ? AND asin IS NOT NULL AND asin != ''
    ''', (paleta_id,)).fetchall()
    
    added = 0
    for p in products:
        asin = p['asin']
        if not asin or asin in ('N/A', 'None'):
            continue
        
        # Sprawdź czy już istnieje w scraped
        existing = conn.execute('SELECT asin FROM scraped WHERE asin = ?', (asin,)).fetchone()
        if not existing:
            try:
                conn.execute('''
                    INSERT INTO scraped (asin, nazwa, zdjecie_url, cena_amazon, status, data_scrape)
                    VALUES (?, ?, ?, ?, 'nowy', datetime('now'))
                ''', (asin, p['nazwa'], p['zdjecie_url'] or '', p['cena_brutto'] or 0))
                added += 1
            except:
                pass
    
    conn.commit()
    
    return redirect(f'/paletomat/scraper?added={added}')


# ============================================================
# ETYKIETA PALETY — druk
# ============================================================
@magazynier_bp.route('/paleta-etykieta/<int:paleta_id>')
def paleta_etykieta(paleta_id):
    """Printable label page for a pallet — contents, QR code, totals."""
    conn = get_db()

    # Fetch pallet info
    paleta = conn.execute(
        'SELECT id, nazwa, dostawca, data_zakupu, cena_zakupu FROM palety WHERE id = ?',
        (paleta_id,)
    ).fetchone()
    if not paleta:
        return redirect('/magazyn/palety')

    p_nazwa = paleta['nazwa'] or ''
    p_dostawca = paleta['dostawca'] or ''
    try:
        p_data = paleta['data_zakupu'] or ''
    except Exception:
        p_data = ''
    try:
        p_cena = paleta['cena_zakupu'] or 0
    except Exception:
        p_cena = 0

    # Fetch products on this pallet
    products = conn.execute(
        'SELECT nazwa, ilosc, ean, stan_przyjecia, klasa_jakosci FROM produkty WHERE paleta_id = ?',
        (paleta_id,)
    ).fetchall()

    total_types = len(products)
    total_pieces = sum((p['ilosc'] or 0) for p in products)

    # Build product rows
    prod_rows = ''
    for idx, pr in enumerate(products, 1):
        name = (pr['nazwa'] or 'Brak nazwy')[:30]
        qty = pr['ilosc'] or 0
        ean = pr['ean'] or ''
        stan = _format_stan_label(pr['stan_przyjecia'] if 'stan_przyjecia' in pr.keys() else '', pr['klasa_jakosci'] if 'klasa_jakosci' in pr.keys() else '')
        prod_rows += f'''<tr>
            <td style="padding:4px 8px;border-bottom:1px solid #ccc;text-align:center">{idx}</td>
            <td style="padding:4px 8px;border-bottom:1px solid #ccc">{name}</td>
            <td style="padding:4px 8px;border-bottom:1px solid #ccc;text-align:center">{qty}</td>
            <td style="padding:4px 8px;border-bottom:1px solid #ccc;font-size:0.75rem">{ean}</td>
            <td style="padding:4px 8px;border-bottom:1px solid #ccc;font-size:0.75rem">{stan}</td>
        </tr>'''

    from .database import get_config as _gc
    ngrok_domain = _gc('ngrok_domain', '')
    base_url = ngrok_domain and f"https://{ngrok_domain}" or request.host_url.rstrip('/')
    qr_url = f"{base_url}/magazyn/paleta-id/{paleta_id}"

    html = f'''<!DOCTYPE html><html lang="pl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Etykieta palety — {p_nazwa}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#fff;color:#000;padding:20px}}
@media screen {{
    body{{background:#0f172a;color:#e2e8f0}}
    .print-page{{background:#fff;color:#000;border-radius:12px;padding:30px;max-width:800px;margin:0 auto}}
}}
@media print {{
    .no-print{{display:none !important}}
    body{{padding:10px}}
    .print-page{{padding:10px}}
}}
.no-print{{text-align:center;margin-bottom:16px}}
.no-print button{{padding:10px 28px;background:#7c3aed;color:#fff;border:none;border-radius:10px;font-size:1rem;font-weight:700;cursor:pointer;margin:4px}}
.no-print a{{color:#94a3b8;text-decoration:none;margin:0 10px;font-size:0.9rem}}
h1{{font-size:2rem;font-weight:900;margin-bottom:4px}}
.meta{{font-size:0.95rem;color:#555;margin-bottom:16px}}
.qr-box{{text-align:center;margin:16px 0}}
.qr-box svg{{width:180px;height:180px}}
table{{width:100%;border-collapse:collapse;font-size:0.85rem;margin-top:12px}}
th{{background:#f0f0f0;padding:6px 8px;text-align:left;border-bottom:2px solid #333;font-weight:700}}
.totals{{margin-top:14px;font-size:1rem;font-weight:700;display:flex;gap:20px;flex-wrap:wrap}}
.totals span{{background:#f0f0f0;padding:6px 14px;border-radius:8px}}
</style>
</head><body>

<div class="no-print">
    <button onclick="window.print()">Drukuj</button>
    <a href="/magazyn/paleta-id/{paleta_id}">Wróć do palety</a>
</div>

<div class="print-page">
    <h1>#{paleta_id} — {p_nazwa}</h1>
    <div class="meta">
        Dostawca: <b>{p_dostawca}</b> &nbsp;|&nbsp;
        Data zakupu: <b>{p_data}</b> &nbsp;|&nbsp;
        Cena zakupu: <b>{p_cena:.2f} zł</b>
    </div>

    <div class="qr-box" id="qrBox" data-url="{qr_url}"></div>

    <table>
        <tr>
            <th>#</th><th>Produkt</th><th>Szt.</th><th>EAN</th><th>Stan</th>
        </tr>
        {prod_rows}
    </table>

    <div class="totals">
        <span>Produktów: {total_types}</span>
        <span>Łącznie sztuk: {total_pieces}</span>
    </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/qrcode-generator@1.4.4/qrcode.min.js" integrity="sha384-lQXOAyZwHXE55JFyrOMB7nY2Wv+m5ZWNtJcHrd1rceRQXAYNLak8ukN5TjBTcIwz" crossorigin="anonymous"></script>
<script>
(function(){{
    var box = document.getElementById('qrBox');
    var url = box.dataset.url;
    if(!url) return;
    var qr = qrcode(0, 'M');
    qr.addData(url);
    qr.make();
    box.innerHTML = qr.createSvgTag(6, 0);
    var svg = box.querySelector('svg');
    if(svg){{ svg.style.width='180px'; svg.style.height='180px'; }}
}})();
</script>

</body></html>'''

    return html


@magazynier_bp.route('/paleta/<path:n>')
def paleta_detail(n):
    from urllib.parse import unquote, quote
    n = unquote(n)
    conn = get_db()
    if n == 'brak':
        products = conn.execute('SELECT * FROM produkty WHERE paleta="" OR paleta IS NULL').fetchall()
        nazwa_palety = 'Bez palety'
        paleta_id = None
        cena_zakupu = 0
        cena_zakupu_netto = 0
    else:
        products = conn.execute('SELECT * FROM produkty WHERE paleta=?', (n,)).fetchall()
        nazwa_palety = n
        # Znajdź paleta_id z pierwszego produktu (jeśli istnieje)
        paleta_id = products[0]['paleta_id'] if products and products[0]['paleta_id'] else None
        # Pobierz cenę zakupu z tabeli palety
        cena_zakupu = 0
        cena_zakupu_netto = 0
        if paleta_id:
            try:
                pr = conn.execute('SELECT cena_zakupu, cena_zakupu_netto FROM palety WHERE id = ?', (paleta_id,)).fetchone()
                cena_zakupu = (pr['cena_zakupu'] or 0) if pr else 0
                cena_zakupu_netto = (pr['cena_zakupu_netto'] or 0) if pr else 0
                # cena_zakupu w bazie = BRUTTO
                if cena_zakupu > 0 and cena_zakupu_netto == 0:
                    cena_zakupu_netto = round(cena_zakupu / 1.23, 2)
            except:
                try:
                    pr = conn.execute('SELECT cena_zakupu FROM palety WHERE id = ?', (paleta_id,)).fetchone()
                    cena_zakupu = (pr['cena_zakupu'] or 0) if pr else 0
                    cena_zakupu_netto = round(cena_zakupu / 1.23, 2)
                except:
                    pass
    
    # Statystyki palety
    stats = conn.execute('''
        SELECT COUNT(*) as cnt, SUM(ilosc) as items, 
               SUM(cena_allegro * ilosc) as allegro_all,
               SUM(cena_brutto) as suma_produktow_brutto
        FROM produkty WHERE paleta=?
    ''', (n if n != 'brak' else '',)).fetchone()
    
    # cena_zakupu w bazie = BRUTTO
    brutto = cena_zakupu
    netto = cena_zakupu_netto if cena_zakupu_netto > 0 else round(cena_zakupu / 1.23, 2)
    
    # Fallback: jeśli brak ceny zakupu w palecie, użyj sumy produktów
    if brutto == 0 and (stats['suma_produktow_brutto'] or 0) > 0:
        brutto = stats['suma_produktow_brutto']
        netto = round(brutto / 1.23, 2)
    
    allegro = stats['allegro_all'] or 0
    zysk = allegro - brutto - (allegro * 0.11)
    
    paleta_encoded = quote(nazwa_palety, safe='')
    
    html = f'''<div class="hdr"><h1><span class=material-symbols-outlined>inventory_2</span> {nazwa_palety}</h1><small>{len(products)} produktów</small></div>
    
    <div class="stats" style="margin-bottom:15px">
        <div class="stat">
            <div class="stat-v">{stats['cnt'] or 0}</div>
            <div class="stat-l">Produktów</div>
        </div>
        <div class="stat">
            <div class="stat-v">{netto:.0f} zł</div>
            <div class="stat-l">Zakup netto</div>
        </div>
        <div class="stat">
            <div class="stat-v">{brutto:.0f} zł</div>
            <div class="stat-l">Zakup brutto</div>
        </div>
        <div class="stat">
            <div class="stat-v green">{allegro:.0f} zł</div>
            <div class="stat-l">Allegro (suma)</div>
        </div>
        <div class="stat">
            <div class="stat-l">Zysk</div>
        </div>
    </div>
    '''
    
    # Przyciski akcji na palecie
    if paleta_id or (n and n != 'brak'):
        html += '<div style="display:flex;gap:10px;margin-bottom:15px;flex-wrap:wrap">'
        
        # Masowe wystawianie (stary system - bezpośrednio)
        if paleta_id:
            html += f'<a href="/palety/{paleta_id}/mass-edit" class="btn" style="background:var(--purple);flex:1"><span class=material-symbols-outlined>shopping_cart</span> Wystaw bezpośrednio</a>'
        
        # NOWE: Przenieś do Paletomat (ze scrapowaniem!)
        html += f'<a href="/magazyn/paleta/{paleta_encoded}/to-paletomat" class="btn btn-ok" style="flex:1"><span class=material-symbols-outlined>sync</span> PALETOMAT (scrapuj)</a>'
        
        html += '</div>'
    
    for p in products:
        img = _resolve_product_image(p, size='sm')
        pcode = get_product_code(p)
        _ean_c = p['ean'] if p['ean'] and p['ean'].upper() not in ('N/A','NAN','NONE') else ''
        display_code = _ean_c or p['asin'] or f"#{p['id']}"

        # Cena zakupu produktu ZA SZTUKĘ (cena_netto/cena_brutto w bazie JUŻ są jednostkowe!)
        cena_za_sztuke = p['cena_netto'] if p['cena_netto'] and p['cena_netto'] > 0 else (p['cena_brutto'] or 0)
        
        html += f'''<a href="/magazyn/produkt/{pcode}" class="item">
            <img src="{img}" onerror="this.src='data:image/svg+xml,%3Csvg xmlns=%27http://www.w3.org/2000/svg%27 width=%2745%27 height=%2745%27%3E%3Crect fill=%27%2312121a%27 width=%2745%27 height=%2745%27/%3E%3Ctext x=%2722%27 y=%2728%27 fill=%27%23555%27 text-anchor=%27middle%27 font-size=%2716%27%3E%F0%9F%93%A6%3C/text%3E%3C/svg%3E'">
            <div class="item-info">
                <div class="item-name">{p['nazwa'][:30]}...</div>
                <div class="item-meta">{display_code} • Zakup: {cena_za_sztuke:.2f} zł/szt</div>
            </div>
            <div class="item-qty">{p['ilosc']}</div>
        </a>'''
    
    # Przycisk usuwania palety (tylko jeśli nie jest to "Bez palety")
    if n != 'brak' and nazwa_palety:
        html += f'''
        <div style="margin-top:20px;padding:15px;background:#12121a;border-radius:12px">
            <form action="/magazyn/paleta/{paleta_encoded}/usun" method="POST" onsubmit="return confirm('Na pewno usunąć paletę {nazwa_palety} i wszystkie jej produkty?')">
                <button type="submit" class="btn btn-err" style="width:100%"><span class=material-symbols-outlined>delete</span> USUŃ PALETĘ + PRODUKTY</button>
            </form>
            <form action="/magazyn/paleta/{paleta_encoded}/wyczysc" method="POST" onsubmit="return confirm('Usunąć tylko przypisanie do palety (produkty zostaną)?')" style="margin-top:10px">
                <button type="submit" class="btn btn-warn" style="width:100%"><span class=material-symbols-outlined>upload</span> WYCZYŚĆ PRZYPISANIE</button>
            </form>
        </div>
        '''
    
    html += '<a href="/magazyn/palety" class="back">← Powrót</a>'
    return render(html)


@magazynier_bp.route('/paleta/<path:n>/to-paletomat')
def paleta_to_paletomat(n):
    """Przenosi wszystkie produkty z palety do Paletomat (tabela scraped) ze scrapowaniem"""
    from urllib.parse import unquote
    n = unquote(n)
    
    conn = get_db()
    
    if n == 'brak':
        products = conn.execute('SELECT * FROM produkty WHERE paleta="" OR paleta IS NULL').fetchall()
    else:
        products = conn.execute('SELECT * FROM produkty WHERE paleta=?', (n,)).fetchall()
    
    added_count = 0
    updated_count = 0
    skipped_count = 0
    
    for p in products:
        p = dict(p)
        
        # Ustal identyfikator (ASIN > EAN > MAG{id})
        asin = p.get('asin', '') or ''
        ean = p.get('ean', '') or ''
        
        if asin and asin not in ('N/A', 'None', ''):
            identyfikator = asin
        elif ean and ean not in ('N/A', 'None', ''):
            identyfikator = ean
        else:
            identyfikator = f"MAG{p['id']}"
        
        # Sprawdź czy już istnieje w scraped
        existing = conn.execute('SELECT asin FROM scraped WHERE asin=?', (identyfikator,)).fetchone()
        
        if not existing:
            # Dodaj do tabeli scraped
            conn.execute('''
                INSERT INTO scraped (asin, nazwa, zdjecie_url, cena_amazon, status, data_scrape, ean)
                VALUES (?, ?, ?, ?, 'nowy', datetime('now'), ?)
            ''', (
                identyfikator,
                p.get('nazwa', f'Produkt {identyfikator}'),
                p.get('zdjecie_url', ''),
                p.get('cena_brutto', 0) or 0,
                ean
            ))
            added_count += 1
        else:
            # Aktualizuj istniejący
            conn.execute('''
                UPDATE scraped SET nazwa=?, zdjecie_url=?, status='nowy'
                WHERE asin=?
            ''', (
                p.get('nazwa', f'Produkt {identyfikator}'),
                p.get('zdjecie_url', ''),
                identyfikator
            ))
            updated_count += 1
    
    conn.commit()
    
    # Pokaż komunikat i przekieruj
    from .magazynier import render
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> PRZENIESIONO DO PALETOMAT</h1></div>
    
    <div class="alert alert-ok">
        <span class=material-symbols-outlined>inventory_2</span> Paleta: {n}<br>
        <span class=material-symbols-outlined>check_circle</span> Dodano: {added_count} produktów<br>
        <span class=material-symbols-outlined>sync</span> Zaktualizowano: {updated_count} produktów<br>
        ⏭ Pominięto: {skipped_count} produktów
    </div>
    
    <div class="card" style="padding:15px;margin-top:15px">
        <div style="font-weight:600;margin-bottom:12px"><span class=material-symbols-outlined>target</span> CO DALEJ?</div>
        <div style="color:#64748b;font-size:0.85rem;margin-bottom:15px">
            1. Przejdź do Paletomat → Generator<br>
            2. Zobaczysz swoje produkty<br>
            3. Kliknij na produkt → Zescrapuje z Amazona (zdjęcia, opisy)<br>
            4. Wystaw masowo z AI opisami!
        </div>
        <a href="/paletomat/generator" class="btn btn-ok" style="width:100%">
            <span class='material-symbols-outlined' style='font-size:1rem;vertical-align:middle'>rocket_launch</span> OTWÓRZ PALETOMAT GENERATOR
        </a>
    </div>
    
    <a href="/magazyn/paleta/{n}" class="btn btn-2" style="margin-top:15px">← Powrót do palety</a>
    <a href="/magazyn" class="back">← Magazyn</a>
    '''
    return render(html)


@magazynier_bp.route('/paleta/<path:n>/usun', methods=['POST'])
def paleta_usun(n):
    """Usuwa paletę wraz z produktami (także ze scraped/Paletomat)"""
    from urllib.parse import unquote
    n = unquote(n)
    
    conn = get_db()
    if n and n != 'brak':
        # [LOCA] NOWE: Pobierz ASINy produktów z tej palety
        asiny = conn.execute('SELECT DISTINCT asin FROM produkty WHERE paleta = ? AND asin != ""', (n,)).fetchall()
        asiny_list = [row[0] for row in asiny if row[0]]
        
        # [LOCA] NOWE: Usuń te produkty ze scraped (Paletomat)
        if asiny_list:
            placeholders = ','.join(['?' for _ in asiny_list])
            conn.execute('DELETE FROM scraped WHERE asin IN (' + placeholders + ')', asiny_list)  # noqa: B608 - placeholders are only ?
            print(f"[DELETE] Usunięto {len(asiny_list)} produktów ze scraped (Paletomat)")
        
        # Usuń wszystkie produkty z tej palety (Magazynier)
        conn.execute('DELETE FROM produkty WHERE paleta = ?', (n,))
        
        conn.commit()
        print(f"[CHECK_CIRCLE] Usunięto paletę: {n}")
    
    return redirect('/magazyn/palety')


@magazynier_bp.route('/paleta/<path:n>/wyczysc', methods=['POST'])
def paleta_wyczysc(n):
    """Usuwa przypisanie do palety (produkty zostają)"""
    from urllib.parse import unquote
    n = unquote(n)
    
    conn = get_db()
    if n and n != 'brak':
        # Wyczyść przypisanie palety
        conn.execute('UPDATE produkty SET paleta = "", paleta_id = NULL WHERE paleta = ?', (n,))
        conn.commit()
    
    return redirect('/magazyn/palety')


# ============================================================
# POBIERANIE ZDJĘĆ Z AMAZON
# ============================================================

@magazynier_bp.route('/fetch-images')
def fetch_images_page():
    """Strona do pobierania zdjęć z Amazon dla produktów z ASIN"""
    conn = get_db()
    
    # Policz produkty bez zdjęć z ASIN
    stats = conn.execute('''
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN asin != '' AND asin IS NOT NULL THEN 1 ELSE 0 END) as with_asin,
            SUM(CASE WHEN (zdjecie_url IS NULL OR zdjecie_url = '') AND asin != '' AND asin IS NOT NULL THEN 1 ELSE 0 END) as no_image_with_asin
        FROM produkty
    ''').fetchone()
    
    total = stats['total'] or 0
    with_asin = stats['with_asin'] or 0
    no_image = stats['no_image_with_asin'] or 0
    
    html = f'''
    <div class="hdr">
        <h1><span class=material-symbols-outlined>photo_camera</span> POBIERZ ZDJĘCIA</h1>
        <small>Automatyczne pobieranie z Amazon</small>
    </div>
    
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;text-align:center">
            <div>
                <div style="font-size:1.8rem;font-weight:700;color:#8ff5ff">{total}</div>
                <div style="font-size:0.75rem;color:#64748b">Produktów</div>
            </div>
            <div>
                <div style="font-size:1.8rem;font-weight:700;color:#beee00">{with_asin}</div>
                <div style="font-size:0.75rem;color:#64748b">Z ASIN</div>
            </div>
            <div>
                <div style="font-size:1.8rem;font-weight:700;color:#f59e0b">{no_image}</div>
                <div style="font-size:0.75rem;color:#64748b">Bez zdjęcia</div>
            </div>
        </div>
    </div>
    
    <div class="card" style="padding:15px;margin-bottom:15px">
        <div style="font-weight:600;margin-bottom:10px"><span class=material-symbols-outlined>warning</span> Uwaga</div>
        <div style="font-size:0.85rem;color:#94a3b8">
            Pobieranie zdjęć wymaga scrapowania Amazona.<br>
            • Każdy produkt = ~3-5 sekund<br>
            • {no_image} produktów = ~{no_image * 4 // 60} minut<br>
            • Może być blokowane przez Amazon (captcha)
        </div>
    </div>
    
    <div id="progress" style="display:none;margin-bottom:15px">
        <div class="card" style="padding:15px">
            <div style="font-weight:600;margin-bottom:10px">⏳ Pobieranie w toku...</div>
            <div id="progress-text" style="font-size:0.85rem;color:#94a3b8">0 / {no_image}</div>
            <div style="background:#1e1e2e;border-radius:6px;height:10px;margin-top:10px;overflow:hidden">
                <div id="progress-bar" style="background:#beee00;width:0%;height:100%;transition:width 0.3s"></div>
            </div>
            <div id="progress-log" style="margin-top:10px;font-size:0.75rem;color:#64748b;max-height:150px;overflow-y:auto"></div>
        </div>
    </div>
    
    <button onclick="startFetch()" id="start-btn" class="btn btn-ok" style="width:100%;padding:14px;font-size:1rem">
        <span class=material-symbols-outlined>photo_camera</span> POBIERZ ZDJĘCIA ({no_image} produktów)
    </button>
    
    <a href="/magazyn" class="back">← Powrót</a>
    
    <script>
    let running = false;
    
    async function startFetch() {{
        if (running) return;
        running = true;
        
        document.getElementById('start-btn').style.display = 'none';
        document.getElementById('progress').style.display = 'block';
        
        const response = await fetch('/magazyn/api/fetch-images-start');
        const data = await response.json();
        
        if (data.success) {{
            pollProgress();
        }} else {{
            alert('Błąd: ' + data.error);
            running = false;
        }}
    }}
    
    async function pollProgress() {{
        while (running) {{
            const response = await fetch('/magazyn/api/fetch-images-status');
            const data = await response.json();
            
            document.getElementById('progress-text').textContent = data.done + ' / ' + data.total;
            document.getElementById('progress-bar').style.width = (data.done / data.total * 100) + '%';
            
            if (data.log) {{
                document.getElementById('progress-log').innerHTML = data.log.slice(-10).map(l => '<div>' + l + '</div>').join('');
                document.getElementById('progress-log').scrollTop = 9999;
            }}
            
            if (!data.running) {{
                running = false;
                alert('<span class=material-symbols-outlined>check_circle</span> Zakończono! Pobrano ' + data.done + ' zdjęć.');
                location.reload();
                break;
            }}
            
            await new Promise(r => setTimeout(r, 1000));
        }}
    }}
    </script>
    '''
    return render(html)


# Globalny stan pobierania zdjęć
_fetch_images_state = {
    'running': False,
    'total': 0,
    'done': 0,
    'log': []
}

@magazynier_bp.route('/api/fetch-images-start')
def api_fetch_images_start():
    """Rozpocznij pobieranie zdjęć w tle"""
    global _fetch_images_state
    
    if _fetch_images_state['running']:
        return jsonify({'success': False, 'error': 'Już działa'})
    
    import threading
    
    def fetch_worker():
        global _fetch_images_state
        from .utils import scrape_amazon_product
        
        _fetch_images_state['running'] = True
        _fetch_images_state['done'] = 0
        _fetch_images_state['log'] = []
        
        try:
            conn = get_db()
            products = conn.execute('''
                SELECT id, asin FROM produkty 
                WHERE asin != '' AND asin IS NOT NULL 
                AND (zdjecie_url IS NULL OR zdjecie_url = '')
                LIMIT 100
            ''').fetchall()
            
            _fetch_images_state['total'] = len(products)
            
            for p in products:
                if not _fetch_images_state['running']:
                    break
                
                asin = p['asin']
                try:
                    result = scrape_amazon_product(asin)
                    if result and result.get('image_url'):
                        conn = get_db()
                        conn.execute('UPDATE produkty SET zdjecie_url = ? WHERE id = ?', 
                            (result['image_url'], p['id']))
                        conn.commit()
                        _fetch_images_state['log'].append(f'<span class=material-symbols-outlined>check_circle</span> {asin}: OK')
                    else:
                        _fetch_images_state['log'].append(f'<span class=material-symbols-outlined>warning</span> {asin}: brak zdjęcia')
                except Exception as e:
                    _fetch_images_state['log'].append(f'<span class=material-symbols-outlined>cancel</span> {asin}: {str(e)[:30]}')
                
                _fetch_images_state['done'] += 1
                
        except Exception as e:
            _fetch_images_state['log'].append(f'<span class=material-symbols-outlined>cancel</span> Błąd: {str(e)}')
        finally:
            _fetch_images_state['running'] = False
    
    thread = threading.Thread(target=fetch_worker, daemon=True)
    thread.start()
    
    return jsonify({'success': True})


@magazynier_bp.route('/api/fetch-images-status')
def api_fetch_images_status():
    """Status pobierania zdjęć"""
    return jsonify(_fetch_images_state)


@magazynier_bp.route('/api/fetch-images-stop')
def api_fetch_images_stop():
    """Zatrzymaj pobieranie"""
    global _fetch_images_state
    _fetch_images_state['running'] = False
    return jsonify({'success': True})


@magazynier_bp.route('/dostawcy')
def dostawcy():
    conn = get_db()
    result = conn.execute('''SELECT dostawca, COUNT(*) as cnt, SUM(ilosc) as items 
        FROM produkty GROUP BY dostawca ORDER BY dostawca''').fetchall()
    
    html = '<div class="hdr"><h1><span class=material-symbols-outlined>local_shipping</span> DOSTAWCY</h1></div>'
    
    for d in result:
        html += f'''<a href="/magazyn/dostawca/{d['dostawca'] or 'brak'}" class="item">
            <div style="font-size:1.5rem;margin-right:10px"><span class=material-symbols-outlined>local_shipping</span></div>
            <div class="item-info">
                <div class="item-name dostawca-name">{d['dostawca'] or 'Nieznany'}</div>
                <div class="item-meta">{d['cnt']} produktów</div>
            </div>
            <div class="item-qty">{d['items'] or 0}</div>
        </a>'''
    
    html += '<a href="/magazyn" class="back">← Powrót</a>'
    return render(html)

@magazynier_bp.route('/dostawca/<n>')
def dostawca_detail(n):
    conn = get_db()
    if n == 'brak':
        products = conn.execute('SELECT * FROM produkty WHERE dostawca="" OR dostawca IS NULL').fetchall()
        nazwa_dostawcy = 'Nieznany dostawca'
    else:
        products = conn.execute('SELECT * FROM produkty WHERE dostawca=?', (n,)).fetchall()
        nazwa_dostawcy = n
    
    html = f'''<div class="hdr"><h1><span class=material-symbols-outlined>local_shipping</span> {nazwa_dostawcy}</h1><small>{len(products)} produktów</small></div>'''
    
    for p in products:
        img = _resolve_product_image(p, size='sm')
        pcode = get_product_code(p)
        _ean_c = p['ean'] if p['ean'] and p['ean'].upper() not in ('N/A','NAN','NONE') else ''
        display_code = _ean_c or p['asin'] or f"#{p['id']}"
        html += f'''<a href="/magazyn/produkt/{pcode}" class="item">
            <img src="{img}" onerror="this.src='{_PLACEHOLDER_IMG_SM}'">
            <div class="item-info">
                <div class="item-name">{p['nazwa'][:30]}...</div>
                <div class="item-meta">{display_code} | <span class=material-symbols-outlined>inventory_2</span> {p['paleta'] or '—'}</div>
            </div>
            <div class="item-qty">{p['ilosc']}</div>
        </a>'''
    
    html += '<a href="/magazyn/dostawcy" class="back">← Powrót</a>'
    return render(html)

@magazynier_bp.route('/export')
def export_csv():
    """Export magazynu do Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    conn = get_db()
    products = conn.execute('SELECT * FROM produkty ORDER BY paleta, lokalizacja').fetchall()
    
    # Utwórz workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Magazyn"
    
    # Style
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Nagłówki
    headers = ['EAN', 'ASIN', 'Nazwa', 'Ilość', 'Cena brutto', 'Cena Allegro', 
               'Lokalizacja', 'Paleta', 'Dostawca', 'Stan', 'Status', 'Kategoria', 'Data dodania']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dane
    for row_num, p in enumerate(products, 2):
        data = [
            p['ean'] or '',
            p['asin'] or '',
            p['nazwa'] or '',
            p['ilosc'] or 0,
            p['cena_brutto'] or 0,
            p['cena_allegro'] or 0,
            p['lokalizacja'] or '',
            p['paleta'] or '',
            p['dostawca'] or '',
            p['stan'] or '',
            p['status'] if p['status'] else 'nowy',
            p['kategoria'] or '',
            p['data_dodania'] or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Wyrównanie
            if col_num in [4, 5, 6]:  # Liczby
                cell.alignment = Alignment(horizontal='right')
            elif col_num in [1, 2, 7, 8]:  # Kody i lokalizacje
                cell.alignment = Alignment(horizontal='center')
    
    # Autofit kolumn
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Zamroź pierwszy wiersz
    ws.freeze_panes = 'A2'
    
    # Dodaj arkusz ze statystykami
    stats_ws = wb.create_sheet("Statystyki")
    stats = get_stats()
    
    stats_data = [
        ['Statystyka', 'Wartość'],
        ['Produktów', stats['produkty']],
        ['Sztuk', stats['sztuki']],
        ['Wartość zakupu (brutto)', f"{stats['wartosc_zakupu']:.2f} zł"],
        ['Wartość zakupu (netto)', f"{stats['wartosc_netto']:.2f} zł"],
        ['Wartość Allegro', f"{stats['wartosc_allegro']:.2f} zł"],
        ['Palet', stats['palety']],
        ['Dostawców', stats['dostawcy']],
        ['Data exportu', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_num, (label, value) in enumerate(stats_data, 1):
        cell_a = stats_ws.cell(row=row_num, column=1, value=label)
        cell_b = stats_ws.cell(row=row_num, column=2, value=value)
        
        if row_num == 1:
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_b.fill = header_fill
            cell_b.font = header_font
        else:
            cell_a.font = Font(bold=True)
        
        cell_a.border = border
        cell_b.border = border
    
    stats_ws.column_dimensions['A'].width = 30
    stats_ws.column_dimensions['B'].width = 20
    
    # Zapisz do BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'magazyn_{datetime.now():%Y%m%d_%H%M}.xlsx'
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def _mag_dostawca_options(selected=''):
    """Generuje opcje <option> dostawcy dla formularzy magazyniera"""
    from modules.database import get_dostawcy_list
    _dlist = get_dostawcy_list()
    opts = ''
    for d in _dlist:
        sel = ' selected' if d == selected else ''
        opts += f'<option value="{d}"{sel}>{d}</option>'
    opts += '<option value="__custom__">+ Dodaj nowego...</option>'
    return opts

@magazynier_bp.route('/import')
def import_page():
    # Pobierz istniejące palety
    conn = get_db()
    palety = conn.execute('SELECT id, nazwa, dostawca, data_zakupu FROM palety ORDER BY data_dodania DESC').fetchall()
    
    # Generuj opcje palet
    palety_options = '<option value="">-- Bez palety (luźne produkty) --</option>'
    palety_options += '<option value="__NEW__">➕ Utwórz nową paletę...</option>'
    for p in palety:
        p_id, p_nazwa, p_dostawca, p_data = p
        label = f"{p_nazwa}" if p_nazwa else f"Paleta #{p_id}"
        if p_dostawca:
            label += f" ({p_dostawca})"
        if p_data:
            label += f" - {p_data}"
        palety_options += f'<option value="{p_id}">{label}</option>'
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>download</span> IMPORT</h1></div>
    
    <form action="/magazyn/import/preview" method="POST" enctype="multipart/form-data" id="importForm">
        
        <!-- WYBÓR PALETY -->
        <div class="card" style="padding:15px;margin-bottom:15px">
            <div style="font-weight:600;margin-bottom:10px;color:#f59e0b"><span class=material-symbols-outlined>inventory_2</span> Przypisz do palety:</div>
            <select name="paleta_id" id="paletaSelect" class="form-ctrl" style="width:100%;padding:12px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff;font-size:1rem" onchange="toggleNewPaleta()">
                {palety_options}
            </select>
            
            <!-- NOWA PALETA -->
            <div id="newPaletaFields" style="display:none;margin-top:15px;padding:15px;background:#0a0a0f;border-radius:8px;border:1px solid #f59e0b">
                <div style="font-weight:600;margin-bottom:10px;color:#f59e0b"><span class=material-symbols-outlined>auto_awesome</span> Nowa paleta:</div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                    <div>
                        <label style="font-size:0.8rem;color:#64748b">Nazwa palety</label>
                        <input type="text" name="new_paleta_nazwa" class="form-ctrl" placeholder="np. Paleta Warrington #15" style="width:100%;padding:10px;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:6px;color:#fff">
                    </div>
                    <div>
                        <label style="font-size:0.8rem;color:#64748b">Dostawca</label>
                        <select name="new_paleta_dostawca" class="form-ctrl" style="width:100%;padding:10px;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:6px;color:#fff" onchange="if(this.value==='__custom__'){{this.nextElementSibling.style.display='block';this.nextElementSibling.focus()}}else{{this.nextElementSibling.style.display='none'}}">
                            {_mag_dostawca_options()}
                        </select>
                        <input type="text" name="new_paleta_dostawca_custom" placeholder="Wpisz nazwe dostawcy" style="display:none;margin-top:8px;width:100%;padding:10px;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:6px;color:#fff">
                    </div>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:10px">
                    <div>
                        <label style="font-size:0.8rem;color:#64748b"><span class=material-symbols-outlined>paid</span> Cena zakupu (aukcja/faktura)</label>
                        <input type="number" name="new_paleta_cena" class="form-ctrl" placeholder="np. 144.80" step="0.01" style="width:100%;padding:10px;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:6px;color:#fff">
                        <div style="font-size:0.7rem;color:#f59e0b;margin-top:4px"><span class=material-symbols-outlined>warning</span> NIE cenę produktów z Excela</div>
                    </div>
                    <div>
                        <label style="font-size:0.8rem;color:#64748b">Typ</label>
                        <select name="new_paleta_typ" class="form-ctrl" style="width:100%;padding:10px;background:#1e1e2e;border:1px solid #2a2a3a;border-radius:6px;color:#fff">
                            <option value="paleta">📦 Paleta</option>
                            <option value="box">inbox Box</option>
                        </select>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- WYBÓR PLIKU -->
        <div class="card" style="padding:30px;text-align:center;cursor:pointer" onclick="document.getElementById('file').click()">
            <div style="font-size:3rem;margin-bottom:10px"><span class=material-symbols-outlined>folder</span></div>
            <div style="font-weight:600">Wybierz plik Excel</div>
            <div style="font-size:0.8rem;color:#64748b;margin-top:5px">.xlsx, .csv</div>
            <input type="file" id="file" name="file" style="display:none" accept=".xlsx,.csv" onchange="this.form.submit()">
        </div>
    </form>
    
    <div class="card" style="padding:15px;margin-top:15px">
        <div style="font-weight:600;color:#eab308;margin-bottom:10px"><span class=material-symbols-outlined>lightbulb</span> Obsługiwane formaty:</div>
        <div style="font-size:0.85rem;color:#94a3b8">
            • <strong>Warrington</strong> - auto-detekcja kolumn<br>
            • <strong>Miglo</strong> - ASIN, nazwa, ilość<br>
            • <strong>Jobalots</strong> - manifest produktów<br>
            • <strong>Własny format</strong> - wybierz kolumny ręcznie
        </div>
    </div>
    
    <a href="/magazyn" class="back">← Powrót</a>
    
    <script>
    function toggleNewPaleta() {{
        var sel = document.getElementById('paletaSelect');
        var fields = document.getElementById('newPaletaFields');
        fields.style.display = (sel.value === '__NEW__') ? 'block' : 'none';
    }}
    </script>
    '''
    return render(html)

@magazynier_bp.route('/import/preview', methods=['POST'])
def import_preview():
    """Podgląd pliku przed importem"""
    if 'file' not in request.files:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    file = request.files['file']
    if file.filename == '':
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    # === OBSŁUGA PALETY ===
    paleta_id = request.form.get('paleta_id', '')
    paleta_nazwa = ''
    paleta_dostawca = ''  # NOWE: przechowujemy dostawcę
    
    if paleta_id == '__NEW__':
        # Utwórz nową paletę
        new_nazwa = request.form.get('new_paleta_nazwa', '').strip()
        new_dostawca = request.form.get('new_paleta_dostawca', '').strip()
        if new_dostawca == '__custom__':
            new_dostawca = request.form.get('new_paleta_dostawca_custom', '').strip()
            if new_dostawca:
                from modules.database import save_custom_dostawca
                save_custom_dostawca(new_dostawca)
        new_cena = request.form.get('new_paleta_cena', '0').strip()
        
        try:
            new_cena = float(new_cena) if new_cena else 0
        except:
            new_cena = 0
        
        if not new_nazwa:
            new_nazwa = f"Import {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Jeśli nie wybrano dostawcy, domyślnie Jobalots
        if not new_dostawca:
            new_dostawca = 'Jobalots'
        
        conn = get_db()
        try:
            new_cena_f = float(new_cena) if new_cena else 0
        except:
            new_cena_f = 0
        # cena_zakupu = brutto
        new_typ = request.form.get('new_paleta_typ', 'paleta').strip()
        cursor = conn.execute('''INSERT INTO palety (nazwa, dostawca, cena_zakupu, cena_zakupu_netto, data_zakupu, typ)
            VALUES (?, ?, ?, ?, date('now'), ?)''', (new_nazwa, new_dostawca, new_cena_f, round(new_cena_f / 1.23, 2), new_typ or 'paleta'))
        paleta_id = str(cursor.lastrowid)
        paleta_nazwa = new_nazwa
        paleta_dostawca = new_dostawca  # NOWE: zapisz dostawcę
        conn.commit()
    elif paleta_id:
        # Pobierz nazwę i dostawcę istniejącej palety
        conn = get_db()
        row = conn.execute('SELECT nazwa, dostawca FROM palety WHERE id = ?', (paleta_id,)).fetchone()
        if row:
            paleta_nazwa = row[0] or f"Paleta #{paleta_id}"
            paleta_dostawca = row[1] or ''  # NOWE: pobierz dostawcę
    
    filename = file.filename.lower()
    headers = []
    preview_rows = []
    total_rows = 0
    
    try:
        if filename.endswith('.xlsx'):
            import openpyxl
            import tempfile
            import os as os_module
            
            tmp_path = os_module.path.join(tempfile.gettempdir(), f'preview_{os_module.getpid()}.xlsx')
            file.save(tmp_path)
            
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True)
                ws = wb.active
                
                # Znajdź wiersz z nagłówkami
                for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                    row_str = [str(c or '').strip() for c in row]
                    if any(row_str):
                        headers = row_str
                        break
                
                # Pobierz podgląd (max 5 wierszy)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    total_rows += 1
                    if len(preview_rows) < 5:
                        preview_rows.append([str(c or '')[:50] for c in row])
                
                wb.close()
            finally:
                try:
                    os_module.remove(tmp_path)
                except:
                    pass
                    
        elif filename.endswith('.csv'):
            import csv
            import io
            
            raw_data = file.read()
            content = None
            for encoding in ['utf-8-sig', 'utf-8', 'cp1250', 'latin-1']:
                try:
                    content = raw_data.decode(encoding)
                    break
                except:
                    continue
            
            if content:
                delimiter = ';' if ';' in content[:500] else ','
                reader = csv.reader(io.StringIO(content), delimiter=delimiter)
                rows = list(reader)
                
                if rows:
                    headers = [str(h).strip() for h in rows[0]]
                    for row in rows[1:]:
                        total_rows += 1
                        if len(preview_rows) < 5:
                            preview_rows.append([str(c)[:50] for c in row])
        
        # Wykryj automatycznie kolumny (ulepszone wykrywanie EAN)
        auto_ean = auto_asin = auto_nazwa = auto_ilosc = auto_cena = -1
        detected_ean_col_name = ""
        detected_asin_col_name = ""
        
        for i, h in enumerate(headers):
            h_clean = h.lower().replace(' ', '').replace('_', '').replace('-', '')
            h_up = h.upper().replace(' ', '').replace('_', '')
            
            # === INTELIGENTNE WYKRYWANIE EAN ===
            # Szukaj: ean, barcode, kod_kreskowy, gtin, kodean
            if auto_ean == -1 and any(x in h_clean for x in ['ean', 'barcode', 'kodkreskowy', 'gtin', 'kodean']):
                auto_ean = i
                detected_ean_col_name = h
            
            # === OSOBNO WYKRYWANIE ASIN ===
            elif auto_asin == -1 and any(x in h_clean for x in ['asin', 'sku', 'code', 'kod']) and 'kreskowy' not in h_clean:
                auto_asin = i
                detected_asin_col_name = h
            
            # Nazwa produktu
            elif auto_nazwa == -1 and any(x in h_up for x in ['NAZWA', 'NAME', 'TYTUL', 'TITLE', 'PRODUCT', 'OPIS']):
                auto_nazwa = i
            
            # Ilość
            elif auto_ilosc == -1 and any(x in h_up for x in ['ILOSC', 'ILOŚĆ', 'QTY', 'QUANTITY', 'SZT', 'SZTUK']):
                auto_ilosc = i
            
            # Cena
            elif auto_cena == -1 and any(x in h_up for x in ['CENA', 'PRICE', 'KOSZT', 'COST']):
                auto_cena = i
        
        # UWAGA: NIE kopiuj ASIN do EAN! To różne pola.
        # ASIN (B0XXXXXXXX) powinien być w osobnej kolumnie
        
        # === AUTO-DETEKCJA DOSTAWCY Z NAGŁÓWKÓW ===
        if not paleta_dostawca and headers:
            detected_dostawca = detect_supplier(headers)
            if detected_dostawca:
                paleta_dostawca = detected_dostawca
                # Aktualizuj paletę jeśli została utworzona bez dostawcy
                if paleta_id:
                    try:
                        conn = get_db()
                        conn.execute('UPDATE palety SET dostawca = ? WHERE id = ? AND (dostawca IS NULL OR dostawca = "")', 
                            (detected_dostawca, paleta_id))
                        conn.commit()
                    except:
                        pass
        
        # Logi wykrywania (niebieskie)
        detection_logs = ""
        if paleta_dostawca:
            detection_logs += f'<div style="color:#beee00;padding:4px 0"><span class=material-symbols-outlined>check_circle</span> <span class=material-symbols-outlined>info</span> Dostawca: <strong>{paleta_dostawca}</strong></div>'
        if detected_ean_col_name:
            detection_logs += f'<div style="color:#8ff5ff;padding:4px 0"><span class=material-symbols-outlined style=color:#3b82f6>info</span> <span class=material-symbols-outlined>info</span> Wykryto kolumnę EAN: "{detected_ean_col_name}"</div>'
        if detected_asin_col_name and auto_asin != auto_ean:
            detection_logs += f'<div style="color:#8ff5ff;padding:4px 0"><span class=material-symbols-outlined style=color:#3b82f6>info</span> <span class=material-symbols-outlined>info</span> Wykryto kolumnę ASIN: "{detected_asin_col_name}"</div>'
        
        # Generuj opcje select
        def make_options(selected):
            opts = '<option value="-1">-- Brak --</option>'
            for i, h in enumerate(headers):
                sel = 'selected' if i == selected else ''
                opts += f'<option value="{i}" {sel}>{i+1}. {h[:30]}</option>'
            return opts
        
        # Tabela podglądu
        preview_table = '<div style="overflow-x:auto;margin:15px 0"><table style="width:100%;font-size:0.75rem;border-collapse:collapse">'
        preview_table += '<tr style="background:#1e1e2e">'
        for i, h in enumerate(headers):
            preview_table += f'<th style="padding:8px;border:1px solid #2a2a3a;white-space:nowrap">{i+1}. {h[:20]}</th>'
        preview_table += '</tr>'
        
        for row in preview_rows:
            preview_table += '<tr>'
            for c in row:
                preview_table += f'<td style="padding:6px;border:1px solid #1e1e2e;color:#94a3b8">{c[:30]}</td>'
            preview_table += '</tr>'
        preview_table += '</table></div>'
        
        # Info o palecie
        paleta_info = ''
        if paleta_id:
            paleta_info = f'<div class="alert" style="background:#f59e0b22;border:1px solid #f59e0b;color:#f59e0b;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>inventory_2</span> Produkty zostaną przypisane do: <strong>{paleta_nazwa}</strong></div>'
        else:
            paleta_info = '<div class="alert" style="background:#64748b22;border:1px solid #64748b;color:#94a3b8;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>warning</span> Produkty będą bez przypisanej palety (luźne)</div>'
        
        html = f'''
        <div class="hdr"><h1><span class=material-symbols-outlined>list_alt</span> PODGLĄD IMPORTU</h1><small>{total_rows} wierszy</small></div>
        
        {paleta_info}
        
        <div class="alert alert-ok" style="font-size:0.85rem">Znaleziono {len(headers)} kolumn, {total_rows} produktów</div>
        
        {f'<div class="card" style="padding:10px;font-family:monospace;font-size:0.8rem;background:#0a0a0f">{detection_logs}</div>' if detection_logs else ''}
        
        <div class="card" style="padding:15px">
            <div style="font-weight:600;margin-bottom:10px"><span class=material-symbols-outlined>bar_chart</span> Podgląd danych:</div>
            {preview_table}
        </div>
        
        <form action="/magazyn/import/execute" method="POST">
            <input type="hidden" name="filename" value="{file.filename}">
            <input type="hidden" name="paleta_id" value="{paleta_id}">
            <input type="hidden" name="dostawca" value="{paleta_dostawca}">
            
            <div class="card" style="padding:15px">
                <div style="font-weight:600;margin-bottom:15px"><span class=material-symbols-outlined>build</span> Mapowanie kolumn:</div>
                
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                    <div class="form-group">
                        <label style="font-size:0.8rem;color:#64748b">Kolumna z EAN</label>
                        <select name="col_ean" class="form-ctrl" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                            <option value="-1">-- Brak --</option>
                            {make_options(auto_ean)}
                        </select>
                        <div style="font-size:0.7rem;color:#64748b;margin-top:4px">Kod kreskowy (8-14 cyfr)</div>
                    </div>
                    <div class="form-group">
                        <label style="font-size:0.8rem;color:#64748b">Kolumna z ASIN</label>
                        <select name="col_asin" class="form-ctrl" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                            <option value="-1">-- Brak --</option>
                            {make_options(-1)}
                        </select>
                        <div style="font-size:0.7rem;color:#64748b;margin-top:4px">Amazon ASIN (B0XXXXXXXX)</div>
                    </div>
                </div>
                
                <div class="form-group">
                    <label style="font-size:0.8rem;color:#64748b">Kolumna z nazwą</label>
                    <select name="col_nazwa" class="form-ctrl" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                        {make_options(auto_nazwa)}
                    </select>
                </div>
                
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                    <div class="form-group">
                        <label style="font-size:0.8rem;color:#64748b">Ilość</label>
                        <select name="col_ilosc" class="form-ctrl" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                            {make_options(auto_ilosc)}
                        </select>
                    </div>
                    <div class="form-group">
                        <label style="font-size:0.8rem;color:#64748b">Cena</label>
                        <select name="col_cena" class="form-ctrl" style="width:100%;padding:10px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#fff">
                            {make_options(auto_cena)}
                        </select>
                    </div>
                </div>
            </div>
            
            <button type="submit" class="btn btn-ok"><span class=material-symbols-outlined>check_circle</span> IMPORTUJ {total_rows} PRODUKTÓW</button>
        </form>
        
        <a href="/magazyn/import" class="back">← Powrót</a>
        '''
        return render(html)
        
    except Exception as e:
        return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">{str(e)}</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')

@magazynier_bp.route('/import/execute', methods=['POST'])
def import_execute():
    """Wykonaj import z wybranymi kolumnami - wymaga ponownego przesłania pliku"""
    # Ze względu na bezstanowość, przekieruj do bezpośredniego importu
    # z parametrami kolumn w URL
    col_ean = request.form.get('col_ean', '-1')
    col_asin = request.form.get('col_asin', '-1')
    col_nazwa = request.form.get('col_nazwa', '1')
    col_ilosc = request.form.get('col_ilosc', '-1')
    col_cena = request.form.get('col_cena', '-1')
    paleta_id = request.form.get('paleta_id', '')
    dostawca = request.form.get('dostawca', '')  # NOWE: pobierz dostawcę
    
    # Info o palecie
    paleta_info = ''
    if paleta_id:
        conn = get_db()
        row = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (paleta_id,)).fetchone()
        paleta_nazwa = row[0] if row else f'Paleta #{paleta_id}'
        paleta_info = f'<div class="alert" style="background:#f59e0b22;border:1px solid #f59e0b;color:#f59e0b;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>inventory_2</span> Produkty zostaną przypisane do: <strong>{paleta_nazwa}</strong></div>'
    
    # Mapowanie info
    ean_info = f'EAN=kol.{int(col_ean)+1}' if int(col_ean) >= 0 else 'EAN=brak'
    asin_info = f'ASIN=kol.{int(col_asin)+1}' if int(col_asin) >= 0 else 'ASIN=brak'
    
    # NOWE: dodaj col_asin do URL
    return render(f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>download</span> IMPORT</h1><small>Krok 2</small></div>
    
    {paleta_info}
    
    <div class="alert alert-warn">Wybierz ponownie ten sam plik żeby zaimportować z wybranymi kolumnami.</div>
    
    <form action="/magazyn/import/final?col_ean={col_ean}&col_asin={col_asin}&col_nazwa={col_nazwa}&col_ilosc={col_ilosc}&col_cena={col_cena}&paleta_id={paleta_id}&dostawca={dostawca}" method="POST" enctype="multipart/form-data">
        <div class="card" style="padding:30px;text-align:center;cursor:pointer" onclick="document.getElementById('file2').click()">
            <div style="font-size:3rem;margin-bottom:10px"><span class=material-symbols-outlined>folder</span></div>
            <div style="font-weight:600">Wybierz ten sam plik</div>
            <input type="file" id="file2" name="file" style="display:none" accept=".xlsx,.csv" onchange="this.form.submit()">
        </div>
    </form>
    
    <div class="card" style="padding:15px;margin-top:15px">
        <div style="font-size:0.85rem;color:#94a3b8">
            Mapowanie: {ean_info}, {asin_info}, Nazwa=kol.{int(col_nazwa)+1}, Ilość=kol.{int(col_ilosc)+1 if int(col_ilosc)>=0 else 'brak'}, Cena=kol.{int(col_cena)+1 if int(col_cena)>=0 else 'brak'}
        </div>
    </div>
    
    <a href="/magazyn/import" class="back">← Powrót</a>
    ''')

@magazynier_bp.route('/import/final', methods=['POST'])
def import_final():
    """Finalny import z określonymi kolumnami"""
    if 'file' not in request.files:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    file = request.files['file']
    if file.filename == '':
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    col_ean = int(request.args.get('col_ean', -1))
    col_asin = int(request.args.get('col_asin', -1))
    col_nazwa = int(request.args.get('col_nazwa', 1))
    col_ilosc = int(request.args.get('col_ilosc', -1))
    col_cena = int(request.args.get('col_cena', -1))
    paleta_id = request.args.get('paleta_id', '')
    dostawca = request.args.get('dostawca', '')  # NOWE: pobierz dostawcę z URL
    
    # Konwertuj paleta_id na int lub None
    paleta_id_int = int(paleta_id) if paleta_id and paleta_id.isdigit() else None
    
    # Pobierz nazwę palety i dostawcę dla podsumowania
    paleta_nazwa = ''
    if paleta_id_int:
        conn = get_db()
        row = conn.execute('SELECT nazwa, dostawca FROM palety WHERE id = ?', (paleta_id_int,)).fetchone()
        if row:
            paleta_nazwa = row[0] if row[0] else f'Paleta #{paleta_id_int}'
            # Jeśli dostawca nie był przekazany, pobierz z palety
            if not dostawca and row[1]:
                dostawca = row[1]
    
    filename = file.filename.lower()
    added = 0
    errors = []
    
    try:
        if filename.endswith('.xlsx'):
            import openpyxl
            import tempfile
            import os as os_module
            
            tmp_path = os_module.path.join(tempfile.gettempdir(), f'import_{os_module.getpid()}.xlsx')
            file.save(tmp_path)
            
            try:
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                
                conn = get_db()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        if not row:
                            continue
                        
                        # === POBIERZ EAN I ASIN Z OSOBNYCH KOLUMN ===
                        ean = ''
                        asin = ''
                        
                        # Kolumna EAN
                        if col_ean >= 0 and col_ean < len(row) and row[col_ean]:
                            ean_raw = str(row[col_ean]).strip()
                            if ean_raw.endswith('.0'):
                                ean_raw = ean_raw[:-2]
                            ean_raw = ean_raw.replace(' ', '').replace('-', '').replace('_', '')
                            if ean_raw and ean_raw.upper() not in ['NONE', 'NAN', '']:
                                ean = ean_raw
                        
                        # Kolumna ASIN
                        if col_asin >= 0 and col_asin < len(row) and row[col_asin]:
                            asin_raw = str(row[col_asin]).strip().upper()
                            if asin_raw and asin_raw not in ['NONE', 'NAN', '']:
                                asin = asin_raw
                        
                        # Pomiń jeśli nie ma ani EAN ani ASIN
                        if not ean and not asin:
                            continue
                        
                        # === AUTO-DETEKCJA JEŚLI TYLKO JEDNA KOLUMNA ===
                        # Jeśli EAN wygląda jak ASIN, przenieś go
                        if ean and not asin and len(ean) == 10 and ean.upper().startswith('B0'):
                            asin = ean.upper()
                            ean = ''
                        
                        nazwa = str(row[col_nazwa] if col_nazwa >= 0 and col_nazwa < len(row) and row[col_nazwa] else (asin or ean)).strip()
                        if nazwa.upper() in ['NONE', 'NAN']:
                            nazwa = asin or ean
                        
                        ilosc = 1
                        if col_ilosc >= 0 and col_ilosc < len(row) and row[col_ilosc]:
                            try:
                                ilosc = int(float(str(row[col_ilosc]).replace(',', '.')) or 1)
                            except:
                                ilosc = 1
                        
                        cena = 0
                        if col_cena >= 0 and col_cena < len(row) and row[col_cena]:
                            try:
                                cena = float(str(row[col_cena]).replace(',', '.').replace(' ', '') or 0)
                            except:
                                cena = 0
                        
                        # Cena z Excela to cena JEDNOSTKOWA (brutto/RRP per item)
                        # NIE mnożymy × ilość - cena_brutto w DB to cena za sztukę
                        cena_jednostkowa = cena

                        # NIE pobieramy zdjęć przy imporcie (za wolne)
                        # Użyj przycisku "Pobierz zdjęcia" po imporcie
                        zdjecie = ''

                        # Sprawdź czy produkt już istnieje NA TEJ SAMEJ PALECIE
                        existing = None
                        if ean:
                            existing = conn.execute('SELECT id FROM produkty WHERE ean=? AND paleta_id=?', (ean, paleta_id_int)).fetchone()
                        if not existing and asin:
                            existing = conn.execute('SELECT id FROM produkty WHERE asin=? AND paleta_id=?', (asin, paleta_id_int)).fetchone()

                        if existing:
                            # Aktualizuj istniejący na tej palecie
                            conn.execute('''UPDATE produkty SET nazwa=?, ilosc=?, cena_brutto=?, dostawca=? WHERE id=?''',
                                (nazwa, ilosc, cena_jednostkowa, dostawca, existing['id']))
                        else:
                            # Nowy produkt — zawsze INSERT, nie dotykamy innych palet
                            conn.execute('''INSERT INTO produkty (ean, asin, nazwa, ilosc, cena_brutto, zdjecie_url, paleta_id, paleta, dostawca)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (ean, asin, nazwa, ilosc, cena_jednostkowa, zdjecie, paleta_id_int, paleta_nazwa, dostawca))
                        added += 1
                    except Exception as e:
                        errors.append(str(e))
                
                conn.commit()
                wb.close()
            finally:
                try:
                    os_module.remove(tmp_path)
                except:
                    pass
                    
        elif filename.endswith('.csv'):
            import csv
            import io
            
            raw_data = file.read()
            content = None
            for encoding in ['utf-8-sig', 'utf-8', 'cp1250', 'latin-1']:
                try:
                    content = raw_data.decode(encoding)
                    break
                except:
                    continue
            
            if content:
                delimiter = ';' if ';' in content[:500] else ','
                reader = csv.reader(io.StringIO(content), delimiter=delimiter)
                rows = list(reader)
                
                conn = get_db()
                for row in rows[1:]:  # Skip header
                    try:
                        if not row:
                            continue
                        
                        # === POBIERZ EAN I ASIN Z OSOBNYCH KOLUMN ===
                        ean = ''
                        asin = ''
                        
                        # Kolumna EAN
                        if col_ean >= 0 and col_ean < len(row) and row[col_ean]:
                            ean_raw = str(row[col_ean]).strip()
                            if ean_raw.endswith('.0'):
                                ean_raw = ean_raw[:-2]
                            ean_raw = ean_raw.replace(' ', '').replace('-', '').replace('_', '')
                            if ean_raw and ean_raw.upper() not in ['NONE', 'NAN', '']:
                                ean = ean_raw
                        
                        # Kolumna ASIN
                        if col_asin >= 0 and col_asin < len(row) and row[col_asin]:
                            asin_raw = str(row[col_asin]).strip().upper()
                            if asin_raw and asin_raw not in ['NONE', 'NAN', '']:
                                asin = asin_raw
                        
                        # Pomiń jeśli nie ma ani EAN ani ASIN
                        if not ean and not asin:
                            continue
                        
                        # === AUTO-DETEKCJA JEŚLI TYLKO JEDNA KOLUMNA ===
                        # Jeśli EAN wygląda jak ASIN, przenieś go
                        if ean and not asin and len(ean) == 10 and ean.upper().startswith('B0'):
                            asin = ean.upper()
                            ean = ''
                        
                        nazwa = str(row[col_nazwa] if col_nazwa >= 0 and col_nazwa < len(row) else (asin or ean)).strip() or (asin or ean)
                        
                        ilosc = 1
                        if col_ilosc >= 0 and col_ilosc < len(row) and row[col_ilosc]:
                            try:
                                ilosc = int(float(row[col_ilosc].replace(',', '.')) or 1)
                            except:
                                ilosc = 1
                        
                        cena = 0
                        if col_cena >= 0 and col_cena < len(row) and row[col_cena]:
                            try:
                                cena = float(row[col_cena].replace(',', '.').replace(' ', '') or 0)
                            except:
                                cena = 0
                        
                        # WAŻNE: Zapisujemy CAŁKOWITĄ cenę zakupu = cena_za_sztukę * ilość
                        cena_jednostkowa = cena  # NIE mnożymy × ilość
                        
                        # NIE pobieramy zdjęć przy imporcie (za wolne)
                        zdjecie = ''
                        
                        # Sprawdź czy produkt już istnieje NA TEJ SAMEJ PALECIE
                        existing = None
                        if ean:
                            existing = conn.execute('SELECT id FROM produkty WHERE ean=? AND paleta_id=?', (ean, paleta_id_int)).fetchone()
                        if not existing and asin:
                            existing = conn.execute('SELECT id FROM produkty WHERE asin=? AND paleta_id=?', (asin, paleta_id_int)).fetchone()
                        
                        if existing:
                            # Aktualizuj istniejący na tej palecie
                            conn.execute('''UPDATE produkty SET nazwa=?, ilosc=?, cena_brutto=?, dostawca=? WHERE id=?''',
                                (nazwa, ilosc, cena_jednostkowa, dostawca, existing['id']))
                        else:
                            # Nowy produkt — zawsze INSERT, nie dotykamy innych palet
                            conn.execute('''INSERT INTO produkty (ean, asin, nazwa, ilosc, cena_brutto, zdjecie_url, paleta_id, paleta, dostawca)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (ean, asin, nazwa, ilosc, cena_jednostkowa, zdjecie, paleta_id_int, paleta_nazwa, dostawca))
                        added += 1
                    except Exception as e:
                        errors.append(str(e))
                
                conn.commit()
    
    except Exception as e:
        return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">{str(e)}</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    # Zaktualizuj liczbę produktów w palecie
    if paleta_id_int and added > 0:
        try:
            conn = get_db()
            conn.execute('UPDATE palety SET ilosc_produktow = ilosc_produktow + ? WHERE id = ?', (added, paleta_id_int))
            conn.commit()
        except:
            pass
    
    # Auto-generuj polskie meta tytuły w tle po imporcie
    if added > 0 and paleta_id_int:
        import threading
        def _auto_translate_titles(pid_int):
            try:
                import time as _t
                _t.sleep(3)  # Poczekaj aż commit się zakończy
                from modules.database import get_db as _gdb2, get_config as _gc2
                from modules.utils import translate_product_name
                _conn2 = _gdb2()
                _prods = _conn2.execute(
                    "SELECT id, nazwa, meta_title FROM produkty WHERE paleta_id = ? AND (meta_title IS NULL OR meta_title = '')",
                    (pid_int,)
                ).fetchall()
                _updated = 0
                for _p in _prods:
                    try:
                        _new_name = translate_product_name(_p['nazwa'], use_ai=True)
                        if _new_name and _new_name != _p['nazwa']:
                            _conn2.execute('UPDATE produkty SET meta_title = ?, nazwa = ? WHERE id = ?',
                                         (_new_name, _new_name, _p['id']))
                            _updated += 1
                            _t.sleep(0.5)  # Rate limit AI
                    except:
                        pass
                if _updated:
                    _conn2.commit()
                    print(f"[AUTO] Przetłumaczono {_updated}/{len(_prods)} tytułów dla palety #{pid_int}")
            except Exception as _e:
                print(f"[WARN] Auto-translate error: {_e}")
        threading.Thread(target=_auto_translate_titles, args=(paleta_id_int,), daemon=True).start()

    # Info o palecie dla podsumowania
    paleta_info = ''
    if paleta_nazwa:
        dostawca_info = f' ({dostawca})' if dostawca else ''
        paleta_info = f'<div class="alert" style="background:#f59e0b22;border:1px solid #f59e0b;color:#f59e0b;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>inventory_2</span> Przypisano do palety: <strong>{paleta_nazwa}</strong>{dostawca_info}</div>'

    translate_info = f'<div class="alert" style="background:#8ff5ff22;border:1px solid #8ff5ff;color:#8ff5ff;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>translate</span> Tłumaczenie tytułów na polski uruchomione w tle...</div>' if added > 0 and paleta_id_int else ''

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> IMPORT ZAKOŃCZONY</h1></div>
    {paleta_info}
    {translate_info}
    <div class="alert alert-ok">Zaimportowano {added} produktów</div>
    '''
    if errors:
        html += f'<div class="alert alert-warn">Błędy: {len(errors)}</div>'
    html += '''
    <a href="/magazyn/import" class="btn btn-p"><span class=material-symbols-outlined>download</span> Importuj więcej</a>
    <a href="/magazyn" class="btn btn-2"><span class=material-symbols-outlined>inventory_2</span> Magazyn</a>
    <a href="/magazyn" class="back">← Powrót</a>
    '''
    return render(html)

@magazynier_bp.route('/import/upload', methods=['POST'])
def import_upload():
    """Import pliku Excel/CSV"""
    if 'file' not in request.files:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')

    file = request.files['file']
    if file.filename == '':
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie wybrano pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')

    # Pobierz paleta_id z formularza lub query string
    paleta_id = request.form.get('paleta_id', '') or request.args.get('paleta_id', '')
    paleta_id_int = int(paleta_id) if paleta_id and paleta_id.isdigit() else None

    # Pobierz nazwę palety
    paleta_nazwa = ''
    if paleta_id_int:
        conn = get_db()
        row = conn.execute('SELECT nazwa, dostawca FROM palety WHERE id = ?', (paleta_id_int,)).fetchone()
        if row:
            paleta_nazwa = row[0] if row[0] else f'Paleta #{paleta_id_int}'

    filename = file.filename.lower()
    added = 0
    errors = []
    
    try:
        if filename.endswith('.csv'):
            # Import CSV - próbuj różne kodowania
            import csv
            import io
            
            raw_data = file.read()
            content = None
            
            # Próbuj różne kodowania (polskie pliki często są w cp1250)
            for encoding in ['utf-8-sig', 'utf-8', 'cp1250', 'latin-1', 'iso-8859-2']:
                try:
                    content = raw_data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nie można odczytać pliku - nieznane kodowanie</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
            
            # Auto-wykryj separator (przecinek lub średnik)
            delimiter = ';' if ';' in content[:500] else ','
            
            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            
            conn = get_db()
            for row in reader:
                try:
                    # Szukaj kolumny z kodem produktu
                    ean = ''
                    for key in ['ean', 'EAN', 'ASIN', 'asin', 'Kod', 'KOD', 'SKU', 'sku', 'Kod produktu', 'kod']:
                        if key in row and row[key]:
                            ean = str(row[key]).strip()
                            break
                    
                    if not ean:
                        continue
                    
                    # Szukaj nazwy
                    nazwa = ean
                    for key in ['nazwa', 'Nazwa', 'NAZWA', 'name', 'Name', 'NAME', 'Tytuł', 'tytul', 'Product', 'Produkt']:
                        if key in row and row[key]:
                            nazwa = str(row[key]).strip()
                            break
                    
                    # Szukaj ilości
                    ilosc = 1
                    for key in ['ilosc', 'Ilosc', 'ILOSC', 'qty', 'Qty', 'QTY', 'Ilość', 'szt', 'Szt']:
                        if key in row and row[key]:
                            try:
                                ilosc = int(float(str(row[key]).replace(',', '.').strip()) or 1)
                            except:
                                ilosc = 1
                            break
                    
                    # Szukaj ceny
                    cena = 0
                    for key in ['cena', 'Cena', 'CENA', 'cena_brutto', 'price', 'Price', 'PRICE', 'Koszt', 'koszt']:
                        if key in row and row[key]:
                            try:
                                cena = float(str(row[key]).replace(',', '.').replace(' ', '').strip() or 0)
                            except:
                                cena = 0
                            break
                    
                    # WAŻNE: Zapisujemy CAŁKOWITĄ cenę zakupu = cena_za_sztukę * ilość
                    cena_jednostkowa = cena  # NIE mnożymy × ilość
                    
                    # Zawsze INSERT — nie scalaj produktów z różnych palet
                    conn.execute('''INSERT INTO produkty (ean, nazwa, ilosc, cena_brutto, zdjecie_url, paleta_id, paleta)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''', (ean, nazwa, ilosc, cena_jednostkowa, '', paleta_id_int, paleta_nazwa))
                    added += 1
                except Exception as e:
                    errors.append(str(e))
            
            conn.commit()
            
        elif filename.endswith('.xlsx'):
            # Import Excel - wymaga openpyxl
            try:
                import openpyxl
                import tempfile
                import os as os_module
                
                # Zapisz plik tymczasowo (Windows-compatible)
                tmp_path = os_module.path.join(tempfile.gettempdir(), f'magazynier_{os_module.getpid()}.xlsx')
                file.save(tmp_path)
                
                try:
                    wb = openpyxl.load_workbook(tmp_path)
                    ws = wb.active
                    
                    # Znajdź wiersz z nagłówkami (może nie być pierwszy)
                    header_row = 1
                    for row_idx in range(1, min(6, ws.max_row + 1)):
                        row_values = [str(cell.value or '').strip().upper() for cell in ws[row_idx]]
                        if any(h in ['EAN', 'ASIN', 'KOD', 'KOD 2', 'SKU', 'NAZWA', 'NAME'] for h in row_values):
                            header_row = row_idx
                            break
                    
                    # Pobierz nagłówki
                    headers = [str(cell.value or '').strip().upper() for cell in ws[header_row]]
                    
                    # Wykryj dostawcę
                    dostawca = detect_supplier(headers)
                    
                    # Mapowanie kolumn - bardziej elastyczne
                    col_map = {'ean': 0, 'nazwa': 1, 'ilosc': 2, 'cena': 3}  # Domyślne pozycje
                    
                    for i, h in enumerate(headers):
                        h_clean = h.replace(' ', '').replace('_', '')
                        # Kod produktu
                        if any(x in h_clean for x in ['EAN', 'ASIN', 'KOD', 'SKU', 'CODE', 'BARCODE']):
                            col_map['ean'] = i
                        # Nazwa
                        elif any(x in h_clean for x in ['NAZWA', 'NAME', 'TYTUL', 'TITLE', 'PRODUCT', 'OPIS']):
                            col_map['nazwa'] = i
                        # Ilość
                        elif any(x in h_clean for x in ['ILOSC', 'ILOŚĆ', 'QTY', 'QUANTITY', 'SZT', 'SZTUK']):
                            col_map['ilosc'] = i
                        # Cena
                        elif any(x in h_clean for x in ['CENA', 'PRICE', 'KOSZT', 'COST', 'BRUTTO']):
                            col_map['cena'] = i
                    
                    conn = get_db()
                    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                        try:
                            if not row or len(row) == 0:
                                continue
                            
                            # Pobierz EAN - sprawdź czy indeks istnieje
                            ean_idx = col_map.get('ean', 0)
                            ean = str(row[ean_idx] if ean_idx < len(row) else '').strip()
                            
                            # Wyczyść EAN z .0 na końcu (Excel traktuje jako liczbę)
                            if ean.endswith('.0'):
                                ean = ean[:-2]
                            ean = ean.replace('.0', '').replace(' ', '')
                            
                            if not ean or ean.upper() in ['NONE', 'NAN', '']:
                                continue
                            
                            # Nazwa
                            nazwa_idx = col_map.get('nazwa', 1)
                            nazwa = str(row[nazwa_idx] if nazwa_idx < len(row) and row[nazwa_idx] else ean).strip()
                            if nazwa.upper() in ['NONE', 'NAN']:
                                nazwa = ean
                            
                            # Ilość
                            ilosc = 1
                            ilosc_idx = col_map.get('ilosc', 2)
                            if ilosc_idx < len(row) and row[ilosc_idx]:
                                try:
                                    ilosc = int(float(str(row[ilosc_idx]).replace(',', '.')) or 1)
                                except:
                                    ilosc = 1
                            
                            # Cena
                            cena = 0
                            cena_idx = col_map.get('cena', 3)
                            if cena_idx < len(row) and row[cena_idx]:
                                try:
                                    cena = float(str(row[cena_idx]).replace(',', '.').replace(' ', '') or 0)
                                except:
                                    cena = 0
                            
                            # WAŻNE: Zapisujemy CAŁKOWITĄ cenę zakupu = cena_za_sztukę * ilość
                            cena_jednostkowa = cena  # NIE mnożymy × ilość
                            
                            # Zawsze INSERT — nie scalaj produktów z różnych palet
                            conn.execute('''INSERT INTO produkty (ean, nazwa, ilosc, cena_brutto, dostawca, zdjecie_url, paleta_id, paleta)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', (ean, nazwa, ilosc, cena_jednostkowa, dostawca or '', '', paleta_id_int, paleta_nazwa))
                            added += 1
                        except Exception as row_err:
                            errors.append(str(row_err))
                    
                    conn.commit()
                    wb.close()
                finally:
                    # Usuń plik tymczasowy
                    try:
                        os_module.remove(tmp_path)
                    except:
                        pass
                    
            except ImportError:
                return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
        else:
            return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Nieobsługiwany format pliku</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')
    
    except Exception as e:
        return render(f'<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">{str(e)}</div><a href="/magazyn/import" class="btn btn-p">← Powrót</a>')

    # Zaktualizuj liczbę produktów w palecie
    if paleta_id_int and added > 0:
        try:
            conn = get_db()
            conn.execute('UPDATE palety SET ilosc_produktow = ilosc_produktow + ? WHERE id = ?', (added, paleta_id_int))
            conn.commit()
        except:
            pass

    # Info o palecie
    paleta_info = ''
    if paleta_nazwa:
        paleta_info = f'<div class="alert" style="background:#f59e0b22;border:1px solid #f59e0b;color:#f59e0b;padding:10px;border-radius:8px;margin-bottom:15px"><span class=material-symbols-outlined>inventory_2</span> Przypisano do palety: <strong>{paleta_nazwa}</strong></div>'

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>check_circle</span> IMPORT ZAKOŃCZONY</h1></div>
    {paleta_info}
    <div class="alert alert-ok">Zaimportowano {added} produktów</div>
    '''
    if errors:
        html += f'<div class="alert alert-warn">Błędy: {len(errors)}</div>'
    html += '''
    <a href="/magazyn/import" class="btn btn-p"><span class=material-symbols-outlined>download</span> Importuj więcej</a>
    <a href="/magazyn" class="btn btn-2"><span class=material-symbols-outlined>inventory_2</span> Magazyn</a>
    <a href="/magazyn" class="back">← Powrót</a>
    '''
    return render(html)

@magazynier_bp.route('/dodaj')
def dodaj():
    """Przekierowanie do dodawania produktu"""
    html = '''
    <div class="hdr"><h1><span class=material-symbols-outlined>add</span> DODAJ PRODUKT</h1></div>
    
    <div class="card" style="padding:15px">
        <form action="/magazyn/szukaj" method="GET">
            <div class="form-group">
                <label>Wpisz EAN / ASIN / SKU</label>
                <input type="text" name="q" class="form-ctrl" placeholder="np. B0CFQBBT7G" autofocus required>
            </div>
            <button type="submit" class="btn btn-ok"><span class=material-symbols-outlined>search</span> SZUKAJ / DODAJ</button>
        </form>
    </div>
    
    <div style="text-align:center;color:#64748b;padding:15px">lub</div>
    
    <a href="/magazyn/skanuj" class="btn btn-p"><span class=material-symbols-outlined>photo_camera</span> SKANUJ KAMERĄ</a>
    <a href="/magazyn/import" class="btn btn-2" style="margin-top:10px"><span class=material-symbols-outlined>download</span> IMPORT Z PLIKU</a>
    
    <a href="/magazyn" class="back">← Powrót</a>
    '''
    return render(html)

@magazynier_bp.route('/skanuj')
def skanuj_kamera():
    """Skaner QR/Barcode kamerą telefonu"""
    from .database import get_config
    base_url = get_config('app_base_url', 'http://localhost:5000')
    
    html = '''
    <div class="hdr">
        <h1><span class=material-symbols-outlined>photo_camera</span> SKANER</h1>
        <small>Skanuj QR lub kod kreskowy</small>
    </div>
    
    <div class="card" style="padding:0;overflow:hidden;border-radius:16px;background:#000">
        <div id="reader" style="width:100%;min-height:300px"></div>
    </div>
    
    <div id="result" class="card" style="display:none;padding:15px;margin-top:15px">
        <div style="font-weight:600;margin-bottom:8px"><span class=material-symbols-outlined>inventory_2</span> Znaleziono:</div>
        <div id="resultText" style="font-size:1.1rem;word-break:break-all"></div>
    </div>
    
    <div id="notFound" class="alert alert-warn" style="display:none;margin-top:15px">
        <span class=material-symbols-outlined>warning</span> Nie znaleziono produktu o tym kodzie
        <div style="margin-top:10px">
            <a id="addNewLink" href="#" class="btn btn-p" style="display:inline-block;padding:10px 20px"><span class=material-symbols-outlined>add</span> DODAJ NOWY</a>
        </div>
    </div>
    
    <div style="display:flex;gap:10px;margin-top:15px">
        <button onclick="switchCamera()" class="btn btn-2" style="flex:1"><span class=material-symbols-outlined>sync</span> Zmień kamerę</button>
        <button onclick="toggleFlash()" class="btn btn-2" style="flex:1">🔦 Latarka</button>
    </div>
    
    <div style="margin-top:10px;font-size:0.75rem;color:#64748b;text-align:center">
        Obsługuje: QR, EAN-13, EAN-8, Code128, UPC-A
    </div>
    
    <a href="/magazyn/dodaj" class="back">← Powrót</a>
    
    <!-- html5-qrcode library -->
    <script src="https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js" integrity="sha384-c9d8RFSL+u3exBOJ4Yp3HUJXS4znl9f+z66d1y54ig+ea249SpqR+w1wyvXz/lk+" crossorigin="anonymous"></script>
    
    <script>
    const baseUrl = "''' + base_url + '''";
    let html5QrCode;
    let currentCamera = 'environment';
    let isProcessing = false;
    let scanCount = 0;

    // Debug info
    const dbg = document.createElement('div');
    dbg.id = 'scanDebug';
    dbg.style.cssText = 'font-size:11px;color:#94a3b8;text-align:center;margin-top:8px';
    dbg.textContent = 'Uruchamianie skanera...';
    document.getElementById('reader').parentNode.after(dbg);

    function onScanSuccess(decodedText, decodedResult) {
        if (isProcessing) return;
        isProcessing = true;

        const fmt = decodedResult.result.format?.formatName || '?';
        console.log('Zeskanowano:', decodedText, fmt);
        dbg.textContent = 'Odczytano: ' + decodedText + ' (' + fmt + ')';

        if (navigator.vibrate) navigator.vibrate([50, 30, 100]);

        document.getElementById('resultText').textContent = decodedText;
        document.getElementById('result').style.display = 'block';

        if (html5QrCode) html5QrCode.pause(true);

        // URL z naszej apki - przekieruj
        if (decodedText.includes('/magazyn/qr/') || decodedText.includes('/magazyn/produkt/')) {
            window.location.href = decodedText;
            return;
        }

        // Obetnij prefix MAG:
        let searchCode = decodedText;
        if (decodedText.toUpperCase().startsWith('MAG:')) {
            searchCode = decodedText.substring(4);
        }

        fetch('/magazyn/api/szukaj?q=' + encodeURIComponent(searchCode))
            .then(r => r.json())
            .then(data => {
                if (data.found && data.product_id) {
                    window.location.href = '/magazyn/produkt/' + data.product_id;
                } else {
                    showNotFound(searchCode);
                }
            })
            .catch(err => {
                console.error(err);
                showNotFound(searchCode);
            });
    }

    function showNotFound(code) {
        document.getElementById('notFound').style.display = 'block';
        document.getElementById('addNewLink').href = '/magazyn/dodaj?ean=' + encodeURIComponent(code);
        setTimeout(() => {
            isProcessing = false;
            document.getElementById('notFound').style.display = 'none';
            document.getElementById('result').style.display = 'none';
            if (html5QrCode) html5QrCode.resume();
        }, 3000);
    }

    function onScanFailure(errorMessage) {
        // Liczy próby skanowania (debug)
        scanCount++;
        if (scanCount % 100 === 0) {
            dbg.textContent = 'Skanuje... (' + scanCount + ' klatek) - skieruj kamerę na kod';
        }
    }

    function startScanner() {
        html5QrCode = new Html5Qrcode("reader", {
            experimentalFeatures: { useBarCodeDetectorIfSupported: true },
            verbose: false
        });

        const config = {
            fps: 10,
            qrbox: function(viewfinderWidth, viewfinderHeight) {
                let size = Math.floor(Math.min(viewfinderWidth, viewfinderHeight) * 0.8);
                return { width: size, height: size };
            },
            formatsToSupport: [
                Html5QrcodeSupportedFormats.QR_CODE,
                Html5QrcodeSupportedFormats.EAN_13,
                Html5QrcodeSupportedFormats.EAN_8,
                Html5QrcodeSupportedFormats.CODE_128,
                Html5QrcodeSupportedFormats.CODE_39,
                Html5QrcodeSupportedFormats.UPC_A,
                Html5QrcodeSupportedFormats.UPC_E
            ]
        };

        html5QrCode.start(
            { facingMode: currentCamera },
            config,
            onScanSuccess,
            onScanFailure
        ).then(() => {
            dbg.textContent = 'Kamera aktywna - skieruj na kod QR';
        }).catch(err => {
            console.error('Blad kamery:', err);
            dbg.textContent = 'Blad: ' + err;
            document.getElementById('reader').innerHTML = '<div style="padding:40px;text-align:center;color:#ef4444">Brak dostepu do kamery<br><small style="color:#64748b">Zezwol w ustawieniach przegladarki.<br>Wymagany HTTPS lub localhost.</small></div>';
        });
    }

    function switchCamera() {
        if (html5QrCode) {
            html5QrCode.stop().then(() => {
                currentCamera = currentCamera === 'environment' ? 'user' : 'environment';
                scanCount = 0;
                startScanner();
            });
        }
    }

    function toggleFlash() {
        if (html5QrCode) {
            html5QrCode.applyVideoConstraints({
                advanced: [{ torch: true }]
            }).catch(e => {
                alert('Latarka nie jest obslugiwana na tym urzadzeniu');
            });
        }
    }

    startScanner();
    </script>
    '''
    return render(html)

@magazynier_bp.route('/api/szukaj')
def api_szukaj():
    """API do szukania produktu po EAN/ASIN"""
    from flask import jsonify
    
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'found': False})
    
    # Obsługa kodów QR z prefiksem MAG: (format: MAG:B0B6NDX3SS)
    if q.upper().startswith('MAG:'):
        q = q[4:]  # Usuń prefiks "MAG:" - zostaje sam kod np. "B0B6NDX3SS"

    conn = get_db()
    # Szukaj po kodzie magazynowym, EAN, ASIN lub nazwie
    product = conn.execute('''
        SELECT id FROM produkty
        WHERE kod_magazynowy = ? OR ean = ? OR asin = ? OR nazwa LIKE ?
        LIMIT 1
    ''', (q.upper(), q, q, f'%{q}%')).fetchone()
    
    if product:
        return jsonify({'found': True, 'product_id': product['id']})
    else:
        return jsonify({'found': False})

@magazynier_bp.route('/etykiety')
def etykiety():
    """Drukowanie etykiet - wybór produktów z filtrem po palecie"""
    conn = get_db()

    # Filtr po palecie (opcjonalny)
    paleta_filter = request.args.get('paleta_id', '')

    if paleta_filter:
        try:
            products = conn.execute('SELECT * FROM produkty WHERE paleta_id = ? AND ilosc > 0 ORDER BY id', (int(paleta_filter),)).fetchall()
        except:
            products = conn.execute('SELECT * FROM produkty WHERE ilosc > 0 ORDER BY data_dodania DESC LIMIT 100').fetchall()
    else:
        products = conn.execute('SELECT * FROM produkty WHERE ilosc > 0 ORDER BY data_dodania DESC LIMIT 100').fetchall()

    # Lista palet do filtra
    try:
        palety = conn.execute('SELECT id, nazwa FROM palety ORDER BY id DESC LIMIT 50').fetchall()
    except:
        palety = []

    paleta_nazwa = ''
    if paleta_filter:
        try:
            pn = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (int(paleta_filter),)).fetchone()
            if pn:
                paleta_nazwa = pn['nazwa']
        except:
            pass

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>label</span> ETYKIETY</h1><small>Drukuj etykiety Niimbot z QR kodem</small></div>

    <!-- Filtr palety -->
    <div style="display:flex;gap:8px;margin-bottom:12px;align-items:center;flex-wrap:wrap">
        <select onchange="window.location='/magazyn/etykiety'+(this.value ? '?paleta_id='+this.value : '')"
            style="flex:1;padding:10px 12px;backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:10px;color:#e2e8f0;font-size:0.85rem">
            <option value="">📦 Wszystkie produkty</option>'''

    for pal in palety:
        selected = 'selected' if str(pal['id']) == str(paleta_filter) else ''
        html += f'<option value="{pal["id"]}" {selected}>{pal["nazwa"]}</option>'

    html += f'''
        </select>
        <input type="text" id="searchInput" oninput="filterProducts()" placeholder="Szukaj..."
            style="width:180px;padding:10px 12px;backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:10px;color:#e2e8f0;font-size:0.85rem">
    </div>

    <!-- Akcje masowe -->
    <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center">
        <button onclick="toggleAll(true)" style="padding:6px 12px;background:#1e293b;border:1px solid #334155;border-radius:8px;color:#94a3b8;font-size:0.8rem;cursor:pointer"><span class="material-symbols-outlined" style="font-size:0.9rem;vertical-align:middle">check_box</span> Zaznacz wszystkie</button>
        <button onclick="toggleAll(false)" style="padding:6px 12px;background:#1e293b;border:1px solid #334155;border-radius:8px;color:#94a3b8;font-size:0.8rem;cursor:pointer">◻ Odznacz</button>
        <span id="countLabel" style="flex:1;text-align:right;font-size:0.8rem;color:#64748b">0 zaznaczonych</span>
    </div>

    <form action="/magazyn/etykiety/drukuj" method="POST" id="printForm">
    <input type="hidden" name="drukarka" id="drukarkaInput" value="niimbot">
    '''

    stan_colors = {
        'Nowy': '#beee00', 'Jak nowy': '#8ff5ff', 'Dobry': '#eab308',
        'Uszkodzony': '#f97316', 'Zniszczony': '#ef4444'
    }

    for p in products:
        img_url = p['zdjecie_url'] or ''
        img_html = f'<img src="{img_url}" style="width:50px;height:50px;object-fit:cover;border-radius:6px;background:#1e1e2e" onerror="this.style.display=\'none\'">' if img_url else '<div style="width:50px;height:50px;background:#1e1e2e;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:1.2rem"><span class=material-symbols-outlined>inventory_2</span></div>'

        # Stan przyjęcia badge
        try:
            stan_p = p['stan_przyjecia'] or ''
        except:
            stan_p = ''
        stan_badge = ''
        if stan_p:
            sc = stan_colors.get(stan_p, '#94a3b8')
            stan_badge = f'<span style="font-size:0.65rem;color:{sc};border:1px solid {sc};padding:1px 5px;border-radius:4px;margin-left:4px">{stan_p}</span>'

        html += f'''<div class="item prod-row" data-name="{p['nazwa'].lower()}" style="cursor:pointer;display:flex;align-items:center;gap:10px" onclick="this.querySelector('input').click();updateCount()">
            <input type="checkbox" name="produkty" value="{p['id']}" style="width:20px;height:20px;flex-shrink:0" onclick="event.stopPropagation();updateCount()">
            {img_html}
            <div class="item-info" style="flex:1;min-width:0">
                <div class="item-name">{p['nazwa'][:40]}{stan_badge}</div>
                <div class="item-meta">{p['ean'] or 'N/A'} | <span class=material-symbols-outlined>pin_drop</span>{p['lokalizacja'] or '—'}</div>
            </div>
            <div class="item-right" style="text-align:right;flex-shrink:0">
                <div class="item-qty">{p['ilosc']}</div>
                <div class="item-price">{p['cena_allegro']:.0f} zł</div>
            </div>
        </div>'''

    html += f'''
    </form>

    <div style="height:20px"></div>
    <div style="position:sticky;bottom:0;padding:12px 0;background:linear-gradient(transparent 0%, #0a0a0f 25%);z-index:50">
        <button onclick="drukuj()" class="btn btn-purple" style="width:100%;padding:16px;font-size:1.1rem;font-weight:700;border-radius:12px"><span class="material-symbols-outlined" style="vertical-align:middle">smartphone</span> DRUKUJ NIIMBOT</button>
    </div>

    <script>
    function drukuj() {{
        var form = document.getElementById('printForm');
        var checked = form.querySelectorAll('input[name="produkty"]:checked');
        if (checked.length === 0) {{
            alert('Wybierz co najmniej jeden produkt!');
            return;
        }}
        form.submit();
    }}

    function toggleAll(state) {{
        document.querySelectorAll('.prod-row').forEach(row => {{
            if (row.style.display !== 'none') {{
                row.querySelector('input[type="checkbox"]').checked = state;
            }}
        }});
        updateCount();
    }}

    function updateCount() {{
        setTimeout(() => {{
            const n = document.querySelectorAll('input[name="produkty"]:checked').length;
            document.getElementById('countLabel').textContent = n + ' zaznaczonych';
        }}, 50);
    }}

    function filterProducts() {{
        const q = document.getElementById('searchInput').value.toLowerCase();
        document.querySelectorAll('.prod-row').forEach(row => {{
            row.style.display = row.dataset.name.includes(q) ? '' : 'none';
        }});
    }}
    </script>

    <a href="/magazyn" class="back">← Powrót</a>
    '''
    return render(html)


@magazynier_bp.route('/etykiety/drukuj', methods=['POST'])
def etykiety_drukuj():
    """Generowanie etykiet PDF z QR kodami"""
    produkty_ids = request.form.getlist('produkty')
    drukarka = request.form.get('drukarka', 'vretti')
    
    if not produkty_ids:
        return redirect('/magazyn/etykiety')
    
    conn = get_db()
    products = []
    for pid in produkty_ids:
        p = conn.execute('SELECT * FROM produkty WHERE id=?', (pid,)).fetchone()
        if p:
            products.append(dict(p))
    
    if drukarka == 'niimbot':
        return etykiety_niimbot_page(products)
    else:
        return etykiety_vretti_pdf(products)


def etykiety_vretti_pdf(products):
    """Generuje PDF z etykietami dla Vretti 420B (100x150mm)"""
    
    # Funkcja zamiany polskich znaków na ASCII (Helvetica nie obsługuje)
    def pl_to_ascii(text):
        if not text:
            return text
        replacements = {
            'ą': 'a', 'ć': 'c', 'ę': 'e', 'ł': 'l', 'ń': 'n', 
            'ó': 'o', 'ś': 's', 'ź': 'z', 'ż': 'z',
            'Ą': 'A', 'Ć': 'C', 'Ę': 'E', 'Ł': 'L', 'Ń': 'N',
            'Ó': 'O', 'Ś': 'S', 'Ź': 'Z', 'Ż': 'Z'
        }
        for pl, ascii in replacements.items():
            text = text.replace(pl, ascii)
        return text
    
    try:
        from reportlab.lib.pagesizes import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        import qrcode
        import io
        
        # Pobierz bazowy URL z konfiguracji (domyślnie localhost)
        base_url = get_config('app_base_url', 'http://localhost:5000')
        
        # Rozmiar etykiety: 100x150mm
        width, height = 100*mm, 150*mm
        
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=(width, height))
        
        for product in products:
            ilosc = max(int(product.get('ilosc') or 1), 1)
            mag_code = f"MAG-{product['id']:05d}"

            for szt_nr in range(1, ilosc + 1):
                # Numer seryjny: MAG-00464/1, MAG-00464/2, ...
                serial = f"{mag_code}/{szt_nr}" if ilosc > 1 else mag_code

                # QR kod z linkiem do produktu + numer seryjny
                qr = qrcode.QRCode(version=1, box_size=10, border=2)
                qr.add_data(f"{base_url}/magazyn/qr/{product['id']}?sn={szt_nr}")
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")

                # Zapisz QR do bufora
                qr_buffer = io.BytesIO()
                qr_img.save(qr_buffer, format='PNG')
                qr_buffer.seek(0)

                # Rysuj etykietę
                c.setFont("Helvetica-Bold", 16)

                # Nazwa produktu (max 2 linie) - zamień polskie znaki
                nazwa = pl_to_ascii(product['nazwa'][:60])
                if len(nazwa) > 30:
                    c.drawString(10*mm, 135*mm, nazwa[:30])
                    c.drawString(10*mm, 128*mm, nazwa[30:60])
                else:
                    c.drawString(10*mm, 135*mm, nazwa)

                # Numer seryjny (duży, widoczny)
                if ilosc > 1:
                    c.setFont("Helvetica-Bold", 14)
                    c.drawString(10*mm, 118*mm, f"S/N: {serial}")
                    c.setFont("Helvetica", 9)
                    c.drawString(10*mm, 112*mm, f"Sztuka {szt_nr} z {ilosc}")

                # QR kod (lewa strona)
                qr_reader = ImageReader(qr_buffer)
                c.drawImage(qr_reader, 8*mm, 65*mm, width=35*mm, height=35*mm)

                # Info obok QR - zamień polskie znaki
                c.setFont("Helvetica", 11)
                y_info = 98*mm
                c.drawString(48*mm, y_info, f"Szt: {ilosc}")
                y_info -= 9*mm

                # Stan przyjęcia (jeśli jest)
                stan_przyjecia = product.get('stan_przyjecia') or ''
                stan_display = stan_przyjecia or pl_to_ascii(product.get('stan') or 'Nowy')
                c.drawString(48*mm, y_info, f"Stan: {pl_to_ascii(stan_display)}")
                y_info -= 9*mm

                c.drawString(48*mm, y_info, f"Polka: {pl_to_ascii(product['lokalizacja'] or '—')}")
                y_info -= 9*mm

                # Cena
                cena = product.get('cena_allegro') or 0
                if cena:
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(48*mm, y_info, f"{cena:.0f} zl")
                    y_info -= 9*mm

                # EAN/kod
                c.setFont("Helvetica", 10)
                ean = product.get('ean') or product.get('asin') or ''
                if ean:
                    c.drawString(10*mm, 55*mm, f"EAN: {ean}")
                else:
                    c.drawString(10*mm, 55*mm, "Brak EAN")

                # Barcode EAN (jeśli jest)
                if ean and len(ean) >= 8:
                    try:
                        from reportlab.graphics.barcode import code128
                        barcode = code128.Code128(ean, barWidth=0.4*mm, barHeight=15*mm)
                        barcode.drawOn(c, 10*mm, 20*mm)
                    except:
                        c.drawString(10*mm, 30*mm, ean)

                # Data wydruku + serial
                c.setFont("Helvetica", 8)
                c.drawString(10*mm, 8*mm, f"{serial} | {datetime.now().strftime('%Y-%m-%d')}")

                c.showPage()
        
        c.save()
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'inline; filename=etykiety_vretti.pdf'}
        )
        
    except ImportError as e:
        return render(f'''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div>
            <div class="alert alert-err">Brak biblioteki: {e}<br><br>
            Zainstaluj: <code>pip install reportlab qrcode pillow</code></div>
            <a href="/magazyn/etykiety" class="back">← Powrót</a>
        ''')


def etykiety_niimbot_page(products):
    """Strona do drukowania na Niimbot B1 - masowe drukowanie"""
    import json
    from .printer_manager import get_printer_manager as get_printer, BLEAK_AVAILABLE, IMAGING_AVAILABLE
    
    # Generuj podglądy etykiet
    previews = []
    pm = get_printer()
    base_url = get_config('app_base_url', 'http://localhost:5000')
    
    conn = get_db()
    for p in products:
        from .printer_manager import ProductLabel
        # Dane palety
        paleta_nazwa = ''
        koszt_szt = 0
        if p.get('paleta_id'):
            koszt_szt = _paleta_koszt_szt(conn, p['paleta_id'])
            pal_row = conn.execute('SELECT nazwa FROM palety WHERE id=?', (p['paleta_id'],)).fetchone()
            if pal_row:
                paleta_nazwa = pal_row['nazwa'] or ''

        kod_mag = p.get('kod_magazynowy', '') or ''
        label = ProductLabel(
            nazwa=p['nazwa'][:70],
            qr_data=kod_mag if kod_mag else f"MAG:{p.get('ean') or p.get('asin') or p['id']}",
            lokalizacja=p.get('lokalizacja', ''),
            ean=p.get('ean', ''),
            ilosc=p.get('ilosc', 1) or 1,
            dostawca=p.get('dostawca', '') or '',
            data_zakupu=p.get('data_zakupu', '') or p.get('data_dodania', '') or '',
            paleta=paleta_nazwa,
            koszt_szt=koszt_szt,
            cena_allegro=0,
            kod_magazynowy=kod_mag,
            stan_przyjecia=_format_stan_label(p.get('stan_przyjecia', ''), p.get('klasa_jakosci', ''))
        )
        preview = pm.generate_label_preview(label) if IMAGING_AVAILABLE else ''
        previews.append({
            'id': p['id'],
            'nazwa': p['nazwa'],
            'preview': preview,
            'ean': p.get('ean', ''),
            'lokalizacja': p.get('lokalizacja', ''),
            'ilosc': p.get('ilosc', 1),
            'dostawca': p.get('dostawca', '')
        })
    
    products_json = json.dumps(products)
    
    # Status backendu
    backend_status = '<span class=material-symbols-outlined>check_circle</span> Gotowe' if BLEAK_AVAILABLE else '<span class=material-symbols-outlined>cancel</span> Brak biblioteki bleak'
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>label</span> ETYKIETY NIIMBOT</h1><small>{len(products)} etykiet do druku</small></div>

    <!-- Instrukcja -->
    <div class="card" style="background:linear-gradient(135deg,rgba(190,238,0,0.15),rgba(143,245,255,0.15));border:1px solid rgba(190,238,0,0.3);padding:12px;margin-bottom:12px">
        <div style="font-size:0.85rem;color:#e2e8f0;line-height:1.5">
            Kliknij <span class=material-symbols-outlined>print</span> → pobierze PNG → otworz w apce Niimbot
        </div>
    </div>

    <!-- Masowe akcje -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:15px">
        <button onclick="downloadAll()" class="btn btn-p" style="padding:14px;font-size:0.95rem" id="btnAll">
            <span class=material-symbols-outlined>download</span> POBIERZ WSZ.
        </button>
        <a href="/magazyn/etykiety/niimbot/zip?ids={','.join(str(p['id']) for p in products)}" class="btn btn-2" style="padding:14px;font-size:0.95rem;display:flex;align-items:center;justify-content:center;text-decoration:none">
            <span class=material-symbols-outlined>inventory_2</span> ZIP
        </a>
        <button onclick="openNiimbot()" class="btn btn-purple" style="padding:14px;font-size:0.95rem">
            [SMARTPHONE] NIIMBOT
        </button>
    </div>

    <!-- LISTA ETYKIET -->
    <div id="printList" style="display:flex;flex-direction:column;gap:10px">
    '''

    for i, pv in enumerate(previews):
        html += f'''
        <div class="card" style="padding:12px" id="card-{i}">
            <div style="display:flex;gap:12px;align-items:center">
                <img src="{pv['preview']}" style="width:100px;height:auto;border:1px solid #2d2d3e;border-radius:8px;cursor:pointer;background:#fff"
                     alt="Etykieta" onclick="showPreview({pv['id']}, this.src)">
                <div style="flex:1;min-width:0">
                    <div style="font-weight:600;font-size:0.9rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{pv['nazwa'][:35]}</div>
                    <div style="font-size:0.75rem;color:#64748b;margin-top:4px">{pv['ean'] or 'Brak EAN'} | <span class=material-symbols-outlined>pin_drop</span> {pv['lokalizacja'] or '—'}</div>
                    <div style="font-size:0.7rem;color:#ff6b9b;margin-top:2px">x{pv['ilosc']} szt.</div>
                </div>
                <button onclick="printLabel({pv['id']}, '{(pv['ean'] or str(pv['id']))}', {i})"
                   style="min-width:60px;padding:14px 18px;background:rgba(190,238,0,0.15);border:1px solid rgba(190,238,0,0.3);color:#beee00;border:none;border-radius:12px;font-size:1.1rem;font-weight:700;cursor:pointer"
                   id="btn-{i}">
                    <span class=material-symbols-outlined>save</span>
                </button>
            </div>
        </div>'''

    html += f'''
    </div>

    <!-- Fullscreen preview overlay -->
    <div id="previewOverlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.95);z-index:1000;flex-direction:column;align-items:center;justify-content:center;padding:20px"
         onclick="closePreview()">
        <img id="previewImg" style="max-width:90%;max-height:70vh;border-radius:8px;background:#fff">
        <div style="margin-top:20px;display:flex;gap:12px">
            <a id="previewDownloadBtn" download onclick="event.stopPropagation()"
               style="padding:16px 32px;background:rgba(190,238,0,0.15);border:1px solid rgba(190,238,0,0.3);color:#beee00;border:none;border-radius:12px;font-size:1.1rem;font-weight:700;cursor:pointer;text-decoration:none;display:flex;align-items:center">
                <span class=material-symbols-outlined>download</span> POBIERZ PNG
            </a>
            <button onclick="event.stopPropagation();openNiimbot()"
                    style="padding:16px 32px;background:rgba(255,107,155,0.15);border:1px solid rgba(255,107,155,0.3);color:#ff6b9b;border:none;border-radius:12px;font-size:1.1rem;font-weight:700;cursor:pointer">
                [SMARTPHONE] OTWORZ NIIMBOT
            </button>
        </div>
    </div>

    <!-- Licznik -->
    <div id="counter" style="display:none;position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:rgba(190,238,0,0.15);border:1px solid rgba(190,238,0,0.3);color:#beee00;padding:10px 20px;border-radius:20px;font-weight:600;z-index:100"></div>

    <a href="/magazyn/etykiety" class="back">← Powrot</a>

    <script>
    const products = {products_json};
    let printed = 0;

    // Pobierz i zapisz PNG — potem user otwiera w Niimbot
    async function printLabel(productId, ean, btnIdx) {{
        const btn = document.getElementById('btn-' + btnIdx);
        btn.textContent = '';
        btn.disabled = true;

        try {{
            // Pobierz PNG
            const link = document.createElement('a');
            link.href = '/magazyn/etykiety/niimbot/png/' + productId;
            link.download = 'etykieta_' + ean + '.png';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);

            btn.textContent = '';
            btn.style.background = '#64748b';
            markPrinted();
        }} catch(e) {{
            btn.textContent = '';
            btn.disabled = false;
        }}
    }}

    // Podglad pelnoekranowy — klik w miniaturke
    function showPreview(productId, imgSrc) {{
        const overlay = document.getElementById('previewOverlay');
        overlay.style.display = 'flex';
        document.getElementById('previewImg').src = imgSrc;

        document.getElementById('previewDownloadBtn').href = '/magazyn/etykiety/niimbot/png/' + productId;
        document.getElementById('previewDownloadBtn').download = 'etykieta_' + productId + '.png';
    }}

    function closePreview() {{
        document.getElementById('previewOverlay').style.display = 'none';
    }}

    // Pobierz wszystkie po kolei
    async function downloadAll() {{
        const btn = document.getElementById('btnAll');
        btn.disabled = true;

        for (let i = 0; i < products.length; i++) {{
            const p = products[i];
            const link = document.createElement('a');
            link.href = '/magazyn/etykiety/niimbot/png/' + p.id;
            link.download = 'etykieta_' + (p.ean || p.id) + '.png';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);

            const cardBtn = document.getElementById('btn-' + i);
            if (cardBtn) {{ cardBtn.textContent = ''; cardBtn.style.background = '#64748b'; }}
            markPrinted();

            btn.textContent = ' ' + (i+1) + '/' + products.length;
            await new Promise(r => setTimeout(r, 600));
        }}

        btn.disabled = false;
        btn.textContent = ' POBIERZ WSZ.';
        showToast('Wszystkie pobrane! Otworz apke Niimbot');
    }}

    // Otworz apke Niimbot na Androidzie
    function openNiimbot() {{
        const a = document.createElement('a');
        a.href = 'intent://#Intent;package=com.gengcon.android.jccloudprinter;S.browser_fallback_url=https%3A%2F%2Fplay.google.com%2Fstore%2Fapps%2Fdetails%3Fid%3Dcom.gengcon.android.jccloudprinter;end';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
    }}

    function markPrinted() {{
        printed++;
        showToast('Pobrano ' + printed + '/' + products.length);
    }}

    function showToast(msg) {{
        const c = document.getElementById('counter');
        c.style.display = 'block';
        c.textContent = msg;
        if (printed >= products.length) {{
            c.style.background = '#ff6b9b';
        }}
        clearTimeout(c._timer);
        c._timer = setTimeout(() => {{ c.style.display = 'none'; }}, 4000);
    }}
    </script>
    '''
    return render(html)


@magazynier_bp.route('/api/print/niimbot', methods=['POST'])
def api_print_niimbot():
    """API do drukowania pojedynczej etykiety na Niimbot przez backend (bleak)"""
    from flask import jsonify
    import asyncio
    
    try:
        from .printer_manager import get_printer_manager as get_printer, ProductLabel, BLEAK_AVAILABLE
    except ImportError:
        return jsonify({'success': False, 'message': 'Brak modułu printer_manager'})
    
    if not BLEAK_AVAILABLE:
        return jsonify({'success': False, 'message': 'Brak biblioteki bleak - zainstaluj: pip install bleak'})
    
    data = request.get_json()
    product_id = data.get('product_id')
    
    if not product_id:
        return jsonify({'success': False, 'message': 'Brak product_id'})
    
    # Pobierz produkt
    conn = get_db()
    product = conn.execute('SELECT * FROM produkty WHERE id=?', (product_id,)).fetchone()

    if not product:
        return jsonify({'success': False, 'message': 'Produkt nie znaleziony'})

    p = dict(product)

    # Dane palety
    paleta_nazwa = ''
    koszt_szt = 0
    if p.get('paleta_id'):
        koszt_szt = _paleta_koszt_szt(conn, p['paleta_id'])
        pal_row = conn.execute('SELECT nazwa FROM palety WHERE id=?', (p['paleta_id'],)).fetchone()
        if pal_row:
            paleta_nazwa = pal_row['nazwa'] or ''

    # Przygotuj etykietę
    kod_mag = p.get('kod_magazynowy', '') or ''
    label = ProductLabel(
        nazwa=p['nazwa'][:35],
        qr_data=kod_mag if kod_mag else f"MAG:{p.get('ean') or p.get('asin') or p['id']}",
        lokalizacja=p.get('lokalizacja', ''),
        ean=p.get('ean', ''),
        ilosc=p.get('ilosc', 1) or 1,
        dostawca=p.get('dostawca', '') or '',
        data_zakupu=p.get('data_zakupu', '') or p.get('data_dodania', '') or '',
        paleta=paleta_nazwa,
        koszt_szt=koszt_szt,
        cena_allegro=0,
        kod_magazynowy=kod_mag,
        stan_przyjecia=_format_stan_label(p.get('stan_przyjecia', ''), p.get('klasa_jakosci', ''))
    )
    
    # Drukuj przez BleakTransport + niimprint (prawidłowy protokół Niimbot)
    pm = get_printer()
    bt_address = pm.device_address or ''

    # Pobierz adres z config jeśli brak
    if not bt_address:
        try:
            from .database import get_config
            bt_address = get_config('niimbot_bt_address', '')
        except:
            pass

    try:
        from .printer_manager import print_niimbot_ble_sync
        result = print_niimbot_ble_sync(
            nazwa=label.nazwa,
            qr_data=label.qr_data,
            lokalizacja=label.lokalizacja,
            ean=label.ean,
            bt_address=bt_address,
            copies=1,
            ilosc=label.ilosc,
            dostawca=label.dostawca,
            data_zakupu=label.data_zakupu,
            paleta=label.paleta,
            koszt_szt=label.koszt_szt,
            cena_allegro=label.cena_allegro
        )
        if not result.get('success'):
            result['message'] = f"[ROUTE:api_print_niimbot] {result.get('message','')}"
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': f'[ROUTE:api_print_niimbot] {e}'})


@magazynier_bp.route('/etykiety/niimbot/png/<int:product_id>')
def etykiety_niimbot_png(product_id):
    """Pobierz pojedynczy PNG etykiety dla Niimbot"""
    import io
    
    try:
        from .printer_manager import get_printer_manager as get_printer, ProductLabel, IMAGING_AVAILABLE
    except ImportError:
        return "Brak modułu printer_manager", 500
    
    if not IMAGING_AVAILABLE:
        return "Brak biblioteki Pillow", 500
    
    # Pobierz produkt
    conn = get_db()
    p = conn.execute('SELECT * FROM produkty WHERE id=?', (product_id,)).fetchone()

    if not p:
        return "Produkt nie znaleziony", 404

    p = dict(p)

    # Dane palety
    paleta_nazwa = ''
    koszt_szt = 0
    if p.get('paleta_id'):
        koszt_szt = _paleta_koszt_szt(conn, p['paleta_id'])
        pal_row = conn.execute('SELECT nazwa FROM palety WHERE id=?', (p['paleta_id'],)).fetchone()
        if pal_row:
            paleta_nazwa = pal_row['nazwa'] or ''

    # Generuj etykietę
    pm = get_printer()
    kod_mag = p.get('kod_magazynowy', '') or ''
    label = ProductLabel(
        nazwa=p['nazwa'][:35],
        qr_data=kod_mag if kod_mag else f"MAG:{p.get('ean') or p.get('asin') or p['id']}",
        lokalizacja=p.get('lokalizacja', ''),
        ean=p.get('ean', ''),
        ilosc=p.get('ilosc', 1),
        dostawca=p.get('dostawca', ''),
        data_zakupu=p.get('data_zakupu', '') or p.get('data_dodania', ''),
        paleta=paleta_nazwa,
        koszt_szt=koszt_szt,
        cena_allegro=0,
        kod_magazynowy=kod_mag,
        stan_przyjecia=_format_stan_label(p.get('stan_przyjecia', ''), p.get('klasa_jakosci', ''))
    )

    # Generuj obraz
    img = pm._generate_label_image(label)
    
    # Konwertuj do PNG
    img_buffer = io.BytesIO()
    from PIL import Image
    img_rgb = Image.new('RGB', img.size, 'white')
    img_rgb.paste(img.convert('L'))
    img_rgb.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Nazwa pliku
    filename = f"etykieta_{p.get('ean') or p.get('asin') or p['id']}.png"
    
    return Response(
        img_buffer.getvalue(),
        mimetype='image/png',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Cache-Control': 'no-cache'
        }
    )


STAN_DO_KLASY = {'Nowy': 'A', 'Jak nowy': 'A-', 'Powystawowy': 'A-', 'Uzywany': 'B',
                 'Uszkodzony': 'C', 'Zniszczony': 'D', 'Odnowiony': 'B'}


@magazynier_bp.route('/etykiety/sztuka/<int:sztuka_id>/png')
def etykieta_sztuka_png(sztuka_id):
    """Pobierz PNG etykiety dla konkretnej sztuki z jej klasą jakości."""
    import io
    conn = get_db()
    sztuka = conn.execute('SELECT * FROM sztuki WHERE id=?', (sztuka_id,)).fetchone()
    if not sztuka:
        return 'Nie znaleziono sztuki', 404
    sztuka = dict(sztuka)

    p = conn.execute('SELECT * FROM produkty WHERE id=?', (sztuka['produkt_id'],)).fetchone()
    if not p:
        return 'Nie znaleziono produktu', 404
    p = dict(p)

    paleta_nazwa = ''
    koszt_szt = 0
    if p.get('paleta_id'):
        koszt_szt = _paleta_koszt_szt(conn, p['paleta_id'])
        pal_row = conn.execute('SELECT nazwa FROM palety WHERE id=?', (p['paleta_id'],)).fetchone()
        if pal_row:
            paleta_nazwa = pal_row['nazwa'] or ''

    from .printer_manager import get_printer_manager as get_pm, ProductLabel as PL
    pm = get_pm()
    kod_mag = p.get('kod_magazynowy', '') or ''
    stan = sztuka.get('stan', '') or ''
    klasa = STAN_DO_KLASY.get(stan, '')
    stan_label = _format_stan_label(stan, klasa)

    label = PL(
        nazwa=p['nazwa'][:35],
        qr_data=f"{kod_mag}#{sztuka['numer']}" if kod_mag else f"MAG:{p.get('ean') or p['id']}#{sztuka['numer']}",
        ean=p.get('ean', '') or '',
        ilosc=1,
        paleta=paleta_nazwa,
        koszt_szt=koszt_szt,
        kod_magazynowy=kod_mag,
        stan_przyjecia=stan_label
    )
    img = pm._generate_label_image(label)
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    filename = f"etykieta_{kod_mag}_{sztuka['numer']}_{klasa or stan}.png"
    return Response(img_buffer.getvalue(), mimetype='image/png',
        headers={'Content-Disposition': f'attachment; filename={filename}', 'Cache-Control': 'no-cache'})


@magazynier_bp.route('/etykiety/produkt/<int:produkt_id>/klasy')
def etykiety_produkt_klasy(produkt_id):
    """Strona z etykietami pogrupowanymi po klasie — 1 PNG per unikalna klasa do pobrania."""
    conn = get_db()
    p = conn.execute('SELECT * FROM produkty WHERE id=?', (produkt_id,)).fetchone()
    if not p:
        return 'Nie znaleziono produktu', 404
    p = dict(p)

    sztuki = conn.execute('SELECT * FROM sztuki WHERE produkt_id=? ORDER BY numer', (produkt_id,)).fetchall()
    if not sztuki:
        return render('<div class="alert alert-err">Brak sztuk dla tego produktu</div><a href="/magazyn" class="back">← Powrót</a>')

    from collections import Counter
    grupy = Counter()
    for s in sztuki:
        s = dict(s)
        stan = s.get('stan', '') or ''
        klasa = STAN_DO_KLASY.get(stan, '')
        grupy[_format_stan_label(stan, klasa)] += 1

    klasa_colors = {'A': '#beee00', 'B': '#eab308', 'C': '#f97316', 'D': '#ef4444', 'A-': '#8ff5ff'}

    rows = ''
    for stan_label, ilosc in sorted(grupy.items()):
        klasa_czesc = stan_label.split('/')[0] if '/' in stan_label else stan_label
        color = klasa_colors.get(klasa_czesc, '#8ff5ff')
        safe = stan_label.replace('/', '-')
        rows += f'''
        <div style="display:flex;align-items:center;gap:16px;padding:14px;background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);margin-bottom:8px">
            <div style="font-size:1.4rem;font-weight:800;color:{color};font-family:'Space Grotesk',sans-serif;min-width:60px">{stan_label}</div>
            <div style="flex:1;color:#94a3b8">× <strong style="color:#e2e8f0;font-size:1.1rem">{ilosc}</strong> sztuk</div>
            <a href="/magazyn/etykiety/produkt/{produkt_id}/klasy/{safe}/png" download="etykieta_{safe}.png"
               style="display:flex;align-items:center;gap:6px;padding:8px 16px;background:{color}18;border:1px solid {color}44;color:{color};font-weight:600;font-size:0.85rem;text-decoration:none">
                <span class="material-symbols-outlined" style="font-size:1rem">download</span> Pobierz PNG
            </a>
        </div>'''

    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>label</span> ETYKIETY PO KLASIE</h1>
    <small>{p['nazwa'][:50]}</small></div>
    <div class="card" style="padding:16px;margin-bottom:12px;background:rgba(143,245,255,0.05);border:1px solid rgba(143,245,255,0.15)">
        <div style="color:#8ff5ff;font-size:0.85rem">Drukuj każdą etykietę <strong>N razy</strong> w aplikacji Niimbot odpowiadając liczbie sztuk danej klasy.</div>
    </div>
    {rows}
    <a href="/magazyn/produkt/{p.get('kod_magazynowy') or p['id']}" class="back">← Powrót do produktu</a>
    '''
    return render(html)


@magazynier_bp.route('/etykiety/produkt/<int:produkt_id>/klasy/<path:stan_label>/png')
def etykieta_klasa_png(produkt_id, stan_label):
    """PNG etykiety dla konkretnej klasy (z liczbą sztuk danej klasy)."""
    import io
    stan_label = stan_label.replace('-', '/', 1)  # przywróć / z -
    conn = get_db()
    p = conn.execute('SELECT * FROM produkty WHERE id=?', (produkt_id,)).fetchone()
    if not p:
        return 'Nie znaleziono produktu', 404
    p = dict(p)

    sztuki = conn.execute('SELECT * FROM sztuki WHERE produkt_id=?', (produkt_id,)).fetchall()
    from collections import Counter
    grupy = Counter()
    for s in sztuki:
        s = dict(s)
        st = s.get('stan', '') or ''
        kl = STAN_DO_KLASY.get(st, '')
        grupy[_format_stan_label(st, kl)] += 1

    ilosc = grupy.get(stan_label, 1)

    paleta_nazwa = ''
    koszt_szt = 0
    if p.get('paleta_id'):
        koszt_szt = _paleta_koszt_szt(conn, p['paleta_id'])
        pal_row = conn.execute('SELECT nazwa FROM palety WHERE id=?', (p['paleta_id'],)).fetchone()
        if pal_row:
            paleta_nazwa = pal_row['nazwa'] or ''

    from .printer_manager import get_printer_manager as get_pm, ProductLabel as PL
    pm = get_pm()
    kod_mag = p.get('kod_magazynowy', '') or ''
    label = PL(
        nazwa=p['nazwa'][:35],
        qr_data=kod_mag or f"MAG:{p.get('ean') or p['id']}",
        ean=p.get('ean', '') or '',
        ilosc=ilosc,
        paleta=paleta_nazwa,
        koszt_szt=koszt_szt,
        kod_magazynowy=kod_mag,
        stan_przyjecia=stan_label
    )
    img = pm._generate_label_image(label)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    safe = stan_label.replace('/', '-')
    return Response(buf.getvalue(), mimetype='image/png',
        headers={'Content-Disposition': f'attachment; filename=etykieta_{safe}.png', 'Cache-Control': 'no-cache'})


@magazynier_bp.route('/etykiety/niimbot/zip')
def etykiety_niimbot_zip():
    """Pobierz ZIP z etykietami PNG dla Niimbot"""
    import zipfile
    import io
    
    try:
        from .printer_manager import get_printer_manager as get_printer, ProductLabel, IMAGING_AVAILABLE
    except ImportError:
        return "Brak modułu printer_manager", 500
    
    if not IMAGING_AVAILABLE:
        return "Brak biblioteki Pillow - zainstaluj: pip install pillow qrcode", 500
    
    ids_str = request.args.get('ids', '')
    if not ids_str:
        return redirect('/magazyn/etykiety')
    
    ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    
    if not ids:
        return redirect('/magazyn/etykiety')
    
    # Pobierz produkty
    conn = get_db()
    products = []
    for pid in ids:
        p = conn.execute('SELECT * FROM produkty WHERE id=?', (pid,)).fetchone()
        if p:
            products.append(dict(p))
    
    if not products:
        return "Brak produktów", 404
    
    # Generuj ZIP z PNG
    pm = get_printer()
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, p in enumerate(products):
            kod_mag = p.get('kod_magazynowy', '') or ''
            label = ProductLabel(
                nazwa=p['nazwa'][:35],
                qr_data=kod_mag if kod_mag else f"MAG:{p.get('ean') or p.get('asin') or p['id']}",
                lokalizacja=p.get('lokalizacja', ''),
                ean=p.get('ean', ''),
                ilosc=p.get('ilosc', 1),
                dostawca=p.get('dostawca', ''),
                data_zakupu=p.get('data_zakupu', '') or p.get('data_dodania', ''),
                kod_magazynowy=kod_mag,
                cena_allegro=0,
                stan_przyjecia=_format_stan_label(p.get('stan_przyjecia', ''), p.get('klasa_jakosci', ''))
            )
            
            # Generuj obraz
            img = pm._generate_label_image(label)
            
            # Konwertuj do PNG
            img_buffer = io.BytesIO()
            # Konwertuj 1-bit do RGB dla lepszego PNG
            from PIL import Image
            img_rgb = Image.new('RGB', img.size, 'white')
            img_rgb.paste(img.convert('L'))
            img_rgb.save(img_buffer, format='PNG')
            
            # Nazwa pliku
            safe_name = ''.join(c if c.isalnum() else '_' for c in p['nazwa'][:20])
            filename = f"etykieta_{i+1:03d}_{safe_name}.png"
            
            zf.writestr(filename, img_buffer.getvalue())
    
    zip_buffer.seek(0)
    
    return Response(
        zip_buffer.getvalue(),
        mimetype='application/zip',
        headers={'Content-Disposition': f'attachment; filename=etykiety_niimbot_{len(products)}szt.zip'}
    )


@magazynier_bp.route('/qr/<int:product_id>')
def qr_product_view(product_id):
    """Strona wyświetlana po zeskanowaniu QR kodu"""
    conn = get_db()
    product = conn.execute('SELECT * FROM produkty WHERE id=?', (product_id,)).fetchone()
    
    if not product:
        return render('''
            <div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> NIE ZNALEZIONO</h1></div>
            <div class="alert alert-err">Produkt nie istnieje w bazie</div>
            <a href="/magazyn" class="back">← Magazyn</a>
        ''')
    
    p = dict(product)

    # Pobierz nazwę palety z tabeli palety
    _paleta_nazwa = ''
    if p.get('paleta_id'):
        _pal_row = conn.execute('SELECT nazwa FROM palety WHERE id = ?', (p['paleta_id'],)).fetchone()
        if _pal_row:
            _paleta_nazwa = _pal_row['nazwa']
    if not _paleta_nazwa:
        _paleta_nazwa = p.get('paleta', '') or ''

    # Sprawdź czy produkt jest sprzedany (sprawdź w Allegro - TODO)
    sprzedany = (p.get('status') or 'nowy') == 'sprzedany'
    
    # Status badge
    if sprzedany:
        status_html = '<span class="badge" style="background:#beee00"><span class=material-symbols-outlined>check_circle</span> SPRZEDANY</span>'
        action_html = '''
            <div class="alert alert-ok" style="margin-bottom:15px">
                <b>Zamówienie do wysyłki!</b><br>
                <span id="buyerInfo">Ładowanie danych kupującego...</span>
            </div>
            <a href="#" class="btn btn-ok"><span class=material-symbols-outlined>inventory_2</span> OZNACZ JAKO WYSŁANE</a>
            <a href="#" class="btn btn-2"><span class=material-symbols-outlined>print</span> DRUKUJ ETYKIETĘ INPOST</a>
        '''
    else:
        status_html = '<span class="badge" style="background:#8ff5ff"><span class=material-symbols-outlined>inventory_2</span> W MAGAZYNIE</span>'
        action_html = f'''
            <a href="/magazyn/produkt/{p['id']}/edit" class="btn btn-2"><span class=material-symbols-outlined>edit</span> EDYTUJ</a>
            <a href="/paletomat/generator/from-magazyn/{p['id']}" class="btn btn-p"><span class=material-symbols-outlined>shopping_cart</span> WYSTAW NA ALLEGRO</a>
        '''
    
    html = f'''
    <div class="hdr">
        <h1><span class=material-symbols-outlined>inventory_2</span> PRODUKT</h1>
        {status_html}
    </div>
    
    <div class="card">
        <div class="card-body">
            <div class="card-name">{p['nazwa']}</div>
            
            <div class="loc">
                <div class="loc-title"><span class=material-symbols-outlined>pin_drop</span> LOKALIZACJA</div>
                <div class="loc-grid">
                    <div><div class="loc-v">{p.get('regal', '—') or '—'}</div><div class="loc-l">Regał</div></div>
                    <div><div class="loc-v">{_paleta_nazwa or '—'}</div><div class="loc-l">Paleta</div></div>
                    <div><div class="loc-v dostawca-name">{p.get('dostawca', '—') or '—'}</div><div class="loc-l">Dostawca</div></div>
                </div>
            </div>
            
            <div class="det-grid">
                <div class="det"><div class="det-l">Cena Allegro</div><div class="det-v green">{p['cena_allegro']:.2f} zł</div></div>
                <div class="det"><div class="det-l">Ilość</div><div class="det-v">{p['ilosc']} szt</div></div>
                <div class="det"><div class="det-l">EAN</div><div class="det-v">{p.get('ean', '—') or '—'}</div></div>
                <div class="det"><div class="det-l">Dostawca</div><div class="det-v dostawca-name">{p.get('dostawca', '—') or '—'}</div></div>
            </div>
        </div>
    </div>
    
    {action_html}
    
    <a href="/magazyn" class="back">← Powrót</a>
    </div>
    '''
    return render(html)

@magazynier_bp.route('/historia/<int:historia_id>/edytuj', methods=['GET', 'POST'])
def edytuj_historie(historia_id):
    """Edycja wpisu w historii produktu"""
    conn = get_db()
    
    if request.method == 'POST':
        opis = request.form.get('opis', '').strip()
        akcja = request.form.get('akcja', 'edytowano').strip()
        
        if opis:
            conn.execute('''
                UPDATE historia_produktu 
                SET opis = ?, akcja = ?
                WHERE id = ?
            ''', (opis, akcja, historia_id))
            conn.commit()
        
        # Pobierz produkt_id aby przekierować z powrotem
        h = conn.execute('SELECT produkt_id FROM historia_produktu WHERE id = ?', (historia_id,)).fetchone()
        
        if h:
            p = get_db().execute('SELECT * FROM produkty WHERE id = ?', (h['produkt_id'],)).fetchone()
            if p:
                product_code = get_product_code(p)
                return redirect(f'/magazyn/produkt/{product_code}?msg=Zaktualizowano+wpis+historii')
        
        return redirect('/magazyn')
    
    # GET - formularz edycji
    h = conn.execute('SELECT * FROM historia_produktu WHERE id = ?', (historia_id,)).fetchone()
    if not h:
        return render('<div class="hdr"><h1><span class=material-symbols-outlined>cancel</span> BŁĄD</h1></div><div class="alert alert-err">Wpis nie istnieje</div><a href="/magazyn" class="back">← Powrót</a>')
    
    h = dict(h)
    
    # Dostępne akcje
    akcje = [
        ('dodano', '<span class=material-symbols-outlined>download</span> Dodano'),
        ('edytowano', '<span class=material-symbols-outlined>edit</span> Edytowano'),
        ('wystawiono', '<span class=material-symbols-outlined>shopping_cart</span> Wystawiono'),
        ('sprzedano', '<span class=material-symbols-outlined>paid</span> Sprzedano'),
        ('wyslano', '<span class=material-symbols-outlined>inventory_2</span> Wysłano'),
        ('zmiana_ceny', '<span class=material-symbols-outlined>payments</span> Zmiana ceny'),
        ('zmiana_lokalizacji', '<span class=material-symbols-outlined>pin_drop</span> Zmiana lokalizacji'),
        ('zmiana_ilosci', '<span class=material-symbols-outlined>bar_chart</span> Zmiana ilości'),
        ('drukowano', '<span class=material-symbols-outlined>label</span> Drukowano'),
        ('skanowano', '[SMARTPHONE] Skanowano'),
        ('importowano', '[FOLDER] Importowano'),
        ('scrapowano', '<span class=material-symbols-outlined>search</span> Scrapowano'),
        ('wygenerowano_opis', '<span class=material-symbols-outlined>auto_awesome</span> Wygenerowano opis'),
        ('dodano_zdjecia', '<span class=material-symbols-outlined>photo_camera</span> Dodano zdjęcia'),
        ('przeniesiono', '<span class=material-symbols-outlined>sync</span> Przeniesiono'),
        ('oznaczono', '<span class=material-symbols-outlined>label</span> Oznaczono')
    ]
    
    akcje_options = ''.join([f'<option value="{a[0]}" {"selected" if h["akcja"] == a[0] else ""}>{a[1]}</option>' for a in akcje])
    
    
    html = f'''
    <div class="hdr"><h1><span class=material-symbols-outlined>edit</span> EDYCJA WPISU HISTORII</h1></div>
    
    <form method="POST" class="card">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <div class="form-g">
            <label>Typ akcji</label>
            <select name="akcja" class="form-input">
                {akcje_options}
            </select>
        </div>
        
        <div class="form-g">
            <label>Opis</label>
            <textarea name="opis" class="form-input" rows="3" required>{h['opis']}</textarea>
        </div>
        
        <div style="display:flex;gap:10px;margin-top:20px">
            <button type="submit" class="btn btn-ok"><span class=material-symbols-outlined>save</span> Zapisz</button>
            <a href="/magazyn" class="btn btn-p">✖ Anuluj</a>
        </div>
    </form>
    '''
    return render(html)

@magazynier_bp.route('/historia/<int:historia_id>/usun')
def usun_historie(historia_id):
    """Usunięcie wpisu z historii produktu"""
    conn = get_db()
    
    # Pobierz produkt_id przed usunięciem
    h = conn.execute('SELECT produkt_id FROM historia_produktu WHERE id = ?', (historia_id,)).fetchone()
    
    # Usuń wpis
    conn.execute('DELETE FROM historia_produktu WHERE id = ?', (historia_id,))
    conn.commit()
    
    # Przekieruj z powrotem do produktu
    redirect_url = request.args.get('redirect', '/magazyn')
    if redirect_url.startswith('/magazyn/produkt/'):
        redirect_url += '?msg=Usunięto+wpis+z+historii'
    
    return redirect(redirect_url)

@magazynier_bp.route('/produkty/masowa-edycja', methods=['POST'])
def masowa_edycja_statusu():
    """Masowa zmiana statusu produktów"""
    from .database import add_historia
    
    product_ids = request.form.getlist('product_ids')
    new_status = request.form.get('new_status', '').strip()
    new_stan = request.form.get('new_stan', '').strip()
    new_lokalizacja = request.form.get('new_lokalizacja', '').strip()
    new_cena_str = request.form.get('new_cena_allegro', '').strip()
    new_cena_allegro = float(new_cena_str) if new_cena_str else None

    if not product_ids:
        return redirect('/magazyn/produkty?msg=Nie+zaznaczono+produktów')

    if not new_status and not new_stan and not new_lokalizacja and new_cena_allegro is None:
        return redirect('/magazyn/produkty?msg=Nie+wybrano+żadnej+zmiany')

    # Nazwy statusów dla logów
    status_names = {
        'magazyn': 'Magazyn',
        'wystawiony': 'Wystawiony (na Allegro)',
        'sprzedany': 'Sprzedany',
        'uszkodzony': 'Uszkodzony',
        'zwrot': 'Zwrot'
    }
    
    conn = get_db()
    updated_count = 0
    
    try:
        for product_id in product_ids:
            # Pobierz aktualny produkt
            p = conn.execute('SELECT * FROM produkty WHERE id = ?', (product_id,)).fetchone()
            if not p:
                continue
            
            old_status = p['status'] or 'magazyn'

            # Buduj dynamicznie pola do aktualizacji
            updates = []
            vals = []
            changes = []

            if new_status:
                updates.append('status = ?')
                vals.append(new_status)
                changes.append(f"status: {status_names.get(old_status, old_status)} → {status_names.get(new_status, new_status)}")
            if new_stan:
                updates.append('stan = ?')
                vals.append(new_stan)
                changes.append(f"stan: {p['stan'] or '—'} → {new_stan}")
            if new_lokalizacja:
                updates.append('lokalizacja = ?')
                vals.append(new_lokalizacja)
                changes.append(f"lokalizacja: {p['lokalizacja'] or '—'} → {new_lokalizacja}")
            if new_cena_allegro is not None:
                updates.append('cena_allegro = ?')
                vals.append(new_cena_allegro)
                changes.append(f"cena: {p['cena_allegro'] or 0:.0f} → {new_cena_allegro:.0f} zł")

            if updates:
                # updates contains only whitelisted 'column = ?' fragments (status, stan, lokalizacja, cena_allegro)
                ALLOWED_SET_CLAUSES = {'status = ?', 'stan = ?', 'lokalizacja = ?', 'cena_allegro = ?'}
                if not all(u in ALLOWED_SET_CLAUSES for u in updates):
                    return jsonify({'error': 'Niedozwolone pole'}), 400
                vals.append(product_id)
                conn.execute("UPDATE produkty SET " + ', '.join(updates) + " WHERE id = ?", vals)
                opis = "Masowa edycja: " + " | ".join(changes)
                add_historia(product_id, 'edytowano', opis, {'zmiany': changes})
                updated_count += 1

        conn.commit()

        return redirect(f'/magazyn/produkty?msg=Zaktualizowano+{updated_count}+produktów')
        
    except Exception as e:
        return redirect(f'/magazyn/produkty?msg=Błąd:+{str(e)[:50]}')


# ======================== AUTO-WYCENA ========================

_nbp_cache = {}  # {waluta: (kurs, timestamp)}

def _get_nbp_rate(currency_code):
    """Pobiera aktualny kurs średni NBP. Cache na 6h."""
    import time as _time
    import requests as _req

    now = _time.time()
    cached = _nbp_cache.get(currency_code)
    if cached and (now - cached[1]) < 6 * 3600:
        return cached[0]

    # Fallbacki gdyby API nie odpowiedziało
    _fallback = {'EUR': 4.3, 'USD': 4.0, 'GBP': 5.1}

    try:
        resp = _req.get(
            f'https://api.nbp.pl/api/exchangerates/rates/a/{currency_code}/?format=json',
            timeout=5
        )
        if resp.status_code == 200:
            rate = resp.json()['rates'][0]['mid']
            _nbp_cache[currency_code] = (rate, now)
            print(f"[NBP] Kurs {currency_code}: {rate} PLN")
            return rate
    except Exception as e:
        print(f"[NBP] Błąd pobierania kursu {currency_code}: {e}")

    fallback = _fallback.get(currency_code, 4.3)
    _nbp_cache[currency_code] = (fallback, now)
    return fallback


def _amazon_price_to_pln(price, domain):
    """Konwertuje cenę z Amazon na PLN wg domeny — kurs z NBP"""
    if not price or price <= 0:
        return 0

    _domain_currency = {
        'amazon.pl': None,       # Już PLN
        'amazon.de': 'EUR',
        'amazon.fr': 'EUR',
        'amazon.it': 'EUR',
        'amazon.es': 'EUR',
        'amazon.com': 'USD',
        'amazon.co.uk': 'GBP',
    }

    currency = _domain_currency.get(domain)
    if currency is None:
        return price  # PLN lub nieznana domena — bez konwersji

    rate = _get_nbp_rate(currency)
    return price * rate


@magazynier_bp.route('/api/autowycena/<int:product_id>', methods=['POST'])
def api_autowycena(product_id):
    """Pobiera cenę z Amazon i sugeruje cenę Allegro"""
    from .utils import scrape_amazon_product
    import json
    
    conn = get_db()
    p = conn.execute('SELECT * FROM produkty WHERE id = ?', (product_id,)).fetchone()

    if not p:
        return json.dumps({'error': 'Produkt nie znaleziony'}), 404

    asin = p['asin']
    # Koszt jednostkowy = własna cena produktu (jednostkowa z importu), fallback na średnią z palety
    cena_brutto_szt = float(p['cena_brutto'] or 0) if p['cena_brutto'] and p['cena_brutto'] > 0 else 0
    if cena_brutto_szt == 0 and p['paleta_id']:
        cena_brutto_szt = _paleta_koszt_szt(conn, p['paleta_id'])

    result = {
        'product_id': product_id,
        'asin': asin,
        'cena_brutto_szt': round(cena_brutto_szt, 2),
        'cena_amazon': None,
        'cena_allegro_sugerowana': None,
        'zrodlo': None
    }
    
    # Próbuj pobrać cenę z Amazon
    if asin:
        try:
            amazon_data = scrape_amazon_product(asin)
            if amazon_data and amazon_data.get('price'):
                cena_amazon = amazon_data['price']
                result['cena_amazon'] = round(cena_amazon, 2)

                # Konwersja wg domeny Amazon
                cena_pln = _amazon_price_to_pln(cena_amazon, amazon_data.get('domain'))

                # Sugerowana cena Allegro = 85% ceny w PLN
                result['cena_allegro_sugerowana'] = round(cena_pln * 0.85, 2)
                result['zrodlo'] = f"amazon ({amazon_data.get('domain', '?')})"
        except Exception as e:
            print(f"[Auto-wycena] Błąd pobierania z Amazon: {e}")
    
    # Fallback: brutto × 2.5
    if not result['cena_allegro_sugerowana']:
        if cena_brutto_szt > 0:
            result['cena_allegro_sugerowana'] = round(cena_brutto_szt * 2.5, 2)
            result['zrodlo'] = 'szacowana (brutto × 2.5)'
        else:
            result['zrodlo'] = 'brak danych'
    
    
    return json.dumps(result), 200, {'Content-Type': 'application/json'}


@magazynier_bp.route('/api/autowycena/zastosuj/<int:product_id>', methods=['POST'])
def api_autowycena_zastosuj(product_id):
    """Zapisuje sugerowaną cenę jako cena_allegro"""
    import json
    
    data = request.get_json() or {}
    nowa_cena = data.get('cena')
    
    if not nowa_cena or nowa_cena <= 0:
        return json.dumps({'error': 'Brak ceny'}), 400
    
    conn = get_db()
    p = conn.execute('SELECT cena_allegro FROM produkty WHERE id = ?', (product_id,)).fetchone()
    
    if not p:
        return json.dumps({'error': 'Produkt nie znaleziony'}), 404
    
    stara_cena = p['cena_allegro'] or 0
    
    conn.execute('UPDATE produkty SET cena_allegro = ? WHERE id = ?', (nowa_cena, product_id))
    
    # Dodaj do historii
    add_historia(product_id, 'zmiana_ceny', 
        f'Auto-wycena: {stara_cena:.2f} → {nowa_cena:.2f} zł',
        {'stara_cena': stara_cena, 'nowa_cena': nowa_cena, 'metoda': 'autowycena'})
    
    conn.commit()
    
    return json.dumps({'success': True, 'nowa_cena': nowa_cena}), 200, {'Content-Type': 'application/json'}


@magazynier_bp.route('/api/autowycena/paleta/<int:paleta_id>', methods=['POST'])
def api_autowycena_paleta(paleta_id):
    """Auto-wycena (legacy, redirect do stream)"""
    return api_autowycena_paleta_stream(paleta_id)


@magazynier_bp.route('/api/autowycena-stream/paleta/<int:paleta_id>', methods=['POST'])
def api_autowycena_paleta_stream(paleta_id):
    """Auto-wycena STREAMOWANA — Gemini AI batch wycena"""
    from .utils import optimize_title_allegro
    import json
    import time
    import re
    import requests as _req

    def generate():
        conn = get_db()
        from modules.database import get_config as _get_cfg
        gemini_key = _get_cfg('gemini_api_key', '')
        _wycena_model = _get_cfg('ai_model_wycena', _get_cfg('gemini_model', 'gemini-2.5-flash'))

        produkty = conn.execute(
            'SELECT id, asin, nazwa, ilosc, cena_brutto, cena_allegro, paleta_id FROM produkty WHERE paleta_id = ?',
            (paleta_id,)
        ).fetchall()

        total = len(produkty)
        stats = {
            'type': 'done', 'total': total, 'updated': 0,
            'from_amazon': 0, 'from_estimate': 0,
            'titles_optimized': 0, 'errors': 0
        }

        # Batch Gemini AI — po 15 produktów
        BATCH_SIZE = 15
        batches = [produkty[i:i+BATCH_SIZE] for i in range(0, len(produkty), BATCH_SIZE)]
        processed = 0

        for batch_idx, batch in enumerate(batches):
            # Zbuduj listę produktów do wyceny
            prod_list = []
            for p in batch:
                nazwa = p['nazwa'] or f'Produkt #{p["id"]}'
                prod_list.append(f'{p["id"]}. {nazwa} (szt: {p["ilosc"] or 1})')

            prompt = f"""Jesteś ekspertem sprzedaży na Allegro.pl w Polsce. Dla KAŻDEGO produktu z listy podaj cenę sprzedaży na Allegro w PLN.

WAŻNE:
- Podaj cenę za jaką ten produkt AKTUALNIE się sprzedaje na Allegro.pl (nie zaniżaj!)
- Produkty są NOWE, oryginalne, z Amazon/hurtowni
- Uwzględnij że na Allegro ceny są często WYŻSZE niż na Amazon (bo Allegro to polski rynek)
- Nie zaniżaj cen — sprzedawcy na Allegro mają marżę 30-100% ponad cenę zakupu
- MUSISZ podać cenę dla KAŻDEGO produktu z listy (wszystkie {len(prod_list)} pozycji)

Produkty:
{chr(10).join(prod_list)}

Odpowiedz DOKŁADNIE w tym formacie, po jednej linii na produkt, BEZ dodatkowego tekstu:
ID:CENA
Przykład:
123:299
456:89.99"""

            ai_prices = {}

            if gemini_key:
                try:
                    _wycena_schema = {
                        "type": "ARRAY",
                        "items": {
                            "type": "OBJECT",
                            "properties": {
                                "id": {"type": "INTEGER"},
                                "cena": {"type": "NUMBER"}
                            },
                            "required": ["id", "cena"]
                        }
                    }
                    api_url = f'https://generativelanguage.googleapis.com/v1beta/models/{_wycena_model}:generateContent?key={gemini_key}'
                    resp = _req.post(
                        api_url,
                        json={
                            'contents': [{'parts': [{'text': prompt}]}],
                            'generationConfig': {
                                'response_mime_type': 'application/json',
                                'response_schema': _wycena_schema
                            }
                        },
                        timeout=30
                    )
                    if resp.status_code == 200:
                        ai_text = resp.json().get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                        print(f"[Auto-wycena] Gemini batch {batch_idx+1} response:\n{ai_text}")
                        try:
                            items = __import__('json').loads(ai_text)
                            for item in items:
                                pid = int(item.get('id', 0))
                                price = float(item.get('cena', 0))
                                if pid and 1 < price < 50000:
                                    ai_prices[pid] = round(price, 2)
                        except Exception as _pe:
                            print(f"[Auto-wycena] JSON parse fallback: {_pe}")
                            for line in ai_text.strip().split('\n'):
                                line = line.strip()
                                m = re.match(r'(\d+)\s*[:\-]\s*(\d+(?:[.,]\d+)?)', line)
                                if m:
                                    pid = int(m.group(1))
                                    price = float(m.group(2).replace(',', '.'))
                                    if 1 < price < 50000:
                                        ai_prices[pid] = round(price, 2)
                    else:
                        print(f"[Auto-wycena] Gemini error {resp.status_code}: {resp.text[:200]}")
                except Exception as e:
                    print(f"[Auto-wycena] Gemini batch error: {e}")
                    stats['errors'] += 1

            print(f"[Auto-wycena] Batch {batch_idx+1}: ai_prices = {ai_prices}")

            # Zapisz wyniki
            for p in batch:
                processed += 1
                nazwa = p['nazwa'] or ''
                cena_allegro = ai_prices.get(p['id'])
                zrodlo = 'gemini' if cena_allegro else None

                if cena_allegro:
                    stats['from_estimate'] += 1

                # Zapisz cenę do DB (bez optymalizacji tytułu — za wolne)
                if cena_allegro:
                    try:
                        conn.execute("UPDATE produkty SET cena_allegro = ? WHERE id = ?", (cena_allegro, p['id']))
                        conn.commit()
                        stats['updated'] += 1
                        print(f"[Auto-wycena] Zapisano {p['id']}: {cena_allegro} zł")
                    except Exception as e:
                        print(f"[Auto-wycena] Błąd zapisu {p['id']}: {e}")
                        stats['errors'] += 1

                # Wyślij progress
                ev = {
                    'type': 'progress',
                    'current': processed,
                    'total': total,
                    'name': (nazwa or f'Produkt #{p["id"]}')[:50],
                    'price': cena_allegro,
                    'source': zrodlo
                }
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

        # Podsumowanie
        yield f"data: {json.dumps(stats, ensure_ascii=False)}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        }
    )


# ======================== GPSR API ========================

@magazynier_bp.route('/api/rescrape-image/<int:product_id>', methods=['POST'])
def api_rescrape_image(product_id):
    """Rescrape zdjęcia dla produktu z Amazona"""
    from modules.database import get_db
    from modules.utils import scrape_amazon_product
    import os, requests as _req

    conn = get_db()
    p = conn.execute('SELECT id, asin, nazwa, zdjecie_url FROM produkty WHERE id = ?', (product_id,)).fetchone()
    if not p:
        return jsonify({'ok': False, 'error': 'Nie znaleziono produktu'})

    asin = (p['asin'] or '').strip().upper()
    if not asin or len(asin) < 10:
        return jsonify({'ok': False, 'error': 'Brak ASIN — nie mozna pobrac zdjecia'})

    # Scrape w tle — odpowiedz natychmiast (ngrok timeout = 60s, scraping = 30-120s)
    import threading

    def _bg_scrape(pid, asin_code):
        try:
            from modules.database import get_db as _gdb
            from modules.utils import scrape_amazon_product as _scrape
            import requests as _rq

            data = _scrape(asin_code)
            if not data or not data.get('image_url'):
                print(f'[PHOTO_CAMERA] Scrape {asin_code}: brak zdjec')
                return

            all_images = data.get('all_images', []) or [data['image_url']]
            asin_dir = os.path.join('static', 'downloads', asin_code)
            os.makedirs(asin_dir, exist_ok=True)

            _headers = {'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.google.com/'}
            local_img = ''
            for idx, url in enumerate(all_images[:5], 1):
                try:
                    resp = _rq.get(url, headers=_headers, timeout=10)
                    if resp.status_code == 200 and len(resp.content) > 500:
                        path = os.path.join(asin_dir, f'image_{idx}.jpg')
                        with open(path, 'wb') as f:
                            f.write(resp.content)
                        if idx == 1:
                            local_img = '/' + path.replace('\\', '/')
                except:
                    pass

            final_url = local_img or data['image_url']
            _conn = _gdb()
            _conn.execute('UPDATE produkty SET zdjecie_url = ? WHERE id = ?', (final_url, pid))
            _conn.commit()
            print(f'[PHOTO_CAMERA] Scrape OK: {asin_code} -> {final_url}')
        except Exception as e:
            print(f'[PHOTO_CAMERA] Scrape error: {e}')

    threading.Thread(target=_bg_scrape, args=(product_id, asin), daemon=True).start()
    return jsonify({'ok': True, 'img': '', 'note': 'Pobieranie w tle — odśwież za ~30s'})


@magazynier_bp.route('/api/utworz-box', methods=['POST'])
def api_utworz_box():
    """Utwórz nowy box z zaznaczonych produktów"""
    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])
        nazwa = data.get('nazwa', '').strip()
        cena_zakupu = float(data.get('cena_zakupu', 0))
        cena_sprzedazy = float(data.get('cena_sprzedazy', 0))

        if not product_ids or len(product_ids) < 2:
            return jsonify({'ok': False, 'error': 'Zaznacz minimum 2 produkty'})
        if not nazwa:
            return jsonify({'ok': False, 'error': 'Podaj nazwę boxa'})
        if cena_zakupu <= 0:
            return jsonify({'ok': False, 'error': 'Podaj cenę zakupu'})

        conn = get_db()

        # Policz łączną ilość sztuk
        placeholders = ','.join(['?' for _ in product_ids])
        rows = conn.execute(
            f'SELECT id, ilosc, dostawca FROM produkty WHERE id IN ({placeholders})',
            product_ids
        ).fetchall()

        if not rows:
            return jsonify({'ok': False, 'error': 'Nie znaleziono produktów'})

        ilosc_produktow = len(rows)
        ilosc_sztuk = sum(r['ilosc'] or 1 for r in rows)
        dostawca = rows[0]['dostawca'] or ''

        # Utwórz nową paletę typu 'box'
        from datetime import datetime
        conn.execute('''
            INSERT INTO palety (nazwa, dostawca, cena_zakupu, ilosc_produktow, ilosc_sztuk,
                               data_zakupu, data_dodania, typ, cena_zakupu_netto)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'box', ?)
        ''', (
            nazwa, dostawca, cena_zakupu, ilosc_produktow, ilosc_sztuk,
            datetime.now().strftime('%Y-%m-%d'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            round(cena_zakupu / 1.23, 2)
        ))
        box_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        # Przenieś produkty do boxa
        conn.execute(
            f'UPDATE produkty SET paleta_id = ?, paleta = ? WHERE id IN ({placeholders})',
            [box_id, nazwa] + product_ids
        )

        # Ustaw cenę allegro jeśli podana
        if cena_sprzedazy > 0:
            conn.execute(
                f'UPDATE produkty SET cena_allegro = ? WHERE id IN ({placeholders})',
                [cena_sprzedazy] + product_ids
            )

        conn.commit()
        return jsonify({'ok': True, 'box_id': box_id})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:200]})


@magazynier_bp.route('/api/zgrupuj-palety-box', methods=['POST'])
def api_zgrupuj_palety_box():
    """Zgrupuj produkty z kilku palet w jeden nowy box"""
    try:
        data = request.get_json()
        paleta_ids = [int(x) for x in data.get('paleta_ids', [])]
        nazwa = data.get('nazwa', '').strip()
        cena_sprzedazy = float(data.get('cena_sprzedazy', 0) or 0)
        cena_zakupu_manual = float(data.get('cena_zakupu', 0) or 0)

        if not paleta_ids:
            return jsonify({'ok': False, 'error': 'Brak zaznaczonych palet'})
        if not nazwa:
            return jsonify({'ok': False, 'error': 'Podaj nazwę boxa'})

        conn = get_db()
        from datetime import datetime

        # Cena zakupu: ręczna jeśli podana, lub suma z palet
        placeholders = ','.join(['?' for _ in paleta_ids])
        if cena_zakupu_manual > 0:
            total_cena = cena_zakupu_manual
        else:
            total_cena = conn.execute(
                f'SELECT COALESCE(SUM(cena_zakupu), 0) FROM palety WHERE id IN ({placeholders})',
                paleta_ids
            ).fetchone()[0]

        # Policz produkty
        total_prod = conn.execute(
            f'SELECT COUNT(*) FROM produkty WHERE paleta_id IN ({placeholders})',
            paleta_ids
        ).fetchone()[0]
        total_szt = conn.execute(
            f'SELECT COALESCE(SUM(ilosc), 0) FROM produkty WHERE paleta_id IN ({placeholders})',
            paleta_ids
        ).fetchone()[0]

        if total_prod == 0:
            return jsonify({'ok': False, 'error': 'Brak produktów w zaznaczonych paletach'})

        # Pobierz dostawcę z pierwszej palety
        first = conn.execute('SELECT dostawca FROM palety WHERE id = ?', (paleta_ids[0],)).fetchone()
        dostawca = first['dostawca'] if first else ''

        # Utwórz nowy box
        conn.execute('''
            INSERT INTO palety (nazwa, dostawca, cena_zakupu, cena_zakupu_netto, ilosc_produktow, ilosc_sztuk,
                               data_zakupu, data_dodania, typ)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'box')
        ''', (
            nazwa, dostawca, total_cena, round(total_cena / 1.23, 2),
            total_prod, total_szt,
            datetime.now().strftime('%Y-%m-%d'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))
        box_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        # Przenieś produkty do boxa
        conn.execute(
            f'UPDATE produkty SET paleta_id = ?, paleta = ? WHERE paleta_id IN ({placeholders})',
            [box_id, nazwa] + paleta_ids
        )

        # Ustaw cenę allegro jeśli podana
        if cena_sprzedazy > 0:
            conn.execute(
                f'UPDATE produkty SET cena_allegro = ? WHERE paleta_id = ?',
                (cena_sprzedazy, box_id)
            )

        # Usuń stare puste palety (produkty już przeniesione do boxa)
        conn.execute(
            f'DELETE FROM palety WHERE id IN ({placeholders})',
            paleta_ids
        )

        conn.commit()
        return jsonify({'ok': True, 'box_id': box_id, 'produktow': total_prod, 'sztuk': total_szt})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:200]})


@magazynier_bp.route('/api/check-image/<int:product_id>')
def api_check_image(product_id):
    """Sprawdza czy produkt ma lokalne zdjęcie"""
    import os
    conn = get_db()
    p = conn.execute('SELECT zdjecie_url FROM produkty WHERE id = ?', (product_id,)).fetchone()
    if not p:
        return jsonify({'has_image': False})
    url = p['zdjecie_url'] or ''
    if url.startswith('/static/downloads/'):
        return jsonify({'has_image': os.path.exists(url.lstrip('/'))})
    if url.startswith('http') and 'media-amazon' in url:
        return jsonify({'has_image': True})
    return jsonify({'has_image': bool(url)})


@magazynier_bp.route('/api/gpsr/<int:product_id>')
def api_gpsr(product_id):
    """Generuje informacje GPSR dla produktu"""
    from .utils import generuj_gpsr_info
    import json
    
    conn = get_db()
    p = conn.execute('SELECT nazwa, kategoria FROM produkty WHERE id = ?', (product_id,)).fetchone()
    
    if not p:
        return json.dumps({'error': 'Produkt nie znaleziony'}), 404
    
    gpsr = generuj_gpsr_info(p['nazwa'] or '', p['kategoria'] or '')
    
    return json.dumps({'gpsr': gpsr, 'nazwa': p['nazwa']}), 200, {'Content-Type': 'application/json'}


# ======================== RAPORT SPRZEDAŻY EXCEL ========================

@magazynier_bp.route('/raport-sprzedazy')
def raport_sprzedazy_excel():
    """Generuje raport sprzedaży Excel z podziałem na miesiące"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import os
    
    conn = get_db()

    # Pobierz wszystkie sprzedaże z prawdziwym kosztem z palety
    # Koszt = paleta.cena_zakupu / MAX(ile_sprzedazy_z_palety, ile_produktow_w_palecie)
    sprzedaze = conn.execute('''
        SELECT
            s.id,
            s.data_sprzedazy,
            s.nazwa,
            s.cena,
            s.ilosc,
            s.status,
            p.asin,
            p.ean,
            p.paleta,
            p.dostawca,
            p.kategoria,
            pal.cena_zakupu as paleta_cena_zakupu,
            CASE
                WHEN sc.sale_cnt > 0 AND pal.cena_zakupu > 0
                THEN pal.cena_zakupu / sc.sale_cnt
                ELSE 0
            END as koszt_jednostkowy
        FROM sprzedaze s
        LEFT JOIN produkty p ON s.produkt_id = p.id
        LEFT JOIN palety pal ON p.paleta_id = pal.id
        LEFT JOIN (
            SELECT pal2.id as paleta_id,
                   CASE
                       WHEN COALESCE(SUM(s2.ilosc), 0) > pal2.ilosc_produktow
                       THEN COALESCE(SUM(s2.ilosc), 0)
                       ELSE pal2.ilosc_produktow
                   END as sale_cnt
            FROM palety pal2
            LEFT JOIN produkty p3 ON p3.paleta_id = pal2.id
            LEFT JOIN sprzedaze s2 ON s2.produkt_id = p3.id
                AND s2.status NOT IN ('anulowana', 'zwrot')
            GROUP BY pal2.id
        ) sc ON pal.id = sc.paleta_id
        WHERE s.status NOT IN ('anulowana', 'zwrot')
        ORDER BY s.data_sprzedazy DESC
    ''').fetchall()
    
    # Utwórz workbook
    wb = Workbook()
    
    # ============ ARKUSZ 1: PODSUMOWANIE MIESIĘCZNE ============
    ws_summary = wb.active
    ws_summary.title = "Podsumowanie"
    
    # Style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1e40af')
    money_fill = PatternFill('solid', fgColor='dcfce7')
    border = Side(style='thin', color='e5e7eb')
    cell_border = Border(left=border, right=border, top=border, bottom=border)
    
    # Nagłówki podsumowania
    headers_sum = ['Miesiąc', 'Sprzedaży', 'Przychód', 'Koszt zakupu', 'Prowizja Allegro (11%)', 'Zysk netto', 'ROI %']
    for col, header in enumerate(headers_sum, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = cell_border
    
    # Grupuj sprzedaże po miesiącach
    miesiace = {}
    for s in sprzedaze:
        data = s['data_sprzedazy'] or ''
        if data:
            # Format: YYYY-MM-DD lub YYYY-MM-DD HH:MM:SS
            miesiac = data[:7]  # YYYY-MM
        else:
            miesiac = 'Brak daty'
        
        if miesiac not in miesiace:
            miesiace[miesiac] = {'sprzedazy': 0, 'przychod': 0, 'koszt': 0}
        
        cena = s['cena'] or 0
        ilosc = s['ilosc'] or 1
        przychod = cena * ilosc
        koszt = (s['koszt_jednostkowy'] or 0) * ilosc

        miesiace[miesiac]['sprzedazy'] += 1
        miesiace[miesiac]['przychod'] += przychod
        miesiace[miesiac]['koszt'] += koszt
    
    # Sortuj miesiące (najnowsze na górze)
    sorted_miesiace = sorted(miesiace.keys(), reverse=True)
    
    # Wypełnij dane
    row = 2
    total_sprzedazy = 0
    total_przychod = 0
    total_koszt = 0
    total_prowizja = 0
    total_zysk = 0
    
    for miesiac in sorted_miesiace:
        data = miesiace[miesiac]
        przychod = data['przychod']
        koszt = data['koszt']
        prowizja = przychod * 0.11
        zysk = przychod - koszt - prowizja
        roi = (zysk / koszt * 100) if koszt > 0 else 0
        
        # Formatuj nazwę miesiąca
        try:
            dt = datetime.strptime(miesiac, '%Y-%m')
            nazwa_miesiaca = dt.strftime('%B %Y').capitalize()
            # Polski format
            pl_months = {'January': 'Styczeń', 'February': 'Luty', 'March': 'Marzec', 
                        'April': 'Kwiecień', 'May': 'Maj', 'June': 'Czerwiec',
                        'July': 'Lipiec', 'August': 'Sierpień', 'September': 'Wrzesień',
                        'October': 'Październik', 'November': 'Listopad', 'December': 'Grudzień'}
            for en, pl in pl_months.items():
                nazwa_miesiaca = nazwa_miesiaca.replace(en, pl)
        except:
            nazwa_miesiaca = miesiac
        
        ws_summary.cell(row=row, column=1, value=nazwa_miesiaca).border = cell_border
        ws_summary.cell(row=row, column=2, value=data['sprzedazy']).border = cell_border
        ws_summary.cell(row=row, column=3, value=round(przychod, 2)).border = cell_border
        ws_summary.cell(row=row, column=4, value=round(koszt, 2)).border = cell_border
        ws_summary.cell(row=row, column=5, value=round(prowizja, 2)).border = cell_border
        
        zysk_cell = ws_summary.cell(row=row, column=6, value=round(zysk, 2))
        zysk_cell.border = cell_border
        if zysk > 0:
            zysk_cell.fill = money_fill
        
        roi_cell = ws_summary.cell(row=row, column=7, value=f"{roi:.1f}%")
        roi_cell.border = cell_border
        
        total_sprzedazy += data['sprzedazy']
        total_przychod += przychod
        total_koszt += koszt
        total_prowizja += prowizja
        total_zysk += zysk
        
        row += 1
    
    # Wiersz SUMA
    row += 1
    suma_font = Font(bold=True)
    suma_fill = PatternFill('solid', fgColor='fef3c7')
    
    ws_summary.cell(row=row, column=1, value='RAZEM').font = suma_font
    ws_summary.cell(row=row, column=1).fill = suma_fill
    ws_summary.cell(row=row, column=2, value=total_sprzedazy).font = suma_font
    ws_summary.cell(row=row, column=2).fill = suma_fill
    ws_summary.cell(row=row, column=3, value=round(total_przychod, 2)).font = suma_font
    ws_summary.cell(row=row, column=3).fill = suma_fill
    ws_summary.cell(row=row, column=4, value=round(total_koszt, 2)).font = suma_font
    ws_summary.cell(row=row, column=4).fill = suma_fill
    ws_summary.cell(row=row, column=5, value=round(total_prowizja, 2)).font = suma_font
    ws_summary.cell(row=row, column=5).fill = suma_fill
    ws_summary.cell(row=row, column=6, value=round(total_zysk, 2)).font = suma_font
    ws_summary.cell(row=row, column=6).fill = suma_fill
    total_roi = (total_zysk / total_koszt * 100) if total_koszt > 0 else 0
    ws_summary.cell(row=row, column=7, value=f"{total_roi:.1f}%").font = suma_font
    ws_summary.cell(row=row, column=7).fill = suma_fill
    
    # Szerokość kolumn
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 14
    ws_summary.column_dimensions['D'].width = 14
    ws_summary.column_dimensions['E'].width = 20
    ws_summary.column_dimensions['F'].width = 14
    ws_summary.column_dimensions['G'].width = 10
    
    # ============ ARKUSZ 2: WSZYSTKIE SZCZEGÓŁY ============
    ws_details = wb.create_sheet("Wszystkie")

    headers_det = ['Data', 'Nazwa produktu', 'Kategoria', 'ASIN', 'EAN', 'Paleta', 'Dostawca',
                   'Ilość', 'Cena sprzedaży', 'Przychód', 'Koszt zakupu', 'Prowizja 11%', 'Zysk']

    def _fill_detail_headers(ws):
        for col, header in enumerate(headers_det, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = cell_border

    def _fill_detail_row(ws, row_num, s):
        cena = s['cena'] or 0
        ilosc = s['ilosc'] or 1
        przychod = cena * ilosc
        koszt = (s['koszt_jednostkowy'] or 0) * ilosc
        prowizja = przychod * 0.11
        zysk = przychod - koszt - prowizja

        ws.cell(row=row_num, column=1, value=s['data_sprzedazy'] or '').border = cell_border
        ws.cell(row=row_num, column=2, value=(s['nazwa'] or '')[:60]).border = cell_border
        ws.cell(row=row_num, column=3, value=s['kategoria'] or '').border = cell_border
        ws.cell(row=row_num, column=4, value=s['asin'] or '').border = cell_border
        ws.cell(row=row_num, column=5, value=s['ean'] or '').border = cell_border
        ws.cell(row=row_num, column=6, value=s['paleta'] or '').border = cell_border
        ws.cell(row=row_num, column=7, value=s['dostawca'] or '').border = cell_border
        ws.cell(row=row_num, column=8, value=ilosc).border = cell_border
        ws.cell(row=row_num, column=9, value=round(cena, 2)).border = cell_border
        ws.cell(row=row_num, column=10, value=round(przychod, 2)).border = cell_border
        ws.cell(row=row_num, column=11, value=round(koszt, 2)).border = cell_border
        ws.cell(row=row_num, column=12, value=round(prowizja, 2)).border = cell_border

        zysk_cell = ws.cell(row=row_num, column=13, value=round(zysk, 2))
        zysk_cell.border = cell_border
        if zysk > 0:
            zysk_cell.fill = money_fill
        elif zysk < 0:
            zysk_cell.font = Font(color='FF0000')
        return przychod, koszt, prowizja, zysk

    def _set_detail_widths(ws):
        widths = [12, 45, 14, 12, 14, 15, 12, 8, 14, 12, 14, 14, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Wypełnij arkusz "Wszystkie"
    _fill_detail_headers(ws_details)
    row = 2
    for s in sprzedaze:
        _fill_detail_row(ws_details, row, s)
        row += 1
    _set_detail_widths(ws_details)

    # ============ ARKUSZE PER MIESIĄC ============
    # Grupuj sprzedaże po miesiącach
    sprzedaze_per_month = {}
    for s in sprzedaze:
        data = s['data_sprzedazy'] or ''
        miesiac = data[:7] if data else 'Brak-daty'
        if miesiac not in sprzedaze_per_month:
            sprzedaze_per_month[miesiac] = []
        sprzedaze_per_month[miesiac].append(s)

    for miesiac in sorted(sprzedaze_per_month.keys(), reverse=True):
        sales = sprzedaze_per_month[miesiac]
        # Nazwa arkusza (max 31 znaków, bez niedozwolonych)
        sheet_name = miesiac.replace('/', '-')[:31]
        ws_month = wb.create_sheet(sheet_name)
        _fill_detail_headers(ws_month)

        row = 2
        m_przychod = 0
        m_koszt = 0
        m_prowizja = 0
        m_zysk = 0
        for s in sales:
            p, k, pr, z = _fill_detail_row(ws_month, row, s)
            m_przychod += p
            m_koszt += k
            m_prowizja += pr
            m_zysk += z
            row += 1

        # Wiersz podsumowania miesiąca
        row += 1
        suma_font_m = Font(bold=True)
        suma_fill_m = PatternFill('solid', fgColor='fef3c7')
        ws_month.cell(row=row, column=1, value='RAZEM').font = suma_font_m
        ws_month.cell(row=row, column=1).fill = suma_fill_m
        ws_month.cell(row=row, column=8, value=len(sales)).font = suma_font_m
        ws_month.cell(row=row, column=10, value=round(m_przychod, 2)).font = suma_font_m
        ws_month.cell(row=row, column=10).fill = suma_fill_m
        ws_month.cell(row=row, column=11, value=round(m_koszt, 2)).font = suma_font_m
        ws_month.cell(row=row, column=11).fill = suma_fill_m
        ws_month.cell(row=row, column=12, value=round(m_prowizja, 2)).font = suma_font_m
        ws_month.cell(row=row, column=12).fill = suma_fill_m
        ws_month.cell(row=row, column=13, value=round(m_zysk, 2)).font = suma_font_m
        ws_month.cell(row=row, column=13).fill = suma_fill_m if m_zysk >= 0 else PatternFill('solid', fgColor='fecaca')
        _set_detail_widths(ws_month)
    
    # Zapisz plik
    filename = f"raport_sprzedazy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'cloud_exports', filename)
    
    # Upewnij się, że folder istnieje
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Zapisz i zamknij workbook
    wb.save(filepath)
    wb.close()
    
    # Poczekaj chwilę na zakończenie zapisu
    import time
    time.sleep(0.1)
    
    # Sprawdź czy plik istnieje
    if not os.path.exists(filepath):
        from flask import jsonify
        return jsonify({'error': f'Nie udało się utworzyć pliku: {filepath}'}), 500
    
    # Zwróć plik do pobrania
    from flask import send_file
    return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@magazynier_bp.route('/raport-sprzedazy-page')
def raport_sprzedazy_page():
    """Strona z przyciskiem do pobrania raportu"""
    html = '''
    <div class="hdr"><h1><span class=material-symbols-outlined>bar_chart</span> Raport sprzedaży</h1></div>
    
    <div style="padding:20px;text-align:center">
        <p style="color:#94a3b8;margin-bottom:20px">
            Pobierz raport sprzedaży w formacie Excel (.xlsx)<br>
            Zawiera podsumowanie miesięczne i szczegóły wszystkich transakcji
        </p>
        
        <a href="/magazyn/raport-sprzedazy" class="btn btn-ok" style="font-size:18px;padding:15px 30px">
            <span class=material-symbols-outlined>download</span> POBIERZ RAPORT EXCEL
        </a>
        
        <div style="margin-top:30px;text-align:left;max-width:500px;margin-left:auto;margin-right:auto">
            <h3 style="color:#beee00"><span class=material-symbols-outlined>list_alt</span> Co zawiera raport:</h3>
            <ul style="color:#94a3b8;line-height:2">
                <li><b>Arkusz "Podsumowanie"</b> - przychody, koszty, zyski per miesiąc</li>
                <li><b>Arkusz "Szczegóły"</b> - wszystkie sprzedaże z datami</li>
                <li>Automatyczne obliczenie prowizji Allegro (11%)</li>
                <li>ROI dla każdego miesiąca</li>
            </ul>
        </div>
    </div>
    
    <a href="/analityka" class="back">← Powrót do Analityki</a>
    '''
    return render(html)


@magazynier_bp.route('/lezaki')
def lezaki():
    """Lista produktów stojących >30 dni z kosztem z palety i sugestią obniżki"""
    from datetime import datetime, timedelta
    
    conn = get_db()
    days_30_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # Pobierz lezaki z dołączoną paletą
    produkty = conn.execute('''
        SELECT 
            p.id, p.nazwa, p.ean, p.asin, p.ilosc, p.cena_brutto, p.cena_netto,
            p.cena_allegro, p.status, p.lokalizacja, p.paleta, p.paleta_id,
            p.zdjecie_url, p.data_dodania,
            CAST(julianday('now') - julianday(p.data_dodania) AS INTEGER) as dni_stoi,
            pal.cena_zakupu as paleta_cena_zakupu,
            COALESCE((SELECT SUM(pr2.ilosc) FROM produkty pr2 WHERE pr2.paleta_id = pal.id), 0)
            + COALESCE((SELECT SUM(pr2.sprzedano_offline) FROM produkty pr2 WHERE pr2.paleta_id = pal.id), 0)
            + COALESCE((SELECT SUM(sp2.ilosc) FROM sprzedaze sp2
                JOIN produkty pp2 ON sp2.produkt_id = pp2.id
                WHERE pp2.paleta_id = pal.id
                AND sp2.status NOT IN ('zwrot','anulowane','anulowana')), 0)
            as paleta_ilosc_sztuk,
            pal.nazwa as paleta_nazwa
        FROM produkty p
        LEFT JOIN palety pal ON p.paleta_id = pal.id
        WHERE p.status IN ('magazyn', 'wystawiony')
          AND date(p.data_dodania) < ?
          AND p.ilosc > 0
        ORDER BY dni_stoi DESC
    ''', (days_30_ago,)).fetchall()
    
    
    total_wartosc = 0
    
    html = '''
    <div class="hdr"><h1>⏳ LEŻAKI</h1><small>Produkty stojące >30 dni</small></div>
    '''
    
    if not produkty:
        html += '<div class="alert alert-ok"><span class=material-symbols-outlined>check_circle</span> Brak leżaków — wszystko się kręci!</div>'
        return render(html)
    
    # Karty produktów
    rows_html = ''
    for p in produkty:
        dni = p['dni_stoi'] or 0
        
        # Koszt jednostkowy z palety lub fallback
        paleta_cena = p['paleta_cena_zakupu'] or 0
        paleta_szt = p['paleta_ilosc_sztuk'] or 0
        if paleta_cena > 0 and paleta_szt > 0:
            koszt_szt_brutto = paleta_cena / paleta_szt
            koszt_szt_netto = koszt_szt_brutto / 1.23
            koszt_src = f"z palety ({p['paleta_nazwa'] or '—'})"
        else:
            koszt_szt_brutto = p['cena_brutto'] or 0  # JEDNOSTKOWA
            koszt_szt_netto = p['cena_netto'] or 0          # JEDNOSTKOWA
            koszt_src = "z produktu"
        
        cena_allegro = p['cena_allegro'] or 0
        
        # Sugestia obniżki — im dłużej stoi, tym agresywniejsza
        if dni < 45:
            obnizka_pct = 10
            obnizka_kolor = '#eab308'
            obnizka_ikona = '●'
        elif dni < 60:
            obnizka_pct = 20
            obnizka_kolor = '#f97316'
            obnizka_ikona = '●'
        elif dni < 90:
            obnizka_pct = 30
            obnizka_kolor = '#ef4444'
            obnizka_ikona = '<span class=material-symbols-outlined>fiber_manual_record</span>'
        else:
            obnizka_pct = 40
            obnizka_kolor = '#dc2626'
            obnizka_ikona = '🚨'
        
        cena_sugerowana = round(cena_allegro * (1 - obnizka_pct / 100), 2)
        marza_po_obnizce = cena_sugerowana - koszt_szt_brutto
        
        # Kolor dni
        if dni > 90:
            dni_kolor = '#dc2626'
        elif dni > 60:
            dni_kolor = '#ef4444'
        elif dni > 45:
            dni_kolor = '#f97316'
        else:
            dni_kolor = '#eab308'
        
        # Wartość leżąca
        wartosc_lezaca = koszt_szt_brutto * (p['ilosc'] or 0)
        total_wartosc += wartosc_lezaca
        
        img_html = ''
        if p['zdjecie_url']:
            img_html = f'<img src="{p["zdjecie_url"]}" style="width:60px;height:60px;object-fit:contain;border-radius:8px;background:#0a0a0f;flex-shrink:0">'
        else:
            img_html = '<div style="width:60px;height:60px;background:#1e1e2e;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1.5rem;flex-shrink:0"><span class=material-symbols-outlined>inventory_2</span></div>'
        
        status_color = '#8ff5ff' if p['status'] == 'wystawiony' else '#eab308'
        status_text = 'WYSTAWIONY' if p['status'] == 'wystawiony' else 'MAGAZYN'
        
        rows_html += f'''
        <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:14px;margin-bottom:10px">
            <div style="display:flex;gap:12px;align-items:flex-start">
                {img_html}
                <div style="flex:1;min-width:0">
                    <div style="font-weight:600;font-size:0.9rem;margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
                        <a href="/magazyn/produkt/{p['id']}" style="color:#fff;text-decoration:none">{p['nazwa']}</a>
                    </div>
                    <div style="font-size:0.72rem;color:#64748b;margin-bottom:8px">
                        {p['lokalizacja'] or '—'} • {p['paleta'] or '—'} • {p['ilosc']} szt • <span style="color:{status_color}">{status_text}</span>
                    </div>
                    
                    <!-- Dni stoi -->
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
                        <div style="background:{dni_kolor}22;border:1px solid {dni_kolor};border-radius:20px;padding:3px 12px;font-size:0.8rem;font-weight:700;color:{dni_kolor}">
                            ⏳ {dni} dni
                        </div>
                        <div style="font-size:0.7rem;color:#64748b">od {(p['data_dodania'] or '')[:10]}</div>
                    </div>
                    
                    <!-- Ceny -->
                    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px">
                        <div style="background:#1e1e2e;border-radius:8px;padding:8px;text-align:center">
                            <div style="font-size:0.65rem;color:#64748b;margin-bottom:2px">KOSZT/szt brutto</div>
                            <div style="font-weight:700;color:#94a3b8">{koszt_szt_brutto:.2f} zł</div>
                            <div style="font-size:0.6rem;color:#475569">{koszt_src}</div>
                        </div>
                        <div style="background:#1e1e2e;border-radius:8px;padding:8px;text-align:center">
                            <div style="font-size:0.65rem;color:#64748b;margin-bottom:2px">CENA ALLEGRO</div>
                            <div style="font-weight:700;color:#beee00">{cena_allegro:.2f} zł</div>
                            <div style="font-size:0.6rem;color:#475569">marża: {cena_allegro - koszt_szt_brutto:.2f} zł</div>
                        </div>
                    </div>
                    
                    <!-- Sugestia obniżki -->
                    <div style="background:{obnizka_kolor}15;border:1px solid {obnizka_kolor}44;border-radius:8px;padding:10px">
                        <div style="font-size:0.75rem;font-weight:600;color:{obnizka_kolor};margin-bottom:6px">
                            {obnizka_ikona} Sugerowana obniżka -{obnizka_pct}%
                        </div>
                        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px">
                            <div>
                                <div style="font-size:0.7rem;color:#94a3b8">Nowa cena: <b style="color:#fff">{cena_sugerowana:.2f} zł</b></div>
                                <div style="font-size:0.75rem;color:#cbd5e1">Marża po obniżce: <b>{marza_po_obnizce:+.2f} zł</b></div>
                            </div>
                            <a href="/magazyn/produkt/{p['id']}/edytuj" 
                               style="padding:6px 12px;background:{obnizka_kolor};border-radius:6px;color:#000;font-size:0.72rem;font-weight:700;text-decoration:none;white-space:nowrap;flex-shrink:0">
                                <span class=material-symbols-outlined>edit</span> Zmień cenę
                            </a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        '''
    
    # Podsumowanie na górze
    html += f'''
    <div style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:15px;margin-bottom:15px">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;text-align:center">
            <div>
                <div style="font-size:1.4rem;font-weight:700;color:#eab308">{len(produkty)}</div>
                <div style="font-size:0.65rem;color:#64748b">PRODUKTÓW</div>
            </div>
            <div>
                <div style="font-size:1.4rem;font-weight:700;color:#ef4444">{sum(p['ilosc'] or 0 for p in produkty)}</div>
                <div style="font-size:0.65rem;color:#64748b">SZTUK</div>
            </div>
            <div>
                <div style="font-size:1.1rem;font-weight:700;color:#f97316">{total_wartosc:.0f} zł</div>
                <div style="font-size:0.65rem;color:#64748b">ZAMROŻONE</div>
            </div>
        </div>
    </div>
    
    <div style="display:flex;gap:8px;margin-bottom:15px">
        <div style="background:#eab30822;border:1px solid #eab308;border-radius:8px;padding:6px 12px;font-size:0.72rem;color:#eab308">● &lt;45 dni → -10%</div>
        <div style="background:#f9731622;border:1px solid #f97316;border-radius:8px;padding:6px 12px;font-size:0.72rem;color:#f97316">● &lt;60 dni → -20%</div>
        <div style="background:#ef444422;border:1px solid #ef4444;border-radius:8px;padding:6px 12px;font-size:0.72rem;color:#ef4444"><span class=material-symbols-outlined>fiber_manual_record</span> &lt;90 dni → -30%</div>
        <div style="background:#dc262622;border:1px solid #dc2626;border-radius:8px;padding:6px 12px;font-size:0.72rem;color:#dc2626">🚨 90+ dni → -40%</div>
    </div>
    '''
    
    html += rows_html
    html += '<a href="/analityka" class="back">← Powrót</a>'
    
    return render(html)




@magazynier_bp.route('/koszty', methods=['GET', 'POST'])
def koszty_page():
    from .database import get_db
    conn = get_db()
    
    # Upewnij się że tabela istnieje
    conn.execute('''CREATE TABLE IF NOT EXISTS koszty (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nazwa TEXT NOT NULL,
        kwota REAL NOT NULL,
        kategoria TEXT DEFAULT 'inne',
        data DATE DEFAULT CURRENT_DATE,
        notatka TEXT DEFAULT '',
        data_dodania TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    
    msg = ''
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'dodaj':
            nazwa  = request.form.get('nazwa', '').strip()
            kwota  = float(request.form.get('kwota', 0) or 0)
            kat    = request.form.get('kategoria', 'inne')
            data   = request.form.get('data', datetime.now().strftime('%Y-%m-%d'))
            notatka = request.form.get('notatka', '').strip()
            if nazwa and kwota > 0:
                conn.execute('INSERT INTO koszty (nazwa, kwota, kategoria, data, notatka) VALUES (?,?,?,?,?)',
                    (nazwa, kwota, kat, data, notatka))
                conn.commit()
                conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
                msg = f'<span class=material-symbols-outlined>check_circle</span> Dodano koszt: {nazwa} — {kwota:.2f} zł'
        elif action == 'usun':
            kid = request.form.get('id')
            conn.execute('DELETE FROM koszty WHERE id=?', (kid,))
            conn.commit()
            conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
            msg = '<span class=material-symbols-outlined>delete</span> Usunięto koszt'
    
    # Pobierz wszystkie koszty
    koszty = conn.execute('SELECT * FROM koszty ORDER BY data DESC, id DESC').fetchall()
    
    # Suma per miesiąc (bieżący rok)
    year = datetime.now().year
    koszty_miesiecznie = conn.execute('''
        SELECT strftime('%m', data) as m, SUM(kwota) as suma
        FROM koszty WHERE strftime('%Y', data) = ?
        GROUP BY m
    ''', (str(year),)).fetchall()
    koszty_msc = {int(r['m']): float(r['suma']) for r in koszty_miesiecznie}
    
    # Suma per kategoria
    koszty_kat = conn.execute('''
        SELECT kategoria, SUM(kwota) as suma, COUNT(*) as cnt
        FROM koszty GROUP BY kategoria ORDER BY suma DESC
    ''').fetchall()
    
    # Suma całkowita i bieżący miesiąc
    suma_total = sum(float(k['kwota']) for k in koszty)
    biezacy_m = datetime.now().month
    suma_msc = koszty_msc.get(biezacy_m, 0)
    
    
    KATEGORIE = [
        ('allegro', '<span class=material-symbols-outlined>shopping_cart</span> Prowizje Allegro'),
        ('wysylka', '<span class=material-symbols-outlined>inventory_2</span> Wysyłka / InPost'),
        ('reklama', '📣 Reklama'),
        ('magazyn', '<span class=material-symbols-outlined>factory</span> Magazyn / najem'),
        ('zakup', '<span class=material-symbols-outlined>paid</span> Zakup towaru'),
        ('ksiegowosc', '<span class=material-symbols-outlined>list_alt</span> Księgowość / ZUS'),
        ('inne', '<span class=material-symbols-outlined>bolt</span> Inne'),
    ]
    
    nazwy_m = ['Sty','Lut','Mar','Kwi','Maj','Cze','Lip','Sie','Wrz','Paź','Lis','Gru']
    
    # Icon mapping per category
    KAT_ICONS = {
        'allegro': 'shopping_cart',
        'wysylka': 'local_shipping',
        'reklama': 'campaign',
        'magazyn': 'inventory',
        'zakup': 'payments',
        'ksiegowosc': 'receipt_long',
        'inne': 'bolt',
    }

    # Tabela kosztów HTML
    koszty_html = ''
    for k in koszty:
        kat_label = dict(KATEGORIE).get(k['kategoria'], k['kategoria'])
        kat_icon = KAT_ICONS.get(k['kategoria'], 'bolt')
        notatka_html = f' • <span style="color:var(--kz-muted)">{k["notatka"]}</span>' if k['notatka'] else ''
        koszty_html += f'''
        <div class="kz-list-item">
            <div style="display:flex;align-items:center;gap:14px;flex:1;min-width:0">
                <div class="kz-list-icon"><span class=material-symbols-outlined style=font-size:20px>{kat_icon}</span></div>
                <div style="min-width:0">
                    <div style="font-weight:600;font-size:0.9rem;color:var(--kz-text)">{k['nazwa']}</div>
                    <div style="font-size:0.75rem;color:var(--kz-muted);margin-top:2px">{kat_label} &middot; {k['data']}{notatka_html}</div>
                </div>
            </div>
            <div style="font-weight:700;color:var(--kz-pink);white-space:nowrap;font-family:'Space Grotesk',sans-serif;font-size:0.95rem">-{k['kwota']:.2f} zl</div>
            <form action="/magazyn/koszty" method="POST" style="margin:0">
                <input type="hidden" name="action" value="usun">
                <input type="hidden" name="id" value="{k['id']}">
                <button type="submit" onclick="return confirm('Usun?')" class="kz-del-btn">
                    <span class=material-symbols-outlined style=font-size:16px>close</span>
                </button>
            </form>
        </div>'''

    if not koszty_html:
        koszty_html = '<div style="text-align:center;color:var(--kz-muted);padding:40px 20px;font-size:0.9rem">Brak kosztow. Dodaj pierwszy!</div>'
    
    # Opcje kategorii
    kat_options = ''.join([f'<option value="{v}">{l}</option>' for v, l in KATEGORIE])
    
    msg_html = f'<div style="background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);border-radius:0;padding:12px 18px;margin-bottom:20px;color:var(--kz-cyan);font-family:\'Manrope\',sans-serif;font-size:0.85rem">{msg}</div>' if msg else ''

    # Category breakdown pills
    kat_pills = ''
    if koszty_kat:
        for r in koszty_kat:
            pill_icon = KAT_ICONS.get(r['kategoria'], 'bolt')
            kat_pills += f'<div style="display:inline-flex;align-items:center;gap:8px;background:var(--kz-card2);border:1px solid var(--kz-border);padding:8px 14px;margin:0 8px 8px 0;font-size:0.8rem;font-family:\'Manrope\',sans-serif"><span class=material-symbols-outlined style=font-size:16px;color:var(--kz-pink)>{pill_icon}</span><span style="color:var(--kz-pink);font-weight:700">{float(r["suma"]):.0f} zl</span><span style="color:var(--kz-muted)">{dict(KATEGORIE).get(r["kategoria"],r["kategoria"])}</span></div>'

    html = f'''
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700;900&family=Manrope:wght@300;400;500;600;700;800&display=swap');
        @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap');
        .material-symbols-outlined{{font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 24}}

        :root{{
            --kz-cyan:#8ff5ff;--kz-pink:#ff6b9b;--kz-lime:#cafd00;--kz-lime-dim:#beee00;
            --kz-bg:#0e0e10;--kz-card:#131315;--kz-card2:#19191c;--kz-card3:#1f1f22;--kz-card4:#262528;
            --kz-border:rgba(255,255,255,0.06);--kz-text:#f9f5f8;--kz-muted:#adaaad;
        }}

        .kz-wrap{{max-width:900px;margin:0 auto;font-family:'Manrope',sans-serif;position:relative;z-index:1;padding:0 8px}}
        .kz-headline{{font-family:'Space Grotesk',sans-serif}}
        .kz-label{{font-family:'Manrope',sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:0.2em;color:var(--kz-muted)}}

        /* Cyber grid bg */
        .kz-grid-bg{{background-size:50px 50px;background-image:linear-gradient(to right,rgba(143,245,255,0.04) 1px,transparent 1px),linear-gradient(to bottom,rgba(143,245,255,0.04) 1px,transparent 1px);position:fixed;inset:0;pointer-events:none;z-index:0}}

        /* Bento metric cards */
        .kz-bento{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:24px}}
        .kz-metric{{position:relative;overflow:hidden;padding:24px 20px;background:var(--kz-card);border:1px solid var(--kz-border);transition:all 0.2s}}
        .kz-metric:hover{{background:var(--kz-card3)}}
        .kz-metric-accent{{position:absolute;left:0;top:0;bottom:0;width:3px}}
        .kz-metric-icon{{position:absolute;right:12px;top:12px;font-size:48px!important;opacity:0.06;color:var(--kz-text)}}
        .kz-metric-val{{font-family:'Space Grotesk',sans-serif;font-size:1.8rem;font-weight:800;line-height:1;margin-bottom:6px}}
        .kz-metric-label{{font-family:'Manrope',sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:0.2em;color:var(--kz-muted)}}

        /* Form section */
        .kz-form-card{{background:var(--kz-card);border:1px solid var(--kz-border);padding:24px;margin-bottom:24px}}
        .kz-form-header{{font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;color:var(--kz-text);margin-bottom:18px;display:flex;align-items:center;gap:10px}}
        .kz-form-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px}}
        .kz-form-full{{grid-column:1/-1}}
        .kz-input-label{{font-family:'Manrope',sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:0.2em;color:var(--kz-muted);margin-bottom:6px;display:block}}
        .kz-input{{width:100%;padding:10px 14px;background:rgba(10,10,15,0.6);border:1px solid var(--kz-border);color:var(--kz-text);font-family:'Manrope',sans-serif;font-size:0.85rem;transition:all 0.2s;box-sizing:border-box}}
        .kz-input:focus{{border-color:var(--kz-cyan);box-shadow:0 0 12px rgba(143,245,255,0.15);outline:none}}
        .kz-input::placeholder{{color:rgba(173,170,173,0.4)}}
        .kz-submit{{font-family:'Space Grotesk',sans-serif;font-weight:800;text-transform:uppercase;letter-spacing:0.05em;padding:12px 28px;border:none;cursor:pointer;background:var(--kz-lime);color:#1a1a00;font-size:0.9rem;box-shadow:0 4px 20px rgba(202,253,0,0.25);transition:all 0.15s}}
        .kz-submit:hover{{transform:scale(1.02);box-shadow:0 6px 25px rgba(202,253,0,0.35)}}
        .kz-submit:active{{transform:scale(0.97)}}

        /* List section */
        .kz-list-header{{font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;color:var(--kz-text);margin-bottom:14px;display:flex;align-items:center;gap:10px}}
        .kz-list-item{{display:flex;align-items:center;gap:14px;padding:16px 20px;background:var(--kz-card);border:1px solid var(--kz-border);margin-bottom:6px;transition:all 0.2s}}
        .kz-list-item:hover{{background:var(--kz-card3);border-color:rgba(255,255,255,0.1)}}
        .kz-list-icon{{width:36px;height:36px;display:flex;align-items:center;justify-content:center;background:var(--kz-card3);border:1px solid var(--kz-border);color:var(--kz-cyan);flex-shrink:0}}
        .kz-del-btn{{width:32px;height:32px;display:flex;align-items:center;justify-content:center;background:rgba(255,107,155,0.08);border:1px solid rgba(255,107,155,0.15);color:var(--kz-pink);cursor:pointer;transition:all 0.15s;flex-shrink:0}}
        .kz-del-btn:hover{{background:rgba(255,107,155,0.2);border-color:rgba(255,107,155,0.4)}}

        .kz-back{{display:inline-flex;align-items:center;gap:6px;color:var(--kz-muted);text-decoration:none;font-size:0.85rem;font-family:'Manrope',sans-serif;margin-top:24px;transition:color 0.2s}}
        .kz-back:hover{{color:var(--kz-cyan)}}

        /* Responsive */
        @media(max-width:768px){{
            .kz-bento{{grid-template-columns:1fr}}
            .kz-form-grid{{grid-template-columns:1fr 1fr}}
        }}
        @media(max-width:480px){{
            .kz-form-grid{{grid-template-columns:1fr}}
        }}
    </style>

    <div class="kz-grid-bg"></div>

    <div class="kz-wrap">

    <!-- Header -->
    <div style="margin-bottom:28px">
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">
            <div style="width:8px;height:8px;border-radius:50%;background:var(--kz-lime-dim);box-shadow:0 0 10px rgba(190,238,0,0.5)"></div>
            <span class="kz-label" style="color:var(--kz-lime-dim);font-weight:900">Wydatki operacyjne</span>
        </div>
        <h1 class="kz-headline" style="font-size:clamp(2rem,5vw,3rem);font-weight:900;font-style:italic;letter-spacing:-0.04em;line-height:1;color:var(--kz-text);margin:0">
            KOSZTY
        </h1>
    </div>

    {msg_html}

    <!-- Bento metrics -->
    <div class="kz-bento">
        <div class="kz-metric">
            <div class="kz-metric-accent" style="background:var(--kz-cyan);box-shadow:0 0 15px rgba(143,245,255,0.3)"></div>
            <span class="material-symbols-outlined kz-metric-icon">account_balance_wallet</span>
            <div class="kz-metric-val" style="color:var(--kz-cyan);text-shadow:0 0 20px rgba(143,245,255,0.4)">{suma_total:.0f} zl</div>
            <div class="kz-metric-label">Lacznie koszty</div>
        </div>
        <div class="kz-metric">
            <div class="kz-metric-accent" style="background:var(--kz-pink);box-shadow:0 0 15px rgba(255,107,155,0.3)"></div>
            <span class="material-symbols-outlined kz-metric-icon">calendar_month</span>
            <div class="kz-metric-val" style="color:var(--kz-pink)">{suma_msc:.0f} zl</div>
            <div class="kz-metric-label">Ten miesiac</div>
        </div>
        <div class="kz-metric">
            <div class="kz-metric-accent" style="background:var(--kz-lime);box-shadow:0 0 15px rgba(202,253,0,0.3)"></div>
            <span class="material-symbols-outlined kz-metric-icon">format_list_numbered</span>
            <div class="kz-metric-val" style="color:var(--kz-lime)">{len(koszty)}</div>
            <div class="kz-metric-label">Wszystkich wpisow</div>
        </div>
    </div>

    <!-- Category breakdown -->
    <div style="margin-bottom:24px">
        {kat_pills}
    </div>

    <!-- Add cost form -->
    <div class="kz-form-card">
        <div class="kz-form-header">
            <span class=material-symbols-outlined style=color:var(--kz-lime);font-size:22px>add_circle</span>
            Dodaj koszt
        </div>
        <form action="/magazyn/koszty" method="POST">
            <input type="hidden" name="action" value="dodaj">
            <div class="kz-form-grid">
                <div>
                    <label class="kz-input-label">Nazwa</label>
                    <input type="text" name="nazwa" class="kz-input" placeholder="np. Prowizja Allegro Luty" required>
                </div>
                <div>
                    <label class="kz-input-label">Kwota (zl)</label>
                    <input type="number" name="kwota" step="0.01" class="kz-input" placeholder="0.00" required>
                </div>
                <div>
                    <label class="kz-input-label">Kategoria</label>
                    <select name="kategoria" class="kz-input">{kat_options}</select>
                </div>
                <div>
                    <label class="kz-input-label">Data</label>
                    <input type="date" name="data" class="kz-input" value="{datetime.now().strftime('%Y-%m-%d')}">
                </div>
                <div class="kz-form-full">
                    <label class="kz-input-label">Notatka (opcjonalnie)</label>
                    <input type="text" name="notatka" class="kz-input" placeholder="np. faktura FV/2026/02/001">
                </div>
            </div>
            <button type="submit" class="kz-submit">Dodaj koszt</button>
        </form>
    </div>

    <!-- Cost list -->
    <div style="margin-bottom:20px">
        <div class="kz-list-header">
            <span class=material-symbols-outlined style=color:var(--kz-cyan);font-size:22px>receipt_long</span>
            Wszystkie koszty
        </div>
        {koszty_html}
    </div>

    <a href="/magazyn" class="kz-back">
        <span class=material-symbols-outlined style=font-size:18px>arrow_back</span>
        Powrot do magazynu
    </a>

    </div>
    '''
    return render(html)



@magazynier_bp.route('/sprzedaz-prywatna', methods=['GET', 'POST'])
def sprzedaz_prywatna_page():
    from .database import get_db
    conn = get_db()
    conn.execute('''CREATE TABLE IF NOT EXISTS sprzedaze_prywatne (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        opis TEXT NOT NULL, kwota REAL NOT NULL,
        data DATE DEFAULT CURRENT_DATE, notatka TEXT DEFAULT '',
        data_dodania TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    
    msg = ''
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'dodaj':
            opis   = request.form.get('opis', '').strip()
            kwota  = float(request.form.get('kwota', 0) or 0)
            data   = request.form.get('data', datetime.now().strftime('%Y-%m-%d'))
            notatka = request.form.get('notatka', '').strip()
            if opis and kwota > 0:
                conn.execute('INSERT INTO sprzedaze_prywatne (opis, kwota, data, notatka) VALUES (?,?,?,?)',
                    (opis, kwota, data, notatka))
                conn.commit()
                conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
                msg = f'<span class=material-symbols-outlined>check_circle</span> Dodano: {opis} — {kwota:.2f} zł'
        elif action == 'usun':
            conn.execute('DELETE FROM sprzedaze_prywatne WHERE id=?', (request.form.get('id'),))
            conn.commit()
            conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
            msg = '<span class=material-symbols-outlined>delete</span> Usunięto'
    
    sprzedaze = conn.execute('SELECT * FROM sprzedaze_prywatne ORDER BY data DESC, id DESC').fetchall()
    
    year = datetime.now().year
    suma_rok = conn.execute(
        'SELECT COALESCE(SUM(kwota),0) FROM sprzedaze_prywatne WHERE strftime(\'%Y\', data)=?', (str(year),)
    ).fetchone()[0]
    biezacy_m = datetime.now().month
    suma_msc = conn.execute(
        'SELECT COALESCE(SUM(kwota),0) FROM sprzedaze_prywatne WHERE strftime(\'%Y-%m\', data)=?',
        (f'{year}-{biezacy_m:02d}',)
    ).fetchone()[0]
    
    msg_html = f'<div style="background:rgba(143,245,255,0.08);border:1px solid rgba(143,245,255,0.2);padding:12px 18px;margin-bottom:20px;color:#8ff5ff;font-family:Manrope,sans-serif;font-size:0.85rem;font-weight:600">{msg}</div>' if msg else ''

    rows_html = ''
    for s in sprzedaze:
        rows_html += f'''
        <div style="display:flex;align-items:center;gap:14px;padding:18px 20px;background:#131315;border-left:3px solid #cafd00;margin-bottom:2px;transition:background 0.2s" onmouseover="this.style.background='#1f1f22'" onmouseout="this.style.background='#131315'">
            <div style="width:36px;height:36px;display:flex;align-items:center;justify-content:center;background:rgba(202,253,0,0.08);flex-shrink:0;font-size:1.1rem">[HANDSHAKE]</div>
            <div style="flex:1;min-width:0">
                <div style="font-weight:600;font-family:Manrope,sans-serif;color:#f9f5f8;font-size:0.9rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{s['opis']}</div>
                <div style="font-size:0.7rem;color:#adaaad;font-family:Manrope,sans-serif;margin-top:2px">{s['data']}{f' &middot; {s["notatka"]}' if s['notatka'] else ''}</div>
            </div>
            <div style="font-weight:800;color:#cafd00;white-space:nowrap;font-family:Space Grotesk,sans-serif;font-size:1rem;letter-spacing:-0.02em;text-shadow:0 0 20px rgba(202,253,0,0.3)">+{s['kwota']:.2f} zl</div>
            <form action="/magazyn/sprzedaz-prywatna" method="POST" style="margin:0">
                <input type="hidden" name="action" value="usun">
                <input type="hidden" name="id" value="{s['id']}">
                <button type="submit" onclick="return confirm('Usuń?')"
                    style="background:rgba(255,107,155,0.1);border:1px solid rgba(255,107,155,0.2);color:#ff6b9b;padding:6px 12px;cursor:pointer;font-size:0.7rem;font-family:Manrope,sans-serif;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;transition:opacity 0.15s;opacity:0.7" onmouseover="this.style.opacity='1'" onmouseout="this.style.opacity='0.7'">&#x2715;</button>
            </form>
        </div>'''

    if not rows_html:
        rows_html = '<div style="text-align:center;color:#adaaad;padding:40px 20px;font-family:Manrope,sans-serif;font-size:0.85rem">Brak wpisów</div>'

    html = f'''
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700;900&family=Manrope:wght@300;400;500;600;700;800&display=swap');
        .sp-grid-bg{{background-size:50px 50px;background-image:linear-gradient(to right,rgba(143,245,255,0.04) 1px,transparent 1px),linear-gradient(to bottom,rgba(143,245,255,0.04) 1px,transparent 1px);position:fixed;inset:0;pointer-events:none;z-index:0}}
        .sp-wrap{{max-width:800px;margin:0 auto;position:relative;z-index:1;font-family:Manrope,sans-serif}}
        .sp-label{{font-family:Manrope,sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:0.2em;color:rgba(255,255,255,0.45);font-weight:700}}
        .sp-input{{width:100%;padding:12px 16px;background:#262528;border:1px solid rgba(255,255,255,0.06);color:#f9f5f8;font-family:Manrope,sans-serif;font-size:0.9rem;transition:all 0.2s;outline:none;box-sizing:border-box}}
        .sp-input:focus{{border-color:#8ff5ff;box-shadow:0 0 12px rgba(143,245,255,0.15)}}
        .sp-input::placeholder{{color:#adaaad}}
    </style>
    <div class="sp-grid-bg"></div>
    <div class="sp-wrap">
        <!-- Header -->
        <div style="margin-bottom:32px">
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">
                <div style="width:8px;height:8px;border-radius:50%;background:#beee00;box-shadow:0 0 10px rgba(190,238,0,0.5)"></div>
                <span class="sp-label" style="color:#beee00;font-weight:900">Prywatna</span>
            </div>
            <h1 style="font-family:Space Grotesk,sans-serif;font-size:clamp(2rem,5vw,3rem);font-weight:900;font-style:italic;letter-spacing:-0.04em;line-height:1;color:#f9f5f8;margin:0">
                SPRZEDAZ <span style="color:#8ff5ff">PRYWATNA</span>
            </h1>
            <p style="color:#adaaad;font-size:0.85rem;margin:8px 0 0;font-family:Manrope,sans-serif">Sprzedaz poza Allegro — OLX, Facebook, cash itp.</p>
        </div>

        {msg_html}

        <!-- Stats bento grid -->
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:2px;margin-bottom:28px">
            <div style="background:#131315;padding:24px 20px;border-left:3px solid #8ff5ff">
                <div class="sp-label" style="margin-bottom:8px">Rok {year}</div>
                <div style="font-family:Space Grotesk,sans-serif;font-size:1.8rem;font-weight:800;color:#8ff5ff;letter-spacing:-0.03em;text-shadow:0 0 30px rgba(143,245,255,0.4),0 0 60px rgba(143,245,255,0.15)">{suma_rok:.0f} zl</div>
            </div>
            <div style="background:#131315;padding:24px 20px;border-left:3px solid #ff6b9b">
                <div class="sp-label" style="margin-bottom:8px">Ten miesiac</div>
                <div style="font-family:Space Grotesk,sans-serif;font-size:1.8rem;font-weight:800;color:#ff6b9b;letter-spacing:-0.03em;text-shadow:0 0 25px rgba(255,107,155,0.3)">{suma_msc:.0f} zl</div>
            </div>
            <div style="background:#131315;padding:24px 20px;border-left:3px solid #cafd00">
                <div class="sp-label" style="margin-bottom:8px">Transakcji</div>
                <div style="font-family:Space Grotesk,sans-serif;font-size:1.8rem;font-weight:800;color:#cafd00;letter-spacing:-0.03em;text-shadow:0 0 25px rgba(202,253,0,0.3)">{len(sprzedaze)}</div>
            </div>
        </div>

        <!-- Add form - glass panel -->
        <div style="backdrop-filter:blur(12px);background:rgba(19,19,21,0.8);border:1px solid rgba(255,255,255,0.06);padding:28px;margin-bottom:28px">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:20px">
                <span style="font-size:1.2rem">&#x2795;</span>
                <span style="font-family:Space Grotesk,sans-serif;font-weight:700;font-size:1rem;color:#f9f5f8;text-shadow:0 0 20px rgba(143,245,255,0.2)">Dodaj sprzedaz</span>
            </div>
            <form action="/magazyn/sprzedaz-prywatna" method="POST">
                <input type="hidden" name="action" value="dodaj">
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
                    <div>
                        <label class="sp-label" style="display:block;margin-bottom:6px">Opis</label>
                        <input type="text" name="opis" class="sp-input" placeholder="np. Odkurzacz Dyson, OLX" required>
                    </div>
                    <div>
                        <label class="sp-label" style="display:block;margin-bottom:6px">Kwota (zl)</label>
                        <input type="number" name="kwota" step="0.01" class="sp-input" placeholder="0.00" required>
                    </div>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px">
                    <div>
                        <label class="sp-label" style="display:block;margin-bottom:6px">Data</label>
                        <input type="date" name="data" class="sp-input" value="{datetime.now().strftime('%Y-%m-%d')}">
                    </div>
                    <div>
                        <label class="sp-label" style="display:block;margin-bottom:6px">Notatka</label>
                        <input type="text" name="notatka" class="sp-input" placeholder="opcjonalnie">
                    </div>
                </div>
                <button type="submit" style="font-family:Space Grotesk,sans-serif;font-weight:900;text-transform:uppercase;letter-spacing:-0.02em;font-style:italic;padding:14px 32px;border:none;cursor:pointer;background:#cafd00;color:#0e0e10;font-size:0.95rem;box-shadow:0 4px 20px rgba(202,253,0,0.3);transition:all 0.15s" onmouseover="this.style.transform='scale(1.02)';this.style.boxShadow='0 6px 30px rgba(202,253,0,0.45)'" onmouseout="this.style.transform='';this.style.boxShadow='0 4px 20px rgba(202,253,0,0.3)'">&#x1F4BE; DODAJ</button>
            </form>
        </div>

        <!-- History section -->
        <div style="margin-bottom:28px">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px">
                <span style="font-size:1.1rem">&#x1F4CB;</span>
                <span style="font-family:Space Grotesk,sans-serif;font-weight:700;font-size:1rem;color:#f9f5f8;text-shadow:0 0 20px rgba(143,245,255,0.2)">Historia</span>
                <div style="flex:1;height:1px;background:rgba(255,255,255,0.06)"></div>
            </div>
            <div style="border:1px solid rgba(255,255,255,0.06);overflow:hidden">
                {rows_html}
            </div>
        </div>

        <!-- Back link -->
        <a href="/magazyn" style="display:inline-flex;align-items:center;gap:8px;color:#adaaad;text-decoration:none;font-family:Manrope,sans-serif;font-size:0.85rem;font-weight:600;padding:10px 0;transition:color 0.15s" onmouseover="this.style.color='#8ff5ff'" onmouseout="this.style.color='#adaaad'">&larr; Powrot</a>
    </div>
    '''
    return render(html)



@magazynier_bp.route('/remanent')
def remanent_page():
    """Strona remanentu z podglądem i przyciskiem do pobrania Excela"""
    from datetime import datetime
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')

    # Podsumowanie per paleta
    # Używa _paleta_koszt_szt() — dynamicznie liczy magazyn + offline + allegro
    palety_all = conn.execute('''
        SELECT
            pal.id, pal.nazwa, pal.dostawca, pal.cena_zakupu, pal.data_zakupu,
            COALESCE(SUM(p.ilosc), 0) as szt_magazyn,
            COALESCE(SUM(p.cena_allegro * p.ilosc), 0) as wartosc_allegro
        FROM palety pal
        LEFT JOIN produkty p ON p.paleta_id = pal.id
        GROUP BY pal.id
        ORDER BY CAST(SUBSTR(pal.nazwa, INSTR(pal.nazwa,'#')+1) AS INTEGER) DESC, pal.data_zakupu DESC
    ''').fetchall()

    # Przelicz dane per paleta
    palety_data = []
    for p in palety_all:
        cena_zak = float(p['cena_zakupu'] or 0)
        szt_magazyn = int(p['szt_magazyn'] or 0)
        # Koszt/szt z helpera (dynamicznie: magazyn + offline + allegro sprzedaże)
        koszt_szt = _paleta_koszt_szt(conn, p['id'])
        wartosc_kosztowa = koszt_szt * szt_magazyn
        # Ile sztuk było łącznie w palecie
        szt_wszystkich = round(cena_zak / koszt_szt) if koszt_szt > 0 else szt_magazyn
        szt_sprzedano = max(0, szt_wszystkich - szt_magazyn)
        wartosc_allegro = float(p['wartosc_allegro'] or 0)
        palety_data.append({
            'nazwa': p['nazwa'], 'dostawca': p['dostawca'],
            'data_zakupu': p['data_zakupu'], 'cena_zakupu': cena_zak,
            'szt_magazyn': szt_magazyn, 'szt_sprzedano': szt_sprzedano,
            'szt_wszystkich': szt_wszystkich,
            'wartosc_kosztowa': wartosc_kosztowa,
            'wartosc_allegro': wartosc_allegro,
        })

    # Sumy ogólne
    suma_zakupu = sum(p['cena_zakupu'] for p in palety_data)
    suma_kosztowa = sum(p['wartosc_kosztowa'] for p in palety_data)
    suma_allegro = sum(p['wartosc_allegro'] for p in palety_data)
    suma_magazyn = sum(p['szt_magazyn'] for p in palety_data)
    suma_sprzedano = sum(p['szt_sprzedano'] for p in palety_data)

    # Podział per kategoria (produkty w magazynie)
    kategorie = conn.execute('''
        SELECT 
            COALESCE(NULLIF(kategoria,''),'inne') as kat,
            COUNT(*) as cnt,
            SUM(ilosc) as sztuki,
            COALESCE(SUM(cena_netto), 0) as wartosc_netto,
            COALESCE(SUM(cena_allegro * ilosc), 0) as wartosc_detal
        FROM produkty
        WHERE status != 'sprzedany'
        GROUP BY kat
        ORDER BY wartosc_detal DESC
    ''').fetchall()

    # Convert kategorie rows to dicts for Jinja2 template access
    kategorie_data = [dict(k) for k in kategorie]

    return render_template('remanent.html',
        today=today,
        palety_count=len(palety_all),
        palety_data=palety_data,
        kategorie=kategorie_data,
        suma_zakupu=suma_zakupu,
        suma_kosztowa=suma_kosztowa,
        suma_allegro=suma_allegro,
        suma_magazyn=suma_magazyn,
        suma_sprzedano=suma_sprzedano,
        brand_name=current_app.config.get('BRAND_NAME', 'Akces Hub'),
        version=current_app.config.get('VERSION', ''),
        current_user=session.get('username')
    )


@magazynier_bp.route('/remanent/excel')
def remanent_excel():
    """Generuje remanent magazynowy - stan na dziś w formacie Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import io
    from flask import send_file
    
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Pobierz wszystkie produkty w magazynie (nie sprzedane)
    produkty = conn.execute('''
        SELECT 
            p.id,
            p.nazwa,
            p.kategoria,
            p.ilosc,
            p.cena_netto,
            p.cena_allegro,
            p.status,
            p.stan,
            p.regal,
            p.ean,
            p.asin,
            p.data_dodania,
            pal.nazwa as paleta_nazwa,
            pal.cena_zakupu as paleta_cena_zakupu,
            pal.dostawca,
            pal.data_zakupu,
            COALESCE(
                (SELECT COUNT(*) FROM sprzedaze s WHERE s.produkt_id = p.id AND s.status NOT IN ('zwrot','anulowana')),
                0
            ) as ilosc_sprzedanych
        FROM produkty p
        LEFT JOIN palety pal ON p.paleta_id = pal.id
        WHERE p.status NOT IN ('sprzedany')
        ORDER BY CAST(SUBSTR(pal.nazwa, INSTR(pal.nazwa,'#')+1) AS INTEGER) DESC, pal.nazwa, p.kategoria, p.nazwa
    ''').fetchall()
    
    # Podsumowanie
    palety_all = conn.execute('''
        SELECT 
            pal.id,
            pal.nazwa,
            pal.dostawca,
            pal.cena_zakupu,
            pal.data_zakupu,
            COUNT(p.id) as ilosc_produktow,
            COUNT(CASE WHEN p.status = 'sprzedany' THEN 1 END) as sprzedanych,
            COUNT(CASE WHEN p.status != 'sprzedany' THEN 1 END) as pozostalo,
            COALESCE(SUM(CASE WHEN p.status != 'sprzedany' THEN p.cena_allegro * p.ilosc ELSE 0 END), 0) as wartosc_detaliczna,
            COALESCE(SUM(CASE WHEN p.status != 'sprzedany' THEN p.cena_netto * p.ilosc ELSE 0 END), 0) as wartosc_netto
        FROM palety pal
        LEFT JOIN produkty p ON p.paleta_id = pal.id
        GROUP BY pal.id
        ORDER BY CAST(SUBSTR(pal.nazwa, INSTR(pal.nazwa,'#')+1) AS INTEGER) DESC, pal.data_zakupu DESC
    ''').fetchall()
    
    
    # === TWORZENIE EXCELA ===
    wb = Workbook()
    
    # Kolory
    COL_NAGLOWEK = "1e3a5f"
    COL_SUBNAGLOWEK = "2d5a8e" 
    COL_ZIELONY = "e8f5e9"
    COL_ZOLTY = "fff9c4"
    COL_CZERWONY = "ffebee"
    COL_SZARY = "f5f5f5"
    COL_BIALY = "ffffff"
    
    def header_style(cell, bg=COL_NAGLOWEK, bold=True, color="FFFFFF", size=11, center=True):
        cell.font = Font(bold=bold, color=color, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def data_style(cell, bg=COL_BIALY, bold=False, color="000000", center=False, num_format=None):
        cell.font = Font(bold=bold, color=color, size=10, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
        if num_format:
            cell.number_format = num_format
    
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def set_border(cell):
        cell.border = border
    
    # ============================================================
    # ARKUSZ 1: REMANENT SZCZEGÓŁOWY
    # ============================================================
    ws1 = wb.active
    ws1.title = "Remanent szczegółowy"
    ws1.freeze_panes = "A3"
    
    # Tytuł
    ws1.merge_cells("A1:O1")
    c = ws1["A1"]
    c.value = f"REMANENT MAGAZYNOWY — STAN NA DZIEŃ {today}"
    header_style(c, bg=COL_NAGLOWEK, size=13)
    ws1.row_dimensions[1].height = 30
    
    # Nagłówki kolumn
    headers = [
        "Lp.", "Paleta", "Dostawca", "Nazwa produktu", "Kategoria",
        "Stan", "Status", "Ilosc w mag.", "Cena netto (zl)",
        "Cena Allegro (zl)", "Wartosc netto (zl)", "Wartosc detal. (zl)",
        "Regal", "EAN", "Data zakupu palety"
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=2, column=col, value=h)
        header_style(c, bg=COL_SUBNAGLOWEK, size=10)
        set_border(c)
    ws1.row_dimensions[2].height = 35
    
    # Szerokości kolumn
    widths = [5, 20, 18, 40, 18, 12, 14, 8, 12, 12, 14, 14, 10, 16, 18]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    
    # Dane
    row = 3
    for lp, p in enumerate(produkty, 1):
        ilosc = p["ilosc"] or 1
        cena_netto_szt = p["cena_netto"] or 0  # JEDNOSTKOWA - nie dzielić przez ilosc
        cena_allegro = p["cena_allegro"] or 0
        
        # Kolor wiersza
        if p["status"] == "wystawiony":
            bg = COL_ZIELONY
        elif p["status"] == "uszkodzony":
            bg = COL_CZERWONY
        else:
            bg = COL_BIALY if lp % 2 == 0 else COL_SZARY
        
        dane = [
            lp,
            p["paleta_nazwa"] or "-",
            p["dostawca"] or "-",
            p["nazwa"] or "-",
            p["kategoria"] or "-",
            p["stan"] or "-",
            p["status"] or "-",
            ilosc,
            cena_netto_szt,      # cena jednostkowa netto
            cena_allegro,
            f"=H{row}*I{row}",  # wartość netto = ilosc × cena_netto_szt
            f"=H{row}*J{row}",  # wartość detal = ilosc × cena_allegro
            p["regal"] or "-",
            p["ean"] or "-",
            p["data_zakupu"] or "-",
        ]
        
        for col, val in enumerate(dane, 1):
            c = ws1.cell(row=row, column=col, value=val)
            data_style(c, bg=bg, center=(col in [1, 7, 8, 9, 10, 11, 12, 13]))
            set_border(c)
            if col in [9, 10, 11, 12]:
                c.number_format = '#,##0.00'
        row += 1
    
    # Wiersz podsumowania
    ws1.merge_cells(f"A{row}:G{row}")
    c = ws1.cell(row=row, column=1, value="RAZEM")
    header_style(c, bg=COL_NAGLOWEK, size=11)
    set_border(c)
    
    for col in [8, 11, 12]:
        c = ws1.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{row-1})")
        header_style(c, bg=COL_NAGLOWEK, size=11)
        set_border(c)
        if col in [11, 12]:
                c.number_format = '#,##0.00'
    
    # ============================================================
    # ARKUSZ 2: PODSUMOWANIE PER PALETA
    # ============================================================
    ws2 = wb.create_sheet("Podsumowanie palet")
    ws2.freeze_panes = "A3"
    
    ws2.merge_cells("A1:J1")
    c = ws2["A1"]
    c.value = f"PODSUMOWANIE PER PALETA — {today}"
    header_style(c, bg=COL_NAGLOWEK, size=13)
    ws2.row_dimensions[1].height = 30
    
    headers2 = [
        "Lp.", "Paleta", "Dostawca", "Data zakupu",
        "Koszt zakupu palety (zl)", "Produktow lacznie",
        "Sprzedanych", "Pozostalo w mag.",
        "Wartosc netto w mag. (zl)", "Wartosc detal. w mag. (zl)"
    ]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        header_style(c, bg=COL_SUBNAGLOWEK, size=10)
        set_border(c)
    ws2.row_dimensions[2].height = 35
    
    widths2 = [5, 25, 18, 14, 16, 12, 12, 12, 18, 18]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    row2 = 3
    for lp, pal in enumerate(palety_all, 1):
        bg = COL_BIALY if lp % 2 == 0 else COL_SZARY
        dane2 = [
            lp,
            pal["nazwa"] or "-",
            pal["dostawca"] or "-",
            pal["data_zakupu"] or "-",
            pal["cena_zakupu"] or 0,
            pal["ilosc_produktow"] or 0,
            pal["sprzedanych"] or 0,
            pal["pozostalo"] or 0,
            pal["wartosc_netto"] or 0,
            pal["wartosc_detaliczna"] or 0,
        ]
        for col, val in enumerate(dane2, 1):
            c = ws2.cell(row=row2, column=col, value=val)
            data_style(c, bg=bg, center=(col in [1, 5, 6, 7, 8, 9, 10]))
            set_border(c)
            if col in [5, 9, 10]:
                c.number_format = '#,##0.00'
        row2 += 1
    
    # Podsumowanie
    ws2.merge_cells(f"A{row2}:D{row2}")
    c = ws2.cell(row=row2, column=1, value="RAZEM")
    header_style(c, bg=COL_NAGLOWEK)
    set_border(c)
    for col in [5, 6, 7, 8, 9, 10]:
        c = ws2.cell(row=row2, column=col, value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{row2-1})")
        header_style(c, bg=COL_NAGLOWEK)
        set_border(c)
        if col in [5, 9, 10]:
                c.number_format = '#,##0.00'
    
    # ============================================================
    # ARKUSZ 3: PODSUMOWANIE KATEGORII
    # ============================================================
    ws3 = wb.create_sheet("Wg kategorii")
    ws3.freeze_panes = "A3"
    
    conn2 = get_db()
    kategorie = conn2.execute('''
        SELECT 
            COALESCE(kategoria, 'Brak kategorii') as kategoria,
            COUNT(*) as ilosc_prod,
            SUM(ilosc) as ilosc_szt,
            COALESCE(SUM(cena_netto), 0) as wartosc_netto,
            COALESCE(SUM(cena_allegro * ilosc), 0) as wartosc_detal
        FROM produkty
        WHERE status NOT IN ('sprzedany')
        GROUP BY kategoria
        ORDER BY wartosc_detal DESC
    ''').fetchall()

    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = f"STAN MAGAZYNU WG KATEGORII — {today}"
    header_style(c, bg=COL_NAGLOWEK, size=13)
    ws3.row_dimensions[1].height = 30
    
    headers3 = ["Lp.", "Kategoria", "Ilosc produktow", "Ilosc sztuk", "Wartosc netto (zl)", "Wartosc detal. (zl)"]




    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        header_style(c, bg=COL_SUBNAGLOWEK, size=10)
        set_border(c)
    ws3.row_dimensions[2].height = 35
    
    widths3 = [5, 30, 14, 12, 18, 18]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    
    row3 = 3
    for lp, kat in enumerate(kategorie, 1):
        bg = COL_BIALY if lp % 2 == 0 else COL_SZARY
        for col, val in enumerate([lp, kat["kategoria"], kat["ilosc_prod"], kat["ilosc_szt"], kat["wartosc_netto"], kat["wartosc_detal"]], 1):
            c = ws3.cell(row=row3, column=col, value=val)
            data_style(c, bg=bg, center=(col != 2))
            set_border(c)
            if col in [5, 6]:
                c.number_format = '#,##0.00'
        row3 += 1
    
    # Podsumowanie
    ws3.merge_cells(f"A{row3}:B{row3}")
    c = ws3.cell(row=row3, column=1, value="RAZEM")
    header_style(c, bg=COL_NAGLOWEK)
    set_border(c)
    for col in [3, 4, 5, 6]:
        c = ws3.cell(row=row3, column=col, value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{row3-1})")
        header_style(c, bg=COL_NAGLOWEK)
        set_border(c)
        if col in [5, 6]:
                c.number_format = '#,##0.00'
    
    # Zapisz do bufora
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"remanent_{today}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@magazynier_bp.route('/statystyki-zakupow')
def statystyki_zakupow():
    """Przeniesione do /analityka/zakupy-dostawcy"""
    from flask import redirect
    return redirect('/analityka/zakupy-dostawcy')



# ============================================================
# PRZYJĘCIE PALETY - Quick condition assessment
# ============================================================

@magazynier_bp.route('/przyjecie/<int:paleta_id>')
def przyjecie_palety(paleta_id):
    """Ekran przyjęcia palety - szybka ocena stanu produktów"""
    conn = get_db()
    paleta = conn.execute('SELECT * FROM palety WHERE id = ?', (paleta_id,)).fetchone()
    if not paleta:
        return redirect(url_for('magazynier.palety_lista'))

    try:
        produkty = conn.execute(
            'SELECT id, nazwa, ean, zdjecie_url, ilosc, stan_przyjecia, notatki_przyjecia, lokalizacja FROM produkty WHERE paleta_id = ? ORDER BY id',
            (paleta_id,)
        ).fetchall()
    except:
        produkty = conn.execute(
            'SELECT id, nazwa, ean, zdjecie_url, ilosc, "" as stan_przyjecia, "" as notatki_przyjecia, "" as lokalizacja FROM produkty WHERE paleta_id = ? ORDER BY id',
            (paleta_id,)
        ).fetchall()

    # Lista regałów do datalist
    try:
        from modules.warehouse_heatmap import WAREHOUSE_CONFIG
        regal_list = WAREHOUSE_CONFIG.get('shelves', [])
    except:
        regal_list = []

    total_sztuk = sum(p['ilosc'] or 1 for p in produkty)

    html = f'''
    <div style="padding:15px;max-width:900px;margin:0 auto">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
            <div>
                <h2 style="margin:0;font-size:1.3rem"><span class=material-symbols-outlined>list_alt</span> Przyjęcie palety #{paleta_id}</h2>
                <div style="color:#64748b;font-size:0.85rem;margin-top:4px">{paleta['nazwa']} • {len(produkty)} prod. • {total_sztuk} szt.</div>
            </div>
            <a href="/magazyn/paleta-id/{paleta_id}" style="background:#1e293b;color:#94a3b8;padding:8px 16px;border-radius:8px;text-decoration:none;font-size:0.85rem">← Powrót</a>
        </div>

        <div id="progress-bar" style="background:#1e1e2e;border-radius:8px;height:8px;margin-bottom:20px;overflow:hidden">
            <div id="progress-fill" style="height:100%;background:linear-gradient(90deg,#beee00,#2dd85a);width:0%;transition:width 0.3s"></div>
        </div>
        <div id="progress-text" style="text-align:center;color:#64748b;font-size:0.8rem;margin-bottom:20px">0 / {len(produkty)} ocenionych</div>

        <datalist id="regal-list">'''
    for r in regal_list:
        html += f'<option value="{r}">'
    html += '''</datalist>

        <div id="products-list">'''

    stany = [
        ('Nowy', 'A', '#beee00'),
        ('Jak nowy', 'A-', '#8ff5ff'),
        ('Dobry', 'B', '#eab308'),
        ('Uszkodzony', 'C', '#f97316'),
        ('Zniszczony', 'D', '#ef4444'),
    ]

    for p in produkty:
        pid = p['id']
        ilosc = p['ilosc'] or 1
        current_stan = p['stan_przyjecia'] or ''
        current_notatki = p['notatki_przyjecia'] or ''
        zdjecie = p['zdjecie_url'] or ''
        img_html = f'<img src="{zdjecie}" style="width:60px;height:60px;object-fit:cover;border-radius:8px">' if zdjecie else '<div style="width:60px;height:60px;background:#1e1e2e;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1.5rem"><span class=material-symbols-outlined>inventory_2</span></div>'

        # Ilość badge
        ilosc_badge = f'<span style="background:#8ff5ff33;color:#8ff5ff;padding:2px 8px;border-radius:6px;font-size:0.75rem;font-weight:700">{ilosc} szt.</span>' if ilosc > 1 else '<span style="color:#64748b;font-size:0.75rem">1 szt.</span>'

        html += f'''
        <div class="prod-card" id="prod-{pid}" data-ilosc="{ilosc}" style="backdrop-filter:blur(16px);background:rgba(15,15,30,0.65);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:15px;margin-bottom:12px">
            <div style="display:flex;gap:12px;align-items:flex-start">
                {img_html}
                <div style="flex:1;min-width:0">
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
                        <div style="font-size:0.9rem;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex:1;min-width:0">{p['nazwa']}</div>
                        {ilosc_badge}
                    </div>
                    <div style="font-size:0.75rem;color:#64748b">EAN: {p['ean'] or '—'}</div>
                </div>
            </div>

            <!-- Tryb prosty: 1 stan dla wszystkich sztuk -->
            <div id="simple-mode-{pid}">
                <div style="display:flex;gap:6px;margin-top:12px;flex-wrap:wrap">'''

        for stan_name, stan_icon, stan_color in stany:
            is_active = 'true' if current_stan == stan_name else 'false'
            html += f'''
                    <button onclick="selectStan({pid}, '{stan_name}', this)"
                        class="stan-btn-{pid}"
                        style="padding:8px 14px;border-radius:8px;border:2px solid {stan_color if current_stan == stan_name else '#1e1e2e'};
                        background:{'rgba(190,238,0,0.1)' if current_stan == stan_name else '#0a0a0f'};
                        color:{stan_color};font-size:0.8rem;cursor:pointer;transition:all 0.2s;flex:1;min-width:0;text-align:center"
                        data-active="{is_active}" data-color="{stan_color}">
                        {stan_icon} {stan_name}
                    </button>'''

        # Przycisk "Podziel" tylko gdy ilosc > 1
        split_btn = ''
        if ilosc > 1:
            split_btn = f'''<button onclick="showSplitMode({pid}, {ilosc})" style="padding:8px 10px;border-radius:8px;border:2px solid #8ff5ff;background:#8ff5ff22;color:#8ff5ff;font-size:0.75rem;cursor:pointer;white-space:nowrap" title="Różne stany dla poszczególnych sztuk">✂ Podziel</button>'''

        html += f'''
                    {split_btn}
                </div>
            </div>

            <!-- Tryb podzielony: różne stany dla różnych sztuk -->
            <div id="split-mode-{pid}" style="display:none;margin-top:12px">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
                    <div style="font-size:0.8rem;color:#a78bfa;font-weight:600">✂ Podział {ilosc} szt. wg stanu:</div>
                    <button onclick="hideSplitMode({pid})" style="padding:4px 10px;border:1px solid #334155;background:#1e293b;border-radius:6px;color:#94a3b8;font-size:0.7rem;cursor:pointer">← Prosty tryb</button>
                </div>
                <div style="display:grid;grid-template-columns:1fr auto;gap:6px;align-items:center">'''

        for stan_name, stan_icon, stan_color in stany:
            html += f'''
                    <div style="display:flex;align-items:center;gap:6px">
                        <div style="width:10px;height:10px;border-radius:3px;background:{stan_color};flex-shrink:0"></div>
                        <span style="font-size:0.8rem;color:{stan_color}">{stan_icon} {stan_name}</span>
                    </div>
                    <input type="number" id="split-{pid}-{stan_name.replace(' ', '_')}" min="0" max="{ilosc}" value="0"
                        onchange="validateSplit({pid}, {ilosc})"
                        style="width:60px;padding:6px 8px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:6px;color:#e2e8f0;font-size:0.85rem;text-align:center">'''

        html += f'''
                </div>
                <div id="split-sum-{pid}" style="margin-top:6px;font-size:0.75rem;color:#64748b;text-align:right">Suma: 0 / {ilosc}</div>
            </div>

            <div style="display:flex;gap:8px;margin-top:10px;align-items:center;flex-wrap:wrap">
                <input type="text" id="regal-{pid}" value="{p['lokalizacja'] if 'lokalizacja' in p.keys() else ''}" placeholder="Regał np. A2"
                    style="width:100px;padding:8px 12px;background:#0a0a0f;border:1px solid #8ff5ff33;border-radius:8px;color:#8ff5ff;font-size:0.8rem;font-weight:600" list="regal-list">
                <input type="text" id="notatki-{pid}" value="{current_notatki}" placeholder="Notatki (wady, braki...)"
                    style="flex:1;min-width:150px;padding:8px 12px;background:#0a0a0f;border:1px solid #1e1e2e;border-radius:8px;color:#e2e8f0;font-size:0.8rem">
                <button onclick="openCamera({pid})" style="padding:8px 12px;background:#7c3aed;border:none;border-radius:8px;color:white;cursor:pointer;font-size:0.85rem" title="Zrób zdjęcie i oceń AI">
                    <span class=material-symbols-outlined>photo_camera</span> AI
                </button>
            </div>
            <div id="ai-result-{pid}" style="display:none;margin-top:8px;padding:10px;background:#0a0a0f;border:1px solid #7c3aed;border-radius:8px;font-size:0.8rem;color:#c4b5fd"></div>
        </div>'''

    html += f'''
        </div>

        <div style="position:sticky;bottom:0;padding:15px 0;background:linear-gradient(transparent, #0a0a0f 30%)">
            <div style="display:flex;gap:10px">
                <button onclick="savePartial()" id="save-partial-btn"
                    style="flex:1;padding:14px;background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.3);border-radius:12px;color:#f59e0b;font-size:0.9rem;font-weight:600;cursor:pointer">
                    <span class=material-symbols-outlined>save</span> Zapisz i wróć później
                </button>
                <button onclick="saveAll()" id="save-btn"
                    style="flex:1;padding:14px;background:rgba(190,238,0,0.15);border:1px solid rgba(190,238,0,0.3);border-radius:12px;color:#beee00;font-size:0.9rem;font-weight:600;cursor:pointer">
                    <span class=material-symbols-outlined>check_circle</span> Zapisz i zakończ przyjęcie
                </button>
            </div>
        </div>
    </div>

    <!-- Hidden camera input -->
    <input type="file" id="camera-input" accept="image/*" capture="environment" style="display:none" onchange="handlePhoto(event)">

    <script>
    let currentStany = {{}};
    let splitModes = {{}};  // pid -> true if split mode active
    let currentPhotoProductId = null;
    const totalProducts = {len(produkty)};
    const stanNames = ['Nowy', 'Jak_nowy', 'Dobry', 'Uszkodzony', 'Zniszczony'];

    function selectStan(pid, stan, btn) {{
        document.querySelectorAll('.stan-btn-' + pid).forEach(b => {{
            b.style.border = '2px solid #1e1e2e';
            b.style.background = '#0a0a0f';
            b.dataset.active = 'false';
        }});
        btn.style.border = '2px solid ' + btn.dataset.color;
        btn.style.background = 'rgba(190,238,0,0.1)';
        btn.dataset.active = 'true';
        currentStany[pid] = stan;
        splitModes[pid] = false;
        updateProgress();
    }}

    function showSplitMode(pid, ilosc) {{
        document.getElementById('simple-mode-' + pid).style.display = 'none';
        document.getElementById('split-mode-' + pid).style.display = 'block';
        splitModes[pid] = true;
        // Jeśli był wybrany stan prosty, wstaw wszystkie szt w ten stan
        if (currentStany[pid]) {{
            const stanKey = currentStany[pid].replace(' ', '_');
            const inp = document.getElementById('split-' + pid + '-' + stanKey);
            if (inp) inp.value = ilosc;
            validateSplit(pid, ilosc);
        }}
        delete currentStany[pid];
        updateProgress();
    }}

    function hideSplitMode(pid) {{
        document.getElementById('simple-mode-' + pid).style.display = 'block';
        document.getElementById('split-mode-' + pid).style.display = 'none';
        splitModes[pid] = false;
    }}

    function validateSplit(pid, maxIlosc) {{
        let sum = 0;
        stanNames.forEach(s => {{
            const inp = document.getElementById('split-' + pid + '-' + s);
            if (inp) {{
                let v = parseInt(inp.value) || 0;
                if (v < 0) v = 0;
                sum += v;
            }}
        }});
        const sumDiv = document.getElementById('split-sum-' + pid);
        const ok = sum === maxIlosc;
        sumDiv.textContent = 'Suma: ' + sum + ' / ' + maxIlosc + (ok ? ' ' : sum > maxIlosc ? '  za dużo!' : '');
        sumDiv.style.color = ok ? '#beee00' : sum > maxIlosc ? '#ef4444' : '#f59e0b';

        if (ok) {{
            currentStany[pid] = 'split';
            updateProgress();
        }} else {{
            delete currentStany[pid];
            updateProgress();
        }}
    }}

    function updateProgress() {{
        const assessed = Object.keys(currentStany).length;
        const pct = totalProducts > 0 ? (assessed / totalProducts * 100) : 0;
        document.getElementById('progress-fill').style.width = pct + '%';
        document.getElementById('progress-text').textContent = assessed + ' / ' + totalProducts + ' ocenionych';
    }}

    function openCamera(pid) {{
        currentPhotoProductId = pid;
        document.getElementById('camera-input').click();
    }}

    function handlePhoto(event) {{
        const file = event.target.files[0];
        if (!file || !currentPhotoProductId) return;
        const pid = currentPhotoProductId;

        const resultDiv = document.getElementById('ai-result-' + pid);
        resultDiv.style.display = 'block';
        resultDiv.innerHTML = '<div style="color:#a78bfa">⏳ Analizuję zdjęcie AI...</div>';

        const reader = new FileReader();
        reader.onload = function(e) {{
            const base64 = e.target.result.split(',')[1];
            fetch('/magazyn/api/ai-ocena-stanu', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify({{ image_base64: base64, product_id: pid }})
            }})
            .then(r => r.json())
            .then(data => {{
                if (data.success) {{
                    resultDiv.innerHTML = `
                        <div style="margin-bottom:6px"><strong><span class=material-symbols-outlined>smart_toy</span> AI ocena:</strong> <span style="color:${{data.stan_color || '#beee00'}}">${{data.stan}}</span></div>
                        <div style="color:#94a3b8">${{data.opis}}</div>
                    `;
                    if (data.stan && !splitModes[pid]) {{
                        const btns = document.querySelectorAll('.stan-btn-' + pid);
                        btns.forEach(b => {{
                            if (b.textContent.includes(data.stan)) {{
                                selectStan(pid, data.stan, b);
                            }}
                        }});
                    }}
                    if (data.opis) {{
                        document.getElementById('notatki-' + pid).value = data.opis;
                    }}
                }} else {{
                    resultDiv.innerHTML = '<div style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> ' + (data.error || 'Błąd AI') + '</div>';
                }}
            }})
            .catch(err => {{
                resultDiv.innerHTML = '<div style="color:#ef4444"><span class=material-symbols-outlined>cancel</span> Błąd połączenia</div>';
            }});
        }};
        reader.readAsDataURL(file);
        event.target.value = '';
    }}

    function collectAssessments() {{
        const assessments = [];
        document.querySelectorAll('.prod-card').forEach(card => {{
            const pid = parseInt(card.id.replace('prod-', ''));
            const notatki = document.getElementById('notatki-' + pid)?.value || '';
            const regal = document.getElementById('regal-' + pid)?.value || '';

            if (splitModes[pid]) {{
                const split = {{}};
                stanNames.forEach(s => {{
                    const inp = document.getElementById('split-' + pid + '-' + s);
                    if (inp) {{
                        const v = parseInt(inp.value) || 0;
                        if (v > 0) split[s.replace('_', ' ')] = v;
                    }}
                }});
                assessments.push({{ product_id: pid, split: split, notatki: notatki, regal: regal }});
            }} else {{
                const stan = currentStany[pid] || '';
                if (stan) assessments.push({{ product_id: pid, stan: stan, notatki: notatki, regal: regal }});
            }}
        }});
        return assessments;
    }}

    function savePartial() {{
        const btn = document.getElementById('save-partial-btn');
        btn.disabled = true;
        btn.textContent = ' Zapisuję...';
        const assessments = collectAssessments();
        fetch('/magazyn/api/przyjecie-save', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json', 'ngrok-skip-browser-warning': '1' }},
            body: JSON.stringify({{ paleta_id: {paleta_id}, assessments: assessments, partial: true }})
        }})
        .then(r => {{ if (!r.ok) throw new Error('HTTP ' + r.status); return r.json(); }})
        .then(data => {{
            if (data.success) {{
                btn.textContent = ' Zapisano! Wróć później...';
                btn.style.borderColor = '#beee00';
                btn.style.color = '#beee00';
                setTimeout(() => window.location.href = '/magazyn/paleta-id/{paleta_id}', 1000);
            }} else {{
                btn.textContent = ' ' + (data.error || 'Błąd');
                btn.disabled = false;
            }}
        }}).catch(e => {{ btn.textContent = ' ' + e.message; btn.disabled = false; }});
    }}

    function saveAll() {{
        const btn = document.getElementById('save-btn');
        btn.disabled = true;
        btn.textContent = ' Zapisuję...';
        const assessments = collectAssessments();

        fetch('/magazyn/api/przyjecie-save', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json', 'ngrok-skip-browser-warning': '1' }},
            body: JSON.stringify({{ paleta_id: {paleta_id}, assessments: assessments }})
        }})
        .then(r => {{
            if (!r.ok) throw new Error('HTTP ' + r.status);
            return r.json();
        }})
        .then(data => {{
            if (data.success) {{
                btn.textContent = ' Zapisano!';
                btn.style.background = '#16a34a';
                setTimeout(() => window.location.href = '/magazyn/paleta-id/{paleta_id}', 1000);
            }} else {{
                btn.textContent = ' ' + (data.error || 'Błąd zapisu');
                btn.disabled = false;
            }}
        }})
        .catch(e => {{
            btn.textContent = ' ' + e.message;
            btn.disabled = false;
        }});
    }}

    // Init: load already assessed
    document.addEventListener('DOMContentLoaded', () => {{
        document.querySelectorAll('.prod-card').forEach(card => {{
            const pid = parseInt(card.id.replace('prod-', ''));
            const activeBtn = card.querySelector('[data-active="true"]');
            if (activeBtn && activeBtn.textContent.trim()) {{
                const stanText = activeBtn.textContent.replace(/[●●●●]/g, '').trim();
                currentStany[pid] = stanText;
            }}
        }});
        updateProgress();
    }});
    </script>
    '''
    return render(html)


@magazynier_bp.route('/api/przyjecie-save', methods=['POST'])
def przyjecie_save():
    """Zapisz oceny stanu produktów i oznacz paletę jako dostarczoną.
    Obsługuje tryb prosty (1 stan) i podzielony (split - różne stany per sztuka)."""
    try:
        data = request.get_json()
        paleta_id = data.get('paleta_id')
        assessments = data.get('assessments', [])
        is_partial = data.get('partial', False)
        conn = get_db()

        # Mapowanie stan_przyjecia → klasa_jakosci
        STAN_TO_KLASA = {
            'Nowy': 'A', 'Jak nowy': 'A-', 'Dobry': 'B',
            'Uszkodzony': 'C', 'Zniszczony': 'D'
        }
        STAN_TO_CONDITION = {
            'Nowy': 'Nowy', 'Jak nowy': 'Jak nowy', 'Dobry': 'Używany',
            'Uszkodzony': 'Uszkodzony', 'Zniszczony': 'Zniszczony'
        }

        # Pobierz wszystkie produkty palety
        all_product_ids = set(r['id'] for r in conn.execute(
            'SELECT id FROM produkty WHERE paleta_id = ?', (paleta_id,)
        ).fetchall())
        assessed_ids = set()

        for a in assessments:
            pid = a.get('product_id')
            notatki = a.get('notatki', '')
            regal = a.get('regal', '')
            split = a.get('split')  # dict: {"Nowy": 3, "Uszkodzony": 2} lub None

            if split and isinstance(split, dict):
                # TRYB PODZIELONY — rozdziel produkt na osobne rekordy per stan
                orig = conn.execute('SELECT * FROM produkty WHERE id = ?', (pid,)).fetchone()
                if not orig:
                    continue
                assessed_ids.add(pid)

                # Zbierz kolumny do kopiowania
                cols = [k for k in orig.keys() if k not in ('id', 'ilosc', 'stan_przyjecia', 'notatki_przyjecia', 'klasa_jakosci', 'stan')]

                first = True
                for stan_name, qty in split.items():
                    qty = int(qty)
                    if qty <= 0:
                        continue

                    klasa = STAN_TO_KLASA.get(stan_name, '')
                    condition = STAN_TO_CONDITION.get(stan_name, stan_name)

                    if first:
                        # Pierwszy stan — aktualizuj oryginalny rekord
                        conn.execute(
                            'UPDATE produkty SET ilosc = ?, stan_przyjecia = ?, notatki_przyjecia = ?, klasa_jakosci = ?, stan = ?, status = ?, lokalizacja = COALESCE(NULLIF(?, ""), lokalizacja) WHERE id = ?',
                            (qty, stan_name, notatki, klasa, condition, 'magazyn', regal, pid)
                        )
                        first = False
                    else:
                        # Kolejne stany — utwórz kopię produktu z nową ilością i stanem
                        col_names = ', '.join(cols)
                        placeholders = ', '.join(['?' for _ in cols])
                        values = [orig[c] for c in cols]

                        conn.execute(
                            f'INSERT INTO produkty ({col_names}, ilosc, stan_przyjecia, notatki_przyjecia, klasa_jakosci, stan, status) VALUES ({placeholders}, ?, ?, ?, ?, ?, ?)',
                            values + [qty, stan_name, notatki, klasa, condition, 'magazyn']
                        )

                # Utwórz sztuki w tabeli ewidencji
                conn.execute('''CREATE TABLE IF NOT EXISTS sztuki (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, produkt_id INTEGER NOT NULL,
                    numer INTEGER NOT NULL, stan TEXT DEFAULT 'Nowy',
                    status TEXT DEFAULT 'magazyn', opis_naprawy TEXT DEFAULT '',
                    data_naprawy DATE DEFAULT NULL, zdjecie TEXT DEFAULT '')''')
                conn.execute('DELETE FROM sztuki WHERE produkt_id = ?', (pid,))
                numer = 1
                for stan_name, qty in split.items():
                    condition = STAN_TO_CONDITION.get(stan_name, stan_name)
                    for _ in range(int(qty)):
                        conn.execute(
                            'INSERT INTO sztuki (produkt_id, numer, stan, status) VALUES (?, ?, ?, ?)',
                            (pid, numer, condition, 'magazyn')
                        )
                        numer += 1
            else:
                # TRYB PROSTY — jeden stan dla wszystkich sztuk
                stan = a.get('stan', '')
                if pid and stan:
                    assessed_ids.add(pid)
                    klasa = STAN_TO_KLASA.get(stan, '')
                    condition = STAN_TO_CONDITION.get(stan, stan)
                    conn.execute(
                        'UPDATE produkty SET stan_przyjecia = ?, notatki_przyjecia = ?, klasa_jakosci = ?, stan = ?, status = ?, lokalizacja = COALESCE(NULLIF(?, ""), lokalizacja) WHERE id = ?',
                        (stan, notatki, klasa, condition, 'magazyn', regal, pid)
                    )

        # Produkty nieocenione → stan 'nieoceniony' tylko przy pełnym zapisie
        if not is_partial:
            unassessed = all_product_ids - assessed_ids
            for uid in unassessed:
                conn.execute(
                    "UPDATE produkty SET stan_przyjecia = 'nieoceniony', klasa_jakosci = '', stan = 'nieoceniony', status = 'magazyn' WHERE id = ? AND (stan_przyjecia IS NULL OR stan_przyjecia = '')",
                    (uid,)
                )

        # Oznacz paletę jako dostarczoną
        # Paleta dostarczona + status oceny
        conn.execute('UPDATE palety SET dostarczona = 1 WHERE id = ?', (paleta_id,))
        if is_partial:
            conn.execute("UPDATE palety SET ocena_status = 'częściowa' WHERE id = ?", (paleta_id,))
        else:
            conn.execute("UPDATE palety SET ocena_status = 'zakończona' WHERE id = ?", (paleta_id,))
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@magazynier_bp.route('/api/ocen-produkt', methods=['POST'])
def ocen_produkt():
    """Oceń stan pojedynczego produktu (lub split na partie).
    Pozwala wrócić do oceny nieocenionych produktów bez ponownego przyjęcia palety."""
    try:
        data = request.get_json()
        pid = data.get('product_id')
        stan = data.get('stan', '')
        notatki = data.get('notatki', '')
        split = data.get('split')  # {"Nowy": 3, "Uszkodzony": 2}
        conn = get_db()

        STAN_TO_KLASA = {
            'Nowy': 'A', 'Jak nowy': 'A-', 'Dobry': 'B',
            'Uszkodzony': 'C', 'Zniszczony': 'D'
        }
        STAN_TO_CONDITION = {
            'Nowy': 'Nowy', 'Jak nowy': 'Jak nowy', 'Dobry': 'Używany',
            'Uszkodzony': 'Uszkodzony', 'Zniszczony': 'Zniszczony'
        }

        if split and isinstance(split, dict):
            orig = conn.execute('SELECT * FROM produkty WHERE id = ?', (pid,)).fetchone()
            if not orig:
                return jsonify({'success': False, 'error': 'Produkt nie znaleziony'})

            # Exclude id, kod_magazynowy (auto-generated by trigger), and fields we set manually
            cols = [k for k in orig.keys() if k not in ('id', 'kod_magazynowy', 'ilosc', 'stan_przyjecia', 'notatki_przyjecia', 'klasa_jakosci', 'stan')]
            first = True
            created_product_ids = []  # Track all product IDs (original + new)
            for stan_name, qty in split.items():
                qty = int(qty)
                if qty <= 0:
                    continue
                klasa = STAN_TO_KLASA.get(stan_name, '')
                condition = STAN_TO_CONDITION.get(stan_name, stan_name)
                if first:
                    conn.execute(
                        'UPDATE produkty SET ilosc = ?, stan_przyjecia = ?, notatki_przyjecia = ?, klasa_jakosci = ?, stan = ? WHERE id = ?',
                        (qty, stan_name, notatki, klasa, condition, pid)
                    )
                    created_product_ids.append(pid)
                    first = False
                else:
                    col_names = ', '.join(cols)
                    placeholders = ', '.join(['?' for _ in cols])
                    values = [orig[c] for c in cols]
                    cursor = conn.execute(
                        f'INSERT INTO produkty ({col_names}, ilosc, stan_przyjecia, notatki_przyjecia, klasa_jakosci, stan, status) VALUES ({placeholders}, ?, ?, ?, ?, ?, ?)',
                        values + [qty, stan_name, notatki, klasa, condition, 'magazyn']
                    )
                    new_pid = cursor.lastrowid
                    created_product_ids.append(new_pid)

            # Add historia entries for split products
            from .database import add_historia
            for cpid in created_product_ids:
                add_historia(cpid, 'oznaczono', f'Ocena stanu: split na {len(created_product_ids)} partii')

            # Utwórz sztuki w ewidencji per product
            conn.execute('''CREATE TABLE IF NOT EXISTS sztuki (
                id INTEGER PRIMARY KEY AUTOINCREMENT, produkt_id INTEGER NOT NULL,
                numer INTEGER NOT NULL, stan TEXT DEFAULT 'Nowy',
                status TEXT DEFAULT 'magazyn', opis_naprawy TEXT DEFAULT '',
                data_naprawy DATE DEFAULT NULL, zdjecie TEXT DEFAULT '')''')
            # Create sztuki for each split product
            for i, cpid in enumerate(created_product_ids):
                conn.execute('DELETE FROM sztuki WHERE produkt_id = ?', (cpid,))
                stan_name = list(split.keys())[i] if i < len(split) else 'Nowy'
                condition = STAN_TO_CONDITION.get(stan_name, stan_name)
                split_qty = int(split.get(stan_name, 0))
                for nr in range(1, split_qty + 1):
                    conn.execute('INSERT INTO sztuki (produkt_id, numer, stan, status) VALUES (?, ?, ?, ?)',
                        (cpid, nr, condition, 'magazyn'))
        elif stan:
            klasa = STAN_TO_KLASA.get(stan, '')
            condition = STAN_TO_CONDITION.get(stan, stan)
            conn.execute(
                'UPDATE produkty SET stan_przyjecia = ?, notatki_przyjecia = ?, klasa_jakosci = ?, stan = ? WHERE id = ?',
                (stan, notatki, klasa, condition, pid)
            )

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@magazynier_bp.route('/api/ai-ocena-stanu', methods=['POST'])
def ai_ocena_stanu():
    """AI photo analysis - ocena stanu produktu ze zdjęcia"""
    try:
        data = request.get_json()
        image_base64 = data.get('image_base64', '')
        if not image_base64:
            return jsonify({'success': False, 'error': 'Brak zdjęcia'})

        # Get Gemini API key from config
        conn = get_db()
        api_key = conn.execute("SELECT wartosc FROM config WHERE klucz = 'gemini_api_key'").fetchone()
        if not api_key or not api_key['wartosc']:
            return jsonify({'success': False, 'error': 'Brak klucza Gemini API w konfiguracji'})

        from modules.database import get_config as _get_cfg
        _zdjecia_model = _get_cfg('ai_model_zdjecia', _get_cfg('gemini_model', 'gemini-2.5-flash-lite'))

        import requests as req, json as _json
        _stan_schema = {
            "type": "OBJECT",
            "properties": {
                "stan": {"type": "STRING", "enum": ["Nowy", "Jak nowy", "Dobry", "Uszkodzony", "Zniszczony"]},
                "opis": {"type": "STRING"}
            },
            "required": ["stan", "opis"]
        }
        response = req.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/{_zdjecia_model}:generateContent?key={api_key['wartosc']}",
            headers={'Content-Type': 'application/json'},
            json={
                'contents': [{
                    'parts': [
                        {'text': 'Jesteś ekspertem od oceny stanu produktów zwrotowych. Oceń stan produktu na zdjęciu. Zwróć JSON z polami: stan (jedna z wartości: Nowy/Jak nowy/Dobry/Uszkodzony/Zniszczony) i opis (krótki opis stanu, wady, braki, uszkodzenia - max 2 zdania po polsku).'},
                        {'inline_data': {'mime_type': 'image/jpeg', 'data': image_base64}}
                    ]
                }],
                'generationConfig': {
                    'maxOutputTokens': 200,
                    'response_mime_type': 'application/json',
                    'response_schema': _stan_schema
                }
            },
            timeout=30
        )

        result = response.json()
        content = result.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')

        # Parse response — schema wymusza JSON, fallback na text parsing
        stan = ''
        opis = ''
        try:
            parsed = _json.loads(content)
            stan = parsed.get('stan', '')
            opis = parsed.get('opis', '')
        except Exception:
            for line in content.split('\n'):
                line = line.strip()
                if line.upper().startswith('STAN:'):
                    stan = line.split(':', 1)[1].strip()
                elif line.upper().startswith('OPIS:'):
                    opis = line.split(':', 1)[1].strip()

        stan_colors = {
            'Nowy': '#beee00', 'Jak nowy': '#8ff5ff', 'Dobry': '#eab308',
            'Uszkodzony': '#f97316', 'Zniszczony': '#ef4444'
        }

        return jsonify({
            'success': True,
            'stan': stan,
            'opis': opis or content,
            'stan_color': stan_colors.get(stan, '#94a3b8')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ============================================================
# RODO - ANONIMIZACJA DANYCH KLIENTA
# ============================================================

@magazynier_bp.route('/magazyn/api/check-duplicate', methods=['POST'])
def check_duplicate():
    """Sprawdza czy produkt o danym ASIN/EAN już istnieje w magazynie."""
    from .database import find_duplicate_product
    body = request.get_json(silent=True) or {}
    asin = body.get('asin', '').strip()
    ean = body.get('ean', '').strip()
    nazwa = body.get('nazwa', '').strip()

    if not asin and not ean:
        return jsonify({'found': False})

    existing = find_duplicate_product(asin=asin, ean=ean, nazwa=nazwa)
    if existing:
        return jsonify({
            'found': True,
            'product': {
                'id': existing['id'],
                'nazwa': existing['nazwa'],
                'asin': existing.get('asin', ''),
                'ean': existing.get('ean', ''),
                'ilosc': existing['ilosc'],
                'lokalizacja': existing.get('lokalizacja', ''),
                'regal': existing.get('regal', ''),
                'paleta': existing.get('paleta_nazwa', ''),
            }
        })
    return jsonify({'found': False})


@magazynier_bp.route('/magazyn/api/add-quantity', methods=['POST'])
def add_quantity():
    """Dodaje ilość do istniejącego produktu (zamiast tworzyć duplikat)."""
    from .database import add_quantity_to_existing
    body = request.get_json(silent=True) or {}
    product_id = body.get('product_id')
    quantity = int(body.get('quantity', 1))

    if not product_id or quantity < 1:
        return jsonify({'success': False, 'error': 'Brak product_id lub quantity'}), 400

    result = add_quantity_to_existing(product_id, quantity)
    if result:
        return jsonify({'success': True, 'product': result})
    return jsonify({'success': False, 'error': 'Nie znaleziono produktu'}), 404


@magazynier_bp.route('/magazyn/api/anonimizuj-klienta', methods=['POST'])
def anonimizuj_klienta():
    """
    RODO: Anonimizuje dane osobowe klienta w tabeli sprzedaze.
    Zachowuje: cena, ilosc, data_sprzedazy, produkt_id (cele ksiegowe).
    Accepts JSON: {buyer_name: "..."} or {sprzedaz_id: N}
    """
    data = request.get_json(silent=True) or {}
    buyer_name = data.get('buyer_name', '').strip()
    sprzedaz_id = data.get('sprzedaz_id')

    if not buyer_name and not sprzedaz_id:
        return jsonify({'ok': False, 'error': 'Podaj buyer_name lub sprzedaz_id'}), 400

    conn = get_db()
    try:
        if sprzedaz_id:
            # Anonimizuj po ID sprzedazy — znajdz kupujacego i anonimizuj wszystkie jego rekordy
            row = conn.execute('SELECT kupujacy FROM sprzedaze WHERE id = ?', (int(sprzedaz_id),)).fetchone()
            if not row:
                return jsonify({'ok': False, 'error': 'Nie znaleziono sprzedazy o podanym ID'}), 404
            buyer_name = row['kupujacy']
            if not buyer_name or buyer_name == 'Dane zanonimizowane':
                return jsonify({'ok': True, 'count': 0, 'message': 'Dane juz zanonimizowane'})

        cursor = conn.execute(
            "UPDATE sprzedaze SET kupujacy='Dane zanonimizowane', adres='Zanonimizowane' WHERE kupujacy=?",
            (buyer_name,)
        )
        conn.commit()
        count = cursor.rowcount
        return jsonify({'ok': True, 'count': count})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ============================================================
# STUDIO FOTO — Photo Pipeline UI
# ============================================================

@magazynier_bp.route('/studio-foto')
def studio_foto():
    """Widok kolejki photo daemon + statusy zdjęć produktów."""
    conn = get_db()

    # Inicjalizuj tabele photo_jobs jeśli jeszcze nie istnieją
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS photo_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_path TEXT NOT NULL,
                work_path TEXT,
                product_id INTEGER NULL,
                sku TEXT NULL,
                status TEXT NOT NULL DEFAULT 'new',
                error_msg TEXT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS processed_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER NOT NULL,
                product_id INTEGER NULL,
                sku TEXT NULL,
                variant TEXT NOT NULL,
                path TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        try:
            conn.execute("ALTER TABLE produkty ADD COLUMN images_ready INTEGER DEFAULT 0")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE produkty ADD COLUMN photo_job_id INTEGER NULL")
        except Exception:
            pass
        conn.commit()
    except Exception as _e:
        print(f"[studio-foto] table init: {_e}")

    # Statystyki kolejki
    stats = {
        'total': 0, 'new': 0, 'processing': 0,
        'done': 0, 'error': 0
    }
    try:
        for row in conn.execute("SELECT status, COUNT(*) as cnt FROM photo_jobs GROUP BY status").fetchall():
            s = row['status']
            cnt = row['cnt']
            stats['total'] += cnt
            if s in stats:
                stats[s] += cnt
    except Exception:
        pass

    # Ostatnie 50 jobów
    jobs = []
    try:
        jobs = [dict(r) for r in conn.execute(
            """SELECT j.*, p.nazwa as produkt_nazwa, p.ean, p.kod_magazynowy
               FROM photo_jobs j
               LEFT JOIN produkty p ON j.product_id = p.id
               ORDER BY j.created_at DESC LIMIT 50"""
        ).fetchall()]
    except Exception:
        pass

    # Produkty bez zdjęć (images_ready = 0 lub NULL)
    products_no_photo = []
    try:
        products_no_photo = [dict(r) for r in conn.execute(
            """SELECT id, nazwa, ean, zdjecie_url, ilosc, status,
                      (images_ready IS NULL OR images_ready = 0) as needs_photo
               FROM produkty
               WHERE (images_ready IS NULL OR images_ready = 0)
                 AND status NOT IN ('sprzedany','wyslany','zlomowany','uszkodzony')
                 AND ilosc > 0
               ORDER BY ilosc DESC LIMIT 100"""
        ).fetchall()]
    except Exception:
        pass

    return render_template('studio_foto.html',
        stats=stats,
        jobs=jobs,
        products_no_photo=products_no_photo,
    )


def _scrape_amazon_images(asin: str, zdjecie_url: str = '') -> list:
    """
    Pobiera wszystkie zdjęcia produktu Amazon po ASIN.
    Zwraca listę URL (max 8). Fallback: [zdjecie_url] jeśli scraping się nie uda.
    """
    import re, requests as _req

    if not asin or len(asin) < 8:
        return [zdjecie_url] if zdjecie_url else []

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept': 'text/html,application/xhtml+xml',
    }

    # Próbuj różne domeny Amazon
    for domain in ['amazon.com', 'amazon.de', 'amazon.co.uk', 'amazon.com.be']:
        try:
            url = f'https://www.{domain}/dp/{asin}'
            resp = _req.get(url, headers=headers, timeout=10, allow_redirects=True)
            if resp.status_code != 200:
                continue

            html = resp.text

            # Metoda 1: hiRes images z colorImages JSON
            hi_res = re.findall(r'"hiRes"\s*:\s*"(https://[^"]+\.jpg[^"]*)"', html)
            if hi_res:
                seen = []
                for u in hi_res:
                    # Zamień _SL75_ / _SY88_ itp. na _SL1500_ (max jakość)
                    clean = re.sub(r'\._[A-Z]{2}\d+_', '._AC_SL1500_', u)
                    if clean not in seen:
                        seen.append(clean)
                if seen:
                    return seen[:8]

            # Metoda 2: large images
            large = re.findall(r'"large"\s*:\s*"(https://[^"]+\.jpg[^"]*)"', html)
            if large:
                seen = list(dict.fromkeys(large))
                return seen[:8]

            # Metoda 3: data-old-hires attribute
            old_hires = re.findall(r'data-old-hires="(https://[^"]+)"', html)
            if old_hires:
                return list(dict.fromkeys(old_hires))[:8]

        except Exception:
            continue

    # Fallback: oryginalne zdjęcie
    return [zdjecie_url] if zdjecie_url else []


@magazynier_bp.route('/photo-request/<int:product_id>', methods=['POST'])
def photo_request(product_id):
    """
    Tworzy zlecenia przetworzenia zdjęć dla produktu.
    Scrape Amazon po ASIN żeby pobrać wszystkie zdjęcia (do 8 szt).
    """
    conn = get_db()
    try:
        p = conn.execute("SELECT id, ean, asin, nazwa, zdjecie_url FROM produkty WHERE id=?", (product_id,)).fetchone()
        if not p:
            return jsonify({'success': False, 'error': 'Produkt nie znaleziony'}), 404

        now = __import__('datetime').datetime.now().isoformat(sep=' ', timespec='seconds')

        # Sprawdź czy jest już aktywny job dla tego produktu
        existing_count = conn.execute(
            "SELECT COUNT(*) FROM photo_jobs WHERE product_id=? AND status IN ('new','processing')",
            (product_id,)
        ).fetchone()[0]
        if existing_count > 0:
            return jsonify({
                'success': False,
                'error': f'Produkt ma już {existing_count} aktywnych zleceń',
            }), 409

        # Pobierz wszystkie zdjęcia Amazon
        asin = p['asin'] or ''
        zdjecie_url = p['zdjecie_url'] or ''

        img_urls = _scrape_amazon_images(asin, zdjecie_url)
        if not img_urls:
            return jsonify({'success': False, 'error': 'Brak URL zdjęcia dla produktu'}), 400

        # Dodaj kolumnę image_index jeśli nie istnieje
        try:
            conn.execute("ALTER TABLE photo_jobs ADD COLUMN image_index INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass

        added = 0
        for idx, img_url in enumerate(img_urls):
            if not img_url:
                continue
            conn.execute(
                """INSERT INTO photo_jobs (original_path, product_id, sku, status, image_index, created_at, updated_at)
                   VALUES (?, ?, ?, 'new', ?, ?, ?)""",
                (img_url, product_id, p['ean'], idx, now, now)
            )
            added += 1

        conn.commit()
        return jsonify({
            'success': True,
            'message': f'Dodano {added} zleceń dla "{p["nazwa"]}" ({added} zdjęć Amazon)',
            'count': added,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@magazynier_bp.route('/photo-clear-and-requeue', methods=['POST'])
def photo_clear_and_requeue():
    """
    Usuwa wszystkie przetworzone zdjęcia (pliki + DB), usuwa wszystkie joby,
    i tworzy nowe zlecenia dla wszystkich produktów z ASIN (8 zdjęć/produkt).
    """
    import os
    from datetime import datetime as _dt
    conn = get_db()
    try:
        # 1. Usuń pliki
        paths = [r['path'] for r in conn.execute("SELECT path FROM processed_photos").fetchall()]
        deleted_files = 0
        for p in paths:
            try:
                if os.path.exists(p):
                    os.remove(p)
                    deleted_files += 1
            except Exception:
                pass

        # 2. Usuń wszystko z DB i od razu commituj — żeby zwolnić lock
        conn.execute("DELETE FROM processed_photos")
        conn.execute("DELETE FROM photo_jobs")
        conn.execute("UPDATE produkty SET images=NULL, images_ready=0 WHERE images IS NOT NULL")
        try:
            conn.execute("ALTER TABLE photo_jobs ADD COLUMN image_index INTEGER DEFAULT 0")
        except Exception:
            pass
        conn.commit()  # ← zwolnij lock przed scrapowaniem

        # 3. Pobierz listę produktów
        products = list(conn.execute(
            "SELECT id, ean, asin, nazwa, zdjecie_url FROM produkty "
            "WHERE (asin IS NOT NULL AND asin != '') OR (zdjecie_url IS NOT NULL AND zdjecie_url != '')"
        ).fetchall())

        # 4. Scrapuj Amazon i wstawiaj joby — commit po każdym produkcie
        import time as _time
        added_total = 0
        now = _dt.now().isoformat(sep=' ', timespec='seconds')

        for i, p in enumerate(products):
            asin = p['asin'] or ''
            zdjecie_url = p['zdjecie_url'] or ''

            if not asin or len(asin) < 8:
                if zdjecie_url:
                    conn.execute(
                        """INSERT INTO photo_jobs (original_path, product_id, sku, status, image_index, created_at, updated_at)
                           VALUES (?, ?, ?, 'new', 0, ?, ?)""",
                        (zdjecie_url, p['id'], p['ean'], now, now)
                    )
                    conn.commit()
                    added_total += 1
                continue

            img_urls = _scrape_amazon_images(asin, zdjecie_url)
            if img_urls:
                for idx, img_url in enumerate(img_urls):
                    if not img_url:
                        continue
                    conn.execute(
                        """INSERT INTO photo_jobs (original_path, product_id, sku, status, image_index, created_at, updated_at)
                           VALUES (?, ?, ?, 'new', ?, ?, ?)""",
                        (img_url, p['id'], p['ean'], idx, now, now)
                    )
                    added_total += 1
                conn.commit()  # commit per produkt — DB wolna między requestami

            if i % 5 == 4:
                _time.sleep(2)
            else:
                _time.sleep(0.8)
        return jsonify({
            'success': True,
            'added': added_total,
            'deleted_files': deleted_files,
            'message': f'Usunięto {deleted_files} plików, utworzono {added_total} nowych zleceń (8 zdjęć/produkt)'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@magazynier_bp.route('/photo-add-gallery', methods=['POST'])
def photo_add_gallery():
    """
    Dodaje brakujące zdjęcia galerii (index 1-7) dla produktów które mają
    już allegro_main ale NIE mają allegro_gallery_1.
    NIE usuwa istniejących zdjęć ani miniatur.
    """
    import time as _time
    from datetime import datetime as _dt
    conn = get_db()
    try:
        try:
            conn.execute("ALTER TABLE photo_jobs ADD COLUMN image_index INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass

        # Produkty które mają allegro_main w processed_photos ale brak allegro_gallery_1
        products = conn.execute(
            """SELECT DISTINCT p.id, p.ean, p.asin, p.zdjecie_url
               FROM produkty p
               JOIN processed_photos pp ON pp.product_id = p.id AND pp.variant = 'allegro_main'
               WHERE p.asin IS NOT NULL AND p.asin != ''
               AND NOT EXISTS (
                   SELECT 1 FROM processed_photos pp2
                   WHERE pp2.product_id = p.id AND pp2.variant = 'allegro_gallery_1'
               )
               AND NOT EXISTS (
                   SELECT 1 FROM photo_jobs pj
                   WHERE pj.product_id = p.id AND pj.image_index > 0
                   AND pj.status IN ('new','processing')
               )"""
        ).fetchall()

        added_total = 0
        now = _dt.now().isoformat(sep=' ', timespec='seconds')

        for i, p in enumerate(products):
            asin = p['asin'] or ''
            if not asin or len(asin) < 8:
                continue

            img_urls = _scrape_amazon_images(asin, p['zdjecie_url'] or '')
            if not img_urls or len(img_urls) < 2:
                continue

            # Tylko index 1..7 (galeria) — miniatura (0) już istnieje
            for idx, img_url in enumerate(img_urls[1:], start=1):
                if not img_url:
                    continue
                conn.execute(
                    """INSERT INTO photo_jobs (original_path, product_id, sku, status, image_index, created_at, updated_at)
                       VALUES (?, ?, ?, 'new', ?, ?, ?)""",
                    (img_url, p['id'], p['ean'], idx, now, now)
                )
                added_total += 1
            conn.commit()

            if i % 5 == 4:
                _time.sleep(2)
            else:
                _time.sleep(0.8)

        return jsonify({
            'success': True,
            'added': added_total,
            'products': len(products),
            'message': f'Dodano {added_total} zleceń galerii dla {len(products)} produktów (miniatury zachowane)'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@magazynier_bp.route('/photo-requeue-all', methods=['POST'])
def photo_requeue_all():
    """
    Kolejkuje WSZYSTKIE produkty z ASIN lub zdjecie_url które nie mają aktywnych jobów.
    Używa _scrape_amazon_images do pobrania wszystkich 8 URL-i Amazon.
    """
    from datetime import datetime as _dt
    conn = get_db()
    try:
        # Dodaj kolumnę image_index jeśli nie istnieje
        try:
            conn.execute("ALTER TABLE photo_jobs ADD COLUMN image_index INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass

        # Pobierz produkty z ASIN lub zdjecie_url
        products = conn.execute(
            "SELECT id, ean, asin, nazwa, zdjecie_url FROM produkty "
            "WHERE (asin IS NOT NULL AND asin != '') OR (zdjecie_url IS NOT NULL AND zdjecie_url != '')"
        ).fetchall()

        added_total = 0
        skipped = 0
        now = _dt.now().isoformat(sep=' ', timespec='seconds')

        for p in products:
            # Pomiń jeśli ma aktywny job
            active = conn.execute(
                "SELECT COUNT(*) FROM photo_jobs WHERE product_id=? AND status IN ('new','processing')",
                (p['id'],)
            ).fetchone()[0]
            if active > 0:
                skipped += 1
                continue

            asin = p['asin'] or ''
            zdjecie_url = p['zdjecie_url'] or ''
            img_urls = _scrape_amazon_images(asin, zdjecie_url)
            if not img_urls:
                continue

            for idx, img_url in enumerate(img_urls):
                if not img_url:
                    continue
                conn.execute(
                    """INSERT INTO photo_jobs (original_path, product_id, sku, status, image_index, created_at, updated_at)
                       VALUES (?, ?, ?, 'new', ?, ?, ?)""",
                    (img_url, p['id'], p['ean'], idx, now, now)
                )
                added_total += 1

        conn.commit()
        return jsonify({
            'success': True,
            'added': added_total,
            'skipped': skipped,
            'message': f'Dodano {added_total} nowych zleceń ({skipped} produktów pominięto — mają aktywne joby)'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@magazynier_bp.route('/photo-stats')
def photo_stats():
    """JSON ze statystykami kolejki photo_jobs — do live pollingu."""
    conn = get_db()
    stats = {'total': 0, 'new': 0, 'processing': 0, 'done': 0, 'error': 0}
    try:
        for row in conn.execute("SELECT status, COUNT(*) as cnt FROM photo_jobs GROUP BY status").fetchall():
            s = row['status']
            if s in stats:
                stats[s] += row['cnt']
            stats['total'] += row['cnt']
        # Ostatnie błędy
        last_errors = [dict(r) for r in conn.execute(
            "SELECT j.error_msg, p.nazwa FROM photo_jobs j LEFT JOIN produkty p ON j.product_id=p.id WHERE j.status='error' ORDER BY j.updated_at DESC LIMIT 3"
        ).fetchall()]
        stats['last_errors'] = last_errors
        # Ostatnio przetworzone
        last_done = [dict(r) for r in conn.execute(
            "SELECT p.nazwa, j.updated_at FROM photo_jobs j LEFT JOIN produkty p ON j.product_id=p.id WHERE j.status='done' ORDER BY j.updated_at DESC LIMIT 3"
        ).fetchall()]
        stats['last_done'] = last_done
    except Exception as e:
        stats['error_msg'] = str(e)
    stats['worker_running'] = _worker_running
    return jsonify(stats)


# ============================================================
# PHOTO WORKER — uruchomienie z poziomu UI
# ============================================================

_worker_running = False
_worker_last_log = ""

@magazynier_bp.route('/photo-worker-run', methods=['POST'])
def photo_worker_run():
    """Uruchamia photo_worker w tle (jeden run, max 10 jobów)."""
    global _worker_running, _worker_last_log
    import threading
    import sys
    from pathlib import Path

    if _worker_running:
        return jsonify({'success': False, 'error': 'Worker już działa — poczekaj chwilę'}), 429

    app = current_app._get_current_object()

    def run_worker():
        global _worker_running, _worker_last_log
        _worker_running = True
        try:
            import subprocess
            root = Path(app.root_path)
            worker_script = str(root / 'photo_daemon' / 'photo_worker.py')
            cfg_path = str(root / 'photo_daemon' / 'config.yaml')
            python_exe = sys.executable
            print(f"[photo-worker-run] {python_exe} {worker_script} --config {cfg_path}")

            result = subprocess.run(
                [python_exe, worker_script, '--config', cfg_path],
                capture_output=True, text=True, timeout=300,
                cwd=str(root)
            )
            out = (result.stdout or '')[-1000:]
            err = (result.stderr or '')[-1000:]
            _worker_last_log = f"=== STDOUT ===\n{out}\n=== STDERR ===\n{err}\n=== CODE: {result.returncode} ==="
            print(f"[photo-worker-run] rc={result.returncode}")
            print(f"[photo-worker-run] stdout: {out[-300:]}")
            if err:
                print(f"[photo-worker-run] stderr: {err[-300:]}")
        except Exception as e:
            _worker_last_log = f"EXCEPTION: {e}"
            print(f"[photo-worker-run] ERROR: {e}")
            import traceback; traceback.print_exc()
        finally:
            _worker_running = False

    t = threading.Thread(target=run_worker, daemon=True)
    t.start()
    return jsonify({'success': True, 'message': 'Worker uruchomiony w tle — odśwież stronę za chwilę'})


@magazynier_bp.route('/photo-by-id/<int:photo_id>')
def photo_by_id(photo_id):
    """Serwuje przetworzony plik zdjęcia po ID rekordu processed_photos."""
    import os
    from flask import send_file, abort
    conn = get_db()
    try:
        row = conn.execute("SELECT path FROM processed_photos WHERE id=?", (photo_id,)).fetchone()
        if not row or not os.path.exists(row['path']):
            abort(404)
        # SECURITY: walidacja ścieżki — tylko pliki z katalogu aplikacji
        _app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        _real = os.path.realpath(row['path'])
        if not _real.startswith(os.path.realpath(_app_dir)):
            abort(403)
        return send_file(_real, mimetype='image/jpeg')
    except Exception:
        abort(404)


@magazynier_bp.route('/photo-file/<int:product_id>/<variant>')
def photo_file(product_id, variant):
    """Serwuje przetworzony plik zdjęcia (allegro_main / vinted / thumb)."""
    import os
    from flask import send_file, abort
    conn = get_db()
    allowed = {'allegro_main', 'vinted', 'thumb'}
    if variant not in allowed:
        abort(404)
    try:
        row = conn.execute(
            "SELECT path FROM processed_photos WHERE product_id=? AND variant=? ORDER BY id DESC LIMIT 1",
            (product_id, variant)
        ).fetchone()
        if not row:
            abort(404)
        path = row['path']
        if not os.path.exists(path):
            abort(404)
        return send_file(path, mimetype='image/jpeg')
    except Exception:
        abort(404)


@magazynier_bp.route('/photo-worker-log')
def photo_worker_log():
    """Zwraca ostatni log workera — do diagnostyki."""
    global _worker_last_log, _worker_running
    return f"<pre>running={_worker_running}\n\n{_worker_last_log or '(brak logu — kliknij Uruchom worker najpierw)'}</pre>"


@magazynier_bp.route('/studio-foto/galeria')
def studio_foto_galeria():
    """Galeria wszystkich przetworzonych zdjęć produktów."""
    conn = get_db()

    search = request.args.get('q', '').strip()
    page = max(0, int(request.args.get('page', 0)))
    per_page = 30  # produktów na stronę

    try:
        # Znajdź produkty które mają przetworzone zdjęcia
        search_where = ""
        params = []
        if search:
            search_where = "AND (p.nazwa LIKE ? OR p.ean LIKE ? OR p.kod_magazynowy LIKE ?)"
            params = [f'%{search}%', f'%{search}%', f'%{search}%']

        # Unikalne product_id z processed_photos
        product_ids_rows = conn.execute(
            f"""SELECT DISTINCT pp.product_id
                FROM processed_photos pp
                LEFT JOIN produkty p ON pp.product_id = p.id
                WHERE pp.product_id IS NOT NULL {search_where}
                ORDER BY pp.product_id DESC
                LIMIT ? OFFSET ?""",
            params + [per_page, page * per_page]
        ).fetchall()

        total_products = conn.execute(
            f"""SELECT COUNT(DISTINCT pp.product_id)
                FROM processed_photos pp
                LEFT JOIN produkty p ON pp.product_id = p.id
                WHERE pp.product_id IS NOT NULL {search_where}""",
            params
        ).fetchone()[0]

        # Dla każdego produktu pobierz wszystkie warianty zdjęć
        products = []
        for row in product_ids_rows:
            pid = row['product_id']
            # Info o produkcie
            p_info = conn.execute(
                "SELECT id, nazwa, ean, kod_magazynowy, status, ilosc FROM produkty WHERE id=?", (pid,)
            ).fetchone()
            if not p_info:
                continue
            # Wszystkie przetworzone zdjęcia produktu (posortowane: main, potem gallery_1..7)
            photos = conn.execute(
                """SELECT pp.id, pp.variant, pp.path, pp.created_at
                   FROM processed_photos pp
                   WHERE pp.product_id = ?
                   ORDER BY CASE
                       WHEN pp.variant = 'allegro_main' THEN 0
                       WHEN pp.variant = 'vinted' THEN 99
                       WHEN pp.variant = 'thumb' THEN 100
                       ELSE CAST(REPLACE(pp.variant, 'allegro_gallery_', '') AS INTEGER)
                   END ASC""",
                (pid,)
            ).fetchall()
            products.append({
                'product': dict(p_info),
                'photos': [dict(ph) for ph in photos],
            })

    except Exception as e:
        products = []
        total_products = 0
        print(f"[galeria] error: {e}")

    return render_template('studio_foto_galeria.html',
        products=products,
        total=total_products,
        page=page,
        per_page=per_page,
        pages=max(1, (total_products + per_page - 1) // per_page),
        search=search,
    )
